"""
TezBozor Mini App backend (FastAPI) — Faza 1.

- Bot bilan AYNAN bir xil PostgreSQL bazaga ulanadi (database.py qayta ishlatiladi)
  → ma'lumot avtomatik 1:1, sinxronlash kerak emas.
- Telegram WebApp `initData` imzosini bot tokeni bilan tekshiradi (xavfsizlik).
- Mahsulot rasmlari Telegram file_id — ularni web'da ko'rsatish uchun rasm-proksi
  (getFile → yuklab → disk-cache → stream).

Ishga tushirish (lokal/sinov):
    uvicorn webapp_server:app --host 127.0.0.1 --port 8080

Kerakli paketlar:  fastapi  uvicorn[standard]  httpx
"""
import os
import re
import html
import json
import hashlib
import hmac
import logging
import asyncio
import base64
import math
from typing import Optional, List, Any
from datetime import datetime, timezone, timedelta

try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

from fastapi import FastAPI, Header, HTTPException, Query, File, UploadFile, Request, BackgroundTasks
from fastapi.responses import Response, FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
import httpx

from database import Database
from webapp_auth import validate_init_data
from languages import t, get_user_lang, DEFAULT_LANG
from tezbozor_design import (fmt_order_id, fmt_price, best_location_text, effective_unit_price,
                             wholesale_info, normalize_schedule, next_shop_open_datetime)
import ai_assistant
import ai_vision
import ad_design
import ad_video

BOT_TOKEN = os.getenv("BOT_TOKEN")
# Rasmiy kanal havolasi (app'dagi "kanalga o'tish" tugmasi — bot bilan bir xil manba)
CHANNEL_ID = os.getenv("CHANNEL_ID")
CHANNEL_URL = (f"https://t.me/{str(CHANNEL_ID).lstrip('@')}"
               if CHANNEL_ID and str(CHANNEL_ID).startswith('@') else None)
try:
    ADMIN_ID = int(os.getenv("ADMIN_ID", "0") or "0")
except ValueError:
    ADMIN_ID = 0
STATIC_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "webapp_static")
IMG_CACHE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "img_cache")
os.makedirs(IMG_CACHE_DIR, exist_ok=True)

db = Database()  # DB_BACKEND / DATABASE_URL .env'dan o'qiladi

# Brend nomi — .env dagi BRAND_NAME. Bu yerda (fayl boshida) e'lon qilinadi,
# chunki quyidagi app = FastAPI(...) import paytida bajariladi.
BRAND_NAME = (os.getenv("BRAND_NAME", "").strip() or "TezBozor")

app = FastAPI(title=f"{BRAND_NAME} Mini App API")
# Tezlik #3 — javoblarni gzip qiladi. index.html (~150KB) va mahsulot JSON ro'yxati
# ~5-7 barobar kichrayadi → sekin tarmoqda sezilarli tezlik. Rasm (image/jpeg)
# allaqachon siqilgan, lekin min_size dan kichik bo'lsa ham zarar yo'q.
from starlette.middleware.gzip import GZipMiddleware  # noqa: E402
app.add_middleware(GZipMiddleware, minimum_size=600)


# ---- Xavfsizlik sarlavhalari (chuqurlashtirilgan himoya — nginx sozlanmagan bo'lsa ham) ----
# DIQQAT: X-Frame-Options / frame-ancestors QO'YILMAYDI — Telegram Mini App iframe ichida
# ishlaydi, DENY bo'lsa ilova ochilmaydi. HSTS va CSP env orqali yoqiladi (inline-JS ko'p
# bo'lgani uchun CSP sukut bo'yicha o'chiq — noto'g'ri siyosat butun ilovani buzishi mumkin).
_SEC_HSTS = os.getenv("SECURITY_HSTS", "").strip() in ("1", "true", "yes")
_SEC_CSP = (os.getenv("SECURITY_CSP", "").strip() or None)


@app.middleware("http")
async def _security_headers(request: Request, call_next):
    resp = await call_next(request)
    resp.headers.setdefault("X-Content-Type-Options", "nosniff")
    resp.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    resp.headers.setdefault("X-XSS-Protection", "0")  # zamonaviy brauzerlar CSP'ga tayanadi
    if _SEC_HSTS:
        resp.headers.setdefault(
            "Strict-Transport-Security", "max-age=31536000; includeSubDomains")
    if _SEC_CSP:
        resp.headers.setdefault("Content-Security-Policy", _SEC_CSP)
    return resp


# ============================================================
# initData TEKSHIRUVI (Telegram WebApp xavfsizligi) — webapp_auth.py'da
# ============================================================
def require_auth(authorization: str):
    """Authorization sarlavhasi: "tma <initData>". To'g'ri bo'lmasa 401."""
    init_data = ""
    if authorization and authorization.startswith("tma "):
        init_data = authorization[4:]
    auth = validate_init_data(init_data, BOT_TOKEN)
    if not auth:
        raise HTTPException(status_code=401, detail="invalid initData")
    return auth


# Xaridorga HECH QACHON yubormaslik kerak bo'lgan maxfiy mahsulot maydonlari (#8 floor).
_SECRET_PRODUCT_FIELDS = ("min_price",)


def _rows(result):
    """Row-larni JSON uchun dict'ga aylantiradi (PG shim _Row ham, sqlite3.Row ham).
    Maxfiy maydonlar (min_price — savdolashish floor'i) olib tashlanadi."""
    if not result:
        return []
    out = []
    for r in result:
        d = dict(r)
        for k in _SECRET_PRODUCT_FIELDS:
            d.pop(k, None)
        # #18 Pro nishon — xom pro_until vaqtini bool'ga aylantiramiz (mahsulot kartasi:
        # seller_pro_until; do'kon kartasi: pro_until). Frontend faqat bool ko'radi.
        if "seller_pro_until" in d:
            d["seller_is_pro"] = _pro_until_active(d.pop("seller_pro_until"))
        if "pro_until" in d:
            d["is_pro"] = _pro_until_active(d.pop("pro_until"))
        out.append(d)
    return out


def _own_shop_seller_id(authorization):
    """Joriy foydalanuvchining O'Z do'koni seller_id si (do'kon egasi yoki xodim bo'lsa),
    aks holda None. Buyer katalogida o'z do'koni mahsulotlarini YASHIRISH uchun: sotuvchi
    o'z mahsulotini sotib ololmaydi (own_product), shuning uchun katalogda ko'rsatmaymiz.
    Oddiy xaridorda do'kon yo'q → o'z id'si qaytadi, lekin u hech bir mahsulot seller_id'siga
    to'g'ri kelmaydi (xaridorda mahsulot yo'q) → hech narsa filtrlanmaydi (xavfsiz)."""
    try:
        auth = require_auth(authorization)
        tg = (auth.get("user") or {}).get("id")
        if not tg:
            return None
        u = db.get_user_by_telegram_id(tg)
        return _owner_id(dict(u)) if u else None
    except Exception:
        return None


def _hide_own(rows, own_seller_id):
    """`rows` (dict ro'yxati) ichidan seller_id == own_seller_id bo'lganlarini olib tashlaydi.
    own_seller_id None bo'lsa — o'zgartirmasdan qaytaradi."""
    if not own_seller_id:
        return rows
    return [d for d in rows if d.get("seller_id") != own_seller_id]


# ============================================================
# API
# ============================================================
@app.get("/api/categories")
def api_categories(authorization: str = Header(None)):
    require_auth(authorization)
    return _rows(db.get_categories())


@app.get("/api/products")
def api_products(
    authorization: str = Header(None),
    q: str = Query(None),
    category_id: int = Query(None),
    sort: str = Query("rating"),
    region_id: int = Query(None),
    seller_id: int = Query(None),
    wholesale: bool = Query(False),
):
    require_auth(authorization)
    items = db.search_products(
        query=q, category_id=category_id, sort_by=sort, region_id=region_id,
        seller_id=seller_id,
    )
    rows = _hide_own(_rows(items), _own_shop_seller_id(authorization))
    # Savdo turi bo'yicha ajratish: "Optom" bo'limi faqat optom (pachka) mahsulotlar;
    # "Mahsulotlar" (dona) bo'limi optomni ko'rsatmaydi.
    if wholesale:
        rows = [r for r in rows if (r.get("sale_mode") == "optom")]
    else:
        rows = [r for r in rows if (r.get("sale_mode") != "optom")]
    return rows


@app.get("/api/products/{product_id}/related")
def api_product_related(product_id: int, authorization: str = Header(None)):
    """AI #10 — "Bular bilan olishadi" (item-to-item cross-sell)."""
    require_auth(authorization)
    return _hide_own(_rows(db.get_frequently_bought_together(product_id, 8)),
                     _own_shop_seller_id(authorization))


@app.get("/api/products/ai-search")
async def api_products_ai_search(
    authorization: str = Header(None),
    q: str = Query(...),
    category_id: int = Query(None),
    region_id: int = Query(None),
):
    """#9 — sheva/slang qidiruv. Oddiy qidiruv natija bermaganda chaqiriladi:
    AI so'rovni standart kalit so'zlarga aylantiradi va qayta qidiradi.
    Qaytaradi {"interpreted": "<topilgan so'zlar>", "items": [...]}."""
    auth = require_auth(authorization)
    query = (q or "").strip()
    if not query:
        return {"interpreted": "", "items": []}
    if not ai_assistant.is_enabled():
        return {"interpreted": "", "items": []}
    uid = (auth.get("user") or {}).get("id")
    _rate_limit("ai_search", uid or 0, 30, 3600)
    lang = DEFAULT_LANG
    try:
        u = db.get_user_by_telegram_id(uid) if uid else None
        if u:
            lang = get_user_lang(dict(u)) or DEFAULT_LANG
    except Exception:
        pass
    cat_names = []
    try:
        cat_names = [dict(c).get("name") for c in (db.get_categories() or [])]
    except Exception:
        cat_names = []
    interp = await ai_assistant.interpret_search_query(query, categories=cat_names, lang=lang)
    if not interp or not interp.get("keywords"):
        return {"interpreted": "", "items": []}
    # AI taklif qilgan kalit so'zlar bo'yicha (transliteratsiyali) qidiramiz; takrorlanmas natija.
    seen, items = set(), []
    for kw in interp["keywords"]:
        try:
            rows = db.search_products(query=kw, category_id=category_id,
                                      sort_by="rating", region_id=region_id)
        except Exception:
            rows = []
        for r in rows:
            rid = dict(r).get("id")
            if rid not in seen:
                seen.add(rid)
                items.append(r)
        if len(items) >= 30:
            break
    return {"interpreted": ", ".join(interp["keywords"]),
            "items": _hide_own(_rows(items[:30]), _own_shop_seller_id(authorization))}


def _shop_facts_text(prod, lang):
    """#3 — RAG uchun do'kon+mahsulot faktlarini matn blokiga yig'adi."""
    uz = lang != "ru"
    L = (lambda u, r: u if uz else r)
    lines = []
    def add(label, val):
        if val not in (None, "", 0):
            lines.append(f"- {label}: {val}")
    add(L("Do'kon", "Магазин"), prod.get("shop_name"))
    add(L("Hudud", "Регион"), db.get_region_label(prod.get("seller_region_id")) or "")
    add(L("Manzil", "Адрес"), prod.get("shop_address"))
    add(L("Mo'ljal", "Ориентир"), prod.get("shop_landmark"))
    add(L("Ish kunlari", "Рабочие дни"), prod.get("working_days"))
    add(L("Ish soatlari", "Рабочие часы"), prod.get("working_hours"))
    add(L("Telefon", "Телефон"), prod.get("phone_number"))
    add(L("Mahsulot", "Товар"), prod.get("name"))
    price = prod.get("price")
    if price:
        unit = prod.get("unit")
        add(L("Narxi", "Цена"), f"{fmt_price(price)}" + (f" / {unit}" if unit else ""))
    sc = prod.get("stock_count")
    if sc is not None:
        add(L("Zahirada", "В наличии"), f"{sc}")
    add(L("Kategoriya", "Категория"), prod.get("category_name"))
    if not lines:
        return L("Ma'lumot yo'q", "Нет данных")
    return "\n".join(lines)


class ShopAskIn(BaseModel):
    question: str


@app.post("/api/product/{product_id}/ask")
async def api_product_ask(product_id: int, body: ShopAskIn, authorization: str = Header(None)):
    """#3 AI-menejer — mijoz savoliga do'kon profili asosida (RAG) avtomatik javob."""
    user = dict(_buyer_from_auth(authorization))
    _rate_limit("shop_ask", user["id"], 20, 3600)
    q = (body.question or "").strip()
    if not q:
        raise HTTPException(status_code=400, detail="empty")
    if len(q) > 500:
        raise HTTPException(status_code=400, detail="too_long")
    if not ai_assistant.is_enabled():
        raise HTTPException(status_code=503, detail="ai_disabled")
    prod = db.get_product_by_id(product_id)
    if not prod:
        raise HTTPException(status_code=404, detail="not_found")
    prod = dict(prod)
    lang = get_user_lang(user) or DEFAULT_LANG
    facts = _shop_facts_text(prod, lang)
    ans = await ai_assistant.answer_shop_question(question=q, facts=facts, lang=lang)
    if not ans:
        raise HTTPException(status_code=502, detail="ai_error")
    return {"answer": ans}


class HaggleIn(BaseModel):
    message: str
    history: Optional[List[dict]] = None


def _haggle_parse_price(text: str):
    """Xaridor xabaridan taklif qilingan narxni o'qiydi. '150 ming', '150000',
    '150k', '1.5 mln', '150 000' kabilarni tushunadi. Topolmasa None."""
    import re as _re2
    s = (text or "").lower().replace(" ", " ")
    # "150 ming" / "150k" / "1.5 mln" — ko'paytma birliklari
    for pat, mult in ((r"(\d+[.,]?\d*)\s*(?:mln|млн|million|миллион)", 1_000_000),
                      (r"(\d+[.,]?\d*)\s*(?:ming|тыс|k\b|минг)", 1_000)):
        m = _re2.search(pat, s)
        if m:
            try:
                return int(float(m.group(1).replace(",", ".")) * mult)
            except (TypeError, ValueError):
                pass
    # oddiy raqamlar — guruh bo'shliqlarini olib tashlab ("150 000" → 150000)
    nums = _re2.findall(r"\d[\d\s]*\d|\d", s)
    best = 0
    for n in nums:
        try:
            v = int(n.replace(" ", ""))
            best = max(best, v)
        except (TypeError, ValueError):
            pass
    return best or None


def _haggle_fallback(*, listed: float, floor: float, buyer_message: str,
                     history=None, lang: str = "uz") -> dict:
    """AI ishlamasa (timeout/bo'sh javob/balans) — savdolashish to'xtab qolmasligi
    uchun deterministik zaxira.

    MUHIM (eng himoyalangan qism): narx BIRDANIGA floor'ga TUSHMAYDI. Har raundda
    bo'shliqning kichik ulushiga (~20%) kamayadi — listed'dan boshlab, xaridor
    qistagan sayin asta-sekin pasayadi. Floor HECH QACHON buzilmaydi, listed'dan
    oshmaydi (server clamp'i ham buni qo'shimcha kafolatlaydi)."""
    ru = (lang == "ru")
    listed_i, floor_i = int(listed), int(floor)
    gap = max(0, listed_i - floor_i)
    # nechta raund o'tgan — tarixda sotuvchining (assistant) oldingi takliflari soni
    rounds = sum(1 for h in (history or []) if (h or {}).get("role") == "assistant")
    # har raundda bo'shliqning ~20% chegirma; ~5 raunddan keyingina floor'ga yetadi
    frac = min(1.0, 0.20 * (rounds + 1))
    target = max(floor_i, int(round(listed_i - gap * frac)))

    offered = _haggle_parse_price(buyer_message)
    if offered is None:
        return {"reply": ("Сколько предложите? Назовите свою цену 🙂" if ru
                          else "Qancha taklif qilasiz? O'z narxingizni ayting 🙂"),
                "offer_price": listed_i, "accepted": False}
    if offered >= listed_i:
        return {"reply": (f"Хорошо, оформляем за {listed_i} 🤝" if ru
                          else f"Mayli, {listed_i} so'mga rasmiylashtiramiz 🤝"),
                "offer_price": listed_i, "accepted": True}
    # xaridor bizning (asta tushayotgan) so'rovimizga yetdi → kelishamiz
    if offered >= target:
        return {"reply": (f"Договорились, пусть будет {target} 🤝" if ru
                          else f"Kelishdik, {target} so'm bo'la qolsin 🤝"),
                "offer_price": target, "accepted": True}
    # hali past — bittada tushmaymiz, ozgina chegirma bilan qarshi-taklif beramiz
    if target <= floor_i:
        reply = ("Это уже самая низкая цена, ниже не могу. Берём? 🙂" if ru
                 else "Bundan past tusholmayman, eng yaxshi narx shu. Olamizmi? 🙂")
    else:
        reply = (f"Столько не выйдет, но уступлю — пусть будет {target} 🙂" if ru
                 else f"Bunga bo'lmaydi-yu, sal yon beraman — {target} so'm bo'lsin 🙂")
    return {"reply": reply, "offer_price": target, "accepted": False}


@app.post("/api/product/{product_id}/haggle")
async def api_product_haggle(product_id: int, body: HaggleIn, authorization: str = Header(None)):
    """#8 AI savdolashish — sotuvchi nomidan, maxfiy floor bilan. Kelishilsa narx
    qisqa muddatga saqlanadi (checkout hurmat qiladi). Floor server'da ham himoyalangan."""
    user = dict(_buyer_from_auth(authorization))
    _rate_limit("haggle", user["id"], 40, 3600)
    msg = (body.message or "").strip()
    if not msg:
        raise HTTPException(status_code=400, detail="empty")
    if len(msg) > 500:
        raise HTTPException(status_code=400, detail="too_long")
    # Arzon validatsiya guardlari AI tekshiruvidan OLDIN: mavjud emas / o'z mahsuloti /
    # savdolashish o'chiq holatlar AI sozlanishidan qat'i nazar bir xil javob berishi
    # kerak (aks holda AI yo'q muhitda — masalan CI — 503 "ai_disabled" own_product'ni
    # to'sib qo'yardi). is_enabled() 503'i faqat haqiqiy AI chaqiruvidan oldin tekshiriladi.
    prod = dict(db.get_product_by_id(product_id) or {})
    if not prod:
        raise HTTPException(status_code=404, detail="not_found")
    if prod.get("seller_id") == user["id"]:
        raise HTTPException(status_code=400, detail="own_product")
    listed = float(prod.get("price") or 0)
    floor = prod.get("min_price")
    if not floor or float(floor) <= 0 or float(floor) >= listed:
        raise HTTPException(status_code=409, detail="haggle_off")
    floor = float(floor)
    lang = get_user_lang(user) or DEFAULT_LANG
    # AI yoqilgan bo'lsa — u bilan savdolashamiz. AI o'chiq YOKI muvaffaqiyatsiz
    # (timeout/bo'sh javob/balans) bo'lsa — savdolashish to'xtab qolmasligi uchun
    # deterministik zaxiraga o'tamiz (floor server'da baribir himoyalangan).
    res = None
    if ai_assistant.is_enabled():
        res = await ai_assistant.haggle(listed_price=listed, floor_price=floor,
                                        history=body.history, buyer_message=msg, lang=lang)
    if not res:
        res = _haggle_fallback(listed=listed, floor=floor, buyer_message=msg,
                               history=body.history, lang=lang)
    # SERVER himoyasi: narx hech qachon floor'dan past / listed'dan baland bo'lmaydi
    price = max(floor, min(float(res["offer_price"]), listed))
    if res["accepted"]:
        db.set_haggle_deal(user["id"], product_id, price, ttl_minutes=60)
        return {"reply": res["reply"], "accepted": True, "agreed_price": int(price)}
    return {"reply": res["reply"], "accepted": False, "offer_price": int(price)}


class GiftIn(BaseModel):
    recipient: Optional[str] = None
    occasion: Optional[str] = None
    budget: Optional[int] = None
    notes: Optional[str] = None


@app.post("/api/gift-assistant")
async def api_gift_assistant(body: GiftIn, authorization: str = Header(None)):
    """#17 — aqlli sovg'a yordamchisi: kimga + sabab + budjet → AI g'oyalar + har
    g'oyaga mos REAL mahsulotlar (budjet ichida)."""
    user = dict(_buyer_from_auth(authorization))
    if not ai_assistant.is_enabled():
        raise HTTPException(status_code=503, detail="ai_disabled")
    _rate_limit("gift", user["id"], 25, 3600)
    recipient = (body.recipient or "").strip()[:120]
    occasion = (body.occasion or "").strip()[:120]
    notes = (body.notes or "").strip()[:200]
    if not (recipient or occasion or notes):
        raise HTTPException(status_code=400, detail="empty")
    budget = None
    if body.budget is not None and body.budget > 0:
        budget = int(body.budget)
    lang = get_user_lang(user) or DEFAULT_LANG
    region_id = user.get("region_id")
    cat_names = []
    try:
        cat_names = [dict(c).get("name") for c in (db.get_categories() or [])]
    except Exception:
        cat_names = []
    budget_txt = f"{budget}" if budget else ("cheklanmagan" if lang == "uz" else "не ограничен")
    res = await ai_assistant.gift_advisor(
        recipient=recipient, occasion=occasion, budget=budget_txt,
        notes=notes, categories=cat_names, lang=lang)
    if not res:
        raise HTTPException(status_code=502, detail="ai_error")
    # Har g'oyaga budjet ichida mos mahsulotlarni biriktiramiz (takrorlanmas)
    used = set()
    ideas_out = []
    for idea in res["ideas"]:
        prods, seen = [], set()
        for kw in idea.get("keywords") or []:
            try:
                rows = db.search_products(query=kw, max_price=budget,
                                          sort_by="rating", region_id=region_id)
            except Exception:
                rows = []
            for r in rows:
                rid = dict(r).get("id")
                if rid in seen or rid in used:
                    continue
                seen.add(rid)
                prods.append(r)
                if len(prods) >= 3:
                    break
            if len(prods) >= 3:
                break
        for r in prods:
            used.add(dict(r).get("id"))
        ideas_out.append({"title": idea["title"], "reason": idea.get("reason", ""),
                          "products": _rows(prods)})
    return {"intro": res.get("intro", ""), "ideas": ideas_out, "budget": budget}


@app.get("/api/favorites")
def api_favorites_list(authorization: str = Header(None)):
    """#16 — xaridorning sevimli mahsulotlari ro'yxati."""
    buyer = _buyer_from_auth(authorization)
    return _rows(db.get_favorites(buyer["id"]))


@app.post("/api/favorites/{product_id}")
def api_favorite_add(product_id: int, authorization: str = Header(None)):
    buyer = _buyer_from_auth(authorization)
    if not db.get_product_by_id(product_id):
        raise HTTPException(status_code=404, detail="not_found")
    db.add_favorite(buyer["id"], product_id)
    return {"ok": True, "is_favorite": True}


@app.delete("/api/favorites/{product_id}")
def api_favorite_remove(product_id: int, authorization: str = Header(None)):
    buyer = _buyer_from_auth(authorization)
    db.remove_favorite(buyer["id"], product_id)
    return {"ok": True, "is_favorite": False}


@app.get("/api/discover")
def api_discover(authorization: str = Header(None)):
    """#15 Kashfiyot — Trend (eng ko'p buyurtma), Chegirmalar, Yaqin atrofda (xaridor
    hududidagi yangi mahsulotlar). Mavjud ma'lumotdan; AI tavsiya keyingi bosqich."""
    auth = require_auth(authorization)
    out = {
        "for_you": [],
        "trending": _rows(db.get_trending_products(12)),
        "discounts": _rows(db.get_discounted_products(12)),
        "nearby": [],
    }
    tg_id = (auth.get("user") or {}).get("id")
    u = db.get_user_by_telegram_id(tg_id) if tg_id else None
    own = None
    if u:
        uid = dict(u)["id"]
        own = _owner_id(dict(u))   # o'z do'koni mahsulotlarini yashirish uchun
        out["for_you"] = _rows(db.get_recommendations(uid, 12))   # #1 shaxsiy tavsiya
        rid = dict(u).get("region_id")
        if rid:
            out["nearby"] = _rows(db.search_products(region_id=rid, sort_by="newest")[:12])
    # Har bo'limdan o'z do'koni mahsulotlarini olib tashlaymiz (buyer katalog parite)
    for k in ("for_you", "trending", "discounts", "nearby"):
        out[k] = _hide_own(out[k], own)
    return out


@app.get("/api/shops")
def api_shops(authorization: str = Header(None), q: str = Query(None),
              region_id: int = Query(None)):
    require_auth(authorization)
    return _rows(db.search_shops(query=q, region_id=region_id))


@app.get("/api/regions")
def api_regions(authorization: str = Header(None), parent_id: int = Query(None)):
    """Viloyatlar (parent_id yo'q) yoki tumanlar (parent_id=viloyat)."""
    require_auth(authorization)
    return _rows(db.get_regions(parent_id))


# AI yordamchi — DeepSeek (ai_assistant.ask qayta ishlatiladi). Tarix xotirada (tg_id bo'yicha).
AI_SESSIONS = {}


class AiAsk(BaseModel):
    text: str
    mode: str = "buyer"   # 'buyer' | 'seller'


@app.post("/api/ai")
async def api_ai(body: AiAsk, authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    _rate_limit("ai", user.get("id"), 20, 60)
    text = (body.text or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="empty")
    if len(text) > 1000:
        raise HTTPException(status_code=400, detail="too_long")
    if not ai_assistant.is_enabled():
        raise HTTPException(status_code=503, detail="ai_disabled")
    # Sotuvchi rejimi — faqat seller/admin uchun; AI sotuvchi vositalarini (mahsulot
    # qo'shish/tahrirlash va h.k.) ishlatadi. Aks holda xaridor rejimi.
    if body.mode == "seller" and user.get("role") in ("seller", "admin"):
        role, seller_id = "seller", user.get("id")
    else:
        role, seller_id = "buyer", None
    ud = AI_SESSIONS.setdefault((user.get("telegram_id"), role), {})
    lang = user.get("language") or "uz"
    res = await ai_assistant.ask(db, lang, role, text, ud,
                                 seller_id=seller_id, user_name=user.get("name") or "",
                                 buyer_id=user.get("id"), channel="app")
    # cart_adds — agent savatga solgan mahsulotlar; frontend localStorage savatiga qo'shadi
    return {"text": res.get("text"), "products": _rows(res.get("products") or []) or None,
            "cart_adds": res.get("cart_adds") or None}


# ============================================================
# UNIVERSAL XABARNOMA + MUROJAAT (support) tizimi
# Foydalanuvchiga kelgan HAR qanday xabar → `notifications` (app top banner) + bir
# martalik Telegram push. Murojaat = 2 tomonlama suhbat (foydalanuvchi↔admin).
# ============================================================
def _notify_db(user_id, kind, title, body="", ref_id=None):
    """Faqat app-banner xabarnomasi (Telegram push'i mavjud nuqtalarда qo'shimcha — bot
    o'zi push yuboradi). title/body string YOKI (uz, ru) tuple bo'lishi mumkin — tuple bo'lsa
    qabul qiluvchi tilida saqlanadi. Sync; xatoda jim o'tadi (asosiy oqimni buzmaydi)."""
    if not user_id:
        return
    try:
        ru = False
        if isinstance(title, tuple) or isinstance(body, tuple):
            u = db.get_user_by_id(user_id)
            ru = (get_user_lang(u) == "ru") if u else False
        t = (title[1] if ru else title[0]) if isinstance(title, tuple) else title
        b = (body[1] if ru else body[0]) if isinstance(body, tuple) else body
        db.create_notification(user_id, kind, t, b, ref_id)
    except Exception as e:
        logging.warning(f"notif_db xato (user {user_id}): {e}")


async def _notify_user(user_id, title, body, kind="info", ref_id=None, push=True):
    """Foydalanuvchiga xabarnoma: DB'ga yozadi (app banner) + Telegram push (bir marta)."""
    if not user_id:
        return
    try:
        db.create_notification(user_id, kind, title, body, ref_id)
    except Exception as e:
        logging.warning(f"notif yozish xato (user {user_id}): {e}")
    if not push:
        return
    try:
        u = db.get_user_by_id(user_id)
        tg = dict(u).get("telegram_id") if u else None
        if tg:
            text = f"<b>{html.escape(title or '')}</b>"
            if body:
                text += f"\n{html.escape(body)}"
            await _tg_call("sendMessage", {"chat_id": tg, "text": text, "parse_mode": "HTML"})
    except Exception as e:
        logging.warning(f"notif push xato (user {user_id}): {e}")


# Sotuvchi Pro/Boost "sotib olish" tugmasini bosganda — unga yuboriladigan TASDIQ
# (real pul shu yerda ko'chmaydi; admin keyin to'lov bo'yicha qo'lda bog'lanadi).
_PURCHASE_PENDING = {
    "boost": {
        "uz": ("🚀 Boost so'rovingiz qabul qilindi",
               "To'lov masalasida admin tez orada siz bilan bog'lanadi."),
        "ru": ("🚀 Запрос на продвижение принят",
               "Администратор скоро свяжется с вами по оплате."),
    },
    "subscription": {
        "uz": ("⭐ Pro obuna so'rovingiz qabul qilindi",
               "To'lov masalasida admin tez orada siz bilan bog'lanadi."),
        "ru": ("⭐ Запрос на Pro-подписку принят",
               "Администратор скоро свяжется с вами по оплате."),
    },
}


async def _notify_purchaser_pending(user, kind, ref_id=None):
    """Pro/Boost sotib olish boshlanganda sotuvchiga 'so'rov qabul qilindi, admin to'lov
    bo'yicha bog'lanadi' tasdig'i (app banner + Telegram push). Xatoda jim o'tadi."""
    msg = _PURCHASE_PENDING.get(kind)
    if not msg:
        return
    try:
        lang = get_user_lang(user) or DEFAULT_LANG
        title, body = msg.get(lang) or msg["uz"]
        await _notify_user(user["id"], title, body, kind="payment", ref_id=ref_id)
    except Exception as e:
        logging.warning(f"purchaser pending notify xato (user {user.get('id')}): {e}")


def _admin_user_ids():
    """Barcha admin foydalanuvchi id'lari (role='admin' + asosiy ADMIN_ID)."""
    ids = set()
    try:
        for a in db.get_all_users(role="admin"):
            ids.add(a["id"])
    except Exception:
        pass
    if ADMIN_ID:
        try:
            au = db.get_user_by_telegram_id(ADMIN_ID)
            if au:
                ids.add(dict(au)["id"])
        except Exception:
            pass
    return ids


async def _notify_admins(title, body, kind="info", ref_id=None):
    for aid in _admin_user_ids():
        await _notify_user(aid, title, body, kind, ref_id)


def _notify_admins_db(kind, title, body="", ref_id=None):
    """Faqat app-banner barcha adminларга (Telegram push'i alohida yuborilgan nuqtalarда)."""
    for aid in _admin_user_ids():
        _notify_db(aid, kind, title, body, ref_id)


def _has_text_content(s):
    """Matnда kamida 2 ta harf/raqam bormi (faqat belgi/emoji bo'lsa False).
    isalnum lotin va kirill (o'zbekcha) harflarni ham, raqamlarni ham qabul qiladi."""
    return sum(1 for c in (s or "") if c.isalnum()) >= 2


# Murojaat sabablari (kalit + UZ/RU). Frontend ham shu kalitlardan foydalanadi.
SUPPORT_REASONS = {
    "order":      {"uz": "Buyurtma muammosi",        "ru": "Проблема с заказом"},
    "payment":    {"uz": "To'lov muammosi",          "ru": "Проблема с оплатой"},
    "seller":     {"uz": "Sotuvchi ustidan shikoyat", "ru": "Жалоба на продавца"},
    "product":    {"uz": "Mahsulot/e'lon muammosi",   "ru": "Проблема с товаром"},
    "account":    {"uz": "Akkaunt/sozlama",           "ru": "Аккаунт/настройки"},
    "suggestion": {"uz": "Taklif/fikr",               "ru": "Предложение/отзыв"},
    "other":      {"uz": "Boshqa",                    "ru": "Другое"},
}


class ContactIn(BaseModel):
    text: str
    reason: Optional[str] = None


@app.post("/api/contact-admin")
async def api_contact_admin(body: ContactIn, background: BackgroundTasks,
                            authorization: str = Header(None)):
    """Murojaat ochadi (support_thread + birinchi xabar) → adminlarga banner+push.
    Sabab MAJBURIY; matn faqat harf/raqamli bo'lishi shart (belgi/emoji-only rad)."""
    user = dict(_buyer_from_auth(authorization))
    _rate_limit("contact", user.get("id"), 5, 600)
    if body.reason not in SUPPORT_REASONS:
        raise HTTPException(status_code=400, detail="reason_required")
    text = (body.text or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="empty")
    if len(text) > 2000:
        raise HTTPException(status_code=400, detail="too_long")
    if not _has_text_content(text):
        raise HTTPException(status_code=400, detail="bad_text")
    reason = body.reason
    tid = db.create_support_thread(user["id"], reason, text)
    rlabel = SUPPORT_REASONS[reason]["uz"]
    await _notify_admins(f"📩 Yangi murojaat — {rlabel}",
                         f"{user.get('name') or ''}: {text[:120]}", kind="support", ref_id=tid)
    # AI birinchi-liniya: oddiy savolga (buyurtma holati kabi) darhol javob beradi;
    # yecholmasa jim qoladi — admin baribir xabardor (yuqorida notify bo'ldi).
    background.add_task(_support_ai_first_line, tid, user["id"])
    return {"ok": True, "thread_id": tid}


async def _support_ai_first_line(thread_id, user_id):
    """AI support: murojaatni foydalanuvchining REAL ma'lumotlari (buyurtma/qarz) bilan
    o'qib, ishonchli bo'lsa 'ai' nomidan javob yozadi + push. Har qanday xatoda JIM —
    admin oqimi buzilmaydi."""
    try:
        if not ai_assistant.is_enabled():
            return
        thread = db.get_support_thread(thread_id)
        if not thread or thread.get("status") == "closed":
            return
        user = db.get_user_by_id(user_id)
        if not user:
            return
        user = dict(user)
        lang = get_user_lang(user) or DEFAULT_LANG
        # Foydalanuvchi faktlari — AI faqat shulardan javob beradi (o'ylab topmaydi)
        orders = [dict(o) for o in (db.get_orders_by_buyer(user_id) or [])[:8]]
        facts = {
            "orders": [{"id": o.get("id"), "product": o.get("product_name"),
                        "status": o.get("status"), "total": o.get("total_price"),
                        "created": str(o.get("created_at") or "")[:16],
                        "delivery": o.get("delivery_type")} for o in orders],
            "open_debts": [{"shop": dict(d).get("shop_name") or dict(d).get("seller_name"),
                            "due": dict(d).get("total_due")}
                           for d in (db.get_buyer_open_debts(user_id) or [])],
        }
        rlabel = SUPPORT_REASONS.get(thread.get("reason") or "other", {}).get(
            "ru" if lang == "ru" else "uz", "")
        res = await ai_assistant.support_first_line(
            reason_label=rlabel, messages=db.get_support_messages(thread_id),
            user_facts=facts, lang=lang)
        if not res or not res.get("can_answer"):
            return
        db.add_support_message(thread_id, "ai", None, res["answer"])
        title = "🤖 AI-помощник ответил" if lang == "ru" else "🤖 AI yordamchi javob berdi"
        await _notify_user(user_id, title, res["answer"][:120],
                           kind="support", ref_id=thread_id)
    except Exception as e:
        logging.warning(f"support AI first-line xato (thread {thread_id}): {e}")


@app.get("/api/support/reasons")
def api_support_reasons(authorization: str = Header(None)):
    require_auth(authorization)
    return {"reasons": [{"key": k, "uz": v["uz"], "ru": v["ru"]} for k, v in SUPPORT_REASONS.items()]}


@app.get("/api/my/notifications")
def api_my_notifications(authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    return {"items": db.get_user_notifications(user["id"]),
            "unread": db.count_unread_notifications(user["id"])}


@app.post("/api/my/notifications/{notif_id}/read")
def api_notif_read(notif_id: int, authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    db.mark_notification_read(notif_id, user["id"])
    return {"ok": True, "unread": db.count_unread_notifications(user["id"])}


@app.post("/api/my/notifications/read-all")
def api_notif_read_all(authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    db.mark_all_notifications_read(user["id"])
    return {"ok": True}


# ============================================================
# ISTAK RO'YXATI (deal hunter) — "shu narsa kerak, chiqsa ayting"
# Moslikni bot skan-jobi tekshiradi (barcha manbalardan kelgan mahsulotlarni qamraydi)
# va topilganda push + app-banner yuboradi.
# ============================================================
class WishIn(BaseModel):
    text: str
    max_price: Optional[float] = None


@app.get("/api/my/wishes")
def api_my_wishes(authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    return {"items": db.get_user_wishes(user["id"])}


@app.post("/api/my/wishes")
def api_wish_create(body: WishIn, authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    text = (body.text or "").strip()
    if len(text) < 3:
        raise HTTPException(status_code=400, detail="too_short")
    if len(text) > 120:
        raise HTTPException(status_code=400, detail="too_long")
    if db.count_user_wishes(user["id"]) >= 5:
        raise HTTPException(status_code=409, detail="wish_limit")
    _rate_limit("wish", user.get("id"), 10, 3600)
    mp = None
    if body.max_price is not None and body.max_price > 0:
        mp = float(body.max_price)
    wid = db.create_wish(user["id"], text, mp)
    return {"ok": True, "id": wid}


@app.delete("/api/my/wishes/{wish_id}")
def api_wish_delete(wish_id: int, authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    if not db.delete_wish(wish_id, user["id"]):
        raise HTTPException(status_code=404, detail="not_found")
    return {"ok": True}


def _support_access_or_403(user, thread):
    """Murojaatga kirish: egasi yoki admin. is_admin qaytaradi."""
    if not thread:
        raise HTTPException(status_code=404, detail="not_found")
    is_admin = user.get("role") == "admin" or user.get("telegram_id") == ADMIN_ID
    if not is_admin and thread["user_id"] != user["id"]:
        raise HTTPException(status_code=403, detail="not_allowed")
    return is_admin


@app.get("/api/support/threads")
def api_support_threads(authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    is_admin = user.get("role") == "admin" or user.get("telegram_id") == ADMIN_ID
    threads = db.list_support_threads(None if is_admin else user["id"])
    return {"threads": threads, "is_admin": is_admin,
            "open_count": (db.count_open_support() if is_admin else None)}


@app.get("/api/support/thread/{thread_id}")
def api_support_thread(thread_id: int, authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    thread = db.get_support_thread(thread_id)
    _support_access_or_403(user, thread)
    return {"thread": thread, "messages": db.get_support_messages(thread_id),
            "reason_label": SUPPORT_REASONS.get(thread.get("reason") or "other", {})}


class SupportMsgIn(BaseModel):
    text: str


@app.post("/api/support/thread/{thread_id}/message")
async def api_support_message(thread_id: int, body: SupportMsgIn, background: BackgroundTasks,
                              authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    thread = db.get_support_thread(thread_id)
    is_admin = _support_access_or_403(user, thread)
    text = (body.text or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="empty")
    if len(text) > 2000:
        raise HTTPException(status_code=400, detail="too_long")
    if not _has_text_content(text):
        raise HTTPException(status_code=400, detail="bad_text")
    _rate_limit("support_msg", user["id"], 20, 600)
    db.add_support_message(thread_id, "admin" if is_admin else "user", user["id"], text)
    if is_admin:
        await _notify_user(thread["user_id"], "📩 Admin javob berdi", text[:120],
                           kind="support", ref_id=thread_id)
    else:
        await _notify_admins("📩 Murojaatга javob", f"{user.get('name') or ''}: {text[:120]}",
                             kind="support", ref_id=thread_id)
        # AI birinchi-liniya faqat ODAM-admin hali qo'shilmagan suhbatda davom etadi
        # (admin yozgan zahoti AI jim bo'ladi — aralashib ketmasin).
        msgs = db.get_support_messages(thread_id)
        if not any(m.get("sender_role") == "admin" for m in msgs):
            background.add_task(_support_ai_first_line, thread_id, user["id"])
    return {"ok": True}


@app.post("/api/support/thread/{thread_id}/close")
def api_support_close(thread_id: int, authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    thread = db.get_support_thread(thread_id)
    is_admin = _support_access_or_403(user, thread)
    if not is_admin:
        raise HTTPException(status_code=403, detail="admin_only")
    db.set_support_status(thread_id, "closed")
    return {"ok": True}


@app.post("/api/support/thread/{thread_id}/ai-reply")
async def api_support_ai_reply(thread_id: int, authorization: str = Header(None)):
    """ADMIN uchun: AI suhbatni O'QIB, foydalanuvchi savoliga mos javob TAKLIF qiladi
    (yubormaydi — admin ko'rib, tahrirlab yoki qayta generatsiya qilib o'zi yuboradi)."""
    user = dict(_buyer_from_auth(authorization))
    thread = db.get_support_thread(thread_id)
    is_admin = _support_access_or_403(user, thread)
    if not is_admin:
        raise HTTPException(status_code=403, detail="admin_only")
    if not ai_assistant.is_enabled():
        raise HTTPException(status_code=503, detail="ai_disabled")
    _rate_limit("support_ai", user["id"], 30, 600)
    msgs = db.get_support_messages(thread_id)
    rlabel = SUPPORT_REASONS.get(thread.get("reason") or "other", {}).get("uz", "")
    lang = get_user_lang(user) or DEFAULT_LANG
    reply = await ai_assistant.generate_support_reply(
        reason_label=rlabel, messages=msgs, lang=lang)
    if not reply:
        raise HTTPException(status_code=502, detail="ai_error")
    return {"reply": reply}


class BecomeSellerIn(BaseModel):
    shop_name: Optional[str] = None
    shop_address: Optional[str] = None
    shop_landmark: Optional[str] = None
    working_days: Optional[str] = None
    working_hours: Optional[str] = None
    phone: Optional[str] = None
    region_id: Optional[int] = None


@app.post("/api/become-seller")
async def api_become_seller(body: BecomeSellerIn = None, authorization: str = Header(None)):
    """Sotuvchi arizasi — do'kon ma'lumotlari bilan (bot become_seller_conv parite).
    Maydonlar user'ga saqlanadi → admin tasdiqlaganda create_shop o'shalardan quradi."""
    user = dict(_buyer_from_auth(authorization))
    _rate_limit("become_seller", user.get("id"), 3, 3600)
    if user.get("role") in ("seller", "admin") or user.get("is_approved"):
        raise HTTPException(status_code=409, detail="already_seller")
    existing = db.get_seller_request_by_user(user["id"])
    if existing and existing.get("status") == "pending":
        raise HTTPException(status_code=409, detail="already_pending")
    body = body or BecomeSellerIn()
    shop_name = (body.shop_name or "").strip()
    if not shop_name:
        raise HTTPException(status_code=400, detail="shop_name_required")
    # Do'kon maydonlarini user'ga saqlaymiz (faqat to'ldirilganlarini)
    fields = {"shop_name": shop_name}
    for attr in ("shop_address", "shop_landmark", "working_days", "working_hours"):
        v = (getattr(body, attr) or "").strip()
        if v:
            fields[attr] = v
    if body.region_id:
        fields["region_id"] = body.region_id
    ph = _normalize_phone(body.phone) if body.phone else None
    if ph and not user.get("phone_number"):
        fields["phone_number"] = ph
    try:
        db.update_user(user["id"], **fields)
    except Exception as e:
        logging.warning(f"become-seller update_user xato: {e}")
    db.create_seller_request(user["id"])
    try:
        if ADMIN_ID:
            loc = best_location_text(fields.get("shop_address"), fields.get("shop_landmark"))
            await _tg_call("sendMessage", {
                "chat_id": ADMIN_ID,
                "text": (f"🏪 Yangi sotuvchi arizasi (Mini App)\n"
                         f"👤 {html.escape(user.get('name') or '')}\n"
                         f"🏬 {html.escape(shop_name)}\n"
                         f"{('📍 ' + html.escape(loc)) if loc else ''}\n"
                         f"🆔 {user.get('telegram_id')}"),
                "parse_mode": "HTML"})
    except Exception as e:
        logging.warning(f"become-seller notify xato: {e}")
    return {"ok": True}


class JoinCodeIn(BaseModel):
    code: str


@app.post("/api/join-with-code")
async def api_join_with_code(body: JoinCodeIn, authorization: str = Header(None)):
    """#5 — taklif kodi bilan do'konga xodim sifatida qo'shilish (bot _handle_staff_deeplink
    pariteti, ro'yxatdan o'tgan foydalanuvchi shoxchasi). Egaga tasdiq xabari yuboriladi."""
    user = dict(_buyer_from_auth(authorization))
    _rate_limit("join_code", user.get("id"), 5, 600)
    code = (body.code or "").strip().upper()
    if not code:
        raise HTTPException(status_code=400, detail="empty")
    if user.get("role") == "admin":
        raise HTTPException(status_code=409, detail="admin_cannot_join")
    invite = db.get_invite_by_code(code)
    if not invite or dict(invite).get("is_used"):
        raise HTTPException(status_code=404, detail="invite_invalid")
    invite = dict(invite)
    shop = db.get_shop_by_id(invite["shop_id"])
    if not shop:
        raise HTTPException(status_code=404, detail="invite_invalid")
    shop = dict(shop)
    existing = db.get_staff_by_user(user["id"])
    if existing:
        existing = dict(existing)
        if existing.get("staff_role") == "owner":
            raise HTTPException(status_code=409, detail="owner_cannot_join")
        if existing.get("shop_id") == shop["id"]:
            raise HTTPException(status_code=409, detail="already_in_this_shop")
        # Eski do'kondan chiqarib, yangisiga o'tkazamiz + eski egaga xabar
        old_shop = db.get_shop_by_id(existing["shop_id"])
        db.remove_staff(existing["id"])
        try:
            if old_shop:
                old_owner = db.get_user_by_id(dict(old_shop)["owner_user_id"])
                if old_owner and dict(old_owner).get("telegram_id"):
                    await _tg_call("sendMessage", {"chat_id": dict(old_owner)["telegram_id"],
                                   "text": t(get_user_lang(dict(old_owner)), "staff_left_old_shop",
                                             name=html.escape(user.get("name") or "—")),
                                   "parse_mode": "HTML"})
        except Exception as e:
            logging.warning(f"join-with-code eski egaga xabar xato: {e}")
    staff_id = db.add_staff(shop["id"], user["id"], staff_role="staff",
                            department=invite.get("department"), is_active=0,
                            added_by=invite.get("created_by"))
    db.update_user(user["id"], role="seller", is_approved=1)
    db.mark_invite_used(code, user["id"])
    # Egaga yangi xodim haqida tasdiq xabari (tugmalar bot callback'ida ishlaydi)
    try:
        owner = db.get_user_by_id(shop["owner_user_id"])
        if owner and dict(owner).get("telegram_id"):
            owner = dict(owner)
            olang = get_user_lang(owner)
            await _tg_call("sendMessage", {
                "chat_id": owner["telegram_id"],
                "text": t(olang, "owner_new_staff_notify",
                          name=html.escape(user.get("name") or "—"),
                          phone=user.get("phone_number") or "—",
                          dept=html.escape(invite.get("department") or "—")),
                "parse_mode": "HTML",
                "reply_markup": {"inline_keyboard": [
                    [{"text": t(olang, "btn_staff_activate"), "callback_data": f"staff_toggle_{staff_id}"}],
                    [{"text": t(olang, "btn_staff_reject"), "callback_data": f"staff_reject_{staff_id}"}],
                    [{"text": t(olang, "btn_manage_staff"), "callback_data": f"staff_detail_{staff_id}"}]]}})
    except Exception as e:
        logging.warning(f"join-with-code egaga xabar xato: {e}")
    return {"ok": True, "shop_name": shop.get("name")}


@app.get("/api/staff/invite-info")
def api_staff_invite_info(code: str = "", authorization: str = Header(None)):
    """App start_param=staff_<kod> (yoki ?staff=) bilan kirilganda — tasdiq ekrani uchun
    do'kon nomi. require_auth: ro'yxatdan o'tmagan yangi user ham ko'ra oladi (initData
    imzosi yetarli). Bog'lashning o'zi /api/join-with-code orqali bo'ladi."""
    require_auth(authorization)
    code = (code or "").strip().upper()
    invite = db.get_invite_by_code(code) if code else None
    if not invite or dict(invite).get("is_used"):
        return {"valid": False}
    invite = dict(invite)
    shop = db.get_shop_by_id(invite["shop_id"])
    if not shop:
        return {"valid": False}
    return {"valid": True, "shop_name": dict(shop).get("name"),
            "department": invite.get("department")}


@app.get("/api/products/{product_id}")
def api_product_detail(product_id: int, authorization: str = Header(None)):
    require_auth(authorization)
    product = db.get_product_by_id(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="not found")
    product = dict(product)
    product["seller_is_pro"] = _is_pro(product.get("seller_id"))   # #18 Pro nishon (do'kon)
    product["images"] = db.get_product_images(product_id)  # file_id ro'yxati
    product["image_variants"] = db.get_product_image_variants(product_id)  # [{file_id,label,position}] — variant tanlash uchun
    # #19 — hudud yorlig'i (kanal e'loni pariteti: app'da ham "Viloyat → Tuman" ko'rinsin)
    product["region_label"] = db.get_region_label(product.get("seller_region_id")) or ""
    # #16 — joriy xaridor uchun sevimli holati (yurakcha tugmasi uchun)
    requester = None
    try:
        requester = db.get_user_by_telegram_id((require_auth(authorization).get("user") or {}).get("id"))
        product["is_favorite"] = db.is_favorite(dict(requester)["id"], product_id) if requester else False
    except Exception:
        product["is_favorite"] = False
    # #8 — maxfiy min_price faqat mahsulot EGASIga (tahrir formasi); boshqalardan olib tashlanadi
    req = dict(requester) if requester else {}
    is_owner_or_admin = req.get("id") == product.get("seller_id") or req.get("role") == "admin"
    # is_own — ko'rayotgan foydalanuvchi shu mahsulot sotuvchisimi (o'zi bilan savdolasholmaydi)
    product["is_own"] = bool(req.get("id") and req.get("id") == product.get("seller_id"))
    product["haggle_on"] = bool(product.get("min_price") and product.get("price")
                                and float(product["min_price"]) < float(product["price"]))
    if not is_owner_or_admin:
        product.pop("min_price", None)
    # #8 — joriy xaridor uchun amaldagi kelishilgan narx (checkout shuni ko'rsatadi)
    product["deal_price"] = None
    if req.get("id"):
        dp = db.get_active_haggle_price(req["id"], product_id)
        if dp and float(dp) <= float(product.get("price") or 0):
            product["deal_price"] = int(dp)
    try:
        product["attributes"] = _rows(db.get_product_attributes(product_id))
    except Exception as e:
        logging.warning(f"product attributes xato (pid {product_id}): {e}")
        product["attributes"] = []
    return product


@app.get("/api/products/{product_id}/ad-caption")
async def api_product_ad_caption(product_id: int, authorization: str = Header(None)):
    """Buyer mahsulot sahifasi uchun chiroyli AI reklama matni (kanaldagi bilan AYNAN bir xil).
    Saqlangani bo'lsa o'shani qaytaradi; bo'lmasa 1 marta yozib, bazaga saqlaydi (lazy).
    is_html=True bo'lsa matn HTML (App innerHTML bilan ko'rsatadi — barcha foydalanuvchi
    matni builder ichida html.escape qilingan, xavfsiz)."""
    require_auth(authorization)
    product = db.get_product_by_id(product_id)
    if not product:
        raise HTTPException(status_code=404, detail="not found")
    product = dict(product)
    cap = (product.get("ad_caption") or "").strip()
    pm = product.get("ad_caption_pm")
    if not cap:
        # Hali e'lon qilinmagan / eski mahsulot — AI ni KUTMASDAN tuzilgan emojili matn
        # (darhol chiqsin). AI takrorlanmas matn kanalga e'lon qilinganda saqlanadi.
        product["region_label"] = db.get_region_label(product.get("seller_region_id")) or ""
        cap, pm = await _build_ad_caption_web(product, "long", DEFAULT_LANG, use_ai=False)
    return {"caption": cap, "is_html": (pm == "HTML")}


# ============================================================
# BUYURTMA YARATISH (to'liq app ichida — sotuvchiga xabarni bot fon job'i yuboradi)
# ============================================================
ORDER_TTL_SECONDS = 1200         # oddiy buyurtma: sotuvchi tasdig'i 20 daqiqa
ORDER_TTL_OPTOM_SECONDS = 1800   # optom (pachka) buyurtma: 30 daqiqa (ko'proq o'ylab ko'rish)
_VALID_DELIVERY = {"delivery", "pickup"}
_VALID_PAYMENT = {"cash", "terminal", "p2p"}


def _order_deadline_pair(working_hours=None, working_days=None, ttl=ORDER_TTL_SECONDS,
                         work_schedule=None):
    """(auto_cancel_deadline, countdown_start) UTC juftligi. Do'kon yopiq bo'lsa sanoq
    keyingi ochilish vaqtidan boshlanadi — buyurtma ish boshlanguncha muzlaydi.
    main.py:_compute_order_deadline bilan bir xil mantiq (webapp tomoni). work_schedule
    berilsa kun-ma-kun jadval (qisqa/dam olish kunlari) hisobga olinadi."""
    now = datetime.now(timezone.utc)
    open_local = next_shop_open_datetime(working_hours, working_days, work_schedule=work_schedule)
    if open_local is not None:
        start = open_local.astimezone(timezone.utc)
        if start > now:
            return start + timedelta(seconds=ttl), start
    return now + timedelta(seconds=ttl), now


def _check_delivery_rules(products, total):
    """delivery_type='delivery' uchun himoya: (1) har bir mahsulot yetkaziladigan
    bo'lishi kerak (sotuvchi mahsulot darajasida belgilaydi), (2) jami summa
    do'konning minimal yetkazish summasiga (delivery_min_total) yetishi kerak —
    arzon buyurtmaga yetkazish sotuvchiga zarar qilmasligi uchun."""
    min_t = 0.0
    for prod in products:
        da = prod.get("delivery_available")
        if da is not None and int(da) == 0:
            raise HTTPException(status_code=400, detail="delivery_not_available")
        try:
            min_t = max(min_t, float(prod.get("delivery_min_total") or 0))
        except (ValueError, TypeError):
            pass
    if min_t > 0 and total < min_t:
        raise HTTPException(status_code=400, detail=f"delivery_min_{int(min_t)}")


class OrderIn(BaseModel):
    product_id: int
    quantity: int = 1
    delivery_type: str = "pickup"
    payment_method: str = "cash"
    address: Optional[str] = None
    lat: Optional[float] = None
    lon: Optional[float] = None


@app.post("/api/order")
def api_create_order(order: OrderIn, authorization: str = Header(None)):
    auth = require_auth(authorization)
    tg_id = (auth.get("user") or {}).get("id")
    if not tg_id:
        raise HTTPException(status_code=401, detail="no user")
    buyer = db.get_user_by_telegram_id(tg_id)
    if not buyer:
        raise HTTPException(status_code=403, detail="not_registered")
    _rate_limit("order", buyer["id"], 15, 60)

    if order.quantity < 1 or order.quantity > 999:
        raise HTTPException(status_code=400, detail="bad_quantity")
    if order.delivery_type not in _VALID_DELIVERY:
        raise HTTPException(status_code=400, detail="bad_delivery_type")
    if order.payment_method not in _VALID_PAYMENT:
        raise HTTPException(status_code=400, detail="bad_payment_method")

    product = db.get_product_by_id(order.product_id)
    if not product or not product.get("in_stock"):
        raise HTTPException(status_code=404, detail="product_unavailable")
    if buyer["id"] == product["seller_id"]:
        raise HTTPException(status_code=400, detail="own_product")
    stock = product.get("stock_count")
    if stock is not None and order.quantity > stock:
        raise HTTPException(status_code=409, detail=f"only_{stock}_available")

    # Optom (ulgurji) narx: buyurtma soni minimumga yetsa, dona narxi optomga tushadi.
    unit_price = effective_unit_price(product, order.quantity)
    # #8 — AI bilan savdolashib kelishilgan narx bo'lsa, shuni qo'llaymiz (floor'dan
    # past bo'lishi mumkin emas — haggle endpoint kafolatlaydi). Optom/listed'dan oshmaydi;
    # optom va savdolashishdan ENG PAST narx xaridorga beriladi.
    deal = db.get_active_haggle_price(buyer["id"], product["id"])
    if deal and 0 < float(deal) <= unit_price:
        unit_price = float(deal)
    total = order.quantity * unit_price
    if order.delivery_type == "delivery":
        _check_delivery_rules([product], total)
        address = (order.address or "").strip() or None
        lat, lon = order.lat, order.lon
        if lat is None or lon is None:   # yetkazib berishda joylashuv MAJBURIY (kuryer topishi uchun)
            raise HTTPException(status_code=400, detail="location_required")
    else:
        address, lat, lon = None, None, None

    order_id = db.create_order(
        buyer_id=buyer["id"], seller_id=product["seller_id"],
        product_id=product["id"], quantity=order.quantity, total_price=total,
        delivery_address=address, buyer_lat=lat, buyer_lon=lon,
        payment_method=order.payment_method, delivery_type=order.delivery_type,
    )
    # Avto-bekor muddati (bot order_confirm'i bilan bir xil) + bot fon job'i xabar yuboradi.
    # Do'kon yopiq bo'lsa sanoq ish boshlanguncha muzlaydi.
    deadline, cd_start = _order_deadline_pair(
        product.get("working_hours"), product.get("working_days"),
        work_schedule=product.get("work_schedule"))
    db.set_order_deadline(order_id, deadline, countdown_start=cd_start)
    db.mark_order_notify_pending(order_id)
    # App-banner: sotuvchiga yangi buyurtma (bot Telegram push'ini fon-job yuboradi)
    pname = product.get("name") or ""
    _notify_db(product["seller_id"], "order", ("🛒 Yangi buyurtma", "🛒 Новый заказ"),
               (f"{pname} · {order.quantity} dona", f"{pname} · {order.quantity} шт"), ref_id=order_id)
    return {"ok": True, "order_id": order_id, "total": total}


class CartItemIn(BaseModel):
    product_id: int
    quantity: int = 1


class CartCheckoutIn(BaseModel):
    seller_id: int
    items: List[CartItemIn]
    delivery_type: str = "pickup"
    payment_method: str = "cash"
    address: Optional[str] = None
    lat: Optional[float] = None
    lon: Optional[float] = None


@app.post("/api/cart/checkout")
def api_cart_checkout(co: CartCheckoutIn, authorization: str = Header(None)):
    """Savat (bitta sotuvchi, ko'p mahsulot) -> guruh buyurtma. Bot fon job'i sotuvchiga
    BITTA guruh bildirishnomasi yuboradi."""
    buyer = dict(_buyer_from_auth(authorization))
    _rate_limit("cart", buyer["id"], 10, 60)
    if not co.items:
        raise HTTPException(status_code=400, detail="empty_cart")
    if co.delivery_type not in _VALID_DELIVERY:
        raise HTTPException(status_code=400, detail="bad_delivery_type")
    if co.payment_method not in _VALID_PAYMENT:
        raise HTTPException(status_code=400, detail="bad_payment_method")
    if co.delivery_type == "delivery":
        address = (co.address or "").strip() or None
        lat, lon = co.lat, co.lon
        if lat is None or lon is None:   # yetkazib berishda joylashuv MAJBURIY
            raise HTTPException(status_code=400, detail="location_required")
    else:
        address, lat, lon = None, None, None

    pending, grand = [], 0.0
    for it in co.items:
        if it.quantity < 1:
            continue
        product = db.get_product_by_id(it.product_id)
        if not product or not product.get("in_stock") or product.get("status") == "deleted":
            continue
        if product.get("seller_id") != co.seller_id:   # savat = bitta sotuvchi
            continue
        if buyer["id"] == product["seller_id"]:
            continue
        qty = it.quantity
        stock = product.get("stock_count")
        if stock is not None:
            qty = min(qty, int(stock))
        if qty <= 0:
            continue
        line = qty * effective_unit_price(product, qty)   # optom narx (qator soni minimumga yetsa)
        grand += line
        pending.append((product, qty, line))

    if not pending:
        raise HTTPException(status_code=409, detail="nothing_available")
    # Yetkazish qoidalari BUTUN savat bo'yicha (yaratishdan OLDIN — yarim guruh qolmasin)
    if co.delivery_type == "delivery":
        _check_delivery_rules([pr for pr, _, _ in pending], grand)

    created = []
    for product, qty, line in pending:
        oid = db.create_order(
            buyer_id=buyer["id"], seller_id=co.seller_id, product_id=product["id"],
            quantity=qty, total_price=line, delivery_address=address,
            buyer_lat=lat, buyer_lon=lon, payment_method=co.payment_method,
            delivery_type=co.delivery_type,
        )
        created.append(oid)
    group_id = str(created[0])
    db.set_orders_group(created, group_id)
    _wh = pending[0][0] if pending else {}
    deadline, cd_start = _order_deadline_pair(_wh.get("working_hours"), _wh.get("working_days"),
                                              work_schedule=_wh.get("work_schedule"))
    db.set_group_deadline(group_id, deadline, countdown_start=cd_start)
    for oid in created:
        db.mark_order_notify_pending(oid)
    return {"ok": True, "group_id": group_id, "count": len(created), "total": grand}


def _min_order_qty(product):
    """Variant-buyurtma uchun JAMI minimal son: FAQAT sotuvchi aniq qo'ygan min_order_qty (>1).
    Optom (ulgurji) zinalar minimal MAJBURLAMAYDI — ular faqat chegirma: kam olgan dona narxida,
    hatto 1 dona ham olishi mumkin; zina mini'ga yetganda narx optomga tushadi."""
    try:
        moq = int(product.get("min_order_qty") or 0)
    except (ValueError, TypeError):
        moq = 0
    return moq if moq > 1 else 1


def _split_opts(s):
    """Atribut qiymatini variantlarga ajratadi. Vergul/nuqtali vergul/slash/yangi qator/'|'/'·'
    bo'lsa shular bo'yicha, aks holda bo'sh joy bo'yicha (AI ko'pincha "36 37 38" deb saqlaydi).
    Frontend `splitOpts` bilan AYNAN bir xil — validatsiya rad etmasligi uchun."""
    s = str(s or "").strip()
    if not s:
        return []
    parts = re.split(r"[,;/\n·|]+", s) if re.search(r"[,;/\n·|]", s) else re.split(r"\s+", s)
    return [p.strip() for p in parts if p.strip()]


def _product_attr_options(product_id, key):
    """Mahsulot atributlaridagi `key` ro'yxati (vergul yoki bo'sh joy bilan), aks holda []."""
    try:
        attrs = db.get_product_attributes(product_id) or []
    except Exception:
        return []
    for a in attrs:
        a = dict(a)
        if a.get("attr_key") == key and (a.get("attr_value") or "").strip():
            return _split_opts(a["attr_value"])
    return []


def _product_size_options(product_id):
    """Mahsulot atributlaridagi razmer ro'yxati."""
    return _product_attr_options(product_id, "size")


def _product_color_options(product_id):
    """Mahsulot atributlaridagi rang ro'yxati."""
    return _product_attr_options(product_id, "color")


class VariantLineIn(BaseModel):
    label: Optional[str] = None    # variant nomi yoki "#N"
    size: Optional[str] = None
    color: Optional[str] = None
    qty: int = 0


class VariantOrderIn(BaseModel):
    product_id: int
    lines: List[VariantLineIn]
    delivery_type: str = "pickup"
    payment_method: str = "cash"
    address: Optional[str] = None
    lat: Optional[float] = None
    lon: Optional[float] = None


@app.post("/api/order/variants")
def api_variant_order(vo: VariantOrderIn, authorization: str = Header(None)):
    """Bitta listing ichidagi bir nechta variant (rasm/hil) bo'yicha BITTA guruh buyurtma.
    Optom narx JAMI songa qarab qo'llanadi; JAMI son minimumdan kam bo'lsa rad etiladi.
    Sotuvchiga bitta guruh bildirishnomasi (variant taqsimoti bilan) ketadi (bot fon job)."""
    buyer = dict(_buyer_from_auth(authorization))
    _rate_limit("order", buyer["id"], 15, 60)
    if vo.delivery_type not in _VALID_DELIVERY:
        raise HTTPException(status_code=400, detail="bad_delivery_type")
    if vo.payment_method not in _VALID_PAYMENT:
        raise HTTPException(status_code=400, detail="bad_payment_method")

    product = db.get_product_by_id(vo.product_id)
    if not product or not product.get("in_stock") or product.get("status") == "deleted":
        raise HTTPException(status_code=404, detail="product_unavailable")
    product = dict(product)
    if buyer["id"] == product["seller_id"]:
        raise HTTPException(status_code=400, detail="own_product")

    # Faqat son > 0 bo'lgan qatorlar
    lines = [ln for ln in (vo.lines or []) if (ln.qty or 0) > 0]
    if not lines:
        raise HTTPException(status_code=400, detail="empty_order")

    total_qty = sum(int(ln.qty) for ln in lines)
    if total_qty > 9999:
        raise HTTPException(status_code=400, detail="bad_quantity")

    # Razmer ro'yxati bor mahsulotda har qatorda razmer MAJBURIY
    size_opts = _product_size_options(vo.product_id)
    if size_opts:
        for ln in lines:
            if not (ln.size or "").strip():
                raise HTTPException(status_code=400, detail="size_required")
            if ln.size.strip() not in size_opts:
                raise HTTPException(status_code=400, detail="bad_size")

    # Rang ro'yxati bor mahsulotda har qatorda rang MAJBURIY
    color_opts = _product_color_options(vo.product_id)
    if color_opts:
        for ln in lines:
            if not (ln.color or "").strip():
                raise HTTPException(status_code=400, detail="color_required")
            if ln.color.strip() not in color_opts:
                raise HTTPException(status_code=400, detail="bad_color")

    # JAMI minimal son (optom)
    min_qty = _min_order_qty(product)
    if total_qty < min_qty:
        raise HTTPException(status_code=400, detail=f"below_min_{min_qty}")

    # Ombor (jami)
    stock = product.get("stock_count")
    if stock is not None and total_qty > int(stock):
        raise HTTPException(status_code=409, detail=f"only_{int(stock)}_available")

    if vo.delivery_type == "delivery":
        address = (vo.address or "").strip() or None
        lat, lon = vo.lat, vo.lon
        if lat is None or lon is None:
            raise HTTPException(status_code=400, detail="location_required")
    else:
        address, lat, lon = None, None, None

    # Optom narx JAMI songa qarab (har variant qatoriga bir xil dona narx)
    unit_price = effective_unit_price(product, total_qty)
    if vo.delivery_type == "delivery":
        _check_delivery_rules([product], total_qty * unit_price)
    created, grand = [], 0.0
    for idx, ln in enumerate(lines):
        qty = int(ln.qty)
        line_total = qty * unit_price
        grand += line_total
        label = (ln.label or "").strip() or f"#{idx + 1}"
        size = (ln.size or "").strip() or None
        color = (ln.color or "").strip() or None
        oid = db.create_order(
            buyer_id=buyer["id"], seller_id=product["seller_id"], product_id=vo.product_id,
            quantity=qty, total_price=line_total, delivery_address=address,
            buyer_lat=lat, buyer_lon=lon, payment_method=vo.payment_method,
            delivery_type=vo.delivery_type, variant_label=label, variant_size=size,
            variant_color=color,
        )
        created.append(oid)

    if not created:
        raise HTTPException(status_code=409, detail="nothing_available")
    group_id = str(created[0])
    db.set_orders_group(created, group_id)
    _is_optom = (product.get("sale_mode") == "optom")
    _ttl = ORDER_TTL_OPTOM_SECONDS if _is_optom else ORDER_TTL_SECONDS  # optom = 30 daqiqa
    deadline, cd_start = _order_deadline_pair(
        product.get("working_hours"), product.get("working_days"), ttl=_ttl,
        work_schedule=product.get("work_schedule"))
    db.set_group_deadline(group_id, deadline, countdown_start=cd_start)
    for oid in created:
        db.mark_order_notify_pending(oid)
    # App-banner: sotuvchiga yangi buyurtma (bot Telegram push'ini fon-job yuboradi)
    pname = product.get("name") or ""
    _u_uz = "pachka" if _is_optom else "dona"
    _u_ru = "упак." if _is_optom else "шт"
    _notify_db(product["seller_id"], "order", ("🛒 Yangi buyurtma", "🛒 Новый заказ"),
               (f"{pname} · {total_qty} {_u_uz}", f"{pname} · {total_qty} {_u_ru}"), ref_id=int(group_id))
    return {"ok": True, "group_id": group_id, "count": len(created),
            "total": grand, "unit_price": unit_price, "total_qty": total_qty}


import time as _time
_RATE = {}  # (bucket, user_id) -> [timestamps]


def _rate_limit(bucket, user_id, max_calls, window):
    """Oddiy xotira-ichi throttle (jarayon bo'yicha). Limitdan oshsa 429.
    Spam/cost himoyasi: AI (DeepSeek puli), admin/seller spam, buyurtma toshqini."""
    now = _time.time()
    key = (bucket, user_id)
    arr = [t for t in _RATE.get(key, ()) if now - t < window]
    if len(arr) >= max_calls:
        # Limit oshdi = flood/spam urinishi → admin statistikasi uchun yozamiz.
        # user_id DB id YOKI telegram_id bo'lishi mumkin (chaqiruv joyiga qarab) —
        # increment_spam_count ikkalasini ham qo'llab-quvvatlaydi.
        try:
            db.increment_spam_count(user_id)
        except Exception:
            pass
        raise HTTPException(status_code=429, detail="too_many_requests")
    arr.append(now)
    _RATE[key] = arr
    # vaqti-vaqti bilan eski kalitlarni tozalaymiz (xotira o'smasin)
    if len(_RATE) > 5000:
        for k in [k for k, v in list(_RATE.items()) if not any(now - t < window for t in v)]:
            _RATE.pop(k, None)


def _buyer_from_auth(authorization):
    """initData'dan xaridorni (DB user) qaytaradi yoki 401/403 ko'taradi."""
    auth = require_auth(authorization)
    tg_id = (auth.get("user") or {}).get("id")
    if not tg_id:
        raise HTTPException(status_code=401, detail="no_user")
    buyer = db.get_user_by_telegram_id(tg_id)
    if not buyer:
        raise HTTPException(status_code=403, detail="not_registered")
    # Faollik kuzatuvi (throttled): har App so'rovi foydalanuvchini "faol" deb belgilaydi.
    try:
        db.touch_user_activity(user_id=buyer["id"])
    except Exception:
        pass
    return buyer


import re as _re


def _normalize_phone(raw):
    """Telefonni standartlashtirish (bot normalize_phone bilan bir xil mantiq).
    '901234567' → '+998901234567'; yaroqsiz bo'lsa None."""
    if not raw:
        return None
    digits = _re.sub(r"\D", "", str(raw))
    if len(digits) == 9:
        digits = "998" + digits
    if len(digits) == 12 and digits.startswith("998"):
        return "+" + digits
    if 10 <= len(digits) <= 15:
        return "+" + digits
    return None


class RegisterIn(BaseModel):
    name: str
    phone: str
    language: str = "uz"
    ref: Optional[str] = None   # referral kodi (/start REF... → ?ref=... → App)
    region_id: Optional[int] = None   # hudud (viloyat/tuman) — ro'yxatdan o'tishda tanlanadi


@app.post("/api/register")
async def api_register(body: RegisterIn, authorization: str = Header(None)):
    """Mini App ichida yangi xaridorni ro'yxatdan o'tkazadi (bot FSM o'rniga).
    initData imzosi tekshiriladi — telegram_id ishonchli manbadan olinadi."""
    auth = require_auth(authorization)
    tg_user = auth.get("user") or {}
    tg_id = tg_user.get("id")
    if not tg_id:
        raise HTTPException(status_code=401, detail="no_user")
    _rate_limit("register", tg_id, 5, 600)
    # Idempotent — allaqachon ro'yxatda bo'lsa, qayta yaratmaymiz
    existing = db.get_user_by_telegram_id(tg_id)
    if existing:
        return {"ok": True, "already": True}
    name = (body.name or "").strip()
    if not name or len(name) > 60:
        raise HTTPException(status_code=400, detail="bad_name")
    phone = _normalize_phone(body.phone)
    if not phone:
        raise HTTPException(status_code=400, detail="bad_phone")
    lang = body.language if body.language in ("uz", "ru") else "uz"
    uid = db.create_user(telegram_id=tg_id, phone_number=phone, name=name, role="buyer")
    fields = {"language": lang}
    uname = tg_user.get("username")
    if uname:
        fields["telegram_username"] = uname
    # Hudud — ro'yxatdan o'tishda tanlanadi (profil keyin "to'ldiring" deb so'ramasin).
    # Yaroqsiz id kelsa jim o'tamiz (ro'yxatdan o'tish buzilmasin).
    if body.region_id:
        try:
            if db.get_region_label(int(body.region_id)):
                fields["region_id"] = int(body.region_id)
        except Exception:
            pass
    db.update_user(uid, **fields)

    # Referral — /start REF... orqali kelgan bo'lsa (bot complete_registration parite)
    ref_code = (body.ref or "").strip()
    if ref_code:
        try:
            referrer = db.get_user_by_referral_code(ref_code)
            if referrer and referrer["id"] != uid:
                db.update_user(uid, referred_by=referrer["id"])
                db.increment_referral_count(referrer["id"])
                if referrer.get("telegram_id"):
                    rlang = get_user_lang(referrer) or DEFAULT_LANG
                    await _tg_call("sendMessage", {
                        "chat_id": referrer["telegram_id"],
                        "text": t(rlang, "new_referral", name=html.escape(name)),
                        "parse_mode": "HTML"})
        except Exception as e:
            logging.warning(f"register referral xato: {e}")
    # Adminga yangi foydalanuvchi haqida xabar (xato yutiladi)
    try:
        if ADMIN_ID:
            await _tg_call("sendMessage", {
                "chat_id": ADMIN_ID,
                "text": (f"👤 <b>Yangi foydalanuvchi (Mini App)</b>\n"
                         f"Ism: {html.escape(name)}\nTelefon: {html.escape(phone)}\n"
                         f"🆔 {tg_id}"),
                "parse_mode": "HTML"})
    except Exception as e:
        logging.warning(f"register notify xato: {e}")
    return {"ok": True, "id": uid}


@app.get("/api/me")
def api_me(authorization: str = Header(None)):
    b = dict(_buyer_from_auth(authorization))
    # #9 — sotuvchi rejimi tugmasida do'kon nomini ko'rsatish uchun. Egada o'z
    # shop_name; xodimda — tegishli do'kon EGASInning shop_name'i. is_owner bilan
    # frontend "ega"/"xodim" farqini ham biladi (#5/#10).
    shop_name = b.get("shop_name")
    own_shop = db.get_shop_by_owner(b["id"])
    is_owner = bool(own_shop)
    shop_for = db.get_shop_for_user(b["id"])
    if not shop_name and shop_for:
        owner = db.get_user_by_id(dict(shop_for)["owner_user_id"])
        shop_name = dict(owner).get("shop_name") if owner else None
    # Multivendor: xodimlik holati + to'lov rejimi + xodim boshqarish huquqi (frontend uchun)
    staff = db.get_staff_by_user(b["id"])
    staff_d = dict(staff) if staff else None
    is_staff = bool(staff_d and staff_d.get("staff_role") != "owner")
    payment_mode = (dict(shop_for).get("payment_mode") if shop_for else None) or "shop"
    can_manage_staff = bool(is_owner or (staff_d and staff_d.get("staff_role") == "manager"
                                         and staff_d.get("perm_add_staff")))
    # Do'kon EGAsining to'ldirilmagan muhim maydonlari — frontend "joniga tegadigan"
    # eslatma bannerini shu ro'yxat asosida ko'rsatadi. To'liq bo'lsa bo'sh ([]).
    profile_gaps = []
    if is_owner:
        def _empty(v):
            return v is None or (isinstance(v, str) and not v.strip())
        if _empty(b.get("shop_name")):
            profile_gaps.append("shop_name")
        if _empty(b.get("shop_address")):
            profile_gaps.append("shop_address")
        if not b.get("region_id"):
            profile_gaps.append("region")
        if _empty(b.get("phone_number")):
            profile_gaps.append("phone")
        if b.get("shop_lat") is None or b.get("shop_lon") is None:
            profile_gaps.append("location")
        if _empty(b.get("telegram_username")):
            profile_gaps.append("username")
    return {
        "id": b.get("id"), "name": b.get("name"), "phone": b.get("phone_number"),
        "username": b.get("telegram_username"), "role": b.get("role"),
        "region_id": b.get("region_id"),   # hudud banneri: yo'q bo'lsa frontend so'raydi
        "language": b.get("language"), "created_at": b.get("created_at"),
        "is_approved": b.get("is_approved"), "shop_name": shop_name, "is_owner": is_owner,
        "is_staff": is_staff, "payment_mode": payment_mode, "can_manage_staff": can_manage_staff,
        "staff_role": (staff_d.get("staff_role") if staff_d else None),
        "is_courier": bool(staff_d and staff_d.get("staff_role") == "courier"),
        "referral_code": b.get("referral_code"), "referral_count": b.get("referral_count"),
        # #22 — monetizatsiya bayroqlari (default hammasi o'chiq); frontend kelajakda
        # tugmalarni shunga qarab ko'rsatadi/yashiradi.
        "monetization": monetization_public(),
        # #18 — Pro-obuna holati + platformaga komissiya qarzi (egasi bo'yicha; sotuvchi
        # bo'lmasa tabiiy ravishda 0 qaytadi)
        "pro": _pro_status(b),
        "commission_owed": db.get_commission_owed_by_seller(_owner_id(b)),
        "commission_paid": db.get_commission_paid_by_seller(_owner_id(b)),
        "profile_gaps": profile_gaps,
    }


# ============================================================
# #16 SODIQLIK — sodiqlik ballari + nishon darajasi. Mavjud ma'lumotdan hisoblanadi
# (xarid + referal), schema/hook YO'Q. Naqd cashback (yechib olish) — to'lov
# integratsiyasiga bog'liq, keyingi bosqich. Ball formulasi:
#   ball = xarajat/10000 + yetkazilgan_buyurtma*5 + referal*50
# ============================================================
LOYALTY_TIERS = [   # (min_ball, kalit, emoji)
    (0,    "bronze",  "🥉"),
    (200,  "silver",  "🥈"),
    (1000, "gold",    "🥇"),
    (5000, "diamond", "💎"),
]


def _loyalty_for(user):
    uid = user["id"]
    tot = db.get_buyer_order_totals(uid)
    spent = float(tot.get("spent") or 0)
    delivered = int(tot.get("delivered_orders") or 0)
    refs = int(user.get("referral_count") or 0)
    points = int(spent // 10000) + delivered * 5 + refs * 50
    # joriy va keyingi daraja
    cur = LOYALTY_TIERS[0]
    nxt = None
    for i, tier in enumerate(LOYALTY_TIERS):
        if points >= tier[0]:
            cur = tier
            nxt = LOYALTY_TIERS[i + 1] if i + 1 < len(LOYALTY_TIERS) else None
    progress = None
    if nxt:
        span = nxt[0] - cur[0]
        progress = round((points - cur[0]) / span * 100) if span > 0 else 100
    return {
        "points": points, "tier": cur[1], "tier_emoji": cur[2],
        "next_tier": (nxt[1] if nxt else None),
        "next_at": (nxt[0] if nxt else None),
        "to_next": (nxt[0] - points if nxt else 0),
        "progress": progress,
        "breakdown": {"spent": spent, "delivered_orders": delivered, "referrals": refs},
    }


@app.get("/api/me/loyalty")
def api_me_loyalty(authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    return _loyalty_for(user)


@app.get("/api/me/pending")
def api_me_pending(authorization: str = Header(None)):
    """#23 — foydalanuvchining e'tibor talab qiladigan ("kutilayotgan") ishlari.
    Chiqishdan oldin eslatish va nav-badge uchun. Panellardagi AYNAN bir xil
    DB chaqiruvlaridan sanaladi — ko'rsatkichlar bir-biriga mos keladi."""
    user = dict(_buyer_from_auth(authorization))
    uid = user["id"]
    items = []  # [{key, n, tab}]

    # Xaridor: sotuvchi bekor qilishni so'ragan — javob berish kerak
    n = sum(1 for o in db.get_buyer_orders_list(uid)
            if dict(o).get("cancel_state") == "requested" and dict(o).get("cancel_by") == "seller")
    if n:
        items.append({"key": "buyer_cancel", "n": n, "tab": "orders"})

    # Sotuvchi (yoki ega/admin)
    if user.get("role") in ("seller", "admin") or user.get("is_approved"):
        so = [dict(o) for o in db.get_seller_orders_list(uid)]
        confirm = sum(1 for o in so if o.get("status") == "pending")
        if confirm:
            items.append({"key": "seller_confirm", "n": confirm, "tab": "seller"})
        crespond = sum(1 for o in so if o.get("cancel_state") == "requested" and o.get("cancel_by") == "buyer")
        if crespond:
            items.append({"key": "seller_cancel", "n": crespond, "tab": "seller"})
        try:
            pend = db.get_seller_products_by_status(uid, "pending_owner")
            if pend:
                items.append({"key": "seller_approve", "n": len(pend), "tab": "seller"})
        except Exception:
            pass

    # Admin
    if user.get("role") == "admin" or user.get("telegram_id") == ADMIN_ID:
        reqs = db.get_pending_seller_requests()
        if reqs:
            items.append({"key": "admin_requests", "n": len(reqs), "tab": "admin"})
        try:
            disp = db.get_disputed_orders()
            if disp:
                items.append({"key": "admin_disputes", "n": len(disp), "tab": "admin"})
        except Exception:
            pass
        # Admin tasdig'ini kutayotgan to'lovlar (boost/Pro dev-confirm) — tasdiqlanguncha
        # statik badge bo'lib qoladi (ochilganda yo'qolmaydi).
        try:
            pays = db.get_payments_admin(state="pending", limit=100)
            if pays:
                items.append({"key": "admin_payments", "n": len(pays), "tab": "admin"})
        except Exception:
            pass

    return {"total": sum(i["n"] for i in items), "items": items}


@app.get("/api/my/orders")
def api_my_orders(authorization: str = Header(None)):
    buyer = _buyer_from_auth(authorization)
    return _rows(db.get_buyer_orders_list(buyer["id"]))


@app.get("/api/my/debts")
def api_my_debts(authorization: str = Header(None)):
    buyer = _buyer_from_auth(authorization)
    return _rows(db.get_buyer_open_debts(buyer["id"]))


@app.get("/api/my/reviews")
def api_my_reviews(authorization: str = Header(None)):
    """#4 — xaridor o'zi qoldirgan sharhlar (bot buyer_reviews pariteti)."""
    buyer = _buyer_from_auth(authorization)
    return _rows(db.get_reviews_by_buyer(buyer["id"], 20))


class ReviewIn(BaseModel):
    seller_rating: int
    product_rating: Optional[int] = None
    comment: Optional[str] = None


@app.post("/api/order/{order_id}/review")
async def api_review(order_id: int, body: ReviewIn, authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    order = db.get_order_by_id(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="not_found")
    if order.get("buyer_id") != user["id"]:
        raise HTTPException(status_code=403, detail="not_your_order")
    if order.get("status") != "delivered":
        raise HTTPException(status_code=409, detail="not_delivered")
    if db.order_review_exists(order_id, user["id"]):
        raise HTTPException(status_code=409, detail="already_reviewed")
    sr = body.seller_rating
    if not isinstance(sr, int) or not (1 <= sr <= 5):
        raise HTTPException(status_code=400, detail="bad_rating")
    pr = body.product_rating
    if pr is not None and not (1 <= pr <= 5):
        raise HTTPException(status_code=400, detail="bad_product_rating")
    comment = (body.comment or "").strip() or None
    db.create_review(order_id, order["seller_id"], user["id"], sr, comment,
                     order.get("product_id"), pr)
    stars = "⭐" * sr
    pname = order.get("product_name") or ""
    _notify_db(order["seller_id"], "review", ("⭐ Yangi baho", "⭐ Новая оценка"),
               (f"{stars} {pname}" + (f" · {comment}" if comment else ""),
                f"{stars} {pname}" + (f" · {comment}" if comment else "")), ref_id=order_id)
    try:
        if order.get("seller_tg"):
            txt = f"{stars} {fmt_order_id(order_id)} — yangi baho"
            if comment:
                txt += f"\n💬 {html.escape(comment)}"
            await _tg_call("sendMessage", {"chat_id": order["seller_tg"], "text": txt, "parse_mode": "HTML"})
    except Exception as e:
        logging.warning(f"review notify xato (order {order_id}): {e}")
    return {"ok": True}


def _norm_cancel_reason(raw):
    """Mijozdan kelgan bekor sababini normallashtiradi: 'code:<kod>'/'text:<matn>' —
    bot va app cancelReasonText/cancel_reason_display bilan AYNAN bir xil format.
    Prefiks bo'lmasa erkin matn deb 'text:' qo'shiladi. Bo'sh bo'lsa None."""
    s = (raw or "").strip()
    if not s:
        return None
    if s.startswith("code:") or s.startswith("text:"):
        return s[:520]
    return ("text:" + s)[:520]


def _cancel_reason_display(reason, lang):
    """Saqlangan 'code:<kod>'/'text:<matn>' sababini o'qiladigan, til-aware matnga
    aylantiradi (bot cancel_reason_display pariteti). Bo'sh/auto bo'lsa '—'."""
    s = (reason or "").strip()
    if not s or s in ("—", "auto_timeout"):
        return "—"
    if s.startswith("code:"):
        return t(lang, "crsn_" + s[5:])
    if s.startswith("text:"):
        return s[5:]
    return s


class BuyerCancelIn(BaseModel):
    reason: Optional[str] = None


@app.post("/api/order/{order_id}/cancel-suggest")
async def api_cancel_suggest(order_id: int, authorization: str = Header(None)):
    """AI bekor qilish sabablarini taklif qiladi (kontekstli). Foydalanuvchining shu
    buyurtmadagi o'rni (xaridor/sotuvchi) va mahsulot/holatga qarab 4 ta qisqa sabab.
    AI o'chiq/xato bo'lsa bo'sh ro'yxat — app/bot oldindan tayyor presetlarga tushadi."""
    user = dict(_buyer_from_auth(authorization))
    order = db.get_order_by_id(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="not_found")
    if user.get("id") == order.get("buyer_id"):
        party = "buyer"
    elif order.get("seller_id") == _owner_id(user) or user.get("role") == "admin":
        party = "seller"
    else:
        raise HTTPException(status_code=403, detail="not_your_order")
    _rate_limit("cancel_suggest", user["id"], 12, 60)
    lang = get_user_lang(user) or DEFAULT_LANG
    reasons = await ai_assistant.suggest_cancel_reasons(
        party=party, product_name=order.get("product_name") or "",
        status=order.get("status") or "", lang=lang)
    return {"reasons": reasons}


@app.post("/api/order/{order_id}/cancel")
async def api_buyer_cancel(order_id: int, body: BuyerCancelIn = None, authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    order = db.get_order_by_id(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="not_found")
    if order.get("buyer_id") != user["id"]:
        raise HTTPException(status_code=403, detail="not_your_order")
    if order.get("status") != "pending":
        raise HTTPException(status_code=409, detail="not_pending")
    reason = _norm_cancel_reason(body.reason if body else None)
    gid = order.get("order_group_id")
    # ATOMIK: faqat hali 'pending' bo'lsa bekor qilamiz. Sotuvchi ayni damda tasdiqlab
    # ulgursa, eski bekor 'confirmed'ni bosib o'tkazib zahirani yo'qotmasin.
    # Variant/savat guruh buyurtmasi — bitta amal bilan butun guruh bekor bo'ladi.
    if gid:
        if not db.transition_group_status(gid, "cancelled", "pending",
                                          cancel_by="buyer", cancel_reason=reason):
            raise HTTPException(status_code=409, detail="not_pending")
    elif not db.transition_order_status(order_id, "cancelled", "pending",
                                        cancel_by="buyer", cancel_reason=reason):
        raise HTTPException(status_code=409, detail="not_pending")
    try:
        seller = db.get_user_by_id(order["seller_id"]) if order.get("seller_id") else None
        slang = get_user_lang(seller) if seller else DEFAULT_LANG
        if order.get("seller_tg"):
            await _tg_call("sendMessage", {
                "chat_id": order["seller_tg"],
                "text": t(slang, "order_cancelled_notify", oid=fmt_order_id(order_id),
                          pname=html.escape(order.get("product_name") or "")),
                "parse_mode": "HTML"})
        # App-banner: sotuvchiga "xaridor bekor qildi"
        pn = order.get("product_name") or ""
        _notify_db(order["seller_id"], "order", ("❌ Xaridor buyurtmani bekor qildi", "❌ Покупатель отменил заказ"),
                   (pn, pn), ref_id=order_id)
        # ADMIN — bekor qilinganidan xabardor bo'lsin (app banner + Telegram push)
        await _notify_admins(
            f"❌ Buyurtma bekor qilindi {fmt_order_id(order_id)}",
            f"{pn} · xaridor bekor qildi", kind="order", ref_id=order_id)
        cid = order.get("notify_chat_id")
        mid = order.get("notify_message_id")
        if cid and mid:
            final = (order.get("notify_caption") or "") + "\n\n❌ Xaridor bekor qildi"
            await _tg_call("editMessageText", {
                "chat_id": cid, "message_id": mid, "text": final,
                "parse_mode": "HTML", "reply_markup": {"inline_keyboard": []}})
    except Exception as e:
        logging.warning(f"buyer cancel notify xato (order {order_id}): {e}")
    return {"ok": True}


class CancelReqIn(BaseModel):
    reason: Optional[str] = None


@app.post("/api/order/{order_id}/request-cancel")
async def api_request_cancel(order_id: int, body: CancelReqIn, authorization: str = Header(None)):
    """Xaridor TASDIQLANGAN buyurtmani bekor qilishni so'raydi (nizo oqimi boshlanishi).
    Sotuvchiga rozilik so'rovi (rozi/rad tugmalari) yuboriladi — bot ularni ushlaydi."""
    user = dict(_buyer_from_auth(authorization))
    _rate_limit("cancel_req", user["id"], 10, 60)
    reason = (body.reason or "").strip() or ("—")
    if len(reason) > 500:
        raise HTTPException(status_code=400, detail="too_long")
    order = db.get_order_by_id(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="not_found")
    if order.get("buyer_id") != user["id"]:
        raise HTTPException(status_code=403, detail="not_your_order")
    if order.get("status") != "confirmed" or (order.get("cancel_state") or ""):
        raise HTTPException(status_code=409, detail="cancel_not_available")
    gid = order.get("order_group_id")
    if gid:
        if not db.request_group_cancel(gid, "buyer", reason):
            raise HTTPException(status_code=409, detail="cancel_not_available")
    elif not db.request_order_cancel(order_id, "buyer", reason):
        raise HTTPException(status_code=409, detail="cancel_not_available")
    try:
        seller = db.get_user_by_id(order["seller_id"]) if order.get("seller_id") else None
        slang = get_user_lang(seller) if seller else DEFAULT_LANG
        if order.get("seller_tg"):
            await _tg_call("sendMessage", {
                "chat_id": order["seller_tg"],
                "text": t(slang, "cancel_request_notify", oid=fmt_order_id(order_id),
                          pname=html.escape(order.get("product_name") or ""),
                          reason=html.escape(_cancel_reason_display(reason, slang))),
                "parse_mode": "HTML",
                "reply_markup": {"inline_keyboard": [
                    [{"text": t(slang, "btn_cancel_agree"), "callback_data": f"cclagree_{order_id}"}],
                    [{"text": t(slang, "btn_cancel_deny"), "callback_data": f"ccldeny_{order_id}"}]]}})
    except Exception as e:
        logging.warning(f"request-cancel notify xato (order {order_id}): {e}")
    return {"ok": True}


# ============================================================
# #13 JONLI YETKAZIB BERISH KUZATUVI — yetkazib beruvchi (hozircha sotuvchi/kuryeri)
# joylashuvini ulashadi; xaridor masofa + ETA + xarita havolasini jonli ko'radi.
# (To'liq embedded jonli xarita keyin — maps provider kerak. Kuryer roli — #13-AI.)
# ============================================================
def _haversine_km(lat1, lon1, lat2, lon2):
    r = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlam / 2) ** 2
    return r * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


class CourierLoc(BaseModel):
    lat: float
    lon: float


@app.post("/api/seller/order/{order_id}/location")
def api_courier_location(order_id: int, body: CourierLoc, authorization: str = Header(None)):
    """Yetkazib beruvchi (sotuvchi/ega/admin) joylashuvini yangilaydi. Faqat yo'ldagi
    (confirmed) yetkazib berish buyurtmasi uchun."""
    user = dict(_buyer_from_auth(authorization))
    _rate_limit("courier_loc", user["id"], 120, 3600)
    order = db.get_order_by_id(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="not_found")
    if not (order.get("seller_id") == _owner_id(user) or user.get("role") == "admin"):
        raise HTTPException(status_code=403, detail="not_your_order")
    if order.get("delivery_type") != "delivery":
        raise HTTPException(status_code=409, detail="not_delivery")
    if order.get("status") != "confirmed":
        raise HTTPException(status_code=409, detail="not_in_delivery")
    if not (-90 <= body.lat <= 90 and -180 <= body.lon <= 180):
        raise HTTPException(status_code=400, detail="bad_coords")
    db.update_courier_location(order_id, body.lat, body.lon)
    return {"ok": True}


@app.get("/api/order/{order_id}/tracking")
def api_order_tracking(order_id: int, authorization: str = Header(None)):
    """Xaridor (yoki sotuvchi/admin) buyurtma kuzatuvini oladi: kuryer joylashuvi,
    manzilgacha masofa va taxminiy ETA (to'g'ri chiziq, shahar tezligi ~18 km/soat)."""
    user = dict(_buyer_from_auth(authorization))
    order = dict(db.get_order_by_id(order_id) or {})
    if not order:
        raise HTTPException(status_code=404, detail="not_found")
    if user["id"] not in (order.get("buyer_id"), order.get("seller_id")) \
            and user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="not_your_order")
    clat, clon = order.get("courier_lat"), order.get("courier_lon")
    dlat, dlon = order.get("buyer_lat"), order.get("buyer_lon")
    out = {
        "status": order.get("status"), "delivery_type": order.get("delivery_type"),
        "courier_lat": clat, "courier_lon": clon,
        "courier_updated_at": order.get("courier_updated_at"),
        "dest_lat": dlat, "dest_lon": dlon,
        "distance_km": None, "eta_min": None,
        # #13 — xaridor kuryer bilan bog'lana olishi uchun kontakt (faqat biriktirilgan bo'lsa)
        "courier_name": order.get("courier_name"),
        "courier_phone": order.get("courier_phone"),
        "courier_username": order.get("courier_username"),
        "has_courier": order.get("courier_id") is not None,
    }
    if None not in (clat, clon, dlat, dlon):
        d = _haversine_km(clat, clon, dlat, dlon)
        out["distance_km"] = round(d, 2)
        out["eta_min"] = max(1, round(d / 18 * 60))   # ~18 km/soat (tirbandlik bilan)
    return out


# ============================================================
# AI #12 AQLLI MARSHRUT — sotuvchining yetkazib berish buyurtmalarini eng qisqa
# ketma-ketlikda tartiblaydi (nearest-neighbor + 2-opt, haversine). Tashqi API yo'q;
# real tirbandlik/yo'l marshruti — keyingi bosqich (routing provider kaliti bilan).
# ============================================================
def _route_dist(a, b):
    return _haversine_km(a[0], a[1], b[0], b[1])


def _optimize_route(start, stops):
    """start: (lat,lon) yoki None. stops: lat/lon bor dict'lar.
    Qaytaradi: (tartiblangan stops, jami_km)."""
    n = len(stops)
    if n == 0:
        return [], 0.0
    pts = [(float(s["buyer_lat"]), float(s["buyer_lon"])) for s in stops]

    def total_len(seq):
        d = 0.0
        if start is not None:
            prev = start
            rng = seq
        else:
            prev = pts[seq[0]]
            rng = seq[1:]
        for k in rng:
            d += _route_dist(prev, pts[k])
            prev = pts[k]
        return d

    # Nearest-neighbor
    unvisited = list(range(n))
    order = []
    cur = start
    if cur is None:
        order.append(unvisited.pop(0))
        cur = pts[order[-1]]
    while unvisited:
        j = min(unvisited, key=lambda i: _route_dist(cur, pts[i]))
        unvisited.remove(j)
        order.append(j)
        cur = pts[j]

    # 2-opt yaxshilash (kichik N uchun yetarli)
    improved = True
    while improved:
        improved = False
        base = total_len(order)
        for i in range(len(order) - 1):
            for k in range(i + 1, len(order)):
                cand = order[:i] + order[i:k + 1][::-1] + order[k + 1:]
                cl = total_len(cand)
                if cl + 1e-9 < base:
                    order, base, improved = cand, cl, True
    return [stops[i] for i in order], total_len(order)


@app.get("/api/seller/route")
def api_seller_route(authorization: str = Header(None)):
    """AI #12 — sotuvchining yo'ldagi (confirmed) yetkazib berish buyurtmalarini eng
    qisqa marshrutda tartiblab beradi (do'kondan boshlab). Koordinatasiz buyurtmalar
    tashlanadi. Ko'p chiqib ketmasligi uchun 25 ta bilan cheklangan."""
    user = dict(_buyer_from_auth(authorization))
    orders = [dict(o) for o in db.get_seller_orders_list(_owner_id(user))]
    stops = [o for o in orders
             if o.get("status") == "confirmed" and o.get("delivery_type") == "delivery"
             and o.get("buyer_lat") is not None and o.get("buyer_lon") is not None][:25]
    start = None
    if user.get("shop_lat") is not None and user.get("shop_lon") is not None:
        start = (float(user["shop_lat"]), float(user["shop_lon"]))
    ordered, total_km = _optimize_route(start, stops)
    out_stops = []
    prev = start
    for i, o in enumerate(ordered):
        pt = (float(o["buyer_lat"]), float(o["buyer_lon"]))
        leg = round(_route_dist(prev, pt), 2) if prev is not None else None
        out_stops.append({
            "seq": i + 1, "order_id": o.get("id"), "buyer_name": o.get("buyer_name"),
            "address": o.get("delivery_address"), "lat": pt[0], "lon": pt[1],
            "leg_km": leg, "buyer_phone": o.get("buyer_phone"),
        })
        prev = pt
    # Google Maps ko'p-to'xtovli yo'nalish havolasi.
    # start bo'lsa: origin=do'kon, waypoints=to'xtovlar[:-1], destination=oxirgi to'xtov.
    # start yo'q bo'lsa: origin=1-to'xtov, waypoints=o'rtadagilar, destination=oxirgi.
    maps_url = None
    if out_stops:
        coords = [f"{s['lat']},{s['lon']}" for s in out_stops]
        dest = coords[-1]
        if start is not None:
            origin = f"{start[0]},{start[1]}"
            mids = coords[:-1]
        else:
            origin = coords[0]
            mids = coords[1:-1]
        wp = ("&waypoints=" + "|".join(mids)) if mids else ""
        maps_url = f"https://www.google.com/maps/dir/?api=1&origin={origin}&destination={dest}{wp}"
    return {
        "stops": out_stops, "total_km": round(total_km, 2),
        "eta_min": (max(1, round(total_km / 18 * 60)) if total_km else 0),
        "has_start": start is not None, "maps_url": maps_url,
    }


class CancelRespondIn(BaseModel):
    agree: bool


@app.post("/api/order/{order_id}/cancel-respond")
async def api_cancel_respond(order_id: int, body: CancelRespondIn, authorization: str = Header(None)):
    """Xaridor SOTUVCHI boshlagan bekor so'roviga javob beradi: rozi (bekor) yoki rad (nizo)."""
    user = dict(_buyer_from_auth(authorization))
    order = db.get_order_by_id(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="not_found")
    if order.get("buyer_id") != user["id"]:
        raise HTTPException(status_code=403, detail="not_your_order")
    if order.get("cancel_state") != "requested":
        raise HTTPException(status_code=409, detail="cancel_already_handled")
    if order.get("cancel_by") == "buyer":
        raise HTTPException(status_code=409, detail="cancel_wait_other")  # o'zi so'ragan
    oid = fmt_order_id(order_id)
    pname = html.escape(order.get("product_name") or "")
    seller = db.get_user_by_id(order["seller_id"]) if order.get("seller_id") else None
    slang = get_user_lang(seller) if seller else DEFAULT_LANG
    gid = order.get("order_group_id")
    if body.agree:
        if gid:
            for o in db.agree_group_cancel(gid):
                try:
                    od = db.get_order_by_id(o)
                    db.restock_on_cancel(od["product_id"], od.get("quantity") or 1)
                except Exception as e:
                    logging.warning(f"cancel-respond restock xato (order {o}): {e}")
        elif db.agree_order_cancel(order_id):
            try:
                db.restock_on_cancel(order["product_id"], order.get("quantity") or 1)
            except Exception as e:
                logging.warning(f"cancel-respond restock xato (order {order_id}): {e}")
        if order.get("seller_tg"):
            await _tg_call("sendMessage", {"chat_id": order["seller_tg"],
                           "text": t(slang, "cancel_agreed_notify", oid=oid, pname=pname),
                           "parse_mode": "HTML"})
        return {"ok": True, "cancelled": True}
    # rad — admin hakamligiga
    if gid:
        db.dispute_group_cancel(gid)
    else:
        db.dispute_order_cancel(order_id)
    _pn = order.get("product_name") or ""
    _notify_admins_db("dispute", ("⚖️ Yangi nizo — hal kutyapti", "⚖️ Новый спор — ждёт решения"),
                      (_pn, _pn), ref_id=order_id)
    if order.get("seller_tg"):
        await _tg_call("sendMessage", {"chat_id": order["seller_tg"],
                       "text": t(slang, "cancel_denied_notify", oid=oid, pname=pname),
                       "parse_mode": "HTML"})
    try:
        if ADMIN_ID:
            await _tg_call("sendMessage", {"chat_id": ADMIN_ID,
                           "text": t(DEFAULT_LANG, "admin_dispute_notify", oid=oid, pname=pname,
                                     by=html.escape(user.get("name") or "xaridor"),
                                     reason=html.escape(order.get("cancel_reason") or "—")),
                           "parse_mode": "HTML"})
    except Exception as e:
        logging.warning(f"cancel-respond admin notify xato (order {order_id}): {e}")
    return {"ok": True, "disputed": True}


def _order_party_or_403(user, order):
    # Ishtirokchilar: xaridor, sotuvchi (ega) va #13 — buyurtmaga biriktirilgan KURYER.
    parties = (order.get("buyer_id"), order.get("seller_id"), order.get("courier_id"))
    if user.get("id") not in parties and user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="not_your_order")


@app.post("/api/order/{order_id}/confirm-pickup")
async def api_confirm_pickup(order_id: int, authorization: str = Header(None)):
    """Xaridor «oldim» tasdig'i (bot buyer_confirm_pickup parite). Status 'confirmed'
    qoladi (sotuvchi to'lovni belgilab yakunlaydi) — faqat buyer_received=1 + sotuvchiga xabar."""
    buyer = dict(_buyer_from_auth(authorization))
    order = db.get_order_by_id(order_id)
    if not order or dict(order).get("buyer_id") != buyer["id"]:
        raise HTTPException(status_code=403, detail="not_your_order")
    order = dict(order)
    if order.get("status") != "confirmed":
        raise HTTPException(status_code=409, detail="not_confirmed")
    if order.get("delivery_type") != "pickup":
        raise HTTPException(status_code=400, detail="not_pickup")
    if order.get("buyer_received"):
        raise HTTPException(status_code=409, detail="already_received")
    gid = order.get("order_group_id")
    if gid:
        db.set_group_buyer_received(gid)
    else:
        db.set_buyer_received(order_id)
    try:
        if order.get("seller_tg"):
            seller = db.get_user_by_id(order["seller_id"])
            slang = get_user_lang(seller) if seller else DEFAULT_LANG
            await _tg_call("sendMessage", {
                "chat_id": order["seller_tg"],
                "text": t(slang, "pickup_seller_finalize",
                          oid=fmt_order_id(order_id),
                          pname=html.escape(order.get("product_name") or ""),
                          buyer=html.escape(order.get("buyer_name") or "")),
                "parse_mode": "HTML",
                "reply_markup": {"inline_keyboard": [[
                    {"text": t(slang, "btn_finalize_payment"),
                     "callback_data": f"seller_order_{order_id}"}]]}})
    except Exception as e:
        logging.warning(f"pickup seller notify (web) xato: {e}")
    return {"ok": True}


# Yakunlangan buyurtmalar — suhbat yopiq (yangi xabar yo'q, faqat tarix ko'rinadi).
_CHAT_CLOSED_STATUSES = ("delivered", "cancelled")


@app.get("/api/order/{order_id}/messages")
def api_order_messages(order_id: int, authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    order = db.get_order_by_id(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="not_found")
    _order_party_or_403(user, order)
    # #13 — xaridor uchun suhbatdosh: kuryer biriktirilgan bo'lsa kuryer, aks holda do'kon.
    if user["id"] == order.get("buyer_id"):
        cp = order.get("courier_name") or order.get("shop_name")
    else:
        cp = order.get("buyer_name")
    # Buyurtma yakunlangach (yetkazildi/bekor) suhbat YOPIQ — yangi xabar yozib bo'lmaydi
    # (tarix ko'rinadi). Bu "chat ochiq qoldi" muammosini yopadi.
    return {"me": user["id"], "counterparty": cp or "—",
            "closed": order.get("status") in _CHAT_CLOSED_STATUSES,
            "messages": _rows(db.get_messages_by_order(order_id))}


class MsgIn(BaseModel):
    text: str


@app.post("/api/order/{order_id}/message")
async def api_send_message(order_id: int, body: MsgIn, authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    text = (body.text or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="empty")
    if len(text) > 2000:
        raise HTTPException(status_code=400, detail="too_long")
    order = db.get_order_by_id(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="not_found")
    _order_party_or_403(user, order)
    # Yakunlangan buyurtmada suhbat yopiq — yangi xabar qabul qilinmaydi.
    if order.get("status") in _CHAT_CLOSED_STATUSES:
        raise HTTPException(status_code=409, detail="chat_closed")

    # #13 — yo'naltirish: xaridordan — kuryer biriktirilgan bo'lsa kuryerga, aks holda
    # sotuvchiga; kuryer/sotuvchidan — har doim xaridorga. Xabarlar bitta tred bo'lib
    # ko'rinadi; receiver faqat Telegram bildirishnomasi manzilini belgilaydi.
    if user["id"] == order.get("buyer_id"):
        if order.get("courier_id"):
            receiver_id = order.get("courier_id")
            receiver_tg = order.get("courier_tg")
        else:
            receiver_id = order.get("seller_id")
            receiver_tg = order.get("seller_tg")
        sender_role = "buyer"
    elif user["id"] == order.get("courier_id"):
        receiver_id = order.get("buyer_id")
        receiver_tg = order.get("buyer_tg")
        sender_role = "courier"
    else:
        receiver_id = order.get("buyer_id")
        receiver_tg = order.get("buyer_tg")
        sender_role = "seller"
    db.create_message(order_id, user["id"], receiver_id, text)
    # App-banner: qabul qiluvchiga yangi chat xabari (Telegram push'i ham quyida yuboriladi)
    _sn = user.get("name") or ""
    _notify_db(receiver_id, "message", ("💬 Yangi xabar", "💬 Новое сообщение"),
               (f"{_sn}: {text[:100]}", f"{_sn}: {text[:100]}"), ref_id=order_id)

    # Qabul qiluvchini Telegram orqali xabardor qilamiz (uning tilida) + bot'dan javob tugmasi
    try:
        receiver = db.get_user_by_id(receiver_id) if receiver_id else None
        rlang = get_user_lang(receiver) if receiver else DEFAULT_LANG
        if sender_role == "buyer":
            sender_label = t(rlang, "sender_label_buyer", name=html.escape(user.get("name") or ""))
        elif sender_role == "courier":
            sender_label = t(rlang, "sender_label_courier", name=html.escape(user.get("name") or ""))
        else:
            sender_label = t(rlang, "sender_label_seller",
                             name=html.escape(user.get("shop_name") or user.get("name") or ""))
        if receiver_tg:
            await _tg_call("sendMessage", {
                "chat_id": receiver_tg,
                "text": t(rlang, "new_message_notify", oid=fmt_order_id(order_id),
                          sender=sender_label, msg=html.escape(text)),
                "parse_mode": "HTML",
                "reply_markup": {"inline_keyboard": [[
                    {"text": t(rlang, "btn_reply"), "callback_data": f"order_msg_{order_id}"}]]},
            })
    except Exception as e:
        logging.error(f"xabar bildirishnomasi xato (order {order_id}): {e}")
    return {"ok": True}


class MeEdit(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    language: Optional[str] = None
    region_id: Optional[int] = None          # 0/-1 = tozalash
    telegram_username: Optional[str] = None   # bo'sh = tozalash


@app.patch("/api/me")
def api_edit_me(p: MeEdit, authorization: str = Header(None)):
    """Foydalanuvchi O'Z profilini tahrirlaydi — har maydon ALOHIDA (faqat kelgan maydon)."""
    user = dict(_buyer_from_auth(authorization))
    fields = {}
    if p.name is not None:
        nm = p.name.strip()
        if not nm or len(nm) > 60:
            raise HTTPException(status_code=400, detail="bad_name")
        fields["name"] = nm
    if p.phone is not None and p.phone.strip():
        ph = p.phone.strip()
        digits = ph.lstrip("+")
        if not digits.isdigit() or not (7 <= len(digits) <= 15):
            raise HTTPException(status_code=400, detail="bad_phone")
        fields["phone_number"] = ph
    if p.language is not None:
        if p.language not in ("uz", "ru"):
            raise HTTPException(status_code=400, detail="bad_language")
        fields["language"] = p.language
    if p.region_id is not None:
        if p.region_id <= 0:
            fields["region_id"] = None
        elif db.get_region_by_id(p.region_id):
            fields["region_id"] = p.region_id
        else:
            raise HTTPException(status_code=400, detail="bad_region")
    if p.telegram_username is not None:
        un = p.telegram_username.strip().lstrip("@").replace(" ", "")
        if not un:
            fields["telegram_username"] = None
        elif _re.fullmatch(r"[A-Za-z0-9_]{5,32}", un):
            fields["telegram_username"] = un
        else:
            raise HTTPException(status_code=400, detail="bad_username")
    if fields:
        db.update_user(user["id"], **fields)
    return {"ok": True}


@app.get("/api/me/full")
def api_me_full(authorization: str = Header(None)):
    """Foydalanuvchining TO'LIQ profili (o'zi uchun) — admin_user_details pariteti.
    Asosiy + do'kon + kanallar + referal + faollik (do'kon kesimida)."""
    user = dict(_buyer_from_auth(authorization))
    uid = user["id"]
    u = dict(db.get_user_by_id(uid) or {})
    out = dict(u)
    out.pop("min_price", None)
    try:
        out["region_label"] = db.get_region_label(u.get("region_id")) or ""
    except Exception:
        out["region_label"] = ""
    try:
        out["buyer_orders_count"] = len(db.get_orders_by_buyer(uid) or [])
    except Exception:
        out["buyer_orders_count"] = 0
    out["is_owner"] = bool(db.get_shop_by_owner(uid))
    staff = db.get_staff_by_user(uid)
    staff_d = dict(staff) if staff else None
    out["is_staff"] = bool(staff_d and staff_d.get("staff_role") != "owner")
    out["staff_role"] = staff_d.get("staff_role") if staff_d else None
    out["is_courier"] = bool(staff_d and staff_d.get("staff_role") == "courier")
    # B — XODIM shaxsiy hissasi (faqat o'zi joylagan: created_by bo'yicha)
    if out["is_staff"]:
        try:
            out["staff_stats"] = dict(db.get_staff_stats(uid) or {})
        except Exception:
            out["staff_stats"] = {}
    # XODIM profili — do'kon bo'limi EGA to'ldirgan ma'lumotlarni ko'rsatsin (do'kon
    # maydonlari EGA yozuvida saqlanadi; xodimnikida bo'sh — "yana to'ldirish kerak"
    # degan noto'g'ri taassurot berardi). Frontend baribir read-only qiladi (canEditShop).
    if out["is_staff"]:
        try:
            sh = db.get_shop_for_user(uid)
            owner_u = dict(db.get_user_by_id(dict(sh)["owner_user_id"]) or {}) if sh else {}
            for k in ("shop_name", "shop_address", "shop_landmark",
                      "working_hours", "working_days", "shop_lat", "shop_lon"):
                if owner_u.get(k) is not None:
                    out[k] = owner_u[k]
        except Exception as e:
            logging.warning(f"me/full xodim do'kon overlay xato (uid {uid}): {e}")
    is_seller = bool(u.get("shop_name")) or u.get("role") == "seller"
    if is_seller:
        oid = _owner_id(user)   # faollik/kanallar — DO'KON kesimida
        try:
            out["seller_stats"] = dict(db.get_seller_stats(oid) or {})
        except Exception:
            out["seller_stats"] = {}
        try:
            out["seller_avg_rating"] = round(float(db.get_seller_avg_rating(oid) or 0), 1)
        except Exception:
            out["seller_avg_rating"] = 0
        try:
            out["channels"] = _rows(db.get_seller_channels(oid))
        except Exception:
            out["channels"] = []
    # referred_by ismi (kim taklif qilgan)
    rb = u.get("referred_by")
    if rb:
        try:
            ref = db.get_user_by_id(rb)
            out["referred_by_name"] = dict(ref).get("name") if ref else None
        except Exception:
            out["referred_by_name"] = None
    return out


# ============================================================
# SOTUVCHI PANELI (D bo'lagi)
# ============================================================
async def _tg_call(method, payload):
    """Telegram Bot API'ga to'g'ridan-to'g'ri chaqiruv (webapp bot token bilan)."""
    if not BOT_TOKEN:
        return None
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            r = await client.post(
                f"https://api.telegram.org/bot{BOT_TOKEN}/{method}", json=payload)
            return r.json()
    except Exception as e:
        logging.warning(f"Telegram {method} xato: {e}")
        return None


async def _notify_price_drop(product_id, name, old_price, new_price):
    """#16 — narx tushganda mahsulotni sevimliga qo'shgan xaridorlarga xabar."""
    try:
        favoriters = db.get_product_favoriters(product_id)
    except Exception as e:
        logging.warning(f"price-drop favoriters xato (pid {product_id}): {e}")
        return
    pname = html.escape(name or "")
    for f in favoriters:
        lang = (f.get("language") or DEFAULT_LANG)
        if lang == "ru":
            txt = (f"📉 <b>Цена снизилась!</b>\n{pname}\n"
                   f"<s>{fmt_price(old_price)}</s> → <b>{fmt_price(new_price)}</b>")
        else:
            txt = (f"📉 <b>Narx tushdi!</b>\n{pname}\n"
                   f"<s>{fmt_price(old_price)}</s> → <b>{fmt_price(new_price)}</b>")
        try:
            await _tg_call("sendMessage", {"chat_id": f["telegram_id"], "text": txt,
                                           "parse_mode": "HTML"})
        except Exception as e:
            logging.warning(f"price-drop xabar xato (buyer {f.get('id')}): {e}")


@app.get("/api/seller/orders")
def api_seller_orders(authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    rows = _rows(db.get_seller_orders_list(_owner_id(user)))   # do'kon (ega+xodimlar) buyurtmalari
    # Bot xabarnomasidagidek do'kon → xaridor masofasini (km) qo'shamiz. shop_lat/shop_lon
    # — do'kon EGAsining koordinatasi (faqat hisob uchun, frontend'ga uzatilmaydi).
    for o in rows:
        slat, slon = o.pop("shop_lat", None), o.pop("shop_lon", None)
        blat, blon = o.get("buyer_lat"), o.get("buyer_lon")
        if o.get("delivery_type") == "delivery" and blat and blon and slat and slon:
            try:
                o["dist_km"] = round(_haversine_km(slat, slon, blat, blon), 1)
            except Exception:
                pass
    return rows


@app.get("/api/seller/products")
def api_seller_products(authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    return _rows(db.get_products_by_seller(_owner_id(user)))   # do'kon mahsulotlari (ega+xodimlar)


@app.get("/api/seller/deleted")
def api_seller_deleted(authorization: str = Header(None)):
    """App «O'chirilgan» tabi — shu do'konning o'chirilgan mahsulotlari jurnali
    (product_audit). Faqat o'qish uchun (tarix)."""
    user = dict(_buyer_from_auth(authorization))
    return _rows(db.get_seller_product_audit(_owner_id(user)))


@app.get("/api/courier/orders")
def api_courier_orders(authorization: str = Header(None)):
    """C — KURYER paneli: AYNAN shu kuryerga biriktirilgan yo'ldagi (confirmed)
    yetkazib berish buyurtmalari (#3 — endi faqat o'ziga biriktirilganlar)."""
    user = dict(_buyer_from_auth(authorization))
    if not _is_courier(user):
        raise HTTPException(status_code=403, detail="not_courier")
    orders = [dict(o) for o in db.get_seller_orders_list(_owner_id(user))]
    return _rows([o for o in orders
                  if o.get("status") == "confirmed" and o.get("delivery_type") == "delivery"
                  and o.get("courier_id") == user["id"]])


@app.get("/api/seller/couriers")
def api_seller_couriers(authorization: str = Header(None)):
    """#3 — biriktirish uchun do'kon kuryerlari ro'yxati (ega yoki tasdiqlash huquqli xodim)."""
    user = dict(_buyer_from_auth(authorization))
    if not _staff_perm(user, "perm_confirm_orders"):
        raise HTTPException(status_code=403, detail="no_perm")
    return _rows(db.get_shop_couriers(_owner_id(user)))


class CourierAssign(BaseModel):
    courier_id: Optional[int] = None   # None = biriktirishni bekor qilish


@app.post("/api/seller/order/{order_id}/assign-courier")
def api_assign_courier(order_id: int, body: CourierAssign, authorization: str = Header(None)):
    """#3 — buyurtmaga kuryer biriktiradi/bekor qiladi. Faqat ega yoki tasdiqlash
    huquqli xodim. Kuryer shu do'konning FAOL kuryeri bo'lishi shart."""
    user = dict(_buyer_from_auth(authorization))
    if not _staff_perm(user, "perm_confirm_orders"):
        raise HTTPException(status_code=403, detail="no_perm")
    owner_id = _owner_id(user)
    if body.courier_id is not None and not db.is_shop_courier(owner_id, body.courier_id):
        raise HTTPException(status_code=400, detail="not_a_courier")
    if not db.assign_order_courier(order_id, body.courier_id, owner_id):
        raise HTTPException(status_code=404, detail="order_not_found")
    return {"ok": True, "courier_id": body.courier_id}


# Excel hisobotda inglizcha enum qiymatlarni (status/yetkazish/to'lov/rol) UZ/RU'ga
# o'giradi — hisobot ichida ingliz so'z qolmasin.
_XL_LABELS = {
    "status": {
        "pending":       ("Yangi", "Новый"),
        "confirmed":     ("Tasdiqlangan", "Подтверждён"),
        "delivered":     ("Yetkazildi", "Доставлен"),
        "cancelled":     ("Bekor qilindi", "Отменён"),
        "active":        ("Sotuvda", "В продаже"),
        "pending_owner": ("Tasdiq kutyapti", "Ждёт подтверждения"),
        "mod_blocked":   ("Moderatsiyada", "На модерации"),
        "scheduled":     ("Rejalashtirilgan", "Запланирован"),
        "deleted":       ("O'chirilgan", "Удалён"),
        "purged":        ("O'chirilgan", "Удалён"),
        "sold":          ("Sotilgan", "Продан"),
    },
    "delivery": {
        "pickup":   ("Olib ketish", "Самовывоз"),
        "delivery": ("Yetkazib berish", "Доставка"),
    },
    "settlement": {
        "paid":        ("To'liq to'landi", "Оплачено полностью"),
        "debt":        ("Qarzga", "В долг"),
        "installment": ("Bo'lib to'lash", "Рассрочка"),
        "partial":     ("Qisman", "Частично"),
    },
    "payment": {
        "cash":     ("Naqd", "Наличные"),
        "terminal": ("Terminal", "Терминал"),
        "p2p":      ("Karta (P2P)", "Карта (P2P)"),
        "card":     ("Karta", "Карта"),
    },
    "role": {
        "buyer":   ("Xaridor", "Покупатель"),
        "seller":  ("Sotuvchi", "Продавец"),
        "admin":   ("Administrator", "Администратор"),
        "courier": ("Kuryer", "Курьер"),
    },
}


def _xl_loc(cat, value, ru):
    """Enum qiymatni UZ (ru=False) yoki RU (ru=True) yorlig'iga o'giradi. Noma'lum → xom."""
    if value is None or value == "":
        return ""
    pair = _XL_LABELS.get(cat, {}).get(str(value))
    if not pair:
        return str(value)
    return pair[1] if ru else pair[0]


def _xl_fill_sheet(ws, report_title, headers, data_rows, money_cols=(), ru=False):
    """Berilgan worksheet'ni Pro ko'rinishdagi hisobotga to'ldiradi: brend sarlavha +
    meta qator + uslublangan header + zebra + chegaralar + pul formati (#,##0) + JAMI
    (SUM) qatori + freeze panes + auto-filter + auto-kenglik. -> qatorlar soni n.
    (Bitta workbook ichida ko'p sahifa yasash uchun ham ishlatiladi.)"""
    import datetime as _dt
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    BRAND, BRAND_DARK, ZEBRA, META = "0E7C3A", "0A5A2A", "EAF6EE", "5A6B5E"
    # Excel varaq nomida ruxsat etilmaydigan belgilar (: \ / ? * [ ]) — tozalanadi
    _safe = report_title
    for _ch in ":\\/?*[]":
        _safe = _safe.replace(_ch, " ")
    ws.title = (_safe.strip()[:31] or "Report")
    ncol = max(1, len(headers))
    thin = Side(style="thin", color="C8D6CC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    n = len(data_rows)
    money_fmt = "#,##0"
    mcols = set(money_cols)

    # 1-qator — brend sarlavha (birlashtirilgan)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    tc = ws.cell(row=1, column=1, value=f"{BRAND_NAME}  ·  {report_title}")
    tc.font = Font(bold=True, size=15, color="FFFFFF")
    tc.fill = PatternFill("solid", fgColor=BRAND)
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    # 2-qator — meta (yaratilgan sana + qatorlar soni)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    gen = _dt.datetime.now().strftime("%Y-%m-%d %H:%M")
    meta = (f"Сформирован: {gen}   ·   {n} записей" if ru
            else f"Yaratilgan: {gen}   ·   {n} qator")
    mc = ws.cell(row=2, column=1, value=meta)
    mc.font = Font(size=10, italic=True, color=META)
    mc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18

    # 3-qator — header
    hrow = 3
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=hrow, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=BRAND_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[hrow].height = 26

    # Ma'lumot qatorlari (zebra + pul formati)
    r = hrow
    for data in data_rows:
        r += 1
        zebra = (r - hrow) % 2 == 0
        for j, val in enumerate(data, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.border = border
            if zebra:
                c.fill = PatternFill("solid", fgColor=ZEBRA)
            if (j - 1) in mcols:
                if isinstance(val, (int, float)):
                    c.number_format = money_fmt
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
    last = r

    # JAMI qatori (faqat pul ustunlari yig'iladi)
    if data_rows and mcols:
        r += 1
        lc = ws.cell(row=r, column=1, value=("ИТОГО" if ru else "JAMI"))
        lc.font = Font(bold=True)
        lc.alignment = Alignment(horizontal="left", indent=1)
        top = Side(style="medium", color=BRAND)
        for j in range(1, ncol + 1):
            cell = ws.cell(row=r, column=j)
            cell.border = Border(top=top)
            if (j - 1) in mcols:
                col = get_column_letter(j)
                cell.value = f"=SUM({col}{hrow+1}:{col}{last})"
                cell.number_format = money_fmt
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="right")

    ws.freeze_panes = ws.cell(row=hrow + 1, column=1)
    ws.auto_filter.ref = f"A{hrow}:{get_column_letter(ncol)}{last}"
    for j in range(1, ncol + 1):
        ml = len(str(headers[j - 1]))
        for data in data_rows:
            v = data[j - 1] if j - 1 < len(data) else ""
            ml = max(ml, len(str(v if v is not None else "")))
        ws.column_dimensions[get_column_letter(j)].width = min(max(ml + 3, 10), 42)

    return n


def _xlsx_report(report_title, headers, data_rows, money_cols=(), lang="uz"):
    """Bitta sahifali Excel hisobot -> (bytes, n). _xl_fill_sheet uslubidan foydalanadi."""
    import io as _io
    import openpyxl
    wb = openpyxl.Workbook()
    n = _xl_fill_sheet(wb.active, report_title, headers, data_rows,
                       money_cols=money_cols, ru=(lang == "ru"))
    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), n


def _build_seller_excel(seller_id, kind, lang):
    """Sotuvchining buyurtmalari yoki mahsulotlarini Pro-Excel'ga yig'adi -> (bytes, fname, n)."""
    import datetime as _dt
    ru = (lang == "ru")
    if kind == "products":
        title = "Отчёт: Товары" if ru else "Mahsulotlar hisoboti"
        headers = (["ID", "Товар", "Цена", "Старая цена", "Статус", "Остаток", "В продаже", "Дата"]
                   if ru else
                   ["ID", "Nom", "Narx", "Eski narx", "Holat", "Zahira", "Sotuvda", "Sana"])
        data = []
        for p in db.get_products_by_seller(seller_id):
            p = dict(p)
            data.append([p.get("id"), p.get("name") or "", p.get("price") or 0,
                         p.get("old_price") or "", _xl_loc("status", p.get("status"), ru),
                         p.get("stock_count") if p.get("stock_count") is not None else "∞",
                         "✓" if p.get("in_stock") else "—", str(p.get("created_at") or "")[:10]])
        content, n = _xlsx_report(title, headers, data, money_cols=(2,), lang=lang)
        fn = "mahsulotlar"
    else:  # orders
        title = "Отчёт: Заказы" if ru else "Buyurtmalar hisoboti"
        headers = (["ID", "Покупатель", "Товар", "Итого", "Статус", "Доставка",
                    "Оплата", "Оплачено", "Долг", "Дата"]
                   if ru else
                   ["ID", "Xaridor", "Mahsulot", "Jami", "Holat", "Yetkazish",
                    "To'lov holati", "To'langan", "Qarz", "Sana"])
        data = []
        for o in db.get_seller_orders_list(seller_id):
            o = dict(o)
            data.append([o.get("id"), o.get("buyer_name") or "", o.get("product_name") or "",
                         o.get("total_price") or o.get("price") or 0, _xl_loc("status", o.get("status"), ru),
                         _xl_loc("delivery", o.get("delivery_type"), ru),
                         _xl_loc("settlement", o.get("settlement_type"), ru),
                         o.get("amount_paid") or 0, o.get("amount_due") or 0,
                         str(o.get("created_at") or "")[:16]])
        content, n = _xlsx_report(title, headers, data, money_cols=(3, 7, 8), lang=lang)
        fn = "buyurtmalar"
    fname = f"tezbozor_{fn}_{_dt.datetime.now().strftime('%Y%m%d')}.xlsx"
    return content, fname, n


def _build_seller_excel_full(seller_id, lang):
    """TO'LIQ sotuvchi hisoboti — bot seller_export_excel bilan to'la parite: bitta
    faylda 4 sahifa (Buyurtmalar to'liq ustunlar, Mahsulotlar sotilgan+daromad,
    Reytinglar, Umumiy). -> (bytes, fname, buyurtmalar_soni)."""
    import io as _io
    import datetime as _dt
    import openpyxl
    ru = (lang == "ru")
    orders = [dict(o) for o in db.get_seller_orders_list(seller_id)]
    products = [dict(p) for p in db.get_products_by_seller(seller_id)]
    reviews = [dict(r) for r in db.get_seller_reviews(seller_id)]
    stats = db.get_seller_stats(seller_id) or {}
    avg_rating = db.get_seller_avg_rating(seller_id) or 0
    owner = dict(db.get_user_by_id(seller_id) or {})

    wb = openpyxl.Workbook()

    # ---- 1) Buyurtmalar (to'liq: telefon, dona narxi, miqdor, to'lov usuli, manzil) ----
    o_headers = (["ID", "Дата", "Покупатель", "Телефон", "Товар", "Цена за шт",
                  "Кол-во", "Сумма", "Статус", "Доставка", "Оплата", "Адрес",
                  "Статус оплаты", "Оплачено", "Долг"]
                 if ru else
                 ["ID", "Sana", "Xaridor", "Telefon", "Mahsulot", "Dona narxi",
                  "Miqdor", "Jami", "Holat", "Yetkazish", "To'lov usuli", "Manzil",
                  "To'lov holati", "To'langan", "Qarz"])
    o_data = []
    for o in orders:
        o_data.append([
            o.get("id"), str(o.get("created_at") or "")[:16],
            o.get("buyer_name") or "", o.get("buyer_phone") or "",
            o.get("product_name") or "", o.get("product_price") or 0,
            o.get("quantity") or 0, o.get("total_price") or 0,
            _xl_loc("status", o.get("status"), ru),
            _xl_loc("delivery", o.get("delivery_type"), ru),
            _xl_loc("payment", o.get("payment_method"), ru),
            o.get("delivery_address") or "",
            _xl_loc("settlement", o.get("settlement_type"), ru),
            o.get("amount_paid") or 0, o.get("amount_due") or 0,
        ])
    _xl_fill_sheet(wb.active, "Заказы" if ru else "Buyurtmalar",
                   o_headers, o_data, money_cols=(5, 7, 13, 14), ru=ru)

    # ---- 2) Mahsulotlar (yetkazilgan buyurtmalardan sotilgan soni + daromad) ----
    sold_count, sold_revenue = {}, {}
    for o in orders:
        if (o.get("status") or "") == "delivered":
            pid = o.get("product_id")
            sold_count[pid] = sold_count.get(pid, 0) + (o.get("quantity") or 0)
            sold_revenue[pid] = sold_revenue.get(pid, 0) + (o.get("total_price") or 0)
    p_headers = (["ID", "Название", "Категория", "Цена", "Статус", "Остаток",
                  "Продано (шт)", "Доход (сум)", "Дата"]
                 if ru else
                 ["ID", "Nom", "Kategoriya", "Narx", "Holat", "Zahira",
                  "Sotilgan (dona)", "Daromad (so'm)", "Sana"])
    p_data = []
    for p in products:
        p_data.append([
            p.get("id"), p.get("name") or "", p.get("category_name") or "",
            p.get("price") or 0, _xl_loc("status", p.get("status"), ru),
            p.get("stock_count") if p.get("stock_count") is not None else "∞",
            sold_count.get(p.get("id"), 0), sold_revenue.get(p.get("id"), 0),
            str(p.get("created_at") or "")[:10],
        ])
    _xl_fill_sheet(wb.create_sheet("Товары" if ru else "Mahsulotlar"),
                   "Товары" if ru else "Mahsulotlar", p_headers, p_data,
                   money_cols=(3, 7), ru=ru)

    # ---- 3) Reytinglar ----
    r_headers = (["Дата", "Покупатель", "Оценка", "Комментарий"] if ru
                 else ["Sana", "Xaridor", "Baho", "Izoh"])
    r_data = []
    for r in reviews:
        r_data.append([str(r.get("created_at") or "")[:16],
                       r.get("buyer_name") or ("Аноним" if ru else "Anonim"),
                       r.get("rating") or 0, r.get("comment") or ""])
    _xl_fill_sheet(wb.create_sheet("Отзывы" if ru else "Reytinglar"),
                   "Отзывы" if ru else "Reytinglar", r_headers, r_data, ru=ru)

    # ---- 4) Umumiy hisobot ----
    total_debt = sum((o.get("amount_due") or 0) for o in orders)
    total_paid = sum((o.get("amount_paid") or 0) for o in orders)
    s_headers = (["Показатель", "Значение"] if ru else ["Ko'rsatkich", "Qiymat"])
    pairs = ([
        ("Магазин", owner.get("shop_name") or "—"),
        ("Кол-во товаров", stats.get("products_count", 0)),
        ("Средний рейтинг", round(avg_rating, 2)),
        ("Кол-во отзывов", len(reviews)),
        ("Всего заказов", stats.get("total_orders", 0)),
        ("Новые", stats.get("pending", 0)),
        ("Подтверждённые", stats.get("confirmed", 0)),
        ("Доставленные", stats.get("delivered", 0)),
        ("Отменённые", stats.get("cancelled", 0)),
        ("7 дней — заказы", stats.get("week_orders", 0)),
        ("7 дней — доход", stats.get("week_revenue", 0)),
        ("30 дней — заказы", stats.get("month_orders", 0)),
        ("30 дней — доход", stats.get("month_revenue", 0)),
        ("Общий доход (доставлено)", stats.get("total_revenue", 0)),
        ("Всего оплачено (при выдаче)", total_paid),
        ("Остаток долга", total_debt),
    ] if ru else [
        ("Do'kon", owner.get("shop_name") or "—"),
        ("Mahsulotlar soni", stats.get("products_count", 0)),
        ("O'rtacha reyting", round(avg_rating, 2)),
        ("Reytinglar soni", len(reviews)),
        ("Jami buyurtmalar", stats.get("total_orders", 0)),
        ("Yangi", stats.get("pending", 0)),
        ("Tasdiqlangan", stats.get("confirmed", 0)),
        ("Yetkazilgan", stats.get("delivered", 0)),
        ("Bekor qilingan", stats.get("cancelled", 0)),
        ("7 kun — buyurtma", stats.get("week_orders", 0)),
        ("7 kun — daromad", stats.get("week_revenue", 0)),
        ("30 kun — buyurtma", stats.get("month_orders", 0)),
        ("30 kun — daromad", stats.get("month_revenue", 0)),
        ("Jami daromad (yetkazilgan)", stats.get("total_revenue", 0)),
        ("Jami to'langan (berishda)", total_paid),
        ("Qolgan qarz", total_debt),
    ])
    _xl_fill_sheet(wb.create_sheet("Сводка" if ru else "Umumiy"),
                   "Сводка" if ru else "Umumiy", s_headers,
                   [list(p) for p in pairs], ru=ru)

    buf = _io.BytesIO()
    wb.save(buf)
    fname = f"tezbozor_toliq_hisobot_{_dt.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return buf.getvalue(), fname, len(orders)


@app.post("/api/seller/export/{kind}")
async def api_seller_export(kind: str, authorization: str = Header(None)):
    """orders|products|full -> Excel yasaydi va sotuvchining Telegram chatiga yuboradi.
    full = bot bilan to'la parite (4 sahifa: buyurtma/mahsulot/reyting/umumiy)."""
    user = dict(_buyer_from_auth(authorization))
    if user.get("role") not in ("seller", "admin") and not user.get("is_approved"):
        raise HTTPException(status_code=403, detail="not_seller")
    if kind not in ("orders", "products", "full"):
        raise HTTPException(status_code=400, detail="bad_kind")
    # #18 Pro — Excel hisobot Pro-obuna imkoniyati (obuna yoqilgan & Pro emas → qulf).
    if _pro_locked(_owner_id(user)):
        raise HTTPException(status_code=403, detail="pro_required")
    _rate_limit("seller_export", user["id"], 10, 600)
    lang = get_user_lang(user) or DEFAULT_LANG
    if kind == "full":
        content, fname, n = await asyncio.to_thread(_build_seller_excel_full, user["id"], lang)
    else:
        content, fname, n = await asyncio.to_thread(_build_seller_excel, user["id"], kind, lang)
    if not user.get("telegram_id"):
        raise HTTPException(status_code=400, detail="no_telegram")
    res = await _tg_send_document(user["telegram_id"], fname, content,
                                  caption=f"📊 {kind} — {n} ta · {BRAND_NAME}")
    if not (res and res.get("ok")):
        raise HTTPException(status_code=502, detail="send_failed")
    return {"ok": True, "rows": n}


# ---- SOTUVCHI QARZ DAFTARI + TO'LOV QABUL (bot seller_debts/debt_pay parite) ----
async def _notify_buyer_debt_paid_web(order, amount, remaining):
    """Xaridorga qarz to'lovi qayd etilgani haqida xabar (bot _notify_buyer_debt_paid parite)."""
    try:
        buyer = db.get_user_by_id(order["buyer_id"])
        if not buyer or not dict(buyer).get("telegram_id"):
            return
        buyer = dict(buyer)
        seller = db.get_user_by_id(order["seller_id"])
        blang = get_user_lang(buyer)
        shop = html.escape((dict(seller).get("shop_name") or dict(seller).get("name") or "") if seller else "")
        if (remaining or 0) <= 0:
            txt = t(blang, "buyer_debt_cleared", shop=shop)
        else:
            txt = t(blang, "buyer_debt_partial", shop=shop,
                    paid=fmt_price(amount), due=fmt_price(remaining))
        await _tg_call("sendMessage", {"chat_id": buyer["telegram_id"],
                                       "text": txt, "parse_mode": "HTML"})
    except Exception as e:
        logging.warning(f"debt paid notify (web) xato: {e}")


@app.get("/api/seller/debts")
def api_seller_debts(authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    owner_id = db.resolve_owner_id(user["id"])
    return {"debts": _rows(db.get_seller_open_debts(owner_id)),
            "total": db.get_seller_debt_total(owner_id)}


@app.get("/api/seller/debts/{buyer_id}")
def api_seller_debt_orders(buyer_id: int, authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    owner_id = db.resolve_owner_id(user["id"])
    orders = _rows(db.get_seller_debt_orders(owner_id, buyer_id))
    buyer = db.get_user_by_id(buyer_id)
    b = dict(buyer) if buyer else {}
    return {"orders": orders,
            "buyer": {"id": buyer_id, "name": b.get("name"),
                      "phone": b.get("phone_number"),
                      "username": b.get("telegram_username")}}


class DebtPayIn(BaseModel):
    amount: float


@app.post("/api/seller/debt/{order_id}/pay")
async def api_seller_debt_pay(order_id: int, body: DebtPayIn, authorization: str = Header(None)):
    """Qarzga to'lov qabul qiladi (to'liq yoki qisman) + xaridorga xabar."""
    user = dict(_buyer_from_auth(authorization))
    owner_id = db.resolve_owner_id(user["id"])
    order = db.get_order_by_id(order_id)
    if not order or dict(order).get("seller_id") != owner_id:
        raise HTTPException(status_code=403, detail="not_your_order")
    order = dict(order)
    due = float(order.get("amount_due") or 0)
    if due <= 0:
        raise HTTPException(status_code=409, detail="no_debt")
    _rate_limit("debt_pay", user["id"], 60, 60)
    pay = max(0.0, min(float(body.amount or 0), due))
    if pay <= 0:
        raise HTTPException(status_code=400, detail="bad_amount")
    remaining = db.record_debt_payment(order_id, pay)
    await _notify_buyer_debt_paid_web(order, pay, remaining or 0)
    return {"ok": True, "paid": pay, "remaining": remaining or 0}


@app.get("/api/seller/reviews")
def api_seller_reviews(authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    return _rows(db.get_seller_reviews(_owner_id(user)))


@app.get("/api/seller/customers")
def api_seller_customers(authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    return _rows(db.get_seller_customers(_owner_id(user)))


@app.get("/api/seller/channels")
def api_seller_channels(authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    return _rows(db.get_seller_channels(_owner_id(user)))


class ChannelRemove(BaseModel):
    channel_id: str


@app.post("/api/seller/channel/remove")
def api_remove_channel(body: ChannelRemove, authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    db.remove_seller_channel(user["id"], body.channel_id)
    return {"ok": True}


# ---- Xodimlar (multivendor) — faqat do'kon egasi ----
def _owner_shop(user):
    shop = db.get_shop_by_owner(user["id"])
    if not shop:
        raise HTTPException(status_code=403, detail="not_owner")
    return dict(shop)


def _owner_id(user):
    """Multivendor: foydalanuvchi (ega yoki xodim) tegishli do'kon EGASIning user id'si.
    Yakka sotuvchi/ega — o'zini qaytaradi. Xodim joylagan mahsulot/buyurtmalar shu id
    (ega) ostida ko'rinadi — do'kon nomidan chiqadi va bir joyda jamlanadi."""
    return db.resolve_owner_id(user["id"])


def _staff_perm(user, perm_col, default=1):
    """Xodim ruxsatini tekshiradi (ega/yakka sotuvchi — har doim ruxsat). default — bot bilan bir xil."""
    staff = db.get_staff_by_user(user["id"])
    if not staff or dict(staff).get("staff_role") == "owner":
        return True
    return bool(dict(staff).get(perm_col, default))


def _is_courier(user):
    """Foydalanuvchi do'kon KURYERImi (staff_role='courier')."""
    s = db.get_staff_by_user(user["id"])
    return bool(s and dict(s).get("staff_role") == "courier")


def _pro_until_dt(pu):
    if not pu:
        return None
    try:
        dt = datetime.fromisoformat(str(pu).replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except Exception:
        return None


def _is_pro(owner_id):
    """Do'kon egasining Pro-obunasi faolmi (pro_until > hozir)."""
    u = db.get_user_by_id(owner_id)
    dt = _pro_until_dt(dict(u).get("pro_until")) if u else None
    return bool(dt and dt > datetime.now(timezone.utc))


def _pro_until_active(pu):
    """pro_until qiymati (DB satri) hozir faolmi — qatorlarga seller_is_pro yopishtirish uchun."""
    dt = _pro_until_dt(pu)
    return bool(dt and dt > datetime.now(timezone.utc))


def _pro_locked(owner_id):
    """Pro-imkoniyat qulflanganmi: obuna monetizatsiyasi YOQILGAN va ega Pro EMAS.
    Obuna o'chiq bo'lsa hech narsa qulflanmaydi — hammasi bepul (eski xulq)."""
    return bool(monetization_public().get("subscription")) and not _is_pro(owner_id)


def _month_key():
    """Joriy oy kaliti (UTC, 'YYYY-MM') — kvota hisoblagichi uchun."""
    return datetime.now(timezone.utc).strftime("%Y-%m")


def _pro_status(user):
    """Frontend uchun Pro holati: egasi bo'yicha (pro_until egada saqlanadi)."""
    owner = db.get_user_by_id(_owner_id(user)) if user else None
    dt = _pro_until_dt(dict(owner).get("pro_until")) if owner else None
    active = bool(dt and dt > datetime.now(timezone.utc))
    return {"active": active, "until": (dt.isoformat() if dt else None)}


def _staff_mgmt_shop(user):
    """Xodim BOSHQARUVI (ro'yxat/taklif/o'chirish) huquqi bor do'kon: EGA yoki
    perm_add_staff berilgan MENEJER. Ruxsat berish (rol/perm) esa faqat egada qoladi."""
    shop = db.get_shop_by_owner(user["id"])
    if shop:
        return dict(shop)
    staff = db.get_staff_by_user(user["id"])
    if staff:
        st = dict(staff)
        if st.get("staff_role") == "manager" and st.get("perm_add_staff"):
            sh = db.get_shop_for_user(user["id"])
            if sh:
                return dict(sh)
    raise HTTPException(status_code=403, detail="not_owner")


def _staff_in_shop(shop_id, staff_id):
    for s in db.get_shop_staff(shop_id):
        if s.get("id") == staff_id:
            return s
    return None


# #3 — xodim ruxsatlari (bot PERM_KEYS pariteti): kalit -> DB ustuni
STAFF_PERM_KEYS = {"add": "perm_add_product", "conf": "perm_confirm_orders",
                   "price": "perm_edit_price", "rev": "perm_reply_reviews",
                   "staff": "perm_add_staff", "ad": "perm_publish_ad"}


@app.get("/api/seller/staff")
def api_staff(authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    shop = _staff_mgmt_shop(user)
    # Har xodimga ko'rsatkich (mahsulot/sotilgan/daromad) — bot staff_stats pariteti
    perf = {r["user_id"]: dict(r) for r in db.get_shop_staff_performance(shop["id"])}
    staff = []
    for s in db.get_shop_staff(shop["id"], include_owner=False):
        s = dict(s)
        pr = perf.get(s["user_id"], {})
        s["products_count"] = pr.get("products_count", 0)
        s["sold"] = pr.get("sold", 0)
        s["revenue"] = pr.get("revenue", 0)
        staff.append(s)
    return {"staff": staff, "invites": _rows(db.get_active_invites(shop["id"]))}


@app.post("/api/seller/staff/invite")
def api_staff_invite(authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    shop = _staff_mgmt_shop(user)
    code = db.create_invite(shop["id"], created_by=user["id"])
    return {"ok": True, "code": code}


@app.delete("/api/seller/staff/invite/{invite_id}")
def api_staff_invite_cancel(invite_id: int, authorization: str = Header(None)):
    """#8 — adashib yuborilgan faol taklifni bekor qiladi (faqat do'kon egasi)."""
    user = dict(_buyer_from_auth(authorization))
    shop = _staff_mgmt_shop(user)
    if not db.delete_invite(invite_id, shop_id=shop["id"]):
        raise HTTPException(status_code=404, detail="not_found")
    return {"ok": True}


@app.post("/api/seller/staff/{staff_id}/toggle")
def api_staff_toggle(staff_id: int, authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    shop = _staff_mgmt_shop(user)
    s = _staff_in_shop(shop["id"], staff_id)
    if not s:
        raise HTTPException(status_code=404, detail="not_found")
    if s.get("staff_role") == "owner":
        raise HTTPException(status_code=400, detail="cant_owner")
    db.set_staff_active(staff_id, 0 if s.get("is_active") else 1)
    return {"ok": True, "is_active": 0 if s.get("is_active") else 1}


@app.delete("/api/seller/staff/{staff_id}")
def api_staff_remove(staff_id: int, authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    shop = _staff_mgmt_shop(user)
    s = _staff_in_shop(shop["id"], staff_id)
    if not s:
        raise HTTPException(status_code=404, detail="not_found")
    if not db.remove_staff(staff_id):
        raise HTTPException(status_code=400, detail="cant_remove")
    return {"ok": True}


# ---- #3 XODIM DETALI + STATISTIKA (bot staff_detail/staff_stats pariteti) ----
@app.get("/api/seller/staff/{staff_id}")
def api_staff_detail(staff_id: int, authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    shop = _staff_mgmt_shop(user)
    s = _staff_in_shop(shop["id"], staff_id)
    if not s:
        raise HTTPException(status_code=404, detail="not_found")
    s = dict(s)
    stats = db.get_staff_stats(s["user_id"])
    perms = {k: bool(s.get(col)) for k, col in STAFF_PERM_KEYS.items()}
    return {"staff": s, "stats": stats, "perms": perms}


class StaffRoleIn(BaseModel):
    role: Optional[str] = None   # 'staff' | 'manager' | 'courier'; bo'sh = manager↔staff toggle


@app.post("/api/seller/staff/{staff_id}/role")
def api_staff_role(staff_id: int, body: Optional[StaffRoleIn] = None,
                   authorization: str = Header(None)):
    """Rol o'rnatadi (staff/manager/courier) yoki body'siz — manager↔staff almashtiradi."""
    user = dict(_buyer_from_auth(authorization))
    shop = _owner_shop(user)
    s = _staff_in_shop(shop["id"], staff_id)
    if not s:
        raise HTTPException(status_code=404, detail="not_found")
    if s.get("staff_role") == "owner":
        raise HTTPException(status_code=400, detail="cant_owner")
    if body and body.role:
        if body.role not in ("staff", "manager", "courier"):
            raise HTTPException(status_code=400, detail="bad_role")
        new_role = body.role
    else:
        new_role = "staff" if s.get("staff_role") == "manager" else "manager"
    db.update_staff(staff_id, staff_role=new_role)
    return {"ok": True, "staff_role": new_role}


class StaffDeptIn(BaseModel):
    department: str = ""


@app.post("/api/seller/staff/{staff_id}/dept")
def api_staff_dept(staff_id: int, body: StaffDeptIn, authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    shop = _staff_mgmt_shop(user)
    s = _staff_in_shop(shop["id"], staff_id)
    if not s:
        raise HTTPException(status_code=404, detail="not_found")
    dept = (body.department or "").strip()[:60] or None
    db.update_staff(staff_id, department=dept)
    return {"ok": True, "department": dept}


class StaffPermIn(BaseModel):
    key: str   # add | conf | price | rev


@app.post("/api/seller/staff/{staff_id}/perm")
def api_staff_perm(staff_id: int, body: StaffPermIn, authorization: str = Header(None)):
    """Bitta ruxsatni toggle qiladi (bot staff_perm pariteti)."""
    user = dict(_buyer_from_auth(authorization))
    shop = _owner_shop(user)
    s = _staff_in_shop(shop["id"], staff_id)
    if not s:
        raise HTTPException(status_code=404, detail="not_found")
    if s.get("staff_role") == "owner":
        raise HTTPException(status_code=400, detail="cant_owner")
    col = STAFF_PERM_KEYS.get(body.key)
    if not col:
        raise HTTPException(status_code=400, detail="bad_key")
    new_val = 0 if s.get(col) else 1
    db.update_staff(staff_id, **{col: new_val})
    return {"ok": True, "key": body.key, "value": bool(new_val)}


# ---- Rejalashtirilgan postlar ----
def _parse_dt(raw):
    from datetime import datetime, timezone
    raw = (raw or "").strip()
    try:
        dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except Exception:
        try:
            return datetime.strptime(raw[:19], "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
        except Exception:
            return None


class ScheduleIn(BaseModel):
    scheduled_at: str
    caption: Optional[str] = None


@app.get("/api/seller/scheduled")
def api_seller_scheduled(authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    return _rows(db.get_seller_scheduled_posts(_owner_id(user)))   # do'kon kesimi (ega+xodim)


@app.post("/api/seller/product/{product_id}/schedule")
def api_schedule_product(product_id: int, body: ScheduleIn, authorization: str = Header(None)):
    from datetime import datetime, timezone
    user = dict(_buyer_from_auth(authorization))
    prod = _own_product_or_403(user, product_id)
    if prod.get("status") in ("deleted", "purged"):
        raise HTTPException(status_code=409, detail="product_unavailable")
    # #18 Pro — Pro = cheksiz rejalashtirish; bepul sotuvchiga bir vaqtda faol limit (limit>0)
    oid = _owner_id(user)
    if _pro_locked(oid):
        limit = int(monetization_public().get("free_scheduled_limit") or 0)
        if db.count_pending_scheduled_posts(oid) >= limit:
            raise HTTPException(status_code=403, detail="pro_required")
    dt = _parse_dt(body.scheduled_at)
    if not dt:
        raise HTTPException(status_code=400, detail="bad_time")
    dt = dt.astimezone(timezone.utc)
    if dt <= datetime.now(timezone.utc):
        raise HTTPException(status_code=400, detail="time_in_past")
    sa = dt.strftime("%Y-%m-%d %H:%M:%S")
    caption = (body.caption or "").strip() or None
    # image_id=None -> bot post vaqtida reklama DIZAYNINI o'zi quradi (narx/badge/do'kon)
    db.create_scheduled_post(product_id, prod["seller_id"], sa,
                             created_by=user["id"], caption=caption,
                             parse_mode=None, image_id=None)
    db.set_product_status(product_id, "scheduled")  # belgilangan vaqtgacha yashiriladi
    return {"ok": True}


@app.post("/api/seller/scheduled/{sched_id}/cancel")
def api_cancel_scheduled(sched_id: int, authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    sp = db.cancel_scheduled_post(sched_id, _owner_id(user))   # do'kon kesimi
    if not sp:
        raise HTTPException(status_code=404, detail="not_found")
    try:
        if sp.get("product_id"):
            db.set_product_status(sp["product_id"], "active")  # bekor -> mahsulot jonli bo'ladi
    except Exception as e:
        logging.warning(f"cancel schedule status xato: {e}")
    return {"ok": True}


# ---- Avto qayta-reklama (har kuni belgilangan soatda kanal/guruhga qayta post) ----
@app.get("/api/seller/autoreposts")
def api_seller_autoreposts(authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    return _rows(db.get_seller_auto_reposts(_owner_id(user)))   # do'kon kesimi


class AutoRepostIn(BaseModel):
    hour: int
    caption: Optional[str] = None


@app.post("/api/seller/product/{product_id}/autorepost")
def api_set_autorepost(product_id: int, body: AutoRepostIn, authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    prod = _own_product_or_403(user, product_id)
    if not isinstance(body.hour, int) or not (0 <= body.hour <= 23):
        raise HTTPException(status_code=400, detail="bad_hour")
    caption = (body.caption or "").strip() or None
    # image_id=None -> bot har post'da reklama DIZAYNINI yangidan quradi
    rid = db.upsert_auto_repost(product_id, prod["seller_id"], body.hour,
                                created_by=user["id"], caption=caption,
                                parse_mode=None, image_id=None)
    return {"ok": True, "id": rid}


@app.post("/api/seller/autorepost/{repost_id}/cancel")
def api_cancel_autorepost(repost_id: int, authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    db.cancel_auto_repost(repost_id, _owner_id(user))   # do'kon kesimi
    return {"ok": True}


# ============================================================
# REKLAMA GENERATORI (bot _build_ad_caption / _build_ad_design_bytes parite)
# Sotuvchi mahsulot uchun reklama matni + dizayn rasmni app ichida ko'radi,
# tahrirlaydi va kanal/guruhlariga darhol e'lon qiladi.
# ============================================================

async def _fetch_image_bytes(file_id):
    """Telegram file_id -> rasm baytlari (disk-cache bilan). Xato bo'lsa None."""
    if not (file_id and BOT_TOKEN):
        return None
    safe = hashlib.sha256(file_id.encode()).hexdigest()
    cache_path = os.path.join(IMG_CACHE_DIR, safe + ".jpg")
    if os.path.exists(cache_path):
        try:
            with open(cache_path, "rb") as f:
                return f.read()
        except Exception:
            pass
    try:
        async with httpx.AsyncClient(timeout=20) as client:
            meta = await client.get(
                f"https://api.telegram.org/bot{BOT_TOKEN}/getFile",
                params={"file_id": file_id})
            data = meta.json()
            if not data.get("ok"):
                return None
            path = data["result"]["file_path"]
            img = await client.get(f"https://api.telegram.org/file/bot{BOT_TOKEN}/{path}")
            img.raise_for_status()
            content = img.content
        try:
            with open(cache_path, "wb") as f:
                f.write(content)
        except Exception:
            pass
        return content
    except Exception as e:
        logging.warning(f"ad image fetch xato ({str(file_id)[:12]}...): {e}")
        return None


def _region_label(product):
    try:
        return db.get_region_label(product.get("seller_region_id")) or ""
    except Exception:
        return ""


def _price_unit(product):
    """Narx + (mavjud bo'lsa) o'lchov birligi — bot price_with_unit pariteti. #20
    Optom mahsulotda birlik = 'pachka' (1 pachka narxi)."""
    s = fmt_price(product.get("price"))
    unit = (product.get("unit") or "").strip()
    if not unit and product.get("sale_mode") == "optom":
        unit = "pachka"
    return f"{s} / {unit}" if unit else s


async def _build_ad_caption_web(product, length, lang, use_ai=True):
    """Reklama matnini qaytaradi: (matn, parse_mode). use_ai=True bo'lsa avval AI
    takrorlanmas matn yozadi; bo'lmasa (yoki use_ai=False) — tuzilgan HTML matn.
    Bot _build_ad_caption bilan bir xil mantiq. Buyer sahifasi use_ai=False bilan
    AI ni KUTMAYDI (darhol chiqsin — AI matni kanal e'lonida saqlangan bo'ladi)."""
    cat = product.get("category_name") or product.get("category")
    cat_emoji = product.get("category_emoji") or "📂"
    cat_line = f"\n{cat_emoji} {html.escape(str(cat))}" if cat else ""
    shop_name = product.get("shop_name")
    shop_line = f"\n🏪 {html.escape(str(shop_name))}" if shop_name else ""
    region_lbl = _region_label(product)
    region_line = f"\n🌍 {html.escape(region_lbl)}" if region_lbl else ""
    loc = best_location_text(product.get("shop_address"), product.get("shop_landmark"))
    loc_line = f"\n📍 {html.escape(loc)}" if loc else ""
    prod_rating = product.get("prod_avg_rating") or 0
    prod_cnt = product.get("prod_review_count") or 0
    rating_line = f"\n⭐ {prod_rating:.1f} ({prod_cnt})" if prod_cnt else ""
    desc = (product.get("description") or "").strip()
    if len(desc) > 300:
        desc = desc[:300].rstrip() + "…"
    desc_line = f"\n\n📝 {html.escape(desc)}" if desc else ""

    # Do'kon murojaat telefoni — reklama matnida (bot _build_ad_caption pariteti).
    # Ijtimoiy tarmoqlar post ostida tugma bo'lib chiqadi (telefon esa matnda).
    shop_phone = (product.get("shop_phone") or "").strip()
    phone_line = f"\n📞 {html.escape(shop_phone)}" if shop_phone else ""

    # Optom (ulgurji/pachka) taklifi — yoqilgan bo'lsa zinalarni ko'rsatamiz (bot pariteti).
    _is_optom = product.get("sale_mode") == "optom"
    _w = wholesale_info(product)
    _unit = "pachka" if _is_optom else (product.get("unit") or "dona")
    _u_esc = html.escape(str(_unit))
    if _w["enabled"]:
        _tier_txt = ", ".join(f"{t['min']}+ {_u_esc} — {html.escape(fmt_price(t['price']))}"
                              for t in _w["tiers"])
        wholesale_line = f"\n📦 Optom narx: {_tier_txt}"
    else:
        wholesale_line = ""

    # Optom (pachka) maxsus qatorlari: pachka belgisi, 1 pachka = N dona, ranglar, razmer.
    optom_lines, extra_facts = "", {}
    if _is_optom:
        _ps = product.get("pack_size")
        try:
            _colors = _product_color_options(product.get("id")) if product.get("id") else []
        except Exception:
            _colors = []
        _sn = (product.get("size_note") or "").strip()
        _pack_line = f"\n📦 1 pachka = {int(_ps)} dona" if _ps else ""
        _colors_line = f"\n🎨 Ranglar: {html.escape(', '.join(_colors))}" if _colors else ""
        _size_line = f"\n📏 Razmer: {html.escape(_sn)}" if _sn else ""
        optom_lines = f"\n🏷 Optom (pachkada sotiladi){_pack_line}{_colors_line}{_size_line}"
        extra_facts = {"savdo turi": "optom (ulgurji, pachkada sotiladi)"}
        if _ps:
            extra_facts["bitta pachka"] = f"{int(_ps)} dona"
        if _colors:
            extra_facts["mavjud ranglar"] = ", ".join(_colors)
        if _sn:
            extra_facts["razmerlar"] = _sn

    caption = (
        f"🆕 <b>{html.escape(product.get('name') or '')}</b>"
        f"\n💵 {_price_unit(product)}{wholesale_line}{optom_lines}"
        f"{cat_line}{shop_line}{region_line}{loc_line}{phone_line}{rating_line}{desc_line}")
    parse_mode = "HTML"
    if not use_ai:
        return caption, parse_mode  # darhol — AI ni kutmaymiz
    try:
        ad_text = await ai_assistant.generate_ad_caption(
            name=product.get("name") or "",
            price_text=_price_unit(product),
            category=str(cat) if cat else "",
            description=(product.get("description") or ""),
            shop=str(shop_name) if shop_name else "",
            region=region_lbl or "",
            location=loc or "",
            lang=lang,
            length=length,
            extra=extra_facts)
    except Exception as e:
        logging.warning(f"reklama matni (web) olinmadi: {e}")
        ad_text = None
    if ad_text:
        ad_text = ad_text.rstrip()
        # Optom: pachka/ranglar/razmer faktlarini AI matniga ham KAFOLATLI qo'shamiz (oddiy matn).
        if _is_optom:
            _plain = []
            if product.get("pack_size"):
                _plain.append(f"📦 1 pachka = {int(product['pack_size'])} dona")
            if extra_facts.get("mavjud ranglar"):
                _plain.append(f"🎨 Ranglar: {extra_facts['mavjud ranglar']}")
            if extra_facts.get("razmerlar"):
                _plain.append(f"📏 Razmer: {extra_facts['razmerlar']}")
            if _plain:
                ad_text += "\n\n" + "\n".join(_plain)
        # Optom taklifini (zina narxlari) AI matniga ham kafolatli qo'shamiz (oddiy matn).
        if _w["enabled"]:
            _tiers_plain = ", ".join(f"{t['min']}+ {_unit} — {fmt_price(t['price'])}" for t in _w["tiers"])
            ad_text = ad_text.rstrip() + f"\n\n📦 Optom narx: {_tiers_plain}"
        # Murojaat telefonini AI matniga ham kafolatli qo'shamiz (bot pariteti — oddiy matn).
        if shop_phone:
            ad_text = ad_text.rstrip() + f"\n📞 {shop_phone}"
        return ad_text, None  # AI matni oddiy matn (HTML emas)
    return caption, parse_mode


async def _build_ad_design_web(product):
    """Mahsulot rasmiga reklama dizayni qo'yib JPEG bytes qaytaradi (yoki None)."""
    photo = product.get("image_url")
    if not (photo and ad_design.is_enabled()):
        return None
    raw = await _fetch_image_bytes(photo)
    if not raw:
        return None
    shop_name = product.get("shop_name")
    region_lbl = _region_label(product)
    # Optom rozetkasi — pachka belgisi (rasm ustida ko'rinadi)
    optom_txt = ""
    if product.get("sale_mode") == "optom":
        _ps = product.get("pack_size")
        optom_txt = f"OPTOM · 1 PACHKA = {int(_ps)} DONA" if _ps else "OPTOM"
    try:
        return await asyncio.to_thread(
            ad_design.build_ad_image, raw,
            price_text=fmt_price(product.get("price")),
            badge_text="",
            shop_text=(str(shop_name) if shop_name else (region_lbl or "")),
            optom_text=optom_txt)
    except Exception as e:
        logging.warning(f"reklama dizayni (web) yasalmadi: {e}")
        return None


@app.get("/api/seller/product/{product_id}/ad-preview")
async def api_ad_preview(product_id: int, length: str = Query("long"),
                         authorization: str = Header(None)):
    """Reklama ko'rinishi: dizayn rasm (base64) + AI reklama matni. Faqat ko'rish."""
    user = dict(_buyer_from_auth(authorization))
    prod = _own_product_or_403(user, product_id)
    length = length if length in ("long", "short") else "long"
    lang = get_user_lang(user) or DEFAULT_LANG
    caption, parse_mode = await _build_ad_caption_web(prod, length, lang)
    # Video bor mahsulotda kanal posti VIDEO bilan chiqadi — dizayn rasm yasalmaydi,
    # preview'da ham video ko'rsatiladi (post bilan 100% mos bo'lsin). 🎬 AI
    # video-reklama klipi sotuvchi videosidan ustun (post_product_to_channel parite).
    video_fid = prod.get("ad_video_file_id") or prod.get("video_file_id")
    design = None if video_fid else await _build_ad_design_web(prod)
    image = None
    if design:
        image = "data:image/jpeg;base64," + base64.b64encode(design).decode("ascii")
    return {"caption": caption, "parse_mode": parse_mode,
            "image": image, "has_design": bool(design),
            "has_video": bool(video_fid), "video_file_id": video_fid}


class AdPublishIn(BaseModel):
    caption: Optional[str] = None
    length: Optional[str] = "long"
    image_id: Optional[str] = None   # 🎨 AI banner file_id (ai-banner endpointi bergan)
    video_id: Optional[str] = None   # 🎬 AI video-reklama klipi file_id (adclip endpointi bergan)


@app.post("/api/seller/product/{product_id}/ad-publish")
def api_ad_publish(product_id: int, body: AdPublishIn, authorization: str = Header(None)):
    """Reklamani DARHOL kanal/guruhlarga e'lon qiladi. Mavjud rejalashtirilgan-post
    mexanizmidan foydalanadi: hozirgi vaqtga 'pending' post yaratiladi, bot
    webapp_scheduled_scan_job (har ~30s) uni topib post_product_to_channel'ni
    ishga tushiradi. image_id=None -> bot reklama DIZAYNINI o'zi quradi."""
    user = dict(_buyer_from_auth(authorization))
    prod = _own_product_or_403(user, product_id)
    if prod.get("status") in ("deleted", "purged"):
        raise HTTPException(status_code=409, detail="product_unavailable")
    _rate_limit("ad_publish", user.get("id"), 10, 3600)
    caption = (body.caption or "").strip() or None
    # Sotuvchi tahrir qilgan matn — oddiy matn sifatida yuboriladi (parse_mode yo'q,
    # buzilgan HTML xavfi yo'q). Bo'sh bo'lsa caption=None -> bot AI matnini quradi.
    image_id = (body.image_id or "").strip()[:200] or None   # AI banner bo'lsa — o'sha rasm
    # 🎬 AI video-reklama klipi tanlangan bo'lsa — mahsulotga bog'laymiz; bot posti
    # (post_product_to_channel) uni video sifatida chiqaradi va keyingi qayta
    # reklamalar (avto-repost ham) shu klip bilan ketadi.
    video_id = (body.video_id or "").strip()[:200] or None
    if video_id:
        db.set_product_ad_video(product_id, video_id)
    sa = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    db.create_scheduled_post(product_id, prod["seller_id"], sa,
                             created_by=user["id"], caption=caption,
                             parse_mode=None, image_id=image_id)
    return {"ok": True}


@app.get("/api/seller/product/{product_id}/reels")
async def api_seller_reels(product_id: int, authorization: str = Header(None)):
    """#4 — mahsulot uchun Reels/TikTok video ssenariysi (AI). Faqat egasi (yoki xodimi)."""
    user = dict(_buyer_from_auth(authorization))
    prod = _own_product_or_403(user, product_id)
    if not ai_assistant.is_enabled():
        raise HTTPException(status_code=503, detail="ai_disabled")
    # #18 Pro — Pro = cheksiz reels; bepul sotuvchiga oylik kvota (limit>0). Obuna o'chiq
    # bo'lsa _pro_locked False — cheklov yo'q (eski xulq).
    oid = _owner_id(user)
    free_capped = _pro_locked(oid)
    if free_capped:
        limit = int(monetization_public().get("free_reels_limit") or 0)
        if db.get_feature_usage(oid, "reels", _month_key()) >= limit:
            raise HTTPException(status_code=403, detail="pro_required")
    _rate_limit("reels", user.get("id"), 20, 3600)
    lang = get_user_lang(user) or DEFAULT_LANG
    cat = prod.get("category_name") or ""
    script = await ai_assistant.generate_video_script(
        name=prod.get("name") or "", price_text=fmt_price(prod.get("price")),
        category=cat, description=prod.get("description") or "", lang=lang)
    if not script:
        raise HTTPException(status_code=502, detail="ai_error")
    if free_capped:   # faqat bepul sotuvchining kvotasini sarflaymiz
        db.incr_feature_usage(oid, "reels", _month_key())
    return {"script": script}


class ReplyIn(BaseModel):
    text: str


@app.get("/api/seller/review/{review_id}/ai-reply")
async def api_review_ai_reply(review_id: int, authorization: str = Header(None)):
    """Sharhga AI javob matnini TAKLIF qiladi (bot ai_review_reply_generate parite)."""
    user = dict(_buyer_from_auth(authorization))
    review = db.get_review_by_id(review_id)
    if not review or dict(review).get("seller_id") != user["id"]:
        raise HTTPException(status_code=403, detail="not_your_review")
    if not ai_assistant.is_enabled():
        raise HTTPException(status_code=503, detail="ai_disabled")
    review = dict(review)
    lang = get_user_lang(user) or DEFAULT_LANG
    reply = await ai_assistant.generate_review_reply(
        product=review.get("product_name") or "", comment=review.get("comment") or "",
        shop_rating=review.get("rating"), product_rating=review.get("product_rating"),
        buyer=review.get("buyer_name") or "", lang=lang)
    if not reply:
        raise HTTPException(status_code=502, detail="ai_error")
    return {"reply": reply}


@app.post("/api/seller/review/{review_id}/reply")
async def api_review_reply(review_id: int, body: ReplyIn, authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    text = (body.text or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="empty")
    if len(text) > 1000:
        raise HTTPException(status_code=400, detail="too_long")
    # Multivendor: sharh do'kon (ega) ostida — ega yoki "sharhga javob" ruxsatli xodim javob yozadi
    if not _staff_perm(user, "perm_reply_reviews"):
        raise HTTPException(status_code=403, detail="no_perm_reply")
    if not db.set_review_reply(review_id, _owner_id(user), text):
        raise HTTPException(status_code=403, detail="not_your_review")
    try:
        rev = db.get_review_by_id(review_id)
        buyer = db.get_user_by_id(rev["buyer_id"]) if rev and rev.get("buyer_id") else None
        if buyer and buyer.get("telegram_id"):
            blang = get_user_lang(buyer)
            pname = html.escape(rev.get("product_name") or "")
            if blang == "ru":
                msg = f"💬 Продавец ответил на ваш отзыв ({pname}):\n\n{html.escape(text)}"
            else:
                msg = f"💬 Sotuvchi sharhingizga javob berdi ({pname}):\n\n{html.escape(text)}"
            await _tg_call("sendMessage", {"chat_id": buyer["telegram_id"], "text": msg,
                                           "parse_mode": "HTML"})
    except Exception as e:
        logging.warning(f"review reply notify xato (review {review_id}): {e}")
    return {"ok": True}


@app.get("/api/seller/pending")
def api_seller_pending(authorization: str = Header(None)):
    """Ega tasdig'ini kutayotgan (xodim joylagan) mahsulotlar."""
    user = dict(_buyer_from_auth(authorization))
    return _rows(db.get_seller_products_by_status(_owner_id(user), "pending_owner"))


class ApproveIn(BaseModel):
    approve: bool = True


@app.post("/api/seller/product/{product_id}/approve")
def api_approve_product(product_id: int, body: ApproveIn, authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    prod = db.get_product_by_id(product_id)
    if not prod:
        raise HTTPException(status_code=404, detail="not_found")
    if prod.get("seller_id") != user.get("id") and user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="not_owner")
    if prod.get("status") != "pending_owner":
        raise HTTPException(status_code=409, detail="not_pending")
    if not body.approve:
        db.set_product_status(product_id, "deleted")
        return {"ok": True, "approved": False}
    # TASDIQLANDI — mahsulot faollashadi VA reklama AVTOMATIK kanal/guruhlarga chiqadi.
    # Ega tasdig'i = "e'lon qil" ruxsati: xodim pending_owner tufayli reklama ekraniga
    # o'tolmagan, shuning uchun tasdiqda uning o'rniga biz e'lonni yuboramiz (aks holda
    # tayyor reklama hech qayerga chiqmasdi). api_ad_publish bilan bir xil mexanizm:
    # hozirgi vaqtga 'pending' post -> bot scan job (~30s) topib post_product_to_channel'ni
    # ishga tushiradi (markaziy kanal + sotuvchining barcha ulangan kanal/guruhlari).
    # image_id/caption=None -> bot AI reklama dizayni + matnini o'zi quradi.
    db.set_product_status(product_id, "active")
    try:
        sa = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        db.create_scheduled_post(product_id, prod["seller_id"], sa,
                                 created_by=(prod.get("created_by") or user["id"]),
                                 caption=None, parse_mode=None, image_id=None)
    except Exception as e:
        logging.warning(f"tasdiqda avto-reklama post yaratilmadi (pid {product_id}): {e}")
    return {"ok": True, "approved": True}


@app.get("/api/seller/stats")
def api_seller_stats(authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    oid = _owner_id(user)   # do'kon (ega) kesimida
    raw = db.get_seller_stats(oid) or {}
    stats = dict(raw)
    try:
        stats["avg_rating"] = round(float(db.get_seller_avg_rating(oid) or 0), 1)
    except Exception:
        stats["avg_rating"] = 0
    debts = db.get_seller_open_debts(oid) or []
    stats["open_debt_total"] = sum(float(d.get("total_due") or 0) for d in debts)
    stats["open_debt_count"] = len(debts)
    return stats


@app.get("/api/seller/product/{product_id}/price-insight")
def api_price_insight(product_id: int, authorization: str = Header(None)):
    """AI #2 — raqobatchi narx maslahati: kategoriya o'rtachasiga nisbatan baho.
    #18 Pro — kengaytirilgan analitika (raqobat narxi) faqat Pro-obunada."""
    user = dict(_buyer_from_auth(authorization))
    prod = _own_product_or_403(user, product_id)
    if _pro_locked(_owner_id(user)):
        raise HTTPException(status_code=403, detail="pro_required")
    cat_id = prod.get("category_id")
    price = float(prod.get("price") or 0)
    stats = db.get_category_price_stats(cat_id, exclude_seller_id=prod.get("seller_id"))
    if not stats or not stats.get("count"):
        return {"available": False}
    avg = float(stats["avg"]) or 0
    verdict, diff_pct = "fair", 0
    if avg > 0:
        diff_pct = round((price - avg) / avg * 100)
        if diff_pct >= 15:
            verdict = "expensive"
        elif diff_pct <= -15:
            verdict = "cheap"
    return {"available": True, "your_price": round(price), "diff_pct": diff_pct,
            "verdict": verdict, **stats}


@app.get("/api/seller/product/{product_id}/dynamic-price")
async def api_dynamic_price(product_id: int, authorization: str = Header(None)):
    """#11 — dinamik narx: talab + raqobat signallaridan AI narx tavsiyasi (ko'tar/tushir/qoldir).
    #18 Pro — AI dinamik narx faqat Pro-obunada (kengaytirilgan analitika)."""
    user = dict(_buyer_from_auth(authorization))
    prod = _own_product_or_403(user, product_id)
    if _pro_locked(_owner_id(user)):
        raise HTTPException(status_code=403, detail="pro_required")
    if not ai_assistant.is_enabled():
        raise HTTPException(status_code=503, detail="ai_disabled")
    _rate_limit("dyn_price", user.get("id"), 30, 3600)
    signals = db.get_product_demand_signals(product_id) or {}
    competitor = db.get_category_price_stats(
        prod.get("category_id"), exclude_seller_id=prod.get("seller_id"))
    lang = get_user_lang(user) or DEFAULT_LANG
    advice = await ai_assistant.dynamic_price_advice(
        name=prod.get("name") or "", price=float(prod.get("price") or 0),
        signals=signals, competitor=competitor,
        category=prod.get("category_name") or "", lang=lang)
    if not advice:
        raise HTTPException(status_code=502, detail="ai_error")
    return {"available": True, "current_price": round(float(prod.get("price") or 0)),
            "signals": signals, "competitor": competitor or None, **advice}


@app.get("/api/seller/analytics")
def api_seller_analytics(authorization: str = Header(None)):
    """#17 — sotuvchi dashboard: top mahsulotlar (nima) + hafta kuni/kunlik (qachon).
    #18 Pro — chuqur tahlil (hafta kuni + 7 kunlik daromad) faqat Pro-obunada ko'rinadi.
    Obuna monetizatsiyasi yoqilmagan bo'lsa hammasi ochiq qoladi (eski xulq)."""
    user = dict(_buyer_from_auth(authorization))
    oid = _owner_id(user)
    perf = [dict(p) for p in db.get_seller_product_performance(oid)]
    top = [p for p in perf if (p.get("sold") or 0) > 0][:5]
    # "Top mahsulotlar" har doim bepul (teaser). Chuqur tahlil — obuna yoqilgan va
    # sotuvchi Pro bo'lmaganda qulflanadi.
    sub_on = bool(monetization_public().get("subscription"))
    is_pro = _is_pro(oid)
    if sub_on and not is_pro:
        # Pro qulfi: BUTUN tahlil (🏆 Eng ko'p sotilgan ham) yopiq — bepul teaser yo'q.
        return {"top_products": [], "by_weekday": None, "daily_7": None,
                "pro_locked": True, "is_pro": False}
    t = db.get_seller_time_analytics(oid)
    return {"top_products": top, "by_weekday": t["by_weekday"], "daily_7": t["daily_7"],
            "pro_locked": False, "is_pro": is_pro}


_SHOP_FIELDS = ("shop_name", "shop_address", "shop_landmark", "working_days",
                "working_hours", "work_schedule", "phone_number", "card_number", "card_owner",
                "card_type", "shop_lat", "shop_lon", "telegram_username",
                "delivery_min_total")
_VALID_CARD = {"uzcard", "humo", "visa", "mastercard"}


@app.get("/api/seller/shop")
def api_get_shop(authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    out = {k: user.get(k) for k in _SHOP_FIELDS}
    out["region_id"] = user.get("region_id")
    # #2 — hudud yorlig'i ("Viloyat → Tuman"); mavjud db helperi (admin'da ham ishlatiladi)
    out["region_label"] = db.get_region_label(user.get("region_id")) or ""
    shop = db.get_shop_by_owner(user["id"])
    out["payment_mode"] = (dict(shop).get("payment_mode") if shop else None) or "shop"
    out["is_owner"] = bool(shop)
    # Do'kon ulashish havolasi uchun — mahsulotlar ostida ko'rinadigan EGA id'si (xodim ham shu id ostida)
    out["seller_id"] = _owner_id(user)
    return out


# ============================================================
# 🧠 AI DIREKTOR — sotuvchiga aqlli hisobot + 1-tap amallar
# ============================================================
def _director_facts(oid):
    """AI uchun kompakt, REAL do'kon faktlari. AI faqat shulardan tavsiya beradi."""
    stats = db.get_seller_stats(oid) or {}
    perf = {p["id"]: dict(p) for p in db.get_seller_product_performance(oid)}
    prods = [dict(p) for p in db.get_products_by_seller(oid)]
    items, stale, low_stock = [], [], []
    for p in prods:
        if p.get("status") != "active":
            continue
        sold = (perf.get(p["id"]) or {}).get("sold", 0)
        it = {"product_id": p["id"], "name": p.get("name"), "price": p.get("price"),
              "sold": sold, "stock": p.get("stock_count"),
              "created": str(p.get("created_at") or "")[:10]}
        items.append(it)
        if sold == 0:
            stale.append(it)
        sc = p.get("stock_count")
        if sc is not None and 0 < sc <= 2:
            low_stock.append(it)
    items.sort(key=lambda x: -(x["sold"] or 0))
    return {
        "shop_stats": {"total_orders": stats.get("total_orders"),
                       "pending_orders": stats.get("pending"),
                       "week_orders": stats.get("week_orders"),
                       "week_revenue": stats.get("week_revenue"),
                       "month_revenue": stats.get("month_revenue"),
                       "active_products": len(items)},
        "top_products": items[:5],
        "not_selling_products": stale[:6],
        "low_stock_products": low_stock[:5],
        "open_debts_count": len(db.get_seller_open_debts(oid) or []),
    }


@app.get("/api/seller/ai-director")
async def api_ai_director(authorization: str = Header(None)):
    """🧠 AI direktor hisoboti: summary + tavsiyalar (har birida 1-tap action bo'lishi mumkin)."""
    user = dict(_buyer_from_auth(authorization))
    if user.get("role") not in ("seller", "admin") and not user.get("is_approved"):
        raise HTTPException(status_code=403, detail="not_seller")
    if not ai_assistant.is_enabled():
        raise HTTPException(status_code=503, detail="ai_disabled")
    _rate_limit("ai_director", user.get("id"), 10, 3600)
    oid = _owner_id(user)
    facts = _director_facts(oid)
    if not facts["shop_stats"]["active_products"] and not facts["shop_stats"]["total_orders"]:
        return {"available": False}
    lang = get_user_lang(user) or DEFAULT_LANG
    rep = await ai_assistant.director_report(facts=facts, lang=lang)
    if not rep:
        raise HTTPException(status_code=502, detail="ai_error")
    # Amallar xavfsizligi: product_id lar shu do'konnikimi — tekshirib chiqamiz
    own_ids = {p["product_id"] for p in facts["top_products"]} | \
              {p["product_id"] for p in facts["not_selling_products"]} | \
              {p["product_id"] for p in facts["low_stock_products"]}
    for r in rep["recs"]:
        a = r.get("action")
        if a and a.get("product_id") not in own_ids:
            r["action"] = None
    return {"available": True, **rep}


@app.get("/api/seller/deep-report")
async def api_deep_report(authorization: str = Header(None)):
    """📊 Chuqur biznes-hisobot: BUTUN savdo tarixi bitta so'rovda AI'ga —
    trend, «qaysi mahsulot qachon sotiladi», harakat rejasi (Gemini katta kontekst)."""
    user = dict(_buyer_from_auth(authorization))
    _seller_or_403(user)
    lang = get_user_lang(user) or DEFAULT_LANG
    if not ai_assistant.is_enabled():
        raise HTTPException(status_code=503, detail="ai_disabled")
    _rate_limit("ai_deep_report", user["id"], 4, 3600)
    oid = _owner_id(user)
    orders = db.get_orders_by_seller(oid) or []
    if len(orders) < 5:
        return {"available": False, "reason": "not_enough_data",
                "orders_count": len(orders)}
    # AI uchun KOMPAKT tarix: shaxsiy ma'lumotlarsiz (ism/telefon yubormaymiz),
    # takroriy xaridorni aniqlash uchun buyer anonim raqam bilan belgilanadi.
    buyer_alias = {}
    hist = []
    for o in orders:
        o = dict(o)
        b = o.get("buyer_id")
        if b not in buyer_alias:
            buyer_alias[b] = f"B{len(buyer_alias) + 1}"
        hist.append({
            "dt": str(o.get("created_at") or "")[:16],
            "product": o.get("product_name"),
            "qty": o.get("quantity"),
            "total": o.get("total_price"),
            "status": o.get("status"),
            "pay": o.get("payment_method"),
            "delivery": o.get("delivery_type"),
            "buyer": buyer_alias[b],
        })
    perf = [{"name": p.get("name"), "price": p.get("price"),
             "sold": p.get("sold"), "revenue": p.get("revenue")}
            for p in (db.get_seller_product_performance(oid) or [])]
    facts = {"shop_stats": db.get_seller_stats(oid) or {},
             "products": perf, "orders": hist}
    rep = await ai_assistant.deep_director_report(facts=facts, lang=lang)
    if not rep:
        raise HTTPException(status_code=502, detail="ai_error")
    return {"available": True, "orders_analyzed": len(hist), **rep}


class DirectorApplyIn(BaseModel):
    type: str                 # 'price' | 'ad'  ('restock' — client o'zi formani ochadi)
    product_id: int
    value: Optional[float] = None   # price uchun yangi narx


@app.post("/api/seller/ai-director/apply")
def api_ai_director_apply(body: DirectorApplyIn, authorization: str = Header(None)):
    """AI direktor tavsiyasini 1 bosishda qo'llash: narx tushirish yoki reklama joylash."""
    user = dict(_buyer_from_auth(authorization))
    prod = _own_product_or_403(user, body.product_id)
    if body.type == "price":
        newp = float(body.value or 0)
        oldp = float(prod.get("price") or 0)
        if newp <= 0 or newp >= oldp:
            raise HTTPException(status_code=400, detail="bad_price")
        if newp < oldp * 0.5:   # AI adashsa ham narx 50%dan ko'p tushmasin (himoya)
            raise HTTPException(status_code=400, detail="too_low")
        db.update_product_fields(body.product_id, price=newp, old_price=oldp)
        return {"ok": True, "applied": "price", "new_price": newp}
    if body.type == "ad":
        sa = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        db.create_scheduled_post(body.product_id, prod["seller_id"], sa,
                                 created_by=user["id"], caption=None,
                                 parse_mode=None, image_id=None)
        return {"ok": True, "applied": "ad"}
    raise HTTPException(status_code=400, detail="bad_type")


# ============================================================
# 🤖 AI SMM — kanalga kunlik avtomatik mavzuli post (sozlama)
# ============================================================
class SmmIn(BaseModel):
    is_active: Optional[bool] = None
    hour: Optional[int] = None            # eski mijoz mosligi (bitta soat)
    hours: Optional[List[int]] = None     # Pro: kunlik post soatlari (0-23, istalgancha)
    days: Optional[List[int]] = None      # Pro: faol hafta kunlari 0-6 ([] = har kuni)
    tone: Optional[str] = None            # Pro: yozish ohangi
    strategy: Optional[str] = None        # Pro: mahsulot tanlash strategiyasi
    custom_note: Optional[str] = None     # Pro: brend-ovoz talabi ('' = tozalash)


_SMM_TONES = {"friendly", "premium", "energetic", "minimal"}
_SMM_STRATEGIES = {"rotation", "random", "new", "discount"}


def _smm_settings_out(oid):
    """Sozlamani App uchun tayyorlaydi: hours/days ro'yxat ko'rinishida."""
    s = dict(db.get_smm_settings(oid) or
             {"hour": 10, "is_active": 0, "last_posted_at": None, "last_product_id": None})
    s["hours"] = db._smm_hours_list(s)
    days = [p for p in str(s.get("days") or "").split(",") if p.strip().isdigit()]
    s["days"] = sorted({int(p) for p in days if 0 <= int(p) <= 6})   # [] = har kuni
    s["tone"] = s.get("tone") or "friendly"
    s["strategy"] = s.get("strategy") or "rotation"
    s["custom_note"] = s.get("custom_note") or ""
    return s


def _smm_require_seller(authorization):
    user = dict(_buyer_from_auth(authorization))
    if user.get("role") not in ("seller", "admin") and not user.get("is_approved"):
        raise HTTPException(status_code=403, detail="not_seller")
    return user


async def _smm_build_caption(oid, s, product, user):
    """Sozlamalarga mos SMM matni (preview va darhol joylash uchun umumiy)."""
    owner = db.get_user_by_id(oid)
    prod = dict(db.get_product_by_id(product["id"]) or product)
    old_p = prod.get("old_price")
    has_disc = bool(old_p and float(old_p) > float(prod.get("price") or 0))
    tz_tk = timezone(timedelta(hours=5))   # Toshkent
    return await ai_assistant.generate_smm_caption(
        name=prod.get("name") or "", price_text=fmt_price(prod.get("price")),
        category=prod.get("category_name") or "",
        description=prod.get("description") or "",
        shop_name=(dict(owner).get("shop_name") if owner else "") or "",
        weekday=datetime.now(tz_tk).weekday(),
        lang=get_user_lang(user) or DEFAULT_LANG,
        tone=s.get("tone") or "friendly",
        custom_note=s.get("custom_note") or "",
        old_price_text=fmt_price(old_p) if has_disc else "",
        stock_count=prod.get("stock_count"))


@app.get("/api/seller/smm")
def api_smm_get(authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    oid = _owner_id(user)
    s = _smm_settings_out(oid)
    chans = len(db.get_active_seller_channels(oid) or [])
    hist = db.get_smm_posts(oid, limit=5)
    prods = [p for p in (db.get_products_by_seller(oid) or [])
             if dict(p).get("status") == "active" and dict(p).get("in_stock")]
    return {**s, "channels": chans, "ai_enabled": ai_assistant.is_enabled(),
            "history": hist["items"], "total_posts": hist["total"],
            "active_products": len(prods)}


@app.post("/api/seller/smm")
def api_smm_set(body: SmmIn, authorization: str = Header(None)):
    user = _smm_require_seller(authorization)
    oid = _owner_id(user)
    kw = {}
    if body.hour is not None:
        if not (0 <= int(body.hour) <= 23):
            raise HTTPException(status_code=400, detail="bad_hour")
        kw["hour"] = int(body.hour)
    if body.hours is not None:
        hrs = sorted({int(h) for h in body.hours})
        if not hrs or any(not (0 <= h <= 23) for h in hrs):
            raise HTTPException(status_code=400, detail="bad_hours")
        kw["hours"] = ",".join(str(h) for h in hrs)
        kw["hour"] = hrs[0]   # eski ustun ham sinxron qoladi
    if body.days is not None:
        dys = sorted({int(d) for d in body.days})
        if any(not (0 <= d <= 6) for d in dys):
            raise HTTPException(status_code=400, detail="bad_days")
        kw["days"] = ",".join(str(d) for d in dys)   # bo'sh = har kuni
    if body.tone is not None:
        if body.tone not in _SMM_TONES:
            raise HTTPException(status_code=400, detail="bad_tone")
        kw["tone"] = body.tone
    if body.strategy is not None:
        if body.strategy not in _SMM_STRATEGIES:
            raise HTTPException(status_code=400, detail="bad_strategy")
        kw["strategy"] = body.strategy
    if body.custom_note is not None:
        kw["custom_note"] = str(body.custom_note).strip()[:200]
    db.upsert_smm_settings(oid, is_active=body.is_active, **kw)
    return {"ok": True, **_smm_settings_out(oid)}


@app.get("/api/seller/smm/preview")
async def api_smm_preview(authorization: str = Header(None)):
    """Navbatdagi post QANDAY chiqishini ko'rsatadi (kanalga yubormaydi)."""
    user = _smm_require_seller(authorization)
    if not ai_assistant.is_enabled():
        raise HTTPException(status_code=503, detail="ai_disabled")
    _rate_limit("smm_preview", user.get("id"), 6, 3600)
    oid = _owner_id(user)
    s = _smm_settings_out(oid)
    nxt = db.get_next_smm_product(oid, s.get("last_product_id"),
                                  strategy=s.get("strategy") or "rotation")
    if not nxt:
        return {"available": False}
    cap = await _smm_build_caption(oid, s, nxt, user)
    if not cap:
        raise HTTPException(status_code=502, detail="ai_error")
    return {"available": True, "caption": cap,
            "product_id": nxt["id"], "product_name": nxt.get("name")}


@app.post("/api/seller/smm/post-now")
async def api_smm_post_now(authorization: str = Header(None)):
    """Navbatdagi mahsulotni AI matn bilan HOZIROQ kanalga joylaydi (bot
    scheduled-post oqimi orqali ~30 soniyada chiqadi). Kunlik limit — 3."""
    user = _smm_require_seller(authorization)
    _rate_limit("smm_post_now", user.get("id"), 3, 86400)
    oid = _owner_id(user)
    s = _smm_settings_out(oid)
    nxt = db.get_next_smm_product(oid, s.get("last_product_id"),
                                  strategy=s.get("strategy") or "rotation")
    if not nxt:
        raise HTTPException(status_code=400, detail="no_product")
    cap = None
    if ai_assistant.is_enabled():
        try:
            cap = await _smm_build_caption(oid, s, nxt, user)
        except Exception:
            cap = None   # AI xatosi post'ni to'xtatmasin — standart reklama matni chiqadi
    sa = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    db.create_scheduled_post(nxt["id"], oid, sa, created_by=user["id"],
                             caption=cap, parse_mode=None, image_id=None)
    # rotatsiya oldinga suriladi; bugungi avtomatik slot holati o'zgarmaydi
    db.mark_smm_posted(oid, nxt["id"], slot_key=dict(db.get_smm_settings(oid) or {}).get("last_slot"))
    db.log_smm_post(oid, nxt["id"], cap)
    return {"ok": True, "product_id": nxt["id"], "product_name": nxt.get("name")}


class ShopEdit(BaseModel):
    shop_name: Optional[str] = None
    shop_address: Optional[str] = None
    shop_landmark: Optional[str] = None
    working_days: Optional[str] = None
    working_hours: Optional[str] = None
    # Haftalik jadval (per-kun soat). dict {"0":"09:00-21:00",...} yoki JSON-matn; "" / {} = o'chirish
    work_schedule: Optional[Any] = None
    phone: Optional[str] = None
    card_number: Optional[str] = None
    card_owner: Optional[str] = None
    card_type: Optional[str] = None
    payment_mode: Optional[str] = None   # 'shop' | 'staff' (multivendor karta yo'nalishi)
    region_id: Optional[int] = None       # #2 — hudud (viloyat/tuman); 0 yoki -1 = tozalash
    lat: Optional[float] = None
    lon: Optional[float] = None
    telegram_username: Optional[str] = None  # sotuvchi kontakt username (bot edit_telegram pariteti)
    delivery_min_total: Optional[float] = None  # yetkazish uchun minimal buyurtma summasi (0 = cheklov yo'q)
    # "Aloqa bloki" — reklamaga avtomatik qo'shiladigan murojaat kanallari
    shop_phone: Optional[str] = None       # murojaat telefoni (reklama matnida ko'rinadi)
    shop_instagram: Optional[str] = None   # Instagram (post ostida tugma)
    shop_telegram: Optional[str] = None    # Telegram kanal/username (post ostida tugma)
    shop_website: Optional[str] = None     # universal havola (post ostida tugma)


@app.patch("/api/seller/shop")
def api_edit_shop(p: ShopEdit, authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    if user.get("role") not in ("seller", "admin") and not user.get("is_approved"):
        raise HTTPException(status_code=403, detail="not_seller")
    # #10 — do'kon sozlamalarini faqat EGA o'zgartira oladi (xodim/sotuvchi emas).
    # Admin (platforma) ham o'tadi. Egasi bo'lmagan sotuvchi → 403 not_owner.
    if user.get("role") != "admin" and not db.get_shop_by_owner(user["id"]):
        raise HTTPException(status_code=403, detail="not_owner")
    fields = {}
    if p.shop_name is not None:
        nm = p.shop_name.strip()
        if not nm:
            raise HTTPException(status_code=400, detail="name_required")
        fields["shop_name"] = nm
    for attr, col in (("shop_address", "shop_address"), ("shop_landmark", "shop_landmark"),
                      ("working_days", "working_days"), ("working_hours", "working_hours"),
                      ("card_owner", "card_owner")):
        v = getattr(p, attr)
        if v is not None:
            fields[col] = v.strip() or None
    # Haftalik jadval (per-kun soat). Client dict/JSON yuboradi; serverda tekshirib kanonik
    # JSON'ga aylantiramiz. Bo'sh/yaroqsiz/{} → NULL (eski yagona working_hours ishlatiladi).
    if p.work_schedule is not None:
        raw = p.work_schedule
        if isinstance(raw, str) and not raw.strip():
            fields["work_schedule"] = None
        elif isinstance(raw, dict) and not raw:
            fields["work_schedule"] = None
        else:
            fields["work_schedule"] = normalize_schedule(raw)
    if p.phone is not None and p.phone.strip():
        ph = p.phone.strip()
        digits = ph.lstrip("+")
        if not digits.isdigit() or not (7 <= len(digits) <= 15):
            raise HTTPException(status_code=400, detail="bad_phone")
        fields["phone_number"] = ph
    if p.card_number is not None:
        cn = p.card_number.replace(" ", "").strip()
        if cn:
            if not cn.isdigit() or not (12 <= len(cn) <= 19):
                raise HTTPException(status_code=400, detail="bad_card")
            fields["card_number"] = cn
        else:
            fields["card_number"] = None
    if p.card_type is not None:
        if p.card_type and p.card_type not in _VALID_CARD:
            raise HTTPException(status_code=400, detail="bad_card_type")
        fields["card_type"] = p.card_type or None
    if p.lat is not None:
        fields["shop_lat"] = p.lat
    if p.lon is not None:
        fields["shop_lon"] = p.lon
    if p.region_id is not None:
        # 0/-1 = hududni tozalash; aks holda mavjud hududgina qabul qilinadi
        if p.region_id <= 0:
            fields["region_id"] = None
        elif db.get_region_by_id(p.region_id):
            fields["region_id"] = p.region_id
        else:
            raise HTTPException(status_code=400, detail="bad_region")
    # Sotuvchi kontakt username (bot edit_telegram pariteti). "@" va bo'sh joylar
    # tozalanadi; bo'sh qiymat → tozalash. Telegram username 5-32 belgi (a-z0-9_).
    if p.telegram_username is not None:
        un = p.telegram_username.strip().lstrip("@").strip()
        if un:
            if not _re.fullmatch(r"[A-Za-z0-9_]{5,32}", un):
                raise HTTPException(status_code=400, detail="bad_username")
            fields["telegram_username"] = un
        else:
            fields["telegram_username"] = None
    # Yetkazish uchun minimal buyurtma summasi (arzon buyurtmaga yetkazish zarar
    # qilmasligi uchun). 0 yoki manfiy → cheklov olib tashlanadi (NULL).
    if p.delivery_min_total is not None:
        try:
            mv = float(p.delivery_min_total)
        except (ValueError, TypeError):
            raise HTTPException(status_code=400, detail="bad_delivery_min")
        if mv > 1e12:
            raise HTTPException(status_code=400, detail="bad_delivery_min")
        fields["delivery_min_total"] = mv if mv > 0 else None
    # "Aloqa bloki" — telefon + ijtimoiy tarmoqlar (reklamaga avtomatik chiqadi).
    # Bo'sh qiymat → o'sha maydonni tozalaydi (NULL). Havolalar bot tomonida
    # to'g'ri https/tg formatga keltiriladi — bu yerda faqat uzunlik cheklovi.
    if p.shop_phone is not None:
        ph = p.shop_phone.strip()
        if ph:
            digits = ph.lstrip("+").replace(" ", "").replace("-", "")
            if not digits.isdigit() or not (7 <= len(digits) <= 15):
                raise HTTPException(status_code=400, detail="bad_phone")
        fields["shop_phone"] = ph or None
    for attr in ("shop_instagram", "shop_telegram", "shop_website"):
        v = getattr(p, attr)
        if v is not None:
            v = v.strip()
            if len(v) > 200:
                raise HTTPException(status_code=400, detail="bad_link")
            fields[attr] = v or None
    if fields:
        db.update_user(user["id"], **fields)
    # payment_mode shops jadvalida (faqat ega)
    if p.payment_mode is not None and p.payment_mode in ("shop", "staff"):
        shop = db.get_shop_by_owner(user["id"])
        if shop:
            db.update_shop(dict(shop)["id"], payment_mode=p.payment_mode)
    return {"ok": True}


# ---- XODIM O'Z TO'LOV KARTASI (faqat payment_mode='staff' bo'lsa) ----
class MyCardIn(BaseModel):
    card_number: Optional[str] = None
    card_owner: Optional[str] = None
    card_type: Optional[str] = None


def _staff_card_ctx(user):
    """Xodim + do'kon to'lov rejimini qaytaradi (ega yoki yakka sotuvchi → None)."""
    staff = db.get_staff_by_user(user["id"])
    if not staff or dict(staff).get("staff_role") == "owner":
        return None
    st = dict(staff)
    shop = db.get_shop_for_user(user["id"])
    mode = (dict(shop).get("payment_mode") if shop else None) or "shop"
    return {"staff": st, "mode": mode}


@app.get("/api/seller/my-card")
def api_my_card(authorization: str = Header(None)):
    """Xodim o'z to'lov kartasi. applicable=True faqat do'kon 'staff' rejimida bo'lsa."""
    user = dict(_buyer_from_auth(authorization))
    ctx = _staff_card_ctx(user)
    if not ctx:
        return {"applicable": False}   # ega/yakka — kartani do'kon sozlamalarida boshqaradi
    st = ctx["staff"]
    return {"applicable": ctx["mode"] == "staff", "payment_mode": ctx["mode"],
            "card_number": st.get("card_number"), "card_owner": st.get("card_owner"),
            "card_type": st.get("card_type")}


@app.patch("/api/seller/my-card")
def api_my_card_set(p: MyCardIn, authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    ctx = _staff_card_ctx(user)
    if not ctx:
        raise HTTPException(status_code=403, detail="not_staff")
    if ctx["mode"] != "staff":
        raise HTTPException(status_code=409, detail="card_not_applicable")  # do'kon kartasi rejimi
    fields = {}
    if p.card_number is not None:
        cn = p.card_number.replace(" ", "").strip()
        if cn:
            if not cn.isdigit() or not (12 <= len(cn) <= 19):
                raise HTTPException(status_code=400, detail="bad_card")
            fields["card_number"] = cn
        else:
            fields["card_number"] = None
    if p.card_owner is not None:
        fields["card_owner"] = p.card_owner.strip() or None
    if p.card_type is not None:
        if p.card_type and p.card_type not in _VALID_CARD:
            raise HTTPException(status_code=400, detail="bad_card_type")
        fields["card_type"] = p.card_type or None
    if fields:
        db.update_staff(ctx["staff"]["id"], **fields)
    return {"ok": True}


class OrderAction(BaseModel):
    action: str  # 'confirm' | 'reject'
    reason: Optional[str] = None  # 'reject' uchun bekor sababi (code:/text:)


@app.post("/api/seller/order/{order_id}/action")
async def api_seller_order_action(order_id: int, body: OrderAction,
                                  authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    if body.action not in ("confirm", "reject"):
        raise HTTPException(status_code=400, detail="bad_action")
    order = db.get_order_by_id(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="not_found")
    # Egalik: shu DO'KON (ega yoki xodim) buyurtmasi yoki admin — bot/kanal buyurtmasi ham
    if not (order.get("seller_id") == _owner_id(user) or user.get("role") == "admin"):
        raise HTTPException(status_code=403, detail="not_your_order")
    # Xodim "buyurtma tasdiqlash" ruxsatiga ega bo'lishi shart (ega — har doim)
    if not _staff_perm(user, "perm_confirm_orders"):
        raise HTTPException(status_code=403, detail="no_perm_confirm")
    new_status = "confirmed" if body.action == "confirm" else "cancelled"
    reason = _norm_cancel_reason(body.reason) if new_status == "cancelled" else None
    gid = order.get("order_group_id")
    # HIMOYA (ATOMIK): faqat 'pending' holatdan o'tkazamiz. Bot va Mini App AYNAN bir
    # buyurtmani bir vaqtda tasdiqlasa/bekor qilsa, faqat BITTA chaqiruv yutadi
    # (rowcount=1) → zahira ikki marta kamaymaydi, takroriy xabar ketmaydi. Yutmagan
    # chaqiruv 409 oladi (ilgari read-then-check atomik emas edi).
    # Variant/savat guruh buyurtmasi — bitta amal butun guruhga (1 shartnoma).
    if gid:
        won_ids = db.transition_group_status(
            gid, new_status, "pending",
            cancel_by="seller" if new_status == "cancelled" else None,
            cancel_reason=reason)
        if not won_ids:
            raise HTTPException(status_code=409, detail="already_processed")
    else:
        if not db.transition_order_status(
                order_id, new_status, "pending",
                cancel_by="seller" if new_status == "cancelled" else None,
                cancel_reason=reason):
            raise HTTPException(status_code=409, detail="already_processed")
        won_ids = [order_id]
    if new_status == "confirmed":
        for wid in won_ids:
            try:
                wo = order if wid == order_id else db.get_order_by_id(wid)
                db.decrement_stock_on_confirm(wo["product_id"], wo["quantity"])
            except Exception as e:
                logging.error(f"stock kamaytirish xato (order {wid}): {e}")

    # Xaridorga bildirishnoma (xaridor tilida) — bot order_confirm bilan bir xil matn
    try:
        buyer = db.get_user_by_id(order["buyer_id"])
        blang = get_user_lang(buyer) if buyer else DEFAULT_LANG
        is_pickup = order.get("delivery_type") == "pickup"
        oid = fmt_order_id(order_id)
        pname = html.escape(order.get("product_name") or "")
        if new_status == "confirmed":
            key = "order_confirmed_pickup" if is_pickup else "order_confirmed_delivery"
        else:
            key = "order_cancelled_notify"
        txt = t(blang, key, oid=oid, pname=pname)
        if order.get("buyer_tg") and txt:
            await _tg_call("sendMessage", {
                "chat_id": order["buyer_tg"], "text": txt, "parse_mode": "HTML"})
        # App-banner: xaridorga buyurtma holati
        pn = order.get("product_name") or ""
        if new_status == "confirmed":
            _notify_db(order["buyer_id"], "order", ("✅ Buyurtma tasdiqlandi", "✅ Заказ подтверждён"),
                       (pn, pn), ref_id=order_id)
        else:
            _notify_db(order["buyer_id"], "order", ("❌ Buyurtma bekor qilindi", "❌ Заказ отменён"),
                       (pn, pn), ref_id=order_id)
            # ADMIN — sotuvchi buyurtmani rad etganidan xabardor bo'lsin
            await _notify_admins(
                f"❌ Sotuvchi buyurtmani rad etdi {oid}",
                f"{pn}", kind="order", ref_id=order_id)
    except Exception as e:
        logging.error(f"xaridorga xabar xato (order {order_id}): {e}")

    # Sotuvchining bildirishnoma xabaridagi tugmalarni olib tashlaymiz (eski tugma bosilmasin)
    try:
        chat_id = order.get("notify_chat_id")
        msg_id = order.get("notify_message_id")
        if chat_id and msg_id:
            final = (order.get("notify_caption") or "")
            final += "\n\n" + ("✅ Tasdiqlandi" if new_status == "confirmed" else "❌ Bekor qilindi")
            await _tg_call("editMessageText", {
                "chat_id": chat_id, "message_id": msg_id, "text": final,
                "parse_mode": "HTML", "reply_markup": {"inline_keyboard": []}})
    except Exception as e:
        logging.warning(f"sotuvchi xabarini tahrirlash xato (order {order_id}): {e}")

    return {"ok": True, "status": new_status}


class CancelReqIn2(BaseModel):
    reason: str = ""


@app.post("/api/seller/order/{order_id}/request-cancel")
async def api_seller_request_cancel(order_id: int, body: CancelReqIn2, authorization: str = Header(None)):
    """Sotuvchi TASDIQLANGAN buyurtmani bekor qilishni so'raydi (nizo oqimi).
    Xaridorga rozilik so'rovi (rozi/rad tugmalari) yuboriladi — bot ularni ham ushlaydi."""
    user = dict(_buyer_from_auth(authorization))
    _rate_limit("cancel_req", user["id"], 10, 60)
    reason = (body.reason or "").strip() or ("—")
    if len(reason) > 500:
        raise HTTPException(status_code=400, detail="too_long")
    order = db.get_order_by_id(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="not_found")
    if not (order.get("seller_id") == _owner_id(user) or user.get("role") == "admin"):
        raise HTTPException(status_code=403, detail="not_your_order")
    if order.get("status") != "confirmed" or (order.get("cancel_state") or ""):
        raise HTTPException(status_code=409, detail="cancel_not_available")
    gid = order.get("order_group_id")
    if gid:
        if not db.request_group_cancel(gid, "seller", reason):
            raise HTTPException(status_code=409, detail="cancel_not_available")
    elif not db.request_order_cancel(order_id, "seller", reason):
        raise HTTPException(status_code=409, detail="cancel_not_available")
    try:
        buyer = db.get_user_by_id(order["buyer_id"]) if order.get("buyer_id") else None
        blang = get_user_lang(buyer) if buyer else DEFAULT_LANG
        if order.get("buyer_tg"):
            await _tg_call("sendMessage", {
                "chat_id": order["buyer_tg"],
                "text": t(blang, "cancel_request_notify", oid=fmt_order_id(order_id),
                          pname=html.escape(order.get("product_name") or ""),
                          reason=html.escape(_cancel_reason_display(reason, blang))),
                "parse_mode": "HTML",
                "reply_markup": {"inline_keyboard": [
                    [{"text": t(blang, "btn_cancel_agree"), "callback_data": f"cclagree_{order_id}"}],
                    [{"text": t(blang, "btn_cancel_deny"), "callback_data": f"ccldeny_{order_id}"}]]}})
    except Exception as e:
        logging.warning(f"seller request-cancel notify xato (order {order_id}): {e}")
    return {"ok": True}


@app.post("/api/seller/order/{order_id}/cancel-respond")
async def api_seller_cancel_respond(order_id: int, body: CancelRespondIn, authorization: str = Header(None)):
    """Sotuvchi XARIDOR boshlagan bekor so'roviga javob beradi: rozi (bekor) yoki rad (nizo)."""
    user = dict(_buyer_from_auth(authorization))
    order = db.get_order_by_id(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="not_found")
    if not (order.get("seller_id") == _owner_id(user) or user.get("role") == "admin"):
        raise HTTPException(status_code=403, detail="not_your_order")
    if order.get("cancel_state") != "requested":
        raise HTTPException(status_code=409, detail="cancel_already_handled")
    if order.get("cancel_by") == "seller":
        raise HTTPException(status_code=409, detail="cancel_wait_other")  # o'zi so'ragan
    oid = fmt_order_id(order_id)
    pname = html.escape(order.get("product_name") or "")
    buyer = db.get_user_by_id(order["buyer_id"]) if order.get("buyer_id") else None
    blang = get_user_lang(buyer) if buyer else DEFAULT_LANG
    gid = order.get("order_group_id")
    if body.agree:
        if gid:
            for o in db.agree_group_cancel(gid):
                try:
                    od = db.get_order_by_id(o)
                    db.restock_on_cancel(od["product_id"], od.get("quantity") or 1)
                except Exception as e:
                    logging.warning(f"seller cancel-respond restock xato (order {o}): {e}")
        elif db.agree_order_cancel(order_id):
            try:
                db.restock_on_cancel(order["product_id"], order.get("quantity") or 1)
            except Exception as e:
                logging.warning(f"seller cancel-respond restock xato (order {order_id}): {e}")
        if order.get("buyer_tg"):
            await _tg_call("sendMessage", {"chat_id": order["buyer_tg"],
                           "text": t(blang, "cancel_agreed_notify", oid=oid, pname=pname),
                           "parse_mode": "HTML"})
        return {"ok": True, "cancelled": True}
    # rad — admin hakamligiga
    if gid:
        db.dispute_group_cancel(gid)
    else:
        db.dispute_order_cancel(order_id)
    _pn = order.get("product_name") or ""
    _notify_admins_db("dispute", ("⚖️ Yangi nizo — hal kutyapti", "⚖️ Новый спор — ждёт решения"),
                      (_pn, _pn), ref_id=order_id)
    if order.get("buyer_tg"):
        await _tg_call("sendMessage", {"chat_id": order["buyer_tg"],
                       "text": t(blang, "cancel_denied_notify", oid=oid, pname=pname),
                       "parse_mode": "HTML"})
    try:
        if ADMIN_ID:
            await _tg_call("sendMessage", {"chat_id": ADMIN_ID,
                           "text": t(DEFAULT_LANG, "admin_dispute_notify", oid=oid, pname=pname,
                                     by=html.escape(user.get("name") or "sotuvchi"),
                                     reason=html.escape(order.get("cancel_reason") or "—")),
                           "parse_mode": "HTML"})
    except Exception as e:
        logging.warning(f"seller cancel-respond admin notify xato (order {order_id}): {e}")
    return {"ok": True, "disputed": True}


class DeliverIn(BaseModel):
    settlement_type: str  # 'paid' | 'debt' | 'installment'
    paid: float = 0


@app.post("/api/seller/order/{order_id}/deliver")
async def api_deliver(order_id: int, body: DeliverIn, authorization: str = Header(None)):
    """Berish + to'lov holati (to'liq/qarz/bo'lib). Tasdiqlangan buyurtmani 'delivered'
    qiladi, settlement saqlaydi, xaridorga xabar (qarz bo'lsa qarz ham) yuboradi.
    Guruh (savat) buyurtmasi bo'lsa — butun guruhga."""
    user = dict(_buyer_from_auth(authorization))
    order = db.get_order_by_id(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="not_found")
    if not (order.get("seller_id") == _owner_id(user) or user.get("role") == "admin"):
        raise HTTPException(status_code=403, detail="not_your_order")
    # Yetkazib berishni yakunlash: tasdiqlash ruxsatli xodim YOKI KURYER
    if not (_staff_perm(user, "perm_confirm_orders") or _is_courier(user)):
        raise HTTPException(status_code=403, detail="no_perm_confirm")
    if order.get("status") != "confirmed":
        raise HTTPException(status_code=409, detail="not_confirmed")
    st = body.settlement_type
    if st not in ("paid", "debt", "installment"):
        raise HTTPException(status_code=400, detail="bad_settlement")

    gid = order.get("order_group_id")
    if gid:
        group_orders = db.get_orders_in_group(gid)
        total = sum(float(o.get("total_price") or 0) for o in group_orders)
    else:
        group_orders = None
        total = float(order.get("total_price") or 0)
    paid = max(0.0, min(float(body.paid or 0), total))
    due = round(total - paid, 2)
    eff = "paid" if due <= 0 else st

    if gid:
        # ATOMIK: faqat 'confirmed' sub-buyurtmalarni 'delivered'ga o'tkazamiz. Bot
        # settlement oqimi bilan bir vaqtda berilsa, faqat YUTGAN tomon settlement/xabar
        # bajaradi (qarama-qarshi to'lov holati + ikki "berildi" xabarining oldini olamiz).
        won = [o for o in group_orders if db.transition_order_status(o["id"], "delivered", "confirmed")]
        if not won:
            raise HTTPException(status_code=409, detail="already_delivered")
        db.set_group_settlement(gid, eff, paid, due)
        disp = fmt_order_id(int(gid))
    else:
        if not db.transition_order_status(order_id, "delivered", "confirmed"):
            raise HTTPException(status_code=409, detail="already_delivered")
        db.set_order_settlement(order_id, eff, paid, due)
        disp = fmt_order_id(order_id)

    # #18 Komissiya — sotuv yakunlanganda platforma ulushini yozamiz (yoqilgan bo'lsa).
    try:
        mc = monetization_config()
        pct = mc["mon_commission_percent"]
        if mc["mon_enabled"] and mc["mon_commission_enabled"] and pct > 0:
            if gid:
                for o in won:
                    db.set_order_commission(o["id"], float(o.get("total_price") or 0) * pct / 100.0)
            else:
                db.set_order_commission(order_id, total * pct / 100.0)
    except Exception as e:
        logging.warning(f"komissiya yozish xato (order {order_id}): {e}")

    try:
        buyer = db.get_user_by_id(order["buyer_id"]) if order.get("buyer_id") else None
        blang = get_user_lang(buyer) if buyer else DEFAULT_LANG
        seller = db.get_user_by_id(order["seller_id"]) if order.get("seller_id") else None
        is_pickup = order.get("delivery_type") == "pickup"
        if gid:
            txt = t(blang, "grp_delivered_pickup" if is_pickup else "grp_delivered_delivery",
                    oid=disp, n=len(group_orders))
        else:
            txt = t(blang, "order_delivered_pickup" if is_pickup else "order_delivered_delivery",
                    oid=disp, pname=html.escape(order.get("product_name") or ""))
        if due > 0:
            shop = html.escape((seller.get("shop_name") or seller.get("name") or "") if seller else "")
            txt += t(blang, "buyer_debt_notify", shop=shop, due=fmt_price(due))
        if order.get("buyer_tg"):
            await _tg_call("sendMessage", {
                "chat_id": order["buyer_tg"], "text": txt, "parse_mode": "HTML",
                "reply_markup": {"inline_keyboard": [[
                    {"text": t(blang, "btn_leave_rating"), "callback_data": f"order_rate_{order_id}"}]]}})
        # App-banner: xaridorga "yetkazildi" (+ qarz bo'lsa eslatma)
        pn = order.get("product_name") or ""
        dbody_uz = pn + (f" · Qarz: {fmt_price(due)}" if due > 0 else "")
        dbody_ru = pn + (f" · Долг: {fmt_price(due)}" if due > 0 else "")
        _notify_db(order["buyer_id"], "order", ("🚚 Buyurtma yetkazildi", "🚚 Заказ доставлен"),
                   (dbody_uz, dbody_ru), ref_id=order_id)
    except Exception as e:
        logging.warning(f"deliver buyer notify xato (order {order_id}): {e}")
    return {"ok": True, "total": total, "paid": paid, "due": due}


# ---- Mahsulot boshqaruvi (D2) ----
MAX_PHOTO_BYTES = 6 * 1024 * 1024


def _own_product_or_403(user, product_id):
    prod = db.get_product_by_id(product_id)
    if not prod:
        raise HTTPException(status_code=404, detail="not_found")
    # Multivendor: mahsulot DO'KONga tegishli bo'lsa — ega ham, xodim ham (shu do'kondan)
    # boshqara oladi. seller_id = do'kon egasi; foydalanuvchining egasi shu bo'lsa — ruxsat.
    if not (prod.get("seller_id") == _owner_id(user) or user.get("role") == "admin"):
        raise HTTPException(status_code=403, detail="not_your_product")
    return prod


@app.post("/api/seller/product/photo")
async def api_product_photo(file: UploadFile = File(...), authorization: str = Header(None)):
    """Rasmni Telegram'ga yuborib file_id oladi (so'ng xabarni o'chiradi — file_id amal qiladi).
    Shunday qilib web upload (baytlar) bot ishlatadigan Telegram file_id'ga aylanadi."""
    user = dict(_buyer_from_auth(authorization))
    seller_tg = user.get("telegram_id")
    if not seller_tg:
        raise HTTPException(status_code=400, detail="no_chat")
    if not BOT_TOKEN:
        raise HTTPException(status_code=503, detail="no_token")
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="empty")
    if len(data) > MAX_PHOTO_BYTES:
        raise HTTPException(status_code=413, detail="too_large")
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            r = await client.post(
                f"https://api.telegram.org/bot{BOT_TOKEN}/sendPhoto",
                data={"chat_id": str(seller_tg)},
                files={"photo": (file.filename or "photo.jpg", data,
                                 file.content_type or "image/jpeg")},
            )
            res = r.json()
    except Exception as e:
        logging.error(f"sendPhoto xato: {e}")
        raise HTTPException(status_code=502, detail="upload_failed")
    if not res.get("ok") or not (res.get("result") or {}).get("photo"):
        raise HTTPException(status_code=502, detail="telegram_rejected")
    msg = res["result"]
    file_id = msg["photo"][-1]["file_id"]
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            await client.post(f"https://api.telegram.org/bot{BOT_TOKEN}/deleteMessage",
                              json={"chat_id": seller_tg, "message_id": msg["message_id"]})
    except Exception:
        pass
    return {"file_id": file_id}


# Video MVP: 1 ta ixtiyoriy mahsulot videosi. 20MB — Bot API getFile yuklab olish
# chegarasi (proksi ishlashi uchun undan oshmasin); 60s — mahsulot klipi normasi.
MAX_VIDEO_BYTES = 20 * 1024 * 1024
MAX_VIDEO_SECONDS = 60


@app.post("/api/seller/product/video")
async def api_product_video(file: UploadFile = File(...), authorization: str = Header(None)):
    """Videoni Telegram'ga yuborib file_id oladi (photo endpoint bilan bir uslub).
    Davomiylik Telegram javobidan tekshiriladi (>60s — rad); xabar darhol o'chiriladi."""
    user = dict(_buyer_from_auth(authorization))
    seller_tg = user.get("telegram_id")
    if not seller_tg:
        raise HTTPException(status_code=400, detail="no_chat")
    if not BOT_TOKEN:
        raise HTTPException(status_code=503, detail="no_token")
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="empty")
    if len(data) > MAX_VIDEO_BYTES:
        raise HTTPException(status_code=413, detail="too_large")
    try:
        async with httpx.AsyncClient(timeout=90) as client:
            r = await client.post(
                f"https://api.telegram.org/bot{BOT_TOKEN}/sendVideo",
                data={"chat_id": str(seller_tg), "supports_streaming": "true"},
                files={"video": (file.filename or "video.mp4", data,
                                 file.content_type or "video/mp4")},
            )
            res = r.json()
    except Exception as e:
        logging.error(f"sendVideo xato: {e}")
        raise HTTPException(status_code=502, detail="upload_failed")
    msg = (res.get("result") or {}) if res.get("ok") else {}
    video = msg.get("video")
    if not video:
        raise HTTPException(status_code=502, detail="telegram_rejected")

    async def _cleanup():
        try:
            async with httpx.AsyncClient(timeout=10) as client:
                await client.post(f"https://api.telegram.org/bot{BOT_TOKEN}/deleteMessage",
                                  json={"chat_id": seller_tg, "message_id": msg["message_id"]})
        except Exception:
            pass

    duration = int(video.get("duration") or 0)
    if duration > MAX_VIDEO_SECONDS + 2:   # +2s — kodeklardagi yaxlitlash farqiga imtiyoz
        await _cleanup()
        raise HTTPException(status_code=400, detail="video_too_long")
    await _cleanup()
    return {"file_id": video["file_id"], "duration": duration}


class AttrItem(BaseModel):
    key: str
    value: Optional[str] = None
    label: Optional[str] = None


class WholesaleTier(BaseModel):
    min_qty: int
    price: float


class ProductIn(BaseModel):
    name: str
    price: float
    category_id: Optional[int] = None
    description: Optional[str] = None
    stock_count: Optional[int] = None
    old_price: Optional[float] = None  # chegirma: eski (chizilgan) narx
    unit: Optional[str] = None  # #20 — o'lchov birligi (dona/kg/litr/metr...)
    min_price: Optional[float] = None  # #8 — MAXFIY oxirgi narx (savdolashish floor'i)
    min_order_qty: Optional[int] = None  # variant-buyurtma JAMI minimal son (optom)
    wholesale_price: Optional[float] = None  # eski yagona zina (moslik)
    wholesale_min_qty: Optional[int] = None  # eski yagona zina (moslik)
    wholesale_tiers: Optional[List[WholesaleTier]] = None  # optom: pachka-zina (bir nechta); dona: ishlatilmaydi
    image_url: Optional[str] = None  # eski: bitta file_id (moslik uchun)
    images: Optional[List[str]] = None  # galereya: file_id ro'yxati (1-chi = asosiy)
    image_labels: Optional[List[str]] = None  # har rasm uchun nom (optom: rang nomi); images bilan bir tartibda
    attributes: Optional[List[AttrItem]] = None  # mahsulot atributlari (klassik/AI)
    sale_mode: Optional[str] = None  # 'dona' (donalab) | 'optom' (pachka). None/noma'lum = 'dona'
    pack_size: Optional[int] = None  # optom: 1 pachkadagi dona soni
    size_note: Optional[str] = None  # optom: razmer matni (butun mahsulotga)
    delivery_available: Optional[int] = None  # 1=yetkaziladi (default), 0=faqat olib ketish
    video_file_id: Optional[str] = None  # ixtiyoriy qisqa video (Telegram file_id)


def _save_attrs(product_id, attributes):
    """AttrItem ro'yxatini DB'ga saqlaydi (key->value, key->label)."""
    if not attributes:
        return
    attrs, labels = {}, {}
    for a in attributes:
        key = (a.key or "").strip()
        val = (a.value or "").strip() if a.value is not None else ""
        if not key or not val:
            continue
        attrs[key] = val
        if a.label:
            labels[key] = a.label.strip()
    if attrs:
        try:
            db.save_product_attributes(product_id, attrs, labels=labels)
        except Exception as e:
            logging.warning(f"save_product_attributes xato (pid {product_id}): {e}")


def _images_list(p):
    """ProductIn/ProductEdit'dan rasm ro'yxatini chiqaradi (images ustun, bo'lmasa image_url)."""
    if p.images is not None:
        return [f for f in p.images if f][:db.MAX_PRODUCT_IMAGES]
    if p.image_url:
        return [p.image_url]
    return None


def _image_labels_for(p, imgs):
    """ProductIn/Edit'dan rasm yorliqlari (optom: rang nomlari) ro'yxatini imgs tartibida
    qaytaradi. Yorliqlar berilmagan bo'lsa None (set_product_images barchasini NULL qiladi).
    Faqat NULL bo'lmagan, file_id bor rasmlarga mos yorliq olinadi (images filtridan keyin)."""
    if p.images is None or not p.image_labels:
        return None
    raw = p.image_labels
    out, j = [], 0
    for f in p.images:           # _images_list bilan bir xil filtr (bo'sh file_id tashlanadi)
        if not f:
            j += 1
            continue
        lbl = raw[j] if j < len(raw) else None
        out.append((str(lbl).strip()[:40] or None) if lbl is not None else None)
        j += 1
    return out[:db.MAX_PRODUCT_IMAGES]


def _norm_sale_mode(val):
    """Savdo turini normallashtiradi: 'optom' bo'lsa shu, aks holda 'dona' (default)."""
    return "optom" if (str(val or "").strip().lower() == "optom") else "dona"


_MAX_WHOLESALE_TIERS = 6


def _wholesale_fields(price, tiers, wp, wq):
    """Optom narx fieldlarini validatsiya qilib qaytaradi:
    {'wholesale_tiers' (JSON yoki None), 'wholesale_price', 'wholesale_min_qty' (eski yagona — moslik)}.

    tiers (List[WholesaleTier]) berilsa — KO'P ZINA rejimi: har zina min>=2, 0<price<dona narx;
    min QAT'IY o'sib, narx QAT'IY kamayib borishi shart; ko'pi bilan 6 zina. Bo'sh ro'yxat = o'chiq.
    tiers None bo'lsa — eski yagona zina (wp/wq) bilan ishlaydi (orqaga moslik).
    Noto'g'ri -> 400 bad_wholesale / wholesale_not_cheaper."""
    pr_listed = float(price) if price not in (None, "") else None
    if tiers is not None:
        items = []
        try:
            for t in tiers:
                items.append((int(t.min_qty), float(t.price)))
        except (ValueError, TypeError, AttributeError):
            raise HTTPException(status_code=400, detail="bad_wholesale")
        if not items:
            return {"wholesale_tiers": None, "wholesale_price": None, "wholesale_min_qty": None}
        if len(items) > _MAX_WHOLESALE_TIERS:
            raise HTTPException(status_code=400, detail="bad_wholesale")
        items.sort(key=lambda x: x[0])
        prev_min, prev_price = 0, None
        for m, p in items:
            if m < 2 or p <= 0:
                raise HTTPException(status_code=400, detail="bad_wholesale")
            if pr_listed is not None and p >= pr_listed:
                raise HTTPException(status_code=400, detail="wholesale_not_cheaper")
            if m <= prev_min:
                raise HTTPException(status_code=400, detail="bad_wholesale")   # min qat'iy o'sishi kerak
            if prev_price is not None and p >= prev_price:
                raise HTTPException(status_code=400, detail="bad_wholesale")   # narx qat'iy kamayishi kerak
            prev_min, prev_price = m, p
        js = json.dumps([{"min": m, "price": p} for m, p in items])
        entry = items[0]   # eski yagona ustunlarni eng past zina bilan sinxronlaymiz (moslik)
        return {"wholesale_tiers": js, "wholesale_price": float(entry[1]), "wholesale_min_qty": int(entry[0])}
    # --- Orqaga moslik: yagona zina (wp/wq) ---
    try:
        has_wp = wp is not None and float(wp) > 0
        has_wq = wq is not None and int(wq) > 0
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="bad_wholesale")
    if not has_wp and not has_wq:
        return {"wholesale_tiers": None, "wholesale_price": None, "wholesale_min_qty": None}   # o'chiq
    if not (has_wp and has_wq):
        raise HTTPException(status_code=400, detail="bad_wholesale")   # ikkalasi ham kerak
    wp, wq = float(wp), int(wq)
    if wq < 2:
        raise HTTPException(status_code=400, detail="bad_wholesale")
    if pr_listed is not None and wp >= pr_listed:
        raise HTTPException(status_code=400, detail="wholesale_not_cheaper")
    return {"wholesale_tiers": None, "wholesale_price": wp, "wholesale_min_qty": wq}


async def _notify_admins_moderation(pid, name, seller, mod):
    """#5 — bloklangan mahsulot haqida adminlarga xabar (ADMIN_ID + role='admin')."""
    recipients = set()
    if ADMIN_ID:
        recipients.add(ADMIN_ID)
    try:
        for a in db.get_all_users(role="admin"):
            tg = dict(a).get("telegram_id")
            if tg:
                recipients.add(tg)
    except Exception:
        pass
    shop = html.escape((seller.get("shop_name") or seller.get("name") or "") if seller else "")
    txt = (f"🛡 <b>Avto-moderatsiya bloklash</b>\n\n"
           f"🛍️ {html.escape(name or '')}\n🏪 {shop}\n"
           f"⚠️ {html.escape(mod.get('category') or '')}: {html.escape(mod.get('reason') or '')}\n\n"
           f"App → Admin → Moderatsiya orqali tasdiqlang yoki rad eting.")
    for chat_id in recipients:
        try:
            await _tg_call("sendMessage", {"chat_id": chat_id, "text": txt, "parse_mode": "HTML"})
        except Exception as e:
            logging.warning(f"moderation admin notify xato: {e}")


@app.post("/api/seller/product")
async def api_create_product(p: ProductIn, background: BackgroundTasks,
                             authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    name = (p.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="name_required")
    if p.price is None or p.price <= 0:
        raise HTTPException(status_code=400, detail="bad_price")
    if p.stock_count is not None and p.stock_count < 0:
        raise HTTPException(status_code=400, detail="bad_stock")
    # MULTIVENDOR: mahsulot do'kon EGASI nomidan chiqadi (seller_id=ega), created_by=xodim.
    # Xodim faol bo'lishi va "mahsulot qo'shish" ruxsatiga ega bo'lishi shart (bot pariteti).
    owner_id = _owner_id(user)
    staff = db.get_staff_by_user(user["id"])
    is_staff = bool(staff and dict(staff).get("staff_role") != "owner")
    if is_staff:
        st = dict(staff)
        if not st.get("is_active", 1):
            raise HTTPException(status_code=403, detail="staff_inactive")
        if not st.get("perm_add_product", 1):
            raise HTTPException(status_code=403, detail="no_perm_add")
    # #18 Pro limit — obuna yoqilgan va bepul limit belgilangan bo'lsa, Pro bo'lmagan
    # do'kon limitdan ortiq faol mahsulot qo'sha olmaydi.
    pub = monetization_public()
    flim = int(pub.get("free_product_limit") or 0)
    if flim > 0 and not _is_pro(owner_id) and db.count_active_products(owner_id) >= flim:
        raise HTTPException(status_code=403, detail="free_limit_reached")
    shop = db.get_shop_for_user(user["id"])
    # perm_publish_ad — "reklamani tasdiqsiz o'zi joylash" ruxsatli xodim ega tasdig'ini
    # CHETLAB O'TADI: mahsuloti darhol faol bo'ladi va reklama AVTOMATIK kanal/guruhlarga
    # chiqadi (qo'lda "Reklama joylash" tugmasini bosish shart emas).
    can_auto_ad = bool(is_staff and dict(staff).get("perm_publish_ad"))
    needs_owner_approval = bool(is_staff and shop
                               and dict(shop).get("moderation") == "owner_approve"
                               and not can_auto_ad)
    # #5 AVTO-MODERATSIYA — taqiqlangan tovar/kontent bo'lsa jonli efirga CHIQARMAYMIZ.
    # AI o'chiq yoki xato bo'lsa mod=None → oddiy o'tadi (hammasi bloklanmaydi).
    lang = get_user_lang(user) or DEFAULT_LANG
    try:
        mod = await ai_assistant.moderate_product(
            name=name, description=(p.description or ""), lang=lang)
    except Exception as e:
        logging.warning(f"moderate_product xato: {e}")
        mod = None
    blocked = bool(mod and mod.get("flagged"))
    imgs = _images_list(p)
    sale_mode = _norm_sale_mode(p.sale_mode)
    # Status ustuvorligi: bloklangan (xavfsizlik) > ega tasdig'i kutilmoqda > faol.
    # DIQQAT: status YARATISH paytidayoq beriladi — aks holda quyidagi update_product_fields
    # yiqilsa AI-bloklangan mahsulot 'active' bo'lib jonli efirga chiqib ketardi (moderatsiya
    # chetlab o'tilishi). Endi bloklangan mahsulot boshidanoq yashirin yaratiladi.
    if blocked:
        status, in_stock = "mod_blocked", 0
    elif needs_owner_approval:
        status, in_stock = "pending_owner", 0
    else:
        status, in_stock = "active", 1
    pid = db.create_product(
        seller_id=owner_id, name=name, price=float(p.price),
        category_id=p.category_id, description=(p.description or "").strip() or None,
        image_url=(imgs[0] if imgs else None), stock_count=p.stock_count, created_by=user["id"],
        status=status,
    )
    if imgs:
        try:
            db.set_product_images(pid, imgs, labels=_image_labels_for(p, imgs))
        except Exception as e:
            logging.warning(f"set_product_images xato (pid {pid}): {e}")
    # status/in_stock yuqorida (create_product'dan oldin) hisoblangan — bu yerda faqat
    # qolgan maydonlar bilan birga yoziladi (in_stock ni ham izchil saqlaymiz).
    fields = {"in_stock": in_stock, "status": status}
    if blocked:
        fields["mod_reason"] = (mod.get("reason") or "")[:300]
    owner_rec = db.get_user_by_id(owner_id)   # mahsulot hududi — DO'KON (ega) hududi
    if owner_rec and dict(owner_rec).get("region_id"):
        fields["region_id"] = dict(owner_rec)["region_id"]
    if p.old_price and p.old_price > 0:
        fields["old_price"] = float(p.old_price)
    if p.unit is not None and p.unit.strip():
        fields["unit"] = p.unit.strip()[:20]
    # Savdo turi: optom (pachka) yoki dona. Optomda savdolashish (min_price) YO'Q.
    fields["sale_mode"] = sale_mode
    if sale_mode == "optom":
        fields["min_price"] = None   # optomda savdolashish yo'q
        fields["pack_size"] = int(p.pack_size) if p.pack_size and p.pack_size > 0 else None
        fields["size_note"] = (p.size_note or "").strip()[:200] or None
    else:
        fields["pack_size"] = None
        fields["size_note"] = None
        if p.min_price is not None:
            fields["min_price"] = float(p.min_price) if p.min_price and p.min_price > 0 else None
    if p.min_order_qty is not None:
        fields["min_order_qty"] = int(p.min_order_qty) if p.min_order_qty and p.min_order_qty > 1 else None
    if p.delivery_available is not None:
        fields["delivery_available"] = 1 if p.delivery_available else 0
    if p.video_file_id is not None:
        fields["video_file_id"] = p.video_file_id.strip() or None
    fields.update(_wholesale_fields(p.price, p.wholesale_tiers, p.wholesale_price, p.wholesale_min_qty))
    try:
        db.update_product_fields(pid, **fields)
    except Exception as e:
        logging.warning(f"product post-fields xato (pid {pid}): {e}")
    _save_attrs(pid, p.attributes)
    if blocked:
        background.add_task(_notify_admins_moderation, pid, name, user, mod)
    # perm_publish_ad'li xodim: reklama AVTOMATIK chiqadi (ega tasdig'isiz). api_ad_publish
    # bilan bir xil mexanizm — hozirgi vaqtga pending post, bot scan-job (~30s) post_product_to_channel'ni
    # ishga tushiradi (markaziy kanal + do'konning barcha ulangan kanal/guruhlari). caption/image=None
    # -> bot AI reklama dizayni + matnini o'zi quradi.
    ad_published = False
    if can_auto_ad and not blocked:
        try:
            sa = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
            db.create_scheduled_post(pid, owner_id, sa, created_by=user["id"],
                                     caption=None, parse_mode=None, image_id=None)
            ad_published = True
        except Exception as e:
            logging.warning(f"perm_publish_ad avto-reklama post yaratilmadi (pid {pid}): {e}")
    return {"ok": True, "product_id": pid, "blocked": blocked,
            "pending_owner": bool(needs_owner_approval and not blocked),
            "ad_published": ad_published,
            "moderation": (mod if blocked else None)}


class ProductEdit(BaseModel):
    name: Optional[str] = None
    price: Optional[float] = None
    description: Optional[str] = None
    stock_count: Optional[int] = None
    old_price: Optional[float] = None
    unit: Optional[str] = None   # #20 — o'lchov birligi
    min_price: Optional[float] = None   # #8 — maxfiy oxirgi narx
    min_order_qty: Optional[int] = None  # variant-buyurtma JAMI minimal son (optom)
    wholesale_price: Optional[float] = None   # eski yagona zina (moslik)
    wholesale_min_qty: Optional[int] = None   # eski yagona zina (moslik)
    wholesale_tiers: Optional[List[WholesaleTier]] = None  # optom: pachka-zina (bir nechta)
    image_url: Optional[str] = None
    images: Optional[List[str]] = None
    image_labels: Optional[List[str]] = None  # har rasm uchun nom (optom: rang nomi)
    attributes: Optional[List[AttrItem]] = None
    sale_mode: Optional[str] = None  # 'dona' | 'optom'
    pack_size: Optional[int] = None  # optom: 1 pachkadagi dona soni
    size_note: Optional[str] = None  # optom: razmer matni
    delivery_available: Optional[int] = None  # 1=yetkaziladi, 0=faqat olib ketish
    video_file_id: Optional[str] = None  # video (bo'sh satr = o'chirish)


@app.patch("/api/seller/product/{product_id}")
async def api_edit_product(product_id: int, p: ProductEdit, background: BackgroundTasks,
                           authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    prod = dict(_own_product_or_403(user, product_id))
    old_price = prod.get("price")
    fields = {}
    if p.name is not None:
        nm = p.name.strip()
        if not nm:
            raise HTTPException(status_code=400, detail="name_required")
        fields["name"] = nm
    if p.price is not None:
        if p.price <= 0:
            raise HTTPException(status_code=400, detail="bad_price")
        fields["price"] = float(p.price)
    if p.description is not None:
        fields["description"] = p.description.strip() or None
    if p.stock_count is not None:
        if p.stock_count < -1:
            raise HTTPException(status_code=400, detail="bad_stock")
        # -1 = cheksiz (NULL) — tahrirda cheksiz qilib qo'yish imkoni
        fields["stock_count"] = None if p.stock_count == -1 else p.stock_count
    if p.old_price is not None:
        # 0/bo'sh -> chegirmani olib tashlaydi (NULL)
        fields["old_price"] = float(p.old_price) if p.old_price and p.old_price > 0 else None
    if p.unit is not None:
        fields["unit"] = p.unit.strip()[:20] or None
    # Savdo turi (dona/optom). Partial patch: berilmasa mavjud tur saqlanadi.
    mode_provided = p.sale_mode is not None
    eff_mode = _norm_sale_mode(p.sale_mode) if mode_provided else _norm_sale_mode(prod.get("sale_mode"))
    if mode_provided:
        fields["sale_mode"] = eff_mode
    if eff_mode == "optom":
        # Optomda savdolashish yo'q — min_price doim NULL.
        fields["min_price"] = None
        if p.pack_size is not None:
            fields["pack_size"] = int(p.pack_size) if p.pack_size and p.pack_size > 0 else None
        if p.size_note is not None:
            fields["size_note"] = (p.size_note or "").strip()[:200] or None
    else:
        if mode_provided:   # optomdan donaga o'tkazilsa — pachka maydonlarini tozalaymiz
            fields["pack_size"] = None
            fields["size_note"] = None
        if p.min_price is not None:
            fields["min_price"] = float(p.min_price) if p.min_price and p.min_price > 0 else None
    if p.min_order_qty is not None:
        fields["min_order_qty"] = int(p.min_order_qty) if p.min_order_qty and p.min_order_qty > 1 else None
    if p.delivery_available is not None:
        fields["delivery_available"] = 1 if p.delivery_available else 0
    if p.video_file_id is not None:
        # bo'sh satr = videoni olib tashlash (NULL)
        fields["video_file_id"] = p.video_file_id.strip() or None
    # Optom narx — berilgan bo'lsa (zinalar yoki eski yagona) validatsiya bilan yangilaymiz.
    # Amaldagi dona narx = yangi narx (berilgan bo'lsa) yoki mavjud narx.
    if p.wholesale_tiers is not None or p.wholesale_price is not None or p.wholesale_min_qty is not None:
        eff_price = fields.get("price", prod.get("price"))
        fields.update(_wholesale_fields(eff_price, p.wholesale_tiers, p.wholesale_price, p.wholesale_min_qty))
    # #5 — TAHRIRDA ham moderatsiya: nom/tavsif o'zgargan bo'lsa qayta tekshiramiz.
    # Taqiqlangan bo'lsa darhol bloklaymiz (active'dan mod_blocked'ga). Toza bo'lsa
    # statusga tegmaymiz (avval bloklangan bo'lsa admin o'zi tasdiqlaydi).
    blocked = False
    if ("name" in fields or "description" in fields) and ai_assistant.is_enabled():
        new_name = fields.get("name", prod.get("name"))
        new_desc = fields.get("description", prod.get("description"))
        try:
            mod = await ai_assistant.moderate_product(
                name=new_name, description=new_desc or "", lang=get_user_lang(user) or DEFAULT_LANG)
        except Exception as e:
            logging.warning(f"edit moderate xato (pid {product_id}): {e}")
            mod = None
        if mod and mod.get("flagged"):
            blocked = True
            fields["status"] = "mod_blocked"
            fields["in_stock"] = 0
            fields["mod_reason"] = (mod.get("reason") or "")[:300]
            background.add_task(_notify_admins_moderation, product_id, new_name, user, mod)
    if fields:
        db.update_product_fields(product_id, **fields)
    # Rasmlar (galereya) — berilgan bo'lsa to'liq almashtiramiz (image_url ham sinxronlanadi)
    if p.images is not None:
        imgs = [f for f in p.images if f][:db.MAX_PRODUCT_IMAGES]
        db.set_product_images(product_id, imgs, labels=_image_labels_for(p, imgs))
    elif p.image_url is not None:
        db.update_product_fields(product_id, image_url=p.image_url)
    _save_attrs(product_id, p.attributes)
    # #16 — narx pasaygan bo'lsa, sevimliga qo'shganlarga xabar (faqat bloklanmagan bo'lsa)
    new_price = fields.get("price")
    if not blocked and new_price is not None and old_price and new_price < old_price:
        background.add_task(_notify_price_drop, product_id, prod.get("name"), old_price, new_price)
    return {"ok": True, "blocked": blocked}


@app.post("/api/seller/product/{product_id}/toggle")
def api_toggle_product(product_id: int, authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    _own_product_or_403(user, product_id)
    return {"ok": True, "in_stock": db.toggle_product_in_stock(product_id)}


@app.delete("/api/seller/product/{product_id}")
def api_delete_product(product_id: int, authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    _own_product_or_403(user, product_id)
    db.delete_product(product_id, deleted_by=user.get("id"), deleted_by_role=user.get("role"))
    return {"ok": True}


# ============================================================
# MAHSULOT SAVOLLARI (atribut shablonlari) — 3 rejim:
#   classic   — kategoriyaning statik savollari (db shablonlari)
#   ai_guided — AI mahsulotga mos savollar tuzadi
#   ai_smart  — AI tavsifdan ajratadi (known) + qolganini so'raydi (questions)
# AI o'chiq/xato bo'lsa — klassik shablonlarga qaytadi (bot bilan bir xil xulq).
# ============================================================
# Kategoriya atributlari uchun tayyor variant tugmalari — qo'lda yozishni bir tegishga
# aylantiradi. Sotuvchi ko'p yozadigan maydonlarni chiplardan tanlaydi; ro'yxatda yo'q
# bo'lsa "qo'lda kiritish" maydoniga yozadi (custom=True). multi=True — bir e'londa bir
# nechta qiymat (razmer 40,41,42 / rang qora,oq) tanlanadi — downstream `_split_opts`
# vergulni allaqachon variantlarga ajratadi, shuning uchun optom/variant bilan mos.
# Qiymatlar (rang/material nomlari) — DATA, saqlanadi va hammaga bir xil ko'rinadi,
# shuning uchun mavjud "select" chiplari kabi lokalizatsiya qilinmaydi (o'zbekcha).
# Data-driven: yangi kategoriya qo'shish uchun shu yerga bitta blok yoziladi.
_ATTR_PRESETS = {
    "Oyoq kiyimlari": {
        "size":     {"options": ["35", "36", "37", "38", "39", "40", "41", "42",
                                  "43", "44", "45", "46"],
                     "multi": True,  "custom": True},
        "color":    {"options": ["Qora", "Oq", "Jigarrang", "Ko'k", "Kulrang", "Qizil",
                                  "Yashil", "Sariq", "Pushti", "Bej", "Kumush", "Oltin"],
                     "multi": True,  "custom": True},
        "gender":   {"options": ["Erkak", "Ayol", "Bolalar", "Uniseks"],
                     "multi": False, "custom": False},
        "brand":    {"options": ["Nike", "Adidas", "Puma", "Reebok", "New Balance",
                                  "Converse", "Vans", "Skechers", "Ecco", "Asics",
                                  "Fila", "Salomon"],
                     "multi": False, "custom": True},
        "material": {"options": ["Charm", "Ekokcharm", "Zamsh", "Tekstil", "Rezina",
                                  "Mesh (to'r)"],
                     "multi": False, "custom": True},
        "season":   {"options": ["Yoz", "Qish", "Demi-sezon", "To'rt fasl"],
                     "multi": False, "custom": True},
    },
    "Kiyimlar": {
        "size":     {"options": ["XS", "S", "M", "L", "XL", "XXL", "XXXL"],
                     "multi": True,  "custom": True},
        "color":    {"options": ["Qora", "Oq", "Jigarrang", "Ko'k", "Kulrang", "Qizil",
                                  "Yashil", "Sariq", "Pushti", "Bej", "Kumush", "Oltin"],
                     "multi": True,  "custom": True},
        "gender":   {"options": ["Erkak", "Ayol", "Bolalar", "Uniseks"],
                     "multi": False, "custom": False},
        "brand":    {"options": ["Nike", "Adidas", "Zara", "Puma", "H&M", "Uniqlo",
                                  "Mango", "Reebok", "Tommy Hilfiger", "Lacoste",
                                  "Levi's", "Gucci"],
                     "multi": False, "custom": True},
        "season":   {"options": ["Bahor", "Yoz", "Kuz", "Qish", "To'rt fasl"],
                     "multi": False, "custom": True},
    },
    "Elektronika": {
        "brand":     {"options": ["Samsung", "Apple", "Xiaomi", "Redmi", "Huawei", "Honor",
                                   "Realme", "Infinix", "Tecno", "Oppo", "OnePlus", "Nokia"],
                      "multi": False, "custom": True},
        "condition": {"options": ["Yangi", "Ishlatilgan", "Qadoqda"],
                      "multi": False, "custom": False},
        "warranty":  {"options": ["Yo'q", "6 oy", "1 yil", "2 yil", "3 yil"],
                      "multi": False, "custom": True},
        "memory":    {"options": ["32GB", "64GB", "128GB", "256GB", "512GB", "1TB"],
                      "multi": False, "custom": True},
    },
    "Ichimliklar": {
        "volume": {"options": ["0.25L", "0.33L", "0.5L", "1L", "1.5L", "2L", "5L"],
                   "multi": False, "custom": True},
        "cold":   {"options": ["Sovuq", "Issiq", "Ikkalasi"],
                   "multi": False, "custom": False},
    },
    "Ehtiyot Qismlar": {
        "car_make":  {"options": ["Nexia", "Cobalt", "Gentra", "Spark", "Malibu",
                                  "Lacetti", "Damas", "Matiz", "Captiva", "Tracker",
                                  "Onix", "Tahoe"],
                      "multi": False, "custom": True},
        "part_type": {"options": ["Dvigatel", "Korobka", "Xodovoy", "Kuzov", "Elektrika",
                                  "Tormoz", "Salon", "Optika"],
                      "multi": False, "custom": True},
        "condition": {"options": ["Yangi", "Ishlatilgan"],
                      "multi": False, "custom": False},
        "oem":       {"options": ["Original", "Kopiya", "Zamena"],
                      "multi": False, "custom": True},
    },
    "Xojalik Mollari": {
        "brand":  {"options": ["Ariel", "Domestos", "Fairy", "Tide", "Persil", "Sarma",
                               "Bref", "Comet", "Myth"],
                   "multi": False, "custom": True},
        "weight": {"options": ["0.5L", "1L", "2L", "5L", "500g", "1kg", "5kg"],
                   "multi": False, "custom": True},
    },
    "Oziq-ovqat": {
        "weight":    {"options": ["100g", "250g", "500g", "1kg", "2kg", "5kg", "10kg"],
                      "multi": False, "custom": True},
        "freshness": {"options": ["Bugungi", "Kechagi", "Yangi", "Muzlatilgan"],
                      "multi": False, "custom": True},
        "origin":    {"options": ["O'zbekiston", "Import", "Rossiya", "Qozog'iston",
                                  "Turkiya", "Xitoy"],
                      "multi": False, "custom": True},
    },
    "Taomlar": {
        "serving":  {"options": ["1 kishi", "2-3 kishi", "4-5 kishi", "Katta oila"],
                     "multi": False, "custom": True},
        "spicy":    {"options": ["Achchiq", "O'rta", "Achchiq emas"],
                     "multi": False, "custom": False},
        "ready_in": {"options": ["Hozir tayyor", "15 daqiqa", "30 daqiqa", "1 soat"],
                     "multi": False, "custom": True},
    },
    "Go'zallik va parfyumeriya": {
        "type":   {"options": ["Parfyum", "Krem", "Pomada", "Tush", "Shampun", "Maska",
                               "Loson", "Atir"],
                   "multi": False, "custom": True},
        "brand":  {"options": ["Chanel", "Dior", "L'Oreal", "Nivea", "Maybelline", "MAC",
                               "Estee Lauder", "Garnier"],
                   "multi": False, "custom": True},
        "volume": {"options": ["30ml", "50ml", "100ml", "150ml", "200ml"],
                   "multi": False, "custom": True},
        "gender": {"options": ["Erkak", "Ayol", "Uniseks"],
                   "multi": False, "custom": False},
    },
    "Salomatlik va dorixona": {
        "type":         {"options": ["Dori", "Vitamin", "BAD", "Bandaj", "Tibbiy asbob",
                                     "Gigiena"],
                         "multi": False, "custom": True},
        "form":         {"options": ["Tabletka", "Kapsula", "Sirop", "Malham", "Ukol",
                                     "Tomchi"],
                         "multi": False, "custom": True},
        "prescription": {"options": ["Retseptsiz", "Retsept bilan"],
                         "multi": False, "custom": False},
    },
    "Bolalar mahsulotlari": {
        "type":   {"options": ["O'yinchoq", "Kiyim", "Aravacha", "Taom", "Gigiena",
                               "Maktab"],
                   "multi": False, "custom": True},
        "age":    {"options": ["0-1 yosh", "1-3 yosh", "3-6 yosh", "6-12 yosh", "12+ yosh"],
                   "multi": False, "custom": True},
        "gender": {"options": ["O'g'il bola", "Qiz bola", "Uniseks"],
                   "multi": False, "custom": False},
    },
    "Sport va dam olish": {
        "type":   {"options": ["Sport kiyim", "Trenajyor", "Velosiped", "Top", "Palatka",
                               "Baliq ovi", "Gantel"],
                   "multi": False, "custom": True},
        "size":   {"options": ["S", "M", "L", "XL", "XXL"],
                   "multi": True,  "custom": True},
        "gender": {"options": ["Erkak", "Ayol", "Uniseks"],
                   "multi": False, "custom": False},
    },
    "Uy va mebel": {
        "type":     {"options": ["Divan", "Stol", "Stul", "Shkaf", "Krovat", "Polka",
                                 "Gilam", "Yoritish"],
                     "multi": False, "custom": True},
        "material": {"options": ["Yog'och", "Metall", "Plastik", "Shisha", "Mato", "Charm"],
                     "multi": False, "custom": True},
        "room":     {"options": ["Yotoqxona", "Mehmonxona", "Oshxona", "Bolalar", "Ofis",
                                 "Hammom"],
                     "multi": False, "custom": True},
    },
    "Kitob va kanstovarlar": {
        "type":     {"options": ["Kitob", "Daftar", "Qalam", "Ruchka", "Papka", "Bo'yoq",
                                 "Ranglar"],
                     "multi": False, "custom": True},
        "language": {"options": ["O'zbek", "Rus", "Ingliz", "Arab"],
                     "multi": False, "custom": True},
    },
    "Qurilish mollari": {
        "type": {"options": ["G'isht", "Sement", "Bo'yoq", "Plitka", "Armatura", "Gips",
                             "Quruq aralashma", "Asbob"],
                 "multi": False, "custom": True},
        "unit": {"options": ["Dona", "Kg", "Tonna", "Metr", "m²", "Qop", "Litr"],
                 "multi": False, "custom": True},
    },
    "Hayvonlar uchun": {
        "animal": {"options": ["It", "Mushuk", "Qush", "Baliq", "Kemiruvchi", "Boshqa"],
                   "multi": False, "custom": True},
        "type":   {"options": ["Ozuqa", "O'yinchoq", "Aksessuar", "Gigiena", "Qafas/Uy",
                               "Dori"],
                   "multi": False, "custom": True},
    },
    "Gul va sovg'alar": {
        "type":     {"options": ["Atirgul", "Buket", "Tuvakli gul", "Sovg'a", "Sharlar",
                                 "Otkritka"],
                     "multi": False, "custom": True},
        "occasion": {"options": ["Tug'ilgan kun", "To'y", "8-mart", "Sevishganlar kuni",
                                 "Yubiley", "Boshqa"],
                     "multi": False, "custom": True},
    },
}

# Preset chip qiymatlarining RU ko'rinishi. MUHIM: bu FAQAT ko'rsatiladigan yorliq —
# bazaga DOIM o'zbekcha kanonik qiymat (option value) saqlanadi, shunda barcha mahsulotda
# rang/material bir xil bo'ladi (ko'rsatish lokal, saqlash kanonik). Raqam/brend/model
# (36, 128GB, Nike, Nexia) ikkala tilda bir xil — bu yerda YO'Q, o'zi qaytadi (fallback).
_ATTR_RU_COMMON = {
    # ranglar
    "Qora": "Чёрный", "Oq": "Белый", "Jigarrang": "Коричневый", "Ko'k": "Синий",
    "Kulrang": "Серый", "Qizil": "Красный", "Yashil": "Зелёный", "Sariq": "Жёлтый",
    "Pushti": "Розовый", "Bej": "Бежевый", "Kumush": "Серебристый", "Oltin": "Золотой",
    # jinsi / kimga
    "Erkak": "Мужской", "Ayol": "Женский", "Bolalar": "Детский", "Uniseks": "Унисекс",
    "O'g'il bola": "Мальчик", "Qiz bola": "Девочка",
    # material
    "Charm": "Кожа", "Ekokcharm": "Экокожа", "Zamsh": "Замша", "Tekstil": "Текстиль",
    "Rezina": "Резина", "Mesh (to'r)": "Сетка", "Yog'och": "Дерево", "Metall": "Металл",
    "Plastik": "Пластик", "Shisha": "Стекло", "Mato": "Ткань",
    # fasl
    "Yoz": "Лето", "Qish": "Зима", "Demi-sezon": "Демисезон", "To'rt fasl": "Всесезон",
    "Bahor": "Весна", "Kuz": "Осень",
    # holat / oem
    "Ishlatilgan": "Б/у", "Qadoqda": "В упаковке",
    "Original": "Оригинал", "Kopiya": "Копия", "Zamena": "Замена",
    # ehtiyot qism turi
    "Dvigatel": "Двигатель", "Korobka": "Коробка", "Xodovoy": "Ходовая", "Kuzov": "Кузов",
    "Elektrika": "Электрика", "Tormoz": "Тормоза", "Salon": "Салон", "Optika": "Оптика",
    # ichimlik / taom
    "Sovuq": "Холодный", "Issiq": "Горячий", "Ikkalasi": "Оба",
    "Achchiq": "Острый", "O'rta": "Средний", "Achchiq emas": "Не острый",
    "1 kishi": "1 человек", "2-3 kishi": "2-3 человека", "4-5 kishi": "4-5 человек",
    "Katta oila": "Большая семья",
    "Hozir tayyor": "Готово сейчас", "15 daqiqa": "15 минут", "30 daqiqa": "30 минут",
    "1 soat": "1 час",
    # oziq-ovqat
    "Bugungi": "Сегодняшний", "Kechagi": "Вчерашний", "Muzlatilgan": "Замороженный",
    "O'zbekiston": "Узбекистан", "Import": "Импорт", "Rossiya": "Россия",
    "Qozog'iston": "Казахстан", "Turkiya": "Турция", "Xitoy": "Китай",
    # kafolat
    "Yo'q": "Нет", "6 oy": "6 мес", "1 yil": "1 год", "2 yil": "2 года", "3 yil": "3 года",
    # go'zallik
    "Parfyum": "Парфюм", "Krem": "Крем", "Pomada": "Помада", "Tush": "Тушь",
    "Shampun": "Шампунь", "Maska": "Маска", "Loson": "Лосьон", "Atir": "Духи",
    # salomatlik
    "Dori": "Лекарство", "Vitamin": "Витамины", "BAD": "БАД", "Bandaj": "Бандаж",
    "Tibbiy asbob": "Мед. прибор", "Gigiena": "Гигиена",
    "Tabletka": "Таблетка", "Kapsula": "Капсула", "Sirop": "Сироп", "Malham": "Мазь",
    "Ukol": "Укол", "Tomchi": "Капли",
    "Retseptsiz": "Без рецепта", "Retsept bilan": "По рецепту",
    # bolalar
    "O'yinchoq": "Игрушка", "Kiyim": "Одежда", "Aravacha": "Коляска", "Taom": "Питание",
    "Maktab": "Школа",
    "0-1 yosh": "0-1 год", "1-3 yosh": "1-3 года", "3-6 yosh": "3-6 лет",
    "6-12 yosh": "6-12 лет", "12+ yosh": "12+ лет",
    # sport
    "Sport kiyim": "Спортивная одежда", "Trenajyor": "Тренажёр", "Velosiped": "Велосипед",
    "Top": "Мяч", "Palatka": "Палатка", "Baliq ovi": "Рыбалка", "Gantel": "Гантели",
    # uy / mebel
    "Divan": "Диван", "Stol": "Стол", "Stul": "Стул", "Shkaf": "Шкаф", "Krovat": "Кровать",
    "Polka": "Полка", "Gilam": "Ковёр", "Yoritish": "Освещение",
    "Yotoqxona": "Спальня", "Mehmonxona": "Гостиная", "Oshxona": "Кухня",
    "Ofis": "Офис", "Hammom": "Ванная",
    # kitob / kanstovar
    "Kitob": "Книга", "Daftar": "Тетрадь", "Qalam": "Карандаш", "Ruchka": "Ручка",
    "Papka": "Папка", "Bo'yoq": "Краска", "Ranglar": "Краски",
    "O'zbek": "Узбекский", "Rus": "Русский", "Ingliz": "Английский", "Arab": "Арабский",
    # qurilish
    "G'isht": "Кирпич", "Sement": "Цемент", "Plitka": "Плитка", "Armatura": "Арматура",
    "Gips": "Гипс", "Quruq aralashma": "Сухая смесь", "Asbob": "Инструмент",
    "Dona": "Штука", "Kg": "Кг", "Tonna": "Тонна", "Metr": "Метр", "Qop": "Мешок",
    "Litr": "Литр",
    # hayvonlar
    "It": "Собака", "Mushuk": "Кошка", "Qush": "Птица", "Baliq": "Рыба",
    "Kemiruvchi": "Грызун", "Boshqa": "Другое",
    "Ozuqa": "Корм", "Aksessuar": "Аксессуар", "Qafas/Uy": "Клетка/Дом",
    # gul / sovg'a
    "Atirgul": "Роза", "Buket": "Букет", "Tuvakli gul": "Горшечный цветок",
    "Sovg'a": "Подарок", "Sharlar": "Шары", "Otkritka": "Открытка",
    "Tug'ilgan kun": "День рождения", "To'y": "Свадьба", "8-mart": "8 марта",
    "Sevishganlar kuni": "День влюблённых", "Yubiley": "Юбилей",
}
# Bir xil o'zbekcha so'z turli maydonda boshqa RU'ga tarjima bo'ladigan holatlar
# (masalan "Yangi": holatda=Новый, lekin yangilikda=Свежий). Maydon bo'yicha ustuvor.
_ATTR_RU_BY_KEY = {
    "freshness":  {"Yangi": "Свежий"},
    "condition":  {"Yangi": "Новый"},
}


def _ru_attr_label(attr_key, value):
    """Preset qiymatining RU yorlig'i (maydon-aniq ustunlik, so'ng umumiy, so'ng o'zi)."""
    by_key = _ATTR_RU_BY_KEY.get(attr_key)
    if by_key and value in by_key:
        return by_key[value]
    return _ATTR_RU_COMMON.get(value, value)


def _norm_question(q):
    """db/AI shablonini frontend uchun barqaror shaklga keltiradi."""
    typ = q.get("attr_type") or "text"
    hint = q.get("hint") or ""
    options = []
    if typ == "select" and hint:
        options = [o.strip() for o in str(hint).split("/") if o.strip()]
    return {"key": q.get("attr_key"), "label": q.get("attr_label") or q.get("attr_key"),
            "type": typ, "required": bool(q.get("is_required")),
            "hint": hint, "options": options, "multi": False, "custom": False}


def _apply_attr_presets(questions, cat_name, lang=DEFAULT_LANG):
    """Kategoriyaga mos tayyor tugma-variantlarni savollarga biriktiradi (chip UX).
    RU tilida option'lar {value, label} bo'ladi — value KANONIK (o'zbekcha, saqlanadi),
    label ko'rsatiladi. UZ tilida oddiy string (label=value)."""
    presets = _ATTR_PRESETS.get(cat_name or "")
    if not presets:
        return questions
    ru = (lang == "ru")
    for q in questions:
        key = q.get("key")
        p = presets.get(key)
        if not p:
            continue
        vals = list(p["options"])
        q["options"] = ([{"value": v, "label": _ru_attr_label(key, v)} for v in vals]
                        if ru else vals)
        q["multi"] = bool(p.get("multi"))
        q["custom"] = bool(p.get("custom", True))
    return questions


def _category_name(category_id, lang):
    if not category_id:
        return ""
    try:
        for c in db.get_all_categories():
            if c[0] == category_id:
                return c[1]
    except Exception:
        pass
    return ""


@app.get("/api/seller/product-questions")
async def api_product_questions(category_id: Optional[int] = Query(None),
                                name: str = Query(""), description: str = Query(""),
                                mode: str = Query("classic"),
                                authorization: str = Header(None)):
    user = dict(_buyer_from_auth(authorization))
    if user.get("role") not in ("seller", "admin") and not user.get("is_approved"):
        raise HTTPException(status_code=403, detail="not_seller")
    mode = mode if mode in ("classic", "ai_guided", "ai_smart") else "classic"
    lang = get_user_lang(user) or DEFAULT_LANG

    def _classic():
        try:
            tmpls = db.get_category_templates(category_id) if category_id else []
        except Exception as e:
            logging.warning(f"category templates xato (cat {category_id}): {e}")
            tmpls = []
        qs = [_norm_question(dict(t)) for t in tmpls]
        return _apply_attr_presets(qs, _category_name(category_id, lang) if category_id else "", lang)

    if mode == "classic" or not ai_assistant.is_enabled():
        return {"questions": _classic(), "known": {}, "source": "classic"}

    try:
        result = await ai_assistant.generate_product_questions(
            name=name or "", category=_category_name(category_id, lang),
            description=description or "", lang=lang, smart=(mode == "ai_smart"))
    except Exception as e:
        logging.warning(f"AI savollar (web) xato: {e}")
        result = None
    if not result:
        return {"questions": _classic(), "known": {}, "source": "classic_fallback"}

    questions = [_norm_question(q) for q in (result.get("questions") or [])]
    # known: {key: {value, label}} — frontend oldindan to'ldiradi
    known = {}
    for k, v in (result.get("known") or {}).items():
        if isinstance(v, dict):
            known[k] = {"value": v.get("value"), "label": v.get("label") or k}
    return {"questions": questions, "known": known, "source": mode}


@app.get("/api/image/{file_id}")
async def api_image(file_id: str, request: Request = None):
    """Telegram file_id'ni haqiqiy rasmga aylantiradi (getFile → yuklab → cache → stream).
    Autentifikatsiyasiz (img-teglar header yubora olmaydi), lekin cache-miss (Telegram'ga
    yangi so'rov) IP bo'yicha cheklanadi — proksi sifatida suiiste'molni cheklaydi (audit #4)."""
    if not BOT_TOKEN:
        raise HTTPException(status_code=503, detail="no token")
    # Disk-cache (getFile rate-limitiga tushmaslik uchun) — keshlangan rasm cheklanmaydi
    safe = hashlib.sha256(file_id.encode()).hexdigest()
    cache_path = os.path.join(IMG_CACHE_DIR, safe + ".jpg")
    if os.path.exists(cache_path):
        return FileResponse(cache_path, media_type="image/jpeg",
                            headers={"Cache-Control": "public, max-age=604800"})
    # Cache-miss: Telegram'ga yangi so'rov — IP bo'yicha 60/min (hammerlashning oldini oladi)
    if request is not None:
        ip = (request.headers.get("x-forwarded-for") or
              (request.client.host if request.client else "?")).split(",")[0].strip()
        _rate_limit("img", ip, 60, 60)
    try:
        async with httpx.AsyncClient(timeout=20) as client:
            meta = await client.get(
                f"https://api.telegram.org/bot{BOT_TOKEN}/getFile",
                params={"file_id": file_id},
            )
            data = meta.json()
            if not data.get("ok"):
                raise HTTPException(status_code=404, detail="file not found")
            path = data["result"]["file_path"]
            img = await client.get(f"https://api.telegram.org/file/bot{BOT_TOKEN}/{path}")
            img.raise_for_status()
            content = img.content
        with open(cache_path, "wb") as f:
            f.write(content)
        return Response(content=content, media_type="image/jpeg",
                        headers={"Cache-Control": "public, max-age=604800"})
    except HTTPException:
        raise
    except Exception as e:
        logging.warning(f"image proxy xatosi ({file_id[:12]}...): {e}")
        raise HTTPException(status_code=502, detail="image fetch failed")


@app.get("/api/video/{file_id}")
async def api_video(file_id: str, request: Request = None):
    """Telegram file_id'ni videoga aylantiradi (rasm proksisi bilan bir uslub: getFile →
    yuklab → disk-cache → stream). Autentifikatsiyasiz (<video src> header yubora olmaydi);
    cache-miss IP bo'yicha cheklanadi. Bot API getFile ≤20MB — upload shu chegarada."""
    if not BOT_TOKEN:
        raise HTTPException(status_code=503, detail="no token")
    safe = hashlib.sha256(file_id.encode()).hexdigest()
    cache_path = os.path.join(IMG_CACHE_DIR, safe + ".mp4")
    if os.path.exists(cache_path):
        return FileResponse(cache_path, media_type="video/mp4",
                            headers={"Cache-Control": "public, max-age=604800"})
    # Video og'ir — cache-miss'ni rasmga qaraganda qattiqroq cheklaymiz (10/min IP)
    if request is not None:
        ip = (request.headers.get("x-forwarded-for") or
              (request.client.host if request.client else "?")).split(",")[0].strip()
        _rate_limit("vid", ip, 10, 60)
    try:
        async with httpx.AsyncClient(timeout=60) as client:
            meta = await client.get(
                f"https://api.telegram.org/bot{BOT_TOKEN}/getFile",
                params={"file_id": file_id},
            )
            data = meta.json()
            if not data.get("ok"):
                raise HTTPException(status_code=404, detail="file not found")
            path = data["result"]["file_path"]
            vid = await client.get(f"https://api.telegram.org/file/bot{BOT_TOKEN}/{path}")
            vid.raise_for_status()
            content = vid.content
        with open(cache_path, "wb") as f:
            f.write(content)
        return Response(content=content, media_type="video/mp4",
                        headers={"Cache-Control": "public, max-age=604800"})
    except HTTPException:
        raise
    except Exception as e:
        logging.warning(f"video proxy xatosi ({file_id[:12]}...): {e}")
        raise HTTPException(status_code=502, detail="video fetch failed")


_bot_username_cache = {"value": None}


@app.get("/api/config")
async def api_config(authorization: str = Header(None)):
    require_auth(authorization)
    # bot username — buyurtma tugmasi bot'ga deep-link qilishi uchun (getMe, keshlanadi)
    if _bot_username_cache["value"] is None and BOT_TOKEN:
        try:
            async with httpx.AsyncClient(timeout=10) as client:
                r = await client.get(f"https://api.telegram.org/bot{BOT_TOKEN}/getMe")
                d = r.json()
                if d.get("ok"):
                    _bot_username_cache["value"] = d["result"]["username"]
        except Exception:
            pass
    return {"bot_username": _bot_username_cache["value"], "channel_url": CHANNEL_URL}


@app.get("/api/health")
def api_health():
    return {"ok": True, "backend": db.backend}


@app.get("/call/{phone}")
def call_page(phone: str):
    """📞 Qo'ng'iroq ko'prigi — kanal reklama TUGMASI uchun. Telegram inline tugma
    `tel:` havolani qabul qilmaydi, shuning uchun tugma shu HTTPS sahifaga keladi;
    sahifa ochilishi bilan telefon tergichini (tel:) avtomatik ishga tushiradi.
    Ochilmasa — katta zaxira tugma + raqam ko'rinadi. Auth talab qilinmaydi."""
    digits = re.sub(r"[^\d]", "", phone or "")
    if not (7 <= len(digits) <= 15):
        raise HTTPException(status_code=404, detail="bad_phone")
    tel = f"tel:+{digits}"
    # +998 90 123 45 67 ko'rinishida chiroyli format (uz raqamlari uchun)
    pretty = f"+{digits}"
    if len(digits) == 12 and digits.startswith("998"):
        pretty = f"+998 {digits[3:5]} {digits[5:8]} {digits[8:10]} {digits[10:12]}"
    page = f"""<!doctype html><html lang="uz"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<meta http-equiv="refresh" content="0;url={tel}">
<title>Qo'ng'iroq — {BRAND_NAME}</title>
<style>
 body{{margin:0;font-family:system-ui,-apple-system,sans-serif;min-height:100vh;display:flex;
      align-items:center;justify-content:center;background:#f4f6f9;color:#111}}
 @media (prefers-color-scheme:dark){{body{{background:#17212b;color:#f5f5f5}}}}
 .c{{text-align:center;padding:28px}}
 .n{{font-size:22px;font-weight:800;margin:14px 0 22px;letter-spacing:.5px}}
 a.b{{display:inline-block;background:#2ea6ff;color:#fff;text-decoration:none;font-size:19px;
     font-weight:700;padding:16px 34px;border-radius:14px;box-shadow:0 4px 14px rgba(46,166,255,.35)}}
 .h{{margin-top:16px;font-size:13px;opacity:.6}}
</style></head><body><div class="c">
 <div style="font-size:52px">📞</div>
 <div class="n">{pretty}</div>
 <a class="b" href="{tel}">📞 Qo'ng'iroq qilish</a>
 <div class="h">{html.escape(BRAND_NAME)} · avtomatik ochilmasa tugmani bosing</div>
</div><script>location.href={tel!r};</script></body></html>"""
    return HTMLResponse(page)


# ============================================================
# ADMIN PANELI (G bo'lagi) — faqat admin rolli foydalanuvchi
# Bot main.py'dagi admin handlerlarning to'liq parite ko'chirmasi.
# ============================================================
def _admin_from_auth(authorization):
    """initData'dan adminni (DB user) qaytaradi yoki 403. Admin = role=='admin' yoki ADMIN_ID."""
    user = dict(_buyer_from_auth(authorization))
    if user.get("role") != "admin" and user.get("telegram_id") != ADMIN_ID:
        raise HTTPException(status_code=403, detail="not_admin")
    return user


# ============================================================
# MONETIZATSIYA SKELETI (#22 + #18)
# Admin yoqadigan/o'chiradigan bayroqlar. DEFAULT HAMMASI O'CHIQ — yoqilmaguncha
# foydalanuvchiga hech narsa o'zgarmaydi (app/bot bepul ishlayveradi). Har "pul
# oluvchi" nuqta kelajakda monetization_config()'ni tekshiradi.
#   (kalit, tip, default)  — tip: 'bool' (0/1) | 'num'
# ============================================================
MONETIZATION_SPEC = [
    ("mon_enabled",              "bool", "0"),  # BOSH RUBILNIK — buni yoqmaguncha hech biri ishlamaydi
    ("mon_commission_enabled",   "bool", "0"),  # har sotuvdan komissiya
    ("mon_commission_percent",   "num",  "0"),  # 0..100
    ("mon_boost_enabled",        "bool", "0"),  # pullik "ko'tarish" reklamasi
    ("mon_boost_price",          "num",  "0"),
    ("mon_boost_days",           "num",  "7"),   # boost necha kun amal qiladi
    ("mon_subscription_enabled", "bool", "0"),  # sotuvchi pro-obunasi
    ("mon_subscription_price",   "num",  "0"),
    ("mon_subscription_days",    "num",  "30"),  # obuna necha kunga
    ("mon_free_product_limit",   "num",  "0"),   # 0 = limitsiz; >0 = bepul sotuvchiga shu sondan ortiq faol mahsulot mumkin emas (Pro = cheksiz)
    ("mon_pro_free_boosts",      "num",  "4"),   # Pro: oyiga shuncha bepul boost (0 = bepul boost yo'q)
    ("mon_free_reels_limit",     "num",  "2"),   # bepul sotuvchi: oyiga AI reels soni (Pro = cheksiz; 0 = reels Pro-only)
    ("mon_free_scheduled_limit", "num",  "2"),   # bepul sotuvchi: bir vaqtda faol rejalashtirilgan post (Pro = cheksiz; 0 = Pro-only)
    ("mon_pay_click_enabled",    "bool", "0"),  # Click to'lov provayderi
    ("mon_pay_payme_enabled",    "bool", "0"),  # Payme to'lov provayderi
    ("mon_pay_paynet_enabled",   "bool", "0"),  # Paynet to'lov provayderi
]
_MON_KEYS = {k: typ for k, typ, _ in MONETIZATION_SPEC}


def monetization_config():
    """Typed config — barcha kalitlar default bilan (admin paneli o'qiydi)."""
    raw = db.get_all_settings()
    out = {}
    for k, typ, dflt in MONETIZATION_SPEC:
        v = raw.get(k, dflt)
        if typ == "bool":
            out[k] = (str(v) == "1")
        else:
            try:
                out[k] = float(v)
            except (TypeError, ValueError):
                out[k] = 0.0
    return out


def monetization_public():
    """Frontend uchun xavfsiz qism — qaysi imkoniyatlar HOZIR yoqilgan (master + har biri)."""
    c = monetization_config()
    on = c["mon_enabled"]
    return {
        "enabled": on,
        "commission": on and c["mon_commission_enabled"],
        "commission_percent": c["mon_commission_percent"] if (on and c["mon_commission_enabled"]) else 0,
        "boost": on and c["mon_boost_enabled"],
        "boost_price": c["mon_boost_price"] if (on and c["mon_boost_enabled"]) else 0,
        "boost_days": int(c["mon_boost_days"]),
        "subscription": on and c["mon_subscription_enabled"],
        "subscription_price": c["mon_subscription_price"] if (on and c["mon_subscription_enabled"]) else 0,
        "subscription_days": int(c["mon_subscription_days"]),
        "free_product_limit": int(c["mon_free_product_limit"]) if (on and c["mon_subscription_enabled"]) else 0,
        "pro_free_boosts": int(c["mon_pro_free_boosts"]) if (on and c["mon_subscription_enabled"]) else 0,
        "free_reels_limit": int(c["mon_free_reels_limit"]) if (on and c["mon_subscription_enabled"]) else 0,
        "free_scheduled_limit": int(c["mon_free_scheduled_limit"]) if (on and c["mon_subscription_enabled"]) else 0,
        "pay_click": on and c["mon_pay_click_enabled"],
        "pay_payme": on and c["mon_pay_payme_enabled"],
        "pay_paynet": on and c["mon_pay_paynet_enabled"],
    }


@app.get("/api/admin/monetization")
def api_admin_monetization_get(authorization: str = Header(None)):
    _admin_from_auth(authorization)
    return monetization_config()


@app.get("/api/admin/payments")
def api_admin_payments(state: str = Query("pending"), authorization: str = Header(None)):
    """Admin to'lovlar ro'yxati (default: kutilayotganlar). Provayder ulanmagan paytda
    admin shu ekranda sotuvchining boost/Pro to'lovini qo'lda tasdiqlaydi (dev-confirm)."""
    _admin_from_auth(authorization)
    st = state if state in ("pending", "paid", "cancelled") else None
    return {"payments": db.get_payments_admin(state=st, limit=100)}


@app.post("/api/admin/monetization")
async def api_admin_monetization_set(request: Request, authorization: str = Header(None)):
    """Faqat kelgan kalitlar yangilanadi (qisman update). Noma'lum kalit e'tiborsiz."""
    _admin_from_auth(authorization)
    body = await request.json()
    if not isinstance(body, dict):
        raise HTTPException(status_code=400, detail="bad_body")
    for k, v in body.items():
        typ = _MON_KEYS.get(k)
        if not typ:
            continue
        if typ == "bool":
            db.set_setting(k, "1" if v in (True, 1, "1", "true") else "0")
        else:
            try:
                num = float(v)
            except (TypeError, ValueError):
                raise HTTPException(status_code=400, detail=f"bad_value:{k}")
            if num < 0:
                raise HTTPException(status_code=400, detail=f"bad_value:{k}")
            if k == "mon_commission_percent" and num > 100:
                raise HTTPException(status_code=400, detail="bad_percent")
            db.set_setting(k, num)
    return {"ok": True, "config": monetization_config()}


@app.get("/api/admin/commissions")
def api_admin_commissions(authorization: str = Header(None)):
    """Sotuvchilar kesimida komissiya: admin qaysi sotuvchidan qancha olishi (owed) va
    allaqachon undirgani (paid). Hozirgi foiz ham qaytadi."""
    _admin_from_auth(authorization)
    mc = monetization_config()
    sellers = db.get_commission_by_sellers()
    total_owed = round(sum(float(s.get("owed") or 0) for s in sellers), 2)
    total_paid = round(sum(float(s.get("paid") or 0) for s in sellers), 2)
    return {
        "active": bool(mc["mon_enabled"] and mc["mon_commission_enabled"]),
        "percent": mc["mon_commission_percent"],
        "sellers": sellers,
        "total_owed": total_owed,
        "total_paid": total_paid,
    }


@app.post("/api/admin/commissions/settle")
async def api_admin_commission_settle(request: Request, authorization: str = Header(None)):
    """Admin sotuvchidan komissiyani naqd/pul orqali undirgach, uning BARCHA ochiq qarzini
    'to'landi' deb yopadi va to'lovlar daftariga (purpose='commission') yozadi."""
    _admin_from_auth(authorization)
    body = await request.json()
    try:
        seller_id = int(body.get("seller_id"))
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="bad_seller_id")
    res = db.settle_seller_commission(seller_id)
    if res["count"] > 0:
        try:
            pid = db.create_payment(seller_id, "commission", res["amount"],
                                    ref_id=seller_id, provider="manual")
            db.set_payment_state(pid, "paid")
        except Exception as e:
            logging.warning(f"komissiya settlement to'lov yozuvi xato (seller {seller_id}): {e}")
    return {"ok": True, **res}


# ============================================================
# TO'LOVLAR — Click / Payme (monetizatsiya #18)
# Boost va Pro-obuna shu rel orqali to'lanadi. Provayder kalitlari .env'dan; yo'q
# bo'lsa endpointlar "sozlanmagan" deydi (sandbox-ready). Kalit yo'q payti sinov
# uchun admin-only "qo'lda tasdiqlash" yo'li bor (/api/pay/dev-confirm).
# ============================================================
import json as _json
import time as _time

CLICK_SERVICE_ID   = os.getenv("CLICK_SERVICE_ID", "")
CLICK_MERCHANT_ID  = os.getenv("CLICK_MERCHANT_ID", "")
CLICK_SECRET_KEY   = os.getenv("CLICK_SECRET_KEY", "")
PAYME_MERCHANT_ID  = os.getenv("PAYME_MERCHANT_ID", "")
PAYME_KEY          = os.getenv("PAYME_KEY", "")  # kassir kaliti (webhook Basic-auth paroli)
PAYME_CHECKOUT_URL = os.getenv("PAYME_CHECKOUT_URL", "https://checkout.paycom.uz")
CLICK_CHECKOUT_URL = os.getenv("CLICK_CHECKOUT_URL", "https://my.click.uz/services/pay")
# Payme kabinetida "hisob (account)" maydonining nomi shu bo'lishi kerak (default: payment_id).
# Kabinetda boshqacha atalса (masalan order_id), .env'da PAYME_ACCOUNT_FIELD bilan moslang.
PAYME_ACCOUNT_FIELD = os.getenv("PAYME_ACCOUNT_FIELD", "payment_id")

# Paynet (merchant JSON-RPC, Payme'ga o'xshash naqsh). Webhook Basic-auth: LOGIN:PASSWORD.
# DIQQAT: Paynet merchant protokoli (metod nomlari/imzo) Payme bilan deyarli bir xil, ammo
# kabinetdan olingan hujjat bilan TASDIQLANISHI shart — bayroq ulanmaguncha o'chiq turadi.
PAYNET_MERCHANT_ID  = os.getenv("PAYNET_MERCHANT_ID", "")
PAYNET_LOGIN        = os.getenv("PAYNET_LOGIN", "")
PAYNET_PASSWORD     = os.getenv("PAYNET_PASSWORD", "")
PAYNET_CHECKOUT_URL = os.getenv("PAYNET_CHECKOUT_URL", "")  # bo'sh = redirect-checkout yo'q


def _click_configured():
    return bool(CLICK_SERVICE_ID and CLICK_MERCHANT_ID and CLICK_SECRET_KEY)


def _payme_configured():
    return bool(PAYME_MERCHANT_ID and PAYME_KEY)


def _paynet_configured():
    return bool(PAYNET_MERCHANT_ID and PAYNET_LOGIN and PAYNET_PASSWORD)


def _fulfill_payment(payment):
    """To'lov 'paid' bo'lganda maqsadni bajaradi (idempotent — chaqiruvchi state'ni
    'paid'ga o'tkazgandan keyin chaqiradi). boost → mahsulotni ko'taradi; subscription → Pro uzaytiradi."""
    cfg = monetization_config()
    purpose = payment.get("purpose")
    try:
        if purpose == "boost" and payment.get("ref_id"):
            db.set_product_boost(payment["ref_id"], int(cfg["mon_boost_days"]) or 7)
        elif purpose == "subscription":
            db.set_pro_until(payment["user_id"], int(cfg["mon_subscription_days"]) or 30)
    except Exception as e:
        logging.error(f"to'lov bajarish xatosi (payment {payment.get('id')}): {e}")


def _revoke_payment_benefit(payment):
    """_fulfill_payment teskarisi: berilgan imtiyozni qaytarib oladi (soxta chek/qaytarish).
    boost → mahsulot boostini o'chiradi; subscription → Pro'ni darhol bekor qiladi."""
    purpose = payment.get("purpose")
    try:
        if purpose == "boost" and payment.get("ref_id"):
            db.clear_product_boost(payment["ref_id"])
        elif purpose == "subscription":
            db.clear_pro(payment["user_id"])
    except Exception as e:
        logging.error(f"to'lov qaytarish xatosi (payment {payment.get('id')}): {e}")


_PAY_REVOKED_TXT = {
    "boost": {"uz": "⚠️ Reklama boost to'lovingiz bekor qilindi (chek tasdiqlanmadi).",
              "ru": "⚠️ Оплата продвижения отменена (чек не подтверждён)."},
    "subscription": {"uz": "⚠️ Pro obuna to'lovingiz bekor qilindi (chek tasdiqlanmadi).",
                     "ru": "⚠️ Оплата Pro-подписки отменена (чек не подтверждён)."},
}


_PAY_DONE_TXT = {
    "boost": {"uz": "🚀 To'lov qabul qilindi — mahsulotingiz endi tepada ko'rinadi!",
              "ru": "🚀 Оплата принята — ваш товар теперь показывается вверху!"},
    "subscription": {"uz": "⭐ To'lov qabul qilindi — Pro obuna faollashtirildi!",
                     "ru": "⭐ Оплата принята — Pro-подписка активирована!"},
}


async def _notify_payment_done(payment):
    """To'lov muvaffaqiyatli bo'lganda foydalanuvchiga botdan xabar (u Payme/Click
    ilovasida to'lasa, Mini App o'zi bilmaydi — shuning uchun push muhim)."""
    try:
        u = db.get_user_by_id(payment["user_id"])
        if not u or not dict(u).get("telegram_id"):
            return
        lang = get_user_lang(u) or DEFAULT_LANG
        msg = _PAY_DONE_TXT.get(payment.get("purpose"), {})
        txt = msg.get(lang) or msg.get("uz")
        if txt:
            await _tg_call("sendMessage", {"chat_id": dict(u)["telegram_id"], "text": txt})
            # To'lovchiга app ichida ham banner ko'rinsin (push allaqachon yuborildi)
            try:
                db.create_notification(payment["user_id"], "payment", txt, "", payment.get("id"))
            except Exception:
                pass
        # ⭐ Pro obuna — adminларни xabardor qilamiz (kim Pro oldi: bilinmay qolmasin)
        if payment.get("purpose") == "subscription":
            pname = dict(u).get("name") or ""
            await _notify_admins("⭐ Yangi Pro obuna",
                                 f"{pname} Pro obuna sotib oldi.", kind="pro", ref_id=payment.get("id"))
    except Exception as e:
        logging.warning(f"to'lov xabar yuborish xato (payment {payment.get('id')}): {e}")


def _mark_paid_and_fulfill(payment, provider, txn_id=None, meta=None):
    """Pending→paid o'tkazadi (faqat bir marta) va maqsadni bajaradi. paid bo'lganini qaytaradi."""
    fresh = db.get_payment(payment["id"])
    if not fresh:
        return False
    if fresh["state"] == "paid":
        return True  # idempotent — webhook qayta keldi (tez yo'l)
    # Atomik da'vo: faqat bu chaqiruv 'paid' qila olsagina fulfillment bajariladi.
    # get_payment o'qish bilan set orasidagi poyga (webhook retry / ko'p worker) yopiladi —
    # aks holda ikkala chaqiruv ham state=pending ko'rib, Pro/boost IKKI marta berilardi.
    if not db.mark_paid_atomic(payment["id"], provider=provider, provider_txn_id=txn_id, meta=meta):
        return True  # boshqa chaqiruv allaqachon paid qildi — idempotent
    _fulfill_payment(fresh)
    # Botdan xabar (async; webhook async kontekstda ishlaydi). Loop bo'lmasa — jim o'tadi.
    try:
        asyncio.get_running_loop().create_task(_notify_payment_done(fresh))
    except RuntimeError:
        pass
    return True


def _payment_checkout_urls(payment, pub):
    """Faol provayderlar uchun checkout URL'lari (so'm → Payme tiyin)."""
    urls = {}
    pid = payment["id"]
    amount = float(payment["amount"])
    if pub.get("pay_click") and _click_configured():
        urls["click"] = (f"{CLICK_CHECKOUT_URL}?service_id={CLICK_SERVICE_ID}"
                         f"&merchant_id={CLICK_MERCHANT_ID}&amount={amount:.0f}"
                         f"&transaction_param={pid}")
    if pub.get("pay_payme") and _payme_configured():
        tiyin = int(round(amount * 100))
        raw = f"m={PAYME_MERCHANT_ID};ac.{PAYME_ACCOUNT_FIELD}={pid};a={tiyin}"
        token = base64.b64encode(raw.encode()).decode()
        urls["payme"] = f"{PAYME_CHECKOUT_URL}/{token}"
    if pub.get("pay_paynet") and _paynet_configured() and PAYNET_CHECKOUT_URL:
        # Paynet'da redirect-checkout bo'lsa link beramiz; bo'lmasa mijoz Paynet ilovasidan
        # to'lov ID (payment_id) bo'yicha to'laydi — link shart emas (webhook baribir keladi).
        urls["paynet"] = (f"{PAYNET_CHECKOUT_URL}?merchant_id={PAYNET_MERCHANT_ID}"
                          f"&payment_id={pid}&amount={amount:.0f}")
    return urls


def _purchase_contact_line(user):
    """Admin xabarnomasi uchun: sotib olmoqchi bo'lganning ismi + kontakt (username/telefon).
    Manzil yozilmagan bo'lsa ham — telegram username yoki ro'yxatdagi telefon ko'rsatiladi."""
    nm = (user.get("name") or "").strip()
    if user.get("telegram_username"):
        ct = "@" + str(user["telegram_username"]).lstrip("@")
    else:
        ct = (user.get("phone_number") or "").strip()
    line = " · ".join([x for x in (nm, ct) if x])
    return line or "Yangi so'rov"


def _new_purchase_response(user, payment, pub):
    """Sotuvchiga to'lov boshlash javobi: checkout URL'lar + (kalit yo'q/admin bo'lsa) dev-confirm."""
    urls = _payment_checkout_urls(payment, pub)
    resp = {"ok": True, "payment_id": payment["id"], "amount": payment["amount"], "providers": urls}
    # Sinov yo'li (qo'lda tasdiqlash) FAQAT adminga — real pul ko'chmaydi. Oddiy sotuvchi
    # provayder ulanmagan bo'lsa to'lay olmaydi (pay_no_provider) — bu to'g'ri (teshik yo'q).
    is_admin = user.get("role") == "admin" or user.get("telegram_id") == ADMIN_ID
    if is_admin:
        resp["dev_confirm"] = True
    return resp


@app.post("/api/seller/boost/{product_id}")
async def api_seller_boost(product_id: int, authorization: str = Header(None)):
    """Mahsulotni boost qilish uchun to'lov boshlaydi. Bayroq yoqilmagan bo'lsa 403.
    #18 Pro — Pro do'kon oyiga `pro_free_boosts` martagacha BEPUL boost qiladi (to'lovsiz)."""
    user = dict(_buyer_from_auth(authorization))
    pub = monetization_public()
    if not pub.get("boost"):
        raise HTTPException(status_code=403, detail="boost_disabled")
    prod = _own_product_or_403(user, product_id)
    oid = _owner_id(user)
    # Pro bepul boost kvotasi — bor bo'lsa to'lovsiz darhol qo'llanadi.
    if _is_pro(oid):
        quota = int(pub.get("pro_free_boosts") or 0)
        used = db.get_feature_usage(oid, "boost_free", _month_key())
        if quota > 0 and used < quota:
            db.set_product_boost(prod["id"], int(pub.get("boost_days") or 7))
            db.incr_feature_usage(oid, "boost_free", _month_key())
            return {"ok": True, "free": True, "boosted": True,
                    "remaining": quota - used - 1}
    price = float(pub.get("boost_price") or 0)
    if price <= 0:
        raise HTTPException(status_code=400, detail="boost_price_unset")
    pid = db.create_payment(oid, "boost", price, ref_id=prod["id"])
    payment = db.get_payment(pid)
    # Adminларга: kimdir Boost sotib olmoqchi (kontakt bilan) — kutilayotgan to'lovni
    # tasdiqlash/bog'lanish uchun. To'lovlar ekranида Telegram/telefon tugmalari bor.
    await _notify_admins("🚀 Boost so'rovi", _purchase_contact_line(user),
                         kind="boost", ref_id=pid)
    await _notify_purchaser_pending(user, "boost", ref_id=pid)   # sotuvchiga tasdiq
    return _new_purchase_response(user, payment, pub)


@app.post("/api/seller/subscribe")
async def api_seller_subscribe(authorization: str = Header(None)):
    """Pro-obuna uchun to'lov boshlaydi. Bayroq yoqilmagan bo'lsa 403."""
    user = dict(_buyer_from_auth(authorization))
    pub = monetization_public()
    if not pub.get("subscription"):
        raise HTTPException(status_code=403, detail="subscription_disabled")
    price = float(pub.get("subscription_price") or 0)
    if price <= 0:
        raise HTTPException(status_code=400, detail="subscription_price_unset")
    pid = db.create_payment(_owner_id(user), "subscription", price)
    payment = db.get_payment(pid)
    await _notify_admins("⭐ Pro obuna so'rovi", _purchase_contact_line(user),
                         kind="pro", ref_id=pid)
    await _notify_purchaser_pending(user, "subscription", ref_id=pid)   # sotuvchiga tasdiq
    return _new_purchase_response(user, payment, pub)


@app.get("/api/seller/payments")
def api_seller_payments(authorization: str = Header(None)):
    """Sotuvchining to'lovlari tarixi (ega bo'yicha)."""
    user = dict(_buyer_from_auth(authorization))
    return {"payments": db.get_payments_by_user(_owner_id(user), limit=50)}


@app.get("/api/seller/commission")
def api_seller_commission(authorization: str = Header(None)):
    """Sotuvchining alohida komissiya ekrani: hozirgi foiz, qoldiq (owed), to'langan (paid)
    va qaysi buyurtmalardan ekani (egasi bo'yicha jamlanadi)."""
    user = dict(_buyer_from_auth(authorization))
    owner = _owner_id(user)
    mc = monetization_config()
    active = bool(mc["mon_enabled"] and mc["mon_commission_enabled"])
    return {
        "active": active,
        "percent": mc["mon_commission_percent"] if active else 0,
        "owed": db.get_commission_owed_by_seller(owner),
        "paid": db.get_commission_paid_by_seller(owner),
        "orders": db.get_commission_orders_by_seller(owner, limit=100),
    }


@app.get("/api/pay/status/{payment_id}")
def api_pay_status(payment_id: int, authorization: str = Header(None)):
    """To'lov holati — Mini App provayder linkini ochgandan keyin shu yerda poll qiladi
    (Payme/Click ilovasida to'lov tugasa, webhook 'paid' qiladi)."""
    user = dict(_buyer_from_auth(authorization))
    payment = db.get_payment(payment_id)
    if not payment:
        raise HTTPException(status_code=404, detail="not_found")
    if not (payment["user_id"] == _owner_id(user)
            or user.get("role") == "admin" or user.get("telegram_id") == ADMIN_ID):
        raise HTTPException(status_code=403, detail="not_allowed")
    return {"state": payment["state"], "purpose": payment["purpose"]}


@app.post("/api/pay/dev-confirm/{payment_id}")
async def api_pay_dev_confirm(payment_id: int, authorization: str = Header(None)):
    """SINOV: to'lovni qo'lda 'paid' qiladi (provider='manual'). FAQAT ADMIN — aks holda
    provayder ulanmagan paytda har kim bepul boost/obuna olishi mumkin edi (teshik).
    Real provayder pulini ko'chirmaydi; faqat oqimni sinash uchun.
    DIQQAT: `async def` bo'lishi SHART — `_mark_paid_and_fulfill` xabarni running loop'da
    `create_task` bilan yuboradi; sync endpoint threadpool'da ishlab loop'siz qolar, natijada
    sotuvchiga Pro/boost faollashgani haqida xabar JIM yutilardi."""
    user = dict(_buyer_from_auth(authorization))
    payment = db.get_payment(payment_id)
    if not payment:
        raise HTTPException(status_code=404, detail="not_found")
    is_admin = user.get("role") == "admin" or user.get("telegram_id") == ADMIN_ID
    if not is_admin:
        raise HTTPException(status_code=403, detail="not_allowed")
    if payment["state"] == "cancelled":
        raise HTTPException(status_code=409, detail="cancelled")
    # Endpoint async bo'lgani uchun _mark_paid_and_fulfill ichidagi create_task running
    # loop'ga tushadi va sotuvchiga xabar (Pro/boost faollashdi) ishonchli yetadi.
    _mark_paid_and_fulfill(payment, "manual")
    return {"ok": True, "payment": db.get_payment(payment_id)}


@app.post("/api/pay/dev-cancel/{payment_id}")
def api_pay_dev_cancel(payment_id: int, authorization: str = Header(None)):
    """Admin kutilayotgan to'lovni QO'LDA bekor qiladi (rad etish). FAQAT ADMIN.
    To'langan to'lovni bekor qilib bo'lmaydi (409) — boost/obuna allaqachon berilgan."""
    user = dict(_buyer_from_auth(authorization))
    payment = db.get_payment(payment_id)
    if not payment:
        raise HTTPException(status_code=404, detail="not_found")
    is_admin = user.get("role") == "admin" or user.get("telegram_id") == ADMIN_ID
    if not is_admin:
        raise HTTPException(status_code=403, detail="not_allowed")
    if payment["state"] == "paid":
        raise HTTPException(status_code=409, detail="already_paid")
    if payment["state"] == "cancelled":
        return {"ok": True, "payment": payment}  # idempotent
    db.set_payment_state(payment_id, "cancelled")
    return {"ok": True, "payment": db.get_payment(payment_id)}


@app.post("/api/pay/dev-revoke/{payment_id}")
async def api_pay_dev_revoke(payment_id: int, authorization: str = Header(None)):
    """TASDIQLANGAN (paid) to'lovni QAYTARIB OLADI — soxta chek bo'lsa. FAQAT ADMIN.
    Berilgan imtiyoz (boost/Pro) olib tashlanadi + to'lov 'cancelled' bo'ladi → platforma
    daromadidan ham tushadi (get_paid_payments_summary faqat state='paid' sanaydi)."""
    user = dict(_buyer_from_auth(authorization))
    payment = db.get_payment(payment_id)
    if not payment:
        raise HTTPException(status_code=404, detail="not_found")
    is_admin = user.get("role") == "admin" or user.get("telegram_id") == ADMIN_ID
    if not is_admin:
        raise HTTPException(status_code=403, detail="not_allowed")
    if payment["state"] != "paid":
        raise HTTPException(status_code=409, detail="not_paid")  # faqat to'langanni qaytarish mumkin
    _revoke_payment_benefit(payment)               # boost/Pro'ni olib tashlash
    db.set_payment_state(payment_id, "cancelled")  # daromaddan ham chiqadi
    try:                                           # sotuvchini xabardor qilish
        u = db.get_user_by_id(payment["user_id"])
        if u and dict(u).get("telegram_id"):
            lang = get_user_lang(u) or DEFAULT_LANG
            msg = _PAY_REVOKED_TXT.get(payment.get("purpose"), {})
            txt = msg.get(lang) or msg.get("uz")
            if txt:
                await _tg_call("sendMessage", {"chat_id": dict(u)["telegram_id"], "text": txt})
    except Exception as e:
        logging.warning(f"qaytarish xabari xato (payment {payment_id}): {e}")
    return {"ok": True, "payment": db.get_payment(payment_id)}


# ---- Click webhook (prepare + complete) ----
def _click_sign(params, keys):
    """Click imzosi: md5 of concat(params[k] for k in keys) + SECRET (tartibga ko'ra)."""
    raw = "".join(str(params.get(k, "")) for k in keys)
    return hashlib.md5((raw + CLICK_SECRET_KEY).encode()).hexdigest()


@app.post("/api/pay/click")
async def api_pay_click(request: Request):
    """Click Merchant API: action=0 (prepare), action=1 (complete). Imzo (sign_string) tekshiriladi."""
    if not _click_configured():
        return {"error": -9, "error_note": "not_configured"}
    form = dict((await request.form()))
    action = str(form.get("action", ""))
    payment_id = form.get("merchant_trans_id") or form.get("transaction_param")
    click_trans_id = form.get("click_trans_id")
    err = {"click_trans_id": click_trans_id, "merchant_trans_id": payment_id}

    # Imzo tekshiruvi
    if action == "0":
        sign_keys = ["click_trans_id", "service_id", "merchant_trans_id", "amount", "action", "sign_time"]
    else:
        sign_keys = ["click_trans_id", "service_id", "merchant_trans_id",
                     "merchant_prepare_id", "amount", "action", "sign_time"]
    # Doimiy-vaqtli solishtiruv (timing-attack himoyasi)
    if not hmac.compare_digest(str(form.get("sign_string") or ""), _click_sign(form, sign_keys)):
        return {**err, "error": -1, "error_note": "SIGN CHECK FAILED"}

    payment = db.get_payment(int(payment_id)) if payment_id and str(payment_id).isdigit() else None
    if not payment:
        return {**err, "error": -5, "error_note": "Order not found"}
    if abs(float(form.get("amount", 0)) - float(payment["amount"])) > 0.5:
        return {**err, "error": -2, "error_note": "Incorrect amount"}
    if payment["state"] == "cancelled":
        return {**err, "error": -9, "error_note": "Transaction cancelled"}

    if action == "0":  # prepare
        return {**err, "merchant_prepare_id": payment["id"], "error": 0, "error_note": "Success"}
    if action == "1":  # complete
        _mark_paid_and_fulfill(payment, "click", txn_id=click_trans_id)
        return {**err, "merchant_confirm_id": payment["id"], "error": 0, "error_note": "Success"}
    return {**err, "error": -3, "error_note": "Action not found"}


# ---- Payme webhook (JSON-RPC) ----
def _payme_auth_ok(authorization):
    """Authorization: Basic base64('Paycom:<PAYME_KEY>')."""
    if not authorization or not authorization.startswith("Basic "):
        return False
    try:
        decoded = base64.b64decode(authorization[6:]).decode()
    except Exception:
        return False
    # Doimiy-vaqtli solishtiruv — timing-attack orqali kalitni bit-bit topishning oldini oladi
    return hmac.compare_digest(decoded.split(":", 1)[-1], PAYME_KEY or "")


def _payme_err(rid, code, msg):
    return {"jsonrpc": "2.0", "id": rid, "error": {"code": code, "message": {"ru": msg, "uz": msg, "en": msg}}}


def _ts_ms(value):
    try:
        dt = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return int(dt.timestamp() * 1000)
    except Exception:
        return int(_time.time() * 1000)


@app.post("/api/pay/payme")
async def api_pay_payme(request: Request, authorization: str = Header(None)):
    """Payme Merchant JSON-RPC: CheckPerformTransaction / CreateTransaction /
    PerformTransaction / CancelTransaction / CheckTransaction."""
    body = await request.json()
    rid = body.get("id") if isinstance(body, dict) else None
    if not _payme_configured() or not _payme_auth_ok(authorization):
        return _payme_err(rid, -32504, "Insufficient privileges")
    method = body.get("method")
    params = body.get("params") or {}

    def _payment_from_account():
        acc = params.get("account") or {}
        pid = acc.get("payment_id")
        if not pid or not str(pid).isdigit():
            return None
        return db.get_payment(int(pid))

    if method == "CheckPerformTransaction":
        p = _payment_from_account()
        if not p:
            return _payme_err(rid, -31050, "Payment not found")
        if int(round(float(p["amount"]) * 100)) != int(params.get("amount", 0)):
            return _payme_err(rid, -31001, "Incorrect amount")
        if p["state"] != "pending":
            return _payme_err(rid, -31008, "Transaction not allowed")
        return {"jsonrpc": "2.0", "id": rid, "result": {"allow": True}}

    if method == "CreateTransaction":
        txn = params.get("id")
        existing = db.get_payment_by_txn("payme", txn)
        if existing:
            meta = _json.loads(existing.get("provider_meta") or "{}")
            return {"jsonrpc": "2.0", "id": rid, "result": {
                "create_time": meta.get("create_time", _ts_ms(existing["created_at"])),
                "transaction": str(existing["id"]),
                "state": 1 if existing["state"] == "pending" else 2}}
        p = _payment_from_account()
        if not p:
            return _payme_err(rid, -31050, "Payment not found")
        if int(round(float(p["amount"]) * 100)) != int(params.get("amount", 0)):
            return _payme_err(rid, -31001, "Incorrect amount")
        if p["state"] != "pending":
            return _payme_err(rid, -31008, "Transaction not allowed")
        ct = int(params.get("time") or _time.time() * 1000)
        db.set_payment_state(p["id"], "pending", provider="payme", provider_txn_id=txn,
                             meta=_json.dumps({"create_time": ct}))
        return {"jsonrpc": "2.0", "id": rid, "result": {
            "create_time": ct, "transaction": str(p["id"]), "state": 1}}

    if method == "PerformTransaction":
        p = db.get_payment_by_txn("payme", params.get("id"))
        if not p:
            return _payme_err(rid, -31003, "Transaction not found")
        meta = _json.loads(p.get("provider_meta") or "{}")
        if p["state"] == "paid":
            return {"jsonrpc": "2.0", "id": rid, "result": {
                "transaction": str(p["id"]), "perform_time": meta.get("perform_time", _ts_ms(p.get("paid_at"))),
                "state": 2}}
        if p["state"] != "pending":
            return _payme_err(rid, -31008, "Transaction not allowed")
        pt = int(_time.time() * 1000)
        meta["perform_time"] = pt
        _mark_paid_and_fulfill(p, "payme", txn_id=params.get("id"), meta=_json.dumps(meta))
        return {"jsonrpc": "2.0", "id": rid, "result": {
            "transaction": str(p["id"]), "perform_time": pt, "state": 2}}

    if method == "CancelTransaction":
        p = db.get_payment_by_txn("payme", params.get("id"))
        if not p:
            return _payme_err(rid, -31003, "Transaction not found")
        meta = _json.loads(p.get("provider_meta") or "{}")
        ct = meta.get("cancel_time") or int(_time.time() * 1000)
        meta["cancel_time"] = ct
        meta["cancel_reason"] = params.get("reason")
        # paid bo'lgandan keyin bekor = state -2, aks holda -1
        post = (p["state"] == "paid")
        if p["state"] != "cancelled":
            db.set_payment_state(p["id"], "cancelled", meta=_json.dumps(meta))
        return {"jsonrpc": "2.0", "id": rid, "result": {
            "transaction": str(p["id"]), "cancel_time": ct, "state": -2 if post else -1}}

    if method == "CheckTransaction":
        p = db.get_payment_by_txn("payme", params.get("id"))
        if not p:
            return _payme_err(rid, -31003, "Transaction not found")
        meta = _json.loads(p.get("provider_meta") or "{}")
        state = {"pending": 1, "paid": 2, "cancelled": -1}[p["state"]]
        return {"jsonrpc": "2.0", "id": rid, "result": {
            "create_time": meta.get("create_time", _ts_ms(p["created_at"])),
            "perform_time": meta.get("perform_time", 0),
            "cancel_time": meta.get("cancel_time", 0),
            "transaction": str(p["id"]), "state": state, "reason": meta.get("cancel_reason")}}

    return _payme_err(rid, -32601, "Method not found")


# ---- Paynet webhook (JSON-RPC, Payme'ga o'xshash) ----
# DIQQAT: Paynet merchant protokoli kabinetdan olingan hujjat bilan TASDIQLANISHI shart.
# Quyidagi metod nomlari (CheckPerformTransaction/CreateTransaction/PerformTransaction/
# CancelTransaction/CheckTransaction) Payme bilan bir xil deb taxmin qilingan — Paynet
# boshqacha atasa, shu yerda moslab o'zgartiriladi (bayroq ulanmaguncha jonli emas).
def _paynet_auth_ok(authorization):
    """Authorization: Basic base64('LOGIN:PASSWORD')."""
    if not authorization or not authorization.startswith("Basic "):
        return False
    try:
        decoded = base64.b64decode(authorization[6:]).decode()
    except Exception:
        return False
    return decoded == f"{PAYNET_LOGIN}:{PAYNET_PASSWORD}"


@app.post("/api/pay/paynet")
async def api_pay_paynet(request: Request, authorization: str = Header(None)):
    """Paynet merchant JSON-RPC. Payme handleri bilan bir xil holat-mashinasi (payments
    daftari), faqat auth LOGIN:PASSWORD. Provider='paynet' bilan yoziladi."""
    body = await request.json()
    rid = body.get("id") if isinstance(body, dict) else None
    if not _paynet_configured() or not _paynet_auth_ok(authorization):
        return _payme_err(rid, -32504, "Insufficient privileges")
    method = body.get("method")
    params = body.get("params") or {}

    def _pmt():
        acc = params.get("account") or {}
        pid = acc.get("payment_id")
        return db.get_payment(int(pid)) if pid and str(pid).isdigit() else None

    if method == "CheckPerformTransaction":
        p = _pmt()
        if not p:
            return _payme_err(rid, -31050, "Payment not found")
        if int(round(float(p["amount"]) * 100)) != int(params.get("amount", 0)):
            return _payme_err(rid, -31001, "Incorrect amount")
        if p["state"] != "pending":
            return _payme_err(rid, -31008, "Transaction not allowed")
        return {"jsonrpc": "2.0", "id": rid, "result": {"allow": True}}

    if method == "CreateTransaction":
        txn = params.get("id")
        existing = db.get_payment_by_txn("paynet", txn)
        if existing:
            meta = _json.loads(existing.get("provider_meta") or "{}")
            return {"jsonrpc": "2.0", "id": rid, "result": {
                "create_time": meta.get("create_time", _ts_ms(existing["created_at"])),
                "transaction": str(existing["id"]),
                "state": 1 if existing["state"] == "pending" else 2}}
        p = _pmt()
        if not p:
            return _payme_err(rid, -31050, "Payment not found")
        if int(round(float(p["amount"]) * 100)) != int(params.get("amount", 0)):
            return _payme_err(rid, -31001, "Incorrect amount")
        if p["state"] != "pending":
            return _payme_err(rid, -31008, "Transaction not allowed")
        ct = int(params.get("time") or _time.time() * 1000)
        db.set_payment_state(p["id"], "pending", provider="paynet", provider_txn_id=txn,
                             meta=_json.dumps({"create_time": ct}))
        return {"jsonrpc": "2.0", "id": rid, "result": {
            "create_time": ct, "transaction": str(p["id"]), "state": 1}}

    if method == "PerformTransaction":
        p = db.get_payment_by_txn("paynet", params.get("id"))
        if not p:
            return _payme_err(rid, -31003, "Transaction not found")
        meta = _json.loads(p.get("provider_meta") or "{}")
        if p["state"] == "paid":
            return {"jsonrpc": "2.0", "id": rid, "result": {
                "transaction": str(p["id"]), "perform_time": meta.get("perform_time", _ts_ms(p.get("paid_at"))),
                "state": 2}}
        if p["state"] != "pending":
            return _payme_err(rid, -31008, "Transaction not allowed")
        pt = int(_time.time() * 1000)
        meta["perform_time"] = pt
        _mark_paid_and_fulfill(p, "paynet", txn_id=params.get("id"), meta=_json.dumps(meta))
        return {"jsonrpc": "2.0", "id": rid, "result": {
            "transaction": str(p["id"]), "perform_time": pt, "state": 2}}

    if method == "CancelTransaction":
        p = db.get_payment_by_txn("paynet", params.get("id"))
        if not p:
            return _payme_err(rid, -31003, "Transaction not found")
        meta = _json.loads(p.get("provider_meta") or "{}")
        ct = meta.get("cancel_time") or int(_time.time() * 1000)
        meta["cancel_time"] = ct
        meta["cancel_reason"] = params.get("reason")
        post = (p["state"] == "paid")
        if p["state"] != "cancelled":
            db.set_payment_state(p["id"], "cancelled", meta=_json.dumps(meta))
        return {"jsonrpc": "2.0", "id": rid, "result": {
            "transaction": str(p["id"]), "cancel_time": ct, "state": -2 if post else -1}}

    if method == "CheckTransaction":
        p = db.get_payment_by_txn("paynet", params.get("id"))
        if not p:
            return _payme_err(rid, -31003, "Transaction not found")
        meta = _json.loads(p.get("provider_meta") or "{}")
        state = {"pending": 1, "paid": 2, "cancelled": -1}[p["state"]]
        return {"jsonrpc": "2.0", "id": rid, "result": {
            "create_time": meta.get("create_time", _ts_ms(p["created_at"])),
            "perform_time": meta.get("perform_time", 0),
            "cancel_time": meta.get("cancel_time", 0),
            "transaction": str(p["id"]), "state": state, "reason": meta.get("cancel_reason")}}

    return _payme_err(rid, -32601, "Method not found")


@app.get("/api/admin/stats")
def api_admin_stats(authorization: str = Header(None)):
    _admin_from_auth(authorization)
    return dict(db.get_admin_stats_summary() or {})


@app.get("/api/admin/analytics")
def api_admin_analytics(authorization: str = Header(None)):
    """A2 — konversiya voronkasi (bot admin_analytics pariteti): get_analytics_funnel."""
    _admin_from_auth(authorization)
    return dict(db.get_analytics_funnel() or {})


@app.get("/api/admin/sentiment")
async def api_admin_sentiment(authorization: str = Header(None)):
    """#6 — mijoz sharhlaridan AI sentiment hisoboti: nimadan norozi / nimani maqtaydi."""
    admin = _admin_from_auth(authorization)
    reviews = db.get_recent_reviews_with_comments(200) or []
    total = len(reviews)
    if not total:
        return {"available": False, "reason": "no_reviews", "total": 0}
    if not ai_assistant.is_enabled():
        return {"available": False, "reason": "ai_disabled", "total": total}
    lang = get_user_lang(admin) or DEFAULT_LANG
    report = await ai_assistant.analyze_sentiment(reviews, lang=lang)
    if not report:
        return {"available": False, "reason": "ai_error", "total": total}
    # Bahosi past sharhlar soni (kontekst uchun)
    low = 0
    for r in reviews:
        rt = r.get("rating") or r.get("product_rating")
        try:
            if rt is not None and float(rt) <= 2:
                low += 1
        except (TypeError, ValueError):
            pass
    report.update({"available": True, "total": total, "low_count": low})
    return report


@app.get("/api/admin/financial")
def api_admin_financial(authorization: str = Header(None)):
    """A3 — moliyaviy hisobot (bot admin_revenue pariteti): yetkazilgan buyurtmalardan
    sotuvchi/to'lov/yetkazish kesimida."""
    _admin_from_auth(authorization)
    from collections import defaultdict
    from datetime import datetime, timedelta, timezone
    orders = [dict(o) for o in (db.get_all_orders() or [])]
    delivered = [o for o in orders if o.get("status") == "delivered"]
    month_ago = (datetime.now(timezone.utc) - timedelta(days=30)).strftime("%Y-%m-%d")
    sellers = defaultdict(lambda: {"count": 0, "revenue": 0.0, "month_count": 0, "month_revenue": 0.0})
    pay = defaultdict(int)
    dlv = defaultdict(int)
    for o in delivered:
        sname = o.get("seller_name") or o.get("shop_name") or "—"
        price = float(o.get("total_price") or 0)
        s = sellers[sname]
        s["count"] += 1
        s["revenue"] += price
        if str(o.get("created_at") or "")[:10] >= month_ago:
            s["month_count"] += 1
            s["month_revenue"] += price
        pay[o.get("payment_method") or "cash"] += 1
        dlv[o.get("delivery_type") or "delivery"] += 1
    total = sum(s["revenue"] for s in sellers.values())
    month_total = sum(s["month_revenue"] for s in sellers.values())
    top = sorted(sellers.items(), key=lambda x: x[1]["revenue"], reverse=True)[:10]
    # #18 Platforma daromadi: komissiya (buyurtmalardan) + boost/obuna to'lovlari
    commission_total = sum(float(o.get("commission_amount") or 0) for o in delivered)
    pay_summary = {p["purpose"]: {"count": p["cnt"], "total": round(p["total"])}
                   for p in (db.get_paid_payments_summary() or [])}
    platform_revenue = round(commission_total + sum(v["total"] for v in pay_summary.values()))
    return {
        "delivered_count": len(delivered),
        "total_revenue": round(total),
        "month_revenue": round(month_total),
        "month_count": sum(s["month_count"] for s in sellers.values()),
        "top_sellers": [{"name": n, "revenue": round(s["revenue"]), "count": s["count"]} for n, s in top],
        "by_payment": dict(pay),
        "by_delivery": dict(dlv),
        # platforma o'z daromadi (monetizatsiya)
        "commission_total": round(commission_total),
        "payments_by_purpose": pay_summary,
        "platform_revenue": platform_revenue,
    }


@app.get("/api/admin/seller-requests")
def api_admin_seller_requests(authorization: str = Header(None)):
    _admin_from_auth(authorization)
    return _rows(db.get_pending_seller_requests())


class SellerReqDecision(BaseModel):
    approve: bool = True


@app.post("/api/admin/seller-request/{user_id}")
async def api_admin_seller_decide(user_id: int, body: SellerReqDecision,
                                  authorization: str = Header(None)):
    """Sotuvchi arizasini tasdiqlash/rad etish — bot approve_seller/reject_seller bilan bir xil."""
    admin = _admin_from_auth(authorization)
    _rate_limit("admin_seller", admin["id"], 30, 60)
    target = db.get_user_by_id(user_id)
    if not target:
        raise HTTPException(status_code=404, detail="user_not_found")
    target = dict(target)
    if body.approve:
        db.update_user(user_id, is_approved=1, role="seller")
        # Tasdiqlangan sotuvchida do'kon bo'lishini KAFOLATLAYMIZ (create_shop idempotent —
        # mavjud bo'lsa qaytaradi, dublikat yaratmaydi). Xato YUTILMAYDI: aks holda sotuvchi
        # do'konsiz "tasdiqlangan" bo'lib qolardi, admin esa buni bilmay ok ko'rardi. 500
        # qaytadi, so'rov 'approved' belgilanmaydi → admin qayta uradi (idempotent, xavfsiz).
        try:
            if not db.get_staff_by_user(user_id):
                db.create_shop(
                    user_id,
                    name=target.get("shop_name"), address=target.get("shop_address"),
                    landmark=target.get("shop_landmark"), lat=target.get("shop_lat"),
                    lon=target.get("shop_lon"), working_days=target.get("working_days"),
                    working_hours=target.get("working_hours"), region_id=target.get("region_id"),
                    card_number=target.get("card_number"), card_owner=target.get("card_owner"),
                    card_type=target.get("card_type"),
                )
        except Exception as e:
            logging.error(f"admin approve: do'kon yaratilmadi (uid {user_id}): {e}")
            raise HTTPException(status_code=500, detail="shop_create_failed")
    else:
        db.update_user(user_id, is_approved=0, role="buyer")
    req = db.get_seller_request_by_user(user_id)
    if req:
        db.update_seller_request(dict(req)["id"], "approved" if body.approve else "rejected")
    # Foydalanuvchiga uning tilida xabar
    try:
        if target.get("telegram_id"):
            tlang = get_user_lang(target)
            key = "approve_seller_notify" if body.approve else "reject_seller_notify"
            await _tg_call("sendMessage", {"chat_id": target["telegram_id"],
                                           "text": t(tlang, key), "parse_mode": "HTML"})
    except Exception as e:
        logging.warning(f"admin seller decide notify xato (uid {user_id}): {e}")
    return {"ok": True, "approved": body.approve}


@app.get("/api/admin/users")
def api_admin_users(authorization: str = Header(None), q: str = Query(None),
                    offset: int = Query(0), role: str = Query(None),
                    inactive: bool = Query(False)):
    _admin_from_auth(authorization)
    q = (q or "").strip()
    if q:
        return {"total": None, "offset": 0, "users": _rows(db.search_users(q, limit=30))}
    if inactive:
        # 💤 Faqat nofaol / bir martalik (30+ kun faollik yo'q yoki hech qachon)
        total, rows = db.get_users_paginated(limit=ADMIN_PAGE, offset=max(0, offset), inactive_only=True)
        return {"total": total, "offset": offset, "users": _rows(rows), "inactive": True}
    role = role if role in ("buyer", "seller", "admin") else None
    if role:
        allr = db.get_all_users(role=role)
        total = len(allr)
        page = allr[max(0, offset):max(0, offset) + ADMIN_PAGE]
        return {"total": total, "offset": offset, "users": _rows(page), "role": role}
    total, rows = db.get_users_paginated(limit=ADMIN_PAGE, offset=max(0, offset))
    return {"total": total, "offset": offset, "users": _rows(rows)}


ADMIN_PAGE = 15


class UserBlockIn(BaseModel):
    block: bool


@app.post("/api/admin/user/{user_id}/block")
def api_admin_block_user(user_id: int, body: UserBlockIn, authorization: str = Header(None)):
    admin = _admin_from_auth(authorization)
    target = db.get_user_by_id(user_id)
    if not target:
        raise HTTPException(status_code=404, detail="user_not_found")
    if dict(target).get("role") == "admin":
        raise HTTPException(status_code=400, detail="cant_block_admin")
    db.update_user(user_id, is_blocked=1 if body.block else 0)
    return {"ok": True, "is_blocked": 1 if body.block else 0}


@app.delete("/api/admin/user/{user_id}")
def api_admin_delete_user(user_id: int, authorization: str = Header(None)):
    """Foydalanuvchini VA uning butun ma'lumotini TIZIMDAN BUTUNLAY o'chiradi
    (test akkauntini '0 ga qaytarish'). Keyin xuddi yangidek qayta ro'yxatdan o'tish
    mumkin. Har qanday admin (jumladan ZAHIRA admin) istalgan foydalanuvchini —
    asosiy adminni (ADMIN_ID) ham — o'chira oladi: zahira admin 100% asosiy admin
    huquqiga ega. Bu xavfsiz, chunki asosiy adminning huquqi DB qatoridan emas,
    ADMIN_ID env'dan keladi — o'chirilsa ham, o'sha Telegram bilan qayta kirsa
    admin avtomatik tiklanadi (lockout bo'lmaydi)."""
    _admin_from_auth(authorization)
    target = db.get_user_by_id(user_id)
    if not target:
        raise HTTPException(status_code=404, detail="user_not_found")
    result = db.delete_user_completely(user_id)
    if not result.get("ok"):
        raise HTTPException(status_code=500, detail="delete_failed")
    return result


# ============================================================
# #21 ZAHIRA ADMIN — asosiy admin yo'qolsa ham nazorat saqlanadi.
# role='admin' qo'yamiz — bot ham, webapp ham buni avtomatik tan oladi.
# Asosiy admin (ADMIN_ID env) himoyalangan: uni olib bo'lmaydi (o'zini qulflab
# qo'ymaslik uchun). Har qanday admin zahira admin qo'sha/olib tashlay oladi.
# ============================================================
def _is_primary_admin(user):
    return bool(ADMIN_ID) and dict(user).get("telegram_id") == ADMIN_ID


@app.get("/api/admin/admins")
def api_admin_admins(authorization: str = Header(None)):
    me = _admin_from_auth(authorization)
    admins = []
    for u in db.get_all_users(role="admin"):
        u = dict(u)
        admins.append({
            "id": u.get("id"), "name": u.get("name"),
            "phone": u.get("phone_number"), "username": u.get("telegram_username"),
            "is_primary": _is_primary_admin(u),
            "is_self": u.get("id") == me.get("id"),
        })
    # Asosiy admin avval (himoyalangan), keyin qolganlari
    admins.sort(key=lambda a: (not a["is_primary"], a["id"] or 0))
    return {"admins": admins}


@app.post("/api/admin/user/{user_id}/grant-admin")
def api_admin_grant(user_id: int, authorization: str = Header(None)):
    _admin_from_auth(authorization)
    target = db.get_user_by_id(user_id)
    if not target:
        raise HTTPException(status_code=404, detail="user_not_found")
    if dict(target).get("role") == "admin":
        raise HTTPException(status_code=409, detail="already_admin")
    db.update_user(user_id, role="admin", is_blocked=0)
    return {"ok": True}


@app.post("/api/admin/user/{user_id}/revoke-admin")
def api_admin_revoke(user_id: int, authorization: str = Header(None)):
    me = _admin_from_auth(authorization)
    target = db.get_user_by_id(user_id)
    if not target:
        raise HTTPException(status_code=404, detail="user_not_found")
    if dict(target).get("role") != "admin":
        raise HTTPException(status_code=400, detail="not_admin")
    if _is_primary_admin(target):
        raise HTTPException(status_code=403, detail="cant_revoke_primary")
    if user_id == me.get("id"):
        raise HTTPException(status_code=400, detail="cant_revoke_self")
    # Do'kon egasi bo'lsa — sotuvchiga, aks holda xaridorga qaytaramiz
    new_role = "seller" if (dict(target).get("shop_name") or db.get_shop_by_owner(user_id)) else "buyer"
    db.update_user(user_id, role=new_role)
    return {"ok": True, "role": new_role}


# ---- FOYDALANUVCHI TO'LIQ MA'LUMOT + AI askfill (bot parite) ----
def _profile_missing_fields(user, lang):
    """Profildagi to'ldirilmagan muhim maydonlar (bot _profile_missing_fields bilan bir xil)."""
    ru = (lang == "ru")
    def empty(v):
        return v is None or (isinstance(v, str) and not v.strip())
    m = []
    if empty(user.get("name")): m.append("Имя" if ru else "Ism")
    if empty(user.get("telegram_username")): m.append("Username (@...)")
    if empty(user.get("phone_number")): m.append("Номер телефона" if ru else "Telefon raqami")
    if user.get("region_id") is None: m.append("Регион (область/район)" if ru else "Hudud (viloyat/tuman)")
    is_seller = bool(user.get("shop_name")) or user.get("role") == "seller"
    if is_seller:
        if empty(user.get("shop_name")): m.append("Название магазина" if ru else "Do'kon nomi")
        if empty(user.get("shop_address")): m.append("Адрес магазина" if ru else "Do'kon manzili")
        if empty(user.get("shop_landmark")): m.append("Ориентир" if ru else "Mo'ljal (orientir)")
        if empty(user.get("working_hours")): m.append("Часы работы" if ru else "Ish vaqti")
        if empty(user.get("working_days")): m.append("Рабочие дни" if ru else "Ish kunlari")
        if empty(user.get("card_number")): m.append("Карта для оплаты" if ru else "To'lov kartasi")
    return m


@app.get("/api/admin/user/{user_id}")
def api_admin_user_detail(user_id: int, authorization: str = Header(None)):
    """Foydalanuvchining TO'LIQ ma'lumoti (bot admin_user_details parite)."""
    _admin_from_auth(authorization)
    u = db.get_user_by_id(user_id)
    if not u:
        raise HTTPException(status_code=404, detail="user_not_found")
    u = dict(u)
    lang = u.get("language") or DEFAULT_LANG
    out = dict(u)
    try:
        out["region_label"] = db.get_region_label(u.get("region_id"))
    except Exception:
        out["region_label"] = None
    out["missing"] = _profile_missing_fields(u, lang)
    out["can_askfill"] = bool(out["missing"]) and ai_assistant.is_enabled()
    try:
        out["buyer_orders_count"] = len(db.get_orders_by_buyer(user_id) or [])
    except Exception:
        out["buyer_orders_count"] = 0
    is_seller = bool(u.get("shop_name")) or u.get("role") == "seller"
    if is_seller:
        try:
            out["seller_stats"] = dict(db.get_seller_stats(user_id) or {})
        except Exception:
            out["seller_stats"] = {}
        try:
            out["channels"] = _rows(db.get_seller_channels(user_id))
        except Exception:
            out["channels"] = []
    return out


class AdminContactsIn(BaseModel):
    shop_phone: Optional[str] = None
    shop_instagram: Optional[str] = None
    shop_telegram: Optional[str] = None
    shop_website: Optional[str] = None


@app.patch("/api/admin/user/{user_id}/contacts")
def api_admin_set_contacts(user_id: int, p: AdminContactsIn, authorization: str = Header(None)):
    """Super admin sotuvchi "Aloqa bloki"ni panel'dan tahrirlaydi (reklamaga chiqadi).
    Bo'sh satr ('') → maydonni tozalaydi; None → tegilmaydi. Multivendor: admin
    ISTALGAN do'konning aloqa kanallarini shu yerdan boshqaradi."""
    _admin_from_auth(authorization)
    target = db.get_user_by_id(user_id)
    if not target:
        raise HTTPException(status_code=404, detail="user_not_found")
    if p.shop_phone is not None and p.shop_phone.strip():
        digits = p.shop_phone.strip().lstrip("+").replace(" ", "").replace("-", "")
        if not digits.isdigit() or not (7 <= len(digits) <= 15):
            raise HTTPException(status_code=400, detail="bad_phone")
    for attr in ("shop_instagram", "shop_telegram", "shop_website"):
        v = getattr(p, attr)
        if v is not None and len(v.strip()) > 200:
            raise HTTPException(status_code=400, detail="bad_link")
    db.update_shop_contacts(
        user_id, phone=p.shop_phone, instagram=p.shop_instagram,
        telegram=p.shop_telegram, website=p.shop_website,
    )
    return {"ok": True}


@app.get("/api/admin/user/{user_id}/fill-preview")
async def api_admin_fill_preview(user_id: int, authorization: str = Header(None)):
    """AI yetishmagan maydonlar bo'yicha foydalanuvchiga yuboriladigan xabarni TAKLIF qiladi."""
    _admin_from_auth(authorization)
    u = db.get_user_by_id(user_id)
    if not u:
        raise HTTPException(status_code=404, detail="user_not_found")
    u = dict(u)
    lang = u.get("language") or DEFAULT_LANG
    missing = _profile_missing_fields(u, lang)
    if not missing:
        raise HTTPException(status_code=409, detail="nothing_missing")
    if not ai_assistant.is_enabled():
        raise HTTPException(status_code=503, detail="ai_disabled")
    is_seller = bool(u.get("shop_name")) or u.get("role") == "seller"
    msg = await ai_assistant.generate_profile_completion_message(
        name=u.get("name") or "", missing_fields=missing, is_seller=is_seller, lang=lang)
    if not msg:
        raise HTTPException(status_code=502, detail="ai_error")
    return {"missing": missing, "message": msg}


class FillSendIn(BaseModel):
    message: str


@app.post("/api/admin/user/{user_id}/sendfill")
async def api_admin_sendfill(user_id: int, body: FillSendIn, authorization: str = Header(None)):
    """AI taklif qilgan (admin tasdiqlagan) xabarni foydalanuvchiga yuboradi."""
    _admin_from_auth(authorization)
    u = db.get_user_by_id(user_id)
    if not u:
        raise HTTPException(status_code=404, detail="user_not_found")
    u = dict(u)
    txt = (body.message or "").strip()
    if not txt:
        raise HTTPException(status_code=400, detail="empty")
    if not u.get("telegram_id"):
        raise HTTPException(status_code=400, detail="no_telegram")
    await _tg_call("sendMessage", {"chat_id": u["telegram_id"], "text": txt})
    return {"ok": True}


# ---- EXCEL EKSPORT (bot admin_export_excel parite) — fayl Telegram'ga yuboriladi ----
async def _tg_send_document(chat_id, filename, content, caption=""):
    if not BOT_TOKEN:
        return None
    try:
        async with httpx.AsyncClient(timeout=40) as client:
            files = {"document": (filename, content,
                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            data = {"chat_id": str(chat_id), "caption": caption[:1000]}
            r = await client.post(
                f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument", data=data, files=files)
            return r.json()
    except Exception as e:
        logging.warning(f"sendDocument xato: {e}")
        return None


def _build_excel(kind, lang):
    """users|products|orders|seller_orders -> Pro-Excel (bytes, filename, rows)."""
    import datetime as _dt
    ru = (lang == "ru")
    if kind == "users":
        title = "Отчёт: Пользователи" if ru else "Foydalanuvchilar hisoboti"
        headers = (["ID", "Telegram ID", "Имя", "Телефон", "Роль", "Магазин", "Регион",
                    "Заказы", "Дата"] if ru else
                   ["ID", "Telegram ID", "Ism", "Telefon", "Rol", "Do'kon", "Hudud",
                    "Buyurtmalar", "Sana"])
        data = []
        for u in db.get_all_users():
            u = dict(u)
            try:
                oc = len(db.get_orders_by_buyer(u["id"]) or [])
            except Exception:
                oc = 0
            data.append([u.get("id"), u.get("telegram_id"), u.get("name") or "",
                         u.get("phone_number") or "", _xl_loc("role", u.get("role"), ru),
                         u.get("shop_name") or "", db.get_region_label(u.get("region_id")) or "",
                         oc, str(u.get("created_at") or "")[:10]])
        content, n = _xlsx_report(title, headers, data, money_cols=(), lang=lang)
        fn = "users"
    elif kind == "products":
        title = "Отчёт: Товары" if ru else "Mahsulotlar hisoboti"
        headers = (["ID", "Продавец", "Категория", "Товар", "Цена", "Статус", "Остаток", "Дата"]
                   if ru else
                   ["ID", "Sotuvchi", "Kategoriya", "Nom", "Narx", "Holat", "Zahira", "Sana"])
        data = []
        for p in db.get_all_products():
            p = dict(p)
            data.append([p.get("id"), p.get("seller_name") or p.get("seller_id") or "",
                         p.get("category_name") or "", p.get("name") or "", p.get("price") or 0,
                         _xl_loc("status", p.get("status"), ru), p.get("stock_count") if p.get("stock_count") is not None else "∞",
                         str(p.get("created_at") or "")[:10]])
        content, n = _xlsx_report(title, headers, data, money_cols=(4,), lang=lang)
        fn = "products"
    elif kind in ("orders", "seller_orders"):
        title = "Отчёт: Заказы" if ru else "Buyurtmalar hisoboti"
        headers = (["ID", "Покупатель", "Продавец", "Товар", "Итого", "Статус", "Оплата",
                    "Доставка", "Дата", "Расчёт", "Оплачено", "Долг"] if ru else
                   ["ID", "Xaridor", "Sotuvchi", "Mahsulot", "Jami", "Holat", "To'lov",
                    "Yetkazish", "Sana", "To'lov holati", "To'langan", "Qarz"])
        data = []
        for o in db.get_all_orders():
            o = dict(o)
            data.append([o.get("id"), o.get("buyer_name") or "", o.get("seller_name") or "",
                         o.get("product_name") or "", o.get("total_price") or o.get("price") or 0,
                         _xl_loc("status", o.get("status"), ru), _xl_loc("payment", o.get("payment_method"), ru),
                         _xl_loc("delivery", o.get("delivery_type"), ru), str(o.get("created_at") or "")[:16],
                         _xl_loc("settlement", o.get("settlement_type"), ru), o.get("paid_amount") or 0,
                         o.get("debt_amount") or 0])
        content, n = _xlsx_report(title, headers, data, money_cols=(4, 10, 11), lang=lang)
        fn = "orders"
    else:
        raise HTTPException(status_code=400, detail="bad_kind")
    fname = f"tezbozor_{fn}_{_dt.datetime.now().strftime('%Y%m%d')}.xlsx"
    return content, fname, n


@app.post("/api/admin/export/{kind}")
async def api_admin_export(kind: str, authorization: str = Header(None)):
    """users|products|orders -> Excel yasaydi va adminning Telegram chatiga yuboradi."""
    admin = _admin_from_auth(authorization)
    if kind not in ("users", "products", "orders"):
        raise HTTPException(status_code=400, detail="bad_kind")
    _rate_limit("admin_export", admin["id"], 10, 600)
    lang = get_user_lang(admin) or DEFAULT_LANG
    content, fname, n = await asyncio.to_thread(_build_excel, kind, lang)
    if not admin.get("telegram_id"):
        raise HTTPException(status_code=400, detail="no_telegram")
    res = await _tg_send_document(admin["telegram_id"], fname, content,
                                  caption=f"📊 {kind} — {n} ta · {BRAND_NAME}")
    if not (res and res.get("ok")):
        raise HTTPException(status_code=502, detail="send_failed")
    return {"ok": True, "rows": n}


@app.get("/api/admin/disputes")
def api_admin_disputes(authorization: str = Header(None)):
    _admin_from_auth(authorization)
    return _rows(db.get_disputed_orders())


@app.get("/api/admin/dispute/{order_id}/messages")
def api_admin_dispute_messages(order_id: int, authorization: str = Header(None)):
    _admin_from_auth(authorization)
    order = db.get_order_by_id(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="not_found")
    return {"messages": _rows(db.get_dispute_messages(order_id))}


class DisputeResolveIn(BaseModel):
    cancel: bool   # True -> buyurtmani bekor; False -> kuchda qoldirish


@app.post("/api/admin/dispute/{order_id}/resolve")
async def api_admin_resolve_dispute(order_id: int, body: DisputeResolveIn,
                                    authorization: str = Header(None)):
    """Nizoni hal qilish — bot admin_resolve_dispute bilan bir xil (ikkala tomonga xabar)."""
    _admin_from_auth(authorization)
    order = db.get_order_by_id(order_id)
    if not order or order.get("cancel_state") != "disputed":
        raise HTTPException(status_code=404, detail="dispute_not_found")
    do_cancel = bool(body.cancel)
    if not db.resolve_order_dispute(order_id, do_cancel):
        raise HTTPException(status_code=409, detail="not_resolved")
    if do_cancel:
        # Bekorda omborni qaytaramiz (confirmed buyurtma stokni kamaytirgan edi)
        try:
            db.restock_on_cancel(order["product_id"], order.get("quantity") or 1)
        except Exception as e:
            logging.warning(f"dispute restock xato (order {order_id}): {e}")
    oid = fmt_order_id(order_id)
    pname = html.escape(order.get("product_name") or "")
    key = "dispute_resolved_cancel" if do_cancel else "dispute_resolved_keep"
    for uid_key, tg_key in (("buyer_id", "buyer_tg"), ("seller_id", "seller_tg")):
        try:
            u = db.get_user_by_id(order[uid_key]) if order.get(uid_key) else None
            ulang = get_user_lang(u) if u else DEFAULT_LANG
            if order.get(tg_key):
                await _tg_call("sendMessage", {"chat_id": order[tg_key],
                               "text": t(ulang, key, oid=oid, pname=pname), "parse_mode": "HTML"})
        except Exception as e:
            logging.warning(f"dispute resolve notify xato (order {order_id}): {e}")
    return {"ok": True, "cancelled": do_cancel}


class DisputeMsgIn(BaseModel):
    text: str
    party: str   # 'buyer' | 'seller' — kimga (qaysi tomonga) yoziladi


@app.post("/api/admin/dispute/{order_id}/message")
async def api_admin_dispute_message(order_id: int, body: DisputeMsgIn,
                                    authorization: str = Header(None)):
    """Admin nizo bo'yicha tomonga xabar yozadi (bot admin_dispute_msg bilan bir xil)."""
    admin = _admin_from_auth(authorization)
    text = (body.text or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="empty")
    if len(text) > 2000:
        raise HTTPException(status_code=400, detail="too_long")
    if body.party not in ("buyer", "seller"):
        raise HTTPException(status_code=400, detail="bad_party")
    order = db.get_order_by_id(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="not_found")
    db.add_dispute_message(order_id, body.party, "admin", admin["id"],
                           admin.get("name") or "Admin", text)
    try:
        tg = order.get("buyer_tg") if body.party == "buyer" else order.get("seller_tg")
        uid = order.get("buyer_id") if body.party == "buyer" else order.get("seller_id")
        u = db.get_user_by_id(uid) if uid else None
        ulang = get_user_lang(u) if u else DEFAULT_LANG
        if tg:
            oid = fmt_order_id(order_id)
            await _tg_call("sendMessage", {"chat_id": tg,
                           "text": t(ulang, "dispute_admin_message", oid=oid, msg=html.escape(text)),
                           "parse_mode": "HTML"})
        # App-banner: nizo bo'yicha admin xabari (tomonga)
        _notify_db(uid, "dispute", ("⚖️ Nizo bo'yicha admin xabari", "⚖️ Сообщение админа по спору"),
                   (text[:120], text[:120]), ref_id=order_id)
    except Exception as e:
        logging.warning(f"dispute admin msg notify xato (order {order_id}): {e}")
    return {"ok": True}


@app.get("/api/admin/shops")
def api_admin_shops(authorization: str = Header(None)):
    _admin_from_auth(authorization)
    return _rows(db.get_all_shops())


@app.get("/api/admin/orders")
def api_admin_orders(authorization: str = Header(None)):
    _admin_from_auth(authorization)
    return _rows(db.get_all_orders())


@app.post("/api/admin/user/{user_id}/verify")
async def api_admin_verify(user_id: int, authorization: str = Header(None)):
    """Sotuvchini muzlatish/faollashtirish (is_approved toggle) — bot admin_verify_user parite."""
    _admin_from_auth(authorization)
    u = db.get_user_by_id(user_id)
    if not u:
        raise HTTPException(status_code=404, detail="user_not_found")
    u = dict(u)
    if bool(u.get("shop_name")):
        if u.get("is_approved"):
            db.update_user(user_id, is_approved=0)
            new_state, key = 0, "seller_frozen_notify"
        else:
            db.update_user(user_id, is_approved=1, role="seller")
            req = db.get_seller_request_by_user(user_id)
            if req:
                db.update_seller_request(dict(req)["id"], "approved")
            new_state, key = 1, "seller_reactivated_notify"
        try:
            if u.get("telegram_id"):
                await _tg_call("sendMessage", {"chat_id": u["telegram_id"],
                                               "text": t(get_user_lang(u), key), "parse_mode": "HTML"})
        except Exception as e:
            logging.warning(f"verify notify xato: {e}")
        return {"ok": True, "is_approved": new_state}
    # oddiy foydalanuvchi — is_verified toggle
    v = db.get_user_is_verified(user_id)
    nv = 0 if v else 1
    db.update_user(user_id, is_verified=nv)
    return {"ok": True, "is_verified": nv}


@app.post("/api/admin/order/{order_id}/force-cancel")
async def api_admin_force_cancel(order_id: int, authorization: str = Header(None)):
    """Admin buyurtmani majburan bekor qiladi + ikkala tomonga xabar (bot parite)."""
    _admin_from_auth(authorization)
    order = db.get_order_by_id(order_id)
    if not order:
        raise HTTPException(status_code=404, detail="not_found")
    order = dict(order)
    if order.get("status") in ("cancelled", "delivered"):
        raise HTTPException(status_code=409, detail="already_closed")
    # ATOMIK bekor: 'confirmed' bo'lsa zahirani QAYTARAMIZ (tasdiqда kamaytirilgan edi),
    # 'pending' bo'lsa shart emas. Poyga: shu orada delivered/cancelled bo'lsa — bekor qilmaymiz.
    if db.transition_order_status(order_id, "cancelled", "confirmed", cancel_by="admin"):
        try:
            db.restock_on_cancel(order["product_id"], order.get("quantity") or 1)
        except Exception as e:
            logging.warning(f"force-cancel restock xato (order {order_id}): {e}")
    elif not db.transition_order_status(order_id, "cancelled", "pending", cancel_by="admin"):
        raise HTTPException(status_code=409, detail="already_closed")
    oid = fmt_order_id(order_id)
    for tg_id, uid in [(order.get("buyer_tg"), order.get("buyer_id")),
                       (order.get("seller_tg"), order.get("seller_id"))]:
        if tg_id:
            u = db.get_user_by_id(uid) if uid else None
            lang = get_user_lang(dict(u)) if u else DEFAULT_LANG
            try:
                await _tg_call("sendMessage", {"chat_id": tg_id,
                                               "text": t(lang, "admin_cancel_notify", oid=oid)})
            except Exception:
                pass
    return {"ok": True}


@app.get("/api/admin/audit")
def api_admin_audit(authorization: str = Header(None)):
    """Mahsulot audit jurnali (kim qaysi mahsulotni o'chirdi/tikladi va h.k.)."""
    _admin_from_auth(authorization)
    return _rows(db.get_product_audit(limit=50))


@app.get("/api/admin/products")
def api_admin_products(q: str = "", offset: int = 0, authorization: str = Header(None)):
    """Admin mahsulotlari — to'liq ma'lumot (rasm, do'kon, status, qoldiq, sotilgan)
    + qidiruv + sahifalash (50 tadan)."""
    _admin_from_auth(authorization)
    return _rows(db.get_admin_products_full(q=(q or "").strip() or None,
                                            limit=50, offset=max(0, offset)))


@app.delete("/api/admin/product/{product_id}")
def api_admin_delete_product(product_id: int, authorization: str = Header(None)):
    admin = _admin_from_auth(authorization)
    prod = db.get_product_by_id(product_id)
    if not prod:
        raise HTTPException(status_code=404, detail="not_found")
    db.delete_product(product_id, deleted_by=admin["id"], deleted_by_role="admin")
    return {"ok": True}


class BroadcastIn(BaseModel):
    text: str
    target: str = "all"   # 'all' | 'buyer' | 'seller'


@app.post("/api/admin/broadcast")
async def api_admin_broadcast(body: BroadcastIn, authorization: str = Header(None)):
    """Ommaviy xabar — barcha (yoki rol bo'yicha) foydalanuvchilarga yuboradi."""
    admin = _admin_from_auth(authorization)
    _rate_limit("admin_broadcast", admin["id"], 3, 600)
    text = (body.text or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="empty")
    if len(text) > 4000:
        raise HTTPException(status_code=400, detail="too_long")
    role = body.target if body.target in ("buyer", "seller") else None
    users = db.get_all_users(role=role)
    sent, failed, app_notified = 0, 0, 0
    title = "📢 Yangilik / Новость"
    for u in users:
        u = dict(u)
        if u.get("is_blocked"):
            continue
        # App-banner: HAMMA foydalanuvchi ko'radi — botni shaxsiy chatda /start qilmagan
        # (app-first ro'yxatdan o'tgan) userlar ham. Telegram push faqat qo'shimcha.
        _notify_db(u.get("id"), "info", title, text)
        app_notified += 1
        tg = u.get("telegram_id")
        if not tg:
            continue
        # parse_mode'siz oddiy matn: apostrof va maxsus belgilar to'g'ri chiqsin
        # (html.escape + parse_mode'siz "&#x27;" kabi belgilarga aylanardi).
        res = await _tg_call("sendMessage", {"chat_id": tg, "text": text})
        if res and res.get("ok"):
            sent += 1
        else:
            failed += 1
    return {"ok": True, "sent": sent, "failed": failed,
            "app_notified": app_notified, "total": len(users)}


# ---- #6 KANALLAR (bot admin_channels pariteti) ----
@app.get("/api/admin/channels")
def api_admin_channels(authorization: str = Header(None)):
    """Barcha sotuvchi kanal/guruhlari — frontend sotuvchi bo'yicha guruhlaydi."""
    _admin_from_auth(authorization)
    return _rows(db.get_all_seller_channels())


# ---- #7 DO'KON DETALI + MODERATSIYA TOGGLE (bot admin_shop_detail/shopmod) ----
@app.get("/api/admin/shop/{shop_id}")
def api_admin_shop_detail(shop_id: int, authorization: str = Header(None)):
    _admin_from_auth(authorization)
    shop = db.get_shop_by_id(shop_id)
    if not shop:
        raise HTTPException(status_code=404, detail="shop_not_found")
    shop = dict(shop)
    owner = db.get_user_by_id(shop.get("owner_user_id")) if shop.get("owner_user_id") else None
    perf = {r["user_id"]: dict(r) for r in db.get_shop_staff_performance(shop_id)}
    staff = []
    for s in db.get_shop_staff(shop_id):
        s = dict(s)
        s["revenue"] = (perf.get(s["user_id"], {}) or {}).get("revenue", 0)
        staff.append(s)
    return {"shop": shop, "owner_name": (dict(owner).get("name") if owner else None),
            "moderation": shop.get("moderation") or "direct",
            "payment_mode": shop.get("payment_mode") or "shop", "staff": staff}


@app.post("/api/admin/shop/{shop_id}/toggle-mod")
def api_admin_shop_toggle_mod(shop_id: int, authorization: str = Header(None)):
    """Moderatsiya siyosatini almashtiradi: direct ↔ owner_approve."""
    _admin_from_auth(authorization)
    shop = db.get_shop_by_id(shop_id)
    if not shop:
        raise HTTPException(status_code=404, detail="shop_not_found")
    new_mod = "owner_approve" if (dict(shop).get("moderation") or "direct") == "direct" else "direct"
    db.update_shop(shop_id, moderation=new_mod)
    return {"ok": True, "moderation": new_mod}


@app.post("/api/admin/shop/{shop_id}/staff/{staff_id}/toggle")
async def api_admin_staff_toggle(shop_id: int, staff_id: int, authorization: str = Header(None)):
    """Admin xodimni faollashtiradi/muzlatadi (bot admin_staff_toggle pariteti)."""
    _admin_from_auth(authorization)
    target = next((dict(s) for s in db.get_shop_staff(shop_id, include_owner=False)
                   if dict(s).get("id") == staff_id), None)
    if not target:
        raise HTTPException(status_code=404, detail="staff_not_found")
    new_active = 0 if target.get("is_active") else 1
    db.set_staff_active(staff_id, new_active)
    try:
        su = db.get_user_by_id(target.get("user_id"))
        if su and dict(su).get("telegram_id"):
            slang = get_user_lang(dict(su))
            await _tg_call("sendMessage", {"chat_id": dict(su)["telegram_id"],
                           "text": t(slang, "staff_you_activated" if new_active else "staff_you_frozen")})
    except Exception as e:
        logging.warning(f"admin staff toggle notify xato (staff {staff_id}): {e}")
    return {"ok": True, "is_active": new_active}


# ---- #5 AVTO-MODERATSIYA NAVBATI (admin tasdiq/rad) ----
async def _notify_seller_moderation(prod, approved):
    tg = prod.get("seller_tg")
    if not tg:
        return
    seller = db.get_user_by_id(prod.get("seller_id")) if prod.get("seller_id") else None
    slang = get_user_lang(seller) if seller else DEFAULT_LANG
    pname = html.escape(prod.get("name") or "")
    if approved:
        txt = (f"✅ Mahsulotingiz tekshiruvdan o'tdi va sotuvga qo'yildi:\n🛍️ {pname}"
               if slang == "uz" else f"✅ Ваш товар прошёл проверку и опубликован:\n🛍️ {pname}")
    else:
        txt = (f"🚫 Mahsulotingiz qoidalarga zid deb topildi va rad etildi:\n🛍️ {pname}"
               if slang == "uz" else f"🚫 Ваш товар отклонён как нарушающий правила:\n🛍️ {pname}")
    try:
        await _tg_call("sendMessage", {"chat_id": tg, "text": txt, "parse_mode": "HTML"})
    except Exception as e:
        logging.warning(f"moderation seller notify xato: {e}")


@app.get("/api/admin/fraud")
def api_admin_fraud(authorization: str = Header(None)):
    """AI #7 — firibgarlik shubhasi signallari (evristik; admin tekshiradi)."""
    _admin_from_auth(authorization)
    return db.get_fraud_signals()


@app.get("/api/admin/moderation")
def api_admin_moderation_queue(authorization: str = Header(None)):
    _admin_from_auth(authorization)
    return _rows(db.get_moderation_queue())


@app.post("/api/admin/moderation/{product_id}/approve")
async def api_admin_mod_approve(product_id: int, authorization: str = Header(None)):
    _admin_from_auth(authorization)
    prod = db.get_product_by_id(product_id)
    if not prod or dict(prod).get("status") != "mod_blocked":
        raise HTTPException(status_code=404, detail="not_found")
    db.update_product_fields(product_id, status="active", in_stock=1, mod_reason=None)
    await _notify_seller_moderation(dict(prod), approved=True)
    return {"ok": True}


@app.post("/api/admin/moderation/{product_id}/reject")
async def api_admin_mod_reject(product_id: int, authorization: str = Header(None)):
    _admin_from_auth(authorization)
    prod = db.get_product_by_id(product_id)
    if not prod or dict(prod).get("status") != "mod_blocked":
        raise HTTPException(status_code=404, detail="not_found")
    db.set_product_status(product_id, "deleted")
    await _notify_seller_moderation(dict(prod), approved=False)
    return {"ok": True}


# ---- #8 O'CHIRILGAN MAHSULOTLAR AUDITI (bot admin_deleted_products — faqat ko'rish) ----
@app.get("/api/admin/deleted-products")
def api_admin_deleted_products(authorization: str = Header(None)):
    _admin_from_auth(authorization)
    return _rows(db.get_product_audit(limit=50))


# ---- #9 SOZLAMALAR (bot admin_settings — hisoblar + backup + tozalash) ----
@app.get("/api/admin/settings")
def api_admin_settings(authorization: str = Header(None)):
    _admin_from_auth(authorization)
    return {"admin_id": ADMIN_ID,
            "users": len(db.get_all_users()),
            "products": len(db.get_all_products(include_hidden=False)),
            "orders": len(db.get_all_orders())}


@app.post("/api/admin/clean-cancelled")
def api_admin_clean_cancelled(authorization: str = Header(None)):
    """30 kundan eski bekor qilingan buyurtmalarni tozalaydi."""
    _admin_from_auth(authorization)
    n = db.clean_old_cancelled_orders(30)
    return {"ok": True, "deleted": n}


@app.post("/api/admin/backup")
async def api_admin_backup(authorization: str = Header(None)):
    """DB backup faylini adminning Telegram chatiga yuboradi (SQLite). PG'da no-op → 400."""
    admin = _admin_from_auth(authorization)
    _rate_limit("admin_backup", admin["id"], 5, 600)
    if not admin.get("telegram_id"):
        raise HTTPException(status_code=400, detail="no_telegram")
    import datetime as _dt
    import tempfile
    ts = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(tempfile.gettempdir(), f"marketplace_backup_{ts}.db")
    ok = await asyncio.to_thread(db.backup, path)
    if not ok:
        raise HTTPException(status_code=400, detail="backup_unavailable")
    try:
        with open(path, "rb") as f:
            content = f.read()
        res = await _tg_send_document(admin["telegram_id"], f"marketplace_backup_{ts}.db",
                                      content, caption=f"💾 Backup · {ts}")
    finally:
        try:
            os.remove(path)
        except OSError:
            pass
    if not (res and res.get("ok")):
        raise HTTPException(status_code=502, detail="send_failed")
    return {"ok": True}


# ============================================================
# AI VISION (Gemini multimodal) — rasm/ovoz qidiruv, rasmdan mahsulot, banner
# ============================================================
MAX_VOICE_BYTES = 4 * 1024 * 1024   # ovozli xabar (30-60s opus ~ 100-500KB)
# BRAND_NAME fayl boshida (app = FastAPI(...) dan oldin) e'lon qilingan.


async def _tg_file_bytes(file_id):
    """Telegram file_id → bytes (rasm disk-keshidan bo'lsa o'shandan). Xatoda None."""
    safe = hashlib.sha256(file_id.encode()).hexdigest()
    cache_path = os.path.join(IMG_CACHE_DIR, safe + ".jpg")
    if os.path.exists(cache_path):
        try:
            with open(cache_path, "rb") as f:
                return f.read()
        except Exception:
            pass
    if not BOT_TOKEN:
        return None
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            meta = await client.get(
                f"https://api.telegram.org/bot{BOT_TOKEN}/getFile",
                params={"file_id": file_id})
            data = meta.json()
            if not data.get("ok"):
                return None
            path = data["result"]["file_path"]
            r = await client.get(f"https://api.telegram.org/file/bot{BOT_TOKEN}/{path}")
            r.raise_for_status()
            return r.content
    except Exception as e:
        logging.warning(f"tg fayl yuklab olish xato ({file_id[:12]}...): {e}")
        return None


async def _tg_upload_photo(chat_id, data, filename="photo.jpg",
                           content_type="image/jpeg"):
    """Baytlarni Telegram'ga yuborib file_id oladi (xabar darhol o'chiriladi).
    Xatoda None."""
    if not BOT_TOKEN or not chat_id:
        return None
    try:
        async with httpx.AsyncClient(timeout=45) as client:
            r = await client.post(
                f"https://api.telegram.org/bot{BOT_TOKEN}/sendPhoto",
                data={"chat_id": str(chat_id)},
                files={"photo": (filename, data, content_type)})
            res = r.json()
    except Exception as e:
        logging.warning(f"AI photo upload xato: {e}")
        return None
    if not res.get("ok") or not (res.get("result") or {}).get("photo"):
        return None
    msg = res["result"]
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            await client.post(f"https://api.telegram.org/bot{BOT_TOKEN}/deleteMessage",
                              json={"chat_id": chat_id, "message_id": msg["message_id"]})
    except Exception:
        pass
    return msg["photo"][-1]["file_id"]


async def _tg_upload_video(chat_id, data, filename="clip.mp4"):
    """Video baytlarni Telegram'ga yuborib file_id oladi (xabar darhol o'chiriladi
    — _tg_upload_photo bilan bir uslub). Xatoda None."""
    if not BOT_TOKEN or not chat_id:
        return None
    try:
        async with httpx.AsyncClient(timeout=120) as client:
            r = await client.post(
                f"https://api.telegram.org/bot{BOT_TOKEN}/sendVideo",
                data={"chat_id": str(chat_id), "supports_streaming": "true"},
                files={"video": (filename, data, "video/mp4")})
            res = r.json()
    except Exception as e:
        logging.warning(f"AI video upload xato: {e}")
        return None
    if not res.get("ok") or not (res.get("result") or {}).get("video"):
        return None
    msg = res["result"]
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            await client.post(f"https://api.telegram.org/bot{BOT_TOKEN}/deleteMessage",
                              json={"chat_id": chat_id, "message_id": msg["message_id"]})
    except Exception:
        pass
    return msg["video"]["file_id"]


def _ai_http_error(code, lang):
    """ai_vision xato kodini HTTPException'ga aylantiradi (401 emas — AI xatosi
    klientni logoutga olib bormasin: hammasi 503, faqat parse 422)."""
    status = 422 if code == "parse" else 503
    return HTTPException(status_code=status, detail={
        "code": f"ai_{code}", "message": ai_vision.error_message(code, lang)})


def _seller_or_403(user):
    if user.get("role") not in ("seller", "admin") and not user.get("is_approved"):
        raise HTTPException(status_code=403, detail="not_seller")


@app.post("/api/seller/product/photo-analyze")
async def api_product_photo_analyze(file: UploadFile = File(...),
                                    authorization: str = Header(None)):
    """✨ Rasmdan mahsulot: sotuvchi rasm yuklaydi — AI nom, tavsif, kategoriya va
    narx taklifini qaytaradi (forma avto-to'ldiriladi)."""
    user = dict(_buyer_from_auth(authorization))
    _seller_or_403(user)
    lang = get_user_lang(user) or DEFAULT_LANG
    if not ai_vision.is_enabled():
        raise _ai_http_error("disabled", lang)
    _rate_limit("ai_photo_analyze", user["id"], 20, 3600)
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="empty")
    if len(data) > MAX_PHOTO_BYTES:
        raise HTTPException(status_code=413, detail="too_large")
    mime = (file.content_type or "").strip()
    if not mime.startswith("image/"):
        mime = "image/jpeg"
    cats = [dict(c) for c in (db.get_categories() or [])]
    res = await ai_vision.analyze_product_photo(
        data, mime=mime, categories=[c.get("name") for c in cats], lang=lang)
    if res.get("error"):
        raise _ai_http_error(res["error"], lang)
    # AI tanlagan kategoriya nomini id'ga bog'laymiz (aniq mos kelsagina)
    cat_id = None
    if res.get("category"):
        want = res["category"].strip().lower()
        for c in cats:
            if (c.get("name") or "").strip().lower() == want:
                cat_id = c.get("id")
                break
    res["category_id"] = cat_id
    return {"ok": True, "suggestion": res}


class AiBannerIn(BaseModel):
    style: Optional[str] = None   # ixtiyoriy uslub tilagi ("bayramona", "minimal"...)


@app.post("/api/seller/product/{product_id}/ai-banner")
async def api_ai_banner(product_id: int, body: AiBannerIn = None,
                        authorization: str = Header(None)):
    """🎨 Gemini reklama banneri: mahsulot rasmi + narx/nomdan tayyor dizayn rasm.
    Model ishlamasa (billing/kvota) — ok=False, frontend PIL ad-preview'ga qaytadi."""
    user = dict(_buyer_from_auth(authorization))
    prod = dict(_own_product_or_403(user, product_id))
    lang = get_user_lang(user) or DEFAULT_LANG
    if not ai_vision.is_enabled():
        raise _ai_http_error("disabled", lang)
    _rate_limit("ai_banner", user["id"], 6, 3600)
    img_fid = prod.get("image_url")
    if not img_fid:
        raise HTTPException(status_code=400, detail="no_photo")
    img = await _tg_file_bytes(img_fid)
    if not img:
        raise HTTPException(status_code=502, detail="image_fetch_failed")
    shop = db.get_shop_by_owner(_owner_id(user))
    shop_name = (dict(shop).get("name") if shop else None) or user.get("shop_name") or BRAND_NAME
    res = await ai_vision.generate_banner_image(
        img, product_name=prod.get("name") or "",
        price_text=fmt_price(prod.get("price")).replace(" ", " "),
        shop_name=shop_name, style_hint=(body.style if body else "") or "", lang=lang)
    if res.get("error"):
        # Fallback signali — frontend mavjud PIL reklama dizaynini ishlatadi
        return {"ok": False, "fallback": True, "code": res["error"],
                "message": ai_vision.error_message(res["error"], lang)}
    ext = "png" if "png" in (res.get("image_mime") or "") else "jpg"
    file_id = await _tg_upload_photo(user.get("telegram_id"), res["image"],
                                     filename=f"banner.{ext}",
                                     content_type=res.get("image_mime") or "image/png")
    if not file_id:
        raise HTTPException(status_code=502, detail="upload_failed")
    return {"ok": True, "file_id": file_id}


# 🎬 AI VIDEO-REKLAMA — ffmpeg render CPU-og'ir, bir vaqtda BITTAdan (navbat)
_ADCLIP_LOCK = asyncio.Lock()


class AdClipIn(BaseModel):
    hook: Optional[str] = None   # sotuvchi o'z hook matnini bersa (bo'sh = AI yozadi)
    music: Optional[bool] = True  # 🎵 fon musiqa (sotuvchi tanlovi; standart — yoqilgan)


@app.post("/api/seller/product/{product_id}/adclip")
async def api_ad_clip(product_id: int, body: AdClipIn = None,
                      authorization: str = Header(None)):
    """🎬 AI video-reklama: mahsulot rasm(lar)i + AI hook matnidan 8-14 soniyalik
    reels-uslub klip (ffmpeg, tashqi API'siz). Klip sotuvchi chatiga yuklanib
    file_id qaytadi; App preview'da ko'rsatadi, ad-publish video_id bilan chiqaradi.
    Diskdagi nusxa /api/video keshiga oldindan yoziladi — preview darhol ochiladi."""
    user = dict(_buyer_from_auth(authorization))
    prod = dict(_own_product_or_403(user, product_id))
    lang = get_user_lang(user) or DEFAULT_LANG
    if not ad_video.is_enabled():
        raise HTTPException(status_code=503, detail="clip_disabled")
    _rate_limit("adclip", user["id"], 5, 3600)
    img_fids = db.get_product_images(product_id)[:3]
    if not img_fids:
        raise HTTPException(status_code=400, detail="no_photo")
    images = []
    for fid in img_fids:
        b = await _tg_file_bytes(fid)
        if b:
            images.append(b)
    if not images:
        raise HTTPException(status_code=502, detail="image_fetch_failed")

    hook = (body.hook if body else "").strip()[:48] if (body and body.hook) else ""
    if not hook:
        hook = await ai_assistant.generate_clip_hook(
            name=prod.get("name") or "", category=prod.get("category_name") or "",
            lang=lang, seed=product_id)
    shop = db.get_shop_by_owner(_owner_id(user))
    shop_name = (dict(shop).get("name") if shop else None) or user.get("shop_name") or BRAND_NAME
    optom_txt = ""
    if prod.get("sale_mode") == "optom":
        _ps = prod.get("pack_size")
        optom_txt = f"OPTOM · 1 PACHKA = {int(_ps)} DONA" if _ps else "OPTOM"
    cta = "SOTIB OLISH" if lang != "ru" else "КУПИТЬ"

    # 🎵 musiqa: None = avto (assets/ad_music.mp3), "" = sotuvchi o'chirgan
    music_path = None if (body is None or body.music is not False) else ""
    async with _ADCLIP_LOCK:
        clip = await asyncio.to_thread(
            ad_video.build_ad_clip, images,
            hook_text=hook,
            price_text=fmt_price(prod.get("price")),
            shop_text=str(shop_name or ""),
            cta_text=cta, optom_text=optom_txt, brand_text=BRAND_NAME,
            music_path=music_path)
    if not clip:
        raise HTTPException(status_code=502, detail="clip_failed")
    file_id = await _tg_upload_video(user.get("telegram_id"), clip,
                                     filename=f"ad_{product_id}.mp4")
    if not file_id:
        raise HTTPException(status_code=502, detail="upload_failed")
    # /api/video disk-keshini oldindan to'ldiramiz — preview getFile'siz darhol ochiladi
    try:
        safe = hashlib.sha256(file_id.encode()).hexdigest()
        with open(os.path.join(IMG_CACHE_DIR, safe + ".mp4"), "wb") as f:
            f.write(clip)
    except Exception:
        pass
    return {"ok": True, "file_id": file_id, "hook": hook}


@app.post("/api/ai/voice-search")
async def api_voice_search(file: UploadFile = File(...),
                           authorization: str = Header(None),
                           category_id: int = Query(None),
                           region_id: int = Query(None)):
    """🎤 Ovozli qidiruv: audio → Gemini transkripsiya → AI kalit so'zlar → mahsulotlar."""
    user = dict(_buyer_from_auth(authorization))
    lang = get_user_lang(user) or DEFAULT_LANG
    if not ai_vision.is_enabled():
        raise _ai_http_error("disabled", lang)
    _rate_limit("ai_voice", user["id"], 15, 3600)
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="empty")
    if len(data) > MAX_VOICE_BYTES:
        raise HTTPException(status_code=413, detail="too_large")
    mime = (file.content_type or "audio/ogg").split(";")[0].strip().lower()
    if not mime.startswith("audio/"):
        mime = "audio/ogg"
    tr = await ai_vision.transcribe_voice(data, mime=mime, lang=lang)
    if tr.get("error") and mime not in ("audio/ogg", "audio/mp3", "audio/wav"):
        # MediaRecorder webm/mp4 bersa va model qabul qilmasa — ogg deb qayta uramiz
        tr = await ai_vision.transcribe_voice(data, mime="audio/ogg", lang=lang)
    if tr.get("error"):
        raise _ai_http_error(tr["error"], lang)
    text = tr["text"][:300]

    cat_names = []
    try:
        cat_names = [dict(c).get("name") for c in (db.get_categories() or [])]
    except Exception:
        pass
    interp = await ai_assistant.interpret_search_query(text, categories=cat_names, lang=lang)
    keywords = (interp or {}).get("keywords") or [text]
    seen, items = set(), []
    for kw in keywords:
        try:
            rows = db.search_products(query=kw, category_id=category_id,
                                      sort_by="rating", region_id=region_id)
        except Exception:
            rows = []
        for r in rows:
            rid = dict(r).get("id")
            if rid not in seen:
                seen.add(rid)
                items.append(r)
        if len(items) >= 30:
            break
    return {"transcript": text, "interpreted": ", ".join(keywords),
            "items": _hide_own(_rows(items[:30]), _own_shop_seller_id(authorization))}


@app.post("/api/ai/image-search")
async def api_image_search(file: UploadFile = File(...),
                           authorization: str = Header(None),
                           category_id: int = Query(None),
                           region_id: int = Query(None)):
    """📷 Rasm orqali qidiruv: xaridor rasm yuklaydi → Gemini mahsulotni aniqlaydi →
    kalit so'zlar bilan katalogdan qidiriladi (ovozli qidiruv bilan bir oqim)."""
    user = dict(_buyer_from_auth(authorization))
    lang = get_user_lang(user) or DEFAULT_LANG
    if not ai_vision.is_enabled():
        raise _ai_http_error("disabled", lang)
    _rate_limit("ai_imgsearch", user["id"], 15, 3600)
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="empty")
    if len(data) > MAX_PHOTO_BYTES:
        raise HTTPException(status_code=413, detail="too_large")
    mime = (file.content_type or "").split(";")[0].strip().lower()
    if not mime.startswith("image/"):
        mime = "image/jpeg"

    cat_names = []
    try:
        cat_names = [dict(c).get("name") for c in (db.get_categories() or [])]
    except Exception:
        pass
    res = await ai_vision.identify_image_for_search(
        data, mime=mime, categories=cat_names, lang=lang)
    if res.get("error"):
        raise _ai_http_error(res["error"], lang)

    seen, items = set(), []
    for kw in (res.get("keywords") or []):
        try:
            rows = db.search_products(query=kw, category_id=category_id,
                                      sort_by="rating", region_id=region_id)
        except Exception:
            rows = []
        for r in rows:
            rid = dict(r).get("id")
            if rid not in seen:
                seen.add(rid)
                items.append(r)
        if len(items) >= 30:
            break
    return {"label": res.get("label") or "",
            "interpreted": ", ".join(res.get("keywords") or []),
            "items": _hide_own(_rows(items[:30]), _own_shop_seller_id(authorization))}


# ============================================================
# FRONTEND (static)
# ============================================================
@app.get("/")
def index():
    # no-cache: Mini App HTML yangilansa, Telegram darhol yangi versiyani olsin
    # (aks holda eski dizayn keshda qolib ketadi).
    return FileResponse(
        os.path.join(STATIC_DIR, "index.html"),
        headers={"Cache-Control": "no-cache, no-store, must-revalidate"},
    )


if os.path.isdir(STATIC_DIR):
    app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
