"""End-to-end RUNTIME oqim testlari — haqiqiy webapp endpoint kodini TestClient bilan
jonli bosadi (real temp-SQLite, imzolangan initData). 2026-06-21 runtime auditidan
rasmiylashtirilgan: atomik guard, settlement, nizo, xodim, to'lov zanjiri, auth himoyasi.

Maqsad: "ulangan ≠ ishlaydi" — bu testlar XULQNI tasdiqlaydi, shunchaki route mavjudligini emas.
Har deployda avtomatik bosiladi → "soxta tugadi" regressiyasini ushlaydi."""
import json
import os
import time

import pytest

os.environ.setdefault("BOT_TOKEN", "123456:TEST-BOT-TOKEN")
os.environ.setdefault("DB_BACKEND", "sqlite")

pytest.importorskip("fastapi")
from fastapi.testclient import TestClient  # noqa: E402

import webapp_auth  # noqa: E402
import webapp_server  # noqa: E402
from database import Database  # noqa: E402

TOKEN = os.environ["BOT_TOKEN"]


def hdr(tg_id):
    init = webapp_auth.build_init_data(TOKEN, {
        "user": json.dumps({"id": tg_id, "first_name": "U"}),
        "auth_date": str(int(time.time()))})
    return {"Authorization": "tma " + init}


def rows(j):
    if isinstance(j, list):
        return j
    if isinstance(j, dict):
        for k in ("items", "orders", "payments", "disputes", "debts", "reviews", "staff"):
            if isinstance(j.get(k), list):
                return j[k]
    return []


@pytest.fixture
def env(tmp_path, monkeypatch):
    """To'liq sozlangan, TASDIQLANGAN sotuvchi + do'kon + xaridor + admin + mahsulot.
    Telegram chaqiruvlari mock (tarmoqqa chiqmaydi)."""
    d = Database(db_path=str(tmp_path / "rt.db"))
    monkeypatch.setattr(webapp_server, "db", d)
    monkeypatch.setattr(webapp_server, "BOT_TOKEN", TOKEN)
    webapp_server._RATE.clear()

    async def _tg(method, payload, **kw):
        return {"ok": True, "result": {"message_id": 1}}
    monkeypatch.setattr(webapp_server, "_tg_call", _tg)
    if hasattr(webapp_server, "_tg_send_document"):
        async def _doc(*a, **k):
            return True
        monkeypatch.setattr(webapp_server, "_tg_send_document", _doc)
    # AI standart holatda O'CHIQ: .env da haqiqiy kalit bo'lsa testlar tarmoqqa
    # chiqib ketmasin va natija kalit borligiga qarab o'zgarmasin. AI xulqini
    # tekshiradigan testlar buni o'zi is_enabled=True qilib qayta yoqadi.
    monkeypatch.setattr(webapp_server.ai_assistant, "is_enabled", lambda: False)

    seller = d.create_user(telegram_id=7002, phone_number="998900000002", name="Seller", role="seller")
    d.update_user(seller, is_approved=1, shop_name="Do'kon")
    admin = d.create_user(telegram_id=7003, phone_number="998900000003", name="Admin", role="admin")
    buyer = d.create_user(telegram_id=7001, phone_number="998900000001", name="Buyer", role="buyer")
    shop = d.create_shop(seller)
    pid = d.create_product(seller_id=seller, name="Mahsulot", price=10000, stock_count=5)
    d.update_product_fields(pid, in_stock=1, status="active")

    c = TestClient(webapp_server.app)
    c.db = d
    c.SELLER, c.ADMIN, c.BUYER, c.SHOP, c.PID = seller, admin, buyer, shop, pid
    return c


def _place_and_confirm(c, qty=2):
    r = c.post("/api/order", headers=hdr(7001), json={
        "product_id": c.PID, "quantity": qty, "delivery_type": "pickup", "payment_method": "cash"})
    assert r.status_code == 200, r.text
    oid = r.json()["order_id"]
    r = c.post(f"/api/seller/order/{oid}/action", headers=hdr(7002), json={"action": "confirm"})
    assert r.status_code == 200, r.text
    return oid


# ---------------- KATALOG / DETAL ----------------
def test_approved_seller_product_visible(env):
    r = env.get("/api/products", headers=hdr(7001))
    assert r.status_code == 200
    assert any(p.get("id") == env.PID for p in rows(r.json()))


def test_secret_min_price_never_exposed(env):
    r = env.get(f"/api/products/{env.PID}", headers=hdr(7001))
    assert r.status_code == 200 and "min_price" not in r.json()


# ---------------- BUYURTMA + TAYMER ----------------
def test_order_total_and_live_timer(env):
    r = env.post("/api/order", headers=hdr(7001), json={
        "product_id": env.PID, "quantity": 2, "delivery_type": "pickup", "payment_method": "cash"})
    assert r.status_code == 200 and r.json()["total"] == 20000
    oid = r.json()["order_id"]
    mo = rows(env.get("/api/my/orders", headers=hdr(7001)).json())
    o = next((x for x in mo if x.get("id") == oid), None)
    assert o and o.get("auto_cancel_at"), "jonli taymer uchun auto_cancel_at ko'rinishi shart"


# ---------------- ATOMIK GUARD (eng muhim) ----------------
def test_double_confirm_atomic_no_double_stock(env):
    r = env.post("/api/order", headers=hdr(7001), json={
        "product_id": env.PID, "quantity": 2, "delivery_type": "pickup", "payment_method": "cash"})
    oid = r.json()["order_id"]
    before = env.db.get_product_by_id(env.PID)["stock_count"]
    r1 = env.post(f"/api/seller/order/{oid}/action", headers=hdr(7002), json={"action": "confirm"})
    r2 = env.post(f"/api/seller/order/{oid}/action", headers=hdr(7002), json={"action": "confirm"})
    assert r1.status_code == 200 and r2.status_code == 409, "takroriy tasdiq 409 bo'lishi shart"
    after = env.db.get_product_by_id(env.PID)["stock_count"]
    assert before - after == 2, "zahira FAQAT bir marta kamayishi shart (poyga himoyasi)"


# ---------------- BERISH + QARZ ----------------
def test_deliver_with_debt_settlement(env):
    oid = _place_and_confirm(env, qty=2)
    r = env.post(f"/api/seller/order/{oid}/deliver", headers=hdr(7002),
                 json={"settlement_type": "debt", "paid": 5000})
    assert r.status_code == 200, r.text
    assert r.json()["due"] == 15000 and r.json()["paid"] == 5000
    assert env.db.get_order_by_id(oid)["status"] == "delivered"
    debts = rows(env.get("/api/my/debts", headers=hdr(7001)).json())
    assert any(float(x.get("total_due") or 0) >= 15000 for x in debts)


# ---------------- SHARH ----------------
def test_review_and_duplicate_block(env):
    oid = _place_and_confirm(env, qty=1)
    env.post(f"/api/seller/order/{oid}/deliver", headers=hdr(7002),
             json={"settlement_type": "paid", "paid": 10000})
    r = env.post(f"/api/order/{oid}/review", headers=hdr(7001),
                 json={"seller_rating": 5, "product_rating": 4, "comment": "Zo'r"})
    assert r.status_code == 200
    r2 = env.post(f"/api/order/{oid}/review", headers=hdr(7001), json={"seller_rating": 5})
    assert r2.status_code == 409, "takroriy sharh 409"
    assert len(rows(env.get("/api/my/reviews", headers=hdr(7001)).json())) >= 1


# ---------------- NIZO ----------------
def test_dispute_flow(env):
    oid = _place_and_confirm(env, qty=1)
    r = env.post(f"/api/seller/order/{oid}/request-cancel", headers=hdr(7002), json={"reason": "Tugadi"})
    assert r.status_code == 200
    assert env.db.get_order_by_id(oid)["cancel_state"] == "requested"
    r = env.post(f"/api/order/{oid}/cancel-respond", headers=hdr(7001), json={"agree": False})
    assert r.status_code == 200 and r.json().get("disputed") is True
    assert env.db.get_order_by_id(oid)["cancel_state"] == "disputed"
    # admin nizo ro'yxatida ko'rinadi
    dl = rows(env.get("/api/admin/disputes", headers=hdr(7003)).json())
    assert any(o.get("id") == oid for o in dl)


# ---------------- XODIM RUXSATLARI ----------------
def test_staff_invite_and_perm_toggle(env):
    r = env.post("/api/seller/staff/invite", headers=hdr(7002))
    assert r.status_code == 200 and r.json().get("code")
    staff_u = env.db.create_user(telegram_id=7004, phone_number="998900000004", name="Xodim", role="seller")
    sid = env.db.add_staff(shop_id=env.SHOP, user_id=staff_u, staff_role="staff")
    slist = rows(env.get("/api/seller/staff", headers=hdr(7002)).json())
    assert any(s.get("id") == sid for s in slist)
    d0 = env.get(f"/api/seller/staff/{sid}", headers=hdr(7002)).json()
    r = env.post(f"/api/seller/staff/{sid}/perm", headers=hdr(7002), json={"key": "price"})
    assert r.status_code == 200
    d1 = env.get(f"/api/seller/staff/{sid}", headers=hdr(7002)).json()
    assert json.dumps(d0) != json.dumps(d1), "ruxsat HAQIQATAN o'zgarishi shart"


# ---------------- MONETIZATSIYA (to'lov → fulfillment zanjiri) ----------------
def test_boost_payment_to_fulfillment_chain(env):
    # default o'chiq → 403
    assert env.post(f"/api/seller/boost/{env.PID}", headers=hdr(7002), json={}).status_code == 403
    # admin yoqadi
    env.post("/api/admin/monetization", headers=hdr(7003),
             json={"mon_enabled": True, "mon_boost_enabled": True, "mon_boost_price": 5000})
    r = env.post(f"/api/seller/boost/{env.PID}", headers=hdr(7002), json={})
    assert r.status_code == 200 and r.json().get("payment_id")
    pay_id = r.json()["payment_id"]
    # dev-confirm FAQAT admin (provayder yo'qda bepul-boost teshigini yopadi)
    assert env.post(f"/api/pay/dev-confirm/{pay_id}", headers=hdr(7002), json={}).status_code == 403
    assert env.post(f"/api/pay/dev-confirm/{pay_id}", headers=hdr(7003), json={}).status_code == 200
    assert env.db.get_payment(pay_id)["state"] == "paid"
    assert env.db.get_product_by_id(env.PID).get("boosted_until"), "to'langach boost qo'llanishi shart"


# ---------------- SOTUVCHI TO'LOVLAR TARIXI (yangi UI endpoint, 2026-06-21) ----------------
def test_seller_payments_history(env):
    env.post("/api/admin/monetization", headers=hdr(7003),
             json={"mon_enabled": True, "mon_boost_enabled": True, "mon_boost_price": 5000})
    r = env.post(f"/api/seller/boost/{env.PID}", headers=hdr(7002), json={})
    env.post(f"/api/pay/dev-confirm/{r.json()['payment_id']}", headers=hdr(7003), json={})
    r = env.get("/api/seller/payments", headers=hdr(7002))
    assert r.status_code == 200
    pays = rows(r.json())
    assert any(p.get("purpose") == "boost" and p.get("state") == "paid" for p in pays), \
        "sotuvchi to'lov tarixi boost to'lovini ko'rsatishi shart"


# ---------------- ADMIN KUTILAYOTGAN TO'LOVNI TASDIQLAYDI (2026-06-21) ----------------
def test_admin_sees_and_confirms_pending_payment(env):
    """Provayder ulanmaganda sotuvchining boost to'lovi 'pending' qoladi → admin
    /api/admin/payments'da ko'radi va dev-confirm bilan tasdiqlaydi → 'paid' bo'ladi."""
    env.post("/api/admin/monetization", headers=hdr(7003),
             json={"mon_enabled": True, "mon_boost_enabled": True, "mon_boost_price": 5000})
    r = env.post(f"/api/seller/boost/{env.PID}", headers=hdr(7002), json={})
    pay_id = r.json()["payment_id"]
    # admin pending ro'yxatida ko'rinadi (sotuvchi ismi bilan)
    pend = rows(env.get("/api/admin/payments?state=pending", headers=hdr(7003)).json())
    row = next((p for p in pend if p.get("id") == pay_id), None)
    assert row and row.get("state") == "pending" and row.get("user_name"), "admin pending to'lovni ko'rishi shart"
    # xaridor bu endpointga kira olmaydi
    assert env.get("/api/admin/payments", headers=hdr(7001)).status_code == 403
    # admin tasdiqlaydi
    assert env.post(f"/api/pay/dev-confirm/{pay_id}", headers=hdr(7003), json={}).status_code == 200
    # endi pending'da yo'q, paid'da bor
    pend2 = rows(env.get("/api/admin/payments?state=pending", headers=hdr(7003)).json())
    paid = rows(env.get("/api/admin/payments?state=paid", headers=hdr(7003)).json())
    assert not any(p.get("id") == pay_id for p in pend2), "tasdiqlangach pending'dan chiqishi shart"
    assert any(p.get("id") == pay_id for p in paid)


def test_admin_can_reject_pending_payment(env):
    """Admin kutilayotgan to'lovni tasdiqlamasdan BEKOR qiladi → 'cancelled',
    boost BERILMAYDI. To'langanini bekor qilib bo'lmaydi (409)."""
    env.post("/api/admin/monetization", headers=hdr(7003),
             json={"mon_enabled": True, "mon_boost_enabled": True, "mon_boost_price": 5000})
    pid_pay = env.post(f"/api/seller/boost/{env.PID}", headers=hdr(7002), json={}).json()["payment_id"]
    # xaridor bekor qila olmaydi
    assert env.post(f"/api/pay/dev-cancel/{pid_pay}", headers=hdr(7001), json={}).status_code == 403
    # admin bekor qiladi
    assert env.post(f"/api/pay/dev-cancel/{pid_pay}", headers=hdr(7003), json={}).status_code == 200
    assert env.db.get_payment(pid_pay)["state"] == "cancelled"
    assert not env.db.get_product_by_id(env.PID).get("boosted_until"), "rad etilgach boost berilmasligi shart"
    # bekor qilingan to'lovni endi tasdiqlab bo'lmaydi
    assert env.post(f"/api/pay/dev-confirm/{pid_pay}", headers=hdr(7003), json={}).status_code == 409
    # to'langanni bekor qilib bo'lmaydi
    pid2 = env.post(f"/api/seller/boost/{env.PID}", headers=hdr(7002), json={}).json()["payment_id"]
    env.post(f"/api/pay/dev-confirm/{pid2}", headers=hdr(7003), json={})
    assert env.post(f"/api/pay/dev-cancel/{pid2}", headers=hdr(7003), json={}).status_code == 409


def test_admin_revokes_paid_payment_and_revenue_drops(env):
    """Soxta chek: admin TASDIQLANGAN (paid) boostni qaytarib oladi → boost olib tashlanadi,
    to'lov 'cancelled', platforma daromadidan ham chiqadi."""
    env.post("/api/admin/monetization", headers=hdr(7003),
             json={"mon_enabled": True, "mon_boost_enabled": True, "mon_boost_price": 5000})
    pay_id = env.post(f"/api/seller/boost/{env.PID}", headers=hdr(7002), json={}).json()["payment_id"]
    env.post(f"/api/pay/dev-confirm/{pay_id}", headers=hdr(7003), json={})
    # to'langach: boost bor + daromadda bor
    assert env.db.get_product_by_id(env.PID).get("boosted_until")
    fin1 = env.get("/api/admin/financial", headers=hdr(7003)).json()
    assert fin1.get("platform_revenue", 0) >= 5000
    # xaridor qaytara olmaydi
    assert env.post(f"/api/pay/dev-revoke/{pay_id}", headers=hdr(7001), json={}).status_code == 403
    # admin qaytaradi
    assert env.post(f"/api/pay/dev-revoke/{pay_id}", headers=hdr(7003), json={}).status_code == 200
    assert env.db.get_payment(pay_id)["state"] == "cancelled"
    assert not env.db.get_product_by_id(env.PID).get("boosted_until"), "qaytarilgach boost olib tashlanishi shart"
    # daromaddan chiqdi
    fin2 = env.get("/api/admin/financial", headers=hdr(7003)).json()
    assert fin2.get("platform_revenue", 0) == fin1["platform_revenue"] - 5000, "bekor summa daromaddan chiqishi shart"
    # pending to'lovni qaytarib bo'lmaydi (faqat paid)
    p2 = env.post(f"/api/seller/boost/{env.PID}", headers=hdr(7002), json={}).json()["payment_id"]
    assert env.post(f"/api/pay/dev-revoke/{p2}", headers=hdr(7003), json={}).status_code == 409


def test_admin_revokes_paid_subscription(env):
    """Pro obuna ham qaytariladi: pro_until olib tashlanadi."""
    env.post("/api/admin/monetization", headers=hdr(7003),
             json={"mon_enabled": True, "mon_subscription_enabled": True, "mon_subscription_price": 20000})
    r = env.post("/api/seller/subscribe", headers=hdr(7002), json={})
    if r.status_code != 200:
        import pytest as _pt
        _pt.skip("subscribe endpoint mavjud emas/o'chiq")
    pay_id = r.json()["payment_id"]
    env.post(f"/api/pay/dev-confirm/{pay_id}", headers=hdr(7003), json={})
    assert env.db.get_user_by_id(env.SELLER).get("pro_until")
    assert env.post(f"/api/pay/dev-revoke/{pay_id}", headers=hdr(7003), json={}).status_code == 200
    assert not env.db.get_user_by_id(env.SELLER).get("pro_until"), "qaytarilgach Pro olib tashlanishi shart"


def test_config_exposes_channel_url(env, monkeypatch):
    """App rasmiy kanal tugmasi uchun /api/config channel_url qaytaradi (bot pariteti)."""
    monkeypatch.setattr(webapp_server, "CHANNEL_URL", "https://t.me/TezBozorUz24")
    r = env.get("/api/config", headers=hdr(7001))
    assert r.status_code == 200 and r.json().get("channel_url") == "https://t.me/TezBozorUz24"


# ---------------- AUTH HIMOYASI ----------------
def test_auth_guards(env):
    assert env.get("/api/me").status_code == 401                      # imzosiz
    bad = {"Authorization": "tma user=%7B%22id%22%3A9%7D&hash=dead&auth_date=" + str(int(time.time()))}
    assert env.get("/api/me", headers=bad).status_code == 401         # soxta imzo
    assert env.get("/api/admin/stats", headers=hdr(7001)).status_code == 403   # xaridor→admin


# ================= PRO IMKONIYATLARI (#18 — 2026-06-21) =================
# Komissiya pasaytirish RAD ETILDI; o'rniga 6 imtiyoz Proga bog'landi:
# cheksiz mahsulot, bepul boost, Pro nishon, kengaytirilgan analitika,
# cheksiz reels, cheksiz rejalashtirilgan post, saralashda ustunlik.

def _enable_subscription(env, **extra):
    body = {"mon_enabled": True, "mon_subscription_enabled": True, "mon_subscription_price": 10000}
    body.update(extra)
    assert env.post("/api/admin/monetization", headers=hdr(7003), json=body).status_code == 200


def test_extended_analytics_pro_gated(env):
    # Obuna o'chiq → bepul (gate yo'q)
    assert env.get(f"/api/seller/product/{env.PID}/price-insight", headers=hdr(7002)).status_code == 200
    # Obuna yoqilgan, sotuvchi Pro emas → 403 pro_required
    _enable_subscription(env)
    r = env.get(f"/api/seller/product/{env.PID}/price-insight", headers=hdr(7002))
    assert r.status_code == 403 and r.json().get("detail") == "pro_required"
    # dynamic-price ham qulflanadi (AI'dan oldin)
    rd = env.get(f"/api/seller/product/{env.PID}/dynamic-price", headers=hdr(7002))
    assert rd.status_code == 403 and rd.json().get("detail") == "pro_required"
    # Pro bo'lgach gate ochiladi (pro_required EMAS)
    env.db.set_pro_until(env.SELLER, 30)
    assert env.get(f"/api/seller/product/{env.PID}/price-insight", headers=hdr(7002)).status_code == 200


def test_analytics_dashboard_pro_locked(env):
    _enable_subscription(env)
    j = env.get("/api/seller/analytics", headers=hdr(7002)).json()
    assert j.get("pro_locked") is True and j.get("by_weekday") is None
    env.db.set_pro_until(env.SELLER, 30)
    j2 = env.get("/api/seller/analytics", headers=hdr(7002)).json()
    assert j2.get("pro_locked") is False and j2.get("by_weekday") is not None


def test_pro_free_boost_quota(env):
    _enable_subscription(env, mon_boost_enabled=True, mon_boost_price=5000, mon_pro_free_boosts=2)
    env.db.set_pro_until(env.SELLER, 30)
    # 1-2 chi boost BEPUL (to'lovsiz, darhol qo'llanadi)
    r1 = env.post(f"/api/seller/boost/{env.PID}", headers=hdr(7002), json={})
    assert r1.status_code == 200 and r1.json().get("free") is True
    assert env.db.get_product_by_id(env.PID).get("boosted_until"), "bepul boost darhol qo'llanishi shart"
    r2 = env.post(f"/api/seller/boost/{env.PID}", headers=hdr(7002), json={})
    assert r2.json().get("free") is True and r2.json().get("remaining") == 0
    # 3-chi — kvota tugadi → pulli oqimga tushadi (payment_id, free yo'q)
    r3 = env.post(f"/api/seller/boost/{env.PID}", headers=hdr(7002), json={})
    assert r3.status_code == 200 and r3.json().get("payment_id") and not r3.json().get("free")


def test_pro_scheduled_limit(env):
    _enable_subscription(env, mon_free_scheduled_limit=1)
    future = "2030-06-01T10:00:00Z"
    # bepul sotuvchi: 1-rejalashtirish OK
    assert env.post(f"/api/seller/product/{env.PID}/schedule", headers=hdr(7002),
                    json={"scheduled_at": future}).status_code == 200
    pid2 = env.db.create_product(seller_id=env.SELLER, name="M2", price=5000, stock_count=3)
    env.db.update_product_fields(pid2, in_stock=1, status="active")
    # 2-rejalashtirish → limit (1) oshib ketdi → 403 pro_required
    r = env.post(f"/api/seller/product/{pid2}/schedule", headers=hdr(7002), json={"scheduled_at": future})
    assert r.status_code == 403 and r.json().get("detail") == "pro_required"
    # Pro bo'lsa cheksiz
    env.db.set_pro_until(env.SELLER, 30)
    assert env.post(f"/api/seller/product/{pid2}/schedule", headers=hdr(7002),
                    json={"scheduled_at": future}).status_code == 200


def test_pro_reels_quota(env, monkeypatch):
    monkeypatch.setattr(webapp_server.ai_assistant, "is_enabled", lambda: True)
    async def _vs(**k):
        return "SSENARIY"
    monkeypatch.setattr(webapp_server.ai_assistant, "generate_video_script", _vs)
    _enable_subscription(env, mon_free_reels_limit=1)
    # bepul sotuvchi: 1-reels OK, kvota sarflanadi
    assert env.get(f"/api/seller/product/{env.PID}/reels", headers=hdr(7002)).status_code == 200
    # 2-reels → 403 pro_required
    r = env.get(f"/api/seller/product/{env.PID}/reels", headers=hdr(7002))
    assert r.status_code == 403 and r.json().get("detail") == "pro_required"
    # Pro = cheksiz
    env.db.set_pro_until(env.SELLER, 30)
    assert env.get(f"/api/seller/product/{env.PID}/reels", headers=hdr(7002)).status_code == 200
    assert env.get(f"/api/seller/product/{env.PID}/reels", headers=hdr(7002)).status_code == 200


def test_pro_badge_in_listings(env):
    # Pro emas → nishon yo'q
    prods = rows(env.get("/api/products", headers=hdr(7001)).json())
    assert prods and prods[0].get("seller_is_pro") is False
    shops = rows(env.get("/api/shops", headers=hdr(7001)).json())
    assert shops and shops[0].get("is_pro") is False
    assert env.get(f"/api/products/{env.PID}", headers=hdr(7001)).json().get("seller_is_pro") is False
    # Pro bo'lgach nishon barcha ro'yxatlarda
    env.db.set_pro_until(env.SELLER, 30)
    assert rows(env.get("/api/products", headers=hdr(7001)).json())[0].get("seller_is_pro") is True
    assert rows(env.get("/api/shops", headers=hdr(7001)).json())[0].get("is_pro") is True
    assert env.get(f"/api/products/{env.PID}", headers=hdr(7001)).json().get("seller_is_pro") is True


# ============= CHAT YOPILISHI + SHARH FLAG + PRO EXCEL (2026-06-21) =============

def test_my_orders_has_review_flag(env):
    oid = _place_and_confirm(env, qty=1)
    env.post(f"/api/seller/order/{oid}/deliver", headers=hdr(7002),
             json={"settlement_type": "paid", "paid": 10000})
    o0 = next(x for x in rows(env.get("/api/my/orders", headers=hdr(7001)).json()) if x["id"] == oid)
    assert not o0.get("has_review"), "sharhdan oldin has_review bo'sh bo'lishi shart"
    env.post(f"/api/order/{oid}/review", headers=hdr(7001), json={"seller_rating": 5, "comment": "Zo'r"})
    o1 = next(x for x in rows(env.get("/api/my/orders", headers=hdr(7001)).json()) if x["id"] == oid)
    assert o1.get("has_review"), "sharhdan keyin has_review TRUE bo'lishi shart (qayta baholash yopiladi)"


def test_chat_closes_after_delivery(env):
    oid = _place_and_confirm(env, qty=1)
    # confirmed holatda suhbat OCHIQ — xaridor xabar yoza oladi
    assert env.post(f"/api/order/{oid}/message", headers=hdr(7001), json={"text": "Salom"}).status_code == 200
    m = env.get(f"/api/order/{oid}/messages", headers=hdr(7001)).json()
    assert m.get("closed") is False
    # yetkazilgach — suhbat YOPIQ
    env.post(f"/api/seller/order/{oid}/deliver", headers=hdr(7002),
             json={"settlement_type": "paid", "paid": 10000})
    m2 = env.get(f"/api/order/{oid}/messages", headers=hdr(7001)).json()
    assert m2.get("closed") is True and len(m2.get("messages") or []) >= 1, "tarix ko'rinadi, lekin yopiq"
    # ikkala taraf ham yangi xabar yoza olmaydi
    assert env.post(f"/api/order/{oid}/message", headers=hdr(7001), json={"text": "yana"}).status_code == 409
    assert env.post(f"/api/order/{oid}/message", headers=hdr(7002), json={"text": "yana"}).status_code == 409


def test_seller_excel_pro_gated(env, monkeypatch):
    async def _ok(*a, **k):
        return {"ok": True}
    monkeypatch.setattr(webapp_server, "_tg_send_document", _ok)
    # obuna o'chiq → Excel bepul ishlaydi
    r = env.post("/api/seller/export/products", headers=hdr(7002), json={})
    assert r.status_code == 200 and r.json().get("rows") >= 1
    # obuna yoqilgan, Pro emas → 403 pro_required
    _enable_subscription(env)
    assert env.post("/api/seller/export/products", headers=hdr(7002), json={}).status_code == 403
    # Pro bo'lgach yana ishlaydi
    env.db.set_pro_until(env.SELLER, 30)
    assert env.post("/api/seller/export/products", headers=hdr(7002), json={}).status_code == 200


def test_xlsx_report_pro_design():
    """Pro Excel generatori: brend sarlavha, header 3-qatorda, freeze, JAMI=SUM formula."""
    import io
    import openpyxl
    data = [[1, "Olma", 12000], [2, "Nok", 8000]]
    content, n = webapp_server._xlsx_report("Test hisoboti", ["ID", "Nom", "Narx"], data,
                                            money_cols=(2,), lang="uz")
    assert n == 2 and len(content) > 2000
    ws = openpyxl.load_workbook(io.BytesIO(content)).active
    assert "TezBozor" in str(ws["A1"].value)              # brend sarlavha
    assert ws["A3"].value == "ID"                          # header 3-qatorda (sarlavha+meta tepada)
    assert ws.freeze_panes == "A4"                         # header muzlatilgan
    assert str(ws.cell(row=6, column=3).value).startswith("=SUM(")  # JAMI = SUM formula


# ============= UNIVERSAL XABARNOMA + MUROJAAT (support) — 2026-06-21 =============

def test_contact_creates_thread_and_notifies_admin(env):
    r = env.post("/api/contact-admin", headers=hdr(7001),
                 json={"text": "Buyurtmam kelmadi", "reason": "order"})
    assert r.status_code == 200 and r.json().get("thread_id")
    tid = r.json()["thread_id"]
    # admin app ichida xabarnoma ko'radi (banner)
    nd = env.get("/api/my/notifications", headers=hdr(7003)).json()
    assert nd["unread"] >= 1 and any(n["kind"] == "support" and n["ref_id"] == tid for n in nd["items"])
    # admin murojaatlar ro'yxatида ko'radi
    sd = env.get("/api/support/threads", headers=hdr(7003)).json()
    assert sd["is_admin"] is True and sd["open_count"] >= 1
    assert any(t["id"] == tid for t in sd["threads"])


def test_support_thread_two_way_and_access(env):
    tid = env.post("/api/contact-admin", headers=hdr(7001),
                   json={"text": "Salom", "reason": "other"}).json()["thread_id"]
    # admin javob beradi → foydalanuvchi xabarnoma oladi
    assert env.post(f"/api/support/thread/{tid}/message", headers=hdr(7003),
                    json={"text": "Yordam beramiz"}).status_code == 200
    nd = env.get("/api/my/notifications", headers=hdr(7001)).json()
    assert any(n["kind"] == "support" and n["ref_id"] == tid for n in nd["items"])
    # foydalanuvchi javob yozadi
    assert env.post(f"/api/support/thread/{tid}/message", headers=hdr(7001),
                    json={"text": "Rahmat"}).status_code == 200
    th = env.get(f"/api/support/thread/{tid}", headers=hdr(7001)).json()
    assert len(th["messages"]) == 3   # boshlang'ich + admin + user
    # BEGONA foydalanuvchi ko'ra olmaydi
    env.db.create_user(telegram_id=7005, phone_number="998900000005", name="Begona", role="buyer")
    assert env.get(f"/api/support/thread/{tid}", headers=hdr(7005)).status_code == 403
    # yopish — faqat admin
    assert env.post(f"/api/support/thread/{tid}/close", headers=hdr(7001), json={}).status_code == 403
    assert env.post(f"/api/support/thread/{tid}/close", headers=hdr(7003), json={}).status_code == 200
    assert env.db.get_support_thread(tid)["status"] == "closed"


def test_notifications_read_flow(env):
    env.post("/api/contact-admin", headers=hdr(7001), json={"text": "Birinchi murojaat", "reason": "order"})
    env.post("/api/contact-admin", headers=hdr(7001), json={"text": "Ikkinchi murojaat", "reason": "payment"})
    before = env.get("/api/my/notifications", headers=hdr(7003)).json()
    assert before["unread"] >= 2
    nid = before["items"][0]["id"]
    r = env.post(f"/api/my/notifications/{nid}/read", headers=hdr(7003), json={})
    assert r.status_code == 200 and r.json()["unread"] == before["unread"] - 1
    env.post("/api/my/notifications/read-all", headers=hdr(7003), json={})
    assert env.get("/api/my/notifications", headers=hdr(7003)).json()["unread"] == 0


def test_pro_purchase_notifies_admin(env):
    import asyncio
    _enable_subscription(env)
    pid = env.db.create_payment(env.SELLER, "subscription", 10000)
    payment = dict(env.db.get_payment(pid))
    asyncio.run(webapp_server._notify_payment_done(payment))
    notifs = env.db.get_user_notifications(env.ADMIN)
    assert any(n["kind"] == "pro" for n in notifs), "admin Pro obuna haqida xabarnoma olishi shart"


def test_contact_validation_requires_reason_and_text(env):
    # sababsiz → 400 reason_required
    assert env.post("/api/contact-admin", headers=hdr(7001),
                    json={"text": "yordam kerak"}).status_code == 400
    # noto'g'ri sabab → 400
    assert env.post("/api/contact-admin", headers=hdr(7001),
                    json={"text": "yordam", "reason": "xxx"}).status_code == 400
    # faqat belgi/emoji → 400 bad_text
    r = env.post("/api/contact-admin", headers=hdr(7001),
                 json={"text": "!!! 😀😀", "reason": "order"})
    assert r.status_code == 400 and r.json()["detail"] == "bad_text"
    # to'g'ri matn+sabab → 200
    assert env.post("/api/contact-admin", headers=hdr(7001),
                    json={"text": "Buyurtmam kelmadi", "reason": "order"}).status_code == 200


def test_support_ai_reply_admin_only(env, monkeypatch):
    monkeypatch.setattr(webapp_server.ai_assistant, "is_enabled", lambda: True)
    async def _gen(**k):
        return "Assalomu alaykum, iltimos buyurtma raqamingizni yuboring."
    monkeypatch.setattr(webapp_server.ai_assistant, "generate_support_reply", _gen)
    tid = env.post("/api/contact-admin", headers=hdr(7001),
                   json={"text": "Buyurtmam kelmadi", "reason": "order"}).json()["thread_id"]
    # admin AI javob TAKLIFini oladi (yuborilmaydi — matn qaytadi)
    r = env.post(f"/api/support/thread/{tid}/ai-reply", headers=hdr(7003), json={})
    assert r.status_code == 200 and "raqam" in r.json()["reply"]
    # egasi bo'lsa ham foydalanuvchi AI-reply'дан foydalana olmaydi (faqat admin)
    assert env.post(f"/api/support/thread/{tid}/ai-reply", headers=hdr(7001), json={}).status_code == 403
    # AI o'chiq → 503
    monkeypatch.setattr(webapp_server.ai_assistant, "is_enabled", lambda: False)
    assert env.post(f"/api/support/thread/{tid}/ai-reply", headers=hdr(7003), json={}).status_code == 503


# ============= BUYURTMA/SHARH/NIZO → BANNER + EXCEL O'ZBEK/RUS (2026-06-21) =============

def test_order_events_create_banner_notifications(env):
    oid = env.post("/api/order", headers=hdr(7001),
                   json={"product_id": env.PID, "quantity": 1, "delivery_type": "pickup",
                         "payment_method": "cash"}).json()["order_id"]
    # yangi buyurtma → SOTUVCHIga banner
    assert any(n["kind"] == "order" for n in env.get("/api/my/notifications", headers=hdr(7002)).json()["items"])
    # tasdiqlash → XARIDORga banner
    env.post(f"/api/seller/order/{oid}/action", headers=hdr(7002), json={"action": "confirm"})
    assert any(n["kind"] == "order" for n in env.get("/api/my/notifications", headers=hdr(7001)).json()["items"])
    # yetkazish + sharh → SOTUVCHIga 'review' banner
    env.post(f"/api/seller/order/{oid}/deliver", headers=hdr(7002), json={"settlement_type": "paid", "paid": 10000})
    env.post(f"/api/order/{oid}/review", headers=hdr(7001), json={"seller_rating": 5, "comment": "Zo'r"})
    assert any(n["kind"] == "review" for n in env.get("/api/my/notifications", headers=hdr(7002)).json()["items"])


def test_excel_values_localized_no_english(env):
    import io
    import openpyxl
    _place_and_confirm(env, qty=1)   # status=confirmed, delivery=pickup
    content, fname, n = webapp_server._build_seller_excel(env.SELLER, "orders", "uz")
    ws = openpyxl.load_workbook(io.BytesIO(content)).active
    txt = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    for eng in ["confirmed", "pickup", "delivery", "pending", "cancelled"]:
        assert eng not in txt, f"Excel'da inglizcha '{eng}' qolmasligi kerak"
    assert "Tasdiqlangan" in txt and "Olib ketish" in txt
    # RU varianti — ruscha qiymatlar, inglizcha yo'q
    content_ru, _, _ = webapp_server._build_seller_excel(env.SELLER, "orders", "ru")
    ws2 = openpyxl.load_workbook(io.BytesIO(content_ru)).active
    txt2 = " ".join(str(c.value) for row in ws2.iter_rows() for c in row if c.value)
    assert "Подтверждён" in txt2 and "Самовывоз" in txt2 and "confirmed" not in txt2


# ============= BOT/APP PRO PARITETI — db helperlar + top_products qulfi (2026-06-21) =============

def test_db_pro_gating_helpers(env):
    oid = env.SELLER
    # monetizatsiya o'chiq → qulf yo'q, limit 0
    assert env.db.pro_locked(oid) is False
    assert env.db.mon_limit("mon_free_product_limit") == 0
    # obuna yoqilgan, Pro emas → qulf; limit o'qiladi
    _enable_subscription(env, mon_free_product_limit=3)
    assert env.db.pro_locked(oid) is True and env.db.is_pro(oid) is False
    assert env.db.mon_limit("mon_free_product_limit") == 3
    # Pro bo'lsa qulf yo'q
    env.db.set_pro_until(oid, 30)
    assert env.db.is_pro(oid) is True and env.db.pro_locked(oid) is False


def test_analytics_top_products_pro_locked(env):
    # "🏆 Eng ko'p sotilgan" endi Pro (bepul teaser emas)
    _enable_subscription(env)
    j = env.get("/api/seller/analytics", headers=hdr(7002)).json()
    assert j["pro_locked"] is True and j["top_products"] == []
    env.db.set_pro_until(env.SELLER, 30)
    assert env.get("/api/seller/analytics", headers=hdr(7002)).json()["pro_locked"] is False


def test_purchase_notifies_admin_with_contact(env):
    _enable_subscription(env)
    env.db.update_user(env.SELLER, telegram_username="seller_un")
    r = env.post("/api/seller/subscribe", headers=hdr(7002), json={})
    assert r.status_code == 200 and r.json().get("payment_id")
    pid = r.json()["payment_id"]
    # admin banner: 'pro' kind + kontakt (@username) body'da
    nd = env.get("/api/my/notifications", headers=hdr(7003)).json()
    pro = [n for n in nd["items"] if n["kind"] == "pro"]
    assert pro and "@seller_un" in (pro[0]["body"] or "")
    # admin to'lovlar — telegram_id + phone (bog'lanish uchun) qaytadi
    pays = env.get("/api/admin/payments?state=pending", headers=hdr(7003)).json()["payments"]
    p = next(x for x in pays if x["id"] == pid)
    assert p.get("telegram_id") == 7002 and p.get("phone_number")


def test_order_chat_creates_banner(env):
    oid = _place_and_confirm(env, qty=1)   # confirmed → chat ochiq
    # xaridor xabar → SOTUVCHIga banner
    env.post(f"/api/order/{oid}/message", headers=hdr(7001), json={"text": "Qachon tayyor?"})
    assert any(n["kind"] == "message" and n["ref_id"] == oid
               for n in env.get("/api/my/notifications", headers=hdr(7002)).json()["items"])
    # sotuvchi javob → XARIDORga banner
    env.post(f"/api/order/{oid}/message", headers=hdr(7002), json={"text": "Bugun tayyor"})
    assert any(n["kind"] == "message" and n["ref_id"] == oid
               for n in env.get("/api/my/notifications", headers=hdr(7001)).json()["items"])
