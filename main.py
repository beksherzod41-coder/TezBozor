import logging
import math
import html
import os
import io
import json
import asyncio

try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

TOKEN = os.getenv("BOT_TOKEN")
if not TOKEN:
    raise SystemExit(
        "❌ BOT_TOKEN topilmadi. Uni .env faylida belgilang:  BOT_TOKEN=...\n"
        "   (Token endi kodda saqlanmaydi — xavfsizlik uchun.)"
    )
# ADMIN_ID — faqat .env'dan o'qiladi. Kodda haqiqiy default qoldirmaymiz:
# aks holda sozlamasdan deploy qilinsa, o'sha ID egasi avtomatik admin bo'lib qoladi.
# O'rnatilmasa 0 (hech kim) — admin amallari hech kimga ochilmaydi.
try:
    ADMIN_ID = int(os.getenv("ADMIN_ID", "0") or "0")
except ValueError:
    ADMIN_ID = 0
if not ADMIN_ID:
    logging.warning("⚠️ ADMIN_ID .env'da o'rnatilmagan — admin huquqlari hech kimga berilmaydi.")

# Markaziy kanal — mahsulotlar avtomatik shu yerga post qilinadi.
# .env faylida belgilang:  CHANNEL_ID=@TezBozorUz24
CHANNEL_ID = os.getenv("CHANNEL_ID")

# Rasmiy kanal havolasi (panellardagi "kanalga o'tish" tugmasi uchun).
# @username bo'lsa — t.me havolasiga aylantiramiz.
CHANNEL_URL = f"https://t.me/{str(CHANNEL_ID).lstrip('@')}" if CHANNEL_ID and str(CHANNEL_ID).startswith('@') else None

# Mini App (Telegram WebApp) katalogi — HTTPS URL. .env'da belgilang:
#   MINIAPP_URL=https://tezbozor.duckdns.org
# O'rnatilmasa — Mini App tugmasi ko'rsatilmaydi (bot avvalgidek ishlayveradi).
MINIAPP_URL = os.getenv("MINIAPP_URL", "").strip() or None
if MINIAPP_URL and not MINIAPP_URL.startswith("https://"):
    logging.warning("⚠️ MINIAPP_URL https:// bilan boshlanishi shart (Telegram WebApp talabi) — o'chirildi.")
    MINIAPP_URL = None


def _product_buy_link(bot_username, product_id):
    """Kanal/guruh/ulashish "Sotib olish" havolasi — bosilganda MINI APP'ni ochadi (botni
    emas) va start_param=product_<id> uzatadi (app o'sha mahsulotni ochadi). startapp uchun
    bot'da "Main Mini App" sozlangan bo'lishi kerak (BotFather). MINIAPP_SHORT_NAME berilsa —
    nomli app. Mini App sozlanmagan bo'lsa — eski bot ?start= xulqiga qaytadi."""
    if MINIAPP_URL:
        short = os.getenv("MINIAPP_SHORT_NAME", "").strip()
        if short:
            return f"https://t.me/{bot_username}/{short}?startapp=product_{product_id}"
        return f"https://t.me/{bot_username}?startapp=product_{product_id}"
    return f"https://t.me/{bot_username}?start=product_{product_id}"


# Mini App'ni "Main Mini App" sifatida ochuvchi t.me havolasi (Desktop'da MINIMIZE tugmasini beradi).
# Inline web_app tugma modal ochadi (faqat X); startapp havolasi esa to'liq rejim → minimize bor.
# BotFather'da Main Mini App yoqilgan bo'lishi shart. Username hali ma'lum bo'lmasa → None (modal zaxira).
BOT_USERNAME = None   # _post_init'da get_me() bilan to'ldiriladi

def _app_deeplink(start_param=""):
    """Valid startapp deep-link qaytaradi, yoki None (username yo'q / belgi noto'g'ri → modal zaxira).
    Telegram start_param faqat [A-Za-z0-9_-] (≤64) qabul qiladi; aks holda tugma butunlay rad etiladi."""
    import re
    if not (MINIAPP_URL and BOT_USERNAME):
        return None
    sp = start_param or "app"
    if not re.fullmatch(r"[A-Za-z0-9_-]{1,64}", sp):
        return None
    short = os.getenv("MINIAPP_SHORT_NAME", "").strip()
    base = f"https://t.me/{BOT_USERNAME}/{short}" if short else f"https://t.me/{BOT_USERNAME}"
    return f"{base}?startapp={sp}"

def _open_app_button(lang, text_key, modal_url, start_param=""):
    """Ilovani ochuvchi inline tugma: imkon bo'lsa Main Mini App havolasi (minimize'li),
    aks holda web_app modal (zaxira). modal_url — startapp ishlamasa ishlatiladigan to'liq URL."""
    dl = _app_deeplink(start_param)
    if dl:
        return InlineKeyboardButton(t(lang, text_key), url=dl)
    return InlineKeyboardButton(t(lang, text_key), web_app=WebAppInfo(url=modal_url))

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, KeyboardButton, ReplyKeyboardMarkup, ReplyKeyboardRemove, InputMediaPhoto, Chat, WebAppInfo, MenuButtonWebApp, MenuButtonCommands
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, ContextTypes, ConversationHandler, PicklePersistence, ChatMemberHandler, TypeHandler, ApplicationHandlerStop
from telegram.error import Forbidden, BadRequest
from database import Database
from languages import t, LANGS, DEFAULT_LANG, get_user_lang, region_name, category_name, all_labels as _lang_labels
from tezbozor_design import (fmt_price, fmt_phone, fmt_order_id, fmt_status, fmt_rating,
                             fmt_datetime, is_shop_open_now, M, TZ_TASHKENT,
                             human_address, best_location_text, maps_link, looks_like_coords,
                             effective_unit_price, wholesale_info)
import ai_assistant
import ad_design
from telegram.constants import ChatAction


def price_with_unit(product):
    """Narx + (mavjud bo'lsa) o'lchov birligi: "5 000 so'm / kg". #20 — app/bot parite.
    Birlik bo'lmasa oddiy narx qaytadi (eski mahsulotlar buzilmaydi)."""
    s = fmt_price(product.get('price'))
    unit = (product.get('unit') or '').strip() if isinstance(product, dict) else ''
    if not unit and isinstance(product, dict) and product.get('sale_mode') == 'optom':
        unit = 'pachka'   # optom: 1 pachka narxi
    return f"{s} / {unit}" if unit else s

# ===== LOGGING + (ixtiyoriy) MONITORING =====
def _setup_logging():
    """Konsol + aylanuvchi fayl loglari. Xatolarni keyinchalik tahlil qilish uchun
    logs/bot.log faylida saqlanadi (har biri 5MB gacha, 5 ta zaxira nusxa)."""
    import logging.handlers
    root = logging.getLogger()
    root.setLevel(logging.INFO)
    fmt = logging.Formatter(
        "%(asctime)s | %(levelname)-8s | %(name)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    console = logging.StreamHandler()
    console.setFormatter(fmt)
    root.addHandler(console)
    try:
        os.makedirs("logs", exist_ok=True)
        fileh = logging.handlers.RotatingFileHandler(
            os.path.join("logs", "bot.log"), maxBytes=5 * 1024 * 1024,
            backupCount=5, encoding="utf-8")
        fileh.setFormatter(fmt)
        root.addHandler(fileh)
    except Exception as e:
        root.warning(f"Fayl logini yoqib bo'lmadi: {e}")
    # Kutubxonalar juda gapdon — biroz tinchlantiramiz
    logging.getLogger("httpx").setLevel(logging.WARNING)
    logging.getLogger("httpcore").setLevel(logging.WARNING)


_SENTRY_ENABLED = False


def _setup_sentry():
    """SENTRY_DSN o'rnatilgan bo'lsa — Sentry xato-kuzatuvini yoqadi. Aks holda hech
    narsa qilmaydi. sentry-sdk o'rnatilmagan bo'lsa ham bot to'xtamaydi."""
    global _SENTRY_ENABLED
    dsn = os.getenv("SENTRY_DSN")
    if not dsn:
        return
    try:
        import sentry_sdk
        sentry_sdk.init(dsn=dsn, traces_sample_rate=0.0,
                        environment=os.getenv("ENV", "production"))
        _SENTRY_ENABLED = True
        logging.info("✅ Sentry monitoring yoqildi.")
    except ImportError:
        logging.warning("SENTRY_DSN berilgan, lekin sentry-sdk o'rnatilmagan "
                        "(pip install sentry-sdk). Sentry o'chiq qoldi.")
    except Exception as e:
        logging.error(f"Sentry init xatosi: {e}")


_setup_logging()
_setup_sentry()

db = Database()


# ============================================================
# TIL (i18n) YORDAMCHILARI
# ============================================================
# Foydalanuvchi tilini ishonchli aniqlash: sessiya keshi -> DB -> default.
# T(update, context, 'kalit', ...) — joriy til bo'yicha matn qaytaradi.

def get_lang(update, context):
    """Joriy foydalanuvchi tili: context kesh -> DB -> 'uz'."""
    try:
        if context is not None:
            cached = context.user_data.get('lang')
            if cached in LANGS:
                return cached
    except Exception:
        pass
    try:
        u = db.get_user_by_telegram_id(update.effective_user.id)
        if u and (u.get('language') in LANGS):
            if context is not None:
                context.user_data['lang'] = u['language']
            return u['language']
    except Exception:
        pass
    return DEFAULT_LANG


def T(update, context, key, **kwargs):
    """Qisqartma: joriy til bo'yicha tarjima qaytaradi."""
    return t(get_lang(update, context), key, **kwargs)


def set_user_lang(update, context, lang):
    """Tilni tanlaydi: sessiya keshi + DB (foydalanuvchi mavjud bo'lsa)."""
    if lang not in LANGS:
        lang = DEFAULT_LANG
    if context is not None:
        context.user_data['lang'] = lang
    try:
        u = db.get_user_by_telegram_id(update.effective_user.id)
        if u:
            db.update_user(u['id'], language=lang)
    except Exception as e:
        logging.error(f"Til saqlanmadi: {e}")
    return lang


# Pastki (Reply) klaviatura tugmalari — har bir kalit ikkala tilda ham mavjud.
# Foydalanuvchi qaysi tilda yozsa ham, tugmani kanonik harakat (action) ga aylantiramiz.
_BOTTOM_BTN_KEYS = [
    'btn_search_menu', 'btn_search', 'btn_categories', 'btn_my_orders', 'btn_profile',
    'btn_add_product', 'btn_my_products', 'btn_orders',
    'btn_home', 'btn_contact_admin',
]


def bottom_action(text):
    """Pastki menyu tugma matni (uz yoki ru) -> kanonik kalit yoki None."""
    for key in _BOTTOM_BTN_KEYS:
        if text in _lang_labels(key):
            return key
    return None


def all_bottom_menu_texts():
    """Barcha tillardagi pastki menyu tugma matnlari (cancel_filter uchun)."""
    texts = []
    for key in _BOTTOM_BTN_KEYS:
        texts.extend(_lang_labels(key))
    return texts


def buyer_bottom_kb(lang):
    """Xaridor rejimi: ish App'ga ko'chirilgan — pastki klaviatura faqat launcher
    ('Ilovani ochish') tugmasini ko'rsatadi."""
    return app_launcher_kb(lang)


def seller_bottom_kb(lang):
    """Sotuvchi rejimi: ish App'ga ko'chirilgan — pastki klaviatura faqat launcher
    ('Ilovani ochish') tugmasini ko'rsatadi."""
    return app_launcher_kb(lang)


def app_launcher_kb(lang):
    """Launcher rejimi: pastki (reply) klaviatura BUTUNLAY olib tashlanadi.
    Eski web_app reply tugmasi xato berardi — endi faqat matn ostidagi inline
    tugma orqali ilovaga kiriladi (app_inline_kb)."""
    return ReplyKeyboardRemove()


def app_inline_kb(lang):
    """Matn ostida ko'rinadigan INLINE 'Ilovaga kirish' tugmasi (web_app)."""
    if MINIAPP_URL:
        return InlineKeyboardMarkup([[_open_app_button(lang, 'btn_open_app', MINIAPP_URL)]])
    return None


async def _go_to_app(update, context):
    """Launcher ekrani: chiroyli kirish matni + matn ostida inline 'Ilovaga kirish'
    tugmasi. Xaridor/sotuvchi botda boshqa hech narsa qilmaydi — faqat shu ekran.
    Callback ham, oddiy xabar ham qo'llanadi. ConversationHandler.END qaytaradi,
    shuning uchun conversation entry_point sifatida ham ishlatiladi (oqim boshlanmaydi)."""
    lang = get_lang(update, context)
    text = t(lang, 'open_app_hint')
    inline = app_inline_kb(lang)
    q = update.callback_query
    if q:
        try:
            await q.answer()
        except Exception:
            pass
        try:
            await q.edit_message_text(text, reply_markup=inline, parse_mode='HTML')
        except Exception:
            try:
                await q.message.reply_text(text, reply_markup=inline, parse_mode='HTML')
            except Exception:
                pass
        return ConversationHandler.END
    msg = getattr(update, 'message', None) or getattr(update, 'effective_message', None)
    if msg:
        # Eski pastki web_app reply tugmasini chatdan bir marta olib tashlaymiz
        # (u bosilganda xato berardi). chat_data /start clear()'dan omon qoladi.
        if not context.chat_data.get('kb_cleared'):
            try:
                await msg.reply_text("🚀", reply_markup=ReplyKeyboardRemove())
            except Exception:
                pass
            context.chat_data['kb_cleared'] = True
        sent = await msg.reply_text(text, reply_markup=inline, parse_mode='HTML')
        # Launcher xabarini chat tepasiga BIR MARTA pin qilamiz — doimiy "Ilovaga
        # kirish" tugmasi yuqorida turadi. disable_notification: ovozsiz.
        if inline is not None and not context.chat_data.get('launcher_pinned'):
            try:
                await context.bot.pin_chat_message(
                    chat_id=sent.chat_id, message_id=sent.message_id,
                    disable_notification=True)
                context.chat_data['launcher_pinned'] = True
            except Exception:
                pass
    return ConversationHandler.END


def haversine_km(lat1, lon1, lat2, lon2):
    """Ikki nuqta orasidagi taxminiy yo'l masofasi (km).
    Haversine (qush uchishi) + 1.35 koeffitsient = yo'l masofasi taxmini."""
    try:
        lat1, lon1, lat2, lon2 = float(lat1), float(lon1), float(lat2), float(lon2)
    except (TypeError, ValueError):
        return None
    R = 6371.0
    dLat = math.radians(lat2 - lat1)
    dLon = math.radians(lon2 - lon1)
    a = (math.sin(dLat / 2) ** 2
         + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dLon / 2) ** 2)
    straight = 2 * R * math.asin(math.sqrt(a))
    # Yo'l masofasi taxminan 35% uzoqroq (O'zbekiston yo'llari uchun)
    return round(straight * 1.35, 1)


# ============================================================
# XARIDOR LOKATSIYASINI ESLAB QOLISH (masofani izchil ko'rsatish uchun)
# ============================================================
# Muammo: masofa qidiruv ro'yxatida ko'rinardi, lekin mahsulot/ do'kon
# sahifasida yo'qolardi (lokatsiya saqlanmagani uchun). Xaridor lokatsiyani
# bir marta yuborsa — uni eslab qolamiz va hamma joyda masofani ko'rsatamiz.

def remember_buyer_geo(context, lat, lon):
    """Xaridor yuborgan lokatsiyani eslab qoladi (keyingi ekranlarda masofa uchun)."""
    try:
        if lat is None or lon is None:
            return
        context.user_data['buyer_geo'] = {'lat': float(lat), 'lon': float(lon)}
    except (TypeError, ValueError):
        pass


def get_buyer_geo(context):
    """Eslab qolingan xaridor lokatsiyasi -> (lat, lon) yoki (None, None)."""
    g = context.user_data.get('buyer_geo') or {}
    return g.get('lat'), g.get('lon')


def distance_line_to_shop(context, shop_lat, shop_lon, prefix="\n📏 Masofa: ~"):
    """Eslab qolingan xaridor lokatsiyasidan do'kongacha masofa qatori.
    Lokatsiya yoki koordinata bo'lmasa — bo'sh satr."""
    b_lat, b_lon = get_buyer_geo(context)
    if b_lat is None or b_lon is None or shop_lat is None or shop_lon is None:
        return ""
    d = haversine_km(b_lat, b_lon, shop_lat, shop_lon)
    if d is None:
        return ""
    return f"{prefix}{d:.1f} km"


# ============================================================
# REVERSE GEOCODING — koordinatani o'qiladigan manzilga aylantirish
# ============================================================
# Lokatsiya yuborilganda uni "Chilonzor t., Bunyodkor ko'chasi" kabi matnga
# aylantiramiz, shunda foydalanuvchi joyni tushunadi (xom raqamlar emas).
#
# Sukut bo'yicha bepul OpenStreetMap Nominatim ishlatiladi (kalit kerak emas).
# O'zbekiston uchun aniqroq natija xohlasangiz, pastdagi YANDEX bloki izohini
# oching va YANDEX_GEOCODER_KEY ni .env ga qo'shing.
#
# MUHIM: bu funksiya HECH QACHON xato tashlamaydi — tarmoq ishlamasa None
# qaytaradi va manzil ko'rsatuvi mo'ljal + xarita havolasiga "tushadi".

GEOCODER_ENABLED = True            # False qilsangiz — tarmoqqa umuman chiqmaydi
_GEO_CACHE = {}                    # {(lat5, lon5): "manzil matni"} — oddiy xotira keshi


async def reverse_geocode(lat, lon):
    """Koordinata -> o'qiladigan qisqa manzil (eng yaxshi harakat). Aks holda None."""
    if not GEOCODER_ENABLED or lat is None or lon is None:
        return None
    try:
        key = (round(float(lat), 5), round(float(lon), 5))
    except (TypeError, ValueError):
        return None
    if key in _GEO_CACHE:
        return _GEO_CACHE[key]

    text = None
    try:
        import httpx  # python-telegram-bot bilan birga keladi

        # --- YANDEX (ixtiyoriy, aniqroq) -------------------------------------
        # yandex_key = os.getenv("YANDEX_GEOCODER_KEY")
        # if yandex_key:
        #     async with httpx.AsyncClient(timeout=6.0) as client:
        #         r = await client.get(
        #             "https://geocode-maps.yandex.ru/1.x/",
        #             params={"apikey": yandex_key, "format": "json",
        #                     "geocode": f"{key[1]},{key[0]}", "lang": "uz_UZ"},
        #         )
        #         if r.status_code == 200:
        #             fm = (r.json()["response"]["GeoObjectCollection"]
        #                   ["featureMember"])
        #             if fm:
        #                 text = fm[0]["GeoObject"]["metaDataProperty"] \
        #                     ["GeocoderMetaData"]["text"]
        # ---------------------------------------------------------------------

        if not text:
            # OpenStreetMap Nominatim (bepul, kalitsiz)
            async with httpx.AsyncClient(timeout=6.0) as client:
                r = await client.get(
                    "https://nominatim.openstreetmap.org/reverse",
                    params={
                        "lat": key[0], "lon": key[1], "format": "jsonv2",
                        "accept-language": "uz,ru,en",
                        "zoom": "18", "addressdetails": "1",
                    },
                    headers={"User-Agent": "TezBozorBot/1.0 (Telegram marketplace)"},
                )
                if r.status_code == 200:
                    data = r.json()
                    addr = data.get("address") or {}
                    road = (addr.get("road") or addr.get("pedestrian")
                            or addr.get("residential"))
                    house = addr.get("house_number")
                    district = (addr.get("city_district") or addr.get("suburb")
                                or addr.get("county") or addr.get("town")
                                or addr.get("village"))
                    city = addr.get("city") or addr.get("state")
                    parts = []
                    if road:
                        parts.append(f"{road} {house}".strip() if house else road)
                    if district:
                        parts.append(district)
                    if city and city not in parts:
                        parts.append(city)
                    text = ", ".join(p for p in parts if p) or data.get("display_name")
    except Exception as e:
        logging.warning(f"reverse_geocode xatosi: {e}")
        text = None

    if text:
        text = text.strip()[:200]
        _GEO_CACHE[key] = text
    return text


async def resolve_shop_address(lat, lon):
    """Lokatsiyadan saqlash uchun manzil matnini qaytaradi.
    Geocoding ishlasa — o'qiladigan manzil; aks holda None (xom koordinata SAQLANMAYDI)."""
    return await reverse_geocode(lat, lon)


import time as _time

# ============================================================
# SPAM / FLOOD HIMOYASI
# ============================================================

# Global flood tracker — bir foydalanuvchi juda ko'p so'rov yuborsa bloklash
_flood_tracker = {}  # {user_id: [timestamp, timestamp, ...]}
FLOOD_WINDOW = 10       # 10 soniya ichida
FLOOD_MAX_REQUESTS = 15  # 15 tadan ko'p so'rov = flood
FLOOD_BAN_DURATION = 60  # 60 soniya ban


def is_flood(user_id: int) -> bool:
    """Foydalanuvchi flood qilyaptimi? True bo'lsa — bloklash kerak."""
    now = _time.monotonic()
    if user_id not in _flood_tracker:
        _flood_tracker[user_id] = []

    # Eski yozuvlarni tozalaymiz
    _flood_tracker[user_id] = [t for t in _flood_tracker[user_id] if now - t < FLOOD_WINDOW]
    _flood_tracker[user_id].append(now)

    return len(_flood_tracker[user_id]) > FLOOD_MAX_REQUESTS


# Flood ban ro'yxati: {user_id: ban_expire_time}
_flood_bans = {}


def check_flood_ban(user_id: int) -> bool:
    """Foydalanuvchi flood ban da bo'lsa True qaytaradi."""
    now = _time.monotonic()
    if user_id in _flood_bans:
        if now < _flood_bans[user_id]:
            return True
        else:
            del _flood_bans[user_id]
    return False


def apply_flood_ban(user_id: int):
    """Foydalanuvchini vaqtincha ban qiladi."""
    _flood_bans[user_id] = _time.monotonic() + FLOOD_BAN_DURATION


def rate_limit(min_interval: float = 1.0):
    """Decorator: foydalanuvchi min_interval soniyadan tez-tez so'rov yuborsa, bloklanadi.
    Faqat CallbackQuery va Message handler'larda ishlaydi."""
    def decorator(func):
        import functools
        @functools.wraps(func)
        async def wrapper(update, context, *args, **kwargs):
            uid = update.effective_user.id if update.effective_user else None
            if uid:
                # Flood ban tekshiruvi
                if check_flood_ban(uid):
                    if update.callback_query:
                        try:
                            await update.callback_query.answer(T(update, context, 'flood_ban'), show_alert=True)
                        except Exception:
                            pass
                    return

                # Flood detection
                if is_flood(uid):
                    apply_flood_ban(uid)
                    if update.callback_query:
                        try:
                            await update.callback_query.answer(T(update, context, 'flood_detected'), show_alert=True)
                        except Exception:
                            pass
                    elif update.message:
                        try:
                            await update.message.reply_text(T(update, context, 'flood_too_many'))
                        except Exception:
                            pass
                    return

                key = f'_rl_{func.__name__}'
                last = context.user_data.get(key, 0)
                now = _time.monotonic()
                if now - last < min_interval:
                    # Telegram'ga "loading" ko'rinmasligi uchun callback_query'ga javob beramiz
                    if update.callback_query:
                        try:
                            await update.callback_query.answer(T(update, context, 'please_wait'))
                        except Exception:
                            pass
                    return
                context.user_data[key] = now
            return await func(update, context, *args, **kwargs)
        return wrapper
    return decorator


# Buyurtma spam oldini olish: bir foydalanuvchi bir mahsulotga 5 daqiqa ichida qayta buyurtma bera olmaydi
ORDER_COOLDOWN = 300  # 5 daqiqa


def check_order_spam(context, buyer_id: int, product_id: int) -> bool:
    """True bo'lsa — spam, buyurtma berish mumkin emas."""
    key = f'_order_cd_{buyer_id}_{product_id}'
    last = context.bot_data.get(key, 0)
    now = _time.monotonic()
    if now - last < ORDER_COOLDOWN:
        return True
    return False


def mark_order_placed(context, buyer_id: int, product_id: int):
    """Buyurtma berilganini belgilaymiz."""
    key = f'_order_cd_{buyer_id}_{product_id}'
    context.bot_data[key] = _time.monotonic()


def is_seller_capable(user):
    """Bir foydalanuvchi sotuvchi bo'lib ishlay oladimi? Do'kon nomi to'ldirilgan bo'lsa — ha."""
    return bool(user and user.get('shop_name'))


def get_active_mode(user, context):
    """Foydalanuvchi hozir qaysi rejimda ishlayapti: 'admin' | 'seller' | 'buyer'.
    Admin har doim admin. Boshqalar — sessiyadagi active_mode yoki role bo'yicha default."""
    if not user:
        return 'buyer'
    if user['role'] == 'admin':
        return 'admin'
    mode = context.user_data.get('active_mode')
    # Saqlangan 'seller' rejimi — faqat haqiqatan sotuvchi bo'lsa amal qiladi (do'koni bor
    # yoki roli seller). Aks holda (sotuvchilik bekor qilingan) xavfsiz xaridorga tushadi.
    if mode == 'seller' and not (user['role'] == 'seller' or is_seller_capable(user)):
        return 'buyer'
    return mode or user['role']


import re as _re


def normalize_phone(raw):
    """Telefonni standartlashtirish. Yaroqsiz bo'lsa None qaytaradi.
    Misol: '901234567' → '+998901234567'; '+998 90 123 45 67' → '+998901234567'"""
    if not raw:
        return None
    digits = _re.sub(r'\D', '', str(raw))
    if len(digits) == 9:
        digits = '998' + digits
    if len(digits) == 12 and digits.startswith('998'):
        return '+' + digits
    # Boshqa mamlakat raqamlari uchun ham bo'sh joy qoldiramiz: 10-15 raqam
    if 10 <= len(digits) <= 15:
        return '+' + digits
    return None


def normalize_telegram_username(raw):
    """Telegram username'ni standartlashtirish. Yaroqsiz bo'lsa None.
    '@user' yoki 'user' → 'user'. Faqat lotin harf/raqam/_; 5–32 belgi."""
    if not raw:
        return None
    u = str(raw).strip().lstrip('@')
    if 5 <= len(u) <= 32 and _re.match(r'^[A-Za-z][A-Za-z0-9_]{4,31}$', u):
        return u
    return None


def normalize_name(raw, min_len=2, max_len=50):
    """Ismni standartlashtirish. Bo'sh joylarni qisqartiradi."""
    if not raw:
        return None
    n = ' '.join(str(raw).split())
    if min_len <= len(n) <= max_len:
        return n
    return None


# Ruxsat etilgan ism belgilari: lotin/kirill harflar, bo'shliq, apostrof, defis, nuqta
_NAME_RE = _re.compile(r"^[A-Za-zА-Яа-яЁёЎўҚқҒғҲҳ'’‘ʻ.\- ]+$")


def validate_fullname(raw, max_len=60):
    """To'liq F.I.SH validatsiyasi.
    Talab: kamida 2 so'z (Familiya + Ism), har biri ≥2 harf, faqat harflar,
    umumiy uzunlik 5–{max_len}. Yaroqsiz bo'lsa None qaytaradi."""
    if not raw:
        return None
    n = ' '.join(str(raw).split())
    if not (5 <= len(n) <= max_len):
        return None
    if not _NAME_RE.match(n):
        return None
    # Kamida 2 ta mazmunli so'z (har biri 2+ harf)
    words = [w for w in n.split() if len(w) >= 2]
    if len(words) < 2:
        return None
    return n

# ============================================================
# CONVERSATION STATES
# Har bir ConversationHandler o'z alohida state raqamlarini ishlatishi kerak
# Aks holda state'lar bir-birini ustidan yozib ketadi (BUG FIX #1)
# ============================================================
(PHONE, NAME, ROLE, SELLER_CATEGORY, SHOP_NAME, SHOP_LANDMARK,
 SHOP_ADDRESS, SHOP_ADDRESS_TEXT, WORKING_DAYS, WORKING_HOURS, TELEGRAM_USERNAME) = range(11)

# Til tanlash (ro'yxatdan o'tishdan oldin) — alohida raqam, boshqa banddagilar bilan to'qnashmaydi
SELECT_LANG = 200

(PRODUCT_NAME, PRODUCT_PRICE, PRODUCT_STOCK, PRODUCT_CATEGORY, PRODUCT_DESC, PRODUCT_PHOTO, PRODUCT_ATTRS) = range(10, 17)
PRODUCT_MODE = 17   # mahsulot joylash usulini tanlash (klassik / AI savollar / AI aqlli)

# Mahsulotni tahrirlash — endi "bir oyna + qaysi qismni tanlash" usulida.
# Har bir maydon alohida fokuslangan tahrir holatiga ega.
(EDIT_FIELD_NAME, EDIT_FIELD_PRICE, EDIT_FIELD_CATEGORY,
 EDIT_FIELD_DESC, EDIT_FIELD_PHOTOS, EDIT_FIELD_ATTR) = range(20, 26)

(ORDER_QUANTITY, ORDER_DELIVERY_TYPE, ORDER_ADDRESS, ORDER_PAYMENT, ORDER_CONFIRM) = range(30, 35)

MESSAGE_TEXT = 40

(PRODUCT_RATING, PRODUCT_COMMENT, SELLER_RATING) = range(50, 53)

(EDIT_PROFILE_NAME, EDIT_PROFILE_PHONE,
 EDIT_SHOP_NAME, EDIT_SHOP_LANDMARK, EDIT_SHOP_ADDRESS, EDIT_SHOP_ADDRESS_TEXT,
 EDIT_WORKING_DAYS, EDIT_WORKING_HOURS, EDIT_TELEGRAM_USERNAME) = range(60, 69)

(EDIT_CARD_TYPE, EDIT_CARD_NUMBER, EDIT_CARD_OWNER) = range(70, 73)

(SELLER_PRODUCT_SEARCH,) = range(80, 81)
(CONTACT_ADMIN_MSG,) = range(90, 91)

# Savat (cart) rasmiylashtirish oqimi — yakka buyurtma oqimidan ALOHIDA holatlar,
# shunda mavjud order_conv'ga umuman tegmaymiz.
(CART_DELIVERY_TYPE, CART_ADDRESS, CART_PAYMENT, CART_CONFIRM) = range(100, 104)
LINK_CHANNEL_WAIT = 110  # Sotuvchi kanalini ulash holati (forward kutilmoqda)

# Shartnomani bekor qilish oqimi (kelishuv + nizo) — sabab tanlash holatlari
(CANCEL_PICK_REASON, CANCEL_REASON_TEXT) = range(120, 122)

# Admin nizo bo'yicha tomonga (bot orqali) xabar yozishi
ADMIN_DISPUTE_MSG = 130
# Xaridor/sotuvchi admin xabariga javob berishi
DMREPLY_MSG = 131


# ============================================================
# START & REGISTRATION
# ============================================================

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = db.get_user_by_telegram_id(update.effective_user.id)

    # Referral kod tekshiruvi — faqat yangi foydalanuvchi uchun
    # /start REF12345 ko'rinishida kelishi mumkin (context.args ichida)
    if not user and context.args:
        ref_code = context.args[0].strip()
        # Deeplink emas — referral (product_/order_/staff_ deeplinklarini chiqarib tashlaymiz)
        if not ref_code.startswith(("product_", "order_", "staff_")):
            referrer = db.get_user_by_referral_code(ref_code)
            if referrer:
                context.user_data['referred_by'] = referrer['id']
                # App ro'yxatiga uzatish uchun xom kodni ham saqlaymiz (?ref=...)
                context.user_data['ref_code'] = ref_code
                logging.info(f"New user referred by {referrer['name']} (code={ref_code})")

    # Deeplink: /start product_123 — mahsulot sahifasiga o'tish.
    # /start order_123 — Mini App'dagi "Sotib olish" tugmasidan keladi (productId).
    # Ikkalasi ham mahsulot sahifasini ochadi (u yerda buyurtma tugmasi bor).
    if user and context.args:
        arg = context.args[0].strip()
        if arg.startswith("product_") or arg.startswith("order_"):
            try:
                product_id = int(arg.split("_", 1)[1])
            except (ValueError, TypeError):
                product_id = None
            if product_id:
                # APP-ONLY: mahsulotni botda emas, to'g'ridan-to'g'ri ilovada ochamiz.
                # Inline web_app tugma ?product=<id> bilan — ilova o'sha sahifani ochadi.
                clang = get_user_lang(user)
                if MINIAPP_URL:
                    url = MINIAPP_URL
                    sep = '&' if '?' in url else '?'
                    url = f"{url}{sep}product={product_id}"
                    await update.message.reply_text(
                        t(clang, 'open_app_hint'),
                        reply_markup=InlineKeyboardMarkup([[_open_app_button(
                            clang, 'btn_open_app', url, f"product_{product_id}")]]),
                        parse_mode='HTML'
                    )
                    return ConversationHandler.END
                # MINIAPP_URL yo'q (zaxira) — eski botdagi ko'rinish
                product = db.get_product_by_id(product_id)
                if product and product.get('in_stock'):
                    context.user_data['active_mode'] = 'buyer'
                    await _show_product_deeplink(update, context, product)
                else:
                    await update.message.reply_text(
                        T(update, context, 'deeplink_product_unavailable'))
                return ConversationHandler.END

    # Deeplink: /start staff_<code> — do'konga sotuvchi-xodim bo'lib qo'shilish
    if context.args and context.args[0].strip().startswith("staff_"):
        code = context.args[0].strip()[len("staff_"):]
        handled = await _handle_staff_deeplink(update, context, code, user)
        if handled:
            return ConversationHandler.END
        # handled=False → yangi foydalanuvchi, staff_invite saqlandi, ro'yxat davom etadi

    # Deeplink: /start contact_<telegram_id> — ADMIN foydalanuvchi bilan bog'lanadi.
    # Bot tg://user?id= mention'ini yuboradi → Telegram uni hal qiladi (user bot bilan
    # muloqotda bo'lgani uchun ishlaydi), admin bosib SHAXSIY chatni ochadi. Faqat admin.
    if user and context.args and context.args[0].strip().startswith("contact_"):
        clang = get_user_lang(user)
        is_admin = user.get('role') == 'admin' or update.effective_user.id == ADMIN_ID
        if not is_admin:
            await update.message.reply_text(t(clang, 'contact_admin_only'))
            return ConversationHandler.END
        try:
            target_tg = int(context.args[0].strip()[len("contact_"):])
        except (ValueError, TypeError):
            target_tg = None
        target = db.get_user_by_telegram_id(target_tg) if target_tg else None
        if not target:
            await update.message.reply_text(t(clang, 'contact_user_not_found'))
            return ConversationHandler.END
        lines = [f"👤 <b>{html.escape(target.get('name') or '—')}</b>"]
        if target.get('shop_name'):
            lines.append(f"🏪 {html.escape(target['shop_name'])}")
        if target.get('telegram_username'):
            lines.append(f"🔗 @{html.escape(str(target['telegram_username']).lstrip('@'))}")
        if target.get('phone_number'):
            lines.append(f"📞 {html.escape(target['phone_number'])}")
        # Bosiladigan mention — shaxsiy chatni ochadi (user botda bo'lgani uchun resolve bo'ladi)
        lines.append(f'\n<a href="tg://user?id={target_tg}">💬 {t(clang, "contact_open_chat")}</a>')
        await update.message.reply_text("\n".join(lines), parse_mode='HTML')
        return ConversationHandler.END

    if user and user['role'] != 'admin' and update.effective_user.id == ADMIN_ID:
        db.update_user(user['id'], role='admin')
        user['role'] = 'admin'
        await update.message.reply_text(T(update, context, 'you_are_admin'))

    if user:
        if user['is_blocked']:
            await update.message.reply_text(t(user, 'blocked'))
            return ConversationHandler.END

        # Avvalgi conversation state'ni tozalaymiz (til + tanlangan rejimni saqlab qolamiz).
        # active_mode'ni asraymiz — foydalanuvchi o'zi almashtirmaguncha oxirgi panelda
        # qoladi (app bilan bir xil). PicklePersistence buni qayta ishga tushishda ham tiklaydi.
        _saved_mode = context.user_data.get('active_mode')
        context.user_data.clear()
        lang = get_user_lang(user)
        context.user_data['lang'] = lang
        if _saved_mode:
            context.user_data['active_mode'] = _saved_mode

        # ReplyKeyboard ni yangilaymiz — har doim to'g'ri tugmalar ko'rinsin
        active = get_active_mode(user, context)
        if user['role'] == 'admin':
            await admin_panel(update, context)
        else:
            # APP-ONLY (Faza 1+2): xaridor va sotuvchi ishi butunlay ilovaga ko'chirilgan.
            # Bot faqat launcher — chiroyli matn + matn ostida inline 'Ilovaga kirish'
            # tugmasi (pastki reply klaviatura olib tashlanadi).
            await _go_to_app(update, context)
        return ConversationHandler.END
    else:
        # APP-ONLY: ro'yxatdan o'tish botdan olib tashlangan — faqat ilovada.
        lang = get_lang(update, context)
        if MINIAPP_URL:
            url = MINIAPP_URL
            ref = context.user_data.get('ref_code')
            if ref:
                sep = '&' if '?' in url else '?'
                url = f"{url}{sep}ref={quote(ref, safe='')}"
            await update.message.reply_text(
                t(lang, 'reg_app_welcome'),
                reply_markup=InlineKeyboardMarkup([[_open_app_button(
                    lang, 'reg_app_btn', url, (f"ref_{ref}" if ref else ""))]]),
                parse_mode='HTML'
            )
            return ConversationHandler.END
        # MINIAPP_URL yo'q (faqat zaxira) — eski ichki ro'yxat
        return await registration_start(update, context)


async def admin_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = db.get_user_by_telegram_id(update.effective_user.id)

    if not user:
        await update.message.reply_text(T(update, context, 'not_registered'))
        return ConversationHandler.END

    if user['role'] != 'admin' and update.effective_user.id == ADMIN_ID:
        db.update_user(user['id'], role='admin')
        user['role'] = 'admin'

    if user['role'] != 'admin':
        await update.message.reply_text(t(user, 'not_admin'))
        return ConversationHandler.END

    # Conversation state'ni tozalaymiz — /admin dan keyin bot ro'yxat jarayonida qolmasin
    context.user_data.clear()

    await admin_panel(update, context)
    return ConversationHandler.END


async def _notify_owner_new_staff(context, shop, staff_user, staff_id, department=None):
    """Do'kon egasiga yangi xodim qo'shilgani haqida xabar (tasdiqlash uchun)."""
    try:
        owner = db.get_user_by_id(shop['owner_user_id'])
        if not owner or not owner.get('telegram_id'):
            return
        olang = get_user_lang(owner)
        await context.bot.send_message(
            chat_id=owner['telegram_id'],
            text=t(olang, 'owner_new_staff_notify',
                   name=html.escape(staff_user.get('name') or '—'),
                   phone=staff_user.get('phone_number') or '—',
                   dept=html.escape(department or '—')),
            parse_mode='HTML',
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(t(olang, 'btn_staff_activate'), callback_data=f"staff_toggle_{staff_id}")],
                [InlineKeyboardButton(t(olang, 'btn_staff_reject'), callback_data=f"staff_reject_{staff_id}")],
                [InlineKeyboardButton(t(olang, 'btn_manage_staff'), callback_data="staff_detail_%d" % staff_id)],
            ])
        )
    except Exception as e:
        logging.error(f"Egaga yangi xodim xabari ketmadi: {e}")


async def _handle_staff_deeplink(update, context, code, user):
    """staff_<code> deeplink. To'liq bajarilsa True (to'xta), yangi foydalanuvchi uchun
    ro'yxat davom etishi kerak bo'lsa False qaytaradi."""
    invite = db.get_invite_by_code(code)
    lang = (get_user_lang(user) if user else context.user_data.get('lang', DEFAULT_LANG))
    if not invite or invite.get('is_used'):
        await update.message.reply_text(t(lang, 'staff_invite_invalid'))
        return True
    shop = db.get_shop_by_id(invite['shop_id'])
    if not shop:
        await update.message.reply_text(t(lang, 'staff_invite_invalid'))
        return True

    # APP-FIRST (2-bosqich): MINIAPP_URL bo'lsa, do'konga bog'lash ilovada bajariladi
    # (/api/join-with-code). Bot faqat ilovani ?staff=<kod> bilan ochuvchi tugma yuboradi.
    # web_app tugmasi start_param uzatmaydi — shuning uchun kodni URL query'da beramiz.
    if MINIAPP_URL:
        from urllib.parse import quote
        sep = '&' if '?' in MINIAPP_URL else '?'
        url = f"{MINIAPP_URL}{sep}staff={quote(code, safe='')}"
        await update.message.reply_text(
            t(lang, 'staff_invite_app_prompt', shop=html.escape(shop.get('name') or '—')),
            reply_markup=InlineKeyboardMarkup([[_open_app_button(
                lang, 'staff_invite_app_btn', url, f"staff_{code}")]]),
            parse_mode='HTML')
        return True

    if user:
        if user.get('role') == 'admin':
            await update.message.reply_text(t(lang, 'staff_admin_cannot_join'))
            return True
        existing = db.get_staff_by_user(user['id'])
        if existing:
            # Ega o'z do'konini tashlab keta olmaydi
            if existing.get('staff_role') == 'owner':
                await update.message.reply_text(t(lang, 'staff_owner_cannot_join'))
                return True
            # Bu do'konning o'ziga qayta urinish — qayta qo'shishning hojati yo'q
            if existing.get('shop_id') == shop['id']:
                await update.message.reply_text(t(lang, 'staff_already_in_this_shop'))
                return True
            # Xodim — eski do'kondan chiqarib, yangisiga o'tkazamiz
            old_shop = db.get_shop_by_id(existing['shop_id'])
            db.remove_staff(existing['id'])
            # Eski do'kon egasiga xabar
            try:
                if old_shop:
                    old_owner = db.get_user_by_id(old_shop['owner_user_id'])
                    if old_owner and old_owner.get('telegram_id'):
                        await context.bot.send_message(
                            chat_id=old_owner['telegram_id'],
                            text=t(get_user_lang(old_owner), 'staff_left_old_shop',
                                   name=html.escape(user.get('name') or '—')),
                            parse_mode='HTML')
            except Exception as e:
                logging.error(f"Eski egaga xodim chiqdi xabari ketmadi: {e}")
        staff_id = db.add_staff(shop['id'], user['id'], staff_role='staff',
                                department=invite.get('department'), is_active=0,
                                added_by=invite.get('created_by'))
        db.update_user(user['id'], role='seller', is_approved=1)
        db.mark_invite_used(code, user['id'])
        await _notify_owner_new_staff(context, shop, user, staff_id, invite.get('department'))
        await update.message.reply_text(
            t(lang, 'staff_joined_pending', shop=html.escape(shop.get('name') or '—')),
            parse_mode='HTML')
        return True

    # Yangi foydalanuvchi — kodni saqlab, ro'yxatdan o'tishni davom ettiramiz
    context.user_data['staff_invite'] = code
    return False


async def _finalize_staff_registration(update, context):
    """Yangi foydalanuvchi ro'yxatdan o'tib bo'lgach — xodim sifatida do'konga bog'laydi."""
    code = context.user_data.get('staff_invite')
    lang = context.user_data.get('lang', DEFAULT_LANG)
    invite = db.get_invite_by_code(code) if code else None
    context.user_data.pop('staff_invite', None)
    if not invite or invite.get('is_used'):
        await update.message.reply_text(t(lang, 'staff_invite_invalid'))
        # Oddiy xaridor sifatida ro'yxatdan o'tkazamiz (kod yaroqsiz bo'lsa ham yo'qotmaymiz)
        uid = db.create_user(telegram_id=update.effective_user.id,
                             phone_number=context.user_data.get('phone'),
                             name=context.user_data.get('name'), role='buyer')
        db.update_user(uid, language=lang)
        await buyer_panel(update, context)
        return ConversationHandler.END
    shop = db.get_shop_by_id(invite['shop_id'])
    uid = db.create_user(telegram_id=update.effective_user.id,
                         phone_number=context.user_data.get('phone'),
                         name=context.user_data.get('name'), role='seller')
    db.update_user(uid, language=lang, is_approved=1)
    staff_id = db.add_staff(shop['id'], uid, staff_role='staff',
                            department=invite.get('department'), is_active=0,
                            added_by=invite.get('created_by'))
    db.mark_invite_used(code, uid)
    staff_user = db.get_user_by_id(uid)
    await _notify_owner_new_staff(context, shop, staff_user, staff_id, invite.get('department'))
    await notify_admins(
        context,
        f"👥 <b>Do'konga yangi sotuvchi-xodim qo'shildi!</b>\n\n"
        f"Ism: {html.escape(context.user_data.get('name') or '')}\n"
        f"Telefon: {context.user_data.get('phone') or ''}\n"
        f"Do'kon: {html.escape(shop.get('name') or '—')}"
    )
    await update.message.reply_text(
        t(lang, 'staff_joined_pending', shop=html.escape(shop.get('name') or '—')),
        parse_mode='HTML')
    return ConversationHandler.END


async def registration_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Ro'yxatdan o'tishning birinchi qadami.

    APP-FIRST: MINIAPP_URL bo'lsa, yangi foydalanuvchi to'g'ridan-to'g'ri Mini App'ga
    yo'naltiriladi — ro'yxatdan o'tish, rol tanlash (xaridor/sotuvchi) va xarid
    hammasi ilovada bo'ladi. Faqat istisnolar bot ichida qoladi:
      • staff_<kod> taklifi — MINIAPP_URL bo'lsa _handle_staff_deeplink allaqachon
        ilovaga uzatgan; bu yerga faqat MINIAPP_URL yo'q paytda staff_invite bilan keladi;
      • MINIAPP_URL o'rnatilmagan — eski ichki FSM (til→telefon→ism→rol) fallback."""
    has_staff_invite = bool(context.user_data.get('staff_invite'))
    if MINIAPP_URL and not has_staff_invite:
        # Telegram tilidan boshlang'ich til (DB hali yo'q) — ilova ichida o'zgartirsa bo'ladi
        lc = (update.effective_user.language_code or '').lower()
        lang = 'ru' if lc.startswith('ru') else DEFAULT_LANG
        url = MINIAPP_URL
        ref_code = context.user_data.get('ref_code')
        if ref_code:
            from urllib.parse import quote
            sep = '&' if '?' in url else '?'
            url = f"{url}{sep}ref={quote(str(ref_code), safe='')}"
        await update.message.reply_text(
            t(lang, 'reg_app_welcome'),
            reply_markup=InlineKeyboardMarkup([[_open_app_button(
                lang, 'reg_app_btn', url, (f"ref_{ref_code}" if ref_code else ""))]]),
            parse_mode='HTML'
        )
        return ConversationHandler.END

    # Fallback (staff taklifi yoki MINIAPP_URL yo'q) — eski ichki ro'yxat: tilni tanlash
    keyboard = [
        [InlineKeyboardButton(LANGS['uz'], callback_data="reglang_uz")],
        [InlineKeyboardButton(LANGS['ru'], callback_data="reglang_ru")],
    ]
    await update.message.reply_text(
        "🌐 Tilni tanlang / Выберите язык:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return SELECT_LANG


async def registration_language(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Til tanlandi — endi telefon so'raymiz (tanlangan tilda)."""
    query = update.callback_query
    await query.answer()
    lang = query.data.split("_")[1]  # "reglang_uz" -> "uz"
    if lang not in LANGS:
        lang = DEFAULT_LANG
    context.user_data['lang'] = lang  # DB hali yo'q — sessiyada saqlaymiz

    keyboard = [[KeyboardButton(t(lang, 'phone_button'), request_contact=True)]]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True)

    await query.message.reply_text(
        t(lang, 'welcome_ask_phone'),
        reply_markup=reply_markup
    )
    return PHONE


async def registration_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.contact:
        raw = update.message.contact.phone_number
    elif update.message.text:
        raw = update.message.text
    else:
        await update.message.reply_text(T(update, context, 'phone_send_prompt'))
        return PHONE

    phone = normalize_phone(raw)
    if not phone:
        await update.message.reply_text(T(update, context, 'phone_invalid'))
        return PHONE

    context.user_data['phone'] = phone
    logging.info(f"Phone normalized: {phone}")

    await update.message.reply_text(
        T(update, context, 'ask_name'),
        reply_markup=ReplyKeyboardRemove()
    )
    return NAME


async def registration_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = validate_fullname(update.message.text, max_len=60)
    if not name:
        await update.message.reply_text(T(update, context, 'name_invalid'))
        return NAME

    context.user_data['name'] = name

    # MULTI-SOTUVCHI: staff_<code> deeplink orqali kelgan bo'lsa — rol so'ramaymiz,
    # to'g'ridan-to'g'ri do'konga xodim sifatida bog'laymiz
    if context.user_data.get('staff_invite'):
        return await _finalize_staff_registration(update, context)

    keyboard = [
        [InlineKeyboardButton(T(update, context, 'role_buyer'), callback_data="reg_buyer")],
        [InlineKeyboardButton(T(update, context, 'role_seller'), callback_data="reg_seller")],
    ]
    await update.message.reply_text(
        T(update, context, 'name_thanks_role', name=name),
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return ROLE


async def registration_role(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    role = query.data.split("_")[1]  # "reg_buyer" -> "buyer"
    context.user_data['role'] = role

    if role == 'seller':
        lang = get_lang(update, context)
        categories = db.get_all_categories()
        keyboard = []
        for cat in categories:
            keyboard.append([InlineKeyboardButton(f"{cat[2]} {category_name(cat[1], lang)}", callback_data=f"regcat_{cat[0]}")])

        await query.edit_message_text(
            T(update, context, 'seller_category_ask'),
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return SELLER_CATEGORY
    else:
        # Xaridor uchun ro'yxatdan o'tishni yakunlaymiz
        user_id = db.create_user(
            telegram_id=update.effective_user.id,
            phone_number=context.user_data['phone'],
            name=context.user_data['name'],
            role=role
        )
        # Tanlangan tilni saqlaymiz
        db.update_user(user_id, language=context.user_data.get('lang', DEFAULT_LANG))

        # Adminlarga yangi foydalanuvchi haqida xabar
        await notify_admins(
            context,
            f"👤 <b>Yangi foydalanuvchi qo'shildi!</b>\n\n"
            f"Ism: {html.escape(context.user_data.get('name') or '')}\n"
            f"Telefon: {context.user_data.get('phone') or ''}\n"
            f"Rol: Xaridor"
        )

        await query.edit_message_text(T(update, context, 'registration_success'))
        await buyer_panel(update, context)
        return ConversationHandler.END


async def registration_seller_category(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    category_id = int(query.data.split("_")[1])
    context.user_data['seller_category'] = category_id

    await query.edit_message_text(T(update, context, 'shop_name_ask'))
    return SHOP_NAME


async def registration_shop_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    if len(name) < 2 or len(name) > 80:
        await update.message.reply_text(T(update, context, 'shop_name_invalid'))
        return SHOP_NAME
    context.user_data['shop_name'] = name
    await update.message.reply_text(T(update, context, 'shop_landmark_ask'))
    return SHOP_LANDMARK


async def registration_shop_landmark(update: Update, context: ContextTypes.DEFAULT_TYPE):
    landmark = update.message.text.strip()
    if len(landmark) > 200:
        await update.message.reply_text(T(update, context, 'shop_landmark_too_long'))
        return SHOP_LANDMARK
    context.user_data['shop_landmark'] = landmark

    keyboard = [[KeyboardButton(T(update, context, 'send_location_button'), request_location=True)]]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True)

    await update.message.reply_text(
        T(update, context, 'shop_location_ask'),
        reply_markup=reply_markup
    )
    return SHOP_ADDRESS


async def registration_shop_address(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """1/2-bosqich: do'kon LOKATSIYASI (xarita uchun). '-' yuborilsa o'tkazib yuboriladi."""
    if update.message.location:
        lat = update.message.location.latitude
        lon = update.message.location.longitude
        context.user_data['shop_lat'] = lat
        context.user_data['shop_lon'] = lon
        # Geocode'dan dastlabki manzil matni (keyingi bosqichda tahrirlanishi mumkin)
        address = await resolve_shop_address(lat, lon)
        context.user_data['shop_address'] = address   # None bo'lishi mumkin
        if address:
            await update.message.reply_text(T(update, context, 'address_detected', address=address))
    else:
        # Matn (odatda "-") — lokatsiyani o'tkazib yuborish
        context.user_data['shop_lat'] = None
        context.user_data['shop_lon'] = None
        context.user_data.setdefault('shop_address', None)

    await update.message.reply_text(
        T(update, context, 'shop_address_text_ask'),
        reply_markup=ReplyKeyboardRemove()
    )
    return SHOP_ADDRESS_TEXT


async def registration_shop_address_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """2/2-bosqich: manzil MATNI. "-" bo'lsa avvalgi (geocode) manzil saqlanadi."""
    text = update.message.text.strip()
    if text != '-':
        if len(text) < 5 or len(text) > 200:
            await update.message.reply_text(T(update, context, 'address_invalid'))
            return SHOP_ADDRESS_TEXT
        context.user_data['shop_address'] = text

    await update.message.reply_text(
        T(update, context, 'working_days_ask'),
        reply_markup=ReplyKeyboardRemove()
    )
    return WORKING_DAYS


async def registration_working_days(update: Update, context: ContextTypes.DEFAULT_TYPE):
    days = update.message.text.strip()
    if len(days) > 100:
        await update.message.reply_text(T(update, context, 'working_days_too_long'))
        return WORKING_DAYS
    context.user_data['working_days'] = days
    await update.message.reply_text(T(update, context, 'working_hours_ask'))
    return WORKING_HOURS


async def registration_working_hours(update: Update, context: ContextTypes.DEFAULT_TYPE):
    hours = update.message.text.strip()
    # Vaqtni qattiq validatsiya qilmaymiz (turli formatlar bo'lishi mumkin: "09-21", "9:00 dan 21:00 gacha")
    # Lekin parse qilib ko'ramiz, agar formatga to'g'ri kelsa — keyinchalik ish vaqti tekshiruvi uchun ishlatamiz.
    if len(hours) > 50:
        await update.message.reply_text(T(update, context, 'working_hours_too_long'))
        return WORKING_HOURS
    context.user_data['working_hours'] = hours
    await update.message.reply_text(T(update, context, 'username_ask'))
    return TELEGRAM_USERNAME


async def registration_telegram_username(update: Update, context: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    if raw == '-' or raw.lower() in ('yoq', "yo'q", "yoʼq", 'no', 'none'):
        context.user_data['telegram_username'] = None
    else:
        u = normalize_telegram_username(raw)
        if not u:
            await update.message.reply_text(T(update, context, 'username_invalid'))
            return TELEGRAM_USERNAME
        context.user_data['telegram_username'] = u

    await complete_registration(update, context)
    return ConversationHandler.END


async def notify_admins(context: ContextTypes.DEFAULT_TYPE, text, reply_markup=None, parse_mode='HTML'):
    """Barcha adminlarga xabar yuboradi: .env'dagi ADMIN_ID va DB'dagi role='admin' foydalanuvchilar.
    Bitta qabul qiluvchiga ketmasa (bloklagan/chat ochmagan), qolganlariga ketaveradi.
    Shu sabab .env'dagi ADMIN_ID adashgan bo'lsa ham, DB'dagi haqiqiy admin xabarni oladi."""
    recipients = set()
    if ADMIN_ID:
        recipients.add(ADMIN_ID)
    try:
        for a in db.get_all_users(role='admin'):
            tid = a.get('telegram_id')
            if tid:
                recipients.add(tid)
    except Exception as e:
        logging.error(f"Adminlar ro'yxatini olishda xato: {e}")
    if not recipients:
        logging.warning("notify_admins: hech qanday admin topilmadi — ADMIN_ID o'rnatilmagan va DB'da admin yo'q.")
        return
    for chat_id in recipients:
        try:
            await context.bot.send_message(
                chat_id=chat_id, text=text,
                parse_mode=parse_mode, reply_markup=reply_markup
            )
        except Exception as e:
            logging.error(f"Adminga ({chat_id}) bildirishnoma ketmadi: {e}")


async def complete_registration(update: Update, context: ContextTypes.DEFAULT_TYPE):
    role = context.user_data.get('role', 'buyer')

    # Admin tekshiruvi
    if update.effective_user.id == ADMIN_ID:
        role = 'admin'

    user_id = db.create_user(
        telegram_id=update.effective_user.id,
        phone_number=context.user_data['phone'],
        name=context.user_data['name'],
        role=role
    )
    # Tanlangan tilni saqlaymiz
    db.update_user(user_id, language=context.user_data.get('lang', DEFAULT_LANG))

    # Referral — agar /start REF... orqali kelgan bo'lsa
    referred_by = context.user_data.get('referred_by')
    if referred_by:
        try:
            db.update_user(user_id, referred_by=referred_by)
            db.increment_referral_count(referred_by)
            # Taklif qiluvchiga bildirishnoma (uning tilida)
            referrer = db.get_user_by_id(referred_by)
            if referrer and referrer.get('telegram_id'):
                try:
                    await context.bot.send_message(
                        chat_id=referrer['telegram_id'],
                        text=t(referrer, 'new_referral',
                               name=html.escape(context.user_data['name'])),
                        parse_mode='HTML'
                    )
                except Exception as e:
                    logging.error(f"Referrerga bildirishnoma ketmadi: {e}")
        except Exception as e:
            logging.error(f"Referral saqlanmadi: {e}")

    shop_name = context.user_data.get('shop_name')
    if shop_name:
        db.update_user(
            user_id,
            shop_name=shop_name,
            shop_address=context.user_data.get('shop_address'),
            shop_landmark=context.user_data.get('shop_landmark'),
            shop_lat=context.user_data.get('shop_lat'),
            shop_lon=context.user_data.get('shop_lon'),
            working_days=context.user_data.get('working_days'),
            working_hours=context.user_data.get('working_hours'),
            telegram_username=context.user_data.get('telegram_username'),
            is_verified=1
        )

    # Sotuvchi uchun — admin tasdiqlashi kerak
    if role == 'seller':
        db.create_seller_request(user_id)
        db.update_user(user_id, is_approved=0)

        # MULTI-SOTUVCHI: yangi sotuvchi uchun do'kon (owner shop_staff bilan) yaratamiz
        try:
            db.create_shop(
                user_id,
                name=shop_name,
                address=context.user_data.get('shop_address'),
                landmark=context.user_data.get('shop_landmark'),
                lat=context.user_data.get('shop_lat'),
                lon=context.user_data.get('shop_lon'),
                working_days=context.user_data.get('working_days'),
                working_hours=context.user_data.get('working_hours'),
            )
        except Exception as e:
            logging.error(f"Yangi sotuvchi uchun do'kon yaratilmadi: {e}")

        await update.message.reply_text(T(update, context, 'reg_success_seller'))

        # Adminlarga bildirishnoma
        user_name = html.escape(context.user_data.get('name') or '')
        shop = html.escape(shop_name or '')
        phone = context.user_data.get('phone') or ''
        await notify_admins(
            context,
            f"🆕 <b>Yangi sotuvchi so'rovi!</b>\n\n"
            f"👤 Ism: {user_name}\n"
            f"📞 Telefon: {phone}\n"
            f"🏪 Do'kon: {shop}\n"
            f"📍 Manzil: {html.escape(context.user_data.get('shop_address') or '')}\n\n"
            f"Tasdiqlash uchun: Admin panel → Sotuvchi so'rovlari",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Tasdiqlash", callback_data=f"approve_seller_{user_id}")],
                [InlineKeyboardButton("❌ Rad etish", callback_data=f"reject_seller_{user_id}")],
            ])
        )

        await buyer_panel(update, context)
    elif role == 'admin':
        await update.message.reply_text(T(update, context, 'registration_success'))
        await admin_panel(update, context)
    else:
        await update.message.reply_text(T(update, context, 'registration_success'))
        await buyer_panel(update, context)

    return ConversationHandler.END


# ============================================================
# ROL ALMASHTIRISH (Xaridor ↔ Sotuvchi)
# Bitta akkaunt — ikkala rejim. role ustuni asosiy rolni saqlaydi,
# active_mode esa hozir qaysi panel ko'rsatilayotganini belgilaydi.
# ============================================================

async def switch_to_buyer_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Rol o'zgartirishdan oldin tasdiqlash so'raydi (sotuvchi → xaridor)."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    await query.edit_message_text(
        t(lang, 'switch_to_buyer_confirm_text'),
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton(t(lang, 'btn_yes_to_buyer'), callback_data="do_switch_buyer")],
            [InlineKeyboardButton(t(lang, 'btn_no_stay'), callback_data="seller_panel")],
        ])
    )


async def switch_to_seller_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Rol o'zgartirishdan oldin tasdiqlash so'raydi (xaridor → sotuvchi)."""
    query = update.callback_query
    await query.answer()
    user = db.get_user_by_telegram_id(update.effective_user.id)

    # Do'kon yo'q bo'lsa — to'g'ridan-to'g'ri become_seller oqimiga (tasdiqsiz)
    if not is_seller_capable(user):
        await switch_to_seller(update, context)
        return

    lang = get_lang(update, context)
    await query.edit_message_text(
        t(lang, 'switch_to_seller_confirm_text'),
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton(t(lang, 'btn_yes_to_seller'), callback_data="do_switch_seller")],
            [InlineKeyboardButton(t(lang, 'btn_no_stay'), callback_data="buyer_panel")],
        ])
    )


async def switch_to_buyer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    context.user_data['active_mode'] = 'buyer'
    lang = get_lang(update, context)
    await buyer_panel(update, context)
    # Pastki menyuni ham xaridor variantiga o'tkazamiz
    await context.bot.send_message(
        chat_id=update.effective_chat.id,
        text=t(lang, 'in_buyer_mode'),
        reply_markup=buyer_bottom_kb(lang)
    )


async def switch_to_seller(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user = db.get_user_by_telegram_id(update.effective_user.id)
    lang = get_lang(update, context)

    if is_seller_capable(user):
        # Do'kon ma'lumotlari bor — to'g'ridan-to'g'ri sotuvchi paneliga
        context.user_data['active_mode'] = 'seller'
        await seller_panel(update, context)
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=t(lang, 'in_seller_mode'),
            reply_markup=seller_bottom_kb(lang)
        )
        return

    # Do'kon yo'q — sotuvchi bo'lishni taklif qilamiz
    kb = [
        [InlineKeyboardButton(t(lang, 'btn_yes_become_seller'), callback_data="become_seller")],
        [InlineKeyboardButton(t(lang, 'btn_cancel'), callback_data="buyer_panel")],
    ]
    await query.edit_message_text(
        t(lang, 'become_seller_prompt'),
        reply_markup=InlineKeyboardMarkup(kb)
    )


async def become_seller_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """'Sotuvchi bo'lish' jarayonini boshlaydi — mavjud akkauntga do'kon ma'lumotlari qo'shamiz."""
    query = update.callback_query
    await query.answer()
    await query.edit_message_text(T(update, context, 'become_seller_start_text'))
    return SHOP_NAME


async def become_seller_finish(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Jarayonning so'nggi qadami — telegram username va DB ga yozish."""
    context.user_data['telegram_username'] = update.message.text.strip()

    user = db.get_user_by_telegram_id(update.effective_user.id)
    db.update_user(
        user['id'],
        role='seller',
        shop_name=context.user_data.get('shop_name'),
        shop_address=context.user_data.get('shop_address'),
        shop_landmark=context.user_data.get('shop_landmark'),
        shop_lat=context.user_data.get('shop_lat'),
        shop_lon=context.user_data.get('shop_lon'),
        working_days=context.user_data.get('working_days'),
        working_hours=context.user_data.get('working_hours'),
        telegram_username=context.user_data.get('telegram_username'),
        is_verified=1,
        is_approved=0,
    )

    # Admin tasdiqlashi uchun so'rov yaratamiz
    db.create_seller_request(user['id'])

    await update.message.reply_text(T(update, context, 'become_seller_saved'))

    # Adminlarga bildirishnoma
    user_name = html.escape(user.get('name') or '')
    shop = html.escape(context.user_data.get('shop_name') or '')
    phone = user.get('phone_number') or ''
    await notify_admins(
        context,
        f"🆕 <b>Yangi sotuvchi so'rovi!</b>\n\n"
        f"👤 Ism: {user_name}\n"
        f"📞 Telefon: {phone}\n"
        f"🏪 Do'kon: {shop}\n"
        f"📍 Manzil: {html.escape(context.user_data.get('shop_address') or '')}\n\n"
        f"Tasdiqlash uchun: Admin panel → Sotuvchi so'rovlari",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ Tasdiqlash", callback_data=f"approve_seller_{user['id']}")],
            [InlineKeyboardButton("❌ Rad etish", callback_data=f"reject_seller_{user['id']}")],
        ])
    )

    context.user_data['active_mode'] = 'buyer'
    await buyer_panel(update, context)
    return ConversationHandler.END


async def reapply_seller(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Rad etilgan sotuvchi qayta so'rov yuboradi."""
    query = update.callback_query
    await query.answer()

    user = db.get_user_by_telegram_id(update.effective_user.id)
    if not user:
        await query.edit_message_text(T(update, context, 'user_not_found'))
        return

    lang = get_lang(update, context)

    # Yangi so'rov yaratish
    db.create_seller_request(user['id'])
    db.update_user(user['id'], is_approved=0, role='seller')

    await query.edit_message_text(
        t(lang, 'reapply_sent'),
        parse_mode='HTML',
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton(t(lang, 'btn_buyer_mode'), callback_data="do_switch_buyer")],
        ])
    )

    # Adminlarga bildirishnoma
    user_name = html.escape(user.get('name') or '')
    shop = html.escape(user.get('shop_name') or '')
    phone = user.get('phone_number') or ''
    await notify_admins(
        context,
        f"🔄 <b>Qayta sotuvchi so'rovi!</b>\n\n"
        f"👤 Ism: {user_name}\n"
        f"📞 Telefon: {phone}\n"
        f"🏪 Do'kon: {shop}\n"
        f"📍 Manzil: {html.escape(user.get('shop_address') or '')}\n\n"
        f"Tasdiqlash uchun: Admin panel → Sotuvchi so'rovlari",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ Tasdiqlash", callback_data=f"approve_seller_{user['id']}")],
            [InlineKeyboardButton("❌ Rad etish", callback_data=f"reject_seller_{user['id']}")],
        ])
    )


# ============================================================
# BUYER PANEL
# ============================================================

async def buyer_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = db.get_user_by_telegram_id(update.effective_user.id)
    context.user_data['active_mode'] = 'buyer'
    context.user_data.pop('shop_ctx', None)  # do'kon konteksti tugadi
    lang = get_lang(update, context)

    # Sotuvchi rejimi tugmasi: agar do'koni bo'lsa — to'g'ri rejimga o'tadi;
    # bo'lmasa — "sotuvchi bo'lish" jarayonini boshlaydi
    seller_btn_label = t(lang, 'btn_seller_mode') if is_seller_capable(user) else t(lang, 'btn_become_seller')

    keyboard = [
        [InlineKeyboardButton(t(lang, 'btn_search_menu'), callback_data="buyer_search_menu")],
        [InlineKeyboardButton(t(lang, 'btn_my_orders'), callback_data="buyer_orders")],
        [InlineKeyboardButton(t(lang, 'btn_my_debts'), callback_data="buyer_debts")],
        [InlineKeyboardButton(t(lang, 'btn_messages'), callback_data="buyer_messages")],
        [InlineKeyboardButton(t(lang, 'btn_reviews'), callback_data="buyer_reviews")],
        [InlineKeyboardButton(t(lang, 'btn_profile'), callback_data="buyer_profile")],
        [InlineKeyboardButton(seller_btn_label, callback_data="switch_to_seller_confirm")],
        [InlineKeyboardButton(t(lang, 'btn_ai_assistant'), callback_data="ai_assistant")],
        [InlineKeyboardButton(t(lang, 'btn_contact_admin'), callback_data="contact_admin")],
    ]
    # Mini App katalogi (WebApp) — rasm gridli "pro" katalog. Faqat MINIAPP_URL o'rnatilgan bo'lsa.
    if MINIAPP_URL:
        keyboard.insert(1, [_open_app_button(lang, 'btn_miniapp_catalog', MINIAPP_URL)])

    # MULTI-SOTUVCHI: do'konga taklif kodi bilan qo'shilish (faqat hali do'konda bo'lmaganlarga)
    if user and not db.get_staff_by_user(user['id']):
        keyboard.insert(6, [InlineKeyboardButton(t(lang, 'btn_join_with_code'), callback_data="join_with_code")])

    if CHANNEL_URL:
        keyboard.append([InlineKeyboardButton(t(lang, 'btn_official_channel'), url=CHANNEL_URL)])

    text = t(lang, 'buyer_panel_title')

    if update.callback_query:
        await update.callback_query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard))
    else:
        await update.message.reply_text(
            text,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        await update.message.reply_text(
            t(lang, 'bottom_hint'),
            reply_markup=buyer_bottom_kb(lang)
        )


async def buyer_search_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Yaxlit qidiruv tugmasi — ichida 3 ta qidiruv bo'limi:
    mahsulot qidirish, do'kon qidirish va kategoriya bo'yicha qidirish."""
    context.user_data.pop('shop_ctx', None)
    lang = get_lang(update, context)

    keyboard = [
        [InlineKeyboardButton(t(lang, 'btn_search'), callback_data="buyer_search")],
        [InlineKeyboardButton(t(lang, 'btn_shop_search'), callback_data="buyer_shop_search")],
        [InlineKeyboardButton(t(lang, 'btn_categories'), callback_data="buyer_categories")],
        [InlineKeyboardButton(t(lang, 'back'), callback_data="buyer_panel")],
    ]
    text = t(lang, 'search_menu_title')

    if update.callback_query:
        await update.callback_query.answer()
        try:
            await update.callback_query.edit_message_text(
                text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML'
            )
        except Exception:
            await update.callback_query.message.reply_text(
                text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML'
            )
    else:
        await update.message.reply_text(
            text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML'
        )


async def buyer_categories(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query

    context.user_data.pop('shop_ctx', None)  # do'kon konteksti tugadi

    lang = get_lang(update, context)
    categories = db.get_all_categories()
    keyboard = []
    for cat in categories:
        keyboard.append([InlineKeyboardButton(f"{cat[2]} {category_name(cat[1], lang)}", callback_data=f"cat_{cat[0]}")])
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="buyer_panel")])

    title = t(lang, 'categories_title')
    if query:
        await query.answer()
        await query.edit_message_text(title, reply_markup=InlineKeyboardMarkup(keyboard))
    else:
        await update.message.reply_text(title, reply_markup=InlineKeyboardMarkup(keyboard))


async def buyer_category_products(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    context.user_data.pop('shop_ctx', None)  # do'kon konteksti tugadi

    # callback_data: 'cat_ID' yoki 'cat_ID_pg_N'
    parts = query.data.split("_")
    category_id = int(parts[1])
    page = int(parts[3]) if len(parts) >= 4 and parts[2] == 'pg' else 0

    lang = get_lang(update, context)
    products = db.search_products(category_id=category_id)

    if not products:
        await query.edit_message_text(
            t(lang, 'category_empty'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="buyer_categories")]])
        )
        return

    total = len(products)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    start = page * PAGE_SIZE
    end = min(start + PAGE_SIZE, total)

    keyboard = []
    for product in products[start:end]:
        rating = product.get('prod_avg_rating') or product.get('avg_rating') or 0
        keyboard.append([InlineKeyboardButton(
            f"⭐{rating:.1f} | {product['name']} — {fmt_price(product['price'])}",
            callback_data=f"prod_{product['id']}"
        )])

    # Pagination
    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton("⬅️", callback_data=f"cat_{category_id}_pg_{page-1}"))
    nav.append(InlineKeyboardButton(f"{page+1}/{total_pages}", callback_data="noop"))
    if page < total_pages - 1:
        nav.append(InlineKeyboardButton("➡️", callback_data=f"cat_{category_id}_pg_{page+1}"))
    if nav:
        keyboard.append(nav)
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="buyer_categories")])

    await query.edit_message_text(
        t(lang, 'products_page', total=total, page=page+1, pages=total_pages),
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def _send_product_card(context, chat_id, images, text, keyboard):
    """Mahsulot kartochkasini yuboradi. images — file_id ro'yxati (0..5 ta).
    • 0 rasm  → faqat matn
    • 1 rasm  → rasm + caption (eski xatti-harakat)
    • 2-5 rasm → albom (media group) + alohida tugmali xabar.
      (Albomga tugma biriktirib bo'lmaydi va caption 1024 belgidan oshmasligi kerak,
       shuning uchun matn alohida xabarda yuboriladi.)"""
    markup = InlineKeyboardMarkup(keyboard) if keyboard else None
    images = [i for i in (images or []) if i][:db.MAX_PRODUCT_IMAGES]
    if len(images) <= 1:
        if images:
            try:
                await context.bot.send_photo(
                    chat_id=chat_id, photo=images[0], caption=text,
                    reply_markup=markup, parse_mode='HTML'
                )
            except BadRequest as e:
                # file_id yaroqsiz bo'lsa (rasm boshqa bot tomonidan yuklangan yoki
                # o'chirilgan) — kartochkani matn bilan ko'rsatamiz, qulamaymiz.
                logging.warning(f"send_photo file_id xatosi, matnga o'tildi: {e}")
                await context.bot.send_message(
                    chat_id=chat_id, text=text, reply_markup=markup, parse_mode='HTML'
                )
        else:
            await context.bot.send_message(
                chat_id=chat_id, text=text, reply_markup=markup, parse_mode='HTML'
            )
        return
    # 2-4 rasm: albomни ham xavfsiz yuboramiz (file_id buzilsa, matn baribir boradi)
    media = [InputMediaPhoto(media=images[0])]
    for fid in images[1:]:
        media.append(InputMediaPhoto(media=fid))
    try:
        await context.bot.send_media_group(chat_id=chat_id, media=media)
    except Exception as e:
        logging.error(f"send_media_group xatosi: {e}")
    await context.bot.send_message(
        chat_id=chat_id, text=text, reply_markup=markup, parse_mode='HTML'
    )


async def buyer_product_details(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    product_id = int(query.data.split("_")[1])
    lang = get_lang(update, context)

    product = db.get_product_by_id(product_id)

    if not product:
        await query.edit_message_text(t(lang, 'product_not_found'))
        return

    # Ko'rish tarixini yangilaymiz
    _track_viewed(context, product_id, product.get('category_id'))

    rating = product['avg_rating'] or 0          # do'kon (sotuvchi) reytingi
    prod_avg = product.get('prod_avg_rating') or 0   # mahsulot reytingi
    prod_count = product.get('prod_review_count') or 0

    map_link = ""
    if product.get('shop_lat') and product.get('shop_lon'):
        _url = f"https://www.google.com/maps/search/?api=1&query={product['shop_lat']},{product['shop_lon']}"
        map_link = t(lang, 'frag_map', url=_url)

    keyboard = []
    # Do'kon qidirish ichida bo'lsa (shu do'kon konteksti) — savatga qo'shish imkoni
    in_shop_ctx = context.user_data.get('shop_ctx') == product.get('seller_id')
    if in_shop_ctx and product.get('in_stock'):
        cart = _cart(context)
        in_cart = None
        if cart and cart.get('seller_id') == product.get('seller_id'):
            in_cart = cart.get('items', {}).get(str(product_id))
        if in_cart:
            keyboard.append([
                InlineKeyboardButton("➖", callback_data=f"cart_dec_{product_id}"),
                InlineKeyboardButton(t(lang, 'btn_in_cart_qty', n=in_cart['qty']), callback_data="cart_view"),
                InlineKeyboardButton("➕", callback_data=f"cart_inc_{product_id}"),
            ])
        else:
            keyboard.append([InlineKeyboardButton(
                t(lang, 'btn_add_to_cart'), callback_data=f"cart_add_{product_id}"
            )])
        keyboard.append([InlineKeyboardButton(
            t(lang, 'btn_view_cart_n', n=_cart_count(context)), callback_data="cart_view"
        )])

    keyboard += [
        [InlineKeyboardButton(t(lang, 'btn_order'), callback_data=f"order_{product_id}")],
        [InlineKeyboardButton(t(lang, 'btn_send_message'), callback_data=f"msg_{product_id}")],
    ]
    if prod_count > 0:
        keyboard.append([InlineKeyboardButton(
            t(lang, 'btn_product_reviews', n=prod_count), callback_data=f"pcomm_{product_id}"
        )])
    if product.get('telegram_username'):
        keyboard.append([InlineKeyboardButton(
            t(lang, 'btn_tg_label', u=product['telegram_username']),
            url=f"https://t.me/{product['telegram_username'].replace('@', '')}"
        )])
    if product.get('phone_number'):
        keyboard.append([InlineKeyboardButton(
            t(lang, 'btn_phone_label', p=product['phone_number']),
            callback_data=f"call_{product['id']}"
        )])

    # Recommendation tugmasi — agar tarix bo'lsa
    history = context.user_data.get('_view_history', [])
    if len(history) >= 2:  # kamida 2 ta ko'rilgan bo'lsa
        keyboard.append([InlineKeyboardButton(
            t(lang, 'btn_recommend'),
            callback_data=f"recommend_{product_id}"
        )])

    # O'xshash mahsulotlar — sotuvni oshirish uchun shu do'kondan taklif
    similar = db.get_similar_products(product_id, limit=3)
    if similar:
        keyboard.append([InlineKeyboardButton(t(lang, 'similar_title'), callback_data="noop")])
        for sp in similar:
            emoji = sp.get('category_emoji') or '🛍'
            keyboard.append([InlineKeyboardButton(
                t(lang, 'btn_similar_item', emoji=emoji,
                  name=(sp['name'] or '')[:25], price=fmt_price(sp['price'])),
                callback_data=f"prod_{sp['id']}"
            )])

    # Orqaga — do'kon kontekstida do'konga, aks holda kategoriyalarga
    if in_shop_ctx:
        keyboard.append([InlineKeyboardButton(
            t(lang, 'btn_back_to_shop'), callback_data=f"shop_products_{product['seller_id']}_0"
        )])
    else:
        keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="buyer_categories")])

    # HTML rejimi — foydalanuvchi matnida `*`, `_`, `[` bo'lsa ham crash bo'lmaydi
    name = html.escape(product['name'] or '')
    shop_name = html.escape(product.get('shop_name') or t(lang, 'unknown_word'))
    verified_badge = " ✅" if product.get('is_verified') else ""
    addr_text = human_address(product.get('shop_address'))
    manzil_line = t(lang, 'frag_address', addr=html.escape(addr_text)) if addr_text else ""
    shop_landmark = html.escape(product.get('shop_landmark') or t(lang, 'not_specified'))
    working_hours = html.escape(product.get('working_hours') or t(lang, 'not_specified'))
    working_days = html.escape(product.get('working_days') or t(lang, 'not_specified'))
    desc = html.escape(product.get('description') or t(lang, 'none_word'))

    # Hudud va masofa
    region_lbl = region_label_l(product.get('seller_region_id'), lang)
    region_line = t(lang, 'frag_region', label=html.escape(region_lbl)) if region_lbl else ""
    dist_raw = distance_line_to_shop(context, product.get('shop_lat'), product.get('shop_lon'),
                                     prefix=t(lang, 'dist_from_you'))
    dist_line = (dist_raw.lstrip("\n") + "\n") if dist_raw else ""

    # Hozir ochiq/yopiqmi (faqat working_hours parse qilinsa)
    open_status = is_shop_open_now(product.get('working_hours'))
    if open_status is True:
        open_line = "  " + t(lang, 'open_now')
    elif open_status is False:
        open_line = "  " + t(lang, 'closed_now')
    else:
        open_line = ""

    stock_line = ""
    if product.get('stock_count') is not None:
        stock_line = t(lang, 'frag_stock', n=product['stock_count'])

    # Dinamik atributlar
    attrs = db.get_product_attributes(product['id'])
    attrs_text = ""
    if attrs:
        lines = []
        for a in attrs:
            label = a.get('attr_label') or a['attr_key']
            lines.append(f"• {label}: {a['attr_value']}")
        attrs_text = t(lang, 'frag_attrs_title') + "\n".join(lines)

    rating_cnt = (t(lang, 'frag_rating_count', n=prod_count) if prod_count
                  else t(lang, 'frag_no_rating'))
    text = t(lang, 'product_card',
             name=name, price=price_with_unit(product), stock=stock_line,
             shop=shop_name, verified=verified_badge,
             region=region_line, address=manzil_line,
             landmark=shop_landmark, map=map_link, dist=dist_line,
             prod_rating=f"{prod_avg:.1f}", rating_cnt=rating_cnt,
             shop_rating=f"{rating:.1f}", wh=working_hours, open=open_line,
             wd=working_days, desc=desc, attrs=attrs_text)

    # Rasm(lar) bo'lsa — alohida xabar(lar) bilan yuboramiz, keyin batafsil ma'lumotni
    # (edit_message_text rasm bilan ishlamaydi, shuning uchun eski xabarni o'chiramiz)
    images = db.get_product_images(product_id)
    if images:
        try:
            await query.message.delete()
        except Exception:
            pass
        await _send_product_card(context, update.effective_chat.id, images, text, keyboard)
    else:
        await query.edit_message_text(
            text,
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode='HTML'
        )


async def product_reviews_view(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mahsulotga yozilgan izohlar ro'yxati (callback: pcomm_{product_id})."""
    query = update.callback_query
    # button_handler allaqachon query.answer() chaqirgan

    try:
        product_id = int(query.data.split("_")[1])
    except (ValueError, IndexError):
        return

    lang = get_lang(update, context)
    product = db.get_product_basic(product_id)
    pname = html.escape(product.get('name')) if product else t(lang, 'product_word')
    prod_avg, prod_count = db.get_product_avg_rating(product_id)
    reviews = db.get_product_reviews(product_id)

    back_kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(t(lang, 'btn_back_to_product'), callback_data=f"prod_{product_id}")]
    ])

    if not reviews:
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=t(lang, 'reviews_none_for_product', name=pname),
            reply_markup=back_kb, parse_mode='HTML'
        )
        return

    lines = [
        t(lang, 'reviews_header', name=pname),
        t(lang, 'reviews_product_rating', avg=f"{prod_avg:.1f}", count=prod_count),
    ]
    for r in reviews:
        pr = r.get('product_rating') or 0
        stars = "⭐" * pr + "☆" * (5 - pr) if pr else ""
        buyer = html.escape(r.get('buyer_name') or t(lang, 'anonymous'))
        comment = html.escape(r.get('comment') or '')
        date = fmt_datetime(r.get('created_at'))
        block = f"\n{stars}\n👤 {buyer} · {date}"
        if comment:
            block += f"\n💬 {comment}"
        reply = html.escape(r.get('seller_reply') or '')
        if reply:
            block += f"\n   ↳ {t(lang, 'review_shop_reply')} {reply}"
        lines.append(block)

    text = "\n".join(lines)
    if len(text) > 4000:
        text = text[:3900] + t(lang, 'reviews_old_cut')

    await context.bot.send_message(
        chat_id=update.effective_chat.id,
        text=text, reply_markup=back_kb, parse_mode='HTML'
    )


async def skip_search_location(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """'Lokatsiyasiz qidirish' tugmasi — lokatsiyasiz natijalarni ko'rsatadi (pagination bilan)."""
    query = update.callback_query
    await query.answer()

    lang = get_lang(update, context)
    q_text = context.user_data.pop('search_query', None)
    context.user_data.pop('search_state', None)

    if not q_text:
        await query.edit_message_text(t(lang, 'search_cancelled'))
        return

    await query.edit_message_text(t(lang, 'searching_for', q=q_text))
    region_id = context.user_data.get('search_region_id')
    products = db.search_products(query=q_text, region_id=region_id)
    await _render_search_page(update.effective_chat.id, context, products,
                              page=0, sort_by='rating', query_text=q_text,
                              buyer_lat=None)


PAGE_SIZE = 10  # sahifadagi mahsulotlar soni
CATALOG_PAGE_SIZE = 8  # do'kon mahsulotlari (rasmli albom grid) sahifasidagi soni (albom maks. 10)
HISTORY_LIMIT = 20  # ko'rilgan mahsulotlar tarixi


async def _show_product_deeplink(update: Update, context: ContextTypes.DEFAULT_TYPE, product):
    """Deeplink orqali kelgan foydalanuvchiga mahsulot sahifasini ko'rsatadi."""
    product_id = product['id']
    _track_viewed(context, product_id, product.get('category_id'))

    lang = get_lang(update, context)
    rating = product['avg_rating'] or 0              # do'kon reytingi
    prod_avg = product.get('prod_avg_rating') or 0   # mahsulot reytingi
    prod_count = product.get('prod_review_count') or 0
    map_link = ""
    if product.get('shop_lat') and product.get('shop_lon'):
        _url = f"https://www.google.com/maps/search/?api=1&query={product['shop_lat']},{product['shop_lon']}"
        map_link = t(lang, 'frag_map', url=_url)

    keyboard = [
        [InlineKeyboardButton(t(lang, 'btn_order'), callback_data=f"order_{product_id}")],
        [InlineKeyboardButton(t(lang, 'btn_send_message'), callback_data=f"msg_{product_id}")],
    ]
    if prod_count > 0:
        keyboard.append([InlineKeyboardButton(
            t(lang, 'btn_product_reviews', n=prod_count), callback_data=f"pcomm_{product_id}"
        )])
    if product.get('telegram_username'):
        keyboard.append([InlineKeyboardButton(
            t(lang, 'btn_tg_label', u=product['telegram_username']),
            url=f"https://t.me/{product['telegram_username'].replace('@', '')}"
        )])
    keyboard.append([InlineKeyboardButton(t(lang, 'btn_home'), callback_data="buyer_panel")])

    name = html.escape(product['name'] or '')
    shop_name = html.escape(product.get('shop_name') or t(lang, 'unknown_word'))
    loc_text = best_location_text(product.get('shop_address'), product.get('shop_landmark'))
    if loc_text:
        manzil_line = t(lang, 'frag_address', addr=html.escape(loc_text) + map_link)
    elif map_link:
        manzil_line = t(lang, 'frag_address_map', map=map_link)
    else:
        manzil_line = ""
    desc = html.escape(product.get('description') or t(lang, 'none_word'))
    working_hours = html.escape(product.get('working_hours') or t(lang, 'not_specified'))

    # Hudud (viloyat → tuman) — xaridor qaysi joydan ekanini ko'rsin
    region_lbl = region_label_l(product.get('seller_region_id'), lang)
    region_line = t(lang, 'frag_region', label=html.escape(region_lbl)) if region_lbl else ""
    # Xaridor lokatsiyasi eslab qolingan bo'lsa — masofa
    dist_raw = distance_line_to_shop(context, product.get('shop_lat'), product.get('shop_lon'),
                                     prefix=t(lang, 'dist_from_you'))
    dist_line = (dist_raw.lstrip("\n") + "\n") if dist_raw else ""

    rating_cnt = (t(lang, 'frag_rating_count', n=prod_count) if prod_count
                  else t(lang, 'frag_no_rating'))
    text = t(lang, 'product_card_deeplink',
             name=name, price=price_with_unit(product), shop=shop_name,
             region=region_line, address=manzil_line, dist=dist_line,
             prod_rating=f"{prod_avg:.1f}", rating_cnt=rating_cnt,
             shop_rating=f"{rating:.1f}", wh=working_hours, desc=desc)

    images = db.get_product_images(product['id'])
    if images:
        await _send_product_card(context, update.effective_chat.id, images, text, keyboard)
    else:
        await update.message.reply_text(
            text,
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode='HTML'
        )

    # Masofa hali noma'lum (xaridor joylashuvi yo'q) va do'kon koordinatasi bor bo'lsa —
    # xaridordan joylashuvini so'raymiz, keyin sotuvchigacha masofani ko'rsatamiz.
    b_lat, b_lon = get_buyer_geo(context)
    if (b_lat is None or b_lon is None) and product.get('shop_lat') and product.get('shop_lon'):
        context.user_data['_dist_pending_product'] = product_id
        loc_kb = ReplyKeyboardMarkup(
            [[KeyboardButton(t(lang, 'btn_send_location'), request_location=True)]],
            resize_keyboard=True, one_time_keyboard=True,
        )
        await update.message.reply_text(t(lang, 'deeplink_ask_location'), reply_markup=loc_kb)


def _track_viewed(context, product_id, category_id):
    """Xaridor ko'rgan mahsulotlarni kuzatib boradi (recommendation uchun)."""
    history = context.user_data.setdefault('_view_history', [])
    # Takrorlashni oldini olamiz
    entry = {'pid': product_id, 'cid': category_id}
    history = [h for h in history if h['pid'] != product_id]
    history.insert(0, entry)
    context.user_data['_view_history'] = history[:HISTORY_LIMIT]


def _get_recommendations(context, db, current_product_id, limit=5):
    """Ko'rish tarixiga qarab tavsiya qilinadigan mahsulotlar."""
    history = context.user_data.get('_view_history', [])
    if not history:
        return []

    # Ko'rilgan kategoriyalar (eng ko'p ko'rilganidan)
    from collections import Counter
    cat_counts = Counter(h['cid'] for h in history if h.get('cid'))
    seen_pids = {h['pid'] for h in history}
    seen_pids.add(current_product_id)

    recommendations = []
    for cat_id, _ in cat_counts.most_common(3):
        products = db.search_products(category_id=cat_id, sort_by='rating')
        for p in products:
            if p['id'] not in seen_pids and len(recommendations) < limit:
                recommendations.append(p)
                seen_pids.add(p['id'])

    return recommendations[:limit]


async def _render_search_page(chat_id, context, products, page=0, sort_by='rating',
                              query_text='', buyer_lat=None):
    """Qidiruv natijalarini sahifa bo'yicha ko'rsatadi (pagination + sort)."""
    lang = get_lang(None, context)
    total = len(products)
    if total == 0:
        await context.bot.send_message(
            chat_id=chat_id,
            text=t(lang, 'search_no_results_q', q=query_text),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="buyer_panel")]])
        )
        return

    total_pages = (total + PAGE_SIZE - 1) // PAGE_SIZE
    page = max(0, min(page, total_pages - 1))
    start = page * PAGE_SIZE
    end = min(start + PAGE_SIZE, total)

    # State'ni saqlaymiz — sahifa va sort almashtirilganda qayta qidirish kerak emas
    context.user_data['_search_products'] = products
    context.user_data['_search_query'] = query_text
    context.user_data['_search_sort'] = sort_by
    context.user_data['_search_buyer_lat'] = buyer_lat
    # buyer_lon ni alohida saqlamasak ham, products ichida _distance_km bor

    for product in products[start:end]:
        rating = product.get('avg_rating') or 0          # do'kon reytingi
        prod_avg = product.get('prod_avg_rating') or 0   # mahsulot reytingi
        prod_count = product.get('prod_review_count') or 0
        map_link = ""
        if product.get('shop_lat') and product.get('shop_lon'):
            _url = (f"https://www.google.com/maps/search/?api=1&"
                    f"query={product['shop_lat']},{product['shop_lon']}")
            map_link = t(lang, 'frag_map', url=_url)

        distance_line = ""
        if product.get('_distance_km') is not None:
            distance_line = t(lang, 'frag_distance', km=f"{product['_distance_km']:.1f}")

        contact_keyboard = []
        if product.get('telegram_username'):
            contact_keyboard.append([InlineKeyboardButton(
                t(lang, 'btn_telegram'),
                url=f"https://t.me/{product['telegram_username'].replace('@', '')}"
            )])
        if product.get('phone_number'):
            contact_keyboard.append([InlineKeyboardButton(
                t(lang, 'btn_phone'), callback_data=f"call_{product['id']}"
            )])
        contact_keyboard.append([InlineKeyboardButton(t(lang, 'btn_details'), callback_data=f"prod_{product['id']}")])
        if prod_count > 0:
            contact_keyboard.append([InlineKeyboardButton(
                t(lang, 'btn_reviews_n', n=prod_count), callback_data=f"pcomm_{product['id']}"
            )])

        emoji = product.get('category_emoji') or '📦'
        name = html.escape(product['name'] or '')
        shop_name = html.escape(product.get('shop_name') or t(lang, 'unknown_word'))
        region_lbl = region_label_l(product.get('seller_region_id'), lang)
        region_line = f"🌍 {html.escape(region_lbl)}\n" if region_lbl else ""
        loc_text = best_location_text(product.get('shop_address'), product.get('shop_landmark'))
        if loc_text:
            manzil_line = f"📍 {html.escape(loc_text)}{map_link}{distance_line}\n"
        elif map_link or distance_line:
            manzil_line = f"{t(lang, 'map_see_line')}{map_link}{distance_line}\n"
        else:
            manzil_line = ""
        rating_cnt = (t(lang, 'srch_rating_cnt', n=prod_count) if prod_count
                      else t(lang, 'srch_no_rating'))
        caption = t(lang, 'search_item_card',
                    emoji=emoji, name=name, price=price_with_unit(product),
                    shop=shop_name, region=region_line, address=manzil_line,
                    prod_rating=f"{prod_avg:.1f}", rating_cnt=rating_cnt,
                    shop_rating=f"{rating:.1f}")

        if product.get('image_url'):
            try:
                await context.bot.send_photo(
                    chat_id=chat_id, photo=product['image_url'],
                    caption=caption, parse_mode='HTML',
                    reply_markup=InlineKeyboardMarkup(contact_keyboard)
                )
            except Exception:
                # Rasm yuborilmasa — oddiy matn
                await context.bot.send_message(
                    chat_id=chat_id, text=caption, parse_mode='HTML',
                    reply_markup=InlineKeyboardMarkup(contact_keyboard)
                )
        else:
            await context.bot.send_message(
                chat_id=chat_id, text=caption, parse_mode='HTML',
                reply_markup=InlineKeyboardMarkup(contact_keyboard)
            )

    # Sort + pagination tugmalari
    sort_labels = {
        'rating': t(lang, 'sort_rating'),
        'price_asc': t(lang, 'sort_price_asc'),
        'price_desc': t(lang, 'sort_price_desc'),
        'newest': t(lang, 'sort_newest'),
    }
    sort_buttons = []
    for key, label in sort_labels.items():
        marker = "✅ " if key == sort_by else ""
        sort_buttons.append(InlineKeyboardButton(f"{marker}{label}", callback_data=f"srt_{key}"))
    # 2x2 grid uchun bo'lamiz
    sort_kb = [sort_buttons[i:i+2] for i in range(0, len(sort_buttons), 2)]

    # Pagination
    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton(t(lang, 'btn_prev'), callback_data=f"pg_{page-1}"))
    nav.append(InlineKeyboardButton(f"{page+1}/{total_pages}", callback_data="noop"))
    if page < total_pages - 1:
        nav.append(InlineKeyboardButton(t(lang, 'btn_next'), callback_data=f"pg_{page+1}"))

    footer_kb = sort_kb + ([nav] if nav else []) + [
        [InlineKeyboardButton(t(lang, 'btn_main_menu'), callback_data="buyer_panel")]
    ]

    await context.bot.send_message(
        chat_id=chat_id,
        text=t(lang, 'search_results_count', q=html.escape(query_text),
               total=total, page=page+1, pages=total_pages),
        reply_markup=InlineKeyboardMarkup(footer_kb),
        parse_mode='HTML'
    )


async def search_sort_change(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Foydalanuvchi sort tartibini o'zgartirsa."""
    query = update.callback_query
    await query.answer()

    sort_by = query.data.replace("srt_", "")
    q_text = context.user_data.get('_search_query', '')
    buyer_lat = context.user_data.get('_search_buyer_lat')

    # Saqlangan natijalardan masofalarni olib o'tish uchun avval eski natijalarni saqlab qo'yamiz
    old_products = context.user_data.get('_search_products') or []
    distances = {p['id']: p.get('_distance_km') for p in old_products}

    # Yangi sort bilan qayta qidiramiz
    products = db.search_products(query=q_text, sort_by=sort_by)
    # Eski masofalarni qayta yopishtiramiz
    for p in products:
        p['_distance_km'] = distances.get(p['id'])

    await _render_search_page(update.effective_chat.id, context, products,
                              page=0, sort_by=sort_by, query_text=q_text,
                              buyer_lat=buyer_lat)


async def search_page_change(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Sahifa o'zgartirish."""
    query = update.callback_query
    await query.answer()

    page = int(query.data.replace("pg_", ""))
    products = context.user_data.get('_search_products') or []
    if not products:
        await query.edit_message_text(T(update, context, 'search_results_gone'))
        return

    sort_by = context.user_data.get('_search_sort', 'rating')
    q_text = context.user_data.get('_search_query', '')
    buyer_lat = context.user_data.get('_search_buyer_lat')

    await _render_search_page(update.effective_chat.id, context, products,
                              page=page, sort_by=sort_by, query_text=q_text,
                              buyer_lat=buyer_lat)


async def _show_search_results(update: Update, context: ContextTypes.DEFAULT_TYPE,
                               query_text: str, buyer_lat=None, buyer_lon=None):
    """Qidiruv natijalarini ko'rsatadi (eski wrapper — pagination versiyasini chaqiradi)."""
    region_id = context.user_data.get('search_region_id')
    products = db.search_products(query=query_text, region_id=region_id)

    # Lokatsiya bo'lsa — masofani hisoblaymiz va saralaymiz
    if buyer_lat is not None and buyer_lon is not None:
        for p in products:
            d = None
            if p.get('shop_lat') and p.get('shop_lon'):
                d = haversine_km(buyer_lat, buyer_lon, p['shop_lat'], p['shop_lon'])
            p['_distance_km'] = d
        products.sort(key=lambda p: (p['_distance_km'] is None, p['_distance_km'] or 0))

    chat_id = update.effective_chat.id if update.effective_chat else update.message.chat.id
    await _render_search_page(chat_id, context, products, page=0,
                              sort_by='rating' if buyer_lat is None else 'distance',
                              query_text=query_text, buyer_lat=buyer_lat)


async def buyer_shop_search(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xaridor do'kon qidiradi."""
    query = update.callback_query
    if query:
        await query.answer()

    context.user_data['shop_search_state'] = 'awaiting_query'

    text = T(update, context, 'shop_search_prompt')
    if query:
        await query.edit_message_text(text)
    else:
        await update.message.reply_text(text)


async def buyer_shop_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Do'konlar ro'yxatini ko'rsatadi (callback: shop_list_{page})."""
    query = update.callback_query
    await query.answer()

    page = int(query.data.split("_")[2])
    shops = context.user_data.get('_shop_results') or []

    if not shops:
        await query.edit_message_text(
            T(update, context, 'shops_not_found'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(T(update, context, 'back'), callback_data="buyer_panel")]])
        )
        return

    await _render_shop_list(query, context, shops, page)


async def _render_shop_list(target, context, shops, page=0):
    """Do'konlar ro'yxatini sahifalab ko'rsatadi."""
    total = len(shops)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    start = page * PAGE_SIZE
    end = min(start + PAGE_SIZE, total)

    lang = get_lang(None, context)
    keyboard = []
    for shop in shops[start:end]:
        rating = shop.get('avg_rating') or 0
        count = shop.get('product_count') or 0
        verified = "✅ " if shop.get('is_verified') else ""
        cnt_lbl = t(lang, 'shop_count_label', n=count)
        label = f"⭐{rating:.1f} | {verified}🏪 {shop['shop_name']} ({cnt_lbl})"[:50]
        keyboard.append([InlineKeyboardButton(label, callback_data=f"shop_{shop['id']}")])

    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton("⬅️", callback_data=f"shop_list_{page-1}"))
    nav.append(InlineKeyboardButton(f"{page+1}/{total_pages}", callback_data="noop"))
    if page < total_pages - 1:
        nav.append(InlineKeyboardButton("➡️", callback_data=f"shop_list_{page+1}"))
    if nav:
        keyboard.append(nav)
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="buyer_panel")])

    text = t(lang, 'shops_page', total=total, page=page+1, pages=total_pages)

    if hasattr(target, 'edit_message_text'):
        await target.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard))
    else:
        await target.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard))


async def buyer_shop_detail(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Do'kon sahifasi — ma'lumotlar va mahsulotlar."""
    query = update.callback_query
    await query.answer()

    seller_id = int(query.data.split("_")[1])

    # Rasmli katalogdan qaytilgan bo'lsa — kartochka xabarlarini tozalab,
    # do'kon sahifasini yangi xabar bilan ko'rsatamiz (rasmni edit qilib bo'lmaydi).
    from_catalog = bool(context.user_data.get('shop_catalog_msgs'))
    if from_catalog:
        await _clear_catalog_messages(context, update.effective_chat.id)
        try:
            await query.message.delete()
        except Exception:
            pass
    lang = get_lang(update, context)
    shop = db.get_seller_public_info(seller_id)

    if not shop:
        not_found_kb = InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="buyer_panel")]])
        if from_catalog:
            await context.bot.send_message(
                chat_id=update.effective_chat.id, text=t(lang, 'shop_not_found'),
                reply_markup=not_found_kb)
        else:
            await query.edit_message_text(t(lang, 'shop_not_found'), reply_markup=not_found_kb)
        return

    # Do'kon konteksti — mahsulot sahifasida "savatga qo'shish" tugmasi shu do'kon uchun chiqsin
    context.user_data['shop_ctx'] = seller_id

    rating = shop.get('avg_rating') or 0
    product_count = shop.get('product_count') or 0
    delivered = shop.get('delivered_count') or 0
    verified_badge = " ✅" if shop.get('is_verified') else ""

    map_link = ""
    if shop.get('shop_lat') and shop.get('shop_lon'):
        _url = f"https://www.google.com/maps/search/?api=1&query={shop['shop_lat']},{shop['shop_lon']}"
        map_link = t(lang, 'frag_map', url=_url)

    open_status = is_shop_open_now(shop.get('working_hours'))
    if open_status is True:
        open_line = "  " + t(lang, 'open_now')
    elif open_status is False:
        open_line = "  " + t(lang, 'closed_now')
    else:
        open_line = ""

    # Hudud (viloyat → tuman) — xaridor do'kon qayerdaligini ko'rsin
    region_lbl = region_label_l(shop.get('region_id'), lang)
    region_line = t(lang, 'frag_region', label=html.escape(region_lbl)) if region_lbl else ""
    # Masofa — xaridor lokatsiyasi bo'lsa
    dist_line = distance_line_to_shop(context, shop.get('shop_lat'), shop.get('shop_lon'),
                                      prefix=t(lang, 'dist_from_you'))
    if dist_line:
        dist_line = dist_line.lstrip("\n") + "\n"

    addr_text = human_address(shop.get('shop_address'))
    manzil_line = t(lang, 'frag_address', addr=html.escape(addr_text) + map_link) if addr_text else ""
    # Haqiqiy manzil bo'lmasa — xarita havolasini mo'ljal qatoriga biriktiramiz
    landmark_map = "" if addr_text else map_link
    text = t(lang, 'shop_detail_card',
             shop=html.escape(shop.get('shop_name') or ''), verified=verified_badge,
             seller=html.escape(shop.get('name') or ''),
             region=region_line, address=manzil_line,
             landmark=html.escape(shop.get('shop_landmark') or t(lang, 'not_specified')),
             landmark_map=landmark_map, dist=dist_line,
             wh=html.escape(shop.get('working_hours') or t(lang, 'not_specified')), open=open_line,
             wd=html.escape(shop.get('working_days') or t(lang, 'not_specified')),
             rating=f"{rating:.1f}", pcount=product_count, delivered=delivered)
    if verified_badge:
        text += t(lang, 'verified_seller_note')

    keyboard = [
        [InlineKeyboardButton(t(lang, 'btn_view_products_n', n=product_count), callback_data=f"shop_products_{seller_id}_0")],
    ]
    # Savatda shu do'kon mahsulotlari bo'lsa — savatga tezkor kirish
    cart = _cart(context)
    if cart and cart.get('seller_id') == seller_id and cart.get('items'):
        keyboard.append([InlineKeyboardButton(
            t(lang, 'btn_my_cart_summary', n=_cart_count(context), total=fmt_price(_cart_total(context))),
            callback_data="cart_view"
        )])
    if shop.get('telegram_username'):
        keyboard.append([InlineKeyboardButton(
            t(lang, 'btn_tg_at', u=shop['telegram_username']),
            url=f"https://t.me/{shop['telegram_username']}"
        )])
    if shop.get('phone_number'):
        keyboard.append([InlineKeyboardButton(
            t(lang, 'btn_phone_plain', p=shop['phone_number']), callback_data=f"noop"
        )])
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="buyer_panel")])

    if from_catalog:
        await context.bot.send_message(
            chat_id=update.effective_chat.id, text=text,
            reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')
    else:
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


async def _clear_catalog_messages(context, chat_id):
    """Oldingi katalog sahifasida yuborilgan rasm-kartochkalar va sarlavha/footer
    xabarlarini o'chiradi (sahifalashda eski sahifa qolib ketmasligi uchun)."""
    context.user_data.pop('catalog_footer', None)
    ids = context.user_data.pop('shop_catalog_msgs', [])
    for mid in ids:
        try:
            await context.bot.delete_message(chat_id=chat_id, message_id=mid)
        except Exception:
            pass


async def _refresh_catalog_footer(context, chat_id, lang):
    """Savatga mahsulot qo'shilganda katalog footer'idagi savat tugmasini (soni+summasi)
    JONLI yangilaydi — foydalanuvchi savatga tushganini darhol ko'rsin."""
    info = context.user_data.get('catalog_footer')
    if not info:
        return
    try:
        markup = _catalog_footer_buttons(lang, info['seller_id'], info['page'], info['pages'], context)
        await context.bot.edit_message_reply_markup(
            chat_id=chat_id, message_id=info['msg_id'], reply_markup=markup)
    except Exception:
        pass


def _product_is_new(created_at, days=7):
    """Mahsulot oxirgi `days` kun ichida qo'shilganmi? Aniqlay olmasa — False."""
    if not created_at:
        return False
    try:
        from datetime import datetime, timezone, timedelta
        dt = datetime.strptime(str(created_at)[:19], "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
        return (datetime.now(timezone.utc) - dt) <= timedelta(days=days)
    except Exception:
        return False


def _format_catalog_card(lang, product, pos, total):
    """Bitta mahsulot kartochkasi matni — Uzum uslubi:
    badge (YANGI yoki kategoriya) + nom + narx + reyting (+ zahira)."""
    # Badge: yangi mahsulot bo'lsa "YANGI", aks holda kategoriya nomi
    if _product_is_new(product.get('created_at')):
        badge = t(lang, 'catalog_badge_new')
    elif product.get('category_name'):
        badge = t(lang, 'catalog_badge_cat',
                  emoji=product.get('category_emoji') or '🏷',
                  name=html.escape(product['category_name']))
    else:
        badge = ""

    # Reyting (mahsulot bo'yicha) + baholar soni
    prod_avg = product.get('prod_avg_rating') or 0
    prod_count = product.get('prod_review_count') or 0
    if prod_count > 0:
        rating = fmt_rating(prod_avg, prod_count)        # "⭐ 4.8 (227 baho)"
    else:
        rating = t(lang, 'catalog_card_new_rating')      # "✨ Yangi mahsulot"

    # Zahira
    stock = ""
    if product.get('stock_count') is not None:
        stock = t(lang, 'catalog_stock_frag', n=product['stock_count'])

    return t(lang, 'catalog_card_feed',
             badge=badge,
             name=html.escape(product.get('name') or ''),
             price=price_with_unit(product),
             rating=rating, stock=stock)


def _catalog_card_buttons(lang, product, cart_items):
    """Kartochka tugmalari:
    1-qator: [🛍 Sotib olish] (yakka buyurtma oqimi) — asosiy amal.
    2-qator: [➕ Savatga / 🛒 N] [📦 Batafsil]."""
    pid = product['id']
    in_cart = cart_items.get(str(pid))
    if in_cart:
        cart_btn = InlineKeyboardButton(
            t(lang, 'btn_cart_qty_short', n=in_cart['qty']), callback_data="cart_view")
    else:
        cart_btn = InlineKeyboardButton(
            t(lang, 'btn_add_to_cart'), callback_data=f"cart_add_{pid}")
    buy_btn = InlineKeyboardButton(t(lang, 'btn_buy_now'), callback_data=f"order_{pid}")
    details_btn = InlineKeyboardButton(t(lang, 'btn_details'), callback_data=f"prod_{pid}")
    return InlineKeyboardMarkup([[buy_btn], [cart_btn, details_btn]])


def _catalog_footer_buttons(lang, seller_id, page, pages, context):
    """Lenta oxiridagi boshqaruv: sahifalash + savatni rasmiylashtirish + AI + do'konga qaytish."""
    kb = []
    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton(t(lang, 'btn_page_prev'),
                                        callback_data=f"shop_products_{seller_id}_{page-1}"))
    if page < pages - 1:
        nav.append(InlineKeyboardButton(t(lang, 'btn_page_next'),
                                        callback_data=f"shop_products_{seller_id}_{page+1}"))
    if nav:
        kb.append(nav)
    # Savat tugmasi — DOIM ko'rinadi (savatni ko'rish uchun orqaga qaytish shart emas).
    # Mahsulot bor bo'lsa — soni + summasi, bo'lmasa — oddiy "Savatni ko'rish".
    cnt = _cart_count(context)
    if cnt > 0:
        cart_label = t(lang, 'btn_my_cart_summary', n=cnt, total=fmt_price(_cart_total(context)))
    else:
        cart_label = t(lang, 'btn_view_cart')
    kb.append([InlineKeyboardButton(cart_label, callback_data="cart_view")])
    if ai_assistant.is_enabled():
        kb.append([InlineKeyboardButton(
            t(lang, 'btn_shop_ai_search'), callback_data=f"shop_ai_{seller_id}")])
    kb.append([InlineKeyboardButton(t(lang, 'btn_back_to_shop'), callback_data=f"shop_{seller_id}")])
    return InlineKeyboardMarkup(kb)


async def buyer_shop_products(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Do'kon mahsulotlari — rasmli LENTA (Uzum uslubi): RASM va ma'lumot BIRGA.

    Har bir mahsulot — rasm + caption (nom + narx + reyting + zahira) bitta kartochkada,
    tagida ixcham [➕ Savatga] [📦 Batafsil]. Pastga aylantirib ko'riladi. Sahifada
    CATALOG_PAGE_SIZE ta; pastda sahifalash + savat + AI + do'konga qaytish.
    callback: shop_products_{seller_id}_{page}  (page — sahifa raqami, 0-based)."""
    query = update.callback_query
    await query.answer()

    parts = query.data.split("_")
    seller_id = int(parts[2])
    page = int(parts[3])

    lang = get_lang(update, context)
    chat_id = update.effective_chat.id
    products = db.get_shop_products(seller_id)
    shop = db.get_seller_public_info(seller_id)
    shop_name = shop.get('shop_name') if (shop and shop.get('shop_name')) else t(lang, 'shop_word')

    # Do'kon konteksti — savat shu do'kon uchun
    context.user_data['shop_ctx'] = seller_id

    # Oldingi lenta xabarlarini va tugma bosilgan xabarni tozalaymiz —
    # rasm-kartochkalarni edit qilib bo'lmaydi, shuning uchun har safar qaytadan yuboriladi.
    await _clear_catalog_messages(context, chat_id)
    try:
        await query.message.delete()
    except Exception:
        pass

    if not products:
        sent = await context.bot.send_message(
            chat_id=chat_id,
            text=t(lang, 'shop_no_products', shop=html.escape(shop_name)),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data=f"shop_{seller_id}")]]),
            parse_mode='HTML'
        )
        context.user_data['shop_catalog_msgs'] = [sent.message_id]
        return

    cart = _cart(context)
    cart_items = cart.get('items', {}) if (cart and cart.get('seller_id') == seller_id) else {}

    total = len(products)
    pages = max(1, (total + CATALOG_PAGE_SIZE - 1) // CATALOG_PAGE_SIZE)
    page = max(0, min(page, pages - 1))
    start = page * CATALOG_PAGE_SIZE
    chunk = products[start:start + CATALOG_PAGE_SIZE]

    sent_ids = []

    # 1) Sarlavha
    try:
        h = await context.bot.send_message(
            chat_id=chat_id,
            text=t(lang, 'catalog_feed_header',
                   shop=html.escape(shop_name), total=total, page=page + 1, pages=pages),
            parse_mode='HTML')
        sent_ids.append(h.message_id)
    except Exception:
        pass

    # 2) Har bir mahsulot — RASM + ma'lumot BIRGA bitta kartochkada (rasm caption'ida
    # nom + narx + reyting), tagida ixcham [➕ Savatga] [📦 Batafsil]. Pastga aylantirib
    # ko'riladi (Uzum uslubidagi lenta).
    for i, product in enumerate(chunk):
        caption = _format_catalog_card(lang, product, start + i + 1, total)
        markup = _catalog_card_buttons(lang, product, cart_items)
        images = db.get_product_images(product['id'])
        photo = images[0] if images else None
        try:
            if photo:
                m = await context.bot.send_photo(
                    chat_id=chat_id, photo=photo, caption=caption,
                    reply_markup=markup, parse_mode='HTML')
            else:
                m = await context.bot.send_message(
                    chat_id=chat_id, text=caption,
                    reply_markup=markup, parse_mode='HTML')
            sent_ids.append(m.message_id)
        except Exception:
            # Rasm yuborilmasa (buzuq file_id va h.k.) — matn bilan davom etamiz
            try:
                m = await context.bot.send_message(
                    chat_id=chat_id, text=caption,
                    reply_markup=markup, parse_mode='HTML')
                sent_ids.append(m.message_id)
            except Exception:
                pass
        await asyncio.sleep(0.04)  # Telegram flud-limitiga ehtiyot

    # 3) Footer — sahifalash + savat + AI + do'kon tugmalari
    try:
        f = await context.bot.send_message(
            chat_id=chat_id,
            text=t(lang, 'catalog_feed_footer', page=page + 1, pages=pages, total=total),
            reply_markup=_catalog_footer_buttons(lang, seller_id, page, pages, context),
            parse_mode='HTML')
        sent_ids.append(f.message_id)
        # Savatga qo'shilganda footer savat tugmasini JONLI yangilash uchun ma'lumot saqlaymiz
        context.user_data['catalog_footer'] = {
            'msg_id': f.message_id, 'seller_id': seller_id, 'page': page, 'pages': pages}
    except Exception:
        pass

    context.user_data['shop_catalog_msgs'] = sent_ids


async def buyer_shop_ai_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Do'kon ichida AI qidiruvni boshlaydi (callback: shop_ai_{seller_id}).
    Keyingi matn xabari faqat shu do'kon mahsulotlari bo'yicha AI orqali qidiriladi."""
    query = update.callback_query
    seller_id = int(query.data.split("_")[2])
    lang = get_lang(update, context)

    if not ai_assistant.is_enabled():
        await query.answer(t(lang, 'ai_disabled_alert'), show_alert=True)
        return
    await query.answer()

    shop = db.get_seller_public_info(seller_id)
    shop_name = (shop.get('shop_name') if shop else None) or t(lang, 'shop_word')

    # AI qidiruv rejimi + do'kon konteksti (topilgan mahsulotni savatga qo'shish uchun)
    context.user_data['ai_chat'] = True
    context.user_data['ai_shop_filter'] = seller_id
    context.user_data['ai_shop_name'] = shop_name
    context.user_data['shop_ctx'] = seller_id
    context.user_data.pop('ai_draft', None)
    ai_assistant.reset_history(context.user_data)

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(t(lang, 'btn_back_to_shop'), callback_data=f"shop_products_{seller_id}_0")],
        [InlineKeyboardButton(t(lang, 'ai_exit'), callback_data="ai_exit")],
    ])
    await context.bot.send_message(
        chat_id=update.effective_chat.id,
        text=t(lang, 'shop_ai_prompt', shop=html.escape(shop_name)),
        reply_markup=kb, parse_mode='HTML'
    )


async def buyer_search(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # 1-bosqich: hudud tanlash yoki o'tkazib yuborish
    context.user_data.pop('shop_ctx', None)  # do'kon konteksti tugadi
    context.user_data['search_state'] = 'awaiting_query'
    context.user_data.pop('search_query', None)
    context.user_data.pop('search_lat', None)
    context.user_data.pop('search_lon', None)
    context.user_data.pop('search_region_id', None)

    lang = get_lang(update, context)
    regions = db.get_regions(parent_id=None)
    kb_rows = []
    for r in regions:
        kb_rows.append([InlineKeyboardButton(region_name(r['name'], lang), callback_data=f"sreg_{r['id']}")])
    kb_rows.append([InlineKeyboardButton(t(lang, 'btn_all_regions'), callback_data="sreg_0")])
    kb_rows.append([InlineKeyboardButton(t(lang, 'back'), callback_data="buyer_panel")])

    text = t(lang, 'search_region_ask')

    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(kb_rows))
    else:
        await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(kb_rows))


async def search_region_select(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Foydalanuvchi viloyat tanladi — tuman ko'rsatamiz yoki to'g'ridan-to'g'ri qidiruv."""
    query = update.callback_query
    await query.answer()

    region_id = int(query.data.replace("sreg_", ""))
    lang = get_lang(update, context)

    if region_id == 0:
        # Barcha hududlar — hududsiz qidiruv
        context.user_data.pop('search_region_id', None)
        await query.edit_message_text(t(lang, 'search_prompt'))
        return

    # Tumanlar bormi?
    districts = db.get_regions(parent_id=region_id)
    if districts:
        context.user_data['search_region_id'] = region_id
        kb_rows = []
        for d in districts:
            kb_rows.append([InlineKeyboardButton(region_name(d['name'], lang), callback_data=f"sdist_{d['id']}")])
        region = db.get_region_by_id(region_id)
        kb_rows.append([InlineKeyboardButton(
            t(lang, 'btn_whole_region', name=region_name(region['name'], lang)), callback_data=f"sdist_0_{region_id}"
        )])
        kb_rows.append([InlineKeyboardButton(t(lang, 'back'), callback_data="buyer_search")])
        await query.edit_message_text(
            t(lang, 'region_pick_district', name=region_name(region['name'], lang)),
            reply_markup=InlineKeyboardMarkup(kb_rows)
        )
    else:
        # Tumansiz viloyat — to'g'ridan-to'g'ri qidiruv
        context.user_data['search_region_id'] = region_id
        region = db.get_region_by_id(region_id)
        await query.edit_message_text(
            t(lang, 'region_then_search', name=region_name(region['name'], lang))
        )


async def search_district_select(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Foydalanuvchi tuman tanladi."""
    query = update.callback_query
    await query.answer()

    data = query.data.replace("sdist_", "")
    lang = get_lang(update, context)

    if "_" in data:
        # "0_region_id" — butun viloyat
        region_id = int(data.split("_")[1])
        context.user_data['search_region_id'] = region_id
        region = db.get_region_by_id(region_id)
        name = region_name(region['name'], lang) if region else t(lang, 'selected_region')
    else:
        district_id = int(data)
        context.user_data['search_region_id'] = district_id
        district = db.get_region_by_id(district_id)
        name = region_name(district['name'], lang) if district else t(lang, 'selected_district')

    await query.edit_message_text(t(lang, 'region_then_search', name=name))


async def buyer_orders(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if query:
        await query.answer()

    lang = get_lang(update, context)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    orders = db.get_orders_by_buyer(user['id'])

    kb = InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="buyer_panel")]])

    if not orders:
        text = t(lang, 'orders_empty')
        if query:
            await query.edit_message_text(text, reply_markup=kb)
        else:
            await update.message.reply_text(text, reply_markup=kb)
        return

    status_emoji = {'pending': '⏳', 'confirmed': '✅', 'delivered': '🚚', 'cancelled': '❌'}

    # Savat buyurtmalarini guruhlash: bir order_group_id bitta qator bo'lib ko'rinadi
    group_agg = {}
    for o in orders:
        gid = o.get('order_group_id')
        if gid:
            g = group_agg.setdefault(gid, {'count': 0, 'sum': 0.0, 'status': o['status']})
            g['count'] += 1
            g['sum'] += float(o['total_price'] or 0)

    keyboard = []
    seen_groups = set()
    shown = 0
    for order in orders:
        if shown >= 15:
            break
        gid = order.get('order_group_id')
        if gid:
            if gid in seen_groups:
                continue
            seen_groups.add(gid)
            g = group_agg[gid]
            emoji = status_emoji.get(g['status'], '❓')
            keyboard.append([InlineKeyboardButton(
                t(lang, 'order_group_row', emoji=emoji, oid=fmt_order_id(int(gid)),
                  count=g['count'], sum=fmt_price(g['sum'])),
                callback_data=f"gorder_detail_{gid}"
            )])
        else:
            emoji = status_emoji.get(order['status'], '❓')
            keyboard.append([InlineKeyboardButton(
                f"{emoji} {fmt_order_id(order['id'])} — {order['product_name'][:25]}",
                callback_data=f"order_detail_{order['id']}"
            )])
        shown += 1
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="buyer_panel")])

    total_display = len(seen_groups) + sum(1 for o in orders if not o.get('order_group_id'))
    text = t(lang, 'my_orders_count', n=total_display)
    if query:
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard))
    else:
        await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard))


async def buyer_messages(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xaridor xabarli buyurtmalarini ko'radi."""
    query = update.callback_query
    if query:
        await query.answer()

    user = db.get_user_by_telegram_id(update.effective_user.id)

    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT DISTINCT o.id, p.name as product_name, u.shop_name,
               MAX(m.created_at) as last_msg,
               COUNT(CASE WHEN m.receiver_id=? AND m.is_read=0 THEN 1 END) as unread
        FROM messages m
        JOIN orders o ON m.order_id=o.id
        JOIN products p ON o.product_id=p.id
        JOIN users u ON o.seller_id=u.id
        WHERE o.buyer_id=?
        GROUP BY o.id
        ORDER BY last_msg DESC
        LIMIT 20
    """, (user['id'], user['id']))
    rows = [dict(r) for r in cursor.fetchall()]
    lang = get_lang(update, context)

    kb_back = InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="buyer_panel")]])

    if not rows:
        text = t(lang, 'messages_empty')
        if query:
            await query.edit_message_text(text, reply_markup=kb_back)
        else:
            await update.message.reply_text(text, reply_markup=kb_back)
        return

    keyboard = []
    for r in rows:
        unread_mark = f" 🔴{r['unread']}" if r['unread'] else ""
        label = f"💬 {fmt_order_id(r['id'])} — {r['product_name'][:20]}{unread_mark}"
        keyboard.append([InlineKeyboardButton(label, callback_data=f"msgs_{r['id']}")])
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="buyer_panel")])

    text = t(lang, 'messages_title')
    if query:
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard))
    else:
        await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard))


# #16 SODIQLIK — app (_loyalty_for) bilan AYNAN bir xil formula va darajalar.
# ball = xarajat/10000 + yetkazilgan_buyurtma*5 + referal*50
LOYALTY_TIERS = [(0, 'bronze', '🥉'), (200, 'silver', '🥈'),
                 (1000, 'gold', '🥇'), (5000, 'diamond', '💎')]


def _loyalty_line(user, lang):
    """Xaridor profiliga qo'shiladigan sodiqlik bloki: ball + daraja + keyingi darajagacha.
    Mavjud ma'lumotdan hisoblanadi (xarid + referal) — alohida schema yo'q."""
    tot = db.get_buyer_order_totals(user['id'])
    spent = float(tot.get('spent') or 0)
    delivered = int(tot.get('delivered_orders') or 0)
    refs = int(user.get('referral_count') or 0)
    points = int(spent // 10000) + delivered * 5 + refs * 50
    cur, nxt = LOYALTY_TIERS[0], None
    for i, tier in enumerate(LOYALTY_TIERS):
        if points >= tier[0]:
            cur = tier
            nxt = LOYALTY_TIERS[i + 1] if i + 1 < len(LOYALTY_TIERS) else None
    if nxt:
        nxt_line = t(lang, 'loyalty_to_next', n=nxt[0] - points, tier=t(lang, 'loy_' + nxt[1]))
    else:
        nxt_line = t(lang, 'loyalty_max')
    return t(lang, 'loyalty_profile', emoji=cur[2], tier=t(lang, 'loy_' + cur[1]),
             points=points, next=nxt_line)


async def buyer_profile(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = db.get_user_by_telegram_id(update.effective_user.id)
    lang = get_lang(update, context)

    keyboard = [
        [InlineKeyboardButton(t(lang, 'btn_edit_name'), callback_data="edit_buyer_name")],
        [InlineKeyboardButton(t(lang, 'btn_edit_phone'), callback_data="edit_buyer_phone")],
        [InlineKeyboardButton(t(lang, 'btn_my_referral'), callback_data="my_referral")],
        [InlineKeyboardButton(t(lang, 'btn_change_language'), callback_data="change_lang")],
        [InlineKeyboardButton(t(lang, 'back'), callback_data="buyer_panel")],
    ]
    text = t(lang, 'buyer_profile_body',
             name=user['name'], phone=user['phone_number'],
             refs=user.get('referral_count') or 0,
             date=fmt_datetime(user['created_at']))
    text += _loyalty_line(user, lang)   # #16 — sodiqlik darajasi (app bilan parite)

    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard))
    else:
        await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard))


async def change_language_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Til tanlash menyusi (profil orqali)."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    active = get_active_mode(user, context)
    if active == 'admin':
        back_cb = "admin_settings"
    elif active == 'seller':
        back_cb = "seller_profile"
    else:
        back_cb = "buyer_profile"
    kb = [
        [InlineKeyboardButton(LANGS['uz'], callback_data="setlang_uz")],
        [InlineKeyboardButton(LANGS['ru'], callback_data="setlang_ru")],
        [InlineKeyboardButton(t(lang, 'back'), callback_data=back_cb)],
    ]
    await query.edit_message_text(
        t(lang, 'choose_language'),
        reply_markup=InlineKeyboardMarkup(kb)
    )


async def set_language(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Tilni tanlaydi va tegishli panelni yangi tilda ko'rsatadi."""
    query = update.callback_query
    new_lang = query.data.split("_")[1]  # "setlang_uz" -> "uz"
    new_lang = set_user_lang(update, context, new_lang)
    try:
        await query.answer(t(new_lang, 'language_changed'))
    except Exception:
        pass
    # Pastki Reply klaviaturani ham yangi tilga moslaymiz
    user = db.get_user_by_telegram_id(update.effective_user.id)
    active = get_active_mode(user, context)
    if active == 'seller':
        await seller_panel(update, context)
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=t(new_lang, 'bottom_hint'),
            reply_markup=seller_bottom_kb(new_lang)
        )
    elif user and user.get('role') == 'admin':
        await admin_panel(update, context)
    else:
        await buyer_panel(update, context)
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=t(new_lang, 'bottom_hint'),
            reply_markup=buyer_bottom_kb(new_lang)
        )


async def my_referral(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Foydalanuvchi o'z taklif havolasini ko'radi va ulashishi mumkin."""
    query = update.callback_query
    await query.answer()

    lang = get_lang(update, context)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    if not user:
        await query.edit_message_text(t(lang, 'user_not_found_start'))
        return

    # Agar referral_code yo'q bo'lsa — hozir yaratib saqlaymiz
    if not user.get('referral_code'):
        import random, string
        code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
        db.update_user(user['id'], referral_code=code)
        user = db.get_user_by_telegram_id(update.effective_user.id)

    from urllib.parse import quote
    bot_me = await context.bot.get_me()
    bot_username = bot_me.username
    ref_link = f"https://t.me/{bot_username}?start={user['referral_code']}"

    # t.me/share/url — faqat ref_link ni encode qilamiz, matn sodda bo'lsin
    share_url = f"https://t.me/share/url?url={quote(ref_link, safe='')}&text={quote(t(lang, 'referral_share_text'), safe='')}"

    text = t(lang, 'referral_link_title',
             link=ref_link, code=user['referral_code'],
             count=user.get('referral_count') or 0)

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(t(lang, 'btn_share_friends'), url=share_url)],
        [InlineKeyboardButton(t(lang, 'back'), callback_data="buyer_profile")],
    ])

    await query.edit_message_text(text, reply_markup=kb, parse_mode='HTML')


async def buyer_order_detail(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    order_id = int(query.data.split("_")[2])
    order = db.get_order_by_id(order_id)

    lang = get_lang(update, context)
    if not order:
        await query.edit_message_text(t(lang, 'order_not_found'))
        return

    # XAVFSIZLIK: bu kartada sotuvchi kontakti/manzili bor — faqat shu buyurtma xaridori (yoki admin) ko'rsin
    actor = db.get_user_by_telegram_id(update.effective_user.id) if update.effective_user else None
    is_owner = bool(actor and actor.get('id') == order.get('buyer_id'))
    is_admin = (update.effective_user and update.effective_user.id == ADMIN_ID) or \
               (actor and actor.get('role') == 'admin')
    if not (is_owner or is_admin):
        logging.warning(
            f"Ruxsatsiz buyurtma ko'rishga urinish (buyer): user_tg={getattr(update.effective_user, 'id', None)} "
            f"order_id={order_id} (xaridori buyer_id={order.get('buyer_id')})"
        )
        await query.edit_message_text(t(lang, 'not_your_order_plain'))
        return

    dlv = order.get('delivery_type', 'delivery')
    pay = order.get('payment_method', 'cash')
    status = order['status']

    # Status tavsifi — xaridorga tushunarli
    # Pending bo'lsa qolgan vaqtni hisoblaymiz
    pending_note = ""
    if status == 'pending' and order.get('created_at'):
        from datetime import datetime, timezone
        try:
            now = datetime.now(timezone.utc)
            # Real muddat DB'dagi auto_cancel_at'dan (optom=30 daq, oddiy=10 daq); yo'q bo'lsa created_at+TTL
            _dl = _order_deadline(order)
            remaining = max(0, (_dl - now).total_seconds()) if _dl else 0
            if remaining > 0:
                mins = int(remaining // 60)
                secs = int(remaining % 60)
                pending_note = t(lang, 'pending_autocancel_left', m=mins, s=f"{secs:02d}")
            else:
                pending_note = t(lang, 'pending_autocancel_soon')
        except Exception:
            pass

    status_guide = {
        'pending':   t(lang, 'status_guide_pending', note=pending_note),
        'confirmed': (t(lang, 'status_guide_confirmed_delivery') if dlv == 'delivery'
                      else t(lang, 'status_guide_confirmed_pickup')),
        'delivered': t(lang, 'status_guide_delivered'),
        'cancelled': t(lang, 'status_guide_cancelled'),
    }

    # Bekor qilish jarayoni ketayotgan bo'lsa — holat izohiga eslatma qo'shamiz
    cstate = order.get('cancel_state') or ''
    if status == 'confirmed' and cstate:
        if cstate == 'disputed':
            status_guide['confirmed'] += t(lang, 'cancel_note_disputed')
        elif cstate == 'requested':
            if order.get('cancel_by') == 'buyer':
                status_guide['confirmed'] += t(lang, 'cancel_note_waiting')
            else:
                status_guide['confirmed'] += t(lang, 'cancel_note_incoming')
    if status == 'cancelled' and order.get('cancel_reason'):
        status_guide['cancelled'] += t(lang, 'cancel_note_reason',
                                       reason=cancel_reason_display(order.get('cancel_reason'), lang))
    if status == 'confirmed' and order.get('buyer_received'):
        status_guide['confirmed'] += "\n" + t(lang, 'buyer_awaiting_finalize')

    # Holat ketma-ketligi vizual
    steps = [t(lang, 'step_new'), t(lang, 'step_confirmed'),
             t(lang, 'step_delivered') if dlv == 'delivery' else t(lang, 'step_picked'),
             t(lang, 'step_rated')]
    step_idx = {'pending': 0, 'confirmed': 1, 'delivered': 2, 'cancelled': 0}.get(status, 0)
    timeline = ""
    for i, step in enumerate(steps):
        if i < step_idx:
            timeline += f"✅ {step}\n"
        elif i == step_idx and status != 'cancelled':
            timeline += f"▶️ {step}{t(lang, 'timeline_now')}\n"
        else:
            timeline += f"⬜ {step}\n"

    keyboard = []

    if status == 'pending':
        keyboard.append([InlineKeyboardButton(
            t(lang, 'btn_cancel_order'), callback_data=f"buyer_cancel_{order_id}"
        )])

    # Pickup: xaridor o'zi "Oldim" tugmasini bosadi. DIQQAT: bu buyurtmani YOPMAYDI —
    # to'lovni sotuvchi yakunlaydi. Shu sababli «oldim» bosilgach tugma qayta ko'rsatilmaydi.
    if status == 'confirmed' and dlv == 'pickup' and not order.get('buyer_received'):
        keyboard.append([InlineKeyboardButton(
            t(lang, 'btn_got_item'), callback_data=f"buyer_confirm_pickup_{order_id}"
        )])

    # Tasdiqlangan shartnomani bekor qilish oqimi (xaridor tomoni)
    if status == 'confirmed':
        cstate = order.get('cancel_state') or ''
        if not cstate:
            keyboard.append([InlineKeyboardButton(
                t(lang, 'btn_request_cancel'), callback_data=f"ccl_req_{order_id}"
            )])
        elif cstate == 'requested' and order.get('cancel_by') == 'seller':
            # Sotuvchi bekor qilishni so'ragan — xaridor javob beradi
            keyboard.append([InlineKeyboardButton(t(lang, 'btn_cancel_agree'), callback_data=f"cclagree_{order_id}")])
            keyboard.append([InlineKeyboardButton(t(lang, 'btn_cancel_deny'), callback_data=f"ccldeny_{order_id}")])

    if status in ('delivered', 'cancelled'):
        keyboard.append([InlineKeyboardButton(
            t(lang, 'btn_reorder'), callback_data=f"order_{order['product_id']}"
        )])

    keyboard.append([InlineKeyboardButton(
        t(lang, 'btn_send_message'), callback_data=f"order_msg_{order_id}"
    )])
    keyboard.append([InlineKeyboardButton(
        t(lang, 'btn_correspondence'), callback_data=f"msgs_{order_id}"
    )])

    # Reyting: delivery — faqat 'delivered' da; pickup — 'confirmed' da ham mumkin
    # Reyting faqat buyurtma YAKUNLANGACH (sotuvchi to'lovni belgilab 'delivered' qilgach).
    # «oldim» bosilganda emas — aks holda reyting ikki marta so'ralardi.
    can_rate = (status == 'delivered')
    if can_rate:
        keyboard.append([InlineKeyboardButton(
            t(lang, 'btn_leave_rating'), callback_data=f"order_rate_{order_id}"
        )])

    dlv_lbl = dlv_label(dlv, lang)
    pay_lbl = pay_label(pay, lang)

    # Sotuvchi lokatsiyasi va masofa
    location_line = ""
    map_button = None
    if order.get('shop_lat') and order.get('shop_lon'):
        slat, slon = order['shop_lat'], order['shop_lon']
        location_line = t(lang, 'frag_order_address',
                          addr=html.escape(order.get('shop_address') or t(lang, 'address_word')))
        if order.get('shop_landmark'):
            location_line += t(lang, 'frag_order_landmark', lm=html.escape(order['shop_landmark']))

        # Masofa — agar xaridorning lokatsiyasi bo'lsa
        if order.get('buyer_lat') and order.get('buyer_lon'):
            dist = haversine_km(order['buyer_lat'], order['buyer_lon'], slat, slon)
            if dist is not None:
                location_line += t(lang, 'frag_order_distance', km=dist)
            # Yo'l xaritasi — Google Maps directions
            map_button = InlineKeyboardButton(
                t(lang, 'btn_show_route'),
                url=f"https://www.google.com/maps/dir/{order['buyer_lat']},{order['buyer_lon']}/{slat},{slon}"
            )
        else:
            # Faqat sotuvchi lokatsiyasi
            map_button = InlineKeyboardButton(
                t(lang, 'btn_shop_location'),
                url=f"https://www.google.com/maps/search/?api=1&query={slat},{slon}"
            )

    # Xarita tugmasini klaviaturaga qo'shish
    if map_button:
        keyboard.append([map_button])

    # Sotuvchi Telegram
    if order.get('seller_username'):
        keyboard.append([InlineKeyboardButton(
            t(lang, 'btn_tg_at', u=order['seller_username']),
            url=f"https://t.me/{order['seller_username']}"
        )])

    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="buyer_orders")])

    text = t(lang, 'order_detail_body',
             oid=fmt_order_id(order['id']),
             pname=html.escape(order.get('product_name') or ''),
             qty=order['quantity'], total=fmt_price(order['total_price']),
             dlv=dlv_lbl, pay=pay_lbl,
             shop=html.escape(order.get('shop_name') or ''),
             location=location_line, phone=order.get('seller_phone') or '—',
             date=fmt_datetime(order.get('created_at')),
             timeline=timeline, guide=status_guide.get(status, ''))
    _sb = _settlement_badge(lang, order)
    if _sb:
        text += "\n" + _sb

    await query.edit_message_text(
        text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML'
    )


async def buyer_confirm_pickup(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xaridor pickup buyurtmasini o'zi olganligi haqida tasdiqlaydi."""
    query = update.callback_query
    await query.answer()

    order_id = int(query.data.split("_")[3])
    order = db.get_order_by_id(order_id)
    buyer = db.get_user_by_telegram_id(update.effective_user.id)

    lang = get_lang(update, context)
    if not order or not buyer or order['buyer_id'] != buyer['id']:
        await query.edit_message_text(t(lang, 'order_not_yours'))
        return

    if order['status'] != 'confirmed':
        await query.edit_message_text(
            t(lang, 'cant_confirm_status', status=fmt_status(order['status']))
        )
        return

    # MUHIM: xaridor «oldim» bossa ham buyurtma YOPILMAYDI. Status 'confirmed' qoladi —
    # sotuvchi to'lov holatini (to'liq/qarz/bo'lib) belgilab, yakunlashi shart.
    db.set_buyer_received(order_id)

    # Sotuvchiga xabar (sotuvchi tilida) — to'lovni belgilab yakunlashga chaqiramiz
    try:
        if order.get('seller_tg'):
            seller = db.get_user_by_id(order['seller_id'])
            slang = get_user_lang(seller) if seller else DEFAULT_LANG
            skb = InlineKeyboardMarkup([[InlineKeyboardButton(
                t(slang, 'btn_finalize_payment'), callback_data=f"seller_order_{order_id}")]])
            await context.bot.send_message(
                chat_id=order['seller_tg'],
                text=t(slang, 'pickup_seller_finalize',
                       oid=fmt_order_id(order_id),
                       pname=html.escape(order.get('product_name') or ''),
                       buyer=html.escape(order.get('buyer_name') or '')),
                reply_markup=skb,
                parse_mode='HTML'
            )
    except Exception as e:
        logging.error(f"Pickup tasdiqlash bildirishnomasi ketmadi: {e}")

    await query.edit_message_text(
        t(lang, 'pickup_received_buyer', oid=fmt_order_id(order_id)),
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton(t(lang, 'btn_orders_back'), callback_data="buyer_orders")],
        ]),
        parse_mode='HTML'
    )


# ============================================================
# SHARTNOMANI BEKOR QILISH (kelishuv + nizo)
# Tasdiqlangan (confirmed) buyurtma bir tomonlama bekor qilinmaydi:
#   1) bir tomon sabab bilan so'raydi  -> cancel_state='requested'
#   2) ikkinchi tomon rozi bo'lsa       -> 'cancelled'
#      rozi bo'lmasa                     -> 'disputed' (admin hakam)
#   3) admin qaror chiqaradi: bekor yoki kuchda qoldirish
# ============================================================

# Sabab kodlari — tarjima kalitlari 'crsn_<code>' ko'rinishida (languages.py)
BUYER_CANCEL_REASONS = ['bchg', 'bprice', 'bfound', 'blate', 'bnoreach']
SELLER_CANCEL_REASONS = ['sstock', 'sprice', 'snoreach', 'snoaddr', 'snopay']


def _order_party(order, user):
    """Foydalanuvchi shu buyurtmada kim: 'buyer' | 'seller' | None."""
    if not user or not order:
        return None
    if user['id'] == order['buyer_id']:
        return 'buyer'
    if user['id'] == order['seller_id']:
        return 'seller'
    return None


def cancel_reason_display(reason, lang):
    """Saqlangan sababni o'qiladigan, til-aware matnga aylantiradi.
    'code:<kod>' -> tarjima; 'text:<matn>' -> erkin matn."""
    if not reason:
        return t(lang, 'crsn_unknown')
    if reason.startswith('code:'):
        return t(lang, 'crsn_' + reason[5:])
    if reason.startswith('text:'):
        return html.escape(reason[5:])
    return html.escape(reason)


async def _maybe_restock_on_cancel(context, order):
    """Tasdiqlangan buyurtma bekor qilinganda zaxirani qaytaradi (xato yutiladi)."""
    try:
        db.restock_on_cancel(order['product_id'], order.get('quantity') or 1)
    except Exception as e:
        logging.error(f"Bekorda zaxirani qaytarish xatosi: {e}")


async def cancel_request_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """'Bekor qilishni so'rash' tugmasi -> sabab tanlash (ccl_req_<oid>)."""
    query = update.callback_query
    await query.answer()
    order_id = int(query.data.split("_")[2])
    order = db.get_order_by_id(order_id)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    lang = get_lang(update, context)

    party = _order_party(order, user)
    if not order or not party:
        await query.edit_message_text(t(lang, 'order_not_yours'))
        return ConversationHandler.END
    if order['status'] != 'confirmed' or (order.get('cancel_state') or ''):
        await query.edit_message_text(t(lang, 'cancel_not_available'))
        return ConversationHandler.END

    context.user_data['cancel_order_id'] = order_id
    context.user_data['cancel_party'] = party

    keyboard = []
    # AI taklif qilgan kontekstli sabablar (matn callback_data'ga sig'maydi → indeks bilan)
    try:
        ai_reasons = await ai_assistant.suggest_cancel_reasons(
            party=party, product_name=order.get('product_name') or '',
            status=order.get('status') or '', lang=lang)
    except Exception as e:
        logging.warning(f"AI bekor sabab taklif xato: {e}")
        ai_reasons = []
    context.user_data['cancel_ai_reasons'] = ai_reasons
    for i, r in enumerate(ai_reasons):
        keyboard.append([InlineKeyboardButton(f"🤖 {r}"[:60], callback_data=f"ccl_air_{i}")])

    codes = BUYER_CANCEL_REASONS if party == 'buyer' else SELLER_CANCEL_REASONS
    keyboard += [[InlineKeyboardButton(t(lang, 'crsn_' + c), callback_data=f"ccl_rsn_{c}")]
                 for c in codes]
    keyboard.append([InlineKeyboardButton(t(lang, 'crsn_other'), callback_data="ccl_rsn_other")])
    keyboard.append([InlineKeyboardButton(t(lang, 'btn_cancel'), callback_data="ccl_abort")])

    await query.edit_message_text(
        t(lang, 'cancel_pick_reason', oid=fmt_order_id(order_id)),
        reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML'
    )
    return CANCEL_PICK_REASON


async def cancel_reason_pick(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Sabab tugmasi tanlandi (ccl_rsn_<code> yoki ccl_abort)."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)

    if query.data == "ccl_abort":
        await query.edit_message_text(t(lang, 'cancel_aborted'))
        context.user_data.pop('cancel_order_id', None)
        context.user_data.pop('cancel_party', None)
        context.user_data.pop('cancel_ai_reasons', None)
        return ConversationHandler.END

    # AI taklif qilgan sabab (indeks bo'yicha) → erkin matn sifatida saqlanadi
    if query.data.startswith("ccl_air_"):
        try:
            idx = int(query.data.rsplit("_", 1)[1])
            ai_reasons = context.user_data.get('cancel_ai_reasons') or []
            chosen = ai_reasons[idx]
        except (ValueError, IndexError):
            await query.edit_message_text(t(lang, 'cancel_not_available'))
            return ConversationHandler.END
        await _finalize_cancel_request(update, context, 'text:' + chosen[:300], via_query=True)
        context.user_data.pop('cancel_ai_reasons', None)
        return ConversationHandler.END

    code = query.data.split("_", 2)[2]
    if code == 'other':
        await query.edit_message_text(t(lang, 'cancel_reason_ask'))
        return CANCEL_REASON_TEXT

    await _finalize_cancel_request(update, context, 'code:' + code, via_query=True)
    return ConversationHandler.END


async def cancel_reason_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """'Boshqa sabab' uchun erkin matn."""
    text = (update.message.text or '').strip()
    lang = get_lang(update, context)
    if not text:
        await update.message.reply_text(t(lang, 'cancel_reason_ask'))
        return CANCEL_REASON_TEXT
    await _finalize_cancel_request(update, context, 'text:' + text[:300], via_query=False)
    return ConversationHandler.END


async def _finalize_cancel_request(update, context, reason, via_query):
    """Bekor so'rovini saqlaydi va ikkinchi tomonga rozilik so'rovini yuboradi."""
    order_id = context.user_data.get('cancel_order_id')
    party = context.user_data.get('cancel_party')
    lang = get_lang(update, context)
    order = db.get_order_by_id(order_id) if order_id else None

    async def _reply(text, kb=None):
        if via_query:
            await update.callback_query.edit_message_text(text, reply_markup=kb, parse_mode='HTML')
        else:
            await update.message.reply_text(text, reply_markup=kb, parse_mode='HTML')

    if not order or not party:
        await _reply(t(lang, 'order_not_found'))
        return

    gid = order.get('order_group_id')
    if gid:
        ok = db.request_group_cancel(gid, party, reason)
    else:
        ok = db.request_order_cancel(order_id, party, reason)
    if not ok:
        await _reply(t(lang, 'cancel_not_available'))
        return

    await _reply(t(lang, 'cancel_requested_sent', oid=fmt_order_id(order_id)))

    # Ikkinchi tomonga rozilik so'rovi (uning tilida)
    if party == 'buyer':
        other_tg, other = order.get('seller_tg'), db.get_user_by_id(order['seller_id'])
    else:
        other_tg, other = order.get('buyer_tg'), db.get_user_by_id(order['buyer_id'])
    olang = get_user_lang(other) if other else DEFAULT_LANG

    if other_tg:
        try:
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton(t(olang, 'btn_cancel_agree'), callback_data=f"cclagree_{order_id}")],
                [InlineKeyboardButton(t(olang, 'btn_cancel_deny'), callback_data=f"ccldeny_{order_id}")],
            ])
            await context.bot.send_message(
                chat_id=other_tg,
                text=t(olang, 'cancel_request_notify',
                       oid=fmt_order_id(order_id),
                       pname=html.escape(order.get('product_name') or ''),
                       reason=cancel_reason_display(reason, olang)),
                reply_markup=kb, parse_mode='HTML'
            )
        except Exception as e:
            logging.error(f"Bekor so'rovi bildirishnomasi ketmadi: {e}")

    context.user_data.pop('cancel_order_id', None)
    context.user_data.pop('cancel_party', None)


async def cancel_respond(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Ikkinchi tomon javobi: cclagree_<oid> (rozi) yoki ccldeny_<oid> (rozi emas)."""
    query = update.callback_query
    await query.answer()
    action, _, oid_s = query.data.partition("_")
    order_id = int(oid_s)
    order = db.get_order_by_id(order_id)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    lang = get_lang(update, context)

    party = _order_party(order, user)
    if not order or not party:
        await query.edit_message_text(t(lang, 'order_not_yours'))
        return
    if order.get('cancel_state') != 'requested':
        await query.edit_message_text(t(lang, 'cancel_already_handled'))
        return
    # Faqat so'ramagan tomon javob beradi
    if order.get('cancel_by') == party:
        await query.answer(t(lang, 'cancel_wait_other'), show_alert=True)
        return

    requester = db.get_user_by_id(order['buyer_id'] if order.get('cancel_by') == 'buyer' else order['seller_id'])
    rlang = get_user_lang(requester) if requester else DEFAULT_LANG
    req_tg = order.get('buyer_tg') if order.get('cancel_by') == 'buyer' else order.get('seller_tg')
    oid = fmt_order_id(order_id)
    pname = html.escape(order.get('product_name') or '')

    gid = order.get('order_group_id')
    if action == "cclagree":
        if gid:
            # Variant/savat guruh — butun guruh bekor bo'ladi, har biriga zahira qaytadi
            for o in db.agree_group_cancel(gid):
                od = db.get_order_by_id(o)
                if od:
                    await _maybe_restock_on_cancel(context, od)
        elif db.agree_order_cancel(order_id):
            await _maybe_restock_on_cancel(context, order)
        await query.edit_message_text(t(lang, 'cancel_agreed_done', oid=oid), parse_mode='HTML')
        if req_tg:
            try:
                await context.bot.send_message(
                    chat_id=req_tg,
                    text=t(rlang, 'cancel_agreed_notify', oid=oid, pname=pname),
                    parse_mode='HTML')
            except Exception as e:
                logging.error(f"Bekor roziligi bildirishnomasi ketmadi: {e}")
    else:  # ccldeny
        if gid:
            db.dispute_group_cancel(gid)
        else:
            db.dispute_order_cancel(order_id)
        await query.edit_message_text(t(lang, 'cancel_denied_done', oid=oid), parse_mode='HTML')
        if req_tg:
            try:
                await context.bot.send_message(
                    chat_id=req_tg,
                    text=t(rlang, 'cancel_denied_notify', oid=oid, pname=pname),
                    parse_mode='HTML')
            except Exception as e:
                logging.error(f"Bekor rad bildirishnomasi ketmadi: {e}")
        # Admin'ga nizo haqida xabar
        await _notify_admin_dispute(context, order_id)


async def _notify_admin_dispute(context, order_id):
    """Admin'ga yangi nizo haqida bildirishnoma + tezkor tugma."""
    order = db.get_order_by_id(order_id)
    if not order:
        return
    try:
        alang = DEFAULT_LANG
        admin_user = db.get_user_by_telegram_id(ADMIN_ID)
        if admin_user:
            alang = get_user_lang(admin_user)
        kb = InlineKeyboardMarkup([[
            InlineKeyboardButton(t(alang, 'btn_open_dispute'), callback_data=f"admin_disp_{order_id}")
        ]])
        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=t(alang, 'admin_dispute_notify',
                   oid=fmt_order_id(order_id),
                   pname=html.escape(order.get('product_name') or ''),
                   by=t(alang, 'party_buyer' if order.get('cancel_by') == 'buyer' else 'party_seller'),
                   reason=cancel_reason_display(order.get('cancel_reason'), alang)),
            reply_markup=kb, parse_mode='HTML')
    except Exception as e:
        logging.error(f"Admin'ga nizo bildirishnomasi ketmadi: {e}")


async def admin_disputes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin — nizodagi (disputed) buyurtmalar ro'yxati."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    disputes = db.get_disputed_orders()
    if not disputes:
        await query.edit_message_text(
            t(lang, 'no_disputes'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="admin_panel")]]))
        return
    keyboard = []
    for o in disputes[:20]:
        label = f"⚖️ {fmt_order_id(o['id'])} — {(o.get('product_name') or '')}"[:45]
        keyboard.append([InlineKeyboardButton(label, callback_data=f"admin_disp_{o['id']}")])
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="admin_panel")])
    await query.edit_message_text(
        t(lang, 'disputes_header', n=len(disputes)),
        reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


async def admin_dispute_detail(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Bitta nizo detali + qaror tugmalari (admin_disp_<oid>)."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    order_id = int(query.data.split("_")[2])
    order = db.get_order_by_id(order_id)
    if not order or order.get('cancel_state') != 'disputed':
        await query.edit_message_text(
            t(lang, 'dispute_not_found'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="admin_disputes")]]))
        return
    by = t(lang, 'party_buyer' if order.get('cancel_by') == 'buyer' else 'party_seller')
    text = t(lang, 'dispute_detail_body',
             oid=fmt_order_id(order_id),
             pname=html.escape(order.get('product_name') or ''),
             qty=order.get('quantity'), total=fmt_price(order.get('total_price')),
             buyer=html.escape(order.get('buyer_name') or ''),
             bphone=fmt_phone(order.get('buyer_phone')),
             seller=html.escape(order.get('shop_name') or order.get('seller_name') or ''),
             sphone=fmt_phone(order.get('seller_phone')),
             by=by, reason=cancel_reason_display(order.get('cancel_reason'), lang))

    keyboard = []
    # Admin -> tomonga bot orqali xabar yozish (doim ishlaydi — telegram_id bor)
    keyboard.append([
        InlineKeyboardButton(t(lang, 'btn_contact_buyer'), callback_data=f"admindm_buyer_{order_id}"),
        InlineKeyboardButton(t(lang, 'btn_contact_seller'), callback_data=f"admindm_seller_{order_id}"),
    ])
    keyboard += [
        [InlineKeyboardButton(t(lang, 'btn_dispute_cancel'), callback_data=f"adisp_cancel_{order_id}")],
        [InlineKeyboardButton(t(lang, 'btn_dispute_keep'), callback_data=f"adisp_keep_{order_id}")],
        [InlineKeyboardButton(t(lang, 'btn_dispute_messages'), callback_data=f"admin_dispmsgs_{order_id}")],
        [InlineKeyboardButton(t(lang, 'btn_correspondence'), callback_data=f"msgs_{order_id}")],
        [InlineKeyboardButton(t(lang, 'back'), callback_data="admin_disputes")],
    ]
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


async def admin_resolve_dispute(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin qarori: adisp_cancel_<oid> (bekor) | adisp_keep_<oid> (kuchda qoldirish)."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    parts = query.data.split("_")
    decision = parts[1]          # cancel | keep
    order_id = int(parts[2])
    order = db.get_order_by_id(order_id)
    if not order or order.get('cancel_state') != 'disputed':
        await query.edit_message_text(
            t(lang, 'dispute_not_found'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="admin_disputes")]]))
        return

    do_cancel = (decision == 'cancel')
    db.resolve_order_dispute(order_id, do_cancel)
    if do_cancel:
        await _maybe_restock_on_cancel(context, order)

    oid = fmt_order_id(order_id)
    pname = html.escape(order.get('product_name') or '')
    # Ikkala tomonga qaror haqida bildirishnoma (har biri o'z tilida)
    for uid_key, tg_key in (('buyer_id', 'buyer_tg'), ('seller_id', 'seller_tg')):
        u = db.get_user_by_id(order[uid_key])
        ulang = get_user_lang(u) if u else DEFAULT_LANG
        tg = order.get(tg_key)
        if not tg:
            continue
        key = 'dispute_resolved_cancel' if do_cancel else 'dispute_resolved_keep'
        try:
            await context.bot.send_message(chat_id=tg, text=t(ulang, key, oid=oid, pname=pname), parse_mode='HTML')
        except Exception as e:
            logging.error(f"Nizo qarori bildirishnomasi ketmadi: {e}")

    await query.edit_message_text(
        t(lang, 'dispute_resolved_admin_cancel' if do_cancel else 'dispute_resolved_admin_keep', oid=oid),
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="admin_disputes")]]))


async def admin_dispute_msg_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin nizo bo'yicha tomonga bot orqali xabar yozadi (admindm_buyer_<oid> | admindm_seller_<oid>)."""
    query = update.callback_query
    lang = get_lang(update, context)
    # Xavfsizlik: bu callback'ni ConversationHandler ushlaydi, shuning uchun
    # admin tekshiruvini shu yerda bajaramiz (button_handler dagi gate ishlamaydi).
    uid = update.effective_user.id
    _u = db.get_user_by_telegram_id(uid)
    if not (_u and (_u.get('role') == 'admin' or uid == ADMIN_ID)):
        await query.answer(t(lang, 'admin_only_action'), show_alert=True)
        return ConversationHandler.END
    await query.answer()
    parts = query.data.split("_")
    target = parts[1]            # buyer | seller
    order_id = int(parts[2])
    order = db.get_order_by_id(order_id)
    if not order:
        await query.edit_message_text(t(lang, 'order_not_found'))
        return ConversationHandler.END

    context.user_data['adm_msg_target'] = target
    context.user_data['adm_msg_order'] = order_id
    who = t(lang, 'party_buyer' if target == 'buyer' else 'party_seller')
    await query.edit_message_text(
        t(lang, 'admin_dm_ask', who=who, oid=fmt_order_id(order_id)),
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton(t(lang, 'btn_cancel'), callback_data=f"admin_disp_{order_id}")
        ]]), parse_mode='HTML')
    return ADMIN_DISPUTE_MSG


async def admin_dispute_msg_send(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin yozgan matnni tegishli tomonga (bot orqali) yetkazadi."""
    lang = get_lang(update, context)
    target = context.user_data.pop('adm_msg_target', None)
    order_id = context.user_data.pop('adm_msg_order', None)
    text = (update.message.text or '').strip()
    if not target or not order_id:
        return ConversationHandler.END
    order = db.get_order_by_id(order_id)
    if not order:
        await update.message.reply_text(t(lang, 'order_not_found'))
        return ConversationHandler.END

    if target == 'buyer':
        tg, u = order.get('buyer_tg'), db.get_user_by_id(order['buyer_id'])
    else:
        tg, u = order.get('seller_tg'), db.get_user_by_id(order['seller_id'])
    ulang = get_user_lang(u) if u else DEFAULT_LANG

    ok = False
    if tg:
        try:
            reply_kb = InlineKeyboardMarkup([[
                InlineKeyboardButton(t(ulang, 'btn_reply_admin'), callback_data=f"dmreply_{order_id}")
            ]])
            await context.bot.send_message(
                chat_id=tg,
                text=t(ulang, 'admin_dm_notify', oid=fmt_order_id(order_id), msg=html.escape(text)),
                reply_markup=reply_kb, parse_mode='HTML')
            ok = True
        except Exception as e:
            logging.error(f"Admin nizo xabari yetkazilmadi: {e}")

    # Audit uchun saqlaymiz (yetkazilgan-yetkazilmaganidan qat'i nazar)
    actor = db.get_user_by_telegram_id(update.effective_user.id)
    db.add_dispute_message(order_id, target, 'admin',
                           actor['id'] if actor else None,
                           (actor.get('name') if actor else None) or 'Admin', text)

    await update.message.reply_text(
        t(lang, 'admin_dm_sent' if ok else 'admin_dm_failed'),
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton(t(lang, 'back'), callback_data=f"admin_disp_{order_id}")
        ]]))
    return ConversationHandler.END


async def dispute_reply_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xaridor/sotuvchi admin xabariga javob beradi (dmreply_<oid>)."""
    query = update.callback_query
    await query.answer()
    order_id = int(query.data.split("_")[1])
    lang = get_lang(update, context)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    order = db.get_order_by_id(order_id)
    party = _order_party(order, user)
    if not order or not party:
        await query.edit_message_text(t(lang, 'order_not_yours'))
        return ConversationHandler.END

    context.user_data['dmreply_order'] = order_id
    context.user_data['dmreply_party'] = party
    back_cb = f"order_detail_{order_id}" if party == 'buyer' else f"seller_order_{order_id}"
    await query.edit_message_text(
        t(lang, 'dmreply_ask', oid=fmt_order_id(order_id)),
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton(t(lang, 'btn_cancel'), callback_data=back_cb)
        ]]), parse_mode='HTML')
    return DMREPLY_MSG


async def dispute_reply_send(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Tomonning javobini admin'ga yetkazadi (admin yana javob bera oladi)."""
    lang = get_lang(update, context)
    order_id = context.user_data.pop('dmreply_order', None)
    party = context.user_data.pop('dmreply_party', None)
    text = (update.message.text or '').strip()
    if not order_id or not party:
        return ConversationHandler.END
    order = db.get_order_by_id(order_id)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    if not order or not user:
        return ConversationHandler.END

    sender_name = user.get('shop_name') if party == 'seller' else user.get('name')
    # Audit uchun saqlaymiz
    db.add_dispute_message(order_id, party, party, user['id'], sender_name or '', text)
    admin_user = db.get_user_by_telegram_id(ADMIN_ID)
    alang = get_user_lang(admin_user) if admin_user else DEFAULT_LANG
    who = t(alang, 'party_buyer' if party == 'buyer' else 'party_seller')

    try:
        kb = InlineKeyboardMarkup([[
            InlineKeyboardButton(t(alang, 'btn_reply'), callback_data=f"admindm_{party}_{order_id}")
        ]])
        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=t(alang, 'dmreply_notify', oid=fmt_order_id(order_id), who=who,
                   name=html.escape(sender_name or ''), msg=html.escape(text)),
            reply_markup=kb, parse_mode='HTML')
    except Exception as e:
        logging.error(f"Adminga javob yetkazilmadi: {e}")

    await update.message.reply_text(t(lang, 'dmreply_sent'))
    return ConversationHandler.END


async def admin_dispute_messages(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin — nizo yozishmalari tarixi (audit). admin_dispmsgs_<oid>."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    order_id = int(query.data.split("_")[2])
    order = db.get_order_by_id(order_id)
    back_cb = (f"admin_disp_{order_id}"
               if (order and order.get('cancel_state') == 'disputed')
               else f"admin_order_{order_id}")
    msgs = db.get_dispute_messages(order_id)
    if not msgs:
        await query.edit_message_text(
            t(lang, 'no_dispute_messages'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data=back_cb)]]))
        return
    lines = [t(lang, 'dispute_messages_header', oid=fmt_order_id(order_id))]
    for m in msgs[-40:]:
        role, party = m.get('sender_role'), m.get('party')
        if role == 'admin':
            arrow = t(lang, 'dm_admin_to_buyer' if party == 'buyer' else 'dm_admin_to_seller')
        else:
            arrow = t(lang, 'dm_buyer_to_admin' if role == 'buyer' else 'dm_seller_to_admin')
        lines.append(f"{arrow} · {fmt_datetime(m.get('created_at'))}\n{html.escape(m.get('message') or '')}\n")
    text = "\n".join(lines)
    if len(text) > 3800:
        text = text[:3800] + t(lang, 'old_messages_cut')
    await query.edit_message_text(
        text,
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data=back_cb)]]),
        parse_mode='HTML')


async def buyer_cancel_order(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xaridor 'pending' buyurtmasini bekor qilishi."""
    query = update.callback_query
    await query.answer()

    order_id = int(query.data.split("_")[2])
    order = db.get_order_by_id(order_id)
    buyer = db.get_user_by_telegram_id(update.effective_user.id)

    lang = get_lang(update, context)
    if not order or not buyer or order['buyer_id'] != buyer['id']:
        await query.edit_message_text(t(lang, 'order_not_yours'))
        return

    if order['status'] != 'pending':
        await query.edit_message_text(
            t(lang, 'cant_cancel_status', status=fmt_status(order['status']))
        )
        return

    # ATOMIK: sotuvchi ayni damda tasdiqlab ulgursa, eski bekor 'confirmed'ni bosib
    # o'tkazib zahirani yo'qotmasin — faqat hali 'pending' bo'lsa bekor qilamiz.
    if not db.transition_order_status(order_id, 'cancelled', 'pending', cancel_by='buyer'):
        cur = db.get_order_by_id(order_id)
        await query.edit_message_text(
            t(lang, 'cant_cancel_status', status=fmt_status((cur or {}).get('status') or 'confirmed'))
        )
        return

    # Avtomatik bekor qilish taymerini o'chiramiz
    if context.application.job_queue:
        jobs = context.application.job_queue.get_jobs_by_name(f"auto_cancel_{order_id}")
        for job in jobs:
            job.schedule_removal()

    await query.edit_message_text(
        t(lang, 'order_cancelled_done', oid=fmt_order_id(order_id)),
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton(t(lang, 'btn_orders_back'), callback_data="buyer_orders")
        ]])
    )

    # Sotuvchiga bildirishnoma (sotuvchi tilida)
    try:
        if order.get('seller_tg'):
            seller = db.get_user_by_id(order['seller_id'])
            await context.bot.send_message(
                chat_id=order['seller_tg'],
                text=t(seller or 'uz', 'seller_notify_cancelled',
                       oid=fmt_order_id(order_id),
                       pname=html.escape(order.get('product_name') or ''))
            )
    except Exception as e:
        logging.error(f"Sotuvchiga 'bekor qilindi' bildirishnomasi ketmadi: {e}")


# ============================================================
# BUYURTMA BERISH (Order Flow)
# Bosqichlar: miqdor → yetkazish turi → (manzil) → to'lov → tasdiq → DB+bildirishnoma
# ============================================================

PAYMENT_LABELS = {
    'cash':     '💵 Naqd pul',
    'terminal': '💳 Terminal (plastik karta)',
    'p2p':      '📲 Karta raqamiga o\'tkazish (P2P)',
}

CARD_TYPE_LABELS = {
    'uzcard':     '🟦 Uzcard',
    'humo':       '🟩 Humo',
    'visa':       '🔵 Visa',
    'mastercard': '🔴 Mastercard',
}
DELIVERY_LABELS = {'delivery': '🚚 Yetkazib berish', 'pickup': '🚶 Olib ketaman'}


# Til-aware yorliqlar (yuqoridagi dict'lar eski moslik uchun qoladi)
def dlv_label(dlv, lang=DEFAULT_LANG):
    return {'delivery': t(lang, 'order_delivery'),
            'pickup': t(lang, 'order_pickup')}.get(dlv, dlv)


def pay_label(pay, lang=DEFAULT_LANG):
    return {'cash': t(lang, 'payment_cash'),
            'terminal': t(lang, 'payment_terminal'),
            'p2p': t(lang, 'payment_p2p')}.get(pay, pay)


def region_label_l(region_id, lang=DEFAULT_LANG):
    """db.get_region_label natijasini ('Viloyat → Tuman') til bo'yicha qaytaradi."""
    lbl = db.get_region_label(region_id)
    if not lbl or lang != 'ru':
        return lbl
    return ' → '.join(region_name(p, lang) for p in lbl.split(' → '))


def status_label(status, lang=DEFAULT_LANG):
    """Buyurtma/so'rov holatining til-aware yorlig'i."""
    return {
        'pending': t(lang, 'st_pending'),
        'confirmed': t(lang, 'st_confirmed'),
        'delivered': t(lang, 'st_delivered'),
        'cancelled': t(lang, 'st_cancelled'),
        'approved': t(lang, 'st_approved'),
        'rejected': t(lang, 'st_rejected'),
    }.get(status, fmt_status(status))


async def order_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mahsulot kartasidagi '🛒 Buyurtma berish' tugmasi."""
    query = update.callback_query
    await query.answer()

    product_id = int(query.data.split("_")[1])
    lang = get_lang(update, context)
    product = db.get_product_by_id(product_id)

    if not product:
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=t(lang, 'product_not_found')
        )
        return ConversationHandler.END

    if not product.get('in_stock'):
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=t(lang, 'product_out_of_stock')
        )
        return ConversationHandler.END

    # O'z mahsulotini buyurtma qilish — oldini olamiz
    buyer = db.get_user_by_telegram_id(update.effective_user.id)
    if buyer and buyer['id'] == product['seller_id']:
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=t(lang, 'product_own')
        )
        return ConversationHandler.END

    # Buyurtma spam tekshiruvi
    if buyer and check_order_spam(context, buyer['id'], product_id):
        try:
            db.increment_spam_count(buyer['id'])  # admin statistikasi: "nechtasi spam qilgan"
        except Exception:
            pass
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=t(lang, 'order_spam')
        )
        return ConversationHandler.END

    # Ish vaqti tashqarisidami?
    closed_note = ""
    if is_shop_open_now(product.get('working_hours')) is False:
        closed_note = t(lang, 'frag_shop_closed_note',
                        wh=html.escape(product.get('working_hours') or ''))

    # Jarayon ma'lumotlarini saqlab qo'yamiz
    context.user_data['order_product'] = product
    context.user_data.pop('order_qty', None)
    context.user_data.pop('order_delivery_type', None)
    context.user_data.pop('order_address', None)
    context.user_data.pop('order_lat', None)
    context.user_data.pop('order_lon', None)
    context.user_data.pop('order_payment', None)

    # edit_message_text foto xabarida ishlamaydi — send_message ishlatamiz
    await context.bot.send_message(
        chat_id=update.effective_chat.id,
        text=t(lang, 'order_qty_prompt',
               name=html.escape(product['name'] or ''),
               price=fmt_price(product['price']), closed=closed_note),
        parse_mode='HTML'
    )
    return ORDER_QUANTITY


async def order_quantity(update: Update, context: ContextTypes.DEFAULT_TYPE):
    lang = get_lang(update, context)
    try:
        qty = int(update.message.text.strip())
        if qty < 1 or qty > 999:
            raise ValueError
    except ValueError:
        await update.message.reply_text(t(lang, 'order_quantity_invalid'))
        return ORDER_QUANTITY

    product = context.user_data['order_product']

    # Stock_count tekshiruvi — agar zahira chegarasi qo'yilgan bo'lsa
    stock = product.get('stock_count')
    if stock is not None and qty > stock:
        await update.message.reply_text(
            t(lang, 'qty_only_n_available', stock=stock)
        )
        return ORDER_QUANTITY

    context.user_data['order_qty'] = qty
    total = qty * effective_unit_price(product, qty)   # optom narx (qty >= min bo'lsa)

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(dlv_label('delivery', lang), callback_data="ord_dlv_delivery")],
        [InlineKeyboardButton(dlv_label('pickup', lang), callback_data="ord_dlv_pickup")],
        [InlineKeyboardButton(t(lang, 'btn_cancel'), callback_data="ord_cancel")],
    ])
    await update.message.reply_text(
        t(lang, 'order_total_delivery_q', total=fmt_price(total)),
        reply_markup=kb,
        parse_mode='HTML'
    )
    return ORDER_DELIVERY_TYPE


async def order_delivery_type(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    lang = get_lang(update, context)
    if query.data == "ord_cancel":
        await query.edit_message_text(t(lang, 'order_cancel'))
        context.user_data.pop('order_product', None)
        return ConversationHandler.END

    choice = query.data.replace("ord_dlv_", "")  # 'delivery' yoki 'pickup'
    context.user_data['order_delivery_type'] = choice

    # Sotuvchi karta ma'lumotini oldindan olamiz (P2P uchun ko'rsatish uchun)
    product = context.user_data.get('order_product', {})
    seller_card = resolve_payment_card(product.get('seller_id'), product.get('created_by'))

    if choice == 'pickup':
        await _ask_payment(query, seller_card_info=seller_card, lang=lang)
        return ORDER_PAYMENT

    # Yetkazib berish — manzil so'raymiz
    await query.edit_message_text(t(lang, 'delivery_address_ask'))
    await query.message.reply_text(
        t(lang, 'delivery_address_hint'),
        reply_markup=ReplyKeyboardMarkup(
            [[KeyboardButton(t(lang, 'btn_send_location'), request_location=True)]],
            resize_keyboard=True, one_time_keyboard=True,
        )
    )
    return ORDER_ADDRESS


async def order_address(update: Update, context: ContextTypes.DEFAULT_TYPE):
    lang = get_lang(update, context)
    if update.message.location:
        loc = update.message.location
        context.user_data['order_lat'] = loc.latitude
        context.user_data['order_lon'] = loc.longitude
        context.user_data['order_address'] = f"{loc.latitude:.5f}, {loc.longitude:.5f}"
        remember_buyer_geo(context, loc.latitude, loc.longitude)
    else:
        # Yetkazib berishda joylashuv MAJBURIY — matn manzil qabul qilinmaydi (kuryer
        # navigatsiyasi GPS'ga bog'liq; app bilan bir xil qoida).
        await update.message.reply_text(
            t(lang, 'delivery_need_location'),
            reply_markup=ReplyKeyboardMarkup(
                [[KeyboardButton(t(lang, 'btn_send_location'), request_location=True)]],
                resize_keyboard=True, one_time_keyboard=True,
            )
        )
        return ORDER_ADDRESS

    await update.message.reply_text(
        t(lang, 'address_accepted'),
        reply_markup=ReplyKeyboardRemove()
    )
    product = context.user_data.get('order_product', {})
    seller_card = resolve_payment_card(product.get('seller_id'), product.get('created_by'))
    await _ask_payment(update.message, seller_card_info=seller_card, lang=lang)
    return ORDER_PAYMENT


def resolve_payment_card(seller_id, created_by=None):
    """Do'kon payment_mode'iga qarab to'lov kartasini qaytaradi.
    seller_id = do'kon EGASIning user id'si; created_by = mahsulotni joylagan xodim (ixtiyoriy).
    'shop' rejim → do'kon (yoki ega) kartasi; 'staff' rejim → xodim kartasi (yo'q bo'lsa egaga qaytadi).
    Qaytaradi: {'card_number','card_owner','card_type'} yoki None."""
    if not seller_id:
        return None
    shop = db.get_shop_by_owner(seller_id)
    mode = (shop.get('payment_mode') if shop else None) or 'shop'
    if mode == 'staff' and created_by and created_by != seller_id:
        staff = db.get_staff_by_user(created_by)
        if staff and staff.get('card_number'):
            return {'card_number': staff['card_number'],
                    'card_owner': staff.get('card_owner'),
                    'card_type': staff.get('card_type')}
    # shop rejim yoki xodim kartasi yo'q → do'kon kartasi, bo'lmasa egasi kartasi
    if shop and shop.get('card_number'):
        return {'card_number': shop['card_number'],
                'card_owner': shop.get('card_owner'),
                'card_type': shop.get('card_type')}
    owner = db.get_user_by_id(seller_id)
    if owner and owner.get('card_number'):
        return {'card_number': owner['card_number'],
                'card_owner': owner.get('card_owner'),
                'card_type': owner.get('card_type')}
    return None


async def _ask_payment(target, seller_card_info=None, cb_prefix="ord_pay_", cancel_cb="ord_cancel",
                       lang=DEFAULT_LANG):
    """To'lov usulini tanlash. seller_card_info — sotuvchining karta ma'lumoti.
    cb_prefix/cancel_cb — yakka buyurtma uchun 'ord_pay_'/'ord_cancel', savat uchun 'cart_pay_'/'cart_cancel'."""

    p2p_note = ""
    if seller_card_info:
        cnum = seller_card_info.get('card_number') or ''
        # Karta raqamini formatlash: 8600 **** **** 1234
        if len(cnum) >= 4:
            masked = f"{cnum[:4]} {'**** ' * ((len(cnum)-8)//4)}{cnum[-4:]}".strip()
        else:
            masked = cnum
        owner = seller_card_info.get('card_owner') or ''
        ctype = CARD_TYPE_LABELS.get(seller_card_info.get('card_type', ''), '')
        p2p_note = t(lang, 'frag_p2p_card_note', ctype=ctype, masked=masked, owner=owner)

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(pay_label('cash', lang),     callback_data=f"{cb_prefix}cash")],
        [InlineKeyboardButton(pay_label('terminal', lang), callback_data=f"{cb_prefix}terminal")],
        [InlineKeyboardButton(pay_label('p2p', lang),      callback_data=f"{cb_prefix}p2p")],
        [InlineKeyboardButton(t(lang, 'btn_cancel'),       callback_data=cancel_cb)],
    ])
    text = t(lang, 'payment_method_choose', p2p_note=p2p_note)

    if hasattr(target, 'edit_message_text'):
        await target.edit_message_text(text, reply_markup=kb, parse_mode='HTML')
    else:
        await target.reply_text(text, reply_markup=kb, parse_mode='HTML')


async def order_payment(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    lang = get_lang(update, context)
    if query.data == "ord_cancel":
        await query.edit_message_text(t(lang, 'order_cancel'))
        return ConversationHandler.END

    method = query.data.replace("ord_pay_", "")  # 'cash' | 'card' | 'click'
    context.user_data['order_payment'] = method

    # Tasdiq ekrani
    product = context.user_data['order_product']
    qty = context.user_data['order_qty']
    total = qty * effective_unit_price(product, qty)   # optom narx (qty >= min bo'lsa)
    dlv = context.user_data['order_delivery_type']

    address_frag = ""
    if dlv == 'delivery':
        address_frag = t(lang, 'frag_summary_address',
                         addr=html.escape(context.user_data.get('order_address') or ''))
    summary = t(lang, 'order_confirm_summary',
                pname=html.escape(product['name'] or ''),
                shop=html.escape(product.get('shop_name') or ''),
                qty=qty, total=fmt_price(total),
                dlv=dlv_label(dlv, lang), address=address_frag,
                pay=pay_label(method, lang))

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(t(lang, 'btn_confirm'), callback_data="ord_confirm_yes")],
        [InlineKeyboardButton(t(lang, 'btn_cancel'), callback_data="ord_cancel")],
    ])
    await query.edit_message_text(summary, reply_markup=kb, parse_mode='HTML')
    return ORDER_CONFIRM


async def _dispatch_order_notification(context, order_id):
    """Sotuvchiga buyurtma bildirishnomasini yuboradi (mahsulot rasmi + bog'lanish +
    jonli teskari sanoq), notify-ref'ni saqlaydi, sanoq jobini qo'yadi va mahsulotni
    joylagan xodimga fan-out qiladi.

    HAMMA ma'lumotni DB'dan (get_order_by_id) oladi — shuning uchun bir xil funksiya
    HAM bot order_confirm'idan, HAM Mini App yaratgan buyurtmalar uchun fon job'idan
    (webapp_order_dispatch_job) chaqirilishi mumkin. deadline DB'dagi auto_cancel_at'dan
    olinadi (chaqirishdan oldin set_order_deadline qilingan bo'lishi kerak)."""
    try:
        order = db.get_order_by_id(order_id)
        if not order:
            return
        seller_tg = order.get('seller_tg')
        if not seller_tg:
            return
        seller = db.get_user_by_id(order['seller_id']) if order.get('seller_id') else None
        slang = get_user_lang(seller) if seller else DEFAULT_LANG
        qty = order['quantity']
        total = order['total_price']
        dlv = order.get('delivery_type')
        buyer_lat = order.get('buyer_lat')
        buyer_lon = order.get('buyer_lon')
        buyer_address = order.get('delivery_address') or ''

        text = t(slang, 'seller_new_order_notify',
                 oid=fmt_order_id(order_id),
                 pname=html.escape(order.get('product_name') or ''),
                 qty=qty, total=fmt_price(total),
                 buyer=html.escape(order.get('buyer_name') or ''),
                 phone=order.get('buyer_phone') or '—',
                 dlv=dlv_label(dlv, slang))

        # Masofa hisoblash (sotuvchi do'koni → xaridor)
        if dlv == 'delivery' and buyer_lat and buyer_lon:
            shop_lat = order.get('shop_lat')
            shop_lon = order.get('shop_lon')
            if shop_lat and shop_lon:
                dist = haversine_km(shop_lat, shop_lon, buyer_lat, buyer_lon)
                if dist is not None:
                    text += t(slang, 'frag_dist_from_shop', km=f"{dist:.1f}")
            _ba = human_address(buyer_address)
            if _ba:
                text += t(slang, 'frag_seller_address', addr=html.escape(_ba))
            text += t(slang, 'seller_client_location_below')
        elif dlv == 'delivery':
            _ba = human_address(buyer_address)
            if _ba:
                text += t(slang, 'frag_seller_address', addr=html.escape(_ba))

        text += "💳 " + pay_label(order.get('payment_method'), slang)

        if order.get('buyer_username'):
            text += t(slang, 'frag_buyer_username', uname=html.escape(order['buyer_username']))

        photo = order.get('product_image')
        deadline = _order_deadline(order)
        kb = _order_notify_kb(slang, order_id=order_id,
                              buyer_tg=order.get('buyer_tg'),
                              buyer_username=order.get('buyer_username'))
        msg_id, is_cap = await _send_order_notification(
            context, seller_tg, slang, photo=photo, static_caption=text,
            kb=kb, deadline=deadline)
        if msg_id:
            db.set_order_notify_ref(order_id, seller_tg, msg_id, is_cap, text)

        # Jonli teskari sanoq jobini ishga tushiramiz (har 60s)
        _schedule_order_countdown(context.application.job_queue, order_id=order_id, first=60)

        # Xaridor lokatsiya yuborgan bo'lsa — alohida Telegram location (yo'l ko'rsatish)
        if dlv == 'delivery' and buyer_lat and buyer_lon:
            await context.bot.send_location(
                chat_id=seller_tg, latitude=buyer_lat, longitude=buyer_lon)

        # MULTI-SOTUVCHI: mahsulotni joylagan xodimga ham (rasm + bog'lanish, sanoqsiz)
        product = db.get_product_by_id(order['product_id'])
        if product:
            await _fanout_order_to_staff(context, product, text, kb,
                                         dlv, buyer_lat, buyer_lon, owner_tg=seller_tg, photo=photo)

        # ADMIN — har yangi buyurtmadan xabardor bo'lsin. Bu funksiya buyurtma uchun
        # YAGONA o'tish nuqtasi (HAM bot order_confirm'i, HAM Mini App fon-job'i shu yerga
        # keladi) — shuning uchun adminga xabar AYNAN BIR MARTA ketadi (spam yo'q).
        try:
            shop = (seller.get('shop_name') or seller.get('name')) if seller else None
            atext = (f"📥 <b>Yangi buyurtma</b> {fmt_order_id(order_id)}\n"
                     f"📦 {html.escape(order.get('product_name') or '')} × {qty}\n"
                     f"💰 {fmt_price(total)}\n"
                     f"👤 {html.escape(order.get('buyer_name') or '—')} · "
                     f"{order.get('buyer_phone') or '—'}\n"
                     f"🏪 {html.escape(shop or '—')}")
            await notify_admins(context, atext)
        except Exception as e:
            logging.error(f"Adminga yangi buyurtma xabari ketmadi (order {order_id}): {e}")
    except Exception as e:
        logging.error(f"Sotuvchiga bildirishnoma ketmadi (order {order_id}): {e}")


async def _dispatch_group_notification(context, group_id):
    """Savat (guruh) buyurtmasi uchun sotuvchiga BITTA bildirishnoma + jonli sanoq.
    Mini App yaratgan guruh buyurtmalari uchun fon job'idan chaqiriladi (mavjud
    _notify_seller_group qayta ishlatiladi — bot va app uchun bir xil)."""
    try:
        orders = db.get_orders_in_group(group_id)
        if not orders:
            return
        first = orders[0]
        seller = db.get_user_by_id(first.get('seller_id')) if first.get('seller_id') else None
        seller_tg = seller.get('telegram_id') if seller else first.get('seller_tg')
        if not seller_tg:
            return
        dlv = first.get('delivery_type') or 'delivery'
        payment = first.get('payment_method')
        addr = first.get('delivery_address')
        b_lat = first.get('buyer_lat')
        b_lon = first.get('buyer_lon')
        deadline = _order_deadline(first)
        await _notify_seller_group(context, group_id, seller_tg, dlv, payment, b_lat, b_lon, addr,
                                   deadline=deadline)
        _schedule_order_countdown(context.application.job_queue, group_id=group_id, first=60)

        # ADMIN — yangi savat (guruh) buyurtmasidan xabardor bo'lsin (BITTA marta)
        try:
            shop = (seller.get('shop_name') or seller.get('name')) if seller else None
            gtotal = sum(float(o.get('total_price') or 0) for o in orders)
            atext = (f"📥 <b>Yangi savat buyurtma</b> {fmt_order_id(int(group_id))}\n"
                     f"🧺 {len(orders)} ta mahsulot\n"
                     f"💰 {fmt_price(gtotal)}\n"
                     f"👤 {html.escape(first.get('buyer_name') or '—')} · "
                     f"{first.get('buyer_phone') or '—'}\n"
                     f"🏪 {html.escape(shop or '—')}")
            await notify_admins(context, atext)
        except Exception as e:
            logging.error(f"Adminga yangi guruh buyurtma xabari ketmadi (group {group_id}): {e}")
    except Exception as e:
        logging.error(f"Guruh bildirishnomasi (group {group_id}) ketmadi: {e}")


async def order_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    lang = get_lang(update, context)
    if query.data == "ord_cancel":
        await query.edit_message_text(t(lang, 'order_cancel'))
        return ConversationHandler.END

    # DB ga yozish
    buyer = db.get_user_by_telegram_id(update.effective_user.id)
    product = context.user_data['order_product']
    qty = context.user_data['order_qty']
    total = qty * effective_unit_price(product, qty)   # optom narx (qty >= min bo'lsa)
    dlv = context.user_data['order_delivery_type']

    order_id = db.create_order(
        buyer_id=buyer['id'],
        seller_id=product['seller_id'],
        product_id=product['id'],
        quantity=qty,
        total_price=total,
        delivery_address=context.user_data.get('order_address'),
        buyer_lat=context.user_data.get('order_lat'),
        buyer_lon=context.user_data.get('order_lon'),
        payment_method=context.user_data.get('order_payment'),
        delivery_type=dlv,
    )

    # Buyurtma spam cooldown belgilash
    mark_order_placed(context, buyer['id'], product['id'])

    await query.edit_message_text(
        t(lang, 'order_placed', oid=fmt_order_id(order_id)),
        parse_mode='HTML'
    )

    # Avto-bekor muddati (deadline) — DB'da saqlanadi. Jonli teskari sanoq shunga
    # bog'lanadi va bot restart bo'lsa ham real (o'zgarmas) qoladi.
    from datetime import datetime as _dt, timezone as _tz, timedelta as _td
    deadline = _dt.now(_tz.utc) + _td(seconds=ORDER_TTL_SECONDS)
    db.set_order_deadline(order_id, deadline)

    # Sotuvchiga bildirishnoma + jonli sanoq + xodim fan-out — endi DB'dan qayta quriladi
    # (umumiy funksiya, Mini App buyurtmalari bilan bir xil yo'l).
    await _dispatch_order_notification(context, order_id)

    # User_data'ni tozalaymiz
    for k in ('order_product', 'order_qty', 'order_delivery_type', 'order_address',
             'order_lat', 'order_lon', 'order_payment'):
        context.user_data.pop(k, None)

    return ConversationHandler.END


async def _fanout_order_to_staff(context, product, text, kb, dlv=None,
                                 buyer_lat=None, buyer_lon=None, owner_tg=None, photo=None):
    """Buyurtma xabarini mahsulotni joylagan xodimga ham yuboradi (ega allaqachon olgan).
    Xodim = ega bo'lsa yoki topilmasa — hech narsa qilmaydi. photo berilsa rasm bilan
    yuboriladi (jonli sanoqsiz — sanoq faqat egadagi asosiy xabarda)."""
    try:
        creator_id = product.get('created_by')
        if not creator_id or creator_id == product.get('seller_id'):
            return
        staff_user = db.get_user_by_id(creator_id)
        if not staff_user or not staff_user.get('telegram_id'):
            return
        staff_tg = staff_user['telegram_id']
        if owner_tg and str(staff_tg) == str(owner_tg):
            return
        await _send_order_notification(context, staff_tg, DEFAULT_LANG, photo=photo,
                                       static_caption=text, kb=kb, deadline=None,
                                       with_countdown=False)
        if dlv == 'delivery' and buyer_lat and buyer_lon:
            await context.bot.send_location(chat_id=staff_tg, latitude=buyer_lat, longitude=buyer_lon)
    except Exception as e:
        logging.error(f"Xodimga buyurtma fan-out ketmadi: {e}")


# ============================================================
# BUYURTMA BILDIRISHNOMASI — rasm + xaridor bilan bog'lanish + jonli teskari sanoq
# ============================================================
ORDER_TTL_SECONDS = 600  # buyurtma tasdiqlash muddati (10 daqiqa)
# ⏰ Muddat tugashidan oldin sotuvchiga ALOHIDA push eslatma (yangi xabar — telefon
# "biqillaydi") yuboriladigan bosqichlar: qolgan vaqt, daqiqada. Har biri BIR marta.
# Jonli sanoq faqat mavjud xabarni tahrirlaydi (push bermaydi) — bu esa qo'shimcha push.
ORDER_REMINDER_MINUTES = [6, 3, 1]


def _reminder_thresholds(total_window_sec):
    """Eslatma bosqichlari (daqiqa). UZUN muddatli buyurtmalar (optom — 30 daqiqa) uchun
    HAR 5 DAQIQA eslatiladi: 25, 20, 15, 10, 5 va oxirgi 1 daqiqa. Qisqa (oddiy — 10 daqiqa)
    buyurtmalar eski xulqni saqlaydi: [6, 3, 1]. Bosqichlar muddatdan KICHIK bo'ladi —
    aks holda boshlanishidayoq hammasi birdan yuborilardi."""
    try:
        mins = int(round(float(total_window_sec) / 60))
    except (ValueError, TypeError):
        return list(ORDER_REMINDER_MINUTES)
    if mins >= 20:
        thr = [m for m in range(5, mins, 5)]   # 5,10,...,(muddatdan kichik): 30→[5..25]
        return sorted(set(thr + [1]), reverse=True)
    return list(ORDER_REMINDER_MINUTES)


def _due_order_reminders(remaining_sec, fired, thresholds=None):
    """Hozir yuborilishi kerak bo'lgan eslatma bosqichlari (vaqti kelgan + hali
    yuborilmagan). `fired` — allaqachon yuborilgan bosqichlar ro'yxati.
    thresholds — bosqichlar ro'yxati (None bo'lsa standart ORDER_REMINDER_MINUTES).

    +3s tolerance: tik daqiqa chegarasidan bir oz kech (yoki erta) tushsa ham bosqich
    o'sha tikda ishga tushsin — aks holda butun bir daqiqa (60s) kechikardi."""
    thresholds = thresholds if thresholds is not None else ORDER_REMINDER_MINUTES
    return [thr for thr in thresholds
            if remaining_sec <= thr * 60 + 3 and thr not in fired]


def _parse_utc(ts):
    from datetime import datetime, timezone
    if not ts:
        return None
    try:
        return datetime.strptime(str(ts)[:19], "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
    except Exception:
        return None


def _order_deadline(order):
    """Buyurtmaning avto-bekor muddati (UTC). auto_cancel_at bo'lsa — o'sha; aks holda
    (eski buyurtmalar) created_at + TTL."""
    from datetime import timedelta
    dl = _parse_utc(order.get('auto_cancel_at'))
    if dl is not None:
        return dl
    ca = _parse_utc(order.get('created_at'))
    if ca is not None:
        return ca + timedelta(seconds=ORDER_TTL_SECONDS)
    return None


def _countdown_line(slang, deadline):
    """'⏳ N daqiqa qoldi (HH:MM gacha)' qatori, yoki muddat tugagan bo'lsa tegishli matn."""
    from datetime import datetime, timezone
    if deadline is None:
        return ""
    rem = (deadline - datetime.now(timezone.utc)).total_seconds()
    if rem <= 0:
        return t(slang, 'countdown_expired')
    mins = max(1, int((rem + 59) // 60))  # yuqoriga yaxlitlangan daqiqa
    until = deadline.astimezone(TZ_TASHKENT).strftime("%H:%M")
    return t(slang, 'countdown_line', mins=mins, until=until)


def _order_contact_rows(slang, relay_order_id, buyer_tg, buyer_username):
    """Xaridor bilan bog'lanish tugmalari: Telegram chat (URL) + bot orqali (relay)."""
    rows = []
    url = None
    if buyer_username:
        url = f"https://t.me/{str(buyer_username).lstrip('@')}"
    elif buyer_tg:
        url = f"tg://user?id={buyer_tg}"
    if url:
        rows.append([InlineKeyboardButton(t(slang, 'btn_contact_tg'), url=url)])
    if relay_order_id:
        rows.append([InlineKeyboardButton(t(slang, 'btn_contact_relay'),
                                          callback_data=f"order_msg_{relay_order_id}")])
    return rows


def _order_notify_kb(slang, *, order_id=None, group_id=None, relay_order_id=None,
                     buyer_tg=None, buyer_username=None):
    """Bildirishnoma klaviaturasi: Tasdiqlash/Bekor + xaridor bilan bog'lanish."""
    if group_id:
        rows = [[InlineKeyboardButton(t(slang, 'btn_confirm'), callback_data=f"gconfirm_{group_id}"),
                 InlineKeyboardButton(t(slang, 'btn_reject'), callback_data=f"gcancel_{group_id}")]]
    else:
        rows = [[InlineKeyboardButton(t(slang, 'btn_confirm'), callback_data=f"confirm_order_{order_id}"),
                 InlineKeyboardButton(t(slang, 'btn_reject'), callback_data=f"cancel_order_{order_id}")]]
    rows += _order_contact_rows(slang, relay_order_id or order_id, buyer_tg, buyer_username)
    return InlineKeyboardMarkup(rows)


async def _send_order_notification(context, recipient_tg, slang, *, photo, static_caption,
                                   kb, deadline, with_countdown=True):
    """Bildirishnomani yuboradi: mahsulot rasmi ALOHIDA xabar, so'ng tugmali MATN xabari.
    (message_id, is_caption=False) qaytaradi. Tugmali xabar doim matn bo'lgani uchun
    tasdiqlash/relay handlerlari (edit_message_text) va jonli sanoq muammosiz ishlaydi."""
    cd = _countdown_line(slang, deadline) if with_countdown else ""
    full = static_caption + (t(slang, 'countdown_sep') + cd if cd else "")
    try:
        if photo:
            try:
                await context.bot.send_photo(chat_id=recipient_tg, photo=photo)
            except Exception as e:
                logging.warning(f"Buyurtma rasmi yuborilmadi (chat {recipient_tg}): {e}")
        msg = await context.bot.send_message(chat_id=recipient_tg, text=full,
                                             parse_mode='HTML', reply_markup=kb)
        return msg.message_id, False
    except Exception as e:
        logging.error(f"Buyurtma bildirishnomasini yuborishda xato: {e}")
        return None, False


def _schedule_order_countdown(job_queue, *, order_id=None, group_id=None, first=60):
    """Buyurtma uchun jonli sanoq (har 60s) jobini qo'yadi. Xotirada — restartda
    _reschedule_pending_order_timers tiklaydi.

    Tiklar DEADLINE'ga moslanadi (`first = qolgan_vaqt % 60`): aks holda tik soati
    buyurtma yaratilgan vaqtdan siljib qoladi (webapp buyurtmasi dispatch-job orqali
    ~12-20s kech rejalanadi) → eslatmalar (6/3/1 daq) va avto-bekor shuncha kech kelardi.
    Moslangach har tik aniq daqiqa chegarasiga tushadi."""
    if not job_queue:
        return
    try:
        from datetime import datetime, timezone
        if group_id:
            _rows = db.get_orders_in_group(group_id)
            _dl = _order_deadline(_rows[0]) if _rows else None
        else:
            _o = db.get_order_by_id(order_id)
            _dl = _order_deadline(_o) if _o else None
        if _dl:
            _rem = (_dl - datetime.now(timezone.utc)).total_seconds()
            if _rem > 0:
                first = max(1, _rem % 60)   # keyingi daqiqa chegarasigacha
    except Exception:
        pass   # deadline o'qilmasa — fallback first (60/5) bilan davom etamiz
    if group_id:
        job_queue.run_repeating(order_countdown_job, interval=60, first=first,
                                data={'group_id': str(group_id)}, name=f"countdown_group_{group_id}")
    else:
        job_queue.run_repeating(order_countdown_job, interval=60, first=first,
                                data={'order_id': order_id}, name=f"countdown_order_{order_id}")


async def _autocancel_order_or_group(context, order, gid, oid, slang):
    """Muddat tugadi — yakka yoki guruh buyurtmani avtomatik bekor qiladi va xabar beradi."""
    try:
        buyer_tg = order.get('buyer_tg')
        seller_tg = order.get('seller_tg')
        if gid:
            orders = db.get_orders_in_group(gid)
            # ATOMIK: faqat hali 'pending' bo'lganlarini bekor qilamiz. Sotuvchi aynan shu
            # lahzada tasdiqlab/bekor qilib ulgursa, transition False qaytaradi → taymer
            # chekinadi (xaridorga soxta "avto-bekor" xabari ketmaydi).
            pend = [o for o in orders if db.transition_order_status(o['id'], 'cancelled', 'pending', cancel_by='system')]
            if not pend:
                return
            disp = fmt_order_id(int(gid))
            buyer = db.get_user_by_id(order['buyer_id'])
            seller = db.get_user_by_id(order['seller_id'])
            if buyer_tg:
                await context.bot.send_message(chat_id=buyer_tg,
                    text=t(buyer or 'uz', 'job_group_autocancel_buyer', oid=disp))
            if seller_tg:
                await context.bot.send_message(chat_id=seller_tg,
                    text=t(seller or 'uz', 'job_group_autocancel_seller', oid=disp))
        else:
            o = db.get_order_by_id(oid)
            if not o or o['status'] != 'pending':
                return
            # ATOMIK: sotuvchi shu lahzada tasdiqlab/bekor qilib ulgurgan bo'lsa, taymer
            # chekinadi (ikki marta ishlov + soxta "avto-bekor" xabarining oldini olamiz).
            if not db.transition_order_status(oid, 'cancelled', 'pending', cancel_by='system'):
                return
            disp = fmt_order_id(oid)
            buyer = db.get_user_by_id(o['buyer_id'])
            seller = db.get_user_by_id(o['seller_id'])
            if buyer_tg:
                await context.bot.send_message(chat_id=buyer_tg,
                    text=t(buyer or 'uz', 'job_autocancel_buyer', oid=disp), parse_mode='HTML')
            if seller_tg:
                await context.bot.send_message(chat_id=seller_tg,
                    text=t(seller or 'uz', 'job_autocancel_seller', oid=disp), parse_mode='HTML')
        # Bildirishnoma xabarini yakuniy holatga keltiramiz (tugmalar olib tashlanadi)
        try:
            chat_id = order.get('notify_chat_id'); msg_id = order.get('notify_message_id')
            if chat_id and msg_id:
                final = (order.get('notify_caption') or '') + "\n" + t(slang, 'countdown_cancelled')
                if order.get('notify_is_caption'):
                    await context.bot.edit_message_caption(chat_id=chat_id, message_id=msg_id,
                                                           caption=final, parse_mode='HTML')
                else:
                    await context.bot.edit_message_text(chat_id=chat_id, message_id=msg_id,
                                                        text=final, parse_mode='HTML')
        except Exception:
            pass
        logging.info(f"Avto-bekor: buyurtma {oid or gid} muddati tugadi.")
    except Exception as e:
        logging.error(f"Avto-bekor (order {oid or gid}) xatosi: {e}")


async def _push_order_reminder(context, order, gid, oid, slang, mins_left):
    """⏰ Muddat tugashidan oldin sotuvchiga (va mahsulotni joylagan xodimga) ALOHIDA
    push eslatma — jonli sanoq faqat xabarni tahrirlaydi, bu esa YANGI xabar (telefon
    push'i 'biqillaydi'). Yakka va guruh buyurtma uchun bir xil. Tugmasiz (amal App'da)."""
    disp = fmt_order_id(int(gid)) if gid else fmt_order_id(oid)
    if slang == 'ru':
        text = (f"⏰ Напоминание! Заказ {disp} ещё не подтверждён — осталось ~{mins_left} мин. "
                f"Подтвердите или отклоните в приложении, иначе он отменится автоматически.")
    else:
        text = (f"⏰ Eslatma! {disp} buyurtma hali tasdiqlanmagan — ~{mins_left} daqiqa qoldi. "
                f"Ilovada tasdiqlang yoki bekor qiling, aks holda avtomatik bekor bo'ladi.")
    recipients = set()
    if order.get('seller_tg'):
        recipients.add(order['seller_tg'])
    # MULTI-SOTUVCHI: mahsulot(lar)ni joylagan xodim(lar)ga ham
    try:
        rows = db.get_orders_in_group(gid) if gid else [order]
        for o in rows:
            prod = db.get_product_basic(o.get('product_id')) if o.get('product_id') else None
            if not prod:
                continue
            cb = prod.get('created_by')
            if cb and cb != prod.get('seller_id'):
                su = db.get_user_by_id(cb)
                if su and su.get('telegram_id'):
                    recipients.add(su['telegram_id'])
    except Exception:
        pass
    for chat_id in recipients:
        try:
            await context.bot.send_message(chat_id=chat_id, text=text)
        except Exception as e:
            logging.warning(f"Buyurtma eslatma push xato (order {oid or gid}, chat {chat_id}): {e}")


async def order_countdown_job(context: ContextTypes.DEFAULT_TYPE):
    """Har ~60s: sotuvchi bildirishnomasidagi jonli teskari sanoqni yangilaydi va
    muddat tugaganda buyurtmani avto-bekor qiladi. Buyurtma 'pending' bo'lmasa to'xtaydi.
    Bundan tashqari muddat tugashidan oldin 3 marta ALOHIDA push eslatma yuboradi."""
    from datetime import datetime, timezone
    try:
        d = context.job.data
        gid = d.get('group_id')
        oid = d.get('order_id')
        if gid:
            orders = db.get_orders_in_group(gid)
            order = orders[0] if orders else None
        else:
            order = db.get_order_by_id(oid)
        if not order:
            context.job.schedule_removal(); return
        if order.get('status') != 'pending':
            # Tasdiqlangan/bekor qilingan — handler xabarni o'zi yangilagan, sanoqni to'xtatamiz
            context.job.schedule_removal(); return

        seller = db.get_user_by_id(order.get('seller_id')) if order.get('seller_id') else None
        slang = get_user_lang(seller) if seller else DEFAULT_LANG
        deadline = _order_deadline(order)
        now = datetime.now(timezone.utc)

        if deadline is None or (deadline - now).total_seconds() <= 0:
            await _autocancel_order_or_group(context, order, gid, oid, slang)
            context.job.schedule_removal()
            return

        # ⏰ Push eslatmalar — har bosqich BIR marta. job.data'da kuzatiladi; restartda
        # 'fired' tozalanadi (kamdan-kam — bir martalik takror mumkin). Bosqichlar
        # buyurtma MUDDATIGA bog'liq: optom (30 daq) — har 5 daqiqa; oddiy (10 daq) — 6/3/1.
        remaining_sec = (deadline - now).total_seconds()
        _created = _parse_utc(order.get('created_at'))
        _window = (deadline - _created).total_seconds() if _created else None
        _thresholds = _reminder_thresholds(_window) if _window else None
        fired = d.setdefault('reminders_fired', [])
        for thr in _due_order_reminders(remaining_sec, fired, _thresholds):
            fired.append(thr)
            await _push_order_reminder(context, order, gid, oid, slang, thr)

        chat_id = order.get('notify_chat_id')
        msg_id = order.get('notify_message_id')
        if not chat_id or not msg_id:
            return  # tahrirlanadigan xabar yo'q (eski buyurtma) — faqat muddatni kutamiz

        static_caption = order.get('notify_caption') or ''
        is_caption = bool(order.get('notify_is_caption'))
        cd = _countdown_line(slang, deadline)
        full = static_caption + (t(slang, 'countdown_sep') + cd if cd else "")
        kb = _order_notify_kb(slang, order_id=(None if gid else oid), group_id=gid,
                              relay_order_id=order.get('id'),
                              buyer_tg=order.get('buyer_tg'), buyer_username=order.get('buyer_username'))
        try:
            if is_caption:
                await context.bot.edit_message_caption(chat_id=chat_id, message_id=msg_id,
                                                       caption=full, parse_mode='HTML', reply_markup=kb)
            else:
                await context.bot.edit_message_text(chat_id=chat_id, message_id=msg_id,
                                                    text=full, parse_mode='HTML', reply_markup=kb)
        except BadRequest as e:
            if 'not modified' not in str(e).lower():
                logging.warning(f"Sanoq tahrir xatosi (order {oid or gid}): {e}")
        except Exception as e:
            logging.warning(f"Sanoq tahrir kutilmagan xato (order {oid or gid}): {e}")
    except Exception as e:
        logging.error(f"order_countdown_job xatosi: {e}")


# ============================================================
# SAVAT (CART) — bir do'kondan bir nechta mahsulotni bitta buyurtmaga
# ============================================================
# Tamoyil: savat faqat BITTA do'kon uchun. Boshqa do'kondan qo'shilsa —
# yangi savat ochiladi (tasdiq bilan). Rasmiylashtirilganda har bir mahsulot
# alohida 'orders' qatori bo'lib yoziladi, lekin barchasi bitta order_group_id
# bilan bog'lanadi — shu sababli mavjud yakka-buyurtma kodi buzilmaydi, ammo
# xaridor ham, sotuvchi ham buni BITTA buyurtma (bitta raqam) sifatida ko'radi.

def _cart(context):
    """Joriy savat (dict) yoki None."""
    return context.user_data.get('cart')


def _cart_count(context):
    """Savatdagi jami dona soni."""
    cart = _cart(context)
    if not cart:
        return 0
    return sum(int(i.get('qty', 0)) for i in cart.get('items', {}).values())


def _cart_total(context):
    """Savatdagi jami summa."""
    cart = _cart(context)
    if not cart:
        return 0
    return sum(float(i.get('price', 0)) * int(i.get('qty', 0))
               for i in cart.get('items', {}).values())


def _cart_clear(context):
    context.user_data.pop('cart', None)
    for k in ('cart_dlv', 'cart_address', 'cart_lat', 'cart_lon', 'cart_payment'):
        context.user_data.pop(k, None)


async def cart_add(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mahsulotni savatga qo'shadi (yoki +1). Boshqa do'kon bo'lsa — tasdiq so'raydi."""
    query = update.callback_query
    product_id = int(query.data.split("_")[2])
    lang = get_lang(update, context)
    product = db.get_product_by_id(product_id)

    if not product or not product.get('in_stock') or product.get('status') == 'deleted':
        await query.answer(t(lang, 'product_out_of_stock'), show_alert=True)
        return

    buyer = db.get_user_by_telegram_id(update.effective_user.id)
    if buyer and buyer['id'] == product['seller_id']:
        await query.answer(t(lang, 'cart_cant_add_own'), show_alert=True)
        return

    cart = _cart(context)
    # Boshqa do'kon mahsuloti — yangi savat kerak
    if cart and cart.get('items') and cart.get('seller_id') != product['seller_id']:
        await query.answer()
        old_shop = cart.get('shop_name') or t(lang, 'other_shop_word')
        new_shop = product.get('shop_name') or t(lang, 'new_shop_word')
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=t(lang, 'cart_other_shop_q',
                   old=html.escape(old_shop), new=html.escape(new_shop)),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(t(lang, 'btn_new_cart_for', shop=new_shop[:20]),
                                      callback_data=f"cart_reset_add_{product_id}")],
                [InlineKeyboardButton(t(lang, 'btn_view_current_cart'), callback_data="cart_view")],
            ]),
            parse_mode='HTML'
        )
        return

    _cart_put(context, product, delta=1)
    cart = _cart(context)
    item = cart['items'].get(str(product_id)) if cart else None
    if not item:
        # Zahira 0 bo'lib qolgan bo'lsa qo'shilmaydi
        await query.answer(t(lang, 'cart_stock_empty'), show_alert=True)
        return
    await query.answer(t(lang, 'cart_added_toast', name=product['name'][:30], qty=item['qty']), show_alert=False)
    # Tugmani darhol "Savatda: N" ga aylantiramiz — xaridor qo'shilganini aniq ko'rsin
    await _refresh_cart_add_button(query, lang, product_id, item['qty'])
    # Katalog footer'idagi savat tugmasini (soni+summasi) JONLI yangilaymiz
    await _refresh_catalog_footer(context, update.effective_chat.id, lang)


async def _refresh_cart_add_button(query, lang, product_id, qty):
    """Mahsulot kartochkasidagi '➕ Savatga qo'shish' tugmasini '🛒 Savatda: N' ga
    almashtiradi (qaysi kontekstda bo'lishidan qat'i nazar — katalog yoki batafsil)."""
    msg = getattr(query, 'message', None)
    if not msg or not msg.reply_markup:
        return
    changed = False
    new_kb = []
    for row in msg.reply_markup.inline_keyboard:
        new_row = []
        for btn in row:
            if btn.callback_data == f"cart_add_{product_id}":
                # Qisqa yorliq — katalog ro'yxati qatori ixcham qolsin; bosilsa savatga o'tadi
                label = ("🛒 " + str(qty)) if len(btn.text or "") <= 3 else t(lang, 'btn_in_cart_manage', n=qty)
                new_row.append(InlineKeyboardButton(label, callback_data="cart_view"))
                changed = True
            else:
                new_row.append(btn)
        new_kb.append(new_row)
    if changed:
        try:
            await query.edit_message_reply_markup(reply_markup=InlineKeyboardMarkup(new_kb))
        except Exception:
            pass


async def cart_reset_add(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Joriy savatni tozalab, yangi do'kon mahsulotini qo'shadi."""
    query = update.callback_query
    product_id = int(query.data.split("_")[3])
    lang = get_lang(update, context)
    product = db.get_product_by_id(product_id)
    if not product or not product.get('in_stock'):
        await query.answer(t(lang, 'product_unavailable_toast'), show_alert=True)
        return
    _cart_clear(context)
    _cart_put(context, product, delta=1)
    await query.answer(t(lang, 'new_cart_toast', name=product['name'][:30]), show_alert=False)
    # Do'kon mahsulotlariga qaytaramiz
    try:
        await query.edit_message_text(
            t(lang, 'new_cart_opened',
              shop=html.escape(product.get('shop_name') or ''),
              name=html.escape(product['name'])),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(t(lang, 'btn_view_cart'), callback_data="cart_view")],
                [InlineKeyboardButton(t(lang, 'btn_shop_products_menu'),
                                      callback_data=f"shop_products_{product['seller_id']}_0")],
            ]),
            parse_mode='HTML'
        )
    except Exception:
        pass


def _cart_put(context, product, delta=1):
    """Savatga mahsulot qo'shadi / sonini o'zgartiradi (zahira chegarasini hisobga olib)."""
    cart = _cart(context)
    if not cart or cart.get('seller_id') != product['seller_id']:
        cart = {
            'seller_id': product['seller_id'],
            'shop_name': product.get('shop_name') or '',
            'items': {},
        }
    pid = str(product['id'])
    item = cart['items'].get(pid, {'name': product['name'], 'price': float(product['price']), 'qty': 0})
    new_qty = item['qty'] + delta
    # Zahira chegarasi
    stock = product.get('stock_count')
    if stock is not None:
        new_qty = min(new_qty, int(stock))
    if new_qty <= 0:
        cart['items'].pop(pid, None)
    else:
        item['qty'] = new_qty
        item['price'] = float(product['price'])  # narx o'zgargan bo'lsa yangilaymiz
        item['name'] = product['name']
        cart['items'][pid] = item
    if cart['items']:
        context.user_data['cart'] = cart
    else:
        _cart_clear(context)


async def cart_inc(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mahsulot sahifasidagi ➕ (toast bilan)."""
    query = update.callback_query
    product_id = int(query.data.split("_")[2])
    lang = get_lang(update, context)
    product = db.get_product_by_id(product_id)
    if not product:
        await query.answer(t(lang, 'product_not_found_toast'), show_alert=True)
        return
    cart = _cart(context)
    cur = cart['items'].get(str(product_id), {}).get('qty', 0) if cart else 0
    stock = product.get('stock_count')
    if stock is not None and cur >= int(stock):
        await query.answer(t(lang, 'only_n_available_toast', stock=stock), show_alert=True)
        return
    _cart_put(context, product, delta=1)
    qty = _cart(context)['items'].get(str(product_id), {}).get('qty', 0) if _cart(context) else 0
    await query.answer(t(lang, 'cart_qty_toast', name=product['name'][:30], qty=qty), show_alert=False)


async def cart_dec(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mahsulot sahifasidagi ➖ (toast bilan)."""
    query = update.callback_query
    product_id = int(query.data.split("_")[2])
    lang = get_lang(update, context)
    product = db.get_product_by_id(product_id)
    if not product:
        await query.answer(t(lang, 'product_not_found_toast'), show_alert=True)
        return
    _cart_put(context, product, delta=-1)
    cart = _cart(context)
    qty = cart['items'].get(str(product_id), {}).get('qty', 0) if cart else 0
    if qty <= 0:
        await query.answer(t(lang, 'removed_from_cart_toast'), show_alert=False)
    else:
        await query.answer(t(lang, 'cart_qty_toast', name=product['name'][:30], qty=qty), show_alert=False)


async def cart_view(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Savat tarkibini ko'rsatadi (miqdorni boshqarish + rasmiylashtirish)."""
    query = update.callback_query
    if query:
        await query.answer()

    lang = get_lang(update, context)
    cart = _cart(context)
    if not cart or not cart.get('items'):
        text = t(lang, 'cart_empty')
        kb = InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'btn_home'), callback_data="buyer_panel")]])
        if query:
            try:
                await query.edit_message_text(text, reply_markup=kb)
            except Exception:
                await context.bot.send_message(update.effective_chat.id, text, reply_markup=kb)
        else:
            await update.message.reply_text(text, reply_markup=kb)
        return

    seller_id = cart['seller_id']
    lines = [t(lang, 'cart_view_header', shop=html.escape(cart.get('shop_name') or ''))]
    keyboard = []
    for pid, item in cart['items'].items():
        subtotal = float(item['price']) * int(item['qty'])
        lines.append(t(lang, 'cart_view_item',
                       name=html.escape(item['name']), qty=item['qty'],
                       price=fmt_price(item['price']), subtotal=fmt_price(subtotal)))
        # Miqdor boshqaruvi
        keyboard.append([
            InlineKeyboardButton("➖", callback_data=f"cvdec_{pid}"),
            InlineKeyboardButton(f"{item['name'][:18]}: {item['qty']}", callback_data="noop"),
            InlineKeyboardButton("➕", callback_data=f"cvinc_{pid}"),
            InlineKeyboardButton("🗑", callback_data=f"cvrm_{pid}"),
        ])

    lines.append(t(lang, 'cart_view_total', total=fmt_price(_cart_total(context)), count=_cart_count(context)))

    keyboard.append([InlineKeyboardButton(t(lang, 'btn_checkout'), callback_data="cart_checkout")])
    keyboard.append([
        InlineKeyboardButton(t(lang, 'btn_add_more'), callback_data=f"shop_products_{seller_id}_0"),
        InlineKeyboardButton(t(lang, 'btn_clear'), callback_data="cart_clear"),
    ])
    keyboard.append([InlineKeyboardButton(t(lang, 'btn_back_to_shop'), callback_data=f"shop_{seller_id}")])

    text = "\n".join(lines)
    if query:
        try:
            await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')
        except Exception:
            await context.bot.send_message(update.effective_chat.id, text,
                                           reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')
    else:
        await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


async def cart_view_inc(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Savat ekranidagi ➕ — qayta render qiladi."""
    query = update.callback_query
    product_id = int(query.data.split("_")[1])
    product = db.get_product_by_id(product_id)
    if product:
        cart = _cart(context)
        cur = cart['items'].get(str(product_id), {}).get('qty', 0) if cart else 0
        stock = product.get('stock_count')
        if stock is not None and cur >= int(stock):
            await query.answer(t(get_lang(update, context), 'only_n_available_toast', stock=stock), show_alert=True)
            return
        _cart_put(context, product, delta=1)
    await cart_view(update, context)


async def cart_view_dec(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Savat ekranidagi ➖ — qayta render qiladi."""
    query = update.callback_query
    product_id = int(query.data.split("_")[1])
    product = db.get_product_by_id(product_id)
    if product:
        _cart_put(context, product, delta=-1)
    await cart_view(update, context)


async def cart_view_remove(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Savat ekranidagi 🗑 — mahsulotni butunlay olib tashlaydi."""
    query = update.callback_query
    product_id = int(query.data.split("_")[1])
    cart = _cart(context)
    if cart:
        cart.get('items', {}).pop(str(product_id), None)
        if not cart.get('items'):
            _cart_clear(context)
        else:
            context.user_data['cart'] = cart
    await cart_view(update, context)


async def cart_clear_ask(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Savatni tozalashdan oldin tasdiq."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    await query.edit_message_text(
        t(lang, 'cart_clear_confirm'),
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton(t(lang, 'btn_yes_clear'), callback_data="cart_clear_yes")],
            [InlineKeyboardButton(t(lang, 'btn_no_back'), callback_data="cart_view")],
        ])
    )


async def cart_clear_yes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    lang = get_lang(update, context)
    await query.answer(t(lang, 'cart_cleared_toast'))
    _cart_clear(context)
    await query.edit_message_text(
        t(lang, 'cart_cleared'),
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'btn_home'), callback_data="buyer_panel")]])
    )


# --- Savatni rasmiylashtirish (alohida conversation) ---

async def cart_checkout_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Rasmiylashtirish: yetkazish turini so'raydi."""
    query = update.callback_query
    await query.answer()

    lang = get_lang(update, context)
    cart = _cart(context)
    if not cart or not cart.get('items'):
        await query.edit_message_text(
            t(lang, 'cart_empty'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'btn_home'), callback_data="buyer_panel")]])
        )
        return ConversationHandler.END

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(dlv_label('delivery', lang), callback_data="cart_dlv_delivery")],
        [InlineKeyboardButton(dlv_label('pickup', lang), callback_data="cart_dlv_pickup")],
        [InlineKeyboardButton(t(lang, 'btn_cancel'), callback_data="cart_cancel")],
    ])
    await query.edit_message_text(
        t(lang, 'cart_checkout_header',
          count=_cart_count(context), total=fmt_price(_cart_total(context))),
        reply_markup=kb, parse_mode='HTML'
    )
    return CART_DELIVERY_TYPE


async def cart_delivery_type(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    lang = get_lang(update, context)
    if query.data == "cart_cancel":
        await query.edit_message_text(
            t(lang, 'checkout_cancelled'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'btn_back_to_cart'), callback_data="cart_view")]])
        )
        return ConversationHandler.END

    choice = query.data.replace("cart_dlv_", "")
    context.user_data['cart_dlv'] = choice

    cart = _cart(context)
    # Savat bitta do'kon uchun; turli xodim mahsuloti bo'lishi mumkin → do'kon/ega kartasi
    seller_card = resolve_payment_card(cart['seller_id']) if cart else None

    if choice == 'pickup':
        await _ask_payment(query, seller_card_info=seller_card,
                           cb_prefix="cart_pay_", cancel_cb="cart_cancel", lang=lang)
        return CART_PAYMENT

    await query.edit_message_text(t(lang, 'delivery_address_ask'))
    await query.message.reply_text(
        t(lang, 'delivery_address_hint'),
        reply_markup=ReplyKeyboardMarkup(
            [[KeyboardButton(t(lang, 'btn_send_location'), request_location=True)]],
            resize_keyboard=True, one_time_keyboard=True,
        )
    )
    return CART_ADDRESS


async def cart_address(update: Update, context: ContextTypes.DEFAULT_TYPE):
    lang = get_lang(update, context)
    if update.message.location:
        loc = update.message.location
        context.user_data['cart_lat'] = loc.latitude
        context.user_data['cart_lon'] = loc.longitude
        context.user_data['cart_address'] = f"{loc.latitude:.5f}, {loc.longitude:.5f}"
        remember_buyer_geo(context, loc.latitude, loc.longitude)
    else:
        # Yetkazib berishda joylashuv MAJBURIY (app bilan bir xil)
        await update.message.reply_text(
            t(lang, 'delivery_need_location'),
            reply_markup=ReplyKeyboardMarkup(
                [[KeyboardButton(t(lang, 'btn_send_location'), request_location=True)]],
                resize_keyboard=True, one_time_keyboard=True,
            )
        )
        return CART_ADDRESS

    await update.message.reply_text(t(lang, 'address_accepted'), reply_markup=ReplyKeyboardRemove())

    cart = _cart(context)
    seller_card = resolve_payment_card(cart['seller_id']) if cart else None
    await _ask_payment(update.message, seller_card_info=seller_card,
                       cb_prefix="cart_pay_", cancel_cb="cart_cancel", lang=lang)
    return CART_PAYMENT


async def cart_payment(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    lang = get_lang(update, context)
    if query.data == "cart_cancel":
        await query.edit_message_text(
            t(lang, 'checkout_cancelled'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'btn_back_to_cart'), callback_data="cart_view")]])
        )
        return ConversationHandler.END

    method = query.data.replace("cart_pay_", "")
    context.user_data['cart_payment'] = method

    cart = _cart(context)
    dlv = context.user_data.get('cart_dlv', 'delivery')

    lines = [t(lang, 'cart_confirm_header'),
             t(lang, 'cart_confirm_shop', shop=html.escape(cart.get('shop_name') or ''))]
    for item in cart['items'].values():
        subtotal = float(item['price']) * int(item['qty'])
        lines.append(t(lang, 'cart_confirm_item', name=html.escape(item['name']),
                       qty=item['qty'], price=fmt_price(item['price']), subtotal=fmt_price(subtotal)))
    lines.append(t(lang, 'cart_confirm_total', total=fmt_price(_cart_total(context))))
    lines.append(t(lang, 'cart_confirm_delivery', dlv=dlv_label(dlv, lang)))
    if dlv == 'delivery':
        lines.append(t(lang, 'cart_confirm_address', addr=html.escape(context.user_data.get('cart_address') or '')))
    lines.append(t(lang, 'cart_confirm_payment', pay=pay_label(method, lang)))

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(t(lang, 'btn_confirm'), callback_data="cart_confirm_yes")],
        [InlineKeyboardButton(t(lang, 'btn_cancel'), callback_data="cart_cancel")],
    ])
    await query.edit_message_text("\n".join(lines), reply_markup=kb, parse_mode='HTML')
    return CART_CONFIRM


async def cart_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    lang = get_lang(update, context)
    if query.data == "cart_cancel":
        await query.edit_message_text(
            t(lang, 'checkout_cancelled'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'btn_back_to_cart'), callback_data="cart_view")]])
        )
        return ConversationHandler.END

    buyer = db.get_user_by_telegram_id(update.effective_user.id)
    cart = _cart(context)
    if not buyer or not cart or not cart.get('items'):
        await query.edit_message_text(t(lang, 'cart_empty_expired'))
        _cart_clear(context)
        return ConversationHandler.END

    dlv = context.user_data.get('cart_dlv', 'delivery')
    payment = context.user_data.get('cart_payment')
    addr = context.user_data.get('cart_address')
    b_lat = context.user_data.get('cart_lat')
    b_lon = context.user_data.get('cart_lon')

    # Har bir mahsulotni qayta tekshiramiz va buyurtma qatori yaratamiz
    created_ids = []
    skipped = []
    grand_total = 0.0
    for pid, item in list(cart['items'].items()):
        product = db.get_product_by_id(int(pid))
        if not product or not product.get('in_stock') or product.get('status') == 'deleted':
            skipped.append(item['name'])
            continue
        qty = int(item['qty'])
        stock = product.get('stock_count')
        if stock is not None:
            qty = min(qty, int(stock))
        if qty <= 0:
            skipped.append(item['name'])
            continue
        price = effective_unit_price(product, qty)   # optom narx (qator soni minimumga yetsa)
        line_total = qty * price
        grand_total += line_total
        oid = db.create_order(
            buyer_id=buyer['id'],
            seller_id=cart['seller_id'],
            product_id=int(pid),
            quantity=qty,
            total_price=line_total,
            delivery_address=addr,
            buyer_lat=b_lat,
            buyer_lon=b_lon,
            payment_method=payment,
            delivery_type=dlv,
        )
        created_ids.append(oid)
        mark_order_placed(context, buyer['id'], int(pid))

    if not created_ids:
        await query.edit_message_text(
            t(lang, 'cart_nothing_available'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'btn_home'), callback_data="buyer_panel")]])
        )
        _cart_clear(context)
        return ConversationHandler.END

    # Guruh kodi = birinchi buyurtma id (do'kon ichida noyob)
    group_id = str(created_ids[0])
    db.set_orders_group(created_ids, group_id)

    seller = db.get_user_by_id(cart['seller_id'])
    seller_tg = seller.get('telegram_id') if seller else None

    skip_note = ""
    if skipped:
        skip_note = t(lang, 'cart_skip_note',
                      names=', '.join(html.escape(s) for s in skipped[:5]))

    await query.edit_message_text(
        t(lang, 'cart_order_placed',
          oid=fmt_order_id(int(group_id)), count=len(created_ids),
          total=fmt_price(grand_total), skip=skip_note),
        parse_mode='HTML'
    )

    # Avto-bekor muddati (deadline) — DB'da saqlanadi (jonli sanoq va restart uchun)
    from datetime import datetime as _dt, timezone as _tz, timedelta as _td
    deadline = _dt.now(_tz.utc) + _td(seconds=ORDER_TTL_SECONDS)
    db.set_group_deadline(group_id, deadline)

    # Sotuvchiga BITTA bildirishnoma (barcha mahsulotlar bilan) + jonli sanoq jobi
    try:
        if seller_tg:
            await _notify_seller_group(context, group_id, seller_tg, dlv, payment, b_lat, b_lon, addr,
                                       deadline=deadline)
        _schedule_order_countdown(context.application.job_queue, group_id=group_id, first=60)
    except Exception as e:
        logging.error(f"Sotuvchiga savat bildirishnomasi ketmadi: {e}")

    _cart_clear(context)
    return ConversationHandler.END


async def _notify_seller_group(context, group_id, seller_tg, dlv, payment, b_lat, b_lon, addr,
                               deadline=None):
    """Sotuvchiga savat buyurtmasi bo'yicha bitta bildirishnoma (mahsulotlar ro'yxati bilan)."""
    orders = db.get_orders_in_group(group_id)
    if not orders:
        return
    first = orders[0]
    grand = sum(float(o['total_price']) for o in orders)

    # Sotuvchi tili
    seller = db.get_user_by_id(first.get('seller_id')) if first.get('seller_id') else None
    slang = get_user_lang(seller) if seller else DEFAULT_LANG

    lines = [
        t(slang, 'seller_group_header', oid=fmt_order_id(int(group_id)), count=len(orders)),
    ]
    for o in orders:
        # Variant-buyurtma qatori (rasm/hil + rang + razmer) — variant_label bo'lsa alohida ko'rsatamiz
        _vlabel = (o.get('variant_label') or '').strip()
        _vsize = (o.get('variant_size') or '').strip()
        _vcolor = (o.get('variant_color') or '').strip()
        if _vlabel or _vsize or _vcolor:
            # Takror qiymatlarni olib tashlaymiz (optomda label=rang=color → "Oq · Oq" bo'lmasin)
            _vseen = []
            for _x in (_vlabel, _vcolor, _vsize):
                if _x and _x not in _vseen:
                    _vseen.append(_x)
            _vtxt = " · ".join(_vseen)
            # dona narx: total/qty (optom narx jamiga qarab hisoblangan)
            _unit = (float(o['total_price']) / o['quantity']) if o.get('quantity') else float(o['product_price'])
            lines.append(t(slang, 'seller_group_item_variant', name=html.escape(o['product_name'] or ''),
                           variant=html.escape(_vtxt), qty=o['quantity'],
                           price=fmt_price(_unit), total=fmt_price(o['total_price'])))
        else:
            lines.append(t(slang, 'seller_group_item', name=html.escape(o['product_name'] or ''),
                           qty=o['quantity'], price=fmt_price(o['product_price']), total=fmt_price(o['total_price'])))
    lines.append("")
    lines.append(t(slang, 'seller_group_total', total=fmt_price(grand)))
    lines.append(t(slang, 'seller_group_buyer', buyer=html.escape(first.get('buyer_name') or '')))
    lines.append(f"📞 {first.get('buyer_phone') or '—'}")
    if first.get('buyer_username'):
        lines.append(f"💬 @{str(first['buyer_username']).lstrip('@')}")
    lines.append(f"🚚 {dlv_label(dlv, slang)}")

    if dlv == 'delivery' and b_lat and b_lon:
        s_lat, s_lon = first.get('shop_lat'), first.get('shop_lon')
        if s_lat and s_lon:
            d = haversine_km(s_lat, s_lon, b_lat, b_lon)
            if d is not None:
                lines.append(t(slang, 'grp_dist_from_shop', km=f"{d:.1f}"))
        _ba = human_address(addr)
        if _ba:
            lines.append(t(slang, 'grp_address', addr=html.escape(_ba)))
        lines.append(t(slang, 'grp_client_location'))
    elif dlv == 'delivery':
        _ba = human_address(addr)
        if _ba:
            lines.append(t(slang, 'grp_address', addr=html.escape(_ba)))

    lines.append("💳 " + pay_label(payment, slang))
    lines.append(t(slang, 'seller_group_confirm_prompt'))

    body = "\n".join(lines)
    photo = first.get('product_image')
    kb = _order_notify_kb(slang, group_id=group_id, relay_order_id=first.get('id'),
                          buyer_tg=first.get('buyer_tg'), buyer_username=first.get('buyer_username'))
    msg_id, is_cap = await _send_order_notification(
        context, seller_tg, slang, photo=photo, static_caption=body, kb=kb, deadline=deadline)
    if msg_id:
        db.set_group_notify_ref(group_id, seller_tg, msg_id, is_cap, body)
    if dlv == 'delivery' and b_lat and b_lon:
        try:
            await context.bot.send_location(chat_id=seller_tg, latitude=b_lat, longitude=b_lon)
        except Exception:
            pass

    # MULTI-SOTUVCHI: guruhdagi mahsulotlarni joylagan xodimlarga ham yuboramiz
    try:
        seen = set()
        for o in orders:
            prod = db.get_product_basic(o.get('product_id')) if o.get('product_id') else None
            if not prod:
                continue
            cb = prod.get('created_by')
            if not cb or cb == prod.get('seller_id') or cb in seen:
                continue
            seen.add(cb)
            su = db.get_user_by_id(cb)
            if not su or not su.get('telegram_id') or str(su['telegram_id']) == str(seller_tg):
                continue
            await _send_order_notification(context, su['telegram_id'], slang, photo=photo,
                                           static_caption=body, kb=kb, deadline=None,
                                           with_countdown=False)
            if dlv == 'delivery' and b_lat and b_lon:
                try:
                    await context.bot.send_location(chat_id=su['telegram_id'], latitude=b_lat, longitude=b_lon)
                except Exception:
                    pass
    except Exception as e:
        logging.error(f"Guruh xodim fan-out ketmadi: {e}")


# --- Guruh (savat) buyurtma — SOTUVCHI tomoni ---

def _seller_manages_group(seller_user, orders):
    """Foydalanuvchi guruh buyurtmasini boshqara oladimi: ega, admin, yoki guruhdagi
    mahsulotni joylagan (ruxsatli, faol) xodim. group_status_action bilan bir xil mantiq."""
    if not orders:
        return False
    is_owner = bool(seller_user and seller_user.get('id') == orders[0].get('seller_id'))
    is_admin = bool(seller_user and seller_user.get('role') == 'admin')
    if is_owner or is_admin:
        return True
    if seller_user:
        staff_rec = db.get_staff_by_user(seller_user['id'])
        if staff_rec and staff_rec.get('perm_confirm_orders', 1) and staff_rec.get('is_active', 1):
            for o in orders:
                prod = db.get_product_basic(o.get('product_id')) if o.get('product_id') else None
                if prod and prod.get('created_by') == seller_user['id']:
                    return True
    return False


async def seller_reject_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Sotuvchi PENDING buyurtmani rad etmoqchi — avval bekor SABABINI so'raymiz
    (AI taklif + preset). Kirish: cancel_order_<oid> yoki gcancel_<gid>."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    data = query.data
    if data.startswith("gcancel_"):
        scope, key = "g", data.split("_", 1)[1]
        orders = db.get_orders_in_group(key)
        if not orders:
            await query.edit_message_text(t(lang, 'order_not_found'))
            return
        seller_user = db.get_user_by_telegram_id(update.effective_user.id)
        if (update.effective_user.id != ADMIN_ID) and not _seller_manages_group(seller_user, orders):
            await query.answer(t(lang, 'not_your_order_toast'), show_alert=True)
            return
        if orders[0].get('status') != 'pending':
            await seller_group_order_detail(update, context, group_id=key)
            return
        prod_name = orders[0].get('product_name') or ''
    else:  # cancel_order_<oid>
        oid = int(data.split("_")[2])
        scope, key = "o", str(oid)
        if not await _ensure_order_seller(update, context, oid):
            return
        order = db.get_order_by_id(oid)
        if not order or order.get('status') != 'pending':
            await seller_order_detail(update, context)
            return
        prod_name = order.get('product_name') or ''

    try:
        ai = await ai_assistant.suggest_cancel_reasons(
            party='seller', product_name=prod_name, status='pending', lang=lang)
    except Exception as e:
        logging.warning(f"AI reject sabab taklif xato: {e}")
        ai = []
    context.user_data[f'rjai_{scope}_{key}'] = ai
    kb = [[InlineKeyboardButton(f"🤖 {r}"[:60], callback_data=f"rjok_{scope}_{key}_a{i}")]
          for i, r in enumerate(ai)]
    kb += [[InlineKeyboardButton(t(lang, 'crsn_' + c), callback_data=f"rjok_{scope}_{key}_{c}")]
           for c in SELLER_CANCEL_REASONS]
    kb.append([InlineKeyboardButton(t(lang, 'btn_cancel'), callback_data=f"rjback_{scope}_{key}")])
    await query.edit_message_text(
        t(lang, 'cancel_pick_reason', oid=fmt_order_id(int(key))),
        reply_markup=InlineKeyboardMarkup(kb), parse_mode='HTML')


async def seller_reject_do(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """rjok_<scope>_<key>_<choice> — tanlangan sabab bilan PENDING buyurtmani rad etadi.
    rjback_<scope>_<key> — sababsiz orqaga (detalga). 'other' → erkin matn so'raydi."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    data = query.data

    if data.startswith("rjback_"):
        _, scope, key = data.split("_", 2)
        context.user_data.pop(f'rjai_{scope}_{key}', None)
        if scope == "g":
            await seller_group_order_detail(update, context, group_id=key)
        else:
            await seller_order_detail(update, context)
        return

    # rjok_<scope>_<key>_<choice>
    body = data[len("rjok_"):]
    scope, key, choice = body.split("_", 2)

    if choice.startswith("a"):
        try:
            idx = int(choice[1:])
            reason = 'text:' + (context.user_data.get(f'rjai_{scope}_{key}') or [])[idx][:300]
        except (ValueError, IndexError):
            reason = None
    else:
        reason = 'code:' + choice
    context.user_data.pop(f'rjai_{scope}_{key}', None)
    await _perform_seller_reject(update, context, scope, key, reason)


async def _perform_seller_reject(update, context, scope, key, reason):
    """PENDING buyurtmani (yakka yoki guruh) sabab bilan rad etadi: atomik o'tkazish,
    taymerlarni o'chirish, xaridorga xabar, detalni yangilash. Pending → zahira tegilmaydi."""
    query = update.callback_query
    lang = get_lang(update, context)
    jq = context.application.job_queue
    if scope == "g":
        orders = db.get_orders_in_group(key)
        if not orders:
            await query.edit_message_text(t(lang, 'order_not_found'))
            return
        seller_user = db.get_user_by_telegram_id(update.effective_user.id)
        if (update.effective_user.id != ADMIN_ID) and not _seller_manages_group(seller_user, orders):
            await query.answer(t(lang, 'not_your_order_toast'), show_alert=True)
            return
        if jq:
            for nm in (f"countdown_group_{key}", f"auto_cancel_group_{key}", f"reminder_group_{key}"):
                for job in jq.get_jobs_by_name(nm):
                    job.schedule_removal()
        won = db.transition_group_status(key, 'cancelled', 'pending',
                                         cancel_by='seller', cancel_reason=reason)
        if not won:
            await seller_group_order_detail(update, context, group_id=key)
            return
        try:
            buyer_tg = orders[0].get('buyer_tg')
            if buyer_tg:
                buyer = db.get_user_by_id(orders[0]['buyer_id'])
                blang = get_user_lang(buyer) if buyer else DEFAULT_LANG
                await context.bot.send_message(
                    chat_id=buyer_tg,
                    text=t(blang, 'grp_cancelled_notify', oid=fmt_order_id(int(key)), n=len(orders)),
                    parse_mode='HTML')
        except Exception as e:
            logging.error(f"Xaridorga guruh bekor bildirishnomasi ketmadi: {e}")
        await seller_group_order_detail(update, context, group_id=key)
    else:
        oid = int(key)
        if not await _ensure_order_seller(update, context, oid):
            return
        if jq:
            for jn in (f"countdown_order_{oid}", f"auto_cancel_{oid}", f"reminder_{oid}"):
                for job in jq.get_jobs_by_name(jn):
                    job.schedule_removal()
        won = db.transition_order_status(oid, 'cancelled', 'pending',
                                         cancel_by='seller', cancel_reason=reason)
        if not won:
            await seller_order_detail(update, context)
            return
        try:
            order = db.get_order_by_id(oid)
            if order and order.get('buyer_tg'):
                buyer = db.get_user_by_id(order['buyer_id'])
                blang = get_user_lang(buyer) if buyer else DEFAULT_LANG
                await context.bot.send_message(
                    chat_id=order['buyer_tg'],
                    text=t(blang, 'order_cancelled_notify', oid=fmt_order_id(oid),
                           pname=html.escape(order.get('product_name') or '')),
                    parse_mode='HTML')
        except Exception as e:
            logging.error(f"Xaridorga bekor bildirishnomasi ketmadi: {e}")
        await seller_order_detail(update, context)


async def group_status_action(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """gconfirm_/gcancel_/gdeliver_ — butun guruh holatini o'zgartiradi."""
    query = update.callback_query
    await query.answer()

    parts = query.data.split("_", 1)
    action = parts[0]  # gconfirm / gcancel / gdeliver
    group_id = parts[1]

    orders = db.get_orders_in_group(group_id)
    if not orders:
        await query.edit_message_text(T(update, context, 'order_not_found'))
        return

    # Egalik tekshiruvi (ega, admin, yoki guruhdagi mahsulotni joylagan xodim — ruxsat bilan)
    seller_user = db.get_user_by_telegram_id(update.effective_user.id)
    is_owner = bool(seller_user and seller_user.get('id') == orders[0].get('seller_id'))
    is_admin = (update.effective_user.id == ADMIN_ID) or (seller_user and seller_user.get('role') == 'admin')
    is_staff_creator = False
    if seller_user and not (is_owner or is_admin):
        staff_rec = db.get_staff_by_user(seller_user['id'])
        if staff_rec and staff_rec.get('perm_confirm_orders', 1) and staff_rec.get('is_active', 1):
            for o in orders:
                prod = db.get_product_basic(o.get('product_id')) if o.get('product_id') else None
                if prod and prod.get('created_by') == seller_user['id']:
                    is_staff_creator = True
                    break
    if not (is_owner or is_admin or is_staff_creator):
        await query.answer(t(get_lang(update, context), 'not_your_order_toast'), show_alert=True)
        return

    # BERISH: avval to'lov holatini so'raymiz (settlement oqimi yakuniy 'delivered' qo'yadi)
    if action == 'gdeliver':
        grand = sum(float(o.get('total_price') or 0) for o in orders)
        await _ask_settlement(update, context, scope='g', key=group_id, total=grand)
        return

    status_map = {'gconfirm': 'confirmed', 'gcancel': 'cancelled', 'gdeliver': 'delivered'}
    new_status = status_map.get(action)
    if not new_status:
        return

    # Guruh taymerlarini o'chiramiz
    if context.application.job_queue:
        for nm in (f"countdown_group_{group_id}", f"auto_cancel_group_{group_id}", f"reminder_group_{group_id}"):
            for job in context.application.job_queue.get_jobs_by_name(nm):
                job.schedule_removal()

    # HIMOYA (ATOMIK): har bir ichki buyurtmani 'pending'dan o'tkazamiz; faqat YUTGAN
    # (haqiqatan o'zgargan) buyurtmalar uchun zahira kamaytiramiz. Bot+Mini App bir vaqtda
    # yoki tugma ikki marta bosilsa, guruh ikki marta ishlanmaydi (ilgari guard yo'q edi →
    # zahira ikki marta kamayardi). Hech biri yutmasa — allaqachon ishlangan, qaytamiz.
    won_orders = [o for o in orders if db.transition_order_status(
        o['id'], new_status, 'pending',
        cancel_by='seller' if new_status == 'cancelled' else None)]
    if not won_orders:
        return

    # Tasdiqlashda har bir (yutgan) mahsulot zahirasini kamaytiramiz
    if new_status == 'confirmed':
        for o in won_orders:
            try:
                left = db.decrement_stock_on_confirm(o['product_id'], o['quantity'])
                if left == 0:
                    await _notify_seller_sold_out(context, o['product_id'])
            except Exception as e:
                logging.error(f"Guruh stock kamaytirish xatosi: {e}")

    # Xaridorga BITTA bildirishnoma (xaridor tilida)
    try:
        buyer_tg = orders[0].get('buyer_tg')
        if buyer_tg:
            buyer = db.get_user_by_id(orders[0]['buyer_id'])
            blang = get_user_lang(buyer) if buyer else DEFAULT_LANG
            is_pickup = orders[0].get('delivery_type') == 'pickup'
            n = len(orders)
            oid = fmt_order_id(int(group_id))
            if new_status == 'confirmed':
                txt = t(blang, 'grp_confirmed_pickup' if is_pickup else 'grp_confirmed_delivery',
                        oid=oid, n=n)
                kb = None
            elif new_status == 'cancelled':
                txt = t(blang, 'grp_cancelled_notify', oid=oid, n=n)
                kb = None
            else:  # delivered
                txt = t(blang, 'grp_delivered_pickup' if is_pickup else 'grp_delivered_delivery',
                        oid=oid, n=n)
                kb = InlineKeyboardMarkup([[InlineKeyboardButton(
                    t(blang, 'btn_leave_rating'), callback_data=f"order_rate_{orders[0]['id']}")]])
            await context.bot.send_message(chat_id=buyer_tg, text=txt, reply_markup=kb, parse_mode='HTML')
    except Exception as e:
        logging.error(f"Xaridorga guruh bildirishnomasi ketmadi: {e}")

    await seller_group_order_detail(update, context, group_id=group_id)


async def seller_group_order_detail(update: Update, context: ContextTypes.DEFAULT_TYPE, group_id: str = None):
    """Sotuvchi uchun savat buyurtmasi sahifasi."""
    query = update.callback_query
    if group_id is None:
        group_id = query.data.split("_", 2)[2]  # seller_gorder_<gid>

    orders = db.get_orders_in_group(group_id)
    lang = get_lang(update, context)
    if not orders:
        await query.edit_message_text(t(lang, 'order_not_found'))
        return

    first = orders[0]
    status = first['status']
    dlv = first.get('delivery_type', 'delivery')
    pay = first.get('payment_method') or 'cash'
    grand = sum(float(o['total_price']) for o in orders)

    lines = [t(lang, 'grp_seller_header', oid=fmt_order_id(int(group_id)), n=len(orders))]
    for o in orders:
        lines.append(t(lang, 'seller_group_item', name=html.escape(o['product_name'] or ''),
                       qty=o['quantity'], price=fmt_price(o['product_price']), total=fmt_price(o['total_price'])))
    lines.append(t(lang, 'grp_total_plain', total=fmt_price(grand)))
    lines.append(t(lang, 'grp_status_line', status=status_label(status, lang)))
    lines.append(f"🚚 {dlv_label(dlv, lang)}")

    pay_note = ""
    if pay == 'p2p':
        card = resolve_payment_card(first.get('seller_id'))
        if card and card.get('card_number'):
            cnum = card['card_number']
            masked = f"{cnum[:4]} **** **** {cnum[-4:]}"
            ctype = CARD_TYPE_LABELS.get(card.get('card_type', ''), '💳')
            pay_note = t(lang, 'p2p_your_card', ctype=ctype, masked=masked)
    lines.append(t(lang, 'grp_pay_line', pay=pay_label(pay, lang), paynote=pay_note))
    lines.append("\n" + t(lang, 'seller_group_buyer', buyer=html.escape(first.get('buyer_name') or '')))
    lines.append(t(lang, 'grp_phone_line', phone=fmt_phone(first.get('buyer_phone'))))

    if dlv == 'delivery':
        b_lat, b_lon = first.get('buyer_lat'), first.get('buyer_lon')
        addr_txt = human_address(first.get('delivery_address'))
        if addr_txt:
            lines.append(t(lang, 'courier_addr', addr=html.escape(addr_txt)))
        if b_lat is not None and b_lon is not None:
            lines.append(t(lang, 'grp_map_line', url=f"https://www.google.com/maps/search/?api=1&query={b_lat},{b_lon}"))
            s_lat, s_lon = first.get('shop_lat'), first.get('shop_lon')
            if s_lat is not None and s_lon is not None:
                d = haversine_km(s_lat, s_lon, b_lat, b_lon)
                if d is not None:
                    lines.append(t(lang, 'courier_dist', km=f"{d:.1f}"))
    lines.append(t(lang, 'grp_date_plain', date=fmt_datetime(first.get('created_at'))))
    _pbadge = _progress_badge(lang, first)
    if _pbadge:
        lines.append(_pbadge)
    if status == 'pending':
        _gcd = _countdown_line(lang, _order_deadline(first))
        if _gcd:
            lines.append(_gcd)
    _badge = _settlement_badge(lang, first)
    if _badge:
        lines.append(_badge)

    keyboard = []
    if status == 'pending':
        keyboard.append([InlineKeyboardButton(t(lang, 'btn_confirm'), callback_data=f"gconfirm_{group_id}")])
        keyboard.append([InlineKeyboardButton(t(lang, 'btn_reject'), callback_data=f"gcancel_{group_id}")])
    elif status == 'confirmed':
        if dlv == 'delivery':
            keyboard.append([InlineKeyboardButton(t(lang, 'btn_delivered'), callback_data=f"gdeliver_{group_id}")])
        else:
            keyboard.append([InlineKeyboardButton(t(lang, 'btn_buyer_received'), callback_data=f"gdeliver_{group_id}")])
    if dlv == 'delivery' and status in ('pending', 'confirmed'):
        keyboard.append([InlineKeyboardButton(t(lang, 'btn_forward_courier'), callback_data=f"gcrfwd_{group_id}")])
    keyboard.append([InlineKeyboardButton(t(lang, 'btn_correspondence'), callback_data=f"msgs_{first['id']}")])
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="seller_orders")])

    await query.edit_message_text("\n".join(lines), reply_markup=InlineKeyboardMarkup(keyboard),
                                  parse_mode='HTML', disable_web_page_preview=True)


async def seller_forward_courier_group(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Savat buyurtmasini kuryerga uzatish (lokatsiya + barcha mahsulotlar)."""
    query = update.callback_query
    chat_id = update.effective_chat.id
    group_id = query.data.split("_", 1)[1]  # gcrfwd_<gid>

    orders = db.get_orders_in_group(group_id)
    lang = get_lang(update, context)
    if not orders:
        await context.bot.send_message(chat_id, t(lang, 'order_not_found'))
        return

    first = orders[0]
    seller_user = db.get_user_by_telegram_id(update.effective_user.id)
    is_owner = bool(seller_user and seller_user.get('id') == first.get('seller_id'))
    is_admin = (update.effective_user.id == ADMIN_ID) or (seller_user and seller_user.get('role') == 'admin')
    if not (is_owner or is_admin):
        await context.bot.send_message(chat_id, t(lang, 'not_your_order_plain'))
        return

    b_lat, b_lon = first.get('buyer_lat'), first.get('buyer_lon')
    s_lat, s_lon = first.get('shop_lat'), first.get('shop_lon')
    pay = first.get('payment_method') or 'cash'
    addr_txt = human_address(first.get('delivery_address'))
    grand = sum(float(o['total_price']) for o in orders)

    if b_lat is not None and b_lon is not None:
        try:
            await context.bot.send_location(chat_id=chat_id, latitude=b_lat, longitude=b_lon)
        except Exception as e:
            logging.warning(f"group courier send_location xatosi: {e}")

    lines = [t(lang, 'courier_group_header', oid=fmt_order_id(int(group_id)), n=len(orders)), ""]
    for o in orders:
        lines.append(t(lang, 'courier_group_item', name=html.escape(o['product_name'] or ''), qty=o['quantity']))
    lines += [
        "",
        t(lang, 'courier_sum', total=fmt_price(grand)),
        t(lang, 'courier_pay', pay=pay_label(pay, lang)),
        "",
        t(lang, 'courier_client', buyer=html.escape(first.get('buyer_name') or '')),
        t(lang, 'courier_phone', phone=fmt_phone(first.get('buyer_phone'))),
    ]
    if addr_txt:
        lines.append(t(lang, 'courier_addr', addr=html.escape(addr_txt)))
    if b_lat is not None and b_lon is not None:
        lines.append(t(lang, 'courier_map', url=f"https://www.google.com/maps/search/?api=1&query={b_lat},{b_lon}"))
        if s_lat is not None and s_lon is not None:
            d = haversine_km(s_lat, s_lon, b_lat, b_lon)
            if d is not None:
                lines.append(t(lang, 'courier_dist', km=f"{d:.1f}"))
            lines.append(t(lang, 'courier_route', url=f"https://www.google.com/maps/dir/?api=1&origin={s_lat},{s_lon}&destination={b_lat},{b_lon}"))
    if not addr_txt and (b_lat is None or b_lon is None):
        lines.append(t(lang, 'courier_no_addr'))

    await context.bot.send_message(chat_id, "\n".join(lines), parse_mode='HTML', disable_web_page_preview=True)
    await context.bot.send_message(
        chat_id,
        t(lang, 'courier_instructions_short'),
        parse_mode='HTML'
    )


# --- Guruh (savat) buyurtma — XARIDOR tomoni ---

async def buyer_group_order_detail(update: Update, context: ContextTypes.DEFAULT_TYPE, group_id: str = None):
    """Xaridor uchun savat buyurtmasi sahifasi."""
    query = update.callback_query
    if group_id is None:
        group_id = query.data.split("_", 2)[2]  # gorder_detail_<gid>

    orders = db.get_orders_in_group(group_id)
    lang = get_lang(update, context)
    if not orders:
        await query.edit_message_text(t(lang, 'order_not_found'))
        return

    first = orders[0]
    status = first['status']
    dlv = first.get('delivery_type', 'delivery')
    pay = first.get('payment_method', 'cash')
    grand = sum(float(o['total_price']) for o in orders)

    steps = [t(lang, 'step_new'), t(lang, 'step_confirmed'),
             t(lang, 'step_delivered') if dlv == 'delivery' else t(lang, 'step_picked'),
             t(lang, 'step_rated')]
    step_idx = {'pending': 0, 'confirmed': 1, 'delivered': 2, 'cancelled': 0}.get(status, 0)
    timeline = ""
    for i, step in enumerate(steps):
        if i < step_idx:
            timeline += f"✅ {step}\n"
        elif i == step_idx and status != 'cancelled':
            timeline += f"▶️ {step}{t(lang, 'timeline_now')}\n"
        else:
            timeline += f"⬜ {step}\n"

    status_guide = {
        'pending': t(lang, 'status_guide_pending_group'),
        'confirmed': (t(lang, 'status_guide_confirmed_delivery') if dlv == 'delivery'
                      else t(lang, 'status_guide_confirmed_pickup')),
        'delivered': t(lang, 'status_guide_delivered'),
        'cancelled': t(lang, 'status_guide_cancelled'),
    }

    lines = [t(lang, 'group_order_header', oid=fmt_order_id(int(group_id)), count=len(orders))]
    for o in orders:
        lines.append(t(lang, 'seller_group_item', name=html.escape(o['product_name'] or ''),
                       qty=o['quantity'], price=fmt_price(o['product_price']), total=fmt_price(o['total_price'])))
    lines.append(t(lang, 'cart_confirm_total', total=fmt_price(grand)))
    lines.append(f"🚚 {dlv_label(dlv, lang)}")
    lines.append(f"💳 {pay_label(pay, lang)}")
    lines.append(f"🏪 {html.escape(first.get('shop_name') or '')}")

    keyboard = []
    map_button = None
    if first.get('shop_lat') and first.get('shop_lon'):
        slat, slon = first['shop_lat'], first['shop_lon']
        loc_line = t(lang, 'frag_order_address',
                     addr=html.escape(first.get('shop_address') or t(lang, 'address_word')))
        if first.get('shop_landmark'):
            loc_line += t(lang, 'frag_order_landmark', lm=html.escape(first['shop_landmark']))
        lines.append(loc_line)
        if first.get('buyer_lat') and first.get('buyer_lon'):
            d = haversine_km(first['buyer_lat'], first['buyer_lon'], slat, slon)
            if d is not None:
                lines.append(t(lang, 'dist_line_plain', km=d))
            map_button = InlineKeyboardButton(t(lang, 'btn_show_route'),
                url=f"https://www.google.com/maps/dir/{first['buyer_lat']},{first['buyer_lon']}/{slat},{slon}")
        else:
            map_button = InlineKeyboardButton(t(lang, 'btn_shop_location'),
                url=f"https://www.google.com/maps/search/?api=1&query={slat},{slon}")

    lines.append(t(lang, 'group_date_line', date=fmt_datetime(first.get('created_at'))))
    _badge = _settlement_badge(lang, first)
    if _badge:
        lines.append(_badge)
    lines.append(f"<b>{t(lang, 'label_status')}:</b>\n{timeline}")
    lines.append(f"<i>{status_guide.get(status, '')}</i>")
    if status == 'confirmed' and first.get('buyer_received'):
        lines.append(t(lang, 'buyer_awaiting_finalize'))

    if status == 'pending':
        keyboard.append([InlineKeyboardButton(t(lang, 'btn_cancel_order'), callback_data=f"gbuyer_cancel_{group_id}")])
    if status == 'confirmed' and dlv == 'pickup' and not first.get('buyer_received'):
        keyboard.append([InlineKeyboardButton(t(lang, 'btn_got_item'), callback_data=f"gbuyer_pickup_{group_id}")])
    keyboard.append([InlineKeyboardButton(t(lang, 'btn_send_message'), callback_data=f"order_msg_{first['id']}")])
    keyboard.append([InlineKeyboardButton(t(lang, 'btn_correspondence'), callback_data=f"msgs_{first['id']}")])
    # Reyting faqat buyurtma YAKUNLANGACH (sotuvchi to'lovni belgilab 'delivered' qilgach).
    # «oldim» bosilganda emas — aks holda reyting ikki marta so'ralardi.
    can_rate = (status == 'delivered')
    if can_rate:
        keyboard.append([InlineKeyboardButton(t(lang, 'btn_leave_rating'), callback_data=f"order_rate_{first['id']}")])
    if map_button:
        keyboard.append([map_button])
    if first.get('seller_username'):
        keyboard.append([InlineKeyboardButton(t(lang, 'btn_tg_at', u=first['seller_username']),
                                              url=f"https://t.me/{first['seller_username']}")])
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="buyer_orders")])

    await query.edit_message_text("\n".join(lines), reply_markup=InlineKeyboardMarkup(keyboard),
                                  parse_mode='HTML', disable_web_page_preview=True)


async def buyer_cancel_group(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xaridor savat buyurtmasini bekor qiladi (faqat pending bo'lsa)."""
    query = update.callback_query
    await query.answer()
    group_id = query.data.split("_", 2)[2]  # gbuyer_cancel_<gid>

    lang = get_lang(update, context)
    orders = db.get_orders_in_group(group_id)
    if not orders:
        await query.edit_message_text(t(lang, 'order_not_found'))
        return

    buyer = db.get_user_by_telegram_id(update.effective_user.id)
    if not buyer or buyer['id'] != orders[0].get('buyer_id'):
        await query.answer(t(lang, 'not_your_order_toast'), show_alert=True)
        return
    # ATOMIK: faqat hali 'pending' bo'lgan sub-buyurtmalarni bekor qilamiz — sotuvchi
    # ayni damda tasdiqlab ulgursa, eski bekor 'confirmed'ni bosib o'tkazib zahirani
    # yo'qotmasin.
    won = [o for o in orders if db.transition_order_status(o['id'], 'cancelled', 'pending', cancel_by='buyer')]
    if not won:
        await query.answer(t(lang, 'cant_cancel_now_toast'), show_alert=True)
        await buyer_group_order_detail(update, context, group_id=group_id)
        return

    # Taymerlarni o'chiramiz
    if context.application.job_queue:
        for nm in (f"countdown_group_{group_id}", f"auto_cancel_group_{group_id}", f"reminder_group_{group_id}"):
            for job in context.application.job_queue.get_jobs_by_name(nm):
                job.schedule_removal()

    # Sotuvchiga xabar (sotuvchi tilida)
    try:
        seller_tg = orders[0].get('seller_tg')
        if seller_tg:
            seller = db.get_user_by_id(orders[0]['seller_id'])
            await context.bot.send_message(
                chat_id=seller_tg,
                text=t(seller or 'uz', 'seller_notify_group_cancelled',
                       oid=fmt_order_id(int(group_id)), count=len(orders))
            )
    except Exception:
        pass

    await buyer_group_order_detail(update, context, group_id=group_id)


async def buyer_pickup_group(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xaridor 'oldim' — savat buyurtmasini yetkazilgan deb belgilaydi (pickup)."""
    query = update.callback_query
    await query.answer()
    group_id = query.data.split("_", 2)[2]  # gbuyer_pickup_<gid>

    lang = get_lang(update, context)
    orders = db.get_orders_in_group(group_id)
    if not orders:
        await query.edit_message_text(t(lang, 'order_not_found'))
        return
    buyer = db.get_user_by_telegram_id(update.effective_user.id)
    if not buyer or buyer['id'] != orders[0].get('buyer_id'):
        await query.answer(t(lang, 'not_your_order_toast'), show_alert=True)
        return
    if orders[0]['status'] != 'confirmed':
        await buyer_group_order_detail(update, context, group_id=group_id)
        return
    # MUHIM: «oldim» bosilsa ham savat buyurtmasi YOPILMAYDI. Status 'confirmed' qoladi —
    # sotuvchi to'lov holatini belgilab yakunlashi shart.
    db.set_group_buyer_received(group_id)

    # Sotuvchiga xabar — to'lovni belgilab yakunlashga chaqiramiz
    try:
        first = orders[0]
        if first.get('seller_tg'):
            seller = db.get_user_by_id(first['seller_id'])
            slang = get_user_lang(seller) if seller else DEFAULT_LANG
            skb = InlineKeyboardMarkup([[InlineKeyboardButton(
                t(slang, 'btn_finalize_payment'), callback_data=f"seller_gorder_{group_id}")]])
            await context.bot.send_message(
                chat_id=first['seller_tg'],
                text=t(slang, 'pickup_seller_finalize_group',
                       oid=fmt_order_id(int(group_id)), n=len(orders),
                       buyer=html.escape(first.get('buyer_name') or '')),
                reply_markup=skb,
                parse_mode='HTML'
            )
    except Exception as e:
        logging.error(f"Guruh pickup bildirishnomasi ketmadi: {e}")

    await buyer_group_order_detail(update, context, group_id=group_id)


async def auto_cancel_group_job(context: ContextTypes.DEFAULT_TYPE):
    """10 daqiqada tasdiqlanmagan savat buyurtmasini avtomatik bekor qiladi."""
    data = context.job.data
    group_id = data.get('group_id')
    orders = db.get_orders_in_group(group_id)
    if not orders:
        return
    # ATOMIK: faqat hali 'pending' bo'lganlarini bekor qilamiz — sotuvchi shu lahzada
    # tasdiqlab ulgursa, taymer 'confirmed'ni bekor qilmasin.
    pending = [o for o in orders if db.transition_order_status(o['id'], 'cancelled', 'pending', cancel_by='system')]
    if not pending:
        return
    try:
        first = orders[0]
        buyer = db.get_user_by_id(first['buyer_id'])
        seller = db.get_user_by_id(first['seller_id'])
        oid = fmt_order_id(int(group_id))
        if data.get('buyer_tg'):
            await context.bot.send_message(
                chat_id=data['buyer_tg'],
                text=t(buyer or 'uz', 'job_group_autocancel_buyer', oid=oid)
            )
        if data.get('seller_tg'):
            await context.bot.send_message(
                chat_id=data['seller_tg'],
                text=t(seller or 'uz', 'job_group_autocancel_seller', oid=oid)
            )
    except Exception as e:
        logging.error(f"auto_cancel_group_job xabar xatosi: {e}")


async def reminder_group_job(context: ContextTypes.DEFAULT_TYPE):
    """5 daqiqadan keyin sotuvchiga eslatma (savat buyurtmasi hali pending bo'lsa)."""
    data = context.job.data
    group_id = data.get('group_id')
    orders = db.get_orders_in_group(group_id)
    if not orders or orders[0]['status'] != 'pending':
        return
    try:
        if data.get('seller_tg'):
            seller = db.get_user_by_id(orders[0]['seller_id'])
            slang = get_user_lang(seller) if seller else DEFAULT_LANG
            grand = sum(float(o['total_price']) for o in orders)
            rtext = t(slang, 'job_group_reminder_seller',
                      oid=fmt_order_id(int(group_id)), n=len(orders), total=fmt_price(grand))
            rkb = InlineKeyboardMarkup([[InlineKeyboardButton(
                t(slang, 'btn_open_order'), callback_data=f"seller_gorder_{group_id}")]])
            await context.bot.send_message(chat_id=data['seller_tg'], text=rtext, reply_markup=rkb)
            # MULTI-SOTUVCHI: eslatma guruhdagi mahsulot egasi xodimlarga ham
            seen = set()
            for o in orders:
                prod = db.get_product_basic(o.get('product_id')) if o.get('product_id') else None
                if not prod:
                    continue
                cb = prod.get('created_by')
                if not cb or cb == prod.get('seller_id') or cb in seen:
                    continue
                seen.add(cb)
                su = db.get_user_by_id(cb)
                if su and su.get('telegram_id') and str(su['telegram_id']) != str(data['seller_tg']):
                    await context.bot.send_message(chat_id=su['telegram_id'], text=rtext, reply_markup=rkb)
    except Exception as e:
        logging.error(f"reminder_group_job xatosi: {e}")


# ============================================================
# SELLER PANEL
# ============================================================

async def seller_channels_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Sotuvchining ulangan kanallari ro'yxati + qo'shish/o'chirish."""
    query = update.callback_query
    if query:
        # answer allaqachon berilgan bo'lishi mumkin (masalan recheck'dan keyin) — xato bermaymiz
        try:
            await query.answer()
        except Exception:
            pass
    lang = get_lang(update, context)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    channels = db.get_seller_channels(user['id']) if user else []

    keyboard = []
    if channels:
        text = t(lang, 'channels_menu_connected') + "\n"
        has_inactive = False
        for ch in channels:
            title = ch.get('channel_title')
            if not title:
                # Sarlavha saqlanmagan bo'lsa — eng yaxshi imkon bilan aniqlaymiz
                try:
                    chat = await context.bot.get_chat(ch['channel_id'])
                    title = chat.title
                    if title:
                        db.update_seller_channel_title(user['id'], ch['channel_id'], title)
                except Exception:
                    title = ch['channel_id']
            is_group = ch.get('chat_type') in ('group', 'supergroup')
            type_icon = "👥" if is_group else "📢"
            is_active = ch.get('is_active', 1)
            if not is_active:
                has_inactive = True
            warn = "  ⚠️" if not is_active else ""
            text += f"\n{type_icon} {html.escape(str(title))}{warn}"
            label = f"🗑 {title}"
            if len(label) > 32:
                label = label[:31] + "…"
            keyboard.append([InlineKeyboardButton(label, callback_data=f"chremove_{ch['channel_id']}")])
        if has_inactive:
            text += "\n\n" + t(lang, 'channels_menu_inactive_hint')
        # Botni guruh/kanaldan chiqarmasdan, hozir post yubora olishini qayta sinab,
        # nofaol bo'lganlarini tiklash uchun.
        keyboard.append([InlineKeyboardButton(t(lang, 'btn_recheck_channels'),
                                              callback_data="seller_channels_recheck")])
    else:
        text = t(lang, 'channels_menu_empty')

    # Qayta tekshiruv natijasi (agar menyu recheck'dan keyin ochilgan bo'lsa) — tepada ko'rsatamiz
    summary = context.user_data.pop('channels_recheck_summary', None)
    if summary:
        text = summary + "\n\n" + text

    keyboard.append([InlineKeyboardButton(t(lang, 'btn_add_channel'), callback_data="seller_link_channel")])
    keyboard.append([InlineKeyboardButton(t(lang, 'btn_add_group'), callback_data="seller_link_group")])
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="seller_panel")])
    markup = InlineKeyboardMarkup(keyboard)

    if query:
        await query.edit_message_text(text, reply_markup=markup, parse_mode='HTML')
    else:
        await update.message.reply_text(text, reply_markup=markup, parse_mode='HTML')


async def seller_channel_remove(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Kanalni o'chirib, menyuni qayta ko'rsatadi."""
    query = update.callback_query
    user = db.get_user_by_telegram_id(update.effective_user.id)
    channel_id = query.data.replace("chremove_", "", 1)
    if user:
        db.remove_seller_channel(user['id'], channel_id)
    await seller_channels_menu(update, context)


# Forum (mavzuli) guruhda topic yaroqsiz bo'lganini bildiruvchi xato belgilari
_TOPIC_ERR_MARKERS = ("thread not found", "topic_closed", "topic was closed",
                      "topic_deleted", "message thread not found")


async def _probe_can_post(context, chat_id, thread_id):
    """Berilgan kanal/guruhga (forum bo'lsa topic ichiga) qisqa probe xabar yuborib,
    bot HOZIR post yubora olishini aniqlaydi. Probe darrov o'chiriladi — guruhda iz
    qolmaydi (delete imkonsiz bo'lsa ham, '✅' xabari zarar qilmaydi).

    Qaytaradi: (ok, used_general)
      ok=True          — bot post yubora oladi;
      used_general=True — forum topic yaroqsiz edi, General topicga muvaffaqiyatli yuborildi
                          (chaqiruvchi thread_id ni tozalashi kerak)."""
    async def _try(th):
        probe = await context.bot.send_message(
            chat_id=chat_id, text="✅", message_thread_id=th,
            disable_notification=True,
        )
        try:
            await context.bot.delete_message(chat_id, probe.message_id)
        except Exception:
            pass

    try:
        await _try(thread_id)
        return True, False
    except (Forbidden, BadRequest) as e:
        # Forum topic yopiq/o'chgan bo'lsa — General topicga urinib ko'ramiz
        if thread_id is not None and any(m in str(e).lower() for m in _TOPIC_ERR_MARKERS):
            try:
                await _try(None)
                return True, True
            except Exception:
                return False, False
        return False, False
    except Exception:
        return False, False


async def seller_channels_recheck(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Ulangan barcha kanal/guruhlarni qayta tekshiradi: bot hozir post yubora olsa —
    nofaol bo'lib qolganlarini QAYTA FAOLLASHTIRADI. Shu sababli sotuvchi botni
    guruhdan chiqarib qayta qo'shishi shart emas (admin qilib qo'ygach shu tugma yetadi)."""
    query = update.callback_query
    lang = get_lang(update, context)
    if query:
        try:
            await query.answer(t(lang, 'channels_recheck_running'))
        except Exception:
            pass
    user = db.get_user_by_telegram_id(update.effective_user.id)
    channels = db.get_seller_channels(user['id']) if user else []

    reactivated = 0  # nofaol edi → endi ishladi
    failing = 0      # hali ham yubora olmadi
    for ch in channels:
        cid = ch.get('channel_id')
        th = ch.get('thread_id')
        th = int(th) if th not in (None, '') else None
        was_active = bool(ch.get('is_active', 1))
        ok, used_general = await _probe_can_post(context, cid, th)
        if ok:
            if used_general:
                # Eski topic yaroqsiz — bundan keyin General topicga yuboramiz
                db.set_seller_channel_thread(user['id'], cid, None)
            if not was_active:
                # add_seller_channel mavjud yozuvni faollashtiradi (is_active=1, xato tozalanadi)
                db.add_seller_channel(user['id'], cid, ch.get('channel_title'))
                reactivated += 1
        else:
            if was_active:
                db.deactivate_seller_channel(user['id'], cid, 'recheck_no_post')
            failing += 1

    # Natijani menyu tepasida bir martalik banner sifatida ko'rsatamiz
    context.user_data['channels_recheck_summary'] = t(
        lang, 'channels_recheck_done', ok=reactivated, bad=failing)
    await seller_channels_menu(update, context)


async def seller_link_channel_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Kanalni ulash: yo'riqnoma ko'rsatadi va kanal postini (forward) kutadi."""
    query = update.callback_query
    if query:
        await query.answer()
    lang = get_lang(update, context)
    bot_me = await context.bot.get_me()
    text = t(lang, 'link_channel_prompt', bot=bot_me.username)
    if query:
        await query.edit_message_text(text, parse_mode='HTML', disable_web_page_preview=True)
    else:
        await update.message.reply_text(text, parse_mode='HTML', disable_web_page_preview=True)
    return LINK_CHANNEL_WAIT


async def link_channel_wait_hint(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Forward o'rniga oddiy matn kelsa — qayta eslatma."""
    lang = get_lang(update, context)
    await update.message.reply_text(t(lang, 'link_channel_not_channel'), parse_mode='HTML')
    return LINK_CHANNEL_WAIT


async def link_channel_receive(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Forward qilingan kanal postidan kanalni aniqlaydi, botning adminligini
    tekshiradi va sotuvchi <-> kanal bog'lanishini saqlaydi."""
    msg = update.message
    lang = get_lang(update, context)
    user = db.get_user_by_telegram_id(update.effective_user.id)

    # Kanalni aniqlash (PTB versiyasiga bardoshli: forward_origin yoki forward_from_chat)
    channel_chat = None
    origin = getattr(msg, 'forward_origin', None)
    if origin is not None and getattr(origin, 'type', None) == 'channel':
        channel_chat = getattr(origin, 'chat', None)
    if channel_chat is None:
        fwd = getattr(msg, 'forward_from_chat', None)
        if fwd is not None and getattr(fwd, 'type', None) == 'channel':
            channel_chat = fwd
    if channel_chat is None:
        await msg.reply_text(t(lang, 'link_channel_not_channel'), parse_mode='HTML')
        return LINK_CHANNEL_WAIT

    bot_me = await context.bot.get_me()

    # Bot o'sha kanalda admin va post yubora oladimi?
    try:
        member = await context.bot.get_chat_member(channel_chat.id, bot_me.id)
    except Exception:
        member = None
    if member is None or member.status not in ('administrator', 'creator'):
        await msg.reply_text(t(lang, 'link_channel_not_admin', bot=bot_me.username), parse_mode='HTML')
        return LINK_CHANNEL_WAIT
    if member.status == 'administrator' and not getattr(member, 'can_post_messages', False):
        await msg.reply_text(t(lang, 'link_channel_no_post_perm'), parse_mode='HTML')
        return LINK_CHANNEL_WAIT

    # Boshqa sotuvchi shu kanalni ulaganmi? (#5 — yumshoq ogohlantirish, bloklamaymiz)
    other_owners = db.find_channel_owners(channel_chat.id, exclude_seller_id=user['id']) if user else []

    # Saqlaymiz (ko'p kanal — seller_channels jadvali)
    title = channel_chat.title or str(channel_chat.id)
    is_new = db.add_seller_channel(user['id'], channel_chat.id, title) if user else False
    if is_new:
        await msg.reply_text(
            t(lang, 'link_channel_success', title=html.escape(title)),
            parse_mode='HTML'
        )
    else:
        await msg.reply_text(t(lang, 'link_channel_already'), parse_mode='HTML')
    if other_owners:
        await msg.reply_text(t(lang, 'link_channel_shared_warn'), parse_mode='HTML')
    await seller_channels_menu(update, context)
    return ConversationHandler.END


async def seller_link_group_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Guruhni ulash yo'riqnomasi.

    Kanaldan farqli — guruh ID si forward orqali ko'rinmaydi. Shu sababli bog'lash
    sotuvchi botni guruhga QO'SHGANDA avtomatik (my_chat_member orqali) amalga oshadi;
    bu yerda alohida holat (state) kerak emas — faqat ko'rsatma beramiz."""
    query = update.callback_query
    if query:
        await query.answer()
    lang = get_lang(update, context)
    bot_me = await context.bot.get_me()
    text = t(lang, 'link_group_prompt', bot=bot_me.username)
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton(t(lang, 'back'), callback_data="seller_channels_menu")]
    ])
    if query:
        await query.edit_message_text(text, reply_markup=keyboard,
                                      parse_mode='HTML', disable_web_page_preview=True)
    else:
        await update.message.reply_text(text, reply_markup=keyboard,
                                        parse_mode='HTML', disable_web_page_preview=True)


async def on_my_chat_member(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Botning guruh/superguruhdagi a'zoligi o'zgarganda ishlaydi.

    Bot guruh YOKI kanalga qo'shilsa — qo'shgan odamni (sotuvchini) aniqlaymiz va
    chatni unga bog'laymiz. Endi yangi mahsulotlar avtomatik o'sha yerga ham chiqadi.

    Eslatma: Telegram 'my_chat_member' yangilanishini guruh, superguruh VA kanal
    uchun ham yuboradi (bot admin qilinganda/qo'shilganda). Ilgari kanallar faqat
    'forward' orqali ulanardi, lekin bot→launcher migratsiyasidan keyin forward
    oqimi App'ga ko'chib o'lik qoldi — shuning uchun kanallarni ham shu yerda
    ushlaymiz. (Alohida 'chat_member' obunasi shart emas.)"""
    cmu = update.my_chat_member
    if cmu is None:
        return
    chat = cmu.chat
    _old = getattr(cmu.old_chat_member, 'status', None)
    _new = getattr(cmu.new_chat_member, 'status', None)
    logging.info(f"my_chat_member: chat={chat.id} type={chat.type} "
                 f"status {_old}->{_new} by={getattr(cmu.from_user, 'id', None)}")
    if chat.type not in ('group', 'supergroup', 'channel'):
        return
    is_channel = (chat.type == 'channel')

    old_status = cmu.old_chat_member.status if cmu.old_chat_member else None
    new_status = cmu.new_chat_member.status if cmu.new_chat_member else None

    OUTSIDE = ('left', 'kicked')
    INSIDE = ('member', 'administrator', 'creator', 'restricted')
    ADMIN = ('administrator', 'creator')

    # Ulanishni (qayta) o'rnatadigan IKKI holat:
    #   1) just_joined — bot guruhga YANGI qo'shildi (tashqaridan ichkariga);
    #   2) promoted    — bot ichkarida turib ADMIN huquqini oldi (member->administrator).
    # 2-holat MUHIM: ko'p guruhlarda a'zolarga yozish taqiqlangan, shuning uchun bot avval
    # oddiy a'zo bo'lib qo'shilganda post yubora olmaydi va guruh "nofaol" bo'lib qoladi.
    # Keyin sotuvchi botni admin qilganda aynan shu yangilanish keladi — uni ham ushlab,
    # guruhni qayta faollashtirib, post yuborishni qaytadan sinaymiz. Aks holda bot
    # abadiy "nofaol" qolib, hech qachon xabar yubora olmaydi.
    just_joined = (new_status in INSIDE) and (old_status in OUTSIDE or old_status is None)
    promoted = (new_status in ADMIN) and (old_status not in ADMIN) and not just_joined
    if not (just_joined or promoted):
        return

    # Eslatma: sotuvchilar guruhga ortiqcha xabar yozilishidan norozi edi. Shu sababli
    # bot guruhga UMUMAN hech narsa yozmaydi — barcha bildirishnomalar sotuvchining
    # shaxsiy chatiga (DM) yuboriladi.
    actor = cmu.from_user  # odatda — botni guruhga qo'shgan kishi
    seller = db.get_user_by_telegram_id(actor.id) if actor else None
    approved = bool(seller and (seller.get('is_approved') or seller.get('role') in ('seller', 'admin')))

    try:
        bot_me = await context.bot.get_me()
    except Exception:
        bot_me = None
    bot_username = bot_me.username if bot_me else "TezBozor"
    title = chat.title or str(chat.id)

    # Qo'shgan odam tasdiqlangan sotuvchi emas — guruhga hech narsa yozmaymiz, jim chiqamiz.
    if not approved:
        return

    slang = get_user_lang(seller)

    # Mavzuli (forum) guruh bo'lsa — postlar "General" topicga emas, alohida ochilgan
    # topicga borishi uchun topic yaratishga harakat qilamiz (bot admin + topic boshqaruvi bo'lsa).
    is_forum = bool(getattr(chat, 'is_forum', False))
    thread_id = None
    if is_forum:
        # Avval shu guruh uchun saqlangan topic bormi (qayta ulanish)?
        for ch in db.get_seller_channels(seller['id']):
            if str(ch.get('channel_id')) == str(chat.id) and ch.get('thread_id'):
                try:
                    thread_id = int(ch['thread_id'])
                except Exception:
                    thread_id = None
                break
        if thread_id is None:
            try:
                topic = await context.bot.create_forum_topic(
                    chat_id=chat.id, name="🛒 Yangi mahsulotlar")
                thread_id = topic.message_thread_id
            except Exception as e:
                # Huquq yo'q — post General topicga boradi (topic yopiq bo'lsa avtomatik unga qaytamiz)
                logging.info(f"Forum topic ochib bo'lmadi (chat {chat.id}): {e} — General ishlatiladi")

    is_new = db.add_seller_channel(seller['id'], chat.id, title,
                                   chat_type=('channel' if is_channel else 'group'),
                                   is_forum=is_forum, thread_id=thread_id)

    # Bot post yubora oladimi — guruhga HECH NARSA yozmasdan, botning a'zolik huquqidan
    # aniqlaymiz (avval bu "jonli test" sifatida guruhga tasdiq xabari yuborilardi, lekin
    # sotuvchilar guruhga ortiqcha xabardan norozi edi).
    can_post = True
    try:
        if bot_me is not None:
            bot_member = await context.bot.get_chat_member(chat.id, bot_me.id)
            bstatus = getattr(bot_member, 'status', None)
            if bstatus in ('administrator', 'creator'):
                can_post = True
            elif bstatus in ('member', 'restricted'):
                # 'restricted' a'zoda can_send_messages aniq beriladi; oddiy 'member'da
                # huquq guruh sozlamasiga bog'liq — True deb hisoblaymiz. Agar baribir
                # yubora olmasa, birinchi mahsulot posti muvaffaqiyatsiz bo'lib,
                # post_product_to_channel kanalni avtomatik o'chiradi va sotuvchini ogohlantiradi.
                can_post = bool(getattr(bot_member, 'can_send_messages', True))
            else:
                can_post = False
    except Exception as e:
        logging.warning(f"Bot a'zolik huquqini tekshirib bo'lmadi (chat {chat.id}): {e}")

    # Bot guruhda post yubora olmasa — uni FAOL ro'yxatda qoldirmaymiz (aks holda har postda
    # behuda urinish bo'ladi). Sotuvchi botni admin qilib qaytadan ulashi mumkin.
    if not can_post:
        db.deactivate_seller_channel(seller['id'], chat.id, 'no_post_permission_on_join')

    # Sotuvchiga shaxsiy xabar (u botni allaqachon ishga tushirgan — DM yetadi).
    try:
        seller_tg = seller.get('telegram_id')
        if seller_tg:
            if not can_post:
                key = 'channel_linked_need_admin' if is_channel else 'group_linked_need_admin'
            elif is_new:
                key = 'channel_linked_notify' if is_channel else 'group_linked_notify'
            else:
                key = 'channel_relinked_notify' if is_channel else 'group_relinked_notify'
            await context.bot.send_message(
                chat_id=seller_tg,
                text=t(slang, key, title=html.escape(title)),
                parse_mode='HTML', disable_web_page_preview=True,
            )
    except Exception as e:
        logging.warning(f"Sotuvchiga guruh-bog'lanish xabari yuborilmadi: {e}")


async def seller_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = db.get_user_by_telegram_id(update.effective_user.id)
    context.user_data['active_mode'] = 'seller'
    lang = get_lang(update, context)

    # MULTI-SOTUVCHI: tasdiqlanmagan (nofaol) xodim — kutish ekrani
    _srec = db.get_staff_by_user(user['id']) if user else None
    if _srec and _srec.get('staff_role') != 'owner' and not _srec.get('is_active'):
        text = t(lang, 'staff_pending_panel')
        keyboard = [
            [InlineKeyboardButton(t(lang, 'btn_buyer_mode'), callback_data="switch_to_buyer_confirm")],
            [InlineKeyboardButton(t(lang, 'btn_contact_admin'), callback_data="contact_admin")],
        ]
        if update.callback_query:
            await update.callback_query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')
        else:
            await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')
        return

    # Tasdiqlanmagan sotuvchi — maxsus panel ko'rsatamiz
    if user and not user.get('is_approved') and user.get('role') != 'admin':
        req = db.get_seller_request_by_user(user['id'])
        req_status = req['status'] if req else None

        if req_status == 'rejected':
            # Rad etilgan — qayta so'rov yuborish imkoniyati
            text = t(lang, 'seller_rejected_panel')
            keyboard = [
                [InlineKeyboardButton(t(lang, 'btn_reapply'), callback_data="reapply_seller")],
                [InlineKeyboardButton(t(lang, 'btn_contact_admin'), callback_data="contact_admin")],
                [InlineKeyboardButton(t(lang, 'btn_buyer_mode'), callback_data="switch_to_buyer_confirm")],
            ]
        else:
            # Kutilmoqda (pending) yoki boshqa holat
            text = t(lang, 'seller_not_approved')
            keyboard = [
                [InlineKeyboardButton(t(lang, 'btn_contact_admin'), callback_data="contact_admin")],
                [InlineKeyboardButton(t(lang, 'btn_buyer_mode'), callback_data="switch_to_buyer_confirm")],
            ]

        if update.callback_query:
            await update.callback_query.edit_message_text(
                text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML'
            )
        else:
            await update.message.reply_text(
                text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML'
            )
        return

    # Tugmalar bo'limlarga guruhlangan — har bo'limga bossa ichidagi tugmalar ochiladi.
    # (Eng ko'p ishlatiladigan «Mahsulot qo'shish» va «AI yordamchi» tepada qoladi.)
    keyboard = [
        [InlineKeyboardButton(t(lang, 'btn_add_product'), callback_data="seller_add_product")],
        [InlineKeyboardButton(t(lang, 'btn_ai_assistant'), callback_data="ai_assistant")],
        [InlineKeyboardButton(t(lang, 'grp_products'), callback_data="sellergrp_products")],
        [InlineKeyboardButton(t(lang, 'grp_sales'), callback_data="sellergrp_sales")],
        [InlineKeyboardButton(t(lang, 'grp_customers'), callback_data="sellergrp_customers")],
        [InlineKeyboardButton(t(lang, 'grp_settings'), callback_data="sellergrp_settings")],
        [InlineKeyboardButton(t(lang, 'btn_buyer_mode'), callback_data="switch_to_buyer_confirm")],
    ]

    if CHANNEL_URL:
        keyboard.append([InlineKeyboardButton(t(lang, 'btn_official_channel'), url=CHANNEL_URL)])

    text = t(lang, 'seller_panel_full',
             shop=user.get('shop_name') or t(lang, 'not_specified'),
             address=user.get('shop_address') or t(lang, 'not_specified'))

    if update.callback_query:
        await update.callback_query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard))
    else:
        await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard))
        await update.message.reply_text(
            t(lang, 'bottom_hint'),
            reply_markup=seller_bottom_kb(lang)
        )


# ============================================================
# SOTUVCHI PANELI — BO'LIM (guruh) MENYULARI
# Asosiy panel tartibli bo'lishi uchun tugmalar 4 bo'limga bo'lingan.
# Har bo'lim — alohida ekran, ichidagi tugmalar mavjud handlerlarga boradi.
# ============================================================
async def _show_seller_group(update, context, title, kb):
    """Bo'lim ekranini ko'rsatadi (Orqaga — asosiy panelga)."""
    kb = kb + [[InlineKeyboardButton(t(get_lang(update, context), 'back'), callback_data="seller_panel")]]
    query = update.callback_query
    markup = InlineKeyboardMarkup(kb)
    if query:
        await query.answer()
        try:
            await query.edit_message_text(title, reply_markup=markup, parse_mode='HTML')
        except Exception:
            await context.bot.send_message(chat_id=update.effective_chat.id, text=title,
                                           reply_markup=markup, parse_mode='HTML')
    else:
        await update.message.reply_text(title, reply_markup=markup, parse_mode='HTML')


async def seller_group_products(update, context):
    """📦 Sotuv va e'lonlar — mahsulot, kanal, rejalashtirilgan post, avto-reklama."""
    lang = get_lang(update, context)
    kb = [
        [InlineKeyboardButton(t(lang, 'btn_my_products'), callback_data="seller_products")],
        [InlineKeyboardButton(t(lang, 'btn_my_channels'), callback_data="seller_channels_menu")],
        [InlineKeyboardButton(t(lang, 'btn_scheduled_posts'), callback_data="seller_scheduled")],
        [InlineKeyboardButton(t(lang, 'btn_autoreposts'), callback_data="seller_autoreposts")],
    ]
    await _show_seller_group(update, context, t(lang, 'grp_products_title'), kb)


async def seller_group_sales(update, context):
    """🛒 Savdo va hisob — buyurtmalar, qarzlar, statistika/Excel."""
    lang = get_lang(update, context)
    kb = [
        [InlineKeyboardButton(t(lang, 'btn_orders'), callback_data="seller_orders")],
        [InlineKeyboardButton(t(lang, 'btn_debts'), callback_data="seller_debts")],
        [InlineKeyboardButton(t(lang, 'btn_stats'), callback_data="seller_stats")],
    ]
    await _show_seller_group(update, context, t(lang, 'grp_sales_title'), kb)


async def seller_group_customers(update, context):
    """💬 Mijozlar — xabarlar, sharhlar."""
    lang = get_lang(update, context)
    kb = [
        [InlineKeyboardButton(t(lang, 'btn_messages'), callback_data="seller_messages")],
        [InlineKeyboardButton(t(lang, 'btn_seller_reviews'), callback_data="seller_reviews")],
    ]
    await _show_seller_group(update, context, t(lang, 'grp_customers_title'), kb)


async def seller_group_settings(update, context):
    """⚙️ Sozlamalar — profil, (ega uchun) sotuvchilar, admin bilan bog'lanish."""
    lang = get_lang(update, context)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    kb = [[InlineKeyboardButton(t(lang, 'btn_profile'), callback_data="seller_profile")]]
    # MULTI-SOTUVCHI: ega/manager bo'lsa — "Sotuvchilar" boshqaruv tugmasi
    _staff_rec = db.get_staff_by_user(user['id']) if user else None
    if _staff_rec and _staff_rec.get('staff_role') in ('owner', 'manager'):
        kb.append([InlineKeyboardButton(t(lang, 'btn_manage_staff'), callback_data="staff_panel")])
    kb.append([InlineKeyboardButton(t(lang, 'btn_contact_admin'), callback_data="contact_admin")])
    await _show_seller_group(update, context, t(lang, 'grp_settings_title'), kb)


# ============================================================
# MULTI-SOTUVCHI: do'kon egasi paneli (sotuvchilarni boshqarish)
# ============================================================

PERM_KEYS = {
    'add':   'perm_add_product',
    'conf':  'perm_confirm_orders',
    'price': 'perm_edit_price',
    'rev':   'perm_reply_reviews',
}


async def _require_owner(update, context):
    """Joriy foydalanuvchi do'kon egasi/manageri ekanini tekshiradi.
    (user, shop, staff_rec) yoki (None, None, None) qaytaradi (xabar bilan)."""
    lang = get_lang(update, context)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    staff_rec = db.get_staff_by_user(user['id']) if user else None
    shop = db.get_shop_for_user(user['id']) if user else None
    if not (staff_rec and shop and staff_rec.get('staff_role') in ('owner', 'manager')):
        q = update.callback_query
        if q:
            await q.answer(t(lang, 'staff_owner_only'), show_alert=True)
        return None, None, None
    return user, shop, staff_rec


async def staff_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Sotuvchilarni boshqarish — asosiy menyu (faqat ega/manager)."""
    query = update.callback_query
    if query:
        await query.answer()
    lang = get_lang(update, context)
    user, shop, _ = await _require_owner(update, context)
    if not shop:
        return
    staff_all = db.get_shop_staff(shop['id'], include_owner=False)
    active_n = sum(1 for s in staff_all if s.get('is_active'))
    pending_n = sum(1 for s in staff_all if not s.get('is_active'))
    pay_mode = shop.get('payment_mode') or 'shop'
    pay_lbl = t(lang, 'paymode_shop' if pay_mode == 'shop' else 'paymode_staff')

    text = t(lang, 'staff_panel_text', total=len(staff_all), active=active_n, pending=pending_n,
             paymode=pay_lbl, mod=t(lang, 'mod_owner' if shop.get('moderation') == 'owner_approve' else 'mod_direct'))
    inv_n = len(db.get_active_invites(shop['id']))
    kb = [
        [InlineKeyboardButton(t(lang, 'btn_staff_list'), callback_data="staff_list")],
        [InlineKeyboardButton(t(lang, 'btn_staff_add'), callback_data="staff_add")],
        [InlineKeyboardButton(t(lang, 'btn_staff_invites', n=inv_n), callback_data="staff_invites")],
        [InlineKeyboardButton(t(lang, 'btn_staff_stats'), callback_data="staff_stats")],
        [InlineKeyboardButton(t(lang, 'btn_paymode', mode=pay_lbl), callback_data="shop_paymode")],
    ]
    if shop.get('moderation') == 'owner_approve':
        cnt = len(db.get_seller_products_by_status(shop['owner_user_id'], 'pending_owner'))
        kb.append([InlineKeyboardButton(t(lang, 'btn_pending_products', n=cnt), callback_data="shop_pending")])
    kb.append([InlineKeyboardButton(t(lang, 'back'), callback_data="seller_panel")])

    if query:
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(kb), parse_mode='HTML')
    else:
        await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(kb), parse_mode='HTML')


async def staff_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Do'kondagi xodimlar ro'yxati."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    user, shop, _ = await _require_owner(update, context)
    if not shop:
        return
    staff_all = db.get_shop_staff(shop['id'], include_owner=False)
    if not staff_all:
        await query.edit_message_text(
            t(lang, 'staff_list_empty'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="staff_panel")]]),
            parse_mode='HTML')
        return
    kb = []
    for s in staff_all:
        mark = "✅" if s.get('is_active') else "⏳"
        dept = s.get('department') or '—'
        kb.append([InlineKeyboardButton(f"{mark} {s.get('name') or '—'} · {dept}",
                                        callback_data=f"staff_detail_{s['id']}")])
    kb.append([InlineKeyboardButton(t(lang, 'back'), callback_data="staff_panel")])
    await query.edit_message_text(t(lang, 'staff_list_header'),
                                  reply_markup=InlineKeyboardMarkup(kb), parse_mode='HTML')


async def staff_detail(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Bitta xodim — ma'lumot, statistika va amallar."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    user, shop, _ = await _require_owner(update, context)
    if not shop:
        return
    staff_id = int(query.data.split("_")[2])
    # Xodim shu do'konga tegishlimi?
    target = next((s for s in db.get_shop_staff(shop['id']) if s['id'] == staff_id), None)
    if not target:
        await query.answer(t(lang, 'staff_not_found'), show_alert=True)
        return
    st = db.get_staff_stats(target['user_id'])
    perms = []
    for k, col in PERM_KEYS.items():
        perms.append(("✅" if target.get(col) else "❌") + " " + t(lang, f'perm_{k}'))
    role_name = t(lang, 'role_manager' if target.get('staff_role') == 'manager' else 'role_staff')
    text = t(lang, 'staff_detail_text',
             name=html.escape(target.get('name') or '—'),
             dept=html.escape(target.get('department') or '—'),
             role=role_name,
             status=t(lang, 'staff_active' if target.get('is_active') else 'staff_pending'),
             products=st['products_count'], delivered=st['delivered'],
             revenue=fmt_price(st['total_revenue']), pending=st['pending'],
             perms="\n".join(perms))
    toggle_lbl = t(lang, 'btn_staff_freeze' if target.get('is_active') else 'btn_staff_activate')
    role_lbl = t(lang, 'btn_staff_make_staff' if target.get('staff_role') == 'manager' else 'btn_staff_make_manager')
    kb = [
        [InlineKeyboardButton(toggle_lbl, callback_data=f"staff_toggle_{staff_id}")],
        [InlineKeyboardButton(t(lang, 'btn_staff_set_dept'), callback_data=f"staff_dept_{staff_id}")],
        [InlineKeyboardButton(role_lbl, callback_data=f"staff_role_{staff_id}")],
        [InlineKeyboardButton(t(lang, 'btn_staff_perms'), callback_data=f"staff_perm_{staff_id}")],
        [InlineKeyboardButton(t(lang, 'btn_staff_remove'), callback_data=f"staff_rm_{staff_id}")],
        [InlineKeyboardButton(t(lang, 'back'), callback_data="staff_list")],
    ]
    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(kb), parse_mode='HTML')


async def staff_toggle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xodimni muzlatish/faollashtirish."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    user, shop, _ = await _require_owner(update, context)
    if not shop:
        return
    staff_id = int(query.data.split("_")[2])
    target = next((s for s in db.get_shop_staff(shop['id'], include_owner=False) if s['id'] == staff_id), None)
    if not target:
        await query.answer(t(lang, 'staff_not_found'), show_alert=True)
        return
    new_active = 0 if target.get('is_active') else 1
    db.set_staff_active(staff_id, new_active)
    # Xodimga xabar
    try:
        su = db.get_user_by_id(target['user_id'])
        if su and su.get('telegram_id'):
            slang = get_user_lang(su)
            await context.bot.send_message(
                chat_id=su['telegram_id'],
                text=t(slang, 'staff_you_activated' if new_active else 'staff_you_frozen'))
    except Exception as e:
        logging.error(f"Xodimga holat xabari ketmadi: {e}")
    await staff_detail(update, context)


async def staff_perm_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xodim ruxsatlarini sozlash menyusi."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    user, shop, _ = await _require_owner(update, context)
    if not shop:
        return
    parts = query.data.split("_")
    staff_id = int(parts[2])
    target = next((s for s in db.get_shop_staff(shop['id'], include_owner=False) if s['id'] == staff_id), None)
    if not target:
        await query.answer(t(lang, 'staff_not_found'), show_alert=True)
        return
    # staff_pset_<id>_<key> — toggle
    if len(parts) >= 4 and parts[1] == 'pset':
        key = parts[3]
        col = PERM_KEYS.get(key)
        if col:
            db.update_staff(staff_id, **{col: 0 if target.get(col) else 1})
            target = next((s for s in db.get_shop_staff(shop['id'], include_owner=False) if s['id'] == staff_id), None)
    kb = []
    for k, col in PERM_KEYS.items():
        mark = "✅" if target.get(col) else "❌"
        kb.append([InlineKeyboardButton(f"{mark} {t(lang, f'perm_{k}')}",
                                        callback_data=f"staff_pset_{staff_id}_{k}")])
    kb.append([InlineKeyboardButton(t(lang, 'back'), callback_data=f"staff_detail_{staff_id}")])
    await query.edit_message_text(t(lang, 'staff_perms_header', name=html.escape(target.get('name') or '—')),
                                  reply_markup=InlineKeyboardMarkup(kb), parse_mode='HTML')


async def staff_remove(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xodimni o'chirish (tasdiq bilan)."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    user, shop, _ = await _require_owner(update, context)
    if not shop:
        return
    parts = query.data.split("_")
    staff_id = int(parts[2])
    confirmed = (len(parts) >= 4 and parts[3] == 'yes')
    target = next((s for s in db.get_shop_staff(shop['id'], include_owner=False) if s['id'] == staff_id), None)
    if not target:
        await query.answer(t(lang, 'staff_not_found'), show_alert=True)
        return
    if not confirmed:
        kb = [
            [InlineKeyboardButton(t(lang, 'btn_staff_remove_yes'), callback_data=f"staff_rm_{staff_id}_yes")],
            [InlineKeyboardButton(t(lang, 'back'), callback_data=f"staff_detail_{staff_id}")],
        ]
        await query.edit_message_text(t(lang, 'staff_remove_confirm', name=html.escape(target.get('name') or '—')),
                                      reply_markup=InlineKeyboardMarkup(kb), parse_mode='HTML')
        return
    db.remove_staff(staff_id)
    await query.edit_message_text(
        t(lang, 'staff_removed_done', name=html.escape(target.get('name') or '—')),
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="staff_list")]]),
        parse_mode='HTML')


async def staff_add(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Yangi sotuvchi taklifi — avval bo'lim nomini so'raydi."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    user, shop, _ = await _require_owner(update, context)
    if not shop:
        return
    context.user_data['staff_invite_dept'] = shop['id']
    kb = [
        [InlineKeyboardButton(t(lang, 'btn_skip_dept'), callback_data="staff_add_nodept")],
        [InlineKeyboardButton(t(lang, 'back'), callback_data="staff_panel")],
    ]
    await query.edit_message_text(t(lang, 'staff_add_ask_dept'),
                                  reply_markup=InlineKeyboardMarkup(kb), parse_mode='HTML')


async def _send_invite_link(update, context, shop, department, user_id):
    """Taklif kodi + deeplink yaratib ko'rsatadi (callback yoki matn javobi orqali)."""
    lang = get_lang(update, context)
    code = db.create_invite(shop['id'], department=department, created_by=user_id)
    bot_me = await context.bot.get_me()
    link = f"https://t.me/{bot_me.username}?start=staff_{code}"
    dept_line = department or t(lang, 'not_specified')
    text = t(lang, 'staff_invite_created', link=link, code=code, dept=html.escape(dept_line))
    kb = [[InlineKeyboardButton(t(lang, 'back'), callback_data="staff_panel")]]
    if update.callback_query:
        await update.callback_query.edit_message_text(
            text, reply_markup=InlineKeyboardMarkup(kb), parse_mode='HTML', disable_web_page_preview=True)
    else:
        await update.message.reply_text(
            text, reply_markup=InlineKeyboardMarkup(kb), parse_mode='HTML', disable_web_page_preview=True)


async def staff_add_nodept(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Bo'limsiz taklif yaratish."""
    query = update.callback_query
    await query.answer()
    context.user_data.pop('staff_invite_dept', None)
    user, shop, _ = await _require_owner(update, context)
    if not shop:
        return
    await _send_invite_link(update, context, shop, None, user['id'])


async def staff_invites(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Faol (ishlatilmagan) takliflar ro'yxati — har birini bekor qilish mumkin."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    user, shop, _ = await _require_owner(update, context)
    if not shop:
        return
    invites = db.get_active_invites(shop['id'])
    if not invites:
        await query.edit_message_text(
            t(lang, 'invites_empty'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="staff_panel")]]),
            parse_mode='HTML')
        return
    kb = []
    for inv in invites[:20]:
        dept = inv.get('department') or t(lang, 'not_specified')
        kb.append([InlineKeyboardButton(f"🔗 {inv['code']} · {dept}"[:50], callback_data="noop")])
        kb.append([InlineKeyboardButton(t(lang, 'btn_invite_cancel'), callback_data=f"inv_cancel_{inv['id']}")])
    kb.append([InlineKeyboardButton(t(lang, 'back'), callback_data="staff_panel")])
    await query.edit_message_text(t(lang, 'invites_header'),
                                  reply_markup=InlineKeyboardMarkup(kb), parse_mode='HTML')


async def invite_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Faol taklifni bekor qiladi (o'chiradi)."""
    query = update.callback_query
    await query.answer()
    user, shop, _ = await _require_owner(update, context)
    if not shop:
        return
    invite_id = int(query.data.split("_")[2])
    db.delete_invite(invite_id, shop_id=shop['id'])
    await staff_invites(update, context)


async def staff_reject(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Qo'shilgan (yoki kutilayotgan) xodimni rad etadi/o'chiradi — notog'ri odam qo'shilsa."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    user, shop, _ = await _require_owner(update, context)
    if not shop:
        return
    staff_id = int(query.data.split("_")[2])
    target = next((s for s in db.get_shop_staff(shop['id'], include_owner=False) if s['id'] == staff_id), None)
    if not target:
        await query.edit_message_text(t(lang, 'staff_not_found'))
        return
    uid = target['user_id']
    db.remove_staff(staff_id)
    try:
        su = db.get_user_by_id(uid)
        if su and su.get('telegram_id'):
            await context.bot.send_message(
                chat_id=su['telegram_id'],
                text=t(get_user_lang(su), 'staff_join_rejected', shop=html.escape(shop.get('name') or '—')),
                parse_mode='HTML')
    except Exception as e:
        logging.error(f"Rad etilgan xodimga xabar ketmadi: {e}")
    await query.edit_message_text(
        t(lang, 'staff_reject_done', name=html.escape(target.get('name') or '—')),
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'btn_manage_staff'), callback_data="staff_panel")]]),
        parse_mode='HTML')


async def join_with_code_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Foydalanuvchi taklif kodini qo'lda kiritib do'konga qo'shilmoqchi."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    if user and db.get_staff_by_user(user['id']):
        await query.answer(t(lang, 'staff_already_member'), show_alert=True)
        return
    context.user_data['joining_with_code'] = True
    kb = [[InlineKeyboardButton(t(lang, 'back'), callback_data="buyer_panel")]]
    await query.edit_message_text(t(lang, 'join_code_ask'),
                                  reply_markup=InlineKeyboardMarkup(kb), parse_mode='HTML')


async def staff_role_toggle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xodim rolini almashtiradi: oddiy xodim ↔ manager."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    user, shop, _ = await _require_owner(update, context)
    if not shop:
        return
    staff_id = int(query.data.split("_")[2])
    target = next((s for s in db.get_shop_staff(shop['id'], include_owner=False) if s['id'] == staff_id), None)
    if not target:
        await query.answer(t(lang, 'staff_not_found'), show_alert=True)
        return
    new_role = 'staff' if target.get('staff_role') == 'manager' else 'manager'
    db.update_staff(staff_id, staff_role=new_role)
    await staff_detail(update, context)


async def staff_set_dept_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xodim bo'limini o'zgartirish — matn so'raydi."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    user, shop, _ = await _require_owner(update, context)
    if not shop:
        return
    staff_id = int(query.data.split("_")[2])
    context.user_data['staff_set_dept_for'] = staff_id
    kb = [[InlineKeyboardButton(t(lang, 'back'), callback_data=f"staff_detail_{staff_id}")]]
    await query.edit_message_text(t(lang, 'staff_set_dept_ask'),
                                  reply_markup=InlineKeyboardMarkup(kb), parse_mode='HTML')


async def staff_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Har bir xodim bo'yicha sotuv ko'rsatkichi."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    user, shop, _ = await _require_owner(update, context)
    if not shop:
        return
    rows = db.get_shop_staff_performance(shop['id'])
    lines = [t(lang, 'staff_stats_header')]
    for r in rows:
        role_mark = "👑" if r.get('staff_role') == 'owner' else "•"
        lines.append(t(lang, 'staff_stats_row',
                       mark=role_mark, name=html.escape(r.get('name') or '—'),
                       products=r.get('products_count', 0), sold=r.get('sold', 0),
                       revenue=fmt_price(r.get('revenue', 0))))
    kb = [[InlineKeyboardButton(t(lang, 'back'), callback_data="staff_panel")]]
    await query.edit_message_text("\n".join(lines), reply_markup=InlineKeyboardMarkup(kb), parse_mode='HTML')


async def shop_paymode_toggle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """To'lov rejimini almashtiradi: do'kon kartasi ↔ har xodim kartasi."""
    query = update.callback_query
    await query.answer()
    user, shop, _ = await _require_owner(update, context)
    if not shop:
        return
    new_mode = 'staff' if (shop.get('payment_mode') or 'shop') == 'shop' else 'shop'
    db.update_shop(shop['id'], payment_mode=new_mode)
    await staff_panel(update, context)


async def shop_pending_products(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Ega tasdig'ini kutayotgan mahsulotlar (moderation='owner_approve')."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    user, shop, _ = await _require_owner(update, context)
    if not shop:
        return
    pend = db.get_seller_products_by_status(shop['owner_user_id'], 'pending_owner')
    if not pend:
        await query.edit_message_text(
            t(lang, 'pending_products_empty'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="staff_panel")]]),
            parse_mode='HTML')
        return
    kb = []
    for p in pend[:20]:
        creator = db.get_user_by_id(p.get('created_by')) if p.get('created_by') else None
        who = (creator.get('name') if creator else '') or '—'
        kb.append([InlineKeyboardButton(f"📦 {p['name'][:30]} · {who}", callback_data=f"ownappr_{p['id']}")])
        kb.append([InlineKeyboardButton(t(lang, 'btn_reject') + f" — {p['name'][:20]}", callback_data=f"ownrej_{p['id']}")])
    kb.append([InlineKeyboardButton(t(lang, 'back'), callback_data="staff_panel")])
    await query.edit_message_text(t(lang, 'pending_products_header'),
                                  reply_markup=InlineKeyboardMarkup(kb), parse_mode='HTML')


async def seller_reviews(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Sotuvchi o'z reytinglarini ko'radi."""
    query = update.callback_query
    if query:
        await query.answer()

    lang = get_lang(update, context)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    reviews = db.get_seller_reviews(user['id'])
    avg = db.get_seller_avg_rating(user['id'])

    if not reviews:
        text = t(lang, 'reviews_none')
        kb = InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="seller_panel")]])
        if query:
            await query.edit_message_text(text, reply_markup=kb)
        else:
            await update.message.reply_text(text, reply_markup=kb)
        return

    lines = [t(lang, 'seller_avg_header', avg=f"{avg:.1f}", count=len(reviews))]
    shop_lbl = t(lang, 'review_shop_to').strip()       # "🏪 Do'konga:"
    prod_lbl = t(lang, 'review_product_to').strip()    # "📦 Mahsulotga:"

    reply_btns = []  # izohi bor, lekin javob berilmagan sharhlarga javob tugmalari
    for idx, r in enumerate(reviews[:20], start=1):  # so'nggi 20 ta
        sr = r['rating'] or 0
        s_stars = "⭐" * sr + "☆" * (5 - sr)
        buyer = html.escape(r.get('buyer_name') or t(lang, 'anonymous'))
        comment = html.escape(r.get('comment') or '')
        date = fmt_datetime(r.get('created_at'))
        product = html.escape(r.get('product_name') or t(lang, 'review_product_unknown'))
        reply = html.escape(r.get('seller_reply') or '')

        # Har bir sharh — alohida, tushunarli karta: qaysi mahsulot/buyurtma aniq ko'rinadi
        block = ["➖➖➖➖➖➖➖➖➖➖"]
        head = f"<b>{idx})</b> 📦 <b>{product}</b>"
        if r.get('order_id'):
            head += f"  ·  {fmt_order_id(r['order_id'])}"
        block.append(head)
        block.append(f"{shop_lbl} {s_stars}")
        pr = r.get('product_rating')
        if pr:
            block.append(f"{prod_lbl} {'⭐' * pr}{'☆' * (5 - pr)}")
        block.append(f"👤 {buyer}  ·  {date}")
        block.append(f"💬 {comment}" if comment else f"💬 <i>{t(lang, 'review_no_comment')}</i>")
        if reply:
            block.append(f"   ↳ {t(lang, 'review_shop_reply')} {reply}")
        elif comment:
            # Javob berilmagan, izohli sharh — javob tugmasi taklif qilamiz
            reply_btns.append(InlineKeyboardButton(
                t(lang, 'review_reply_btn', n=idx), callback_data=f"rvreply_{r['id']}"
            ))
        lines.append("\n".join(block))

    text = "\n".join(lines)
    if len(text) > 4000:
        text = text[:3900] + t(lang, 'reviews_old_cut_seller')

    # Tugmalar: javob tugmalarini 2 tadan qatorga joylaymiz + orqaga
    kb_rows = [reply_btns[i:i + 2] for i in range(0, len(reply_btns), 2)]
    kb_rows.append([InlineKeyboardButton(t(lang, 'back'), callback_data="seller_panel")])
    kb = InlineKeyboardMarkup(kb_rows)

    if query:
        await query.edit_message_text(text, reply_markup=kb, parse_mode='HTML')
    else:
        await update.message.reply_text(text, reply_markup=kb, parse_mode='HTML')


async def buyer_reviews(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xaridor o'zi qoldirgan reytinglarni ko'radi."""
    query = update.callback_query
    if query:
        await query.answer()

    user = db.get_user_by_telegram_id(update.effective_user.id)

    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT r.*, u.shop_name, u.name as seller_name, p.name as product_name
        FROM reviews r
        JOIN users u ON r.seller_id=u.id
        LEFT JOIN orders o ON r.order_id=o.id
        LEFT JOIN products p ON o.product_id=p.id
        WHERE r.buyer_id=?
        ORDER BY r.created_at DESC
        LIMIT 20
    """, (user['id'],))
    reviews = [dict(r) for r in cursor.fetchall()]

    lang = get_lang(update, context)
    if not reviews:
        text = t(lang, 'my_reviews_empty')
        kb = InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="buyer_panel")]])
        if query:
            await query.edit_message_text(text, reply_markup=kb)
        else:
            await update.message.reply_text(text, reply_markup=kb)
        return

    lines = [t(lang, 'my_reviews_header', n=len(reviews))]
    for r in reviews:
        sr = r['rating'] or 0
        s_stars = "⭐" * sr + "☆" * (5 - sr)
        shop = html.escape(r.get('shop_name') or r.get('seller_name') or '—')
        product = html.escape(r.get('product_name') or '—')
        comment = html.escape(r.get('comment') or '')
        date = fmt_datetime(r.get('created_at'))
        line = f"\n📦 {product} · {date}"
        pr = r.get('product_rating')
        if pr:
            line += f"\n{t(lang, 'review_to_product')}{'⭐' * pr}{'☆' * (5 - pr)}"
        line += f"\n🏪 {shop} — {s_stars}"
        if comment:
            line += f"\n💬 {comment}"
        reply = html.escape(r.get('seller_reply') or '')
        if reply:
            line += f"\n   ↳ {t(lang, 'review_shop_reply')} {reply}"
        lines.append(line)

    text = "\n".join(lines)
    kb = InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="buyer_panel")]])

    if query:
        await query.edit_message_text(text, reply_markup=kb, parse_mode='HTML')
    else:
        await update.message.reply_text(text, reply_markup=kb, parse_mode='HTML')


async def review_reply_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Sotuvchi sharhga javob yozishni boshlaydi (callback: rvreply_{review_id}).
    Keyingi matn xabari javob sifatida qabul qilinadi (text_handler orqali)."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    chat_id = update.effective_chat.id
    try:
        review_id = int(query.data.split("_")[1])
    except (ValueError, IndexError):
        return

    # XAVFSIZLIK: faqat shu sharhning sotuvchisi javob yoza oladi
    user = db.get_user_by_telegram_id(update.effective_user.id)
    review = db.get_review_by_id(review_id)
    if not review or not user or review.get('seller_id') != user['id']:
        await context.bot.send_message(chat_id, t(lang, 'review_reply_not_yours'))
        return

    context.user_data['awaiting_review_reply'] = review_id
    product = html.escape(review.get('product_name') or t(lang, 'review_product_unknown'))
    comment = html.escape(review.get('comment') or '')
    # AI yoqilgan bo'lsa — javobni AI yozib berishi uchun tugma taklif qilamiz
    kb = None
    if ai_assistant.is_enabled():
        kb = InlineKeyboardMarkup([[
            InlineKeyboardButton(t(lang, 'ai_review_gen_btn'), callback_data=f"airvgen_{review_id}")
        ]])
    await context.bot.send_message(
        chat_id, t(lang, 'review_reply_prompt', product=product, comment=comment),
        parse_mode='HTML', reply_markup=kb
    )


async def _notify_buyer_of_reply(context, review_id):
    """Sotuvchi sharhga javob yozganda — sharh egasi xaridorga xabar yuboradi.
    Xato bo'lsa jim yutiladi (javob baribir saqlangan va ommaviy ko'rinadi)."""
    try:
        review = db.get_review_by_id(review_id)
        if not review or not (review.get('seller_reply') or '').strip():
            return
        buyer = db.get_user_by_id(review.get('buyer_id'))
        if not buyer or not buyer.get('telegram_id'):
            return
        blang = get_user_lang(buyer)
        seller = db.get_user_by_id(review.get('seller_id'))
        shop = html.escape((seller.get('shop_name') or seller.get('name') or '') if seller else '')
        product = html.escape(review.get('product_name') or t(blang, 'review_product_unknown'))
        comment = html.escape(review.get('comment') or '')
        reply = html.escape(review.get('seller_reply') or '')
        await context.bot.send_message(
            chat_id=buyer['telegram_id'],
            text=t(blang, 'buyer_review_reply_notify',
                   shop=shop, product=product, comment=comment, reply=reply),
            parse_mode='HTML'
        )
    except Exception as e:
        logging.warning(f"Sharh javobi bildirishnomasi ketmadi: {e}")


async def review_reply_submit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Sotuvchi yozgan javob matnini qabul qiladi va saqlaydi."""
    lang = get_lang(update, context)
    review_id = context.user_data.get('awaiting_review_reply')
    if not review_id:
        return
    reply_text = (update.message.text or '').strip()
    if len(reply_text) < 2:
        await update.message.reply_text(t(lang, 'review_reply_too_short'))
        return  # bayroq saqlanadi — yana matn kutamiz
    context.user_data.pop('awaiting_review_reply', None)
    if len(reply_text) > 500:
        reply_text = reply_text[:500].rstrip()
    user = db.get_user_by_telegram_id(update.effective_user.id)
    ok = db.set_review_reply(review_id, user['id'], reply_text) if user else False
    if ok:
        await _notify_buyer_of_reply(context, review_id)
        await update.message.reply_text(t(lang, 'review_reply_saved'))
        await seller_reviews(update, context)
    else:
        await update.message.reply_text(t(lang, 'review_reply_not_yours'))


def _pro_gate_kb(lang):
    """Pro qulfi xabari klaviaturasi — Mini App'ni ochish (Pro o'sha yerda olinadi)."""
    rows = []
    if MINIAPP_URL:
        rows.append([_open_app_button(lang, 'pro_open_app', MINIAPP_URL)])
    rows.append([InlineKeyboardButton(t(lang, 'back'), callback_data="seller_panel")])
    return InlineKeyboardMarkup(rows)


async def seller_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Sotuvchi statistikasi: buyurtmalar, daromad, mahsulotlar — hafta/oy/jami."""
    query = update.callback_query
    await query.answer()

    lang = get_lang(update, context)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    stats = db.get_seller_stats(user['id'])
    avg_rating = db.get_seller_avg_rating(user['id'])

    text = t(lang, 'seller_stats_body',
             shop=html.escape(user.get('shop_name') or ''),
             products=stats['products_count'], rating=f"{avg_rating:.1f}",
             week_orders=stats['week_orders'], week_revenue=fmt_price(stats['week_revenue']),
             month_orders=stats['month_orders'], month_revenue=fmt_price(stats['month_revenue']),
             total_orders=stats['total_orders'], pending=stats['pending'],
             confirmed=stats['confirmed'], delivered=stats['delivered'],
             cancelled=stats['cancelled'], total_revenue=fmt_price(stats['total_revenue']))

    await query.edit_message_text(
        text,
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton(t(lang, 'btn_detailed_excel'), callback_data="seller_export_excel")],
            [InlineKeyboardButton(t(lang, 'back'), callback_data="seller_panel")],
        ]),
        parse_mode='HTML'
    )


async def seller_export_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Sotuvchining barcha ma'lumotlarini kengaytirilgan Excel fayl sifatida yuboradi:
    buyurtmalar (kim, qachon, nima olgani), mahsulotlar, reytinglar va umumiy hisobot."""
    query = update.callback_query
    lang = get_lang(update, context)
    ru = (lang == 'ru')
    await query.answer(t(lang, 'excel_preparing'))

    user = db.get_user_by_telegram_id(update.effective_user.id)
    if not user:
        await query.message.reply_text(t(lang, 'user_not_found'))
        return
    # #18 Pro — Excel yuklab olish Pro imkoniyat (app bilan bir xil qoida; bot bypass YO'Q)
    if db.pro_locked(db.resolve_owner_id(user['id'])):
        await query.message.reply_text(t(lang, 'pro_locked_bot'), parse_mode='HTML',
                                       reply_markup=_pro_gate_kb(lang))
        return

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import datetime as dt
        import os

        seller_id = user['id']
        orders = db.get_seller_orders_list(seller_id)
        products = db.get_products_by_seller(seller_id)
        reviews = db.get_seller_reviews(seller_id)
        stats = db.get_seller_stats(seller_id)
        avg_rating = db.get_seller_avg_rating(seller_id)

        status_label = {
            'pending': 'Новый' if ru else 'Yangi',
            'confirmed': 'Подтверждён' if ru else 'Tasdiqlangan',
            'delivered': 'Доставлен' if ru else 'Yetkazilgan',
            'cancelled': 'Отменён' if ru else 'Bekor qilingan',
        }
        delivery_label = {
            'delivery': 'Доставка' if ru else 'Yetkazib berish',
            'pickup': 'Самовывоз' if ru else 'Olib ketish',
        }
        payment_label = {
            'cash': 'Наличные' if ru else 'Naqd pul',
            'card': 'Терминал' if ru else 'Terminal',
            'terminal': 'Терминал' if ru else 'Terminal',
            'p2p': 'Карта (P2P)' if ru else 'Karta (P2P)',
            'click': 'Click',
        }
        product_status_label = {
            'active': 'В продаже' if ru else 'Sotuvda',
            'reserve': 'В резерве' if ru else 'Zahirada',
            'deleted': 'Удалён' if ru else "O'chirilgan",
        }
        settlement_label = {
            'paid': 'Оплачено' if ru else "To'langan",
            'debt': 'Долг' if ru else 'Qarz',
            'installment': 'Рассрочка' if ru else "Bo'lib to'lash",
        }

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1a8a2e")
        header_align = Alignment(horizontal="center", vertical="center")

        def style_header(ws_):
            for cell in ws_[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

        def auto_width(ws_):
            for col in ws_.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=0)
                ws_.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)

        wb = openpyxl.Workbook()

        # ---- 1) Buyurtmalar ----
        ws = wb.active
        ws.title = "Заказы" if ru else "Buyurtmalar"
        ws.append(
            ["ID заказа", "Дата", "Покупатель", "Телефон покупателя",
             "Товар", "Цена за шт", "Кол-во", "Сумма",
             "Статус", "Тип доставки", "Способ оплаты", "Адрес",
             "Статус оплаты", "Оплачено", "Остаток долга"]
            if ru else
            ["Buyurtma ID", "Sana", "Xaridor", "Xaridor telefoni",
             "Mahsulot", "Dona narxi", "Miqdor", "Jami summa",
             "Holat", "Yetkazish turi", "To'lov usuli", "Manzil",
             "To'lov holati", "To'langan", "Qolgan qarz"]
        )
        style_header(ws)
        for o in orders:
            stt = o.get('settlement_type') or ''
            ws.append([
                o.get('id'),
                fmt_datetime(o.get('created_at')),
                o.get('buyer_name') or '',
                o.get('buyer_phone') or '',
                o.get('product_name') or '',
                o.get('product_price') or 0,
                o.get('quantity') or 0,
                o.get('total_price') or 0,
                status_label.get(o.get('status') or '', o.get('status') or ''),
                delivery_label.get(o.get('delivery_type') or '', o.get('delivery_type') or ''),
                payment_label.get(o.get('payment_method') or '', o.get('payment_method') or ''),
                o.get('delivery_address') or '',
                settlement_label.get(stt, '—' if not stt else stt),
                o.get('amount_paid') if o.get('amount_paid') is not None else '',
                o.get('amount_due') if o.get('amount_due') is not None else '',
            ])
        auto_width(ws)

        # ---- 2) Mahsulotlar (sotilgan soni va daromad bilan) ----
        # Yetkazilgan buyurtmalardan har bir mahsulot bo'yicha jami hisoblaymiz
        sold_count = {}
        sold_revenue = {}
        for o in orders:
            if (o.get('status') or '') == 'delivered':
                pid = o.get('product_id')
                sold_count[pid] = sold_count.get(pid, 0) + (o.get('quantity') or 0)
                sold_revenue[pid] = sold_revenue.get(pid, 0) + (o.get('total_price') or 0)

        ws2 = wb.create_sheet("Товары" if ru else "Mahsulotlar")
        ws2.append(
            ["ID", "Название", "Категория", "Цена", "Статус", "Остаток",
             "Продано (шт)", "Доход (сум)", "Дата добавления"]
            if ru else
            ["ID", "Nom", "Kategoriya", "Narx", "Holat", "Zahira",
             "Sotilgan (dona)", "Daromad (so'm)", "Qo'shilgan sana"]
        )
        style_header(ws2)
        for p in products:
            st = p.get('status') or ('active' if p.get('in_stock') else 'reserve')
            ws2.append([
                p.get('id'),
                p.get('name') or '',
                p.get('category_name') or '',
                p.get('price') or 0,
                product_status_label.get(st, st),
                p.get('stock_count') if p.get('stock_count') is not None else ('Без лимита' if ru else 'Cheksiz'),
                sold_count.get(p.get('id'), 0),
                sold_revenue.get(p.get('id'), 0),
                fmt_datetime(p.get('created_at')),
            ])
        auto_width(ws2)

        # ---- 3) Reytinglar ----
        ws3 = wb.create_sheet("Отзывы" if ru else "Reytinglar")
        ws3.append(["Дата", "Покупатель", "Оценка", "Комментарий"] if ru
                   else ["Sana", "Xaridor", "Baho", "Izoh"])
        style_header(ws3)
        for r in reviews:
            ws3.append([
                fmt_datetime(r.get('created_at')),
                r.get('buyer_name') or ('Аноним' if ru else 'Anonim'),
                r.get('rating') or 0,
                r.get('comment') or '',
            ])
        auto_width(ws3)

        # ---- 4) Umumiy hisobot ----
        ws4 = wb.create_sheet("Сводка" if ru else "Umumiy")
        ws4.append(["Показатель", "Значение"] if ru else ["Ko'rsatkich", "Qiymat"])
        style_header(ws4)
        # To'lov holati bo'yicha jami: qolgan qarz va to'langan summa
        total_debt = sum((o.get('amount_due') or 0) for o in orders)
        total_paid = sum((o.get('amount_paid') or 0) for o in orders)
        if ru:
            summary = [
                ("Название магазина", user.get('shop_name') or '—'),
                ("Кол-во товаров", stats.get('products_count', 0)),
                ("Средний рейтинг", round(avg_rating, 2)),
                ("Кол-во отзывов", len(reviews)),
                ("Всего заказов", stats.get('total_orders', 0)),
                ("Новые (в ожидании)", stats.get('pending', 0)),
                ("Подтверждённые", stats.get('confirmed', 0)),
                ("Доставленные", stats.get('delivered', 0)),
                ("Отменённые", stats.get('cancelled', 0)),
                ("Последние 7 дней — заказы", stats.get('week_orders', 0)),
                ("Последние 7 дней — доход", stats.get('week_revenue', 0)),
                ("Последние 30 дней — заказы", stats.get('month_orders', 0)),
                ("Последние 30 дней — доход", stats.get('month_revenue', 0)),
                ("Общий доход (доставленные)", stats.get('total_revenue', 0)),
                ("Всего оплачено (при выдаче)", total_paid),
                ("Остаток долга (всего)", total_debt),
            ]
        else:
            summary = [
                ("Do'kon nomi", user.get('shop_name') or '—'),
                ("Mahsulotlar soni", stats.get('products_count', 0)),
                ("O'rtacha reyting", round(avg_rating, 2)),
                ("Reytinglar soni", len(reviews)),
                ("Jami buyurtmalar", stats.get('total_orders', 0)),
                ("Yangi (kutilmoqda)", stats.get('pending', 0)),
                ("Tasdiqlangan", stats.get('confirmed', 0)),
                ("Yetkazilgan", stats.get('delivered', 0)),
                ("Bekor qilingan", stats.get('cancelled', 0)),
                ("So'nggi 7 kun — buyurtma", stats.get('week_orders', 0)),
                ("So'nggi 7 kun — daromad", stats.get('week_revenue', 0)),
                ("So'nggi 30 kun — buyurtma", stats.get('month_orders', 0)),
                ("So'nggi 30 kun — daromad", stats.get('month_revenue', 0)),
                ("Jami daromad (yetkazilgan)", stats.get('total_revenue', 0)),
                ("Jami to'langan (berishda)", total_paid),
                ("Qolgan qarz (jami)", total_debt),
            ]
        for k, v in summary:
            ws4.append([k, v])
        auto_width(ws4)

        ts = dt.datetime.now().strftime("%d.%m.%Y %H:%M")
        filename = f"tezbozor_hisobot_{dt.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(filename)

        shop_caption = user.get('shop_name') or t(lang, 'excel_shop_default')
        await context.bot.send_document(
            chat_id=update.effective_chat.id,
            document=open(filename, 'rb'),
            filename=filename,
            caption=t(lang, 'excel_caption', shop=shop_caption,
                      orders=len(orders), products=len(products),
                      reviews=len(reviews), ts=ts)
        )
        try:
            os.remove(filename)
        except Exception:
            pass

    except ImportError:
        await query.message.reply_text(
            t(lang, 'excel_not_installed'),
            parse_mode='HTML'
        )
    except Exception as e:
        logging.error(f"Sotuvchi Excel eksport xatosi: {e}")
        await query.message.reply_text(t(lang, 'error_generic', e=e))


# ============================================================
# MAHSULOT QO'SHISH — bosqichlar va navigatsiya (Orqaga / O'tkazib yuborish)
# ============================================================
def _add_nav_kb(lang, *, back=None, skip=None, extra=None):
    """Bosqich tugmalari + navigatsiya qatori. back/skip — maqsad bosqich nomi."""
    rows = list(extra or [])
    nav = []
    if back:
        nav.append(InlineKeyboardButton(t(lang, 'btn_back_step'), callback_data=f"addnav_{back}"))
    if skip:
        nav.append(InlineKeyboardButton(t(lang, 'btn_skip_step'), callback_data=f"addnav_{skip}"))
    if nav:
        rows.append(nav)
    return InlineKeyboardMarkup(rows) if rows else None


async def _ask_price(update, context):
    lang = get_lang(update, context)
    await context.bot.send_message(
        chat_id=update.effective_chat.id, text=t(lang, 'add_product_price_ask'),
        reply_markup=_add_nav_kb(lang, back='name'))
    return PRODUCT_PRICE


async def _ask_stock(update, context):
    lang = get_lang(update, context)
    extra = [
        [InlineKeyboardButton(t(lang, 'btn_stock_unlimited'), callback_data="apstock_unlim")],
        [InlineKeyboardButton(t(lang, 'btn_stock_limited'), callback_data="apstock_num")],
    ]
    await context.bot.send_message(
        chat_id=update.effective_chat.id, text=t(lang, 'add_product_stock_ask'),
        reply_markup=_add_nav_kb(lang, back='price', extra=extra))
    return PRODUCT_STOCK


async def _ask_category(update, context):
    lang = get_lang(update, context)
    categories = db.get_all_categories()
    extra = [[InlineKeyboardButton(f"{cat[2]} {category_name(cat[1], lang)}",
                                   callback_data=f"prodcat_{cat[0]}")] for cat in categories]
    await context.bot.send_message(
        chat_id=update.effective_chat.id, text=t(lang, 'choose_category'),
        reply_markup=_add_nav_kb(lang, back='stock', extra=extra))
    return PRODUCT_CATEGORY


async def _ask_desc(update, context):
    lang = get_lang(update, context)
    await context.bot.send_message(
        chat_id=update.effective_chat.id, text=t(lang, 'add_product_desc_ask'),
        reply_markup=_add_nav_kb(lang, back='category', skip='photo'))
    return PRODUCT_DESC


async def _ask_photo(update, context):
    lang = get_lang(update, context)
    context.user_data.setdefault('product_photos', [])
    await context.bot.send_message(
        chat_id=update.effective_chat.id, text=t(lang, 'add_photo_ask'),
        reply_markup=_add_nav_kb(lang, back='desc', skip='attrs'))
    return PRODUCT_PHOTO


async def add_product_nav(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Orqaga / O'tkazib yuborish tugmalari — maqsad bosqichga o'tadi.
    callback: addnav_<bosqich>"""
    query = update.callback_query
    await query.answer()
    target = query.data.split("_", 1)[1]
    # Eski xabardagi tugmalarni olib tashlaymiz (chalkashmaslik uchun)
    try:
        await query.edit_message_reply_markup(reply_markup=None)
    except Exception:
        pass
    if target == 'price':
        return await _ask_price(update, context)
    if target == 'stock':
        return await _ask_stock(update, context)
    if target == 'category':
        return await _ask_category(update, context)
    if target == 'desc':
        return await _ask_desc(update, context)
    if target == 'photo':
        return await _ask_photo(update, context)
    if target == 'attrs':
        # Rasmlarni o'tkazib yuborish — atributlar/saqlashga o'tamiz
        return await _proceed_after_photos(update, context)
    if target == 'name':
        lang = get_lang(update, context)
        await context.bot.send_message(chat_id=update.effective_chat.id,
                                       text=t(lang, 'add_product_name_ask'))
        return PRODUCT_NAME
    return None


async def seller_add_product_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if query:
        await query.answer()

    # Callback (inline) yoki Reply tugma (matn) — ikkalasida ham ishlasin
    async def _show(text, **kw):
        if query:
            await query.edit_message_text(text, **kw)
        else:
            await update.message.reply_text(text, **kw)

    # Tasdiqlanmagan sotuvchi mahsulot qo'sha olmaydi
    user = db.get_user_by_telegram_id(update.effective_user.id)
    lang = get_lang(update, context)
    if user and not user.get('is_approved') and user.get('role') != 'admin':
        await _show(
            t(lang, 'seller_not_approved'),
            parse_mode='HTML',
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(t(lang, 'btn_contact_admin'), callback_data="contact_admin")],
                [InlineKeyboardButton(t(lang, 'back'), callback_data="seller_panel")],
            ])
        )
        return ConversationHandler.END

    # #18 Pro — bepul mahsulot limiti (app create_product gate'i bilan bir xil; Pro = cheksiz)
    owner_id = db.resolve_owner_id(user['id']) if user else None
    flim = db.mon_limit('mon_free_product_limit')
    if owner_id and flim > 0 and not db.is_pro(owner_id) and db.count_active_products(owner_id) >= flim:
        await _show(t(lang, 'pro_locked_limit_bot'), parse_mode='HTML', reply_markup=_pro_gate_kb(lang))
        return ConversationHandler.END

    context.user_data['adding_product'] = True
    # Eski qiymatlarni tozalaymiz (oldingi yarim qolgan jarayonni)
    for k in ('product_name', 'product_price', 'product_stock', 'product_category',
              'product_desc', 'product_photo', 'product_photos', 'question_mode',
              'attr_templates', 'attr_index', 'product_attrs'):
        context.user_data.pop(k, None)
    context.user_data['product_photos'] = []   # 5 tagacha rasm shu yerda yig'iladi

    # Joylash usulini tanlash. AI o'chiq bo'lsa — to'g'ridan-to'g'ri klassikka o'tamiz
    # (sotuvchini ishlamaydigan tugmalar bilan chalkashtirmaymiz).
    if not ai_assistant.is_enabled():
        context.user_data['question_mode'] = 'classic'
        await _show(t(lang, 'add_product_name_ask'))
        return PRODUCT_NAME

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(t(lang, 'btn_mode_classic'), callback_data="pmode_classic")],
        [InlineKeyboardButton(t(lang, 'btn_mode_ai_guided'), callback_data="pmode_ai_guided")],
        [InlineKeyboardButton(t(lang, 'btn_mode_ai_smart'), callback_data="pmode_ai_smart")],
    ])
    await _show(t(lang, 'choose_post_mode'), parse_mode='HTML', reply_markup=kb)
    return PRODUCT_MODE


async def seller_add_product_mode(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Joylash usulini tanlash: pmode_classic | pmode_ai_guided | pmode_ai_smart.
    Tanlovdan keyin oddiy nom bosqichiga o'tamiz — qolgan oqim barcha rejimlar uchun bir xil,
    farq faqat rasm bosqichidan keyingi atribut bosqichida (_proceed_after_photos)."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    mode = query.data.split("_", 1)[1]  # 'classic' | 'ai_guided' | 'ai_smart'
    if mode not in ('classic', 'ai_guided', 'ai_smart'):
        mode = 'classic'
    context.user_data['question_mode'] = mode
    try:
        await query.edit_message_reply_markup(reply_markup=None)
    except Exception:
        pass
    await context.bot.send_message(chat_id=update.effective_chat.id,
                                   text=t(lang, 'add_product_name_ask'))
    return PRODUCT_NAME


async def seller_add_product_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    lang = get_lang(update, context)
    # Validatsiya: uzunlik
    if len(name) < 3:
        await update.message.reply_text(t(lang, 'name_too_short'))
        return PRODUCT_NAME
    if len(name) > 100:
        await update.message.reply_text(t(lang, 'name_too_long'))
        return PRODUCT_NAME

    context.user_data['product_name'] = name
    return await _ask_price(update, context)


async def seller_add_product_price(update: Update, context: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.replace(" ", "").replace(",", "").replace("\u00A0", "")
    lang = get_lang(update, context)
    try:
        price = float(raw)
    except ValueError:
        await update.message.reply_text(t(lang, 'price_invalid'))
        return PRODUCT_PRICE

    # Validatsiya: musbat va mantiqiy chegara
    if price <= 0:
        await update.message.reply_text(t(lang, 'price_positive'))
        return PRODUCT_PRICE
    if price > 1_000_000_000:
        await update.message.reply_text(t(lang, 'price_too_big'))
        return PRODUCT_PRICE

    context.user_data['product_price'] = price
    return await _ask_stock(update, context)


async def seller_add_product_stock_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mahsulot qo'shishda zaxira tugmasi: '♾ Cheksiz' yoki '🔢 Aniq miqdor'."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    if query.data == "apstock_unlim":
        context.user_data['product_stock'] = None
        try:
            await query.edit_message_text(t(lang, 'stock_set_unlimited'))
        except Exception:
            pass
        return await _ask_category(update, context)
    # apstock_num — aniq son so'raymiz (orqaga qaytish tugmasi bilan)
    try:
        await query.edit_message_text(
            t(lang, 'add_product_stock_enter'),
            reply_markup=_add_nav_kb(lang, back='price'))
    except Exception:
        pass
    return PRODUCT_STOCK


async def seller_add_product_stock_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """'🔢 Aniq miqdor' tanlangach — sotuvga qo'yiladigan sonni qabul qiladi."""
    lang = get_lang(update, context)
    raw = (update.message.text or "").strip()
    try:
        n = int(raw)
        if n <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text(t(lang, 'stock_enter_invalid'))
        return PRODUCT_STOCK
    context.user_data['product_stock'] = n
    return await _ask_category(update, context)


async def seller_add_product_category(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    category_id = int(query.data.split("_")[1])
    context.user_data['product_category'] = category_id

    # Eski kategoriya tugmalarini olib tashlaymiz, so'ng tavsif bosqichiga o'tamiz
    try:
        await query.edit_message_reply_markup(reply_markup=None)
    except Exception:
        pass
    return await _ask_desc(update, context)


async def seller_add_product_desc(update: Update, context: ContextTypes.DEFAULT_TYPE):
    lang = get_lang(update, context)
    desc = update.message.text.strip()
    if desc == "-":
        desc = None
    elif len(desc) > 500:
        await update.message.reply_text(t(lang, 'desc_too_long'))
        return PRODUCT_DESC

    context.user_data['product_desc'] = desc
    return await _ask_photo(update, context)


async def seller_add_product_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    photo_file_id = None
    lang = get_lang(update, context)

    if update.message.photo:
        # Eng katta o'lchamdagi rasmni olamiz
        photo = update.message.photo[-1]

        # Hajm tekshiruvi: 5 MB dan katta bo'lmasin
        if photo.file_size and photo.file_size > 5 * 1024 * 1024:
            await update.message.reply_text(t(lang, 'photo_too_big'))
            return PRODUCT_PHOTO

        # O'lcham tekshiruvi: kamida 200x200 piksel
        if photo.width and photo.height:
            if photo.width < 200 or photo.height < 200:
                await update.message.reply_text(t(lang, 'photo_too_small'))
                return PRODUCT_PHOTO

        photo_file_id = photo.file_id

    elif update.message.document:
        # Document sifatida yuborilgan rasmni tekshiramiz
        doc = update.message.document
        mime = doc.mime_type or ''

        if not mime.startswith('image/'):
            await update.message.reply_text(t(lang, 'only_images'))
            return PRODUCT_PHOTO

        # Hajm tekshiruvi
        if doc.file_size and doc.file_size > 5 * 1024 * 1024:
            await update.message.reply_text(t(lang, 'photo_too_big_doc'))
            return PRODUCT_PHOTO

        photo_file_id = doc.file_id

    elif update.message.sticker:
        await update.message.reply_text(t(lang, 'no_sticker'))
        return PRODUCT_PHOTO

    elif update.message.text and update.message.text.strip() == "-":
        # '-' = tugatish: yangi rasm qo'shmaymiz (rasmsiz yoki yetarli) — keyingi bosqich
        return await _proceed_after_photos(update, context)
    else:
        await update.message.reply_text(t(lang, 'send_photo_or_skip'))
        return PRODUCT_PHOTO

    # Bu yergacha yetib keldik — demak yangi rasm bor
    photos = context.user_data.setdefault('product_photos', [])
    photos.append(photo_file_id)
    n = len(photos)

    if n >= db.MAX_PRODUCT_IMAGES:
        await update.message.reply_text(t(lang, 'photos_max_added', n=n))
        return await _proceed_after_photos(update, context)

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(t(lang, 'btn_add_more_photo'), callback_data="addphoto_more")],
        [InlineKeyboardButton(t(lang, 'btn_continue'), callback_data="addphoto_done")],
    ])
    await update.message.reply_text(
        t(lang, 'photos_added_n', n=n, max=db.MAX_PRODUCT_IMAGES),
        reply_markup=kb
    )
    return PRODUCT_PHOTO


def _classic_templates(context):
    """Klassik (kategoriyaga bog'langan) statik shablonlar."""
    category_id = context.user_data.get('product_category')
    return db.get_category_templates(category_id) if category_id else []


async def _start_attr_flow(update, context, templates, prefilled=None, prefilled_labels=None):
    """Atribut bosqichini berilgan shablonlar bilan boshlaydi (yoki saqlaydi).
    prefilled — AI aqlli rejimda tavsifdan oldindan to'ldirilgan {key: value}.
    Yorliqlarni (key→label) saqlaymiz — AI savollari uchun shablon yo'q, shu sababli
    saqlanmasa mahsulot kartasida xom kalit ('car_model') ko'rinib qoladi."""
    context.user_data['product_attrs'] = dict(prefilled or {})
    labels = {}
    for tmpl in (templates or []):
        if tmpl.get('attr_key') and tmpl.get('attr_label'):
            labels[tmpl['attr_key']] = tmpl['attr_label']
    for k, lbl in (prefilled_labels or {}).items():
        if lbl:
            labels[k] = lbl
    context.user_data['attr_labels'] = labels
    if templates:
        context.user_data['attr_templates'] = templates
        context.user_data['attr_index'] = 0
        return await _ask_next_attr(update, context)
    # Savol yo'q — to'g'ridan-to'g'ri saqlaymiz
    return await _save_product(update, context)


async def _proceed_after_photos(update, context):
    """Rasm bosqichidan keyin atribut bosqichi. Tanlangan rejimga qarab:
      • classic   — kategoriyaning statik savollari (eski xulq);
      • ai_guided — AI mahsulotga mos savollar tuzadi;
      • ai_smart  — AI tavsifdan ajratadi + qolgan savollarni so'raydi.
    AI o'chiq/xato bo'lsa — har doim klassik shablonlarga qaytadi."""
    mode = context.user_data.get('question_mode', 'classic')
    lang = get_lang(update, context)
    chat_id = update.effective_chat.id

    if mode == 'classic' or not ai_assistant.is_enabled():
        return await _start_attr_flow(update, context, _classic_templates(context))

    # --- AI rejimlari ---
    try:
        await context.bot.send_message(chat_id=chat_id, text=t(lang, 'ai_questions_thinking'))
    except Exception:
        pass

    cat_id = context.user_data.get('product_category')
    cat_name = ''
    if cat_id:
        try:
            for c in db.get_all_categories():
                if c[0] == cat_id:
                    cat_name = category_name(c[1], lang)
                    break
        except Exception:
            cat_name = ''

    result = None
    try:
        result = await ai_assistant.generate_product_questions(
            name=context.user_data.get('product_name', ''),
            category=cat_name,
            description=context.user_data.get('product_desc') or '',
            lang=lang,
            smart=(mode == 'ai_smart'),
        )
    except Exception as e:
        logging.warning(f"AI savollar olinmadi: {e}")
        result = None

    if not result:
        # AI ishlamadi — klassikka qaytamiz, sotuvchi savolsiz qolmaydi
        await context.bot.send_message(chat_id=chat_id, text=t(lang, 'ai_questions_failed'))
        return await _start_attr_flow(update, context, _classic_templates(context))

    known = result.get('known') or {}
    templates = result.get('questions') or []
    prefilled = {k: v['value'] for k, v in known.items()}
    prefilled_labels = {k: v['label'] for k, v in known.items()}

    # AI aqlli: tavsifdan aniqlanganlarni sotuvchiga ko'rsatamiz (shaffoflik uchun)
    if mode == 'ai_smart' and known:
        lines = "\n".join(f"• {v['label']}: {v['value']}" for v in known.values())
        await context.bot.send_message(chat_id=chat_id,
                                       text=t(lang, 'ai_smart_prefilled', lines=lines))
        if not templates:
            await context.bot.send_message(chat_id=chat_id, text=t(lang, 'ai_smart_no_questions'))

    return await _start_attr_flow(update, context, templates,
                                  prefilled=prefilled, prefilled_labels=prefilled_labels)


async def add_photo_more(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """'Yana rasm qo'shish' tugmasi — keyingi rasmni so'raydi."""
    query = update.callback_query
    await query.answer()
    n = len(context.user_data.get('product_photos', []))
    await query.edit_message_text(T(update, context, 'next_photo_ask', n=n, max=db.MAX_PRODUCT_IMAGES))
    return PRODUCT_PHOTO


async def add_photo_done(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """'Davom etish' tugmasi — keyingi bosqichga o'tadi."""
    query = update.callback_query
    await query.answer()
    return await _proceed_after_photos(update, context)


def _attr_nav_row(lang, is_required):
    """Atribut bosqichi uchun navigatsiya qatori:
      [⬅️ Orqaga]  — doim (oldingi atributga yoki rasm bosqichiga qaytadi);
      [⏭ O'tkazish] — faqat ixtiyoriy atributda (yozmasdan keyingisiga o'tish)."""
    row = [InlineKeyboardButton(t(lang, 'btn_back_step'), callback_data="attrnav_back")]
    if not is_required:
        row.append(InlineKeyboardButton(t(lang, 'btn_skip_step'), callback_data="attrnav_skip"))
    return row


async def _ask_next_attr(update, context):
    """Navbatdagi atributni so'raydi."""
    templates = context.user_data.get('attr_templates', [])
    idx = context.user_data.get('attr_index', 0)

    if idx >= len(templates):
        return await _save_product(update, context)

    lang = get_lang(update, context)
    tmpl = templates[idx]
    required_mark = " *" if tmpl['is_required'] else t(lang, 'attr_optional_mark')
    hint = t(lang, 'attr_eg', hint=tmpl['hint']) if tmpl.get('hint') else ""
    skip_note = "" if tmpl['is_required'] else t(lang, 'attr_skip_note')

    # Har bir atribut bosqichida navigatsiya tugmalari (barcha kategoriyalar uchun bir xil):
    # sotuvchi xato kiritsa Orqaga qaytib tuzata oladi, ixtiyoriy maydonni esa yozmasdan
    # O'tkazish tugmasi bilan o'tkazib yuboradi.
    nav_row = _attr_nav_row(lang, tmpl['is_required'])

    if tmpl['attr_type'] == 'select' and tmpl.get('hint'):
        # Tanlov variantlarini tugma sifatida ko'rsatamiz
        options = [o.strip() for o in tmpl['hint'].split('/')]
        kb = [[InlineKeyboardButton(opt, callback_data=f"attr_{opt}")] for opt in options]
        kb.append(nav_row)
        msg = f"📝 {tmpl['attr_label']}{required_mark}{hint}"
        if update.message:
            await update.message.reply_text(msg, reply_markup=InlineKeyboardMarkup(kb))
        else:
            await update.callback_query.edit_message_text(msg, reply_markup=InlineKeyboardMarkup(kb))
    else:
        msg = f"📝 {tmpl['attr_label']}{required_mark}{hint}{skip_note}"
        kb = InlineKeyboardMarkup([nav_row])
        if update.message:
            await update.message.reply_text(msg, reply_markup=kb)
        else:
            await update.callback_query.message.reply_text(msg, reply_markup=kb)

    return PRODUCT_ATTRS


async def seller_add_product_attr_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Matn atribut qabul qilish."""
    templates = context.user_data.get('attr_templates', [])
    idx = context.user_data.get('attr_index', 0)
    tmpl = templates[idx]

    value = update.message.text.strip()
    if value == '-':
        if tmpl['is_required']:
            await update.message.reply_text(T(update, context, 'attr_required_field'))
            return PRODUCT_ATTRS
        value = None

    if value:
        context.user_data['product_attrs'][tmpl['attr_key']] = value

    context.user_data['attr_index'] = idx + 1
    return await _ask_next_attr(update, context)


async def seller_add_product_attr_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Tugma orqali atribut tanlash."""
    query = update.callback_query
    await query.answer()

    templates = context.user_data.get('attr_templates', [])
    idx = context.user_data.get('attr_index', 0)
    tmpl = templates[idx]

    value = query.data.replace("attr_", "")
    if value != '-':
        context.user_data['product_attrs'][tmpl['attr_key']] = value

    context.user_data['attr_index'] = idx + 1
    return await _ask_next_attr(update, context)


async def seller_add_product_attr_nav(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Atribut bosqichidagi navigatsiya: ⬅️ Orqaga / ⏭ O'tkazish.
    callback: attrnav_back | attrnav_skip"""
    query = update.callback_query
    await query.answer()
    action = query.data.split("_", 1)[1]  # 'back' yoki 'skip'
    templates = context.user_data.get('attr_templates', [])
    idx = context.user_data.get('attr_index', 0)

    # Joriy xabardagi tugmalarni olib tashlaymiz (chalkashmaslik uchun)
    try:
        await query.edit_message_reply_markup(reply_markup=None)
    except Exception:
        pass

    if action == 'skip':
        # Ixtiyoriy atributni yozmasdan o'tkazamiz (qiymat saqlanmaydi)
        context.user_data['attr_index'] = idx + 1
        return await _ask_next_attr(update, context)

    # action == 'back'
    if idx <= 0:
        # Birinchi atribut — orqaga rasm bosqichiga qaytamiz
        return await _ask_photo(update, context)

    # Oldingi atributga qaytamiz; uning saqlangan qiymatini tozalaymiz (qayta kiritilsin)
    prev_idx = idx - 1
    context.user_data['attr_index'] = prev_idx
    if prev_idx < len(templates):
        prev_key = templates[prev_idx].get('attr_key')
        if prev_key:
            context.user_data.get('product_attrs', {}).pop(prev_key, None)
    return await _ask_next_attr(update, context)


async def _build_ad_caption(product, length="long"):
    """Mahsulot uchun reklama matnini (caption) qaytaradi: (matn, parse_mode).
    Avval AI takrorlanmas reklama yozishga urinadi; bo'lmasa — tuzilgan HTML matn.
    length — 'long' (uzun) yoki 'short' (qisqa) — sotuvchi preview'da tanlaydi."""
    cat = product.get('category_name') or product.get('category')
    cat_emoji = product.get('category_emoji') or '📂'
    cat_line = f"\n{cat_emoji} {html.escape(str(cat))}" if cat else ""
    shop_name = product.get('shop_name')
    shop_line = f"\n🏪 {html.escape(str(shop_name))}" if shop_name else ""
    region_lbl = region_label_l(product.get('seller_region_id'), DEFAULT_LANG)
    region_line = f"\n🌍 {html.escape(region_lbl)}" if region_lbl else ""
    loc = best_location_text(product.get('shop_address'), product.get('shop_landmark'))
    loc_line = f"\n📍 {html.escape(loc)}" if loc else ""
    prod_rating = product.get('prod_avg_rating') or 0
    prod_cnt = product.get('prod_review_count') or 0
    rating_line = f"\n⭐ {prod_rating:.1f} ({prod_cnt})" if prod_cnt else ""
    desc = (product.get('description') or "").strip()
    if len(desc) > 300:
        desc = desc[:300].rstrip() + "…"
    desc_line = f"\n\n📝 {html.escape(desc)}" if desc else ""

    # Optom (ulgurji/pachka) taklifi — yoqilgan bo'lsa reklamada zinalarni ko'rsatamiz.
    _is_optom = product.get('sale_mode') == 'optom'
    _w = wholesale_info(product)
    _unit = 'pachka' if _is_optom else (product.get('unit') or 'dona')
    _u_esc = html.escape(str(_unit))
    if _w['enabled']:
        _tier_txt = ", ".join(f"{t['min']}+ {_u_esc} — {html.escape(fmt_price(t['price']))}"
                              for t in _w['tiers'])
        wholesale_line = f"\n📦 Optom narx: {_tier_txt}"
    else:
        wholesale_line = ""

    # Optom (pachka) maxsus qatorlari: pachka belgisi, 1 pachka = N dona, ranglar, razmer.
    optom_lines, extra_facts = "", {}
    if _is_optom:
        _ps = product.get('pack_size')
        try:
            _colors = [a['attr_value'] for a in (db.get_product_attributes(product['id']) or [])
                       if a['attr_key'] == 'color' and (a['attr_value'] or '').strip()]
            _colors = _colors[0] if _colors else ""
        except Exception:
            _colors = ""
        _sn = (product.get('size_note') or "").strip()
        _pack_line = f"\n📦 1 pachka = {int(_ps)} dona" if _ps else ""
        _colors_line = f"\n🎨 Ranglar: {html.escape(_colors)}" if _colors else ""
        _size_line = f"\n📏 Razmer: {html.escape(_sn)}" if _sn else ""
        optom_lines = f"\n🏷 Optom (pachkada sotiladi){_pack_line}{_colors_line}{_size_line}"
        extra_facts = {"savdo turi": "optom (ulgurji, pachkada sotiladi)"}
        if _ps:
            extra_facts["bitta pachka"] = f"{int(_ps)} dona"
        if _colors:
            extra_facts["mavjud ranglar"] = _colors
        if _sn:
            extra_facts["razmerlar"] = _sn

    caption = (
        f"🆕 <b>{html.escape(product.get('name') or '')}</b>"
        f"\n💵 {price_with_unit(product)}{wholesale_line}{optom_lines}"
        f"{cat_line}{shop_line}{region_line}{loc_line}{rating_line}{desc_line}"
    )
    parse_mode = 'HTML'

    # AI takrorlanmas reklama matni (faktlardan kelib chiqib)
    try:
        ad_text = await ai_assistant.generate_ad_caption(
            name=product.get('name') or '',
            price_text=price_with_unit(product),
            category=str(cat) if cat else '',
            description=(product.get('description') or ''),
            shop=str(shop_name) if shop_name else '',
            region=region_lbl or '',
            location=loc or '',
            lang=DEFAULT_LANG,
            length=length,
            extra=extra_facts,
        )
    except Exception as e:
        logging.warning(f"Reklama matni olinmadi: {e}")
        ad_text = None
    if ad_text:
        # AI matni — oddiy matn (emoji + bezak). HTML parse qilinmaydi (xavfsiz).
        caption = ad_text.rstrip()
        parse_mode = None
        # Optom: pachka/ranglar/razmer faktlarini AI matniga ham KAFOLATLI qo'shamiz.
        if _is_optom:
            _plain = []
            if product.get('pack_size'):
                _plain.append(f"📦 1 pachka = {int(product['pack_size'])} dona")
            if extra_facts.get('mavjud ranglar'):
                _plain.append(f"🎨 Ranglar: {extra_facts['mavjud ranglar']}")
            if extra_facts.get('razmerlar'):
                _plain.append(f"📏 Razmer: {extra_facts['razmerlar']}")
            if _plain:
                caption += "\n\n" + "\n".join(_plain)
        # Optom zina narxlarini AI matniga ham kafolatli qo'shamiz (oddiy matn).
        if _w['enabled']:
            _tiers_plain = ", ".join(f"{t['min']}+ {_unit} — {fmt_price(t['price'])}" for t in _w['tiers'])
            caption = caption.rstrip() + f"\n\n📦 Optom narx: {_tiers_plain}"
    return caption, parse_mode


async def _build_ad_design_bytes(context, product):
    """Mahsulot rasmiga reklama dizayni qo'yib JPEG bytes qaytaradi (yoki None)."""
    photo = product.get('image_url')
    if not (photo and ad_design.is_enabled()):
        return None
    try:
        tg_file = await context.bot.get_file(photo)
        raw = bytes(await tg_file.download_as_bytearray())
        shop_name = product.get('shop_name')
        region_lbl = region_label_l(product.get('seller_region_id'), DEFAULT_LANG)
        # Optom rozetkasi — pachka belgisi (rasm ustida)
        optom_txt = ''
        if product.get('sale_mode') == 'optom':
            _ps = product.get('pack_size')
            optom_txt = f"OPTOM · 1 PACHKA = {int(_ps)} DONA" if _ps else "OPTOM"
        return await asyncio.to_thread(
            ad_design.build_ad_image, raw,
            price_text=fmt_price(product.get('price')),
            badge_text='',
            shop_text=(str(shop_name) if shop_name else (region_lbl or '')),
            optom_text=optom_txt,
        )
    except Exception as e:
        logging.warning(f"Reklama dizayni yasalmadi: {e}")
        return None


def _is_permanent_channel_error(err) -> bool:
    """Post xatosi DOIMIY (bot chiqarilgan / huquqsiz / kanal o'chgan) ekanini aniqlaydi.

    Faqat shunday xatolarda kanal/guruh "yetim" deb o'chiriladi. Vaqtinchalik yoki
    konfiguratsiya xatolari (forum topic yopiq, caption uzun, file_id, tarmoq/limit)
    DOIMIY hisoblanmaydi — aks holda guruh keraksiz o'chib, post boshqa bormay qoladi."""
    # Forbidden — deyarli har doim doimiy (bot kicked / yozish taqiqlangan / huquq yo'q)
    if isinstance(err, Forbidden):
        return True
    msg = str(err).lower()
    permanent_markers = (
        "bot was kicked", "bot is not a member", "chat not found",
        "user is deactivated", "chat_write_forbidden",
        "not enough rights", "have no rights to send",
        "need administrator rights", "chat_admin_required",
        "peer_id_invalid", "the group chat was deleted",
        "bot was blocked", "group chat was migrated",
    )
    return any(m in msg for m in permanent_markers)


async def _notify_seller_sold_out(context, product_id):
    """Mahsulot sotilib tugab, avtomatik zaxiraga o'tganda sotuvchini ogohlantiradi.
    Xabarda darhol zaxira sonini yangilash tugmasi ham bo'ladi."""
    try:
        prod = db.get_product_by_id(product_id)
        if not prod:
            return
        seller = db.get_user_by_id(prod.get('seller_id'))
        if not seller or not seller.get('telegram_id'):
            return
        slang = get_user_lang(seller)
        kb = InlineKeyboardMarkup([[
            InlineKeyboardButton(t(slang, 'btn_set_stock'), callback_data=f"set_stock_{product_id}")
        ]])
        await context.bot.send_message(
            chat_id=seller['telegram_id'],
            text=t(slang, 'stock_sold_out_notify', name=html.escape(prod.get('name') or '')),
            parse_mode='HTML', reply_markup=kb,
        )
    except Exception as e:
        logging.warning(f"Sold-out notify failed (product {product_id}): {e}")


async def post_product_to_channel(context, product_id, *,
                                  caption_override=None, parse_mode_override=None,
                                  image_override=None, collect_sent=False):
    """Mahsulotni markaziy kanalga VA sotuvchining shaxsiy kanaliga post qiladi.

    caption_override / parse_mode_override — sotuvchi ko'rib tasdiqlagan AYNAN o'sha
    matnni joylash uchun (preview bilan 100% mos bo'lsin).
    image_override — allaqachon yuklangan (dizayn) rasm file_id si; berilsa qayta
    dizayn qilinmaydi va qayta yuklanmaydi.
    collect_sent=True bo'lsa — yuborilgan xabarlar ro'yxatini qaytaradi
    ([{'chat_id':.., 'message_id':..}]) — avto qayta-reklama eski postni o'chirishi uchun.
    Xato yuz bersa ham asosiy oqimga (saqlash/status) ta'sir qilmaydi."""
    sent_refs = []
    try:
        product = db.get_product_by_id(product_id)
        if not product:
            return

        bot_me = await context.bot.get_me()
        deep_link = _product_buy_link(bot_me.username, product_id)   # kanal tugmasi → Mini App

        # === REKLAMA MATNI (A) ===
        if caption_override is not None:
            caption, caption_parse_mode = caption_override, parse_mode_override
        else:
            caption, caption_parse_mode = await _build_ad_caption(product)

        # App buyer sahifasi kanaldagi AYNAN shu reklama matnini ko'rsatadi (parite)
        try:
            db.set_product_ad_caption(product_id, caption, caption_parse_mode)
        except Exception as e:
            logging.warning(f"ad_caption saqlanmadi (pid {product_id}): {e}")

        keyboard = InlineKeyboardMarkup([[
            InlineKeyboardButton("🛒 Sotib olish", url=deep_link)
        ]])

        # Maqsad kanallar: (chat_id, owner_seller_id, thread_id). Markaziy kanal — owner None.
        # thread_id faqat forum (mavzuli) guruhlar uchun — post o'sha topic ichiga boradi.
        # Faqat FAOL sotuvchi kanallariga yuboramiz (yetimlari tashlab ketiladi).
        targets = []
        seen = set()
        if CHANNEL_ID:
            targets.append((CHANNEL_ID, None, None))
            seen.add(str(CHANNEL_ID))
        seller_id = product.get('seller_id')
        if seller_id:
            for ch in db.get_active_seller_channels(seller_id):
                cid = ch.get('channel_id')
                if cid and str(cid) not in seen:
                    th = ch.get('thread_id')
                    th = int(th) if th not in (None, '') else None
                    targets.append((cid, seller_id, th))
                    seen.add(str(cid))

        photo = product.get('image_url')

        # === REKLAMA DIZAYNI (B) ===
        # image_override berilsa — tayyor file_id ni hamma kanalga qayta ishlatamiz.
        # Aks holda dizaynni yasab, birinchi yuborishdan keyin file_id ni eslab qolamiz.
        reusable_id = image_override
        designed_bytes = None
        if reusable_id is None:
            designed_bytes = await _build_ad_design_bytes(context, product)

        # Forum topic bilan bog'liq (vaqtinchalik) xatolar — guruhni o'chirmaymiz, General ga qaytamiz
        TOPIC_ERR_MARKERS = ("message thread not found", "topic_closed",
                             "topic was closed", "topic_deleted", "thread not found")
        # Telegram rasm captioni 1024 belgidan oshmasligi kerak — uzun bo'lsa ajratib yuboramiz
        too_long = bool(caption) and len(caption) > 1024

        async def _send_to(chat_id, thread_id):
            """Bitta chatga (forum bo'lsa topic ichiga) reklamani yuboradi.
            Caption 1024 belgidan uzun bo'lsa — rasm va to'liq matnni ajratib yuboradi."""
            nonlocal reusable_id
            if reusable_id:
                send_photo_arg = reusable_id
            elif designed_bytes is not None:
                send_photo_arg = io.BytesIO(designed_bytes)
            elif photo:
                send_photo_arg = photo
            else:
                send_photo_arg = None

            if send_photo_arg is not None:
                if too_long:
                    # Rasm captionsiz; to'liq matn (va "Sotib olish" tugmasi) alohida xabarda
                    sent = await context.bot.send_photo(
                        chat_id=chat_id, photo=send_photo_arg, message_thread_id=thread_id,
                    )
                    sent_refs.append({'chat_id': chat_id, 'message_id': sent.message_id})
                    sent2 = await context.bot.send_message(
                        chat_id=chat_id, text=caption, parse_mode=caption_parse_mode,
                        reply_markup=keyboard, message_thread_id=thread_id,
                    )
                    sent_refs.append({'chat_id': chat_id, 'message_id': sent2.message_id})
                else:
                    sent = await context.bot.send_photo(
                        chat_id=chat_id, photo=send_photo_arg,
                        caption=caption, parse_mode=caption_parse_mode,
                        reply_markup=keyboard, message_thread_id=thread_id,
                    )
                    sent_refs.append({'chat_id': chat_id, 'message_id': sent.message_id})
                # Birinchi yuborilgan rasmni keyingi chatlar uchun eslab qolamiz
                if reusable_id is None:
                    try:
                        reusable_id = sent.photo[-1].file_id
                    except Exception:
                        pass
            else:
                sent3 = await context.bot.send_message(
                    chat_id=chat_id, text=caption, parse_mode=caption_parse_mode,
                    reply_markup=keyboard, message_thread_id=thread_id,
                )
                sent_refs.append({'chat_id': chat_id, 'message_id': sent3.message_id})

        for chat_id, owner_id, thread_id in targets:
            try:
                try:
                    await _send_to(chat_id, thread_id)
                except BadRequest as e:
                    # Forum topic yopiq/o'chgan — guruhni o'chirmasdan General ga qayta urinamiz
                    if thread_id is not None and any(m in str(e).lower() for m in TOPIC_ERR_MARKERS):
                        logging.warning(f"Forum topic muammosi (chat {chat_id}, thread {thread_id}): {e} — General ga yuboriladi")
                        await _send_to(chat_id, None)
                    else:
                        raise
            except (Forbidden, BadRequest) as e:
                if _is_permanent_channel_error(e):
                    # Doimiy xato: bot kanaldan chiqarilgan / huquqi yo'q / kanal o'chgan.
                    # Sotuvchi kanali bo'lsa — yetim deb belgilaymiz va sotuvchini ogohlantiramiz.
                    logging.warning(f"Channel post permanent error (product {product_id}, chat {chat_id}): {e}")
                    if owner_id is not None and db.deactivate_seller_channel(owner_id, chat_id, str(e)):
                        try:
                            seller = db.get_user_by_id(owner_id)
                            seller_tg = seller.get('telegram_id') if seller else None
                            if seller_tg:
                                slang = get_user_lang(seller)
                                await context.bot.send_message(
                                    chat_id=seller_tg,
                                    text=t(slang, 'channel_deactivated_notify'),
                                    parse_mode='HTML',
                                )
                        except Exception as notify_err:
                            logging.warning(f"Channel deactivation notify failed (seller {owner_id}): {notify_err}")
                else:
                    # Vaqtinchalik/konfiguratsiya xatosi (caption, file_id, topic, limit) —
                    # guruh/kanalni O'CHIRMAYMIZ, faqat loglaymiz. Keyingi postlar baribir boradi.
                    logging.error(f"Channel post non-permanent error (product {product_id}, chat {chat_id}): {e}")
            except Exception as e:
                # Vaqtinchalik xato (tarmoq/limit) — kanalni o'chirmaymiz, faqat loglaymiz.
                logging.error(f"Channel post failed (product {product_id}, chat {chat_id}): {e}")
    except Exception as e:
        logging.error(f"post_product_to_channel failed (product {product_id}): {e}")
    return sent_refs if collect_sent else None


# ============================================================
# REKLAMA KO'RINISHI (preview) — kanalga joylashdan oldin tasdiqlash
# ============================================================
def _ad_preview_control_kb(lang, length="long"):
    """Preview ostidagi boshqaruv tugmalari. length — joriy reklama uzunligi
    ('long'/'short'); tanlangani ✅ bilan belgilanadi."""
    rows = [[InlineKeyboardButton(t(lang, 'ad_confirm_publish'), callback_data="adprev_publish")]]
    rows.append([InlineKeyboardButton(t(lang, 'ad_schedule_btn'), callback_data="adprev_schedule")])
    rows.append([InlineKeyboardButton(t(lang, 'autorep_btn'), callback_data="adprev_autorep")])
    if ai_assistant.is_enabled():
        # Sotuvchi reklama matnini UZUN yoki QISQA qilib tanlaydi (AI qayta yozadi)
        long_lbl = ("✅ " if length == 'long' else "") + t(lang, 'ad_len_long')
        short_lbl = ("✅ " if length == 'short' else "") + t(lang, 'ad_len_short')
        rows.append([
            InlineKeyboardButton(long_lbl, callback_data="adprev_long"),
            InlineKeyboardButton(short_lbl, callback_data="adprev_short"),
        ])
        rows.append([InlineKeyboardButton(t(lang, 'ad_regen'), callback_data="adprev_regen")])
    rows.append([InlineKeyboardButton(t(lang, 'ad_edit_text'), callback_data="adprev_edit")])
    rows.append([InlineKeyboardButton(t(lang, 'ad_skip'), callback_data="adprev_skip")])
    return InlineKeyboardMarkup(rows)


async def _render_ad_preview(context, chat_id, lang, prev):
    """Saqlangan preview holatidan AYNAN kanaldagidek ko'rinishni yuboradi.
    prev['image_id'] mavjud bo'lsa — qayta yuklamasdan ishlatadi."""
    caption = prev['caption']
    pm = prev.get('parse_mode')
    buy_kb = InlineKeyboardMarkup([[InlineKeyboardButton("🛒 Sotib olish", url=prev['deep_link'])]])
    # 1) Aynan reklama ko'rinishi
    try:
        if prev.get('image_id'):
            await context.bot.send_photo(chat_id=chat_id, photo=prev['image_id'],
                                         caption=caption, parse_mode=pm, reply_markup=buy_kb)
        elif prev.get('image_bytes'):
            await context.bot.send_photo(chat_id=chat_id, photo=io.BytesIO(prev['image_bytes']),
                                         caption=caption, parse_mode=pm, reply_markup=buy_kb)
        else:
            await context.bot.send_message(chat_id=chat_id, text=caption,
                                           parse_mode=pm, reply_markup=buy_kb)
    except Exception as e:
        logging.warning(f"Preview render xatosi: {e}")
        await context.bot.send_message(chat_id=chat_id, text=caption)
    # 2) Boshqaruv paneli
    await context.bot.send_message(chat_id=chat_id, text=t(lang, 'ad_preview_question'),
                                   reply_markup=_ad_preview_control_kb(lang, prev.get('length', 'long')))


async def show_ad_preview(update, context, product_id):
    """Mahsulot saqlangandan keyin — kanalga joylashdan OLDIN reklama ko'rinishini
    sotuvchiga ko'rsatadi (dizayn rasm + reklama matni)."""
    lang = get_lang(update, context)
    chat_id = update.effective_chat.id
    product = db.get_product_by_id(product_id)
    if not product:
        return
    await context.bot.send_message(chat_id=chat_id, text=t(lang, 'ad_preview_preparing'))

    caption, pm = await _build_ad_caption(product)
    designed = await _build_ad_design_bytes(context, product)
    photo = product.get('image_url')
    bot_me = await context.bot.get_me()
    deep_link = _product_buy_link(bot_me.username, product_id)   # preview tugmasi → Mini App
    buy_kb = InlineKeyboardMarkup([[InlineKeyboardButton("🛒 Sotib olish", url=deep_link)]])

    # 1) Aynan reklama ko'rinishini yuboramiz va (rasmli bo'lsa) file_id ni eslab qolamiz
    image_id = None
    try:
        if designed is not None:
            sent = await context.bot.send_photo(chat_id=chat_id, photo=io.BytesIO(designed),
                                                caption=caption, parse_mode=pm, reply_markup=buy_kb)
            try:
                image_id = sent.photo[-1].file_id
            except Exception:
                image_id = None
        elif photo:
            sent = await context.bot.send_photo(chat_id=chat_id, photo=photo,
                                                caption=caption, parse_mode=pm, reply_markup=buy_kb)
            try:
                image_id = sent.photo[-1].file_id
            except Exception:
                image_id = None
        else:
            await context.bot.send_message(chat_id=chat_id, text=caption,
                                           parse_mode=pm, reply_markup=buy_kb)
    except Exception as e:
        logging.warning(f"Preview yuborish xatosi: {e}")
        await context.bot.send_message(chat_id=chat_id, text=caption)

    context.user_data['ad_preview'] = {
        'product_id': product_id, 'caption': caption, 'parse_mode': pm,
        'image_id': image_id, 'deep_link': deep_link, 'length': 'long',
    }
    context.user_data.pop('ad_editing_caption', None)

    # 2) Boshqaruv paneli
    await context.bot.send_message(chat_id=chat_id, text=t(lang, 'ad_preview_question'),
                                   reply_markup=_ad_preview_control_kb(lang, 'long'))


async def ad_preview_publish(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """«✅ Kanalga joylash» — preview da ko'rsatilgan AYNAN reklamani joylaydi."""
    query = update.callback_query
    lang = get_lang(update, context)
    prev = context.user_data.get('ad_preview')
    if not prev:
        await query.answer(t(lang, 'ad_preview_expired'), show_alert=True)
        return
    await query.answer()
    try:
        await query.edit_message_text(t(lang, 'ad_publishing'))
    except Exception:
        pass
    await post_product_to_channel(
        context, prev['product_id'],
        caption_override=prev['caption'], parse_mode_override=prev.get('parse_mode'),
        image_override=prev.get('image_id'),
    )
    context.user_data.pop('ad_preview', None)
    context.user_data.pop('ad_editing_caption', None)
    kb = InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'btn_home'), callback_data="seller_panel")]])
    try:
        await query.edit_message_text(t(lang, 'ad_published'), reply_markup=kb)
    except Exception:
        await context.bot.send_message(chat_id=update.effective_chat.id,
                                       text=t(lang, 'ad_published'), reply_markup=kb)


async def ad_preview_regen(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """«🔄 Boshqa variant» — AI yangi reklama matni yozadi, dizayn rasm o'zgarmaydi."""
    query = update.callback_query
    lang = get_lang(update, context)
    prev = context.user_data.get('ad_preview')
    if not prev:
        await query.answer(t(lang, 'ad_preview_expired'), show_alert=True)
        return
    await query.answer()
    try:
        await query.edit_message_text(t(lang, 'ad_preview_preparing'))
    except Exception:
        pass
    product = db.get_product_by_id(prev['product_id'])
    if product:
        caption, pm = await _build_ad_caption(product, length=prev.get('length', 'long'))
        prev['caption'], prev['parse_mode'] = caption, pm
        context.user_data['ad_preview'] = prev
    await _render_ad_preview(context, update.effective_chat.id, lang, prev)


async def ad_preview_set_length(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """«📏 Uzun matn» / «✂️ Qisqa matn» — AI reklamani tanlangan uzunlikda qayta yozadi.
    Dizayn rasm o'zgarmaydi."""
    query = update.callback_query
    lang = get_lang(update, context)
    prev = context.user_data.get('ad_preview')
    if not prev:
        await query.answer(t(lang, 'ad_preview_expired'), show_alert=True)
        return
    length = 'short' if query.data == 'adprev_short' else 'long'
    await query.answer()
    try:
        await query.edit_message_text(t(lang, 'ad_preview_preparing'))
    except Exception:
        pass
    prev['length'] = length
    product = db.get_product_by_id(prev['product_id'])
    if product:
        caption, pm = await _build_ad_caption(product, length=length)
        prev['caption'], prev['parse_mode'] = caption, pm
    context.user_data['ad_preview'] = prev
    await _render_ad_preview(context, update.effective_chat.id, lang, prev)


async def ad_preview_edit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """«✏️ Matnni tahrirlash» — sotuvchidan o'z reklama matnini so'raydi."""
    query = update.callback_query
    lang = get_lang(update, context)
    prev = context.user_data.get('ad_preview')
    if not prev:
        await query.answer(t(lang, 'ad_preview_expired'), show_alert=True)
        return
    await query.answer()
    context.user_data['ad_editing_caption'] = True
    try:
        await query.edit_message_text(t(lang, 'ad_edit_prompt'))
    except Exception:
        await context.bot.send_message(chat_id=update.effective_chat.id, text=t(lang, 'ad_edit_prompt'))


async def ad_preview_caption_input(update, context):
    """Sotuvchi tahrirlangan reklama matnini yubordi — preview ni yangilaymiz."""
    lang = get_lang(update, context)
    prev = context.user_data.get('ad_preview')
    context.user_data.pop('ad_editing_caption', None)
    if not prev:
        await update.message.reply_text(t(lang, 'ad_preview_expired'))
        return
    new_text = (update.message.text or '').strip()
    if new_text:
        prev['caption'] = new_text
        prev['parse_mode'] = None   # sotuvchi matni — oddiy matn
        context.user_data['ad_preview'] = prev
    await _render_ad_preview(context, update.effective_chat.id, lang, prev)


async def ad_preview_skip(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """«⏭ Hozircha joylamayman» — mahsulot saqlangan, kanalga chiqarilmaydi."""
    query = update.callback_query
    lang = get_lang(update, context)
    context.user_data.pop('ad_preview', None)
    context.user_data.pop('ad_editing_caption', None)
    await query.answer()
    kb = InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'btn_home'), callback_data="seller_panel")]])
    try:
        await query.edit_message_text(t(lang, 'ad_skipped'), reply_markup=kb)
    except Exception:
        await context.bot.send_message(chat_id=update.effective_chat.id,
                                       text=t(lang, 'ad_skipped'), reply_markup=kb)


# ============================================================
# REJALASHTIRILGAN POST — sotuvchi/xodim mahsulotni belgilangan sana va soatda
# avtomatik sotuvga qo'yadi (botda faollashadi + kanal/guruhlarga reklama chiqadi).
# Belgilangan vaqtgacha mahsulot 'scheduled' holatda — botda ko'rinmaydi.
# ============================================================
def _sched_date_label(lang, off, d):
    """Sana tugmasi yorlig'i: Bugun / Ertaga / DD.MM."""
    if off == 0:
        return t(lang, 'sched_today')
    if off == 1:
        return t(lang, 'sched_tomorrow')
    return f"{d.day:02d}.{d.month:02d}"


def _sched_date_kb(lang):
    from datetime import datetime, timedelta
    now = datetime.now(TZ_TASHKENT)
    rows, row = [], []
    for off in range(7):
        d = now + timedelta(days=off)
        row.append(InlineKeyboardButton(_sched_date_label(lang, off, d),
                                        callback_data=f"schd_date_{off}"))
        if len(row) == 2:
            rows.append(row); row = []
    if row:
        rows.append(row)
    rows.append([InlineKeyboardButton(t(lang, 'sched_abort_btn'), callback_data="schd_abort")])
    return InlineKeyboardMarkup(rows)


def _sched_hour_kb(lang, off):
    """Soat tugmalari (00..23). Bugun bo'lsa — o'tib ketgan soatlar chiqarilmaydi."""
    from datetime import datetime
    now = datetime.now(TZ_TASHKENT)
    start = now.hour if off == 0 else 0
    rows, row = [], []
    for h in range(start, 24):
        row.append(InlineKeyboardButton(f"{h:02d}", callback_data=f"schd_hour_{h}"))
        if len(row) == 6:
            rows.append(row); row = []
    if row:
        rows.append(row)
    if not rows:  # bugun barcha soatlar o'tib ketgan
        return None
    rows.append([InlineKeyboardButton(t(lang, 'sched_abort_btn'), callback_data="schd_abort")])
    return InlineKeyboardMarkup(rows)


def _sched_minute_kb(lang, off, hour):
    """Daqiqa tugmalari (00/15/30/45). Bugun va shu soat bo'lsa — o'tgan daqiqalar yo'q."""
    from datetime import datetime
    now = datetime.now(TZ_TASHKENT)
    mins = [0, 15, 30, 45]
    if off == 0 and hour == now.hour:
        mins = [m for m in mins if m > now.minute]
    if not mins:
        return None
    row = [InlineKeyboardButton(f":{m:02d}", callback_data=f"schd_min_{m}") for m in mins]
    rows = [row, [InlineKeyboardButton(t(lang, 'sched_abort_btn'), callback_data="schd_abort")]]
    return InlineKeyboardMarkup(rows)


async def ad_preview_schedule_start(update, context):
    """«⏰ Rejalashtirish» — preview'dagi reklamani saqlab, sana tanlashni so'raydi."""
    query = update.callback_query
    lang = get_lang(update, context)
    prev = context.user_data.get('ad_preview')
    if not prev:
        await query.answer(t(lang, 'ad_preview_expired'), show_alert=True)
        return
    await query.answer()
    context.user_data['sched'] = {
        'product_id': prev['product_id'],
        'caption': prev['caption'],
        'parse_mode': prev.get('parse_mode'),
        'image_id': prev.get('image_id'),
    }
    try:
        await query.edit_message_text(t(lang, 'sched_pick_date'),
                                      reply_markup=_sched_date_kb(lang), parse_mode='HTML')
    except Exception:
        await context.bot.send_message(chat_id=update.effective_chat.id,
                                       text=t(lang, 'sched_pick_date'),
                                       reply_markup=_sched_date_kb(lang), parse_mode='HTML')


async def sched_pick_date(update, context):
    """schd_date_{off} — sana tanlandi, soat tanlashni so'raydi."""
    query = update.callback_query
    lang = get_lang(update, context)
    sd = context.user_data.get('sched')
    if not sd:
        await query.answer(t(lang, 'ad_preview_expired'), show_alert=True)
        return
    off = int(query.data.split('_')[-1])
    sd['date_offset'] = off
    await query.answer()
    kb = _sched_hour_kb(lang, off)
    if kb is None:
        await query.answer(t(lang, 'sched_in_past'), show_alert=True)
        await query.edit_message_text(t(lang, 'sched_pick_date'),
                                      reply_markup=_sched_date_kb(lang), parse_mode='HTML')
        return
    await query.edit_message_text(t(lang, 'sched_pick_hour'), reply_markup=kb, parse_mode='HTML')


async def sched_pick_hour(update, context):
    """schd_hour_{HH} — soat tanlandi, daqiqa tanlashni so'raydi."""
    query = update.callback_query
    lang = get_lang(update, context)
    sd = context.user_data.get('sched')
    if not sd:
        await query.answer(t(lang, 'ad_preview_expired'), show_alert=True)
        return
    hour = int(query.data.split('_')[-1])
    sd['hour'] = hour
    await query.answer()
    kb = _sched_minute_kb(lang, sd.get('date_offset', 0), hour)
    if kb is None:
        await query.answer(t(lang, 'sched_in_past'), show_alert=True)
        await query.edit_message_text(t(lang, 'sched_pick_hour'),
                                      reply_markup=_sched_hour_kb(lang, sd.get('date_offset', 0)),
                                      parse_mode='HTML')
        return
    await query.edit_message_text(t(lang, 'sched_pick_minute', hour=f"{hour:02d}"),
                                  reply_markup=kb, parse_mode='HTML')


async def sched_pick_minute(update, context):
    """schd_min_{MM} — daqiqa tanlandi: rejani yaratadi, mahsulotni yashiradi, job qo'yadi."""
    from datetime import datetime, timedelta, timezone
    query = update.callback_query
    lang = get_lang(update, context)
    sd = context.user_data.get('sched')
    if not sd:
        await query.answer(t(lang, 'ad_preview_expired'), show_alert=True)
        return
    minute = int(query.data.split('_')[-1])
    off = sd.get('date_offset', 0)
    hour = sd.get('hour', 0)

    now_local = datetime.now(TZ_TASHKENT)
    target_local = (now_local + timedelta(days=off)).replace(
        hour=hour, minute=minute, second=0, microsecond=0)
    if target_local <= now_local:
        await query.answer(t(lang, 'sched_in_past'), show_alert=True)
        return
    await query.answer()
    target_utc = target_local.astimezone(timezone.utc)

    product_id = sd['product_id']
    product = db.get_product_by_id(product_id)
    if not product:
        await query.edit_message_text(t(lang, 'ad_preview_expired'))
        context.user_data.pop('sched', None)
        return
    owner_id = product.get('seller_id')
    user = db.get_user_by_telegram_id(update.effective_user.id)
    created_by = user['id'] if user else owner_id

    # #18 Pro — bepul rejalashtirish limiti (app schedule gate'i bilan bir xil; Pro = cheksiz)
    slim = db.mon_limit('mon_free_scheduled_limit')
    if slim > 0 and not db.is_pro(owner_id) and db.count_pending_scheduled_posts(owner_id) >= slim:
        await query.edit_message_text(t(lang, 'pro_locked_limit_bot'), parse_mode='HTML',
                                      reply_markup=_pro_gate_kb(lang))
        context.user_data.pop('sched', None)
        return

    sched_id = db.create_scheduled_post(
        product_id, owner_id, target_utc.strftime("%Y-%m-%d %H:%M:%S"),
        created_by=created_by, caption=sd.get('caption'),
        parse_mode=sd.get('parse_mode'), image_id=sd.get('image_id'))

    # Mahsulotni yashiramiz — belgilangan vaqtgacha botda ko'rinmaydi/sotib olib bo'lmaydi
    db.set_product_status(product_id, 'scheduled')

    # Job qo'yamiz (xotirada; restartda _reschedule_scheduled_posts tiklaydi)
    when = max(1, (target_utc - datetime.now(timezone.utc)).total_seconds())
    context.application.job_queue.run_once(
        scheduled_post_job, when=when,
        data={'sched_id': sched_id}, name=f"sched_post_{sched_id}")

    context.user_data.pop('sched', None)
    context.user_data.pop('ad_preview', None)
    when_str = target_local.strftime("%d.%m.%Y %H:%M")
    kb = InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'btn_home'), callback_data="seller_panel")]])
    await query.edit_message_text(
        t(lang, 'sched_confirmed', name=html.escape(product.get('name') or ''), when=when_str),
        reply_markup=kb, parse_mode='HTML')


async def sched_abort_flow(update, context):
    """«Bekor qilish» — rejalashtirish ustasidan chiqadi (mahsulot sotuvda qoladi)."""
    query = update.callback_query
    lang = get_lang(update, context)
    context.user_data.pop('sched', None)
    await query.answer()
    kb = InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'btn_home'), callback_data="seller_panel")]])
    try:
        await query.edit_message_text(t(lang, 'sched_aborted'), reply_markup=kb)
    except Exception:
        pass


async def scheduled_post_job(context: ContextTypes.DEFAULT_TYPE):
    """Belgilangan vaqt kelganda: mahsulotni faollashtiradi + kanal/guruhlarga reklama
    joylaydi va sotuvchini xabardor qiladi."""
    from datetime import datetime as _dt, timezone as _tz
    sched_id = context.job.data.get('sched_id')
    sp = db.get_scheduled_post(sched_id)
    if not sp or sp.get('status') != 'pending':
        return  # bekor qilingan yoki allaqachon bajarilgan
    product_id = sp['product_id']
    product = db.get_product_by_id(product_id)
    if not product or product.get('status') in ('deleted', 'purged'):
        db.mark_scheduled_post(sched_id, 'failed')
        return

    # 1) Mahsulotni sotuvga qo'yamiz (botda faollashadi)
    db.set_product_status(product_id, 'active')

    # 2) Kanal va barcha ulangan guruh/kanallarga reklama (preview'dagi AYNAN matn/rasm)
    try:
        await post_product_to_channel(
            context, product_id,
            caption_override=sp.get('caption'),
            parse_mode_override=sp.get('parse_mode'),
            image_override=sp.get('image_id'))
    except Exception as e:
        logging.error(f"Rejalashtirilgan post joylashda xato (sched {sched_id}): {e}")

    db.mark_scheduled_post(sched_id, 'posted', posted_at=_dt.now(_tz.utc))

    # 3) Sotuvchi (ega) va rejani tuzgan xodimni xabardor qilamiz
    notified = set()
    for uid in (sp.get('seller_id'), sp.get('created_by')):
        if not uid or uid in notified:
            continue
        notified.add(uid)
        try:
            u = db.get_user_by_id(uid)
            if u and u.get('telegram_id'):
                ulang = get_user_lang(u)
                await context.bot.send_message(
                    chat_id=u['telegram_id'],
                    text=t(ulang, 'sched_job_done', name=html.escape(product.get('name') or '')),
                    parse_mode='HTML')
        except Exception as e:
            logging.warning(f"Rejalashtirilgan post bildirishnomasi ketmadi (user {uid}): {e}")


async def seller_scheduled_posts(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """«⏰ Rejalashtirilgan postlar» — kutilayotgan rejalar ro'yxati + bekor qilish."""
    query = update.callback_query
    lang = get_lang(update, context)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    owner_id = db.resolve_owner_id(user['id'])
    posts = db.get_seller_scheduled_posts(owner_id)

    if not posts:
        kb = InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="seller_panel")]])
        text = t(lang, 'scheduled_list_empty')
        if query:
            await query.answer()
            await query.edit_message_text(text, reply_markup=kb)
        else:
            await update.message.reply_text(text, reply_markup=kb)
        return

    lines = [t(lang, 'scheduled_list_title')]
    keyboard = []
    for sp in posts:
        when_str = fmt_datetime(sp.get('scheduled_at'))
        lines.append(t(lang, 'scheduled_list_item',
                       name=html.escape(sp.get('product_name') or ''), when=when_str))
        keyboard.append([InlineKeyboardButton(
            t(lang, 'scheduled_cancel_btn', name=(sp.get('product_name') or '')[:20]),
            callback_data=f"schd_cancel_{sp['id']}")])
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="seller_panel")])
    text = "\n".join(lines)
    if query:
        await query.answer()
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')
    else:
        await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


async def sched_cancel_post(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """schd_cancel_{id} — rejani bekor qiladi: jobni o'chiradi, mahsulotni zaxiraga oladi."""
    query = update.callback_query
    lang = get_lang(update, context)
    sched_id = int(query.data.split('_')[-1])
    user = db.get_user_by_telegram_id(update.effective_user.id)
    owner_id = db.resolve_owner_id(user['id'])
    sp = db.cancel_scheduled_post(sched_id, seller_id=owner_id)
    if not sp:
        await query.answer(t(lang, 'scheduled_cancel_failed'), show_alert=True)
        await seller_scheduled_posts(update, context)
        return
    # Jobni o'chiramiz
    try:
        for job in context.application.job_queue.get_jobs_by_name(f"sched_post_{sched_id}"):
            job.schedule_removal()
    except Exception as e:
        logging.warning(f"Reja jobini o'chirishda xato (sched {sched_id}): {e}")
    # Mahsulotni zaxiraga olamiz (sotuvchi keyin qo'lda faollashtira oladi)
    try:
        db.set_product_status(sp['product_id'], 'reserve')
    except Exception:
        pass
    await query.answer(t(lang, 'scheduled_cancelled'), show_alert=True)
    await seller_scheduled_posts(update, context)


# ============================================================
# AVTO QAYTA-REKLAMA (kuniga bir marta avtomatik qayta chiqarish)
# ============================================================
AUTOREPOST_MAX_DAYS = 30   # avto-to'xtash: shu kundan keyin o'zi to'xtaydi


def _autorep_hour_kb(lang):
    """Soat tugmalari (00..23) — kuniga qaysi soatda qayta chiqsin (Toshkent vaqti)."""
    rows, row = [], []
    for h in range(24):
        row.append(InlineKeyboardButton(f"{h:02d}", callback_data=f"arep_hour_{h}"))
        if len(row) == 6:
            rows.append(row); row = []
    if row:
        rows.append(row)
    rows.append([InlineKeyboardButton(t(lang, 'sched_abort_btn'), callback_data="arep_abort")])
    return InlineKeyboardMarkup(rows)


async def ad_preview_autorepost_start(update, context):
    """«🔁 Avto qayta-reklama» — preview'dagi reklamani saqlab, kunlik soatni so'raydi."""
    query = update.callback_query
    lang = get_lang(update, context)
    prev = context.user_data.get('ad_preview')
    if not prev:
        await query.answer(t(lang, 'ad_preview_expired'), show_alert=True)
        return
    await query.answer()
    context.user_data['autorep'] = {
        'product_id': prev['product_id'],
        'caption': prev['caption'],
        'parse_mode': prev.get('parse_mode'),
        'image_id': prev.get('image_id'),
    }
    try:
        await query.edit_message_text(t(lang, 'autorep_pick_hour'),
                                      reply_markup=_autorep_hour_kb(lang), parse_mode='HTML')
    except Exception:
        await context.bot.send_message(chat_id=update.effective_chat.id,
                                       text=t(lang, 'autorep_pick_hour'),
                                       reply_markup=_autorep_hour_kb(lang), parse_mode='HTML')


def _schedule_autorepost_job(job_queue, repost_id, hour):
    """Kunlik (har kuni shu soatda, Toshkent vaqti) job qo'yadi.
    Avval shu nomli eski jobni o'chiradi (qayta yoqishda ikkilanmasin)."""
    from datetime import time as _time
    name = f"autorep_{repost_id}"
    try:
        for j in job_queue.get_jobs_by_name(name):
            j.schedule_removal()
    except Exception:
        pass
    job_queue.run_daily(
        auto_repost_job, time=_time(hour=hour, minute=0, tzinfo=TZ_TASHKENT),
        data={'repost_id': repost_id}, name=name)


async def autorep_pick_hour(update, context):
    """arep_hour_{HH} — soat tanlandi: mahsulotni HOZIR joylaydi + kunlik avto-reklamani yoqadi."""
    from datetime import datetime, timezone, timedelta
    query = update.callback_query
    lang = get_lang(update, context)
    ar = context.user_data.get('autorep')
    if not ar:
        await query.answer(t(lang, 'ad_preview_expired'), show_alert=True)
        return
    hour = int(query.data.split('_')[-1])
    await query.answer()

    product_id = ar['product_id']
    product = db.get_product_by_id(product_id)
    if not product:
        await query.edit_message_text(t(lang, 'ad_preview_expired'))
        context.user_data.pop('autorep', None)
        return
    owner_id = product.get('seller_id')
    user = db.get_user_by_telegram_id(update.effective_user.id)
    created_by = user['id'] if user else owner_id

    # Mahsulot sotuvda bo'lsin (botda ko'rinsin)
    if product.get('status') != 'active':
        db.set_product_status(product_id, 'active')

    expires = datetime.now(timezone.utc) + timedelta(days=AUTOREPOST_MAX_DAYS)
    repost_id = db.upsert_auto_repost(
        product_id, owner_id, hour, created_by=created_by,
        caption=ar.get('caption'), parse_mode=ar.get('parse_mode'),
        image_id=ar.get('image_id'), expires_at=expires)

    # HOZIR birinchi marta joylaymiz va xabar id larini eslab qolamiz (keyin o'chirish uchun)
    refs = []
    try:
        refs = await post_product_to_channel(
            context, product_id, caption_override=ar.get('caption'),
            parse_mode_override=ar.get('parse_mode'), image_override=ar.get('image_id'),
            collect_sent=True) or []
    except Exception as e:
        logging.error(f"Avto qayta-reklama birinchi joylashda xato (repost {repost_id}): {e}")
    db.update_auto_repost_run(repost_id, json.dumps(refs), last_run_at=datetime.now(timezone.utc))

    # Kunlik jobni qo'yamiz (restartda _reschedule_auto_reposts tiklaydi)
    if context.application.job_queue:
        _schedule_autorepost_job(context.application.job_queue, repost_id, hour)

    context.user_data.pop('autorep', None)
    context.user_data.pop('ad_preview', None)
    kb = InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'btn_home'), callback_data="seller_panel")]])
    await query.edit_message_text(
        t(lang, 'autorep_confirmed', name=html.escape(product.get('name') or ''),
          hour=f"{hour:02d}", days=AUTOREPOST_MAX_DAYS),
        reply_markup=kb, parse_mode='HTML')


async def autorep_abort_flow(update, context):
    """«Bekor qilish» — avto-reklama ustasidan chiqadi (mahsulot holati o'zgarmaydi)."""
    query = update.callback_query
    lang = get_lang(update, context)
    context.user_data.pop('autorep', None)
    await query.answer()
    kb = InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'btn_home'), callback_data="seller_panel")]])
    try:
        await query.edit_message_text(t(lang, 'sched_aborted'), reply_markup=kb)
    except Exception:
        pass


async def _delete_old_repost_messages(context, refs):
    """Oldingi avto-reklama xabarlarini o'chiradi (kanal toza qolsin). Xato — e'tiborsiz."""
    for ref in (refs or []):
        try:
            await context.bot.delete_message(chat_id=ref.get('chat_id'),
                                             message_id=ref.get('message_id'))
        except Exception:
            pass   # 48 soatdan o'tgan / huquq yo'q — muhim emas, yangisi baribir chiqadi


async def _notify_autorepost_stopped(context, ar, *, reason):
    """Avto-reklama avtomatik to'xtaganda sotuvchi (va yoqgan xodim)ni xabardor qiladi."""
    product = db.get_product_by_id(ar.get('product_id'))
    name = (product.get('name') if product else '') or ''
    notified = set()
    for uid in (ar.get('seller_id'), ar.get('created_by')):
        if not uid or uid in notified:
            continue
        notified.add(uid)
        try:
            u = db.get_user_by_id(uid)
            if u and u.get('telegram_id'):
                ulang = get_user_lang(u)
                await context.bot.send_message(
                    chat_id=u['telegram_id'],
                    text=t(ulang, 'autorep_stopped_notify', name=html.escape(name)),
                    parse_mode='HTML')
        except Exception as e:
            logging.warning(f"Avto-reklama to'xtash bildirishnomasi ketmadi (user {uid}): {e}")


async def auto_repost_job(context: ContextTypes.DEFAULT_TYPE):
    """Har kuni belgilangan soatda: eski reklamani o'chirib, yangisini chiqaradi.
    Mahsulot sotilgan/o'chirilgan/zaxira 0 yoki muddat tugagan bo'lsa — avto-to'xtaydi."""
    from datetime import datetime, timezone
    repost_id = context.job.data.get('repost_id')
    ar = db.get_auto_repost(repost_id)
    if not ar or not ar.get('is_active'):
        context.job.schedule_removal()
        return

    # Muddat tugadimi?
    exp = ar.get('expires_at')
    if exp:
        try:
            exp_dt = datetime.strptime(str(exp)[:19], "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
            if datetime.now(timezone.utc) >= exp_dt:
                db.deactivate_auto_repost(repost_id)
                context.job.schedule_removal()
                await _notify_autorepost_stopped(context, ar, reason='expired')
                return
        except Exception:
            pass

    product = db.get_product_by_id(ar['product_id'])
    # Avto-to'xtash sharti: mahsulot yo'q/o'chirilgan/sotuvda emas/zaxira 0
    sc = product.get('stock_count') if product else None
    sold_out = (sc is not None and sc <= 0)
    if (not product or product.get('status') != 'active' or sold_out):
        db.deactivate_auto_repost(repost_id)
        context.job.schedule_removal()
        await _notify_autorepost_stopped(context, ar, reason='unavailable')
        return

    # Eskisini o'chir, yangisini chiqar (kanal toza qolsin, yangi a'zolar tepada ko'radi)
    try:
        old_refs = json.loads(ar.get('last_message_ids') or "[]")
    except Exception:
        old_refs = []
    await _delete_old_repost_messages(context, old_refs)

    refs = []
    try:
        refs = await post_product_to_channel(
            context, ar['product_id'], caption_override=ar.get('caption'),
            parse_mode_override=ar.get('parse_mode'), image_override=ar.get('image_id'),
            collect_sent=True) or []
    except Exception as e:
        logging.error(f"Avto qayta-reklama joylashda xato (repost {repost_id}): {e}")
    db.update_auto_repost_run(repost_id, json.dumps(refs), last_run_at=datetime.now(timezone.utc))


async def seller_auto_reposts(update, context):
    """«🔁 Avto qayta-reklamalar» — faol avto-reklamalar ro'yxati + o'chirish."""
    query = update.callback_query
    lang = get_lang(update, context)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    owner_id = db.resolve_owner_id(user['id'])
    rows = db.get_seller_auto_reposts(owner_id)
    if not rows:
        kb = InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="seller_panel")]])
        text = t(lang, 'autorep_list_empty')
        if query:
            await query.answer()
            await query.edit_message_text(text, reply_markup=kb)
        else:
            await update.message.reply_text(text, reply_markup=kb)
        return
    lines = [t(lang, 'autorep_list_title')]
    keyboard = []
    for ar in rows:
        lines.append(t(lang, 'autorep_list_item',
                       name=html.escape(ar.get('product_name') or ''), hour=f"{ar.get('hour'):02d}"))
        keyboard.append([InlineKeyboardButton(
            t(lang, 'autorep_cancel_btn', name=(ar.get('product_name') or '')[:20]),
            callback_data=f"arep_cancel_{ar['id']}")])
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="seller_panel")])
    text = "\n".join(lines)
    if query:
        await query.answer()
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')
    else:
        await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


async def autorep_cancel(update, context):
    """arep_cancel_{id} — avto-reklamani o'chiradi: jobni ham olib tashlaydi."""
    query = update.callback_query
    lang = get_lang(update, context)
    repost_id = int(query.data.split('_')[-1])
    user = db.get_user_by_telegram_id(update.effective_user.id)
    owner_id = db.resolve_owner_id(user['id'])
    ar = db.cancel_auto_repost(repost_id, seller_id=owner_id)
    if not ar:
        await query.answer(t(lang, 'scheduled_cancel_failed'), show_alert=True)
        await seller_auto_reposts(update, context)
        return
    try:
        for job in context.application.job_queue.get_jobs_by_name(f"autorep_{repost_id}"):
            job.schedule_removal()
    except Exception as e:
        logging.warning(f"Avto-reklama jobini o'chirishda xato (repost {repost_id}): {e}")
    await query.answer(t(lang, 'autorep_cancelled'), show_alert=True)
    await seller_auto_reposts(update, context)


async def product_autorepost_start(update, context):
    """arep_start_{pid} — MAVJUD (eski yoki yangi) mahsulotga avto qayta-reklama yoqish.
    Reklama matni har safar yangidan tuziladi (override yo'q) — har chiqishda yangicha."""
    query = update.callback_query
    lang = get_lang(update, context)
    pid = int(query.data.split("_")[-1])
    user = db.get_user_by_telegram_id(update.effective_user.id)
    owner_id = db.resolve_owner_id(user['id']) if user else None
    product = db.get_product_by_id(pid)
    if not product or product.get('seller_id') != owner_id:
        await query.answer(t(lang, 'product_not_found'), show_alert=True)
        return
    await query.answer()
    # caption=None → auto_repost_job har safar post_product_to_channel'da yangi reklama tuzadi
    context.user_data['autorep'] = {
        'product_id': pid, 'caption': None, 'parse_mode': None, 'image_id': None,
    }
    try:
        await query.edit_message_text(t(lang, 'autorep_pick_hour'),
                                      reply_markup=_autorep_hour_kb(lang), parse_mode='HTML')
    except Exception:
        await context.bot.send_message(chat_id=update.effective_chat.id,
                                       text=t(lang, 'autorep_pick_hour'),
                                       reply_markup=_autorep_hour_kb(lang), parse_mode='HTML')


async def product_autorepost_stop(update, context):
    """arep_off_{pid} — mahsulot menyusidan avto qayta-reklamani o'chiradi."""
    query = update.callback_query
    lang = get_lang(update, context)
    pid = int(query.data.split("_")[-1])
    user = db.get_user_by_telegram_id(update.effective_user.id)
    owner_id = db.resolve_owner_id(user['id']) if user else None
    ar = db.get_auto_repost_by_product(pid)
    if ar and ar.get('seller_id') == owner_id:
        db.cancel_auto_repost(ar['id'], seller_id=owner_id)
        try:
            for job in context.application.job_queue.get_jobs_by_name(f"autorep_{ar['id']}"):
                job.schedule_removal()
        except Exception as e:
            logging.warning(f"Avto-reklama jobini o'chirishda xato (repost {ar['id']}): {e}")
        await query.answer(t(lang, 'autorep_cancelled'), show_alert=True)
    else:
        await query.answer()
    await seller_product_menu(update, context, product_id=pid)


def _reschedule_auto_reposts(job_queue):
    """Restart/deploy'dan keyin faol avto qayta-reklamalar uchun kunlik joblarni tiklaydi."""
    try:
        rows = db.get_active_auto_reposts()
        n = 0
        for ar in rows:
            try:
                _schedule_autorepost_job(job_queue, ar['id'], int(ar['hour']))
                n += 1
            except Exception:
                continue
        logging.info(f"Restart: {n} ta avto qayta-reklama qayta tiklandi.")
    except Exception as e:
        logging.error(f"Avto qayta-reklamalarni qayta tiklash xatosi: {e}")


async def _maybe_preview_on_reactivation(update, context, product_id, was_active):
    """Mahsulot zahiradan (reserve) sotuvga (active) qaytgan bo'lsa — reklama
    ko'rinishini ko'rsatadi. True qaytaradi (preview ko'rsatildi)."""
    if was_active:
        return False   # avval ham sotuvda edi — qayta reklama shart emas
    prod = db.get_product_by_id(product_id)
    if prod and (prod.get('status') == 'active'):
        await show_ad_preview(update, context, product_id)
        return True
    return False


async def _save_product(update, context):
    """Mahsulotni DB ga saqlaydi."""
    user = db.get_user_by_telegram_id(update.effective_user.id)
    lang = get_lang(update, context)
    photos = [p for p in context.user_data.get('product_photos', []) if p][:db.MAX_PRODUCT_IMAGES]
    stock_count = context.user_data.get('product_stock')  # None = cheksiz

    # ===== MULTI-SOTUVCHI: mahsulot do'kon EGASIga tegishli, lekin xodimga attribute qilinadi =====
    staff = db.get_staff_by_user(user['id'])
    is_staff = bool(staff and staff.get('staff_role') != 'owner')
    owner_id = db.resolve_owner_id(user['id'])
    shop = db.get_shop_for_user(user['id'])

    # Nofaol (tasdiqlanmagan) xodim mahsulot joylay olmaydi
    if is_staff and not staff.get('is_active', 1):
        _reply = update.message.reply_text if update.message else update.callback_query.message.reply_text
        await _reply(t(lang, 'staff_inactive_block'))
        for k in ('product_name', 'product_price', 'product_stock', 'product_category',
                  'product_desc', 'product_photo', 'product_photos', 'attr_templates',
                  'attr_index', 'adding_product', 'question_mode', 'product_attrs',
                  'attr_labels'):
            context.user_data.pop(k, None)
        return ConversationHandler.END

    # Ruxsat tekshiruvi — xodim mahsulot qo'sha olmasa
    if is_staff and not staff.get('perm_add_product', 1):
        _reply = update.message.reply_text if update.message else update.callback_query.message.reply_text
        await _reply(t(lang, 'staff_no_perm_add'))
        for k in ('product_name', 'product_price', 'product_stock', 'product_category',
                  'product_desc', 'product_photo', 'product_photos', 'attr_templates',
                  'attr_index', 'adding_product', 'question_mode', 'product_attrs',
                  'attr_labels'):
            context.user_data.pop(k, None)
        return ConversationHandler.END

    # Moderatsiya: ega tasdig'i talab qilinsa — pending_owner holatida saqlanadi
    needs_owner_approval = bool(is_staff and shop and shop.get('moderation') == 'owner_approve')

    product_id = db.create_product(
        seller_id=owner_id,
        name=context.user_data['product_name'],
        price=context.user_data['product_price'],
        category_id=context.user_data.get('product_category'),
        description=context.user_data.get('product_desc'),
        image_url=(photos[0] if photos else None),
        stock_count=stock_count,
        created_by=user['id'],
        status=('pending_owner' if needs_owner_approval else None),
    )

    # Barcha rasmlarni saqlaymiz (image_url ham birinchi rasmga sinxronlanadi)
    if photos and product_id:
        db.set_product_images(product_id, photos)

    # Atributlarni saqlash (AI yorliqlari bilan — xom kalit ko'rinmasin)
    attrs = context.user_data.pop('product_attrs', {})
    attr_labels = context.user_data.pop('attr_labels', {})
    if attrs and product_id:
        db.save_product_attributes(product_id, attrs, labels=attr_labels)

    # State tozalash
    for k in ('product_name', 'product_price', 'product_stock', 'product_category',
              'product_desc', 'product_photo', 'product_photos', 'attr_templates',
              'attr_index', 'adding_product', 'question_mode', 'attr_labels'):
        context.user_data.pop(k, None)

    # Moderatsiya: ega tasdig'i kerak bo'lsa — xodimga xabar, egaga tasdiq so'rovi
    if needs_owner_approval and product_id:
        _reply = update.message.reply_text if update.message else update.callback_query.message.reply_text
        await _reply(t(lang, 'product_sent_for_approval'))
        try:
            owner = db.get_user_by_id(owner_id)
            if owner and owner.get('telegram_id'):
                olang = get_user_lang(owner)
                await context.bot.send_message(
                    chat_id=owner['telegram_id'],
                    text=t(olang, 'owner_product_review',
                           staff=html.escape(user.get('name') or ''),
                           pname=html.escape(context.user_data.get('product_name') or ''),
                           price=fmt_price(context.user_data.get('product_price') or 0)),
                    parse_mode='HTML',
                    reply_markup=InlineKeyboardMarkup([
                        [InlineKeyboardButton(t(olang, 'btn_approve'), callback_data=f"ownappr_{product_id}")],
                        [InlineKeyboardButton(t(olang, 'btn_reject'), callback_data=f"ownrej_{product_id}")],
                    ])
                )
        except Exception as e:
            logging.error(f"Egaga mahsulot tasdig'i xabari ketmadi: {e}")
        return ConversationHandler.END

    msg = t(lang, 'product_saved')
    if photos:
        msg += t(lang, 'frag_photos_saved', n=len(photos))
    if attrs:
        msg += t(lang, 'frag_attrs_saved', n=len(attrs))
    # Zaxira holatini ham ko'rsatamiz
    if stock_count is None:
        msg += t(lang, 'frag_stock_unlim')
    else:
        msg += t(lang, 'frag_stock_saved', n=stock_count)

    if update.message:
        await update.message.reply_text(msg)
    else:
        await update.callback_query.message.reply_text(msg)

    # Kanalga joylashdan OLDIN — reklama ko'rinishini ko'rsatamiz (tasdiq/tahrir)
    if product_id:
        await show_ad_preview(update, context, product_id)

    return ConversationHandler.END


async def owner_review_product(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Do'kon egasi xodim joylagan mahsulotni tasdiqlaydi (ownappr_) yoki rad etadi (ownrej_).
    Faqat shu do'kon egasi amal qila oladi."""
    query = update.callback_query
    await query.answer()
    data = query.data
    approve = data.startswith("ownappr_")
    product_id = int(data.split("_")[1])
    lang = get_lang(update, context)
    actor = db.get_user_by_telegram_id(update.effective_user.id)
    product = db.get_product_by_id(product_id)
    if not product:
        await query.edit_message_text(t(lang, 'product_not_found'))
        return
    # Egalik tekshiruvi: actor shu mahsulot do'koni egasi (yoki admin) bo'lishi shart
    owner_id = db.resolve_owner_id(product['seller_id'])
    is_admin = actor and (actor.get('role') == 'admin' or update.effective_user.id == ADMIN_ID)
    if not (actor and (actor['id'] == owner_id or actor['id'] == product['seller_id'] or is_admin)):
        await query.edit_message_text(t(lang, 'not_your_order_plain'))
        return
    if product.get('status') != 'pending_owner':
        await query.edit_message_text(t(lang, 'owner_review_already'))
        return

    staff_user = db.get_user_by_id(product.get('created_by')) if product.get('created_by') else None
    if approve:
        db.set_product_status(product_id, 'active')
        await query.edit_message_text(t(lang, 'owner_approved_done', pname=html.escape(product['name'] or '')))
        if staff_user and staff_user.get('telegram_id'):
            try:
                slang = get_user_lang(staff_user)
                await context.bot.send_message(
                    chat_id=staff_user['telegram_id'],
                    text=t(slang, 'staff_product_approved', pname=html.escape(product['name'] or '')),
                    parse_mode='HTML')
            except Exception as e:
                logging.error(f"Xodimga tasdiq xabari ketmadi: {e}")
    else:
        db.set_product_status(product_id, 'deleted')
        await query.edit_message_text(t(lang, 'owner_rejected_done', pname=html.escape(product['name'] or '')))
        if staff_user and staff_user.get('telegram_id'):
            try:
                slang = get_user_lang(staff_user)
                await context.bot.send_message(
                    chat_id=staff_user['telegram_id'],
                    text=t(slang, 'staff_product_rejected', pname=html.escape(product['name'] or '')),
                    parse_mode='HTML')
            except Exception as e:
                logging.error(f"Xodimga rad xabari ketmadi: {e}")


async def seller_products(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mahsulotlarim — 3 bo'lim: sotuvda / zahirada / o'chirilgan + qidiruv."""
    query = update.callback_query
    if query:
        await query.answer()

    lang = get_lang(update, context)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    counts = db.count_seller_products_by_status(user['id'])

    text = t(lang, 'my_products_overview',
             active=counts['active'], reserve=counts['reserve'], deleted=counts['deleted'])

    keyboard = [
        [InlineKeyboardButton(t(lang, 'btn_on_sale_n', n=counts['active']), callback_data="sp_list_active_0")],
        [InlineKeyboardButton(t(lang, 'btn_reserve_n', n=counts['reserve']), callback_data="sp_list_reserve_0")],
        [InlineKeyboardButton(t(lang, 'btn_deleted_n', n=counts['deleted']), callback_data="sp_list_deleted_0")],
        [InlineKeyboardButton(t(lang, 'btn_search_product'), callback_data="sp_search")],
        [InlineKeyboardButton(t(lang, 'back'), callback_data="seller_panel")],
    ]

    if query:
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')
    else:
        await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


SP_PAGE = 8  # sahifadagi mahsulotlar soni

async def seller_products_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Status bo'yicha mahsulotlar ro'yxati (paginatsiya bilan)."""
    query = update.callback_query
    await query.answer()

    # callback: sp_list_{status}_{page}
    parts = query.data.split("_")
    status = parts[2]
    page = int(parts[3])

    lang = get_lang(update, context)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    products = db.get_seller_products_by_status(user['id'], status=status)

    status_labels = {'active': t(lang, 'status_on_sale'), 'reserve': t(lang, 'status_reserve_short'),
                     'deleted': t(lang, 'status_deleted_short')}

    if not products:
        await query.edit_message_text(
            t(lang, 'section_empty', status=status_labels.get(status)),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="seller_products")]])
        )
        return

    total = len(products)
    total_pages = max(1, (total + SP_PAGE - 1) // SP_PAGE)
    page = max(0, min(page, total_pages - 1))
    start = page * SP_PAGE
    end = min(start + SP_PAGE, total)

    keyboard = []
    for p in products[start:end]:
        stock_info = ""
        if p.get('stock_count') is not None:
            stock_info = t(lang, 'frag_stock_pieces', n=p['stock_count'])
        keyboard.append([InlineKeyboardButton(
            f"{p['name']} — {fmt_price(p['price'])}{stock_info}",
            callback_data=f"prod_menu_{p['id']}"
        )])

    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton("⬅️", callback_data=f"sp_list_{status}_{page-1}"))
    nav.append(InlineKeyboardButton(f"{page+1}/{total_pages}", callback_data="noop"))
    if page < total_pages - 1:
        nav.append(InlineKeyboardButton("➡️", callback_data=f"sp_list_{status}_{page+1}"))
    if nav:
        keyboard.append(nav)
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="seller_products")])

    await query.edit_message_text(
        t(lang, 'section_page', status=status_labels.get(status), total=total, page=page+1, pages=total_pages),
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def seller_product_search_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mahsulot qidirish — so'rov so'raydi."""
    query = update.callback_query
    await query.answer()
    await query.edit_message_text(
        T(update, context, 'seller_search_prompt'),
        parse_mode='HTML'
    )
    return SELLER_PRODUCT_SEARCH


async def seller_product_search_result(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Qidiruv natijasini ko'rsatadi."""
    lang = get_lang(update, context)
    search_text = update.message.text.strip()
    user = db.get_user_by_telegram_id(update.effective_user.id)
    products = db.search_seller_products(user['id'], search_text)

    if not products:
        await update.message.reply_text(
            t(lang, 'seller_search_none', q=search_text),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'btn_my_products_back'), callback_data="seller_products")]])
        )
        return ConversationHandler.END

    status_emoji = {'active': '✅', 'reserve': '📥', 'deleted': '🗑'}
    keyboard = []
    for p in products[:20]:
        st = p.get('status') or 'active'
        emoji = status_emoji.get(st, '')
        stock_info = t(lang, 'frag_stock_pieces', n=p['stock_count']) if p.get('stock_count') is not None else ""
        keyboard.append([InlineKeyboardButton(
            f"{emoji} {p['name']} — {fmt_price(p['price'])}{stock_info}",
            callback_data=f"prod_menu_{p['id']}"
        )])
    keyboard.append([InlineKeyboardButton(t(lang, 'btn_my_products_back'), callback_data="seller_products")])

    await update.message.reply_text(
        t(lang, 'seller_search_found', q=search_text, n=len(products)),
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return ConversationHandler.END


async def seller_product_menu(update: Update, context: ContextTypes.DEFAULT_TYPE, product_id: int = None):
    query = update.callback_query
    # Agar to'g'ridan-to'g'ri callback dan kelsa — javob beramiz; ichki chaqiriqda esa o'tkazib yuboramiz
    if product_id is None:
        await query.answer()
        product_id = int(query.data.split("_")[2])
    context.user_data['editing_product_id'] = product_id
    lang = get_lang(update, context)

    product = db.get_product_basic(product_id)

    if not product:
        await query.edit_message_text(t(lang, 'product_not_found'))
        return

    status = product.get('status') or 'active'
    status_labels = {
        'active':  t(lang, 'pm_status_active'),
        'reserve': t(lang, 'pm_status_reserve'),
        'deleted': t(lang, 'pm_status_deleted'),
    }
    status_text = status_labels.get(status, status)

    # Statusga qarab tugmalar
    keyboard = [[InlineKeyboardButton(t(lang, 'btn_edit'), callback_data=f"edit_start_{product_id}")]]
    keyboard.append([InlineKeyboardButton(t(lang, 'btn_share_link'), callback_data=f"share_link_{product_id}")])

    # Avto qayta-reklama tugmasi (faol/zaxiradagi mahsulot uchun) — holatga qarab
    # yoqish yoki o'chirish. Eski (oldin joylangan) mahsulotlarga ham qo'yish mumkin.
    def _autorep_row():
        ar = db.get_auto_repost_by_product(product_id)
        if ar:
            return [InlineKeyboardButton(
                t(lang, 'btn_autorep_off', hour=f"{ar.get('hour'):02d}"),
                callback_data=f"arep_off_{product_id}")]
        return [InlineKeyboardButton(t(lang, 'autorep_btn'),
                                     callback_data=f"arep_start_{product_id}")]

    if status == 'active':
        keyboard.append([InlineKeyboardButton(t(lang, 'btn_to_reserve'), callback_data=f"pstatus_reserve_{product_id}")])
        keyboard.append([InlineKeyboardButton(t(lang, 'btn_set_stock'), callback_data=f"set_stock_{product_id}")])
        keyboard.append(_autorep_row())
        keyboard.append([InlineKeyboardButton(t(lang, 'btn_remove_from_sale'), callback_data=f"pstatus_deleted_{product_id}")])
    elif status == 'reserve':
        keyboard.append([InlineKeyboardButton(t(lang, 'btn_return_to_sale'), callback_data=f"pstatus_active_{product_id}")])
        keyboard.append([InlineKeyboardButton(t(lang, 'btn_set_stock'), callback_data=f"set_stock_{product_id}")])
        keyboard.append(_autorep_row())
        keyboard.append([InlineKeyboardButton(t(lang, 'btn_remove_from_sale'), callback_data=f"pstatus_deleted_{product_id}")])
    else:  # deleted
        keyboard.append([InlineKeyboardButton(t(lang, 'btn_repost_sale'), callback_data=f"pstatus_active_{product_id}")])
        keyboard.append([InlineKeyboardButton(t(lang, 'btn_delete_forever'), callback_data=f"delete_prod_{product_id}")])

    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="seller_products")])

    # HTML rejimi + foydalanuvchi matnini escape
    name = html.escape(product['name'] or '')
    desc = html.escape(product.get('description') or t(lang, 'none_word'))
    if product.get('stock_count') is not None:
        stock_line = t(lang, 'pm_stock_line', n=product['stock_count'])
    else:
        stock_line = t(lang, 'pm_stock_unlimited')

    # Atributlar
    attrs = db.get_product_attributes(product_id)
    attrs_text = ""
    if attrs:
        lines = [f"• {a.get('attr_label') or a['attr_key']}: {a['attr_value']}" for a in attrs]
        attrs_text = t(lang, 'pm_attrs_title') + "\n".join(lines)

    body = t(lang, 'product_menu_body',
             name=name, price=fmt_price(product['price']),
             status=status_text, stock=stock_line, desc=desc, attrs=attrs_text)

    # Sotuvchi mahsulotining rasm(lar)ini ham ko'rsatamiz — shunda rasm
    # saqlangani-saqlanmaganini o'z ko'zi bilan tekshira oladi.
    images = db.get_product_images(product_id)
    if images:
        try:
            await query.message.delete()
        except Exception:
            pass
        await _send_product_card(context, update.effective_chat.id, images, body,
                                 keyboard)
    else:
        await query.edit_message_text(
            body, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML'
        )


async def change_product_status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Mahsulot statusini o'zgartiradi: active/reserve/deleted."""
    query = update.callback_query
    await query.answer()

    # callback: pstatus_{status}_{product_id}
    parts = query.data.split("_")
    new_status = parts[1]
    product_id = int(parts[2])

    db.set_product_status(product_id, new_status)

    lang = get_lang(update, context)

    # Qayta sotuvga qo'yilganda — kanalga to'g'ridan chiqarmaymiz, avval reklama
    # ko'rinishini ko'rsatamiz (sotuvchi tasdiqlasa/tahrirlasa, keyin joylanadi).
    if new_status == 'active':
        await query.answer(t(lang, 'pstatus_active_toast'), show_alert=True)
        await seller_product_menu(update, context, product_id=product_id)
        await show_ad_preview(update, context, product_id)
        return

    labels = {
        'reserve': t(lang, 'pstatus_reserve_toast'),
        'deleted': t(lang, 'pstatus_deleted_toast'),
    }
    await query.answer(labels.get(new_status, t(lang, 'status_changed_toast')), show_alert=True)

    # Menyuni yangilaymiz
    await seller_product_menu(update, context, product_id=product_id)


async def toggle_product_stock(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    product_id = int(query.data.split("_")[2])

    db.toggle_product_in_stock(product_id)

    # query.data ni o'zgartirmaymiz (PTB v22'da CallbackQuery muzlatilgan).
    # Buning o'rniga product_id'ni to'g'ridan-to'g'ri uzatamiz.
    await seller_product_menu(update, context, product_id=product_id)


async def set_stock_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Zaxira belgilash — avval tugma: '♾ Cheksiz' yoki '🔢 Aniq miqdor'."""
    query = update.callback_query
    await query.answer()

    product_id = int(query.data.split("_")[2])
    lang = get_lang(update, context)
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(t(lang, 'btn_stock_unlimited'), callback_data=f"setstock_unlim_{product_id}")],
        [InlineKeyboardButton(t(lang, 'btn_stock_limited'), callback_data=f"setstock_num_{product_id}")],
    ])
    try:
        await query.edit_message_text(t(lang, 'set_stock_choose'), reply_markup=kb)
    except Exception:
        await context.bot.send_message(chat_id=update.effective_chat.id,
                                       text=t(lang, 'set_stock_choose'), reply_markup=kb)


async def set_stock_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """'♾ Cheksiz' yoki '🔢 Aniq miqdor' tanlovini qayta ishlaydi.
    callback: setstock_unlim_{id} / setstock_num_{id}"""
    query = update.callback_query
    await query.answer()
    parts = query.data.split("_")  # setstock, unlim|num, id
    mode = parts[1]
    product_id = int(parts[2])
    lang = get_lang(update, context)

    if mode == "num":
        # Aniq son — text orqali kiritiladi (text_handler -> set_stock_submit)
        context.user_data['setting_stock_for'] = product_id
        try:
            await query.edit_message_text(t(lang, 'set_stock_ask'))
        except Exception:
            await context.bot.send_message(chat_id=update.effective_chat.id, text=t(lang, 'set_stock_ask'))
        return

    # unlim — darhol cheksiz qilib belgilaymiz
    before = db.get_product_by_id(product_id)
    was_active = bool(before and before.get('status') == 'active')
    db.set_product_stock_count(product_id, None)
    try:
        await query.edit_message_text(t(lang, 'stock_set_unlimited'))
    except Exception:
        await context.bot.send_message(chat_id=update.effective_chat.id, text=t(lang, 'stock_set_unlimited'))
    # Zaxiradan sotuvga qaytgan bo'lsa — reklama ko'rinishini ko'rsatamiz
    if not await _maybe_preview_on_reactivation(update, context, product_id, was_active):
        await seller_panel(update, context)


async def set_stock_submit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """text_handler tomonidan chaqiriladi — setting_stock_for bo'lsa."""
    product_id = context.user_data.pop('setting_stock_for', None)
    if not product_id:
        return False  # bizning rejimimiz emas

    lang = get_lang(update, context)
    text = update.message.text.strip()
    if text == '-':
        new_stock = None
    else:
        try:
            new_stock = int(text)
            if new_stock < 0:
                raise ValueError
        except ValueError:
            await update.message.reply_text(t(lang, 'stock_invalid'))
            return True

    # Zahira to'ldirishdan oldingi holatni eslab qolamiz (reserve→active aniqlash uchun)
    before = db.get_product_by_id(product_id)
    was_active = bool(before and before.get('status') == 'active')

    db.set_product_stock_count(product_id, new_stock)

    if new_stock is None:
        await update.message.reply_text(t(lang, 'stock_set_unlimited'))
    else:
        await update.message.reply_text(t(lang, 'stock_set_n', n=new_stock))

    # Zahiradan sotuvga qaytgan bo'lsa — reklama ko'rinishini ko'rsatamiz,
    # aks holda oddiy holatda sotuvchi paneliga qaytamiz.
    if not await _maybe_preview_on_reactivation(update, context, product_id, was_active):
        await seller_panel(update, context)
    return True


async def delete_product(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    lang = get_lang(update, context)
    if data.startswith("delete_prod_") and not data.startswith("delete_confirm_"):
        product_id = int(data.split("_")[2])
        keyboard = [
            [InlineKeyboardButton(t(lang, 'btn_yes_delete'), callback_data=f"delete_confirm_{product_id}")],
            [InlineKeyboardButton(t(lang, 'btn_no_cancel'), callback_data=f"prod_menu_{product_id}")],
        ]
        await query.edit_message_text(
            t(lang, 'delete_confirm_ask'),
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return

    product_id = int(data.split("_")[2])
    actor = db.get_user_by_telegram_id(update.effective_user.id)
    hard_deleted = db.delete_product(
        product_id,
        deleted_by=actor['id'] if actor else None,
        deleted_by_role=(actor.get('role') if actor else None),
    )

    msg = t(lang, 'product_deleted') if hard_deleted else t(lang, 'product_deleted_kept_history')
    await query.edit_message_text(
        msg,
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'btn_my_products_back'), callback_data="seller_products")]])
    )


# ============================================================
# MAHSULOTNI TAHRIRLASH — "bir oyna + qaysi qismni tanlash"
# Sotuvchi ma'lumotlarning hammasini bir oynada ko'radi va
# faqat o'zgartirmoqchi bo'lgan qismini tanlab tahrirlaydi.
# ============================================================

def _edit_field_kb(product_id, attrs, lang=DEFAULT_LANG):
    """Tahrir oynasidagi tugmalar — har bir maydon va xususiyat uchun alohida."""
    rows = [
        [InlineKeyboardButton(t(lang, 'ef_btn_name'), callback_data=f"ef_name_{product_id}")],
        [InlineKeyboardButton(t(lang, 'ef_btn_price'), callback_data=f"ef_price_{product_id}")],
        [InlineKeyboardButton(t(lang, 'ef_btn_cat'), callback_data=f"ef_cat_{product_id}")],
        [InlineKeyboardButton(t(lang, 'ef_btn_desc'), callback_data=f"ef_desc_{product_id}")],
        [InlineKeyboardButton(t(lang, 'ef_btn_photos'), callback_data=f"ef_photos_{product_id}")],
    ]
    for a in attrs:
        label = a.get('attr_label') or a['attr_key']
        rows.append([InlineKeyboardButton(
            f"🏷 {label}", callback_data=f"ea_{product_id}_{a['attr_key']}"
        )])
    rows.append([InlineKeyboardButton(t(lang, 'back'), callback_data=f"prod_menu_{product_id}")])
    return InlineKeyboardMarkup(rows)


def _edit_overview_text(product, attrs, img_count, lang=DEFAULT_LANG):
    """Mahsulotning barcha ma'lumotini bitta oynada ko'rsatuvchi matn."""
    name = html.escape(product.get('name') or '—')
    desc = html.escape(product.get('description') or t(lang, 'none_word'))
    cat = html.escape(category_name(product.get('category_name'), lang) or t(lang, 'category_not_selected'))
    lines = [
        t(lang, 'edit_title'),
        "",
        t(lang, 'edit_lbl_name', v=name),
        t(lang, 'edit_lbl_price', v=fmt_price(product.get('price') or 0)),
        t(lang, 'edit_lbl_cat', v=cat),
        t(lang, 'edit_lbl_photos', n=img_count),
        t(lang, 'edit_lbl_desc', v=desc),
    ]
    if attrs:
        lines.append("")
        lines.append(t(lang, 'edit_attrs_title'))
        for a in attrs:
            label = html.escape(a.get('attr_label') or a['attr_key'])
            val = html.escape(a.get('attr_value') or '—')
            lines.append(f"• {label}: {val}")
    lines.append("")
    lines.append(t(lang, 'edit_which_part'))
    return "\n".join(lines)


async def edit_product_hub(update: Update, context: ContextTypes.DEFAULT_TYPE, product_id: int = None):
    """Tahrir oynasi — hamma ma'lumot bir joyda, har biriga alohida tugma."""
    query = update.callback_query
    if product_id is None:
        if query and query.data and query.data.startswith("edit_start_"):
            product_id = int(query.data.split("_")[2])
        else:
            product_id = context.user_data.get('editing_product_id')
    lang = get_lang(update, context)
    if not product_id:
        if query:
            await query.answer(t(lang, 'product_not_found'), show_alert=True)
        return
    context.user_data['editing_product_id'] = product_id

    product = db.get_product_by_id(product_id)
    if not product:
        if query:
            await query.edit_message_text(t(lang, 'product_not_found'))
        else:
            await update.message.reply_text(t(lang, 'product_not_found'))
        return

    attrs = db.get_product_attributes(product_id)
    img_count = db.count_product_images(product_id)
    text = _edit_overview_text(product, attrs, img_count, lang)
    kb = _edit_field_kb(product_id, attrs, lang)

    if query:
        try:
            await query.edit_message_text(text, reply_markup=kb, parse_mode='HTML')
        except Exception:
            # Oldingi xabar rasm bo'lsa edit ishlamaydi — yangi xabar yuboramiz
            await context.bot.send_message(
                chat_id=update.effective_chat.id, text=text,
                reply_markup=kb, parse_mode='HTML'
            )
    else:
        await update.message.reply_text(text, reply_markup=kb, parse_mode='HTML')


def _pid_from_cb(query):
    """ef_name_5 / ef_price_5 / ef_cat_5 / ef_desc_5 / ef_photos_5 → 5"""
    return int(query.data.split("_")[2])


# ---------- NOM ----------
async def edit_field_name_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    context.user_data['editing_product_id'] = _pid_from_cb(query)
    await query.edit_message_text(T(update, context, 'edit_name_ask'))
    return EDIT_FIELD_NAME


async def edit_field_name_save(update: Update, context: ContextTypes.DEFAULT_TYPE):
    lang = get_lang(update, context)
    pid = context.user_data.get('editing_product_id')
    name = update.message.text.strip()
    if len(name) < 3:
        await update.message.reply_text(t(lang, 'edit_name_short'))
        return EDIT_FIELD_NAME
    if len(name) > 100:
        await update.message.reply_text(t(lang, 'edit_name_long'))
        return EDIT_FIELD_NAME
    db.update_product_fields(pid, name=name)
    await update.message.reply_text(t(lang, 'name_updated'))
    await edit_product_hub(update, context, product_id=pid)
    return ConversationHandler.END


# ---------- NARX ----------
async def edit_field_price_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    context.user_data['editing_product_id'] = _pid_from_cb(query)
    await query.edit_message_text(T(update, context, 'edit_price_ask'))
    return EDIT_FIELD_PRICE


async def edit_field_price_save(update: Update, context: ContextTypes.DEFAULT_TYPE):
    pid = context.user_data.get('editing_product_id')
    raw = update.message.text.replace(" ", "").replace(",", "").replace("\u00A0", "")
    lang = get_lang(update, context)
    try:
        price = float(raw)
    except ValueError:
        await update.message.reply_text(t(lang, 'edit_price_invalid'))
        return EDIT_FIELD_PRICE
    if price <= 0 or price > 1_000_000_000:
        await update.message.reply_text(t(lang, 'edit_price_range'))
        return EDIT_FIELD_PRICE
    db.update_product_fields(pid, price=price)
    await update.message.reply_text(t(lang, 'price_updated'))
    await edit_product_hub(update, context, product_id=pid)
    return ConversationHandler.END


# ---------- KATEGORIYA ----------
async def edit_field_cat_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    pid = _pid_from_cb(query)
    context.user_data['editing_product_id'] = pid
    categories = db.get_all_categories()
    lang = get_lang(update, context)
    kb = [[InlineKeyboardButton(f"{cat[2]} {category_name(cat[1], lang)}", callback_data=f"ecat_{cat[0]}")] for cat in categories]
    kb.append([InlineKeyboardButton(t(lang, 'btn_cancel_edit'), callback_data="ecat_cancel")])
    await query.edit_message_text(t(lang, 'edit_cat_ask'), reply_markup=InlineKeyboardMarkup(kb))
    return EDIT_FIELD_CATEGORY


async def edit_field_cat_save(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    pid = context.user_data.get('editing_product_id')
    if query.data != "ecat_cancel":
        cat_id = int(query.data.split("_")[1])
        db.update_product_fields(pid, category_id=cat_id)
    await edit_product_hub(update, context, product_id=pid)
    return ConversationHandler.END


# ---------- TAVSIF ----------
async def edit_field_desc_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    context.user_data['editing_product_id'] = _pid_from_cb(query)
    await query.edit_message_text(T(update, context, 'edit_desc_ask'))
    return EDIT_FIELD_DESC


async def edit_field_desc_save(update: Update, context: ContextTypes.DEFAULT_TYPE):
    lang = get_lang(update, context)
    pid = context.user_data.get('editing_product_id')
    desc = update.message.text.strip()
    if desc == "-":
        desc = None
    elif len(desc) > 500:
        await update.message.reply_text(t(lang, 'edit_desc_long'))
        return EDIT_FIELD_DESC
    db.update_product_fields(pid, description=desc)
    await update.message.reply_text(t(lang, 'desc_updated'))
    await edit_product_hub(update, context, product_id=pid)
    return ConversationHandler.END


# ---------- RASMLAR (barchasini almashtirish, 5 tagacha) ----------
async def edit_field_photos_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    pid = _pid_from_cb(query)
    context.user_data['editing_product_id'] = pid
    context.user_data['edit_photos'] = []
    await query.edit_message_text(T(update, context, 'edit_photos_ask'))
    return EDIT_FIELD_PHOTOS


async def edit_field_photos_collect(update: Update, context: ContextTypes.DEFAULT_TYPE):
    lang = get_lang(update, context)
    pid = context.user_data.get('editing_product_id')
    file_id = None
    if update.message.photo:
        photo = update.message.photo[-1]
        if photo.file_size and photo.file_size > 5 * 1024 * 1024:
            await update.message.reply_text(t(lang, 'edit_photo_too_big'))
            return EDIT_FIELD_PHOTOS
        if photo.width and photo.height and (photo.width < 200 or photo.height < 200):
            await update.message.reply_text(t(lang, 'edit_photo_too_small'))
            return EDIT_FIELD_PHOTOS
        file_id = photo.file_id
    elif update.message.document and (update.message.document.mime_type or '').startswith('image/'):
        doc = update.message.document
        if doc.file_size and doc.file_size > 5 * 1024 * 1024:
            await update.message.reply_text(t(lang, 'edit_photo_too_big'))
            return EDIT_FIELD_PHOTOS
        file_id = doc.file_id
    elif update.message.text and update.message.text.strip() == "-":
        db.set_product_images(pid, [])
        context.user_data.pop('edit_photos', None)
        await update.message.reply_text(t(lang, 'all_photos_deleted'))
        await edit_product_hub(update, context, product_id=pid)
        return ConversationHandler.END
    else:
        await update.message.reply_text(t(lang, 'edit_photo_send_or_dash'))
        return EDIT_FIELD_PHOTOS

    photos = context.user_data.setdefault('edit_photos', [])
    photos.append(file_id)
    n = len(photos)
    if n >= db.MAX_PRODUCT_IMAGES:
        db.set_product_images(pid, photos)
        context.user_data.pop('edit_photos', None)
        await update.message.reply_text(t(lang, 'photos_saved_max', n=n))
        await edit_product_hub(update, context, product_id=pid)
        return ConversationHandler.END

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(t(lang, 'btn_add_more_photo'), callback_data="eph_more")],
        [InlineKeyboardButton(t(lang, 'btn_save'), callback_data="eph_done")],
    ])
    await update.message.reply_text(
        t(lang, 'photos_selected_n', n=n, max=db.MAX_PRODUCT_IMAGES),
        reply_markup=kb
    )
    return EDIT_FIELD_PHOTOS


async def edit_photos_more(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    n = len(context.user_data.get('edit_photos', []))
    await query.edit_message_text(T(update, context, 'next_photo_edit', n=n, max=db.MAX_PRODUCT_IMAGES))
    return EDIT_FIELD_PHOTOS


async def edit_photos_done(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    pid = context.user_data.get('editing_product_id')
    photos = context.user_data.pop('edit_photos', [])
    if photos:
        db.set_product_images(pid, photos)
    await edit_product_hub(update, context, product_id=pid)
    return ConversationHandler.END


# ---------- XUSUSIYAT (atribut) ----------
async def edit_attr_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    # ea_{pid}_{attr_key}  (attr_key ichida '_' bo'lishi mumkin)
    parts = query.data.split("_", 2)
    pid = int(parts[1])
    attr_key = parts[2]
    context.user_data['editing_product_id'] = pid
    context.user_data['editing_attr_key'] = attr_key
    label = attr_key
    for a in db.get_product_attributes(pid):
        if a['attr_key'] == attr_key:
            label = a.get('attr_label') or attr_key
            break
    await query.edit_message_text(
        T(update, context, 'edit_attr_ask', label=html.escape(label)),
        parse_mode='HTML'
    )
    return EDIT_FIELD_ATTR


async def edit_attr_save(update: Update, context: ContextTypes.DEFAULT_TYPE):
    lang = get_lang(update, context)
    pid = context.user_data.get('editing_product_id')
    attr_key = context.user_data.get('editing_attr_key')
    if not pid or not attr_key:
        await update.message.reply_text(t(lang, 'edit_proc_error'))
        return ConversationHandler.END
    value = update.message.text.strip()
    if value != "-" and len(value) > 100:
        await update.message.reply_text(t(lang, 'attr_too_long'))
        return EDIT_FIELD_ATTR
    context.user_data.pop('editing_attr_key', None)
    if value == "-":
        db.delete_product_attribute(pid, attr_key)
        await update.message.reply_text(t(lang, 'attr_deleted'))
    else:
        db.save_product_attributes(pid, {attr_key: value})
        await update.message.reply_text(t(lang, 'attr_updated'))
    await edit_product_hub(update, context, product_id=pid)
    return ConversationHandler.END


async def seller_orders(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if query:
        await query.answer()

    # Callback (inline) yoki Reply tugma (matn) — ikkalasida ham ishlasin
    async def _show(text, **kw):
        if query:
            await query.edit_message_text(text, **kw)
        else:
            await update.message.reply_text(text, **kw)

    lang = get_lang(update, context)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    orders = db.get_orders_by_seller(user['id'])

    if not orders:
        await _show(
            t(lang, 'orders_empty'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="seller_panel")]])
        )
        return

    status_emoji = {'pending': '⏳', 'confirmed': '✅', 'delivered': '🚚', 'cancelled': '❌'}

    # Savat buyurtmalarini guruhlash
    group_agg = {}
    for o in orders:
        gid = o.get('order_group_id')
        if gid:
            g = group_agg.setdefault(gid, {'count': 0, 'sum': 0.0, 'status': o['status'],
                                           'buyer': o.get('buyer_name') or ''})
            g['count'] += 1
            g['sum'] += float(o['total_price'] or 0)

    keyboard = []
    seen_groups = set()
    shown = 0
    for order in orders:
        if shown >= 10:
            break
        gid = order.get('order_group_id')
        if gid:
            if gid in seen_groups:
                continue
            seen_groups.add(gid)
            g = group_agg[gid]
            prog = t(lang, 'row_progress_tag') if g['status'] in ('pending', 'confirmed') else ''
            keyboard.append([InlineKeyboardButton(
                t(lang, 'seller_order_group_row', emoji=status_emoji.get(g['status'], '❓'),
                  buyer=g['buyer'], count=g['count'], sum=fmt_price(g['sum']), prog=prog),
                callback_data=f"seller_gorder_{gid}"
            )])
        else:
            prog = t(lang, 'row_progress_tag') if order['status'] in ('pending', 'confirmed') else ''
            keyboard.append([InlineKeyboardButton(
                t(lang, 'seller_order_row', emoji=status_emoji.get(order['status'], '❓'),
                  buyer=order['buyer_name'], pname=(order.get('product_name') or '—')[:22],
                  qty=order['quantity'], total=fmt_price(order['total_price']), prog=prog),
                callback_data=f"seller_order_{order['id']}"
            )])
        shown += 1
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="seller_panel")])

    # Jarayondagi (yakunlanmagan) buyurtmalar sonini sarlavhada ko'rsatamiz
    in_progress = sum(1 for o in orders if o['status'] in ('pending', 'confirmed'))
    title = t(lang, 'orders_title')
    if in_progress:
        title += t(lang, 'orders_title_inprogress', n=in_progress)

    await _show(title, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


async def seller_profile(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = db.get_user_by_telegram_id(update.effective_user.id)
    avg_rating = db.get_seller_avg_rating(user['id'])
    lang = get_lang(update, context)

    # Karta holati
    card_info = ""
    if user.get('card_number'):
        cnum = user['card_number']
        masked = f"{cnum[:4]} **** **** {cnum[-4:]}" if len(cnum) >= 8 else cnum
        ctype = CARD_TYPE_LABELS.get(user.get('card_type', ''), '💳')
        card_info = f"\n{ctype} {masked} ({user.get('card_owner', '')})"
    else:
        card_info = t(lang, 'card_not_added')

    keyboard = [
        [InlineKeyboardButton(t(lang, 'btn_edit_shop_name'),    callback_data="edit_shop_name")],
        [InlineKeyboardButton(t(lang, 'btn_edit_address'),       callback_data="edit_shop_address")],
        [InlineKeyboardButton(t(lang, 'btn_edit_landmark'),      callback_data="edit_shop_landmark")],
        [InlineKeyboardButton(t(lang, 'btn_select_region'),      callback_data="edit_seller_region")],
        [InlineKeyboardButton(t(lang, 'btn_edit_working_days'),  callback_data="edit_working_days")],
        [InlineKeyboardButton(t(lang, 'btn_edit_working_hours'), callback_data="edit_working_hours")],
        [InlineKeyboardButton(t(lang, 'btn_edit_telegram'),      callback_data="edit_telegram")],
        [InlineKeyboardButton(t(lang, 'btn_card_info'),          callback_data="edit_card_info")],
        [InlineKeyboardButton(t(lang, 'btn_change_language'),    callback_data="change_lang")],
        [InlineKeyboardButton(t(lang, 'back'),                   callback_data="seller_panel")],
    ]

    # Hudud
    region_info = ""
    if user.get('region_id'):
        r = db.get_region_by_id(user['region_id'])
        if r:
            if r.get('parent_id'):
                parent = db.get_region_by_id(r['parent_id'])
                region_info = (f"{region_name(parent['name'], lang)} → {region_name(r['name'], lang)}"
                               if parent else region_name(r['name'], lang))
            else:
                region_info = region_name(r['name'], lang)
        else:
            region_info = t(lang, 'unknown_word')
    else:
        region_info = t(lang, 'region_not_set')

    YOQ = t(lang, 'not_specified')
    text = t(lang, 'seller_profile_body',
             shop=user.get('shop_name') or YOQ,
             address=user.get('shop_address') or YOQ,
             landmark=user.get('shop_landmark') or YOQ,
             region=region_info,
             wd=user.get('working_days') or YOQ,
             wh=user.get('working_hours') or YOQ,
             tg=user.get('telegram_username') or YOQ,
             phone=user.get('phone_number') or YOQ,
             card=card_info, rating=f"{avg_rating:.1f}",
             date=fmt_datetime(user['created_at']))

    if update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard))
    else:
        await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard))


async def product_ask_seller(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xaridor sotuvchiga buyurtmasiz savol yubora oladi."""
    query = update.callback_query
    await query.answer()

    product_id = int(query.data.split("_")[1])
    lang = get_lang(update, context)
    product = db.get_product_basic(product_id)

    if not product:
        await query.answer(t(lang, 'product_not_found'), show_alert=True)
        return

    seller = db.get_user_by_id(product['seller_id'])
    if not seller or not seller.get('telegram_id'):
        await query.answer(t(lang, 'contact_seller_unavailable'), show_alert=True)
        return

    buyer = db.get_user_by_telegram_id(update.effective_user.id)
    if buyer and buyer['id'] == product['seller_id']:
        await query.answer(t(lang, 'this_is_your_product'), show_alert=True)
        return

    # Xaridorga sotuvchi ma'lumotlarini ko'rsatamiz
    seller_phone = seller.get('phone_number') or '—'
    seller_tg = seller.get('telegram_username')
    seller_tg_text = f"@{seller_tg}" if seller_tg else t(lang, 'tg_not_shown')
    shop_name = html.escape(seller.get('shop_name') or seller.get('name') or t(lang, 'seller_word'))

    contact_text = t(lang, 'contact_seller_text',
                     shop=shop_name, phone=seller_phone, tg=seller_tg_text,
                     pname=html.escape(product.get('name') or ''))

    kb = []
    if seller_tg:
        kb.append([InlineKeyboardButton(
            t(lang, 'btn_write_telegram'),
            url=f"https://t.me/{seller_tg.replace('@', '')}"
        )])
    kb.append([InlineKeyboardButton(
        t(lang, 'btn_order'),
        callback_data=f"order_{product_id}"
    )])
    kb.append([InlineKeyboardButton(t(lang, 'back'), callback_data=f"prod_{product_id}")])

    await context.bot.send_message(
        chat_id=update.effective_chat.id,
        text=contact_text,
        reply_markup=InlineKeyboardMarkup(kb),
        parse_mode='HTML'
    )

    # Sotuvchiga ham xabar — kim qiziqdi (sotuvchi tilida)
    try:
        slang = get_user_lang(seller)
        buyer_tg = f"@{buyer['telegram_username']}" if buyer and buyer.get('telegram_username') and buyer.get('telegram_username', '').startswith(('@', 'a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', 'A')) else "—"
        buyer_phone = buyer.get('phone_number') or '—' if buyer else '—'
        buyer_name = html.escape(buyer.get('name') or t(slang, 'anonymous')) if buyer else t(slang, 'anonymous')
        await context.bot.send_message(
            chat_id=seller['telegram_id'],
            text=t(slang, 'seller_interest_notify',
                   pname=html.escape(product.get('name') or ''),
                   buyer=buyer_name, phone=buyer_phone, tg=buyer_tg),
            parse_mode='HTML'
        )
    except Exception as e:
        logging.error(f"product_ask_seller sotuvchi xabari: {e}")


async def contact_admin_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Foydalanuvchi admin bilan bog'lanmoqchi."""
    query = update.callback_query
    await query.answer()
    await query.edit_message_text(
        T(update, context, 'contact_admin_prompt'),
        parse_mode='HTML'
    )
    return CONTACT_ADMIN_MSG


async def contact_admin_send(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Foydalanuvchi xabarini adminga yuboradi."""
    user = db.get_user_by_telegram_id(update.effective_user.id)
    text = update.message.text.strip()

    role = user.get('role', 'buyer') if user else 'buyer'
    role_label = {'buyer': '🛒 Xaridor', 'seller': '🏪 Sotuvchi', 'admin': '🔧 Admin'}.get(role, role)
    name = html.escape(user.get('name') or 'Anonim') if user else 'Anonim'
    phone = user.get('phone_number') or '—' if user else '—'
    tg_id = update.effective_user.id

    try:
        await context.bot.send_message(
            chat_id=ADMIN_ID,
            text=(
                f"📨 <b>Foydalanuvchidan xabar</b>\n\n"
                f"{role_label}: {name}\n"
                f"📞 {phone}\n"
                f"🆔 Telegram ID: <code>{tg_id}</code>\n\n"
                f"💬 Xabar:\n{html.escape(text)}\n\n"
                f"<i>Javob berish uchun: /reply {tg_id} [matn]</i>"
            ),
            parse_mode='HTML'
        )
        await update.message.reply_text(T(update, context, 'contact_admin_sent'))
    except Exception as e:
        logging.error(f"contact_admin_send xatosi: {e}")
        await update.message.reply_text(T(update, context, 'admin_msg_failed'))

    # Panelga qaytamiz
    if update.effective_user:
        u = db.get_user_by_telegram_id(update.effective_user.id)
        if u:
            if get_active_mode(u, context) == 'seller':
                await seller_panel(update, context)
            else:
                await buyer_panel(update, context)
    return ConversationHandler.END


async def admin_reply_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/reply {user_id} {matn} — admin foydalanuvchiga javob beradi."""
    if update.effective_user.id != ADMIN_ID:
        return

    lang = get_lang(update, context)
    args = context.args  # [user_id, ...matn...]
    if not args or len(args) < 2:
        await update.message.reply_text(t(lang, 'admin_reply_usage'))
        return

    try:
        user_id = int(args[0])
        reply_text = ' '.join(args[1:])
        target = db.get_user_by_id(user_id)  # ehtimol topilmaydi (telegram_id bo'lishi mumkin)
        tlang = get_user_lang(target) if target else DEFAULT_LANG
        await context.bot.send_message(
            chat_id=user_id,
            text=t(tlang, 'admin_reply_prefix', text=html.escape(reply_text)),
            parse_mode='HTML'
        )
        await update.message.reply_text(t(lang, 'admin_reply_sent', uid=user_id))
    except ValueError:
        await update.message.reply_text(t(lang, 'admin_reply_format'))
    except Exception as e:
        await update.message.reply_text(t(lang, 'error_generic', e=e))


async def edit_seller_region(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Sotuvchi o'z hududini tanlaydi."""
    query = update.callback_query
    await query.answer()

    lang = get_lang(update, context)
    regions = db.get_regions(parent_id=None)
    kb = []
    for r in regions:
        kb.append([InlineKeyboardButton(region_name(r['name'], lang), callback_data=f"sregset_{r['id']}")])
    kb.append([InlineKeyboardButton(t(lang, 'btn_cancel_edit'), callback_data="seller_profile")])

    await query.edit_message_text(
        t(lang, 'seller_region_ask'),
        reply_markup=InlineKeyboardMarkup(kb)
    )


async def seller_region_set(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Viloyat tanlandi — tuman bor bo'lsa ko'rsatamiz, yo'q bo'lsa saqlaymiz."""
    query = update.callback_query
    await query.answer()

    lang = get_lang(update, context)
    region_id = int(query.data.split("_")[1])
    districts = db.get_regions(parent_id=region_id)

    if districts:
        # Tuman tanlash
        kb = []
        for d in districts:
            kb.append([InlineKeyboardButton(region_name(d['name'], lang), callback_data=f"sregdist_{d['id']}")])
        region = db.get_region_by_id(region_id)
        kb.append([InlineKeyboardButton(
            t(lang, 'btn_whole_region', name=region_name(region['name'], lang)), callback_data=f"sregdist_0_{region_id}"
        )])
        kb.append([InlineKeyboardButton(t(lang, 'back'), callback_data="edit_seller_region")])
        await query.edit_message_text(
            t(lang, 'region_pick_district', name=region_name(region['name'], lang)),
            reply_markup=InlineKeyboardMarkup(kb)
        )
    else:
        # Tumansiz — to'g'ridan-to'g'ri saqlaymiz
        user = db.get_user_by_telegram_id(update.effective_user.id)
        db.update_user(user['id'], region_id=region_id)
        r = db.get_region_by_id(region_id)
        await query.answer(t(lang, 'region_saved_toast', name=region_name(r['name'], lang)), show_alert=True)
        await seller_profile(update, context)


async def seller_district_set(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Tuman tanlandi — saqlaydi."""
    query = update.callback_query
    await query.answer()

    lang = get_lang(update, context)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    data = query.data.replace("sregdist_", "")

    if "_" in data:
        # "0_{region_id}" — butun viloyat
        region_id = int(data.split("_")[1])
    else:
        region_id = int(data)

    db.update_user(user['id'], region_id=region_id)
    r = db.get_region_by_id(region_id)
    await query.answer(t(lang, 'region_saved_toast', name=region_name(r['name'], lang)), show_alert=True)
    await seller_profile(update, context)


async def edit_buyer_name_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_text(T(update, context, 'ask_new_name'))
    return EDIT_PROFILE_NAME


async def edit_buyer_name_submit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    lang = get_lang(update, context)
    name = validate_fullname(update.message.text, max_len=60)
    if not name:
        await update.message.reply_text(t(lang, 'name_invalid'))
        return EDIT_PROFILE_NAME

    user = db.get_user_by_telegram_id(update.effective_user.id)
    db.update_user(user['id'], name=name)
    await update.message.reply_text(t(lang, 'name_updated_excl'))
    await buyer_panel(update, context)
    return ConversationHandler.END


async def edit_buyer_phone_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)

    await query.edit_message_text(t(lang, 'ask_new_phone'))
    # edit_message_text reply_markup qabul qilmaydi ReplyKeyboard uchun,
    # shuning uchun alohida xabar yuboramiz
    await query.message.reply_text(
        t(lang, 'press_phone_btn'),
        reply_markup=ReplyKeyboardMarkup([[KeyboardButton(t(lang, 'phone_button'), request_contact=True)]],
                                         resize_keyboard=True, one_time_keyboard=True)
    )
    return EDIT_PROFILE_PHONE


async def edit_buyer_phone_submit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    lang = get_lang(update, context)
    if update.message.contact:
        phone = update.message.contact.phone_number
    else:
        phone = update.message.text

    normalized = normalize_phone(phone)
    if not normalized:
        await update.message.reply_text(t(lang, 'phone_invalid_2'))
        return EDIT_PROFILE_PHONE

    user = db.get_user_by_telegram_id(update.effective_user.id)
    db.update_user(user['id'], phone_number=normalized)
    await update.message.reply_text(t(lang, 'phone_updated'), reply_markup=ReplyKeyboardRemove())
    await buyer_panel(update, context)
    return ConversationHandler.END


async def edit_seller_field_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    data = query.data  # masalan: "edit_shop_name", "edit_working_days", "edit_telegram"

    # BUG FIX #4: field nomini to'g'ri ajratish
    if data == "edit_telegram":
        field = "telegram"
    elif data.startswith("edit_shop_"):
        field = data[len("edit_"):]   # "shop_name", "shop_address", "shop_landmark"
    elif data.startswith("edit_working_"):
        field = data[len("edit_"):]   # "working_days", "working_hours"
    else:
        field = data[len("edit_"):]

    context.user_data['editing_field'] = field
    lang = get_lang(update, context)

    field_labels = {
        'shop_name': t(lang, 'efl_shop_name'),
        'shop_address': t(lang, 'efl_shop_address'),
        'shop_landmark': t(lang, 'efl_shop_landmark'),
        'working_days': t(lang, 'efl_working_days'),
        'working_hours': t(lang, 'efl_working_hours'),
        'telegram': t(lang, 'efl_telegram'),
    }

    label = field_labels.get(field, field)

    if field == 'shop_address':
        await query.edit_message_text(t(lang, 'edit_field_ask_addr', label=label))
        await query.message.reply_text(
            t(lang, 'shop_location_ask'),
            reply_markup=ReplyKeyboardMarkup(
                [[KeyboardButton(t(lang, 'send_location_button'), request_location=True)]],
                resize_keyboard=True, one_time_keyboard=True
            )
        )
        return EDIT_SHOP_ADDRESS
    else:
        await query.edit_message_text(t(lang, 'edit_field_ask', label=label))

    state_map = {
        'shop_name': EDIT_SHOP_NAME,
        'shop_landmark': EDIT_SHOP_LANDMARK,
        'working_days': EDIT_WORKING_DAYS,
        'working_hours': EDIT_WORKING_HOURS,
        'telegram': EDIT_TELEGRAM_USERNAME,
    }
    return state_map.get(field, ConversationHandler.END)


async def edit_seller_field_submit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    field = context.user_data.get('editing_field')
    user = db.get_user_by_telegram_id(update.effective_user.id)

    # shop_address — ikki bosqichli: 1/2 lokatsiya (xarita), 2/2 matn
    if field == 'shop_address':
        if update.message.location:
            lat = update.message.location.latitude
            lon = update.message.location.longitude
            addr_text = await resolve_shop_address(lat, lon)
            db.update_user(user['id'], shop_lat=lat, shop_lon=lon, shop_address=addr_text)
            if addr_text:
                await update.message.reply_text(T(update, context, 'address_detected', address=addr_text))
        else:
            # "-" yoki boshqa matn — lokatsiyani tozalaymiz
            db.update_user(user['id'], shop_lat=None, shop_lon=None)
        await update.message.reply_text(
            T(update, context, 'shop_address_text_ask'),
            reply_markup=ReplyKeyboardRemove()
        )
        return EDIT_SHOP_ADDRESS_TEXT

    value = update.message.text.strip()
    field_map = {
        'shop_name': 'shop_name',
        'shop_landmark': 'shop_landmark',
        'working_days': 'working_days',
        'working_hours': 'working_hours',
        'telegram': 'telegram_username',
    }
    if field in field_map:
        db.update_user(user['id'], **{field_map[field]: value})

    await update.message.reply_text(T(update, context, 'info_updated'), reply_markup=ReplyKeyboardRemove())
    await seller_panel(update, context)
    return ConversationHandler.END


async def edit_shop_address_text_submit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Manzilni tahrirlashning 2/2-bosqichi: matn. "-" bo'lsa geocode manzil saqlanadi."""
    user = db.get_user_by_telegram_id(update.effective_user.id)
    text = update.message.text.strip()
    if text != '-':
        if len(text) < 5 or len(text) > 200:
            await update.message.reply_text(T(update, context, 'address_invalid'))
            return EDIT_SHOP_ADDRESS_TEXT
        db.update_user(user['id'], shop_address=text)

    await update.message.reply_text(T(update, context, 'info_updated'), reply_markup=ReplyKeyboardRemove())
    await seller_panel(update, context)
    return ConversationHandler.END


async def edit_card_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Sotuvchi karta ma'lumotini qo'shish/tahrirlash."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("🟦 Uzcard",     callback_data="card_type_uzcard")],
        [InlineKeyboardButton("🟩 Humo",       callback_data="card_type_humo")],
        [InlineKeyboardButton("🔵 Visa",       callback_data="card_type_visa")],
        [InlineKeyboardButton("🔴 Mastercard", callback_data="card_type_mastercard")],
        [InlineKeyboardButton(t(lang, 'btn_card_remove'), callback_data="card_type_remove")],
        [InlineKeyboardButton(t(lang, 'btn_cancel_edit'), callback_data="seller_profile")],
    ])
    await query.edit_message_text(
        t(lang, 'card_menu'),
        reply_markup=kb,
        parse_mode='HTML'
    )
    return EDIT_CARD_TYPE


async def edit_card_type(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    lang = get_lang(update, context)
    if query.data == "card_type_remove":
        user = db.get_user_by_telegram_id(update.effective_user.id)
        db.update_user(user['id'], card_number=None, card_owner=None, card_type=None)
        await query.edit_message_text(t(lang, 'card_removed'))
        await seller_profile(update, context)
        return ConversationHandler.END

    card_type = query.data.replace("card_type_", "")
    context.user_data['new_card_type'] = card_type
    await query.edit_message_text(
        t(lang, 'card_number_ask'),
        parse_mode='HTML'
    )
    return EDIT_CARD_NUMBER


async def edit_card_number(update: Update, context: ContextTypes.DEFAULT_TYPE):
    lang = get_lang(update, context)
    raw = update.message.text.replace(" ", "").replace("-", "")

    if not raw.isdigit() or len(raw) not in (16, 18, 20):
        await update.message.reply_text(
            t(lang, 'card_number_invalid'),
            parse_mode='HTML'
        )
        return EDIT_CARD_NUMBER

    context.user_data['new_card_number'] = raw
    await update.message.reply_text(
        t(lang, 'card_owner_ask'),
        parse_mode='HTML'
    )
    return EDIT_CARD_OWNER


async def edit_card_owner(update: Update, context: ContextTypes.DEFAULT_TYPE):
    lang = get_lang(update, context)
    owner = update.message.text.strip().upper()

    if len(owner) < 3 or len(owner) > 50:
        await update.message.reply_text(t(lang, 'card_owner_invalid'))
        return EDIT_CARD_OWNER

    user = db.get_user_by_telegram_id(update.effective_user.id)
    card_type = context.user_data.pop('new_card_type', '')
    card_number = context.user_data.pop('new_card_number', '')

    db.update_user(
        user['id'],
        card_type=card_type,
        card_number=card_number,
        card_owner=owner
    )

    # Tasdiqlash
    cnum = card_number
    masked = f"{cnum[:4]} **** **** {cnum[-4:]}"
    ctype_label = CARD_TYPE_LABELS.get(card_type, card_type)

    await update.message.reply_text(
        t(lang, 'card_saved', ctype=ctype_label, masked=masked, owner=owner),
        parse_mode='HTML'
    )
    await seller_profile(update, context)
    return ConversationHandler.END


async def seller_order_detail(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    order_id = int(query.data.split("_")[2])

    order = db.get_order_by_id(order_id)
    lang = get_lang(update, context)

    if not order:
        await query.edit_message_text(t(lang, 'order_not_found'))
        return

    # XAVFSIZLIK: bu kartada xaridor telefoni/manzili (PII) bor — faqat egasi (sotuvchi),
    # mahsulotni joylagan xodim yoki admin ko'rsin
    actor = db.get_user_by_telegram_id(update.effective_user.id) if update.effective_user else None
    is_admin_id = update.effective_user and update.effective_user.id == ADMIN_ID
    if not (_order_actor_role(actor, order) or is_admin_id):
        logging.warning(
            f"Ruxsatsiz buyurtma ko'rishga urinish: user_tg={getattr(update.effective_user, 'id', None)} "
            f"order_id={order_id} (egasi seller_id={order.get('seller_id')})"
        )
        await query.edit_message_text(t(lang, 'not_your_order_plain'))
        return

    dlv = order.get('delivery_type', 'delivery')
    keyboard = []
    if order['status'] == 'pending':
        keyboard.append([InlineKeyboardButton(t(lang, 'btn_confirm'), callback_data=f"confirm_order_{order_id}")])
        keyboard.append([InlineKeyboardButton(t(lang, 'btn_reject'), callback_data=f"cancel_order_{order_id}")])
    elif order['status'] == 'confirmed':
        if dlv == 'delivery':
            keyboard.append([InlineKeyboardButton(
                t(lang, 'btn_delivered'), callback_data=f"deliver_order_{order_id}"
            )])
        else:
            # Pickup: xaridor o'zi "Oldim" bosadi, lekin sotuvchi ham tasdiqlashi mumkin
            keyboard.append([InlineKeyboardButton(
                t(lang, 'btn_buyer_received'), callback_data=f"deliver_order_{order_id}"
            )])
        # Tasdiqlangan shartnomani bekor qilish oqimi (sotuvchi tomoni)
        cstate = order.get('cancel_state') or ''
        if not cstate:
            keyboard.append([InlineKeyboardButton(
                t(lang, 'btn_request_cancel'), callback_data=f"ccl_req_{order_id}"
            )])
        elif cstate == 'requested' and order.get('cancel_by') == 'buyer':
            # Xaridor bekor qilishni so'ragan — sotuvchi javob beradi
            keyboard.append([InlineKeyboardButton(t(lang, 'btn_cancel_agree'), callback_data=f"cclagree_{order_id}")])
            keyboard.append([InlineKeyboardButton(t(lang, 'btn_cancel_deny'), callback_data=f"ccldeny_{order_id}")])
        elif cstate == 'disputed':
            keyboard.append([InlineKeyboardButton(t(lang, 'btn_dispute_pending'), callback_data="noop")])
    # Kuryerga uzatish — yetkazib berish buyurtmasi hali yopilmagan bo'lsa
    if dlv == 'delivery' and order['status'] in ('pending', 'confirmed'):
        keyboard.append([InlineKeyboardButton(
            t(lang, 'btn_forward_courier'), callback_data=f"crfwd_{order_id}"
        )])
    keyboard.append([InlineKeyboardButton(t(lang, 'btn_correspondence'), callback_data=f"msgs_{order_id}")])
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="seller_orders")])

    pay_method = order.get('payment_method') or 'cash'
    pay_lbl = pay_label(pay_method, lang)

    # P2P bo'lsa — sotuvchiga karta raqamini eslatamiz (do'kon payment_mode bo'yicha)
    pay_note = ""
    if pay_method == 'p2p':
        _prod = db.get_product_basic(order.get('product_id')) if order.get('product_id') else None
        card = resolve_payment_card(order.get('seller_id'),
                                    _prod.get('created_by') if _prod else None)
        if card and card.get('card_number'):
            cnum = card['card_number']
            masked = f"{cnum[:4]} **** **** {cnum[-4:]}"
            ctype = CARD_TYPE_LABELS.get(card.get('card_type', ''), '💳')
            pay_note = t(lang, 'p2p_your_card', ctype=ctype, masked=masked)
        else:
            pay_note = t(lang, 'p2p_no_card')

    dlv_type = dlv_label(dlv, lang)

    # Yetkazib berish manzili — kuryer uchun (faqat delivery buyurtmalarda)
    delivery_block = ""
    if dlv == 'delivery':
        b_lat = order.get('buyer_lat')
        b_lon = order.get('buyer_lon')
        addr_txt = human_address(order.get('delivery_address'))
        parts = []
        if addr_txt:
            parts.append(t(lang, 'seller_order_addr', addr=html.escape(addr_txt)))
        if b_lat is not None and b_lon is not None:
            parts.append(t(lang, 'seller_order_map',
                           url=f"https://www.google.com/maps/search/?api=1&query={b_lat},{b_lon}"))
            s_lat = order.get('shop_lat')
            s_lon = order.get('shop_lon')
            if s_lat is not None and s_lon is not None:
                d = haversine_km(s_lat, s_lon, b_lat, b_lon)
                if d is not None:
                    parts.append(t(lang, 'seller_dist_from_shop', km=f"{d:.1f}"))
        if not parts:
            parts.append(t(lang, 'addr_not_shown'))
        delivery_block = "".join(parts)

    status_disp = status_label(order['status'], lang)
    if order['status'] == 'cancelled' and order.get('cancel_reason'):
        status_disp += t(lang, 'cancel_note_reason',
                         reason=cancel_reason_display(order.get('cancel_reason'), lang))

    body = t(lang, 'seller_order_body',
          oid=fmt_order_id(order['id']), pname=html.escape(order.get('product_name') or ''),
          qty=order['quantity'], total=fmt_price(order['total_price']),
          status=status_disp, dlv=dlv_type,
          pay=pay_lbl, paynote=pay_note,
          buyer=html.escape(order.get('buyer_name') or ''),
          phone=fmt_phone(order.get('buyer_phone')),
          delivery=delivery_block, date=fmt_datetime(order.get('created_at')))
    _pb = _progress_badge(lang, order)
    if _pb:
        body += "\n" + _pb
    # Pending — sotuvchi buyurtmani Buyurtmalar orqali ochsa ham jonli teskari sanoq ko'rinsin
    if order['status'] == 'pending':
        _cd = _countdown_line(lang, _order_deadline(order))
        if _cd:
            body += t(lang, 'countdown_sep') + _cd
    _sb = _settlement_badge(lang, order)
    if _sb:
        body += "\n" + _sb
    await query.edit_message_text(
        body,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode='HTML'
    )


async def seller_forward_courier(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Sotuvchi buyurtma manzilini kuryerga uzatadi.
    Bot sotuvchiga forward qilinadigan paket yuboradi:
    1) mijoz lokatsiyasi (pin),  2) buyurtma + manzil matni,  3) yo'riqnoma."""
    query = update.callback_query
    # Eslatma: button_handler bu callback uchun allaqachon query.answer() chaqirgan.
    chat_id = update.effective_chat.id
    lang = get_lang(update, context)

    try:
        order_id = int(query.data.split("_")[1])
    except (ValueError, IndexError):
        await context.bot.send_message(chat_id, t(lang, 'order_num_invalid'))
        return

    order = db.get_order_by_id(order_id)
    if not order:
        await context.bot.send_message(chat_id, t(lang, 'order_not_found'))
        return

    # Egalik tekshiruvi — faqat shu buyurtma sotuvchisi yoki admin
    seller_user = db.get_user_by_telegram_id(update.effective_user.id)
    is_owner = bool(seller_user and seller_user.get('id') == order.get('seller_id'))
    is_admin = (update.effective_user.id == ADMIN_ID) or (
        seller_user and seller_user.get('role') == 'admin'
    )
    if not (is_owner or is_admin):
        await context.bot.send_message(chat_id, t(lang, 'not_your_order_plain'))
        return

    b_lat = order.get('buyer_lat')
    b_lon = order.get('buyer_lon')
    s_lat = order.get('shop_lat')
    s_lon = order.get('shop_lon')

    pay_method = order.get('payment_method') or 'cash'
    pay_lbl = pay_label(pay_method, lang)
    addr_txt = human_address(order.get('delivery_address'))

    # 1) Mijoz lokatsiyasi (pin) — forward qilish uchun alohida xabar
    if b_lat is not None and b_lon is not None:
        try:
            await context.bot.send_location(chat_id=chat_id, latitude=b_lat, longitude=b_lon)
        except Exception as e:
            logging.warning(f"courier send_location xatosi: {e}")

    # 2) Buyurtma + manzil matni
    lines = [t(lang, 'courier_body',
               oid=fmt_order_id(order['id']),
               pname=html.escape(order.get('product_name') or ''),
               qty=order.get('quantity'), total=fmt_price(order.get('total_price')),
               pay=pay_lbl, buyer=html.escape(order.get('buyer_name') or ''),
               phone=fmt_phone(order.get('buyer_phone')))]
    if addr_txt:
        lines.append(t(lang, 'courier_addr', addr=html.escape(addr_txt)))
    if b_lat is not None and b_lon is not None:
        lines.append(t(lang, 'courier_map',
                       url=f"https://www.google.com/maps/search/?api=1&query={b_lat},{b_lon}"))
        if s_lat is not None and s_lon is not None:
            d = haversine_km(s_lat, s_lon, b_lat, b_lon)
            if d is not None:
                lines.append(t(lang, 'courier_dist', km=f"{d:.1f}"))
            lines.append(t(lang, 'courier_route',
                           url=f"https://www.google.com/maps/dir/?api=1&origin={s_lat},{s_lon}&destination={b_lat},{b_lon}"))
    if not addr_txt and (b_lat is None or b_lon is None):
        lines.append(t(lang, 'courier_no_addr'))

    await context.bot.send_message(
        chat_id,
        "\n".join(lines),
        parse_mode='HTML',
        disable_web_page_preview=True,
    )

    # 3) Yo'riqnoma
    await context.bot.send_message(
        chat_id,
        t(lang, 'courier_instructions'),
        parse_mode='HTML',
    )


def _order_actor_role(actor, order):
    """Buyurtma bo'yicha actor kim ekanini aniqlaydi.
    Qaytaradi: ('owner'|'admin'|'staff'|None). 'staff' — mahsulotni joylagan xodim
    (buyurtma tasdiqlash ruxsati va faol holatda)."""
    if not actor or not order:
        return None
    if actor.get('role') == 'admin':
        return 'admin'
    if actor.get('id') == order.get('seller_id'):
        return 'owner'
    prod = db.get_product_basic(order.get('product_id')) if order.get('product_id') else None
    if prod and prod.get('created_by') == actor.get('id'):
        staff_rec = db.get_staff_by_user(actor['id'])
        if staff_rec and staff_rec.get('perm_confirm_orders', 1) and staff_rec.get('is_active', 1):
            return 'staff'
    return None


async def _ensure_order_seller(update: Update, context: ContextTypes.DEFAULT_TYPE, order_id):
    """Joriy foydalanuvchi shu buyurtmaning sotuvchisi (yoki admin) ekanligini tasdiqlaydi.
    To'g'ri bo'lsa True; aks holda foydalanuvchini ogohlantirib False qaytaradi.
    Buyurtma bo'yicha amallar va xaridor PII'sini begonalardan himoyalaydi."""
    lang = get_lang(update, context)
    chat_id = update.effective_chat.id if update.effective_chat else None
    order = db.get_order_by_id(order_id)
    if not order:
        if chat_id:
            await context.bot.send_message(chat_id, t(lang, 'order_not_found'))
        return False
    actor = db.get_user_by_telegram_id(update.effective_user.id) if update.effective_user else None
    is_admin_id = update.effective_user and update.effective_user.id == ADMIN_ID
    role = _order_actor_role(actor, order)
    if not (role or is_admin_id):
        logging.warning(
            f"Ruxsatsiz buyurtma amaliga urinish: user_tg={getattr(update.effective_user, 'id', None)} "
            f"order_id={order_id} (egasi seller_id={order.get('seller_id')})"
        )
        if chat_id:
            await context.bot.send_message(chat_id, t(lang, 'not_your_order_plain'))
        return False
    return True


# ============================================================
# TO'LOV HOLATI (settlement) — berishda: To'liq to'landi / Qarzga / Bo'lib + qarz daftari
# ============================================================
def _settlement_menu_kb(lang, scope, key):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(t(lang, 'setl_paid_btn'), callback_data=f"setl_paid_{scope}_{key}")],
        [InlineKeyboardButton(t(lang, 'setl_debt_btn'), callback_data=f"setl_debt_{scope}_{key}")],
        [InlineKeyboardButton(t(lang, 'setl_inst_btn'), callback_data=f"setl_inst_{scope}_{key}")],
    ])


def _settle_amount_kb(lang):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(t(lang, 'setl_amt_zero'), callback_data="setlamt_zero")],
        [InlineKeyboardButton(t(lang, 'setl_amt_half'), callback_data="setlamt_half")],
        [InlineKeyboardButton(t(lang, 'setl_amt_custom'), callback_data="setlamt_custom")],
    ])


async def _ask_settlement(update, context, scope, key, total):
    """«Berildi» bosilganda — to'lov holatini so'raydi (delivered hali qo'yilmaydi)."""
    query = update.callback_query
    lang = get_lang(update, context)
    context.user_data['settle'] = {'scope': scope, 'key': str(key), 'total': float(total or 0)}
    context.user_data.pop('awaiting_settle_amount', None)
    txt = t(lang, 'setl_ask', total=fmt_price(total or 0))
    try:
        await query.edit_message_text(txt, reply_markup=_settlement_menu_kb(lang, scope, key), parse_mode='HTML')
    except Exception:
        await context.bot.send_message(chat_id=update.effective_chat.id, text=txt,
                                       reply_markup=_settlement_menu_kb(lang, scope, key), parse_mode='HTML')


async def settle_choice(update, context):
    """setl_paid_/setl_debt_/setl_inst_ — to'lov holati tanlovi."""
    query = update.callback_query
    lang = get_lang(update, context)
    sd = context.user_data.get('settle')
    if not sd:
        await query.answer(t(lang, 'setl_expired'), show_alert=True)
        return
    kind = query.data.split('_')[1]  # paid / debt / inst
    await query.answer()
    if kind == 'paid':
        await _finalize_settlement(update, context, settlement_type='paid', paid=sd['total'])
        return
    sd['type'] = 'debt' if kind == 'debt' else 'installment'
    context.user_data['settle'] = sd
    await query.edit_message_text(
        t(lang, 'setl_amount_ask', total=fmt_price(sd['total'])),
        reply_markup=_settle_amount_kb(lang), parse_mode='HTML')


async def settle_amount_choice(update, context):
    """setlamt_zero/half/custom — qarz/bo'lib uchun hozir to'langan summa."""
    query = update.callback_query
    lang = get_lang(update, context)
    sd = context.user_data.get('settle')
    if not sd:
        await query.answer(t(lang, 'setl_expired'), show_alert=True)
        return
    choice = query.data.replace('setlamt_', '')
    if choice == 'custom':
        await query.answer()
        context.user_data['awaiting_settle_amount'] = True
        await query.edit_message_text(t(lang, 'setl_custom_ask', total=fmt_price(sd['total'])), parse_mode='HTML')
        return
    await query.answer()
    paid = 0.0 if choice == 'zero' else round(sd['total'] / 2.0)
    await _finalize_settlement(update, context, settlement_type=sd.get('type', 'debt'), paid=paid)


async def settle_custom_amount_input(update, context):
    """Sotuvchi qo'lda to'langan summani yozdi (custom)."""
    lang = get_lang(update, context)
    sd = context.user_data.get('settle')
    context.user_data.pop('awaiting_settle_amount', None)
    if not sd:
        await update.message.reply_text(t(lang, 'setl_expired'))
        return
    raw = (update.message.text or '').strip().replace(' ', '').replace(' ', '')
    try:
        paid = float(raw)
    except ValueError:
        context.user_data['awaiting_settle_amount'] = True
        await update.message.reply_text(t(lang, 'setl_amount_invalid'))
        return
    await _finalize_settlement(update, context, settlement_type=sd.get('type', 'debt'),
                               paid=paid, via_message=True)


async def _finalize_settlement(update, context, settlement_type, paid, via_message=False):
    """Berishni yakunlaydi: delivered + settlement saqlash + xaridorga bildirishnoma."""
    lang = get_lang(update, context)
    sd = context.user_data.get('settle') or {}
    scope = sd.get('scope', 'o')
    key = sd.get('key')
    total = float(sd.get('total') or 0)
    paid = max(0.0, min(float(paid or 0), total))
    due = round(total - paid, 2)
    eff_type = 'paid' if due <= 0 else settlement_type
    chat_id = update.effective_chat.id

    async def _already_done():
        # Buyurtma shu lahzada ilova (yoki boshqa qurilma) orqali berilgan — qayta
        # settlement/xabar bajarmaymiz, sotuvchini ortga qaytaramiz.
        context.user_data.pop('settle', None)
        back_cb = f"seller_gorder_{key}" if scope == 'g' else f"seller_order_{key}"
        kb = InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data=back_cb)]])
        msg = t(lang, 'setl_already_done')
        if via_message:
            await update.message.reply_text(msg, reply_markup=kb)
        else:
            try:
                await update.callback_query.edit_message_text(msg, reply_markup=kb)
            except Exception:
                await context.bot.send_message(chat_id=chat_id, text=msg, reply_markup=kb)

    if scope == 'g':
        orders = db.get_orders_in_group(key)
        # ATOMIK: faqat 'confirmed' sub-buyurtmalarni 'delivered'ga o'tkazamiz. App deliver
        # bilan bir vaqtda yoki ikki marta bo'lsa — faqat YUTGAN tomon settlement/xabar
        # bajaradi (qarama-qarshi to'lov holati + ikki marta "berildi" xabarining oldini).
        won = [o for o in orders if db.transition_order_status(o['id'], 'delivered', 'confirmed')]
        if not won:
            await _already_done()
            return
        db.set_group_settlement(key, eff_type, paid, due)
        rep = orders[0] if orders else None
        if rep:
            buyer_tg = rep.get('buyer_tg')
            buyer = db.get_user_by_id(rep['buyer_id'])
            seller = db.get_user_by_id(rep['seller_id'])
            disp = fmt_order_id(int(key))
            is_pickup = rep.get('delivery_type') == 'pickup'
            if buyer_tg:
                blang = get_user_lang(buyer) if buyer else DEFAULT_LANG
                txt = t(blang, 'grp_delivered_pickup' if is_pickup else 'grp_delivered_delivery',
                        oid=disp, n=len(orders))
                if due > 0:
                    txt += t(blang, 'buyer_debt_notify',
                             shop=html.escape((seller.get('shop_name') or seller.get('name') or '') if seller else ''),
                             due=fmt_price(due))
                kb = InlineKeyboardMarkup([[InlineKeyboardButton(
                    t(blang, 'btn_leave_rating'), callback_data=f"order_rate_{rep['id']}")]])
                try:
                    await context.bot.send_message(chat_id=buyer_tg, text=txt, reply_markup=kb, parse_mode='HTML')
                except Exception:
                    pass
    else:
        order = db.get_order_by_id(key)
        if not order:
            context.user_data.pop('settle', None)
            return
        # ATOMIK: faqat 'confirmed' holatdan o'tkazamiz (app deliver bilan poyga himoyasi).
        if not db.transition_order_status(int(key), 'delivered', 'confirmed'):
            await _already_done()
            return
        db.set_order_settlement(int(key), eff_type, paid, due)
        buyer_tg = order.get('buyer_tg')
        buyer = db.get_user_by_id(order['buyer_id'])
        seller = db.get_user_by_id(order['seller_id'])
        disp = fmt_order_id(int(key))
        pname = html.escape(order.get('product_name') or '')
        is_pickup = order.get('delivery_type') == 'pickup'
        if buyer_tg:
            blang = get_user_lang(buyer) if buyer else DEFAULT_LANG
            txt = t(blang, 'order_delivered_pickup' if is_pickup else 'order_delivered_delivery',
                    oid=disp, pname=pname)
            if due > 0:
                txt += t(blang, 'buyer_debt_notify',
                         shop=html.escape((seller.get('shop_name') or seller.get('name') or '') if seller else ''),
                         due=fmt_price(due))
            kb = InlineKeyboardMarkup([[InlineKeyboardButton(
                t(blang, 'btn_leave_rating'), callback_data=f"order_rate_{key}")]])
            try:
                await context.bot.send_message(chat_id=buyer_tg, text=txt, reply_markup=kb, parse_mode='HTML')
            except Exception:
                pass

    context.user_data.pop('settle', None)
    if eff_type == 'paid':
        conf = t(lang, 'setl_done_paid')
    else:
        conf = t(lang, 'setl_done_debt', paid=fmt_price(paid), due=fmt_price(due))
    back_cb = f"seller_gorder_{key}" if scope == 'g' else f"seller_order_{key}"
    kb = InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data=back_cb)]])
    if via_message:
        await update.message.reply_text(conf, reply_markup=kb, parse_mode='HTML')
    else:
        try:
            await update.callback_query.edit_message_text(conf, reply_markup=kb, parse_mode='HTML')
        except Exception:
            await context.bot.send_message(chat_id=chat_id, text=conf, reply_markup=kb, parse_mode='HTML')


def _progress_badge(lang, order):
    """Yakunlanmagan (jarayondagi) buyurtma uchun belgi. delivered/cancelled da bo'sh
    qaytadi — jarayon tugagach belgi yo'qoladi. Xaridor «oldim» bosgan bo'lsa, sotuvchiga
    to'lovni yakunlash kerakligini alohida ta'kidlaymiz."""
    status = order.get('status')
    if status not in ('pending', 'confirmed'):
        return ""
    if status == 'confirmed' and order.get('buyer_received'):
        return t(lang, 'badge_awaiting_settlement')
    return t(lang, 'badge_in_progress')


def _settlement_badge(lang, order):
    """Order detalida ko'rsatiladigan to'lov holati satri (delivered buyurtmalar uchun)."""
    st = order.get('settlement_type')
    if not st:
        return ""
    due = float(order.get('amount_due') or 0)
    paid = float(order.get('amount_paid') or 0)
    if st == 'paid' or due <= 0:
        return t(lang, 'badge_paid')
    label = t(lang, 'badge_debt') if st == 'debt' else t(lang, 'badge_installment')
    return t(lang, 'badge_due', label=label, due=fmt_price(due), paid=fmt_price(paid))


# ===== QARZ DAFTARI EKRANI (sotuvchi + xaridor) =====
async def seller_debts(update, context):
    """Sotuvchi panelidagi «Qarzlar» — kim qancha qarzdor (jamlangan)."""
    query = update.callback_query
    lang = get_lang(update, context)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    owner_id = db.resolve_owner_id(user['id'])
    debts = db.get_seller_open_debts(owner_id)
    total = db.get_seller_debt_total(owner_id)
    if not debts:
        kb = InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="seller_panel")]])
        text = t(lang, 'debts_empty')
        if query:
            await query.answer(); await query.edit_message_text(text, reply_markup=kb)
        else:
            await update.message.reply_text(text, reply_markup=kb)
        return
    lines = [t(lang, 'debts_title', total=fmt_price(total))]
    keyboard = []
    for d in debts:
        lines.append(t(lang, 'debts_buyer_line',
                       name=html.escape(d.get('buyer_name') or '—'),
                       due=fmt_price(d['total_due']), cnt=d['cnt']))
        keyboard.append([InlineKeyboardButton(
            t(lang, 'debts_buyer_btn', name=(d.get('buyer_name') or '—')[:18], due=fmt_price(d['total_due'])),
            callback_data=f"debtbuyer_{d['buyer_id']}")])
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="seller_panel")])
    text = "\n".join(lines)
    if query:
        await query.answer()
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')
    else:
        await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


async def debt_buyer_detail(update, context, buyer_id=None):
    """debtbuyer_{buyer_id} — bir xaridorning qarzli buyurtmalari + to'lov tugmalari."""
    query = update.callback_query
    lang = get_lang(update, context)
    if query and buyer_id is None:
        await query.answer()
        buyer_id = int(query.data.split('_')[1])
    user = db.get_user_by_telegram_id(update.effective_user.id)
    owner_id = db.resolve_owner_id(user['id'])
    orders = db.get_seller_debt_orders(owner_id, buyer_id)
    buyer = db.get_user_by_id(buyer_id)
    bname = html.escape((buyer.get('name') if buyer else '') or '—')
    if not orders:
        kb = InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="seller_debts")]])
        txt = t(lang, 'debts_buyer_clear', name=bname)
        if query:
            await query.edit_message_text(txt, reply_markup=kb, parse_mode='HTML')
        else:
            await update.message.reply_text(txt, reply_markup=kb, parse_mode='HTML')
        return
    total_due = sum(float(o['amount_due']) for o in orders)
    bphone = (buyer.get('phone_number') if buyer else '') or '—'
    lines = [t(lang, 'debts_buyer_header', name=bname, total=fmt_price(total_due),
               phone=fmt_phone(bphone), cnt=len(orders))]
    keyboard = []
    for o in orders:
        disp = fmt_order_id(int(o.get('order_group_id') or o['id']))
        kind = t(lang, 'badge_installment') if o.get('settlement_type') == 'installment' else t(lang, 'badge_debt')
        lines.append(t(lang, 'debts_order_row', oid=disp,
                       pname=html.escape(o.get('product_name') or '—'),
                       kind=kind,
                       total=fmt_price(o.get('total_price') or 0),
                       paid=fmt_price(o.get('amount_paid') or 0),
                       due=fmt_price(o['amount_due']),
                       date=fmt_datetime(o.get('created_at'))))
        keyboard.append([
            InlineKeyboardButton(t(lang, 'debt_pay_full_btn'), callback_data=f"debtpayfull_{o['id']}"),
            InlineKeyboardButton(t(lang, 'debt_pay_part_btn'), callback_data=f"debtpaypart_{o['id']}"),
        ])
    # Xaridor bilan bog'lanish (Telegram) — qarzni eslatish uchun
    if buyer and buyer.get('telegram_username'):
        keyboard.append([InlineKeyboardButton(
            t(lang, 'btn_tg_at', u=buyer['telegram_username']),
            url=f"https://t.me/{buyer['telegram_username']}")])
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="seller_debts")])
    text = "\n".join(lines)
    if query:
        await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')
    else:
        await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


async def _debt_order_guard(update, order_id):
    """Buyurtma shu sotuvchiникими — tekshiradi. order (dict) yoki None qaytaradi."""
    user = db.get_user_by_telegram_id(update.effective_user.id)
    owner_id = db.resolve_owner_id(user['id']) if user else None
    order = db.get_order_by_id(order_id)
    if not order or owner_id is None or order.get('seller_id') != owner_id:
        return None
    return order


async def debt_pay_full(update, context):
    """debtpayfull_{order_id} — qarzni to'liq yopadi."""
    query = update.callback_query
    lang = get_lang(update, context)
    order_id = int(query.data.split('_')[1])
    order = await _debt_order_guard(update, order_id)
    if not order:
        await query.answer(t(lang, 'not_your_order_toast'), show_alert=True)
        return
    due = float(order.get('amount_due') or 0)
    db.record_debt_payment(order_id, due)
    await query.answer(t(lang, 'debt_settled_toast'), show_alert=True)
    await _notify_buyer_debt_paid(context, order, due, remaining=0)
    await debt_buyer_detail(update, context, buyer_id=order['buyer_id'])


async def debt_pay_part_start(update, context):
    """debtpaypart_{order_id} — qisman to'lov summasini so'raydi."""
    query = update.callback_query
    lang = get_lang(update, context)
    order_id = int(query.data.split('_')[1])
    order = await _debt_order_guard(update, order_id)
    if not order:
        await query.answer(t(lang, 'not_your_order_toast'), show_alert=True)
        return
    await query.answer()
    context.user_data['awaiting_debt_payment'] = {
        'order_id': order_id, 'buyer_id': order['buyer_id'], 'due': float(order.get('amount_due') or 0)}
    await query.edit_message_text(t(lang, 'debt_part_ask', due=fmt_price(order.get('amount_due') or 0)),
                                  parse_mode='HTML')


async def debt_payment_input(update, context):
    """Qisman to'lov summasi (matn) — qarzga qo'shadi."""
    lang = get_lang(update, context)
    info = context.user_data.get('awaiting_debt_payment')
    context.user_data.pop('awaiting_debt_payment', None)
    if not info:
        return
    raw = (update.message.text or '').strip().replace(' ', '').replace(' ', '')
    try:
        pay = float(raw)
    except ValueError:
        context.user_data['awaiting_debt_payment'] = info
        await update.message.reply_text(t(lang, 'setl_amount_invalid'))
        return
    order = db.get_order_by_id(info['order_id'])
    if not order:
        return
    pay = max(0.0, min(pay, float(order.get('amount_due') or 0)))
    remaining = db.record_debt_payment(info['order_id'], pay)
    await _notify_buyer_debt_paid(context, order, pay, remaining=remaining or 0)
    kb = InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'),
                              callback_data=f"debtbuyer_{info['buyer_id']}")]])
    if (remaining or 0) <= 0:
        await update.message.reply_text(t(lang, 'debt_settled_msg'), reply_markup=kb, parse_mode='HTML')
    else:
        await update.message.reply_text(
            t(lang, 'debt_part_done', paid=fmt_price(pay), due=fmt_price(remaining)),
            reply_markup=kb, parse_mode='HTML')


async def _notify_buyer_debt_paid(context, order, amount, remaining):
    """Xaridorga qarz to'lovi qayd etilgani haqida xabar."""
    try:
        buyer = db.get_user_by_id(order['buyer_id'])
        if not buyer or not buyer.get('telegram_id'):
            return
        seller = db.get_user_by_id(order['seller_id'])
        blang = get_user_lang(buyer)
        shop = html.escape((seller.get('shop_name') or seller.get('name') or '') if seller else '')
        if (remaining or 0) <= 0:
            txt = t(blang, 'buyer_debt_cleared', shop=shop)
        else:
            txt = t(blang, 'buyer_debt_partial', shop=shop, paid=fmt_price(amount), due=fmt_price(remaining))
        await context.bot.send_message(chat_id=buyer['telegram_id'], text=txt, parse_mode='HTML')
    except Exception as e:
        logging.warning(f"Xaridorga qarz to'lovi xabari ketmadi: {e}")


async def buyer_debts(update, context):
    """Xaridor paneli — «Mening qarzlarim» (kimga qancha qarzdorman)."""
    query = update.callback_query
    lang = get_lang(update, context)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    debts = db.get_buyer_open_debts(user['id'])
    if not debts:
        kb = InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="buyer_panel")]])
        text = t(lang, 'my_debts_empty')
        if query:
            await query.answer(); await query.edit_message_text(text, reply_markup=kb)
        else:
            await update.message.reply_text(text, reply_markup=kb)
        return
    total = sum(float(d['total_due']) for d in debts)
    lines = [t(lang, 'my_debts_title', total=fmt_price(total))]
    for d in debts:
        shop = html.escape(d.get('shop_name') or d.get('seller_name') or '—')
        lines.append(t(lang, 'my_debts_row', shop=shop, due=fmt_price(d['total_due'])))
    kb = InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="buyer_panel")]])
    text = "\n".join(lines)
    if query:
        await query.answer()
        await query.edit_message_text(text, reply_markup=kb, parse_mode='HTML')
    else:
        await update.message.reply_text(text, reply_markup=kb, parse_mode='HTML')


async def update_order_status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    data = query.data
    # confirm_order_ID, cancel_order_ID, deliver_order_ID
    parts = data.split("_")
    order_id = int(parts[2])
    action = parts[0]  # confirm / cancel / deliver

    # XAVFSIZLIK: faqat shu buyurtmaning sotuvchisi (yoki admin) holatini o'zgartira oladi.
    # Telegram callback'lari soxtalashtirilmasa-da, egalikni server tomonda tasdiqlaymiz —
    # begona buyurtmani tasdiqlash/bekor qilish va xaridor ma'lumotlarini ko'rish oldi olinadi.
    if not await _ensure_order_seller(update, context, order_id):
        return

    # BERISH: to'g'ridan-to'g'ri 'delivered' qilmaymiz — avval to'lov holatini so'raymiz
    # (to'liq to'landi / qarzga / bo'lib to'lashga). Yakuniy holatni settlement oqimi qo'yadi.
    if action == 'deliver':
        order = db.get_order_by_id(order_id)
        if order:
            await _ask_settlement(update, context, scope='o', key=order_id,
                                  total=float(order.get('total_price') or 0))
        return

    status_map = {'confirm': 'confirmed', 'cancel': 'cancelled', 'deliver': 'delivered'}
    new_status = status_map.get(action)
    if new_status:
        # HIMOYA (ATOMIK): faqat 'pending' holatdan o'tkazamiz. Bot va Mini App AYNAN bir
        # buyurtmani bir vaqtda tasdiqlasa/bekor qilsa yoki tugma ikki marta bosilsa —
        # `transition_order_status` faqat BITTA chaqiruvga True qaytaradi (rowcount=1).
        # Yutmagan chaqiruv zahirani kamaytirmaydi va xabar yubormaydi, faqat detalni
        # ko'rsatadi (ilgarigi read-then-check atomik emas edi → ikki marta ishlardi).
        won = db.transition_order_status(order_id, new_status, 'pending',
                                         cancel_by='seller' if new_status == 'cancelled' else None)
        if not won:
            await seller_order_detail(update, context)
            return

        # Jonli sanoq + avtomatik bekor taymerini o'chiramiz
        if context.application.job_queue:
            for jn in (f"countdown_order_{order_id}", f"auto_cancel_{order_id}", f"reminder_{order_id}"):
                for job in context.application.job_queue.get_jobs_by_name(jn):
                    job.schedule_removal()

        # Stock kamaytirish — faqat tasdiqlanganda (va biz yutgan bo'lsak)
        if new_status == 'confirmed':
            try:
                order_for_stock = db.get_order_by_id(order_id)
                if order_for_stock:
                    left = db.decrement_stock_on_confirm(
                        order_for_stock['product_id'],
                        order_for_stock['quantity']
                    )
                    if left == 0:
                        await _notify_seller_sold_out(context, order_for_stock['product_id'])
            except Exception as e:
                logging.error(f"Stock kamaytirish xatosi: {e}")

        # Xaridorga bildirishnoma yuboramiz (xaridor tilida)
        try:
            order = db.get_order_by_id(order_id)
            if order and order.get('buyer_tg'):
                buyer = db.get_user_by_id(order['buyer_id'])
                blang = get_user_lang(buyer) if buyer else DEFAULT_LANG
                dlv = order.get('delivery_type', 'delivery')
                is_pickup = dlv == 'pickup'
                oid = fmt_order_id(order_id)
                pname = html.escape(order.get('product_name') or '')

                msg_map = {
                    'confirmed': t(blang, 'order_confirmed_pickup' if is_pickup else 'order_confirmed_delivery',
                                   oid=oid, pname=pname),
                    'cancelled': t(blang, 'order_cancelled_notify', oid=oid, pname=pname),
                    'delivered': t(blang, 'order_delivered_pickup' if is_pickup else 'order_delivered_delivery',
                                   oid=oid, pname=pname),
                }

                txt = msg_map.get(new_status)
                if txt:
                    kb = None
                    if new_status == 'delivered':
                        kb = InlineKeyboardMarkup([[
                            InlineKeyboardButton(
                                t(blang, 'btn_leave_rating'), callback_data=f"order_rate_{order_id}"
                            )
                        ]])
                    await context.bot.send_message(
                        chat_id=order['buyer_tg'], text=txt,
                        reply_markup=kb, parse_mode='HTML'
                    )
        except Exception as e:
            logging.error(f"Xaridorga bildirishnoma ketmadi: {e}")

    await seller_order_detail(update, context)


# ============================================================
# MESSAGING
# ============================================================

async def message_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    order_id = int(query.data.split("_")[2])
    context.user_data['current_order_id'] = order_id

    await query.edit_message_text(T(update, context, 'message_ask'))
    return MESSAGE_TEXT


async def message_send(update: Update, context: ContextTypes.DEFAULT_TYPE):
    message_text = update.message.text
    order_id = context.user_data.get('current_order_id')

    user = db.get_user_by_telegram_id(update.effective_user.id)
    lang = get_lang(update, context)

    # Buyurtma bo'yicha xaridor va sotuvchini olamiz (telegram_id bilan)
    order = db.get_order_by_id(order_id)
    if not order:
        await update.message.reply_text(t(lang, 'order_not_found_x'))
        return ConversationHandler.END

    # Foydalanuvchi shu buyurtmaning xaridori/sotuvchisi/kuryeri — shunga qarab xabar
    # yo'naltiriladi. #13 — kuryer ham ishtirokchi: xaridordan kuryer biriktirilgan bo'lsa
    # kuryerga, aks holda sotuvchiga; kuryer/sotuvchidan — xaridorga.
    if user['id'] == order['buyer_id']:
        if order.get('courier_id'):
            receiver_id = order['courier_id']
            receiver_tg = order.get('courier_tg')
        else:
            receiver_id = order['seller_id']
            receiver_tg = order.get('seller_tg')
        sender_role = 'buyer'
    elif user['id'] == order.get('courier_id'):
        receiver_id = order['buyer_id']
        receiver_tg = order.get('buyer_tg')
        sender_role = 'courier'
    else:
        receiver_id = order['buyer_id']
        receiver_tg = order.get('buyer_tg')
        sender_role = 'seller'

    # Qabul qiluvchi tilini aniqlash uchun
    receiver = db.get_user_by_id(receiver_id) if receiver_id else None
    rlang = get_user_lang(receiver) if receiver else DEFAULT_LANG
    if sender_role == 'buyer':
        sender_label = t(rlang, 'sender_label_buyer', name=html.escape(user.get('name') or ''))
    elif sender_role == 'courier':
        sender_label = t(rlang, 'sender_label_courier', name=html.escape(user.get('name') or ''))
    else:
        sender_label = t(rlang, 'sender_label_seller', name=html.escape(user.get('shop_name') or user.get('name') or ''))

    db.create_message(order_id, user['id'], receiver_id, message_text)
    await update.message.reply_text(t(lang, 'message_sent'))

    # Qabul qiluvchiga real yetkazish — Telegram orqali (qabul qiluvchi tilida)
    if receiver_tg:
        try:
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton(t(rlang, 'btn_reply'), callback_data=f"order_msg_{order_id}")],
                [InlineKeyboardButton(t(rlang, 'btn_correspondence'), callback_data=f"msgs_{order_id}")],
            ])
            await context.bot.send_message(
                chat_id=receiver_tg,
                text=t(rlang, 'new_message_notify', oid=fmt_order_id(order_id),
                       sender=sender_label, msg=html.escape(message_text)),
                reply_markup=kb,
                parse_mode='HTML'
            )
        except Exception as e:
            logging.error(f"Xabar qabul qiluvchiga yetkazilmadi: {e}")

    # Qaysi panelga qaytishni active_mode aniqlaydi
    if get_active_mode(user, context) == 'seller':
        await seller_panel(update, context)
    else:
        await buyer_panel(update, context)

    return ConversationHandler.END


async def view_messages(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Buyurtma bo'yicha xabar tarixini ko'rsatadi."""
    query = update.callback_query
    await query.answer()

    order_id = int(query.data.split("_")[1])
    lang = get_lang(update, context)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    order = db.get_order_by_id(order_id)

    if not order or not user or user['id'] not in (order['buyer_id'], order['seller_id'], order.get('courier_id')):
        await query.edit_message_text(t(lang, 'order_not_yours_full'))
        return

    messages = db.get_messages_by_order(order_id)
    if not messages:
        text = t(lang, 'no_messages_yet')
    else:
        lines = [t(lang, 'messages_history_header', oid=fmt_order_id(order_id))]
        for m in messages[-30:]:  # so'nggi 30 ta xabar
            who = "👤" if m['sender_id'] == order['buyer_id'] else ("🚴" if m['sender_id'] == order.get('courier_id') else "🏪")
            name = html.escape(m.get('sender_name') or '')
            msg = html.escape(m.get('message') or '')
            ts = fmt_datetime(m.get('created_at'))
            lines.append(f"{who} <b>{name}</b> · {ts}\n{msg}\n")
        text = "\n".join(lines)
        # Telegram xabari 4096 belgidan uzun bo'la olmaydi — kesamiz
        if len(text) > 3800:
            text = text[:3800] + t(lang, 'old_messages_cut')

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(t(lang, 'btn_new_message'), callback_data=f"order_msg_{order_id}")],
        [InlineKeyboardButton(t(lang, 'back'), callback_data=f"order_detail_{order_id}"
                              if user['id'] == order['buyer_id'] else f"seller_order_{order_id}")],
    ])
    await query.edit_message_text(text, reply_markup=kb, parse_mode='HTML')


async def seller_messages(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Sotuvchining barcha xabarli buyurtmalari ro'yxati."""
    query = update.callback_query
    await query.answer()

    user = db.get_user_by_telegram_id(update.effective_user.id)
    rows = db.get_seller_messages_summary(user['id'])

    lang = get_lang(update, context)
    if not rows:
        await query.edit_message_text(
            t(lang, 'messages_empty'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="seller_panel")]])
        )
        return

    keyboard = []
    for rec in rows:
        label = f"📜 {fmt_order_id(rec['id'])} — {rec.get('product_name', '')[:25]}"
        keyboard.append([InlineKeyboardButton(label, callback_data=f"msgs_{rec['id']}")])
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="seller_panel")])

    await query.edit_message_text(
        t(lang, 'recent_correspondence'),
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


# ============================================================
# RATING
# ============================================================

async def rating_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    order_id = int(query.data.split("_")[2])
    context.user_data['current_order_id'] = order_id
    context.user_data.pop('product_rating', None)
    context.user_data.pop('product_comment', None)
    context.user_data.pop('seller_rating', None)

    # 1-qadam: mahsulotni baholash
    keyboard = [[InlineKeyboardButton("⭐" * i, callback_data=f"prate_{i}")] for i in range(1, 6)]
    await query.edit_message_text(
        T(update, context, 'rate_product_ask'),
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode='HTML'
    )
    return PRODUCT_RATING


async def rating_product_select(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """1-qadam tugadi: mahsulot reytingi tanlandi -> izoh so'raymiz."""
    query = update.callback_query
    await query.answer()

    context.user_data['product_rating'] = int(query.data.split("_")[1])
    await query.edit_message_text(
        T(update, context, 'rate_product_comment_ask'),
        parse_mode='HTML'
    )
    return PRODUCT_COMMENT


async def rating_product_comment(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """2-qadam tugadi: mahsulot izohi olindi -> sotuvchini baholashga o'tamiz."""
    comment = (update.message.text or "").strip()
    if comment == "-":
        comment = None
    context.user_data['product_comment'] = comment

    # 3-qadam: sotuvchini (do'konni) baholash
    keyboard = [[InlineKeyboardButton("⭐" * i, callback_data=f"srate_{i}")] for i in range(1, 6)]
    await update.message.reply_text(
        T(update, context, 'rate_seller_ask'),
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode='HTML'
    )
    return SELLER_RATING


async def rating_submit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """3-qadam tugadi: sotuvchi reytingi tanlandi -> bazaga yozamiz."""
    query = update.callback_query
    await query.answer()

    seller_rating = int(query.data.split("_")[1])
    product_rating = context.user_data.get('product_rating')
    comment = context.user_data.get('product_comment')
    order_id = context.user_data.get('current_order_id')
    user = db.get_user_by_telegram_id(update.effective_user.id)
    lang = get_lang(update, context)

    if not product_rating or not seller_rating or not order_id:
        await query.edit_message_text(t(lang, 'rate_not_found'))
        await buyer_panel(update, context)
        return ConversationHandler.END

    order = db.get_order_by_id_for_rating(order_id)
    if not order:
        await query.edit_message_text(t(lang, 'order_not_found_x'))
        await buyer_panel(update, context)
        return ConversationHandler.END

    # Bir buyurtmaga bir martadan ko'p baho qoldirishni oldini olamiz
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id FROM reviews WHERE order_id=? AND buyer_id=?",
        (order_id, user['id'])
    )
    if cursor.fetchone():
        await query.edit_message_text(t(lang, 'rate_already'))
        await buyer_panel(update, context)
        return ConversationHandler.END

    db.create_review(
        order_id=order_id,
        seller_id=order['seller_id'],
        buyer_id=user['id'],
        rating=seller_rating,                 # do'kon (sotuvchi) reytingi
        comment=comment,                      # mahsulot haqida izoh
        product_id=order.get('product_id'),   # baho qaysi mahsulotga
        product_rating=product_rating         # mahsulot reytingi
    )

    # Sotuvchiga baho haqida xabar (sotuvchi tilida)
    try:
        full_order = db.get_order_by_id(order_id)
        if full_order and full_order.get('seller_tg'):
            seller = db.get_user_by_id(order['seller_id'])
            slang = get_user_lang(seller) if seller else DEFAULT_LANG
            p_stars = "⭐" * product_rating
            s_stars = "⭐" * seller_rating
            msg = t(slang, 'rate_seller_notify',
                    pname=html.escape(full_order.get('product_name') or ''),
                    pstars=p_stars, prate=product_rating,
                    sstars=s_stars, srate=seller_rating,
                    buyer=html.escape(user.get('name') or ''))
            if comment:
                msg += t(slang, 'rate_comment_line', comment=html.escape(comment))
            await context.bot.send_message(
                chat_id=full_order['seller_tg'], text=msg, parse_mode='HTML'
            )
    except Exception as e:
        logging.error(f"Baho bildirishnomasi ketmadi: {e}")

    await query.edit_message_text(
        t(lang, 'rate_thanks', pstars='⭐' * product_rating, sstars='⭐' * seller_rating)
    )
    await buyer_panel(update, context)
    return ConversationHandler.END


# ============================================================
# ADMIN PANEL
# ============================================================

async def admin_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Admin tekshiruvi
    _uid = update.effective_user.id if update.effective_user else None
    _user = db.get_user_by_telegram_id(_uid) if _uid else None
    lang = get_lang(update, context)
    if not _user or (_user['role'] != 'admin' and _uid != ADMIN_ID):
        if update.callback_query:
            await update.callback_query.answer(t(lang, 'no_access_alert'), show_alert=True)
        else:
            await update.message.reply_text(t(lang, 'not_admin'))
        return

    total_users = len(db.get_all_users())
    total_products = len(db.get_all_products(include_hidden=False))
    total_orders = len(db.get_all_orders())
    pending_requests = len(db.get_pending_seller_requests())
    dispute_count = len(db.get_disputed_orders())

    # Tugmalar bo'limlarga guruhlangan. Harakat talab qiladigan «Sotuvchi so'rovlari»
    # va «Nizolar» (sonli) — tepada, ko'rinib turishi uchun.
    keyboard = [
        [InlineKeyboardButton(t(lang, 'btn_seller_requests_n', n=pending_requests), callback_data="admin_seller_requests")],
        [InlineKeyboardButton(t(lang, 'btn_disputes_n', n=dispute_count), callback_data="admin_disputes")],
        [InlineKeyboardButton(t(lang, 'agrp_people'), callback_data="admingrp_people")],
        [InlineKeyboardButton(t(lang, 'agrp_catalog'), callback_data="admingrp_catalog")],
        [InlineKeyboardButton(t(lang, 'agrp_manage'), callback_data="admingrp_manage")],
        [InlineKeyboardButton(t(lang, 'btn_ai_assistant'), callback_data="ai_assistant")],
    ]

    text = t(lang, 'admin_panel_body', users=total_users, products=total_products, orders=total_orders)

    if update.callback_query:
        try:
            await update.callback_query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')
        except Exception:
            await update.callback_query.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')
    else:
        # Admin faqat inline panel bilan ishlaydi. Boshqa rejimlardan (mas. sotuvchi
        # profilini tahrirlash yoki xaridor qidiruvi) qolib ketgan pastki Reply
        # klaviaturani — jumladan "📍 Lokatsiya yuborish" tugmasini — tozalaymiz.
        try:
            await update.message.reply_text(t(lang, 'admin_kb_cleared'), reply_markup=ReplyKeyboardRemove())
        except Exception:
            pass
        await update.message.reply_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


# ============================================================
# ADMIN PANELI — BO'LIM (guruh) MENYULARI
# ============================================================
async def _show_admin_group(update, context, title, kb):
    """Admin bo'lim ekranini ko'rsatadi (Orqaga — admin panelga)."""
    kb = kb + [[InlineKeyboardButton(t(get_lang(update, context), 'back'), callback_data="admin_panel")]]
    query = update.callback_query
    markup = InlineKeyboardMarkup(kb)
    if query:
        await query.answer()
        try:
            await query.edit_message_text(title, reply_markup=markup, parse_mode='HTML')
        except Exception:
            await context.bot.send_message(chat_id=update.effective_chat.id, text=title,
                                           reply_markup=markup, parse_mode='HTML')
    else:
        await update.message.reply_text(title, reply_markup=markup, parse_mode='HTML')


async def admin_group_people(update, context):
    """👥 Odamlar — foydalanuvchilar, do'konlar, kanallar."""
    lang = get_lang(update, context)
    kb = [
        [InlineKeyboardButton(t(lang, 'btn_admin_users'), callback_data="admin_users")],
        [InlineKeyboardButton(t(lang, 'btn_admin_shops'), callback_data="admin_shops")],
        [InlineKeyboardButton(t(lang, 'btn_admin_channels'), callback_data="admin_channels")],
    ]
    await _show_admin_group(update, context, t(lang, 'agrp_people_title'), kb)


async def admin_group_catalog(update, context):
    """📦 Katalog — mahsulotlar, o'chirilganlar, buyurtmalar."""
    lang = get_lang(update, context)
    kb = [
        [InlineKeyboardButton(t(lang, 'btn_admin_products'), callback_data="admin_products")],
        [InlineKeyboardButton(t(lang, 'btn_deleted_products'), callback_data="admin_deleted_products")],
        [InlineKeyboardButton(t(lang, 'btn_admin_orders'), callback_data="admin_orders")],
    ]
    await _show_admin_group(update, context, t(lang, 'agrp_catalog_title'), kb)


async def admin_group_manage(update, context):
    """🛠 Boshqaruv — statistika, ommaviy xabar, sozlamalar."""
    lang = get_lang(update, context)
    kb = [
        [InlineKeyboardButton(t(lang, 'btn_admin_stats'), callback_data="admin_stats")],
        [InlineKeyboardButton(t(lang, 'btn_admin_broadcast'), callback_data="admin_broadcast")],
        [InlineKeyboardButton(t(lang, 'btn_admin_settings'), callback_data="admin_settings")],
    ]
    await _show_admin_group(update, context, t(lang, 'agrp_manage_title'), kb)


async def admin_shops(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin — barcha do'konlar ro'yxati."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    shops = db.get_all_shops()
    if not shops:
        await query.edit_message_text(
            t(lang, 'admin_shops_empty'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="admin_panel")]]))
        return
    kb = []
    for s in shops[:30]:
        nm = s.get('name') or s.get('owner_name') or '—'
        kb.append([InlineKeyboardButton(f"🏪 {nm[:30]} · 👥{s.get('staff_count', 0)}",
                                        callback_data=f"admin_shop_{s['id']}")])
    kb.append([InlineKeyboardButton(t(lang, 'back'), callback_data="admin_panel")])
    await query.edit_message_text(t(lang, 'admin_shops_header', n=len(shops)),
                                  reply_markup=InlineKeyboardMarkup(kb), parse_mode='HTML')


async def admin_shop_detail(update: Update, context: ContextTypes.DEFAULT_TYPE, shop_id=None):
    """Admin — bitta do'kon: xodimlar daraxti, moderatsiya, faollashtirish."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    if shop_id is None:
        shop_id = int(query.data.split("_")[2])
    shop = db.get_shop_by_id(shop_id)
    if not shop:
        await query.answer(t(lang, 'shop_not_found'), show_alert=True)
        return
    owner = db.get_user_by_id(shop['owner_user_id'])
    staff_all = db.get_shop_staff(shop_id)
    perf = {r['user_id']: r for r in db.get_shop_staff_performance(shop_id)}
    lines = [t(lang, 'admin_shop_title', name=html.escape(shop.get('name') or '—'),
               owner=html.escape((owner.get('name') if owner else '') or '—'),
               mod=t(lang, 'mod_owner' if shop.get('moderation') == 'owner_approve' else 'mod_direct'),
               paymode=t(lang, 'paymode_shop' if (shop.get('payment_mode') or 'shop') == 'shop' else 'paymode_staff'))]
    for s in staff_all:
        mark = "👑" if s.get('staff_role') == 'owner' else ("✅" if s.get('is_active') else "⏳")
        pr = perf.get(s['user_id'], {})
        lines.append(t(lang, 'admin_shop_staff_row',
                       mark=mark, name=html.escape(s.get('name') or '—'),
                       dept=html.escape(s.get('department') or '—'),
                       revenue=fmt_price(pr.get('revenue', 0))))
    kb = [[InlineKeyboardButton(t(lang, 'btn_admin_toggle_mod'), callback_data=f"admin_shopmod_{shop_id}")]]
    # Nofaol xodimlarni tezkor faollashtirish
    for s in staff_all:
        if s.get('staff_role') != 'owner' and not s.get('is_active'):
            kb.append([InlineKeyboardButton(t(lang, 'btn_admin_activate_staff', name=(s.get('name') or '—')[:20]),
                                            callback_data=f"admin_stafftog_{s['id']}_{shop_id}")])
    kb.append([InlineKeyboardButton(t(lang, 'back'), callback_data="admin_shops")])
    await query.edit_message_text("\n".join(lines), reply_markup=InlineKeyboardMarkup(kb), parse_mode='HTML')


async def admin_shop_toggle_mod(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin — do'kon moderatsiya siyosatini almashtiradi (direct ↔ owner_approve)."""
    query = update.callback_query
    await query.answer()
    shop_id = int(query.data.split("_")[2])
    shop = db.get_shop_by_id(shop_id)
    if not shop:
        return
    new_mod = 'owner_approve' if (shop.get('moderation') or 'direct') == 'direct' else 'direct'
    db.update_shop(shop_id, moderation=new_mod)
    # Detalni qayta ko'rsatamiz
    await admin_shop_detail(update, context, shop_id=shop_id)


async def admin_staff_toggle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin — xodimni faollashtiradi/muzlatadi."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    parts = query.data.split("_")
    staff_id = int(parts[2])
    shop_id = int(parts[3])
    target = next((s for s in db.get_shop_staff(shop_id, include_owner=False) if s['id'] == staff_id), None)
    if not target:
        await query.answer(t(lang, 'staff_not_found'), show_alert=True)
        return
    new_active = 0 if target.get('is_active') else 1
    db.set_staff_active(staff_id, new_active)
    try:
        su = db.get_user_by_id(target['user_id'])
        if su and su.get('telegram_id'):
            slang = get_user_lang(su)
            await context.bot.send_message(
                chat_id=su['telegram_id'],
                text=t(slang, 'staff_you_activated' if new_active else 'staff_you_frozen'))
    except Exception as e:
        logging.error(f"Admin: xodimga holat xabari ketmadi: {e}")
    await admin_shop_detail(update, context, shop_id=shop_id)


async def admin_seller_requests(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin — kutilayotgan sotuvchi so'rovlari ro'yxati."""
    query = update.callback_query
    await query.answer()

    lang = get_lang(update, context)
    requests = db.get_pending_seller_requests()

    if not requests:
        await query.edit_message_text(
            t(lang, 'no_pending_requests'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="admin_panel")]])
        )
        return

    keyboard = []
    for req in requests[:20]:
        name = req.get('name') or t(lang, 'anonymous')
        shop = req.get('shop_name') or ''
        label = f"🏪 {name} — {shop}"[:45]
        keyboard.append([InlineKeyboardButton(label, callback_data=f"seller_req_{req['id']}")])
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="admin_panel")])

    await query.edit_message_text(
        t(lang, 'pending_requests_title', n=len(requests)),
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def admin_seller_request_detail(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Bitta sotuvchi so'rovining batafsil ma'lumotlari."""
    query = update.callback_query
    await query.answer()

    request_id = int(query.data.split("_")[2])

    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT sr.*, u.name, u.phone_number, u.shop_name, u.shop_address,
               u.shop_landmark, u.working_days, u.working_hours,
               u.telegram_username, u.telegram_id, u.id as uid
        FROM seller_requests sr
        JOIN users u ON sr.user_id = u.id
        WHERE sr.id=?
    """, (request_id,))
    row = cursor.fetchone()

    lang = get_lang(update, context)
    if not row:
        await query.edit_message_text(
            t(lang, 'request_not_found'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="admin_seller_requests")]])
        )
        return

    req = dict(row)
    user_id = req['uid']

    text = t(lang, 'seller_request_detail',
             name=html.escape(req.get('name') or ''),
             phone=req.get('phone_number') or '—',
             shop=html.escape(req.get('shop_name') or '—'),
             address=html.escape(req.get('shop_address') or '—'),
             landmark=html.escape(req.get('shop_landmark') or '—'),
             wd=html.escape(req.get('working_days') or '—'),
             wh=html.escape(req.get('working_hours') or '—'),
             tg=req.get('telegram_username') or '—',
             date=fmt_datetime(req.get('created_at')))

    keyboard = [
        [InlineKeyboardButton(t(lang, 'btn_confirm'), callback_data=f"approve_seller_{user_id}")],
        [InlineKeyboardButton(t(lang, 'btn_reject'), callback_data=f"reject_seller_{user_id}")],
        [InlineKeyboardButton(t(lang, 'back'), callback_data="admin_seller_requests")],
    ]

    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


async def approve_seller(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin sotuvchini tasdiqlaydi."""
    query = update.callback_query
    await query.answer()

    user_id = int(query.data.split("_")[2])
    user = db.get_user_by_id(user_id)
    lang = get_lang(update, context)

    if not user:
        await query.edit_message_text(t(lang, 'user_not_found'))
        return

    # Foydalanuvchini tasdiqlash
    db.update_user(user_id, is_approved=1, role='seller')

    # MULTI-SOTUVCHI: tasdiqlangan sotuvchida do'kon bo'lishini kafolatlaymiz (idempotent).
    # Bu xodim sifatida boshqa do'konga biriktirilmaganlar uchun — o'z do'konini yaratadi.
    try:
        if not db.get_staff_by_user(user_id):
            db.create_shop(
                user_id,
                name=user.get('shop_name'),
                address=user.get('shop_address'),
                landmark=user.get('shop_landmark'),
                lat=user.get('shop_lat'),
                lon=user.get('shop_lon'),
                working_days=user.get('working_days'),
                working_hours=user.get('working_hours'),
                region_id=user.get('region_id'),
                card_number=user.get('card_number'),
                card_owner=user.get('card_owner'),
                card_type=user.get('card_type'),
            )
    except Exception as e:
        logging.error(f"approve_seller: do'kon yaratilmadi: {e}")

    # seller_requests jadvalini yangilash
    req = db.get_seller_request_by_user(user_id)
    if req:
        db.update_seller_request(req['id'], 'approved')

    # Sotuvchiga xabar (sotuvchi tilida)
    try:
        await context.bot.send_message(
            chat_id=user['telegram_id'],
            text=t(user, 'approve_seller_notify'),
            parse_mode='HTML'
        )
    except Exception as e:
        logging.error(f"Sotuvchiga tasdiqlash xabari ketmadi: {e}")

    await query.edit_message_text(
        t(lang, 'seller_approved_admin', name=html.escape(user.get('name') or '')),
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'btn_requests_back'), callback_data="admin_seller_requests")]])
    )


async def reject_seller(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin sotuvchini rad etadi."""
    query = update.callback_query
    await query.answer()

    user_id = int(query.data.split("_")[2])
    user = db.get_user_by_id(user_id)
    lang = get_lang(update, context)

    if not user:
        await query.edit_message_text(t(lang, 'user_not_found'))
        return

    # So'rovni rad etish
    db.update_user(user_id, is_approved=0, role='buyer')

    req = db.get_seller_request_by_user(user_id)
    if req:
        db.update_seller_request(req['id'], 'rejected')

    # Foydalanuvchiga xabar (uning tilida)
    try:
        await context.bot.send_message(
            chat_id=user['telegram_id'],
            text=t(user, 'reject_seller_notify'),
            parse_mode='HTML'
        )
    except Exception as e:
        logging.error(f"Sotuvchiga rad xabari ketmadi: {e}")

    await query.edit_message_text(
        t(lang, 'seller_rejected_admin', name=html.escape(user.get('name') or '')),
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'btn_requests_back'), callback_data="admin_seller_requests")]])
    )


ADMIN_USERS_PAGE_SIZE = 10


async def admin_users(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    # Qidirish holatini tozalash
    context.user_data.pop('admin_searching_user', None)

    # Sahifa raqami callback_data dan kelishi mumkin: 'admin_users' yoki 'admin_users_pg_N'
    page = 0
    if query.data.startswith("admin_users_pg_"):
        try:
            page = int(query.data.replace("admin_users_pg_", ""))
        except ValueError:
            page = 0

    total, rows_raw = db.get_users_paginated(limit=ADMIN_USERS_PAGE_SIZE, offset=0)
    total_pages = max(1, (total + ADMIN_USERS_PAGE_SIZE - 1) // ADMIN_USERS_PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    offset = page * ADMIN_USERS_PAGE_SIZE
    total, rows_raw = db.get_users_paginated(limit=ADMIN_USERS_PAGE_SIZE, offset=offset)
    rows = rows_raw
    columns = list(rows[0].keys()) if rows else []

    lang = get_lang(update, context)
    if not rows:
        await query.edit_message_text(
            t(lang, 'admin_no_users'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="admin_panel")]])
        )
        return

    keyboard = []
    # Qidirish tugmasi
    keyboard.append([InlineKeyboardButton(t(lang, 'btn_search_user'), callback_data="admin_user_search")])
    for user in rows:
        status = "🟢" if not user.get('is_blocked') else "🔴"
        role_emoji = {"buyer": "🛒", "seller": "🏪", "admin": "🔧"}
        keyboard.append([InlineKeyboardButton(
            f"{status} {role_emoji.get(user.get('role'), '❓')} {user.get('name') or t(lang, 'anonymous')}",
            callback_data=f"admin_user_{user['id']}"
        )])

    # Pagination tugmalari
    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton("⬅️", callback_data=f"admin_users_pg_{page-1}"))
    nav.append(InlineKeyboardButton(f"{page+1}/{total_pages}", callback_data="noop"))
    if page < total_pages - 1:
        nav.append(InlineKeyboardButton("➡️", callback_data=f"admin_users_pg_{page+1}"))
    if nav:
        keyboard.append(nav)
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="admin_panel")])

    await query.edit_message_text(
        t(lang, 'admin_users_title', total=total, page=page+1, pages=total_pages),
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def admin_user_search_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin — foydalanuvchi qidirish boshlash."""
    query = update.callback_query
    await query.answer()
    context.user_data['admin_searching_user'] = True
    lang = get_lang(update, context)
    await query.edit_message_text(
        t(lang, 'admin_user_search_ask'),
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton(t(lang, 'btn_cancel'), callback_data="admin_users")]
        ])
    )


async def admin_user_search_result(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin — foydalanuvchi qidirish natijalari."""
    search_text = update.message.text.strip()
    context.user_data.pop('admin_searching_user', None)
    lang = get_lang(update, context)

    if len(search_text) < 2:
        await update.message.reply_text(
            t(lang, 'admin_search_min2'),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(t(lang, 'btn_retry_search_user'), callback_data="admin_user_search")],
                [InlineKeyboardButton(t(lang, 'back'), callback_data="admin_users")],
            ])
        )
        return

    results = db.search_users(search_text)

    if not results:
        await update.message.reply_text(
            t(lang, 'admin_search_none', q=search_text),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(t(lang, 'btn_retry_search_user'), callback_data="admin_user_search")],
                [InlineKeyboardButton(t(lang, 'back'), callback_data="admin_users")],
            ])
        )
        return

    keyboard = []
    for user in results[:20]:
        status = "🟢" if not user.get('is_blocked') else "🔴"
        role_emoji = {"buyer": "🛒", "seller": "🏪", "admin": "🔧"}
        name = user.get('name') or t(lang, 'anonymous')
        shop = f" · {user['shop_name']}" if user.get('shop_name') else ""
        keyboard.append([InlineKeyboardButton(
            f"{status} {role_emoji.get(user.get('role'), '❓')} {name}{shop}",
            callback_data=f"admin_user_{user['id']}"
        )])

    keyboard.append([InlineKeyboardButton(t(lang, 'btn_retry_search_user'), callback_data="admin_user_search")])
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="admin_users")])

    await update.message.reply_text(
        t(lang, 'admin_search_results', q=search_text, n=len(results)),
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


def _profile_missing_fields(user, lang):
    """Foydalanuvchi profilidagi to'ldirilmagan muhim maydonlar ro'yxati (til bo'yicha label).
    Sotuvchi uchun do'kon maydonlari ham tekshiriladi."""
    ru = (lang == 'ru')
    def empty(v):
        return v is None or (isinstance(v, str) and not v.strip())
    missing = []
    # --- Umumiy ---
    if empty(user.get('name')):
        missing.append("Имя" if ru else "Ism")
    if empty(user.get('telegram_username')):
        missing.append("Username (@...)" if ru else "Username (@...)")
    if empty(user.get('phone_number')):
        missing.append("Номер телефона" if ru else "Telefon raqami")
    if user.get('region_id') is None:
        missing.append("Регион (область/район)" if ru else "Hudud (viloyat/tuman)")
    # --- Sotuvchi (do'koni bor) ---
    is_seller = bool(user.get('shop_name')) or user.get('role') == 'seller'
    if is_seller:
        if empty(user.get('shop_name')):
            missing.append("Название магазина" if ru else "Do'kon nomi")
        if empty(user.get('shop_address')):
            missing.append("Адрес магазина" if ru else "Do'kon manzili")
        if empty(user.get('shop_landmark')):
            missing.append("Ориентир" if ru else "Mo'ljal (orientir)")
        if empty(user.get('working_hours')):
            missing.append("Часы работы" if ru else "Ish vaqti")
        if empty(user.get('working_days')):
            missing.append("Рабочие дни" if ru else "Ish kunlari")
        if empty(user.get('card_number')):
            missing.append("Карта для оплаты" if ru else "To'lov kartasi")
    return missing


async def admin_user_details(update: Update, context: ContextTypes.DEFAULT_TYPE, user_id: int = None):
    query = update.callback_query
    if user_id is None:
        await query.answer()
        user_id = int(query.data.split("_")[2])

    user = db.get_user_by_id(user_id)
    lang = get_lang(update, context)

    if not user:
        await query.edit_message_text(t(lang, 'user_not_found'))
        return
    status = t(lang, 'adu_status_active') if not user['is_blocked'] else t(lang, 'adu_status_blocked')

    # Sotuvchi uchun is_approved ni ko'rsatamiz
    is_seller = bool(user.get('shop_name'))
    if is_seller:
        approved_text = t(lang, 'adu_seller_approved') if user.get('is_approved') else t(lang, 'adu_seller_not_approved')
    else:
        approved_text = t(lang, 'adu_buyer_dash')

    keyboard = [
        [InlineKeyboardButton(
            t(lang, 'btn_unblock') if user['is_blocked'] else t(lang, 'btn_block'),
            callback_data=f"admin_block_{user_id}"
        )],
    ]

    # Sotuvchi tasdiqlash/bekor qilish — faqat do'koni bor foydalanuvchilar uchun
    if is_seller:
        if user.get('is_approved'):
            keyboard.append([InlineKeyboardButton(
                t(lang, 'btn_unverify_seller'),
                callback_data=f"admin_verify_{user_id}"
            )])
        else:
            keyboard.append([InlineKeyboardButton(
                t(lang, 'btn_verify_seller'),
                callback_data=f"admin_verify_{user_id}"
            )])

    # Profilda kamchilik bo'lsa — AI orqali to'ldirishni so'rash tugmasi
    _user_lang = user.get('language') or DEFAULT_LANG
    if _profile_missing_fields(user, _user_lang) and ai_assistant.is_enabled():
        keyboard.append([InlineKeyboardButton(
            t(lang, 'btn_request_fill'),
            callback_data=f"admin_askfill_{user_id}"
        )])

    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="admin_users")])

    seller_flag = is_seller or user.get('role') == 'seller'

    # ===== Asosiy ma'lumotlar =====
    uname = user.get('telegram_username')
    uname_txt = f"@{html.escape(uname)}" if uname else "—"
    verified_txt = t(lang, 'adu_yes') if user.get('is_verified') else t(lang, 'adu_no')
    _none = t(lang, 'word_none_yes')

    text = t(lang, 'adu_main_block',
             id=user['id'], name=html.escape(user.get('name') or t(lang, 'anonymous')),
             tgid=user.get('telegram_id'), username=uname_txt,
             phone=fmt_phone(user.get('phone_number')), role=user.get('role') or '—',
             lang=user.get('language') or '—',
             region=user.get('region_id') if user.get('region_id') is not None else '—',
             status=status, approved=approved_text, verified=verified_txt,
             created=fmt_datetime(user.get('created_at')), updated=fmt_datetime(user.get('updated_at')))

    # ===== Do'kon ma'lumotlari =====
    if seller_flag:
        shop_addr = human_address(user.get('shop_address'))
        addr_line = (html.escape(shop_addr) if shop_addr else _none) + maps_link(user.get('shop_lat'), user.get('shop_lon'))
        text += t(lang, 'adu_shop_block',
                  name=html.escape(user.get('shop_name') or _none), addr=addr_line,
                  lm=html.escape(user.get('shop_landmark') or _none),
                  wh=html.escape(user.get('working_hours') or _none),
                  wd=html.escape(user.get('working_days') or _none))

    # ===== Ulangan kanallar (sotuvchi) =====
    if seller_flag:
        _chans = db.get_seller_channels(user_id)
        if _chans:
            _lines = [t(lang, 'adu_channels_header')]
            for _ch in _chans:
                _ttl = _ch.get('channel_title') or _ch.get('channel_id')
                _lines.append(f"• {html.escape(str(_ttl))}")
            text += "\n\n" + "\n".join(_lines)
        else:
            text += "\n\n" + t(lang, 'adu_channels_none')

    # ===== To'lov kartasi =====
    cnum = user.get('card_number')
    if cnum:
        cnum_str = str(cnum)
        grouped = " ".join(cnum_str[i:i + 4] for i in range(0, len(cnum_str), 4))
        ctype = CARD_TYPE_LABELS.get(user.get('card_type', ''), '💳')
        text += t(lang, 'adu_card_block', ctype=ctype, num=grouped,
                  owner=html.escape(user.get('card_owner') or _none))

    # ===== Referal =====
    text += t(lang, 'adu_referral_block',
              code=user.get('referral_code') or '—', by=user.get('referred_by') or '—',
              count=user.get('referral_count') or 0)

    # ===== Faollik statistikasi =====
    try:
        buyer_orders = db.get_orders_by_buyer(user_id) or []
    except Exception:
        buyer_orders = []
    text += t(lang, 'adu_activity_block', n=len(buyer_orders))
    if seller_flag:
        try:
            st = db.get_seller_stats(user_id) or {}
        except Exception:
            st = {}
        try:
            avg = db.get_seller_avg_rating(user_id) or 0.0
        except Exception:
            avg = 0.0
        text += t(lang, 'adu_seller_activity',
                  total=st.get('total_orders', 0), delivered=st.get('delivered', 0),
                  pending=st.get('pending', 0), cancelled=st.get('cancelled', 0),
                  products=st.get('products_count', 0),
                  revenue=fmt_price(st.get('total_revenue', 0)), rating=f"{avg:.1f}")

    await query.edit_message_text(
        text,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode='HTML',
        disable_web_page_preview=True,
    )


async def admin_block_user(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    user_id = int(query.data.split("_")[2])

    is_blocked = db.get_user_is_blocked(user_id)
    if is_blocked is not None:
        if is_blocked:
            db.unblock_user(user_id)
        else:
            db.block_user(user_id)

    # query.data ni mutatsiya qilmaymiz — user_id ni to'g'ridan-to'g'ri uzatamiz
    await admin_user_details(update, context, user_id=user_id)


async def admin_verify_user(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Sotuvchi uchun is_approved toggle — do'konni muzlatish/qayta tasdiqlash."""
    query = update.callback_query
    await query.answer()

    user_id = int(query.data.split("_")[2])
    user = db.get_user_by_id(user_id)
    lang = get_lang(update, context)

    if not user:
        await query.edit_message_text(t(lang, 'user_not_found'))
        return

    is_seller = bool(user.get('shop_name'))

    if is_seller:
        if user.get('is_approved'):
            # Tasdiqlashni bekor qilish — do'konni muzlatish
            db.update_user(user_id, is_approved=0)
            # Sotuvchiga xabar (sotuvchi tilida)
            try:
                await context.bot.send_message(
                    chat_id=user['telegram_id'],
                    text=t(user, 'seller_frozen_notify'),
                    parse_mode='HTML'
                )
            except Exception as e:
                logging.error(f"Sotuvchiga tasdiqlash bekor xabari ketmadi: {e}")
        else:
            # Qayta tasdiqlash — do'konni ochish
            db.update_user(user_id, is_approved=1, role='seller')
            # seller_requests jadvalini yangilash
            req = db.get_seller_request_by_user(user_id)
            if req:
                db.update_seller_request(req['id'], 'approved')
            # Sotuvchiga xabar (sotuvchi tilida)
            try:
                await context.bot.send_message(
                    chat_id=user['telegram_id'],
                    text=t(user, 'seller_reactivated_notify'),
                    parse_mode='HTML'
                )
            except Exception as e:
                logging.error(f"Sotuvchiga qayta tasdiqlash xabari ketmadi: {e}")
    else:
        # Oddiy foydalanuvchi uchun is_verified toggle (eski xatti-harakat)
        is_verified = db.get_user_is_verified(user_id)
        if is_verified is not None:
            db.update_user(user_id, is_verified=0 if is_verified else 1)

    await admin_user_details(update, context, user_id=user_id)


async def admin_request_fill(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin: profildagi kamchiliklarni AI tahlil qilib, foydalanuvchiga yuboriladigan
    xabarni TAKLIF qiladi (hali yubormaydi). "Qayta yaratish" shu handlerni qayta chaqiradi."""
    query = update.callback_query
    user_id = int(query.data.split("_")[2])
    lang = get_lang(update, context)
    user = db.get_user_by_id(user_id)
    if not user:
        await query.answer(t(lang, 'user_not_found'), show_alert=True)
        return
    if not ai_assistant.is_enabled():
        await query.answer(t(lang, 'fill_ai_off'), show_alert=True)
        return

    user_lang = user.get('language') or DEFAULT_LANG
    missing = _profile_missing_fields(user, user_lang)
    if not missing:
        await query.answer(t(lang, 'fill_none_missing'), show_alert=True)
        return

    await query.answer(t(lang, 'fill_generating'))
    is_seller = bool(user.get('shop_name')) or user.get('role') == 'seller'
    msg = await ai_assistant.generate_profile_completion_message(
        name=user.get('name') or '', missing_fields=missing,
        is_seller=is_seller, lang=user_lang
    )
    if not msg:
        await query.answer(t(lang, 'fill_ai_error'), show_alert=True)
        return

    # Aynan shu matn yuborilishi uchun saqlab qo'yamiz (admin chatiga bog'liq)
    context.user_data[f'_fill_msg_{user_id}'] = msg

    missing_list = "\n".join(f"• {m}" for m in missing)
    preview = t(lang, 'fill_preview',
                name=html.escape(user.get('name') or '—'),
                missing=html.escape(missing_list),
                msg=html.escape(msg))
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(t(lang, 'btn_fill_send'), callback_data=f"admin_sendfill_{user_id}")],
        [InlineKeyboardButton(t(lang, 'btn_fill_regen'), callback_data=f"admin_askfill_{user_id}")],
        [InlineKeyboardButton(t(lang, 'back'), callback_data=f"admin_user_{user_id}")],
    ])
    await query.edit_message_text(preview, reply_markup=kb,
                                  parse_mode='HTML', disable_web_page_preview=True)


async def admin_send_fill(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin tasdiqladi — AI taklif qilgan xabarni foydalanuvchiga yuboradi."""
    query = update.callback_query
    user_id = int(query.data.split("_")[2])
    lang = get_lang(update, context)
    user = db.get_user_by_id(user_id)
    if not user:
        await query.answer(t(lang, 'user_not_found'), show_alert=True)
        return

    msg = context.user_data.get(f'_fill_msg_{user_id}')
    if not msg:
        await query.answer(t(lang, 'fill_expired'), show_alert=True)
        return

    sent_ok = False
    try:
        await context.bot.send_message(chat_id=user['telegram_id'], text=msg)
        sent_ok = True
    except Exception as e:
        logging.error(f"Profil to'ldirish xabari foydalanuvchiga ({user_id}) ketmadi: {e}")

    context.user_data.pop(f'_fill_msg_{user_id}', None)
    await query.answer(t(lang, 'fill_sent_ok') if sent_ok else t(lang, 'fill_send_failed'),
                       show_alert=True)
    await admin_user_details(update, context, user_id=user_id)


ADMIN_PRODUCTS_PAGE_SIZE = 10
_PROD_STATUS_EMOJI = {'active': '✅', 'reserve': '📦', 'deleted': '🗑'}


def _admin_product_rows(products):
    rows = []
    for p in products:
        st = _PROD_STATUS_EMOJI.get(p.get('status') or 'active', '✅')
        rows.append([InlineKeyboardButton(
            f"{st} {p['name']} — {fmt_price(p['price'])}",
            callback_data=f"admin_prod_{p['id']}"
        )])
    return rows


async def _admin_guard(update, context) -> bool:
    """Chaqiruvchi admin bo'lsa True qaytaradi. Aks holda ogohlantirib False qaytaradi.

    XAVFSIZLIK: Pastdagi admin_product_* handler'lari ALOHIDA CallbackQueryHandler
    sifatida ro'yxatdan o'tgan (12568+ qatorlar) — ya'ni umumiy `button_handler`
    ichidagi admin-gate'ni CHETLAB o'tadi. Shu sababli admin tekshiruvini har bir
    shunday handler ICHIDA bajaramiz, aks holda istalgan foydalanuvchi callback
    yuborib mahsulot o'chirishi/ko'rishi mumkin edi."""
    uid = update.effective_user.id if update.effective_user else None
    u = db.get_user_by_telegram_id(uid) if uid else None
    if u and (u.get('role') == 'admin' or uid == ADMIN_ID):
        return True
    q = update.callback_query
    if q:
        try:
            await q.answer(T(update, context, 'admin_only_action'), show_alert=True)
        except Exception:
            pass
    return False


async def admin_products(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _admin_guard(update, context):
        return
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    context.user_data.pop('admin_searching_product', None)

    page = 0
    if query.data.startswith("admin_products_pg_"):
        try:
            page = int(query.data.replace("admin_products_pg_", ""))
        except ValueError:
            page = 0

    products = db.get_all_products(include_hidden=False)
    total = len(products)
    if not products:
        await query.edit_message_text(
            t(lang, 'admin_no_products'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="admin_panel")]])
        )
        return

    total_pages = (total + ADMIN_PRODUCTS_PAGE_SIZE - 1) // ADMIN_PRODUCTS_PAGE_SIZE
    page = max(0, min(page, total_pages - 1))
    start = page * ADMIN_PRODUCTS_PAGE_SIZE
    chunk = products[start:start + ADMIN_PRODUCTS_PAGE_SIZE]

    keyboard = _admin_product_rows(chunk)
    keyboard.append([InlineKeyboardButton(t(lang, 'btn_search_product'), callback_data="admin_product_search")])
    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton("⬅️", callback_data=f"admin_products_pg_{page-1}"))
    if page < total_pages - 1:
        nav.append(InlineKeyboardButton("➡️", callback_data=f"admin_products_pg_{page+1}"))
    if nav:
        keyboard.append(nav)
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="admin_panel")])

    await query.edit_message_text(
        t(lang, 'admin_products_title', n=total),
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def admin_product_search(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _admin_guard(update, context):
        return
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    context.user_data['admin_searching_product'] = True
    await query.edit_message_text(
        t(lang, 'admin_product_search_ask'),
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'btn_cancel'), callback_data="admin_products")]])
    )


async def admin_product_search_result(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.pop('admin_searching_product', None)
    lang = get_lang(update, context)
    search_text = (update.message.text or "").strip()
    if len(search_text) < 2:
        await update.message.reply_text(
            t(lang, 'admin_search_min2'),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(t(lang, 'btn_retry_search_product'), callback_data="admin_product_search")],
                [InlineKeyboardButton(t(lang, 'back'), callback_data="admin_products")],
            ])
        )
        return
    results = db.admin_search_products(search_text)
    if not results:
        await update.message.reply_text(
            t(lang, 'admin_product_search_none', q=html.escape(search_text)),
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(t(lang, 'btn_retry_search_product'), callback_data="admin_product_search")],
                [InlineKeyboardButton(t(lang, 'back'), callback_data="admin_products")],
            ])
        )
        return
    keyboard = _admin_product_rows(results[:20])
    keyboard.append([InlineKeyboardButton(t(lang, 'btn_retry_search_product'), callback_data="admin_product_search")])
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="admin_products")])
    await update.message.reply_text(
        t(lang, 'admin_product_search_results', q=html.escape(search_text), n=len(results)),
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def admin_deleted_products(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin — o'chirilgan mahsulotlar jurnali (audit)."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    entries = db.get_product_audit(limit=50)
    if not entries:
        await query.edit_message_text(
            t(lang, 'no_deleted_products'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="admin_panel")]]))
        return
    keyboard = []
    for e in entries[:20]:
        icon = "🗑" if e.get('action') == 'deleted' else "📦"
        label = f"{icon} {(e.get('name') or '—')} — {(e.get('shop_name') or '')}"[:45]
        keyboard.append([InlineKeyboardButton(label, callback_data=f"admin_audit_{e['id']}")])
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="admin_panel")])
    await query.edit_message_text(
        t(lang, 'deleted_products_header', n=len(entries)),
        reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


async def admin_audit_detail(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Bitta o'chirilgan mahsulot yozuvining to'liq ma'lumoti."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)
    audit_id = int(query.data.split("_")[2])
    e = db.get_product_audit_entry(audit_id)
    if not e:
        await query.edit_message_text(
            t(lang, 'audit_not_found'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="admin_deleted_products")]]))
        return
    action_lbl = t(lang, 'audit_action_purged' if e.get('action') == 'purged' else 'audit_action_deleted')
    by_role = e.get('deleted_by_role')
    by_lbl = (t(lang, 'party_seller') if by_role == 'seller'
              else (t(lang, 'role_admin_word') if by_role == 'admin' else '—'))
    stock = e.get('stock_count')
    text = t(lang, 'audit_detail_body',
             name=html.escape(e.get('name') or '—'),
             price=fmt_price(e.get('price') or 0),
             cat=html.escape(category_name(e.get('category_name'), lang) or '—'),
             shop=html.escape(e.get('shop_name') or e.get('seller_name') or '—'),
             stock=(stock if stock is not None else '∞'),
             orders=e.get('order_count') or 0,
             action=action_lbl, by=by_lbl,
             byname=html.escape(e.get('deleted_by_name') or '—'),
             created=fmt_datetime(e.get('product_created_at')),
             deleted=fmt_datetime(e.get('deleted_at')))
    await query.edit_message_text(
        text,
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="admin_deleted_products")]]),
        parse_mode='HTML')


async def _admin_render_product(update: Update, context: ContextTypes.DEFAULT_TYPE, product_id: int):
    """Mahsulot detalini (matn) ko'rsatadi + admin amallari (olib tashlash/qaytarish/rasm)."""
    query = update.callback_query
    lang = get_lang(update, context)
    product = db.get_product_by_id(product_id)
    if not product:
        await query.edit_message_text(
            t(lang, 'product_not_found'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="admin_products")]])
        )
        return

    status = product.get('status') or 'active'
    status_lbls = {'active': 'status_on_sale', 'reserve': 'status_reserve_short', 'deleted': 'status_removed'}
    status_lbl = t(lang, status_lbls.get(status, 'status_on_sale'))

    region_lbl = region_label_l(product.get('seller_region_id'), lang) or '—'
    prod_cnt = product.get('prod_review_count') or 0
    prod_rating = product.get('prod_avg_rating') or 0
    rating_txt = f"{prod_rating:.1f} ({prod_cnt})" if prod_cnt else '—'
    desc = (product.get('description') or '').strip() or '—'
    _uname = product.get('telegram_username')
    seller_disp = f"@{html.escape(str(_uname))}" if _uname else '—'

    text = t(lang, 'admin_product_body',
             name=html.escape(product.get('name') or '—'),
             price=fmt_price(product.get('price')),
             cat=html.escape(str(product.get('category_name') or '—')),
             status=status_lbl,
             shop=html.escape(str(product.get('shop_name') or '—')),
             seller=seller_disp,
             region=html.escape(str(region_lbl)),
             rating=rating_txt,
             created=fmt_datetime(product.get('created_at')),
             pid=product_id,
             desc=html.escape(desc))

    buttons = []
    if status == 'active':
        buttons.append([InlineKeyboardButton(t(lang, 'btn_remove_from_sale'), callback_data=f"admin_prodrm_{product_id}")])
    else:
        buttons.append([InlineKeyboardButton(t(lang, 'btn_return_to_sale'), callback_data=f"admin_prodrs_{product_id}")])
    if product.get('image_url'):
        buttons.append([InlineKeyboardButton(t(lang, 'btn_view_image'), callback_data=f"admin_prodimg_{product_id}")])
    buttons.append([InlineKeyboardButton(t(lang, 'back'), callback_data="admin_products")])

    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(buttons), parse_mode='HTML')


async def admin_product_detail(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _admin_guard(update, context):
        return
    query = update.callback_query
    await query.answer()
    product_id = int(query.data.rsplit("_", 1)[1])
    await _admin_render_product(update, context, product_id)


async def admin_product_remove(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _admin_guard(update, context):
        return
    query = update.callback_query
    product_id = int(query.data.rsplit("_", 1)[1])
    db.set_product_status(product_id, 'deleted')
    await query.answer(t(get_lang(update, context), 'admin_product_removed'), show_alert=False)
    await _admin_render_product(update, context, product_id)


async def admin_product_restore(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _admin_guard(update, context):
        return
    query = update.callback_query
    product_id = int(query.data.rsplit("_", 1)[1])
    db.set_product_status(product_id, 'active')
    await query.answer(t(get_lang(update, context), 'admin_product_restored'), show_alert=False)
    await _admin_render_product(update, context, product_id)


async def admin_product_image(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _admin_guard(update, context):
        return
    query = update.callback_query
    product_id = int(query.data.rsplit("_", 1)[1])
    product = db.get_product_by_id(product_id)
    lang = get_lang(update, context)
    photo = product.get('image_url') if product else None
    if not photo:
        await query.answer(t(lang, 'no_image'), show_alert=True)
        return
    await query.answer()
    try:
        await context.bot.send_photo(
            chat_id=update.effective_chat.id, photo=photo,
            caption=f"🖼 {html.escape(product.get('name') or '')}"
        )
    except Exception as e:
        logging.error(f"admin_product_image failed: {e}")


async def admin_orders(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    lang = get_lang(update, context)
    orders = db.get_all_orders()

    if not orders:
        await query.edit_message_text(
            t(lang, 'admin_no_orders'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="admin_panel")]])
        )
        return

    status_emoji = {'pending': '⏳', 'confirmed': '✅', 'delivered': '🚚', 'cancelled': '❌'}
    keyboard = [[InlineKeyboardButton(
        f"{status_emoji.get(o['status'], '❓')} #{o['id']} — {fmt_price(o['total_price'])}",
        callback_data=f"admin_order_{o['id']}"
    )] for o in orders[:20]]
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="admin_panel")])

    await query.edit_message_text(
        t(lang, 'admin_orders_title', n=len(orders)),
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def admin_order_detail(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    order_id = int(query.data.split("_")[2])
    order = db.get_order_by_id(order_id)
    lang = get_lang(update, context)

    if not order:
        await query.edit_message_text(t(lang, 'order_not_found'))
        return

    dlv = order.get('delivery_type') or 'delivery'
    pay = order.get('payment_method') or '—'

    text = t(lang, 'admin_order_body',
             oid=fmt_order_id(order_id), pname=html.escape(order.get('product_name') or ''),
             qty=order.get('quantity'), total=fmt_price(order.get('total_price')),
             status=status_label(order.get('status'), lang),
             buyer=html.escape(order.get('buyer_name') or ''), buyer_phone=order.get('buyer_phone') or '—',
             seller=html.escape(order.get('seller_name') or ''), seller_phone=order.get('seller_phone') or '—',
             dlv=dlv_label(dlv, lang), pay=pay_label(pay, lang),
             date=fmt_datetime(order.get('created_at')))
    if order.get('delivery_address'):
        text += t(lang, 'admin_order_addr', addr=html.escape(order.get('delivery_address') or ''))

    keyboard = []
    if order.get('status') == 'pending':
        keyboard.append([
            InlineKeyboardButton(t(lang, 'btn_confirm'), callback_data=f"confirm_order_{order_id}"),
            InlineKeyboardButton(t(lang, 'btn_cancel'), callback_data=f"cancel_order_{order_id}"),
        ])
    elif order.get('status') == 'confirmed':
        keyboard.append([InlineKeyboardButton(t(lang, 'btn_delivered'), callback_data=f"deliver_order_{order_id}")])

    if order.get('status') not in ('cancelled',):
        keyboard.append([InlineKeyboardButton(
            t(lang, 'btn_force_cancel'), callback_data=f"admin_force_cancel_{order_id}"
        )])

    keyboard.append([InlineKeyboardButton(t(lang, 'btn_send_message'), callback_data=f"order_msg_{order_id}")])
    keyboard.append([InlineKeyboardButton(t(lang, 'btn_correspondence'), callback_data=f"msgs_{order_id}")])
    # Nizo yozishmalari bo'lsa — audit ko'rinishi (nizo hal bo'lgandan keyin ham)
    if db.count_dispute_messages(order_id) > 0:
        keyboard.append([InlineKeyboardButton(t(lang, 'btn_dispute_messages'), callback_data=f"admin_dispmsgs_{order_id}")])
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="admin_orders")])

    await query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(keyboard), parse_mode='HTML')


async def admin_force_cancel_order(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin istalgan buyurtmani majburan bekor qiladi."""
    query = update.callback_query
    await query.answer()

    order_id = int(query.data.split("_")[3])
    order = db.get_order_by_id(order_id)

    if order:
        # ATOMIK bekor: 'confirmed' bo'lsa zahirani QAYTARAMIZ (kamaytirilgan edi),
        # 'pending' bo'lsa shart emas. Avval restock yo'q edi → zahira yo'qolardi (bug).
        if db.transition_order_status(order_id, 'cancelled', 'confirmed', cancel_by='admin'):
            await _maybe_restock_on_cancel(context, order)
        else:
            db.transition_order_status(order_id, 'cancelled', 'pending', cancel_by='admin')
        # Ikki tomonga xabar (har biri o'z tilida)
        oid = fmt_order_id(order_id)
        buyer = db.get_user_by_id(order['buyer_id']) if order.get('buyer_id') else None
        seller = db.get_user_by_id(order['seller_id']) if order.get('seller_id') else None
        for tg_id, u in [(order.get('buyer_tg'), buyer), (order.get('seller_tg'), seller)]:
            if tg_id:
                try:
                    await context.bot.send_message(
                        chat_id=tg_id,
                        text=t(u or 'uz', 'admin_cancel_notify', oid=oid)
                    )
                except Exception:
                    pass

    await admin_orders(update, context)


ADMIN_CHANNELS_PAGE_SIZE = 8  # sahifada nechta SOTUVCHI ko'rsatiladi


async def admin_channels(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Admin — qaysi sotuvchi qaysi kanal(lar)ni ulagani. Sotuvchi bo'yicha guruhlangan, sahifalangan."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)

    # Sahifa raqami: 'admin_channels' yoki 'admin_channels_pg_N'
    page = 0
    if query.data.startswith("admin_channels_pg_"):
        try:
            page = int(query.data.replace("admin_channels_pg_", ""))
        except ValueError:
            page = 0

    rows = db.get_all_seller_channels()

    if not rows:
        await query.edit_message_text(
            t(lang, 'admin_channels_none'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="admin_panel")]]),
            parse_mode='HTML',
        )
        return

    # Sotuvchi bo'yicha guruhlaymiz (rows allaqachon shop_name -> created_at bo'yicha tartiblangan)
    grouped = {}
    order = []
    for r in rows:
        sid = r['seller_id']
        if sid not in grouped:
            grouped[sid] = []
            order.append(sid)
        grouped[sid].append(r)

    # Sotuvchilar bo'yicha sahifalash
    total_sellers = len(order)
    total_pages = max(1, (total_sellers + ADMIN_CHANNELS_PAGE_SIZE - 1) // ADMIN_CHANNELS_PAGE_SIZE)
    page = max(0, min(page, total_pages - 1))
    start = page * ADMIN_CHANNELS_PAGE_SIZE
    page_sellers = order[start:start + ADMIN_CHANNELS_PAGE_SIZE]

    lines = [t(lang, 'admin_channels_title', sellers=total_sellers, channels=len(rows)),
             t(lang, 'admin_channels_page', page=page + 1, pages=total_pages)]
    for sid in page_sellers:
        chans = grouped[sid]
        head = chans[0]
        shop = head.get('shop_name') or head.get('seller_name') or t(lang, 'anonymous')
        uname = head.get('telegram_username')
        uname_txt = f" (@{html.escape(uname)})" if uname else ""
        approved = "✅" if head.get('is_approved') else "⏳"
        lines.append(f"\n{approved} <b>{html.escape(str(shop))}</b>{uname_txt} — ID {sid}")
        for c in chans:
            ttl = c.get('channel_title') or c.get('channel_id')
            cmark = "📢" if c.get('is_active', 1) else "⚠️"
            lines.append(f"   {cmark} {html.escape(str(ttl))}  <code>{html.escape(str(c.get('channel_id')))}</code>")

    text = "\n".join(lines)
    if len(text) > 4000:
        text = text[:3900] + "\n\n…"

    # Navigatsiya tugmalari
    keyboard = []
    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton("⬅️", callback_data=f"admin_channels_pg_{page-1}"))
    nav.append(InlineKeyboardButton(f"{page+1}/{total_pages}", callback_data="noop"))
    if page < total_pages - 1:
        nav.append(InlineKeyboardButton("➡️", callback_data=f"admin_channels_pg_{page+1}"))
    if len(nav) > 1:
        keyboard.append(nav)
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data="admin_panel")])

    await query.edit_message_text(
        text,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode='HTML',
        disable_web_page_preview=True,
    )


async def admin_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    lang = get_lang(update, context)
    s = db.get_admin_stats_summary()
    top = t(lang, 'top_seller_fmt', name=s['top_seller'], n=s['top_seller_count']) if s['top_seller'] else "—"

    await query.edit_message_text(
        t(lang, 'admin_stats_body',
          total_users=s['total_users'], buyers=s['buyers'], sellers=s['sellers'],
          products=s['products'], total_orders=s['total_orders'],
          pending=s['pending'], confirmed=s['confirmed'], delivered=s['delivered'], cancelled=s['cancelled'],
          today_rev=fmt_price(s['today_revenue']), today_cnt=s['today_count'],
          week_rev=fmt_price(s['week_revenue']), week_cnt=s['week_count'],
          month_rev=fmt_price(s['month_revenue']), month_cnt=s['month_count'],
          total_rev=fmt_price(s['total_revenue']), top=top),
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton(t(lang, 'btn_conversion_funnel'), callback_data="admin_analytics")],
            [InlineKeyboardButton(t(lang, 'btn_financial_report'), callback_data="admin_revenue")],
            [InlineKeyboardButton(t(lang, 'back'), callback_data="admin_panel")],
        ]),
        parse_mode='HTML'
    )


async def admin_analytics(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Conversion funnel va batafsil analytics."""
    query = update.callback_query
    await query.answer()

    lang = get_lang(update, context)
    a = db.get_analytics_funnel()

    # Funnel vizualizatsiya
    funnel_bar = t(lang, 'analytics_funnel_bar',
                   b1='█' * 20,
                   b2='█' * max(1, int(20 * a['confirm_rate'] / 100)),
                   b3='█' * max(1, int(20 * a['deliver_rate'] / 100)),
                   b4='░' * max(1, int(20 * a['cancel_rate'] / 100)),
                   issued=a.get('total_issued', a['total_orders']),
                   total=a['total_orders'], confirmed=a['confirmed_total'], confirm_rate=a['confirm_rate'],
                   delivered=a['delivered_total'], deliver_rate=a['deliver_rate'],
                   cancelled=a['cancelled_total'], cancel_rate=a['cancel_rate'])

    # Peak soatlar
    peak_hours_text = ""
    if a['peak_hours']:
        hours = [t(lang, 'analytics_peak_hour', h=h, cnt=cnt) for h, cnt in a['peak_hours'][:3]]
        peak_hours_text = ", ".join(hours)

    # Peak kunlar
    peak_days_text = ""
    if a['peak_days']:
        days = [t(lang, 'analytics_peak_day', d=d, cnt=cnt) for d, cnt in a['peak_days'][:3]]
        peak_days_text = ", ".join(days)

    # Top kategoriyalar
    top_cats_text = ""
    if a['top_categories']:
        for emoji, name, cnt in a['top_categories']:
            top_cats_text += t(lang, 'analytics_top_cat', emoji=emoji, name=name, cnt=cnt)

    # Top mahsulotlar
    top_prods_text = ""
    if a['top_products']:
        for i, (name, cnt, rev) in enumerate(a['top_products'], 1):
            top_prods_text += t(lang, 'analytics_top_prod', i=i, name=name, cnt=cnt, rev=fmt_price(rev))

    text = t(lang, 'analytics_body',
             funnel=funnel_bar,
             week_orders=a['week_orders'], week_confirmed=a['week_confirmed'], week_confirm_rate=a['week_confirm_rate'],
             week_delivered=a['week_delivered'], week_deliver_rate=a['week_deliver_rate'], week_cancelled=a['week_cancelled'],
             month_orders=a['month_orders'], month_confirmed=a['month_confirmed'], month_confirm_rate=a['month_confirm_rate'],
             month_delivered=a['month_delivered'], month_deliver_rate=a['month_deliver_rate'], month_cancelled=a['month_cancelled'],
             avg_order=fmt_price(a['avg_order_value']),
             new_week=a['new_users_week'], new_month=a['new_users_month'],
             peak_hours=peak_hours_text, peak_days=peak_days_text,
             top_cats=top_cats_text, top_prods=top_prods_text)

    if len(text) > 4000:
        text = text[:3900] + "\n\n…"

    await query.edit_message_text(
        text,
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton(t(lang, 'btn_general_stats'), callback_data="admin_stats")],
            [InlineKeyboardButton(t(lang, 'back'), callback_data="admin_panel")],
        ]),
        parse_mode='HTML'
    )


async def admin_revenue(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Batafsil moliyaviy hisobot — sotuvchi bo'yicha."""
    query = update.callback_query
    await query.answer()

    lang = get_lang(update, context)
    orders = db.get_all_orders()
    delivered = [o for o in orders if o['status'] == 'delivered']

    if not delivered:
        await query.edit_message_text(
            t(lang, 'revenue_no_delivered'),
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton(t(lang, 'back'), callback_data="admin_stats")
            ]])
        )
        return

    from collections import defaultdict
    from datetime import datetime, timedelta, timezone

    now = datetime.now(timezone.utc)
    month_ago = (now - timedelta(days=30)).strftime('%Y-%m-%d')

    # Sotuvchi bo'yicha hisobot
    seller_stats = defaultdict(lambda: {'count': 0, 'revenue': 0, 'month_count': 0, 'month_revenue': 0})
    for o in delivered:
        sname = o.get('seller_name') or t(lang, 'unknown_seller')
        price = float(o.get('total_price') or 0)
        seller_stats[sname]['count'] += 1
        seller_stats[sname]['revenue'] += price
        if str(o.get('created_at') or '')[:10] >= month_ago:
            seller_stats[sname]['month_count'] += 1
            seller_stats[sname]['month_revenue'] += price

    # To'lov usuli bo'yicha
    pay_stats = defaultdict(int)
    for o in delivered:
        pay = o.get('payment_method') or 'cash'
        pay_stats[pay_label(pay, lang)] += 1

    # Yetkazish turi bo'yicha
    dlv_stats = defaultdict(int)
    for o in delivered:
        dlv = o.get('delivery_type') or 'delivery'
        dlv_stats[dlv_label(dlv, lang)] += 1

    total = sum(s['revenue'] for s in seller_stats.values())
    month_total = sum(s['month_revenue'] for s in seller_stats.values())

    # Top 5 sotuvchi
    top5 = sorted(seller_stats.items(), key=lambda x: x[1]['revenue'], reverse=True)[:5]
    sellers_text = ""
    for i, (name, s) in enumerate(top5, 1):
        sellers_text += t(lang, 'revenue_seller_line', i=i, name=name,
                          revenue=fmt_price(s['revenue']), count=s['count'])

    pay_text = "\n".join(t(lang, 'revenue_pay_line', label=k, n=v) for k, v in pay_stats.items())
    dlv_text = "\n".join(t(lang, 'revenue_pay_line', label=k, n=v) for k, v in dlv_stats.items())

    text = t(lang, 'revenue_body',
             month_total=fmt_price(month_total),
             month_count=sum(s['month_count'] for s in seller_stats.values()),
             total=fmt_price(total), delivered_count=len(delivered),
             sellers=sellers_text, pay=pay_text, dlv=dlv_text)

    if len(text) > 4000:
        text = text[:3900] + "\n\n…"

    await query.edit_message_text(
        text,
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton(t(lang, 'btn_excel_report'), callback_data="excel_orders")],
            [InlineKeyboardButton(t(lang, 'back'), callback_data="admin_stats")],
        ]),
        parse_mode='HTML'
    )


async def admin_broadcast_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    context.user_data['broadcasting'] = True
    await query.edit_message_text(T(update, context, 'broadcast_ask'))


async def admin_settings(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    users = db.get_all_users()
    products = db.get_all_products(include_hidden=False)
    orders = db.get_all_orders()

    lang = get_lang(update, context)
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(t(lang, 'btn_db_backup'), callback_data="admin_backup")],
        [InlineKeyboardButton(t(lang, 'btn_excel_users'), callback_data="excel_users")],
        [InlineKeyboardButton(t(lang, 'btn_excel_products'), callback_data="excel_products")],
        [InlineKeyboardButton(t(lang, 'btn_excel_orders'), callback_data="excel_orders")],
        [InlineKeyboardButton(t(lang, 'btn_clean_cancelled'), callback_data="admin_clean_cancelled")],
        [InlineKeyboardButton(t(lang, 'btn_change_language'), callback_data="change_lang")],
        [InlineKeyboardButton(t(lang, 'back'), callback_data="admin_panel")],
    ])

    await query.edit_message_text(
        t(lang, 'admin_settings_body', admin_id=ADMIN_ID,
          users=len(users), products=len(products), orders=len(orders)),
        reply_markup=kb,
        parse_mode='HTML'
    )


async def admin_export_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Excel eksport — foydalanuvchilar, mahsulotlar yoki buyurtmalar."""
    query = update.callback_query
    lang = get_lang(update, context)
    ru = (lang == 'ru')
    await query.answer(t(lang, 'excel_preparing'))

    export_type = query.data.replace("excel_", "")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import datetime as dt
        import os

        wb = openpyxl.Workbook()
        ws = wb.active

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1a8a2e")
        header_align = Alignment(horizontal="center", vertical="center")

        def style_header(row):
            for cell in row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

        def auto_width(ws):
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=0)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

        ts = dt.datetime.now().strftime("%d.%m.%Y %H:%M")

        if export_type == "users":
            ws.title = "Пользователи" if ru else "Foydalanuvchilar"
            headers = (["ID", "Telegram ID", "Имя", "Телефон", "Роль",
                        "Магазин", "ID региона", "Рейтинг", "Заказы", "Дата регистрации"]
                       if ru else
                       ["ID", "Telegram ID", "Ism", "Telefon", "Rol",
                        "Do'kon", "Hudud ID", "Reyting", "Buyurtmalar", "Ro'yxat sanasi"])
            ws.append(headers)
            style_header(ws[1])

            users = db.get_all_users()
            for u in users:
                avg = db.get_seller_avg_rating(u['id'])
                conn = db.get_connection()
                cur = conn.cursor()
                cur.execute("SELECT COUNT(*) FROM orders WHERE buyer_id=?", (u['id'],))
                order_count = cur.fetchone()[0]
                region = db.get_region_by_id(u['region_id']) if u.get('region_id') else None
                if region:
                    if region.get('parent_id'):
                        parent = db.get_region_by_id(region['parent_id'])
                        region_name = f"{parent['name']} → {region['name']}" if parent else region['name']
                    else:
                        region_name = region['name']
                else:
                    region_name = ''
                ws.append([
                    u['id'], u['telegram_id'], u.get('name') or '',
                    u.get('phone_number') or '', u.get('role') or '',
                    u.get('shop_name') or '', region_name,
                    round(avg, 1), order_count,
                    str(u.get('created_at') or '')[:10]
                ])
            filename = f"tezbozor_users_{dt.datetime.now().strftime('%Y%m%d')}.xlsx"
            caption = (f"👥 Список пользователей\n{len(users)} · {ts}" if ru
                       else f"👥 Foydalanuvchilar ro'yxati\n{len(users)} ta · {ts}")

        elif export_type == "products":
            ws.title = "Товары" if ru else "Mahsulotlar"
            headers = (["ID", "Продавец", "Категория", "Название", "Цена",
                        "Статус", "Остаток", "Рейтинг", "Дата добавления"]
                       if ru else
                       ["ID", "Sotuvchi", "Kategoriya", "Nom", "Narx",
                        "Holat", "Zahira", "Reyting", "Qo'shilgan sana"])
            ws.append(headers)
            style_header(ws[1])

            products = db.get_all_products()
            for p in products:
                status = p.get('status') or ('active' if p.get('in_stock') else 'reserve')
                if ru:
                    status_lbl = {'active': 'В продаже', 'reserve': 'В резерве', 'deleted': 'Удалён'}.get(status, status)
                else:
                    status_lbl = {'active': 'Sotuvda', 'reserve': 'Zahirada', 'deleted': "O'chirilgan"}.get(status, status)
                ws.append([
                    p['id'], p.get('seller_name') or p.get('seller_id') or '',
                    p.get('category_name') or '',
                    p.get('name') or '', p.get('price') or 0,
                    status_lbl,
                    p.get('stock_count') if p.get('stock_count') is not None else ('Без лимита' if ru else 'Cheksiz'),
                    round(p.get('avg_rating') or 0, 1),
                    str(p.get('created_at') or '')[:10]
                ])
            filename = f"tezbozor_products_{dt.datetime.now().strftime('%Y%m%d')}.xlsx"
            caption = (f"📦 Список товаров\n{len(products)} · {ts}" if ru
                       else f"📦 Mahsulotlar ro'yxati\n{len(products)} ta · {ts}")

        elif export_type == "orders":
            ws.title = "Заказы" if ru else "Buyurtmalar"
            headers = (["ID", "Покупатель", "Продавец", "Товар", "Цена",
                        "Кол-во", "Итого", "Статус", "Оплата", "Доставка", "Дата",
                        "Статус оплаты", "Оплачено", "Остаток долга"]
                       if ru else
                       ["ID", "Xaridor", "Sotuvchi", "Mahsulot", "Narx",
                        "Miqdor", "Jami", "Holat", "To'lov", "Yetkazish", "Sana",
                        "To'lov holati", "To'langan", "Qolgan qarz"])
            ws.append(headers)
            style_header(ws[1])

            settlement_lbl_map = ({'paid': 'Оплачено', 'debt': 'Долг', 'installment': 'Рассрочка'}
                                  if ru else
                                  {'paid': "To'langan", 'debt': 'Qarz', 'installment': "Bo'lib to'lash"})
            orders = db.get_all_orders()
            for o in orders:
                if ru:
                    status_lbl = {'pending': 'В ожидании', 'confirmed': 'Подтверждён',
                                  'delivered': 'Доставлен', 'cancelled': 'Отменён'}.get(o.get('status') or '', o.get('status') or '')
                else:
                    status_lbl = {'pending': 'Kutilmoqda', 'confirmed': 'Tasdiqlangan',
                                  'delivered': 'Yetkazildi', 'cancelled': 'Bekor'}.get(o.get('status') or '', o.get('status') or '')
                stt = o.get('settlement_type') or ''
                ws.append([
                    o['id'],
                    o.get('buyer_name') or '', o.get('seller_name') or '',
                    o.get('product_name') or '', o.get('product_price') or 0,
                    o.get('quantity') or 0, o.get('total_price') or 0,
                    status_lbl,
                    o.get('payment_method') or '', o.get('delivery_type') or '',
                    str(o.get('created_at') or '')[:10],
                    settlement_lbl_map.get(stt, '—' if not stt else stt),
                    o.get('amount_paid') if o.get('amount_paid') is not None else '',
                    o.get('amount_due') if o.get('amount_due') is not None else '',
                ])
            filename = f"tezbozor_orders_{dt.datetime.now().strftime('%Y%m%d')}.xlsx"
            caption = (f"🛒 Список заказов\n{len(orders)} · {ts}" if ru
                       else f"🛒 Buyurtmalar ro'yxati\n{len(orders)} ta · {ts}")
        else:
            await query.message.reply_text(t(lang, 'unknown_export'))
            return

        auto_width(ws)
        wb.save(filename)

        await context.bot.send_document(
            chat_id=update.effective_chat.id,
            document=open(filename, 'rb'),
            filename=filename,
            caption=caption
        )
        os.remove(filename)

    except ImportError:
        await query.message.reply_text(
            t(lang, 'excel_not_installed'),
            parse_mode='HTML'
        )
    except Exception as e:
        logging.error(f"Excel eksport xatosi: {e}")
        await query.message.reply_text(t(lang, 'error_generic', e=e))


async def admin_clean_cancelled(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Bekor qilingan eski (30 kundan oshgan) buyurtmalarni DB dan tozalaydi."""
    query = update.callback_query
    await query.answer()
    lang = get_lang(update, context)

    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        DELETE FROM orders
        WHERE status='cancelled'
        AND created_at < datetime('now', '-30 days')
    """)
    deleted = cursor.rowcount
    conn.commit()

    # Natijani alohida ekranda ko'rsatamiz (bitta javob — ikki marta answer qilmaymiz)
    kb = InlineKeyboardMarkup([[InlineKeyboardButton(t(lang, 'back'), callback_data="admin_settings")]])
    await query.edit_message_text(t(lang, 'clean_done', n=deleted), reply_markup=kb)


async def admin_backup(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """DB'ni admin'ga yuboradi — `.db` fayl sifatida."""
    query = update.callback_query
    lang = get_lang(update, context)
    await query.answer(t(lang, 'backup_preparing'))

    import datetime as dt
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = f"marketplace_backup_{ts}.db"

    ok = db.backup(backup_path)
    if not ok:
        await query.message.reply_text(t(lang, 'backup_failed'))
        return

    try:
        await context.bot.send_document(
            chat_id=update.effective_chat.id,
            document=open(backup_path, 'rb'),
            filename=backup_path,
            caption=t(lang, 'backup_caption', ts=ts)
        )
    except Exception as e:
        await query.message.reply_text(t(lang, 'backup_send_failed', e=e))
    finally:
        try:
            import os
            os.remove(backup_path)
        except Exception:
            pass


# ============================================================
# AI RECOMMENDATIONS
# ============================================================

async def share_product_link(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Sotuvchi mahsulot havolasini oladi — xaridorlarga ulashish uchun."""
    query = update.callback_query
    await query.answer()

    product_id = int(query.data.split("_")[2])
    lang = get_lang(update, context)
    product = db.get_product_basic(product_id)

    if not product:
        await query.answer(t(lang, 'product_not_found'), show_alert=True)
        return

    bot_me = await context.bot.get_me()
    bot_username = bot_me.username
    deep_link = _product_buy_link(bot_username, product_id)   # ulashish havolasi → Mini App

    from urllib.parse import quote
    share_text = f"{product.get('name', '')} — {fmt_price(product.get('price', 0))}"
    share_url = f"https://t.me/share/url?url={quote(deep_link, safe='')}&text={quote(share_text, safe='')}"

    text = t(lang, 'share_link_body', name=html.escape(product.get('name') or ''), link=deep_link)

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(t(lang, 'btn_share'), url=share_url)],
        [InlineKeyboardButton(t(lang, 'back'), callback_data=f"prod_menu_{product_id}")],
    ])

    await query.edit_message_text(text, reply_markup=kb, parse_mode='HTML')


async def show_recommendations(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Ko'rish tarixiga qarab tavsiya qilinadigan mahsulotlarni ko'rsatadi."""
    query = update.callback_query
    await query.answer()

    current_pid = int(query.data.split("_")[1])
    lang = get_lang(update, context)
    recs = _get_recommendations(context, db, current_pid, limit=10)

    if not recs:
        await query.answer(t(lang, 'recs_not_enough'), show_alert=True)
        return

    keyboard = []
    for p in recs:
        rating = p.get('prod_avg_rating') or p.get('avg_rating') or 0
        emoji = p.get('category_emoji') or '📦'
        keyboard.append([InlineKeyboardButton(
            f"{emoji} ⭐{rating:.1f} | {p['name']} — {fmt_price(p['price'])}",
            callback_data=f"prod_{p['id']}"
        )])
    keyboard.append([InlineKeyboardButton(t(lang, 'back'), callback_data=f"prod_{current_pid}")])

    await query.edit_message_text(
        t(lang, 'recs_title'),
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode='HTML'
    )


async def ai_recommendations(update: Update, context: ContextTypes.DEFAULT_TYPE):
    lang = get_lang(update, context)
    products = db.get_admin_products_summary(limit=10)

    if not products:
        await update.message.reply_text(t(lang, 'ai_no_products'))
        return

    keyboard = [[InlineKeyboardButton(
        f"🤖 {p['name']} — {fmt_price(p['price'])}",
        callback_data=f"prod_{p['id']}"
    )] for p in products]

    await update.message.reply_text(
        t(lang, 'ai_recs_title'),
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def recommend_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await ai_recommendations(update, context)


# ============================================================
# AI YORDAMCHI (DeepSeek)
# ============================================================

def _ai_exit_kb(lang):
    """AI rejimidan chiqish tugmasi."""
    return InlineKeyboardMarkup([[
        InlineKeyboardButton(t(lang, 'ai_exit'), callback_data="ai_exit")
    ]])


async def ai_assistant_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """AI suhbat rejimini boshlaydi (tugma yoki /ai orqali)."""
    lang = get_lang(update, context)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    if not user:
        msg = T(update, context, 'not_registered_short')
        if update.callback_query:
            await update.callback_query.answer(msg, show_alert=True)
        else:
            await update.message.reply_text(msg)
        return

    if not ai_assistant.is_enabled():
        if update.callback_query:
            await update.callback_query.answer(t(lang, 'ai_disabled_alert'), show_alert=True)
        else:
            await update.message.reply_text(t(lang, 'ai_disabled_alert'))
        return

    context.user_data['ai_chat'] = True
    context.user_data.pop('ai_draft', None)
    ai_assistant.reset_history(context.user_data)

    # Rolga moslangan xush kelibsiz matni (yangi imkoniyatlarni ko'rsatadi)
    role = user.get('role', 'buyer')
    if role != 'admin':
        role = get_active_mode(user, context)
    text = t(lang, f'ai_welcome_{role}') if role in ('seller', 'buyer', 'admin') else t(lang, 'ai_welcome')
    if update.callback_query:
        try:
            await update.callback_query.edit_message_text(
                text, parse_mode='HTML', reply_markup=_ai_exit_kb(lang)
            )
        except Exception:
            await update.callback_query.message.reply_text(
                text, parse_mode='HTML', reply_markup=_ai_exit_kb(lang)
            )
    else:
        await update.message.reply_text(
            text, parse_mode='HTML', reply_markup=_ai_exit_kb(lang)
        )


async def ai_assistant_exit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """AI suhbat rejimidan chiqadi va tegishli panelga qaytaradi."""
    lang = get_lang(update, context)
    context.user_data.pop('ai_chat', None)
    context.user_data.pop('ai_draft', None)
    context.user_data.pop('ai_awaiting_photos', None)
    context.user_data.pop('ai_product_photos', None)
    context.user_data.pop('ai_shop_filter', None)
    context.user_data.pop('ai_shop_name', None)
    ai_assistant.reset_history(context.user_data)

    if update.callback_query:
        try:
            await update.callback_query.edit_message_text(t(lang, 'ai_exited'))
        except Exception:
            pass

    user = db.get_user_by_telegram_id(update.effective_user.id)
    if user:
        if user['role'] == 'admin':
            await admin_panel(update, context)
        elif get_active_mode(user, context) == 'seller':
            await seller_panel(update, context)
        else:
            await buyer_panel(update, context)


async def ai_handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """AI agent rejimi — matnni bazaga ulangan agentga uzatadi va strukturalangan
    javobni (matn + mahsulot kartalari + e'lon qoralamasi) render qiladi."""
    lang = get_lang(update, context)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    role = user.get('role', 'buyer') if user else 'buyer'
    if role != 'admin':
        role = get_active_mode(user, context) if user else 'buyer'
    user_name = (user.get('name') if user else '') or ''
    seller_id = user['id'] if user else None

    # "Yozyapti..." indikatori
    try:
        await context.bot.send_chat_action(
            chat_id=update.effective_chat.id, action=ChatAction.TYPING
        )
    except Exception:
        pass

    # Do'kon ichida AI qidiruv — natijalar faqat shu do'kon bilan cheklanadi
    shop_filter = context.user_data.get('ai_shop_filter')
    shop_name = context.user_data.get('ai_shop_name') or ''

    result = await ai_assistant.ask(
        db, lang=lang, role=role, user_text=update.message.text,
        user_data=context.user_data, seller_id=seller_id, user_name=user_name,
        shop_filter=shop_filter, shop_name=shop_name,
    )

    text = (result.get('text') or '').strip() if isinstance(result, dict) else str(result)
    products = result.get('products') if isinstance(result, dict) else None
    draft = result.get('draft') if isinstance(result, dict) else None
    reactivated_id = result.get('reactivated_id') if isinstance(result, dict) else None
    order_actions = result.get('order_actions') if isinstance(result, dict) else None
    review_replies = result.get('review_replies') if isinstance(result, dict) else None

    # 1) Asosiy javob matni
    if text:
        await update.message.reply_text(text, reply_markup=_ai_exit_kb(lang))

    # 2) Topilgan real mahsulotlar — kartalar/tugmalar bilan
    if products:
        await _ai_send_products(update, context, lang, products)

    # 3) Sotuvchi uchun tayyor e'lon qoralamasi — joylash tugmasi bilan
    if draft:
        await _ai_send_draft(update, context, lang, draft)

    # 4) AI mahsulotni zahiradan sotuvga qaytardi — reklama ko'rinishini ko'rsatamiz
    if reactivated_id:
        await show_ad_preview(update, context, reactivated_id)

    # 5) Sotuvchi buyurtma amali — tasdiq tugmasi (mavjud confirm/cancel/deliver oqimini ishlatadi)
    if order_actions:
        for act in order_actions:
            await _ai_send_order_action(update, context, lang, act)

    # 6) AI tuzgan sharh javobi — e'lon qilish tugmasi bilan
    if review_replies:
        for rr in review_replies:
            await _ai_send_review_reply(update, context, lang, rr)


async def _ai_send_order_action(update, context, lang, act):
    """AI buyurtma amali uchun tasdiq kartasini ko'rsatadi. Tugma mavjud
    confirm_order_/cancel_order_/deliver_order_ oqimini ishga tushiradi —
    stok kamaytirish, xaridorga xabar va taymerlar shu yerda hal bo'ladi."""
    oid = act.get('order_id')
    action = act.get('action')
    if not oid or action not in ('confirm', 'deliver', 'cancel'):
        return
    btn_key = {'confirm': 'ai_order_btn_confirm',
               'deliver': 'ai_order_btn_deliver',
               'cancel': 'ai_order_btn_cancel'}[action]
    cb = f"{action}_order_{oid}"
    text = t(lang, 'ai_order_action_card',
             oid=fmt_order_id(oid),
             product=html.escape(act.get('product') or '—'),
             qty=act.get('qty') or 1,
             buyer=html.escape(act.get('buyer') or '—'),
             price=act.get('price_som') or '—')
    kb = [
        [InlineKeyboardButton(t(lang, btn_key), callback_data=cb)],
        [InlineKeyboardButton(t(lang, 'ai_exit'), callback_data="ai_exit")],
    ]
    await update.message.reply_text(
        text, reply_markup=InlineKeyboardMarkup(kb), parse_mode='HTML'
    )


async def _ai_send_review_reply(update, context, lang, rr):
    """AI tuzgan sharh javobini tasdiq kartasi bilan ko'rsatadi. Matn user_data'da
    saqlanadi; «e'lon qilish» tugmasi bosilganda set_review_reply orqali yoziladi."""
    rid = rr.get('review_id')
    reply = (rr.get('reply') or '').strip()
    if not rid or not reply:
        return
    store = context.user_data.setdefault('ai_review_replies', {})
    store[str(rid)] = reply
    text = t(lang, 'ai_review_reply_card',
             product=html.escape(rr.get('product') or '—'),
             comment=html.escape(rr.get('comment') or ''),
             reply=html.escape(reply))
    kb = [
        [InlineKeyboardButton(t(lang, 'ai_review_publish_btn'), callback_data=f"airvpub_{rid}")],
        [InlineKeyboardButton(t(lang, 'ai_exit'), callback_data="ai_exit")],
    ]
    await update.message.reply_text(
        text, reply_markup=InlineKeyboardMarkup(kb), parse_mode='HTML'
    )


async def ai_review_reply_publish(update, context):
    """AI tuzgan sharh javobini e'lon qiladi (callback: airvpub_{review_id})."""
    query = update.callback_query
    lang = get_lang(update, context)
    try:
        review_id = int(query.data.split("_")[1])
    except (ValueError, IndexError):
        return
    store = context.user_data.get('ai_review_replies') or {}
    reply = store.get(str(review_id))
    user = db.get_user_by_telegram_id(update.effective_user.id)
    if not reply or not user:
        try:
            await query.edit_message_text(t(lang, 'ai_review_reply_expired'))
        except Exception:
            pass
        return
    # set_review_reply egalikni SQL'da tekshiradi (seller_id mos kelmasa yozmaydi)
    ok = db.set_review_reply(review_id, user['id'], reply)
    store.pop(str(review_id), None)
    if ok:
        await _notify_buyer_of_reply(context, review_id)
    msg = t(lang, 'review_reply_saved') if ok else t(lang, 'review_reply_not_yours')
    try:
        await query.edit_message_text(msg)
    except Exception:
        await context.bot.send_message(update.effective_chat.id, msg)


async def ai_review_reply_generate(update, context):
    """Sharhga AI javobini tuzadi va e'lon / boshqa variant tugmalari bilan ko'rsatadi
    (callback: airvgen_{review_id}). Qo'lda javob oynasidagi «🤖 AI yozib bersin» tugmasi."""
    query = update.callback_query
    lang = get_lang(update, context)
    chat_id = update.effective_chat.id
    try:
        review_id = int(query.data.split("_")[1])
    except (ValueError, IndexError):
        return

    # XAVFSIZLIK: faqat shu sharhning sotuvchisi
    user = db.get_user_by_telegram_id(update.effective_user.id)
    review = db.get_review_by_id(review_id)
    if not review or not user or review.get('seller_id') != user['id']:
        await context.bot.send_message(chat_id, t(lang, 'review_reply_not_yours'))
        return

    try:
        await context.bot.send_chat_action(chat_id=chat_id, action=ChatAction.TYPING)
    except Exception:
        pass

    reply = await ai_assistant.generate_review_reply(
        product=review.get('product_name') or '',
        comment=review.get('comment') or '',
        shop_rating=review.get('rating'),
        product_rating=review.get('product_rating'),
        buyer=review.get('buyer_name') or '',
        lang=lang,
    )
    if not reply:
        await context.bot.send_message(chat_id, t(lang, 'ai_review_gen_failed'))
        return

    # Matnni saqlaymiz — «e'lon qilish» tugmasi (airvpub_) shu yerdan oladi
    store = context.user_data.setdefault('ai_review_replies', {})
    store[str(review_id)] = reply
    text = t(lang, 'ai_review_reply_card',
             product=html.escape(review.get('product_name') or '—'),
             comment=html.escape(review.get('comment') or ''),
             reply=html.escape(reply))
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton(t(lang, 'ai_review_publish_btn'), callback_data=f"airvpub_{review_id}")],
        [InlineKeyboardButton(t(lang, 'ai_review_regen_btn'), callback_data=f"airvgen_{review_id}")],
    ])
    await context.bot.send_message(chat_id, text, reply_markup=kb, parse_mode='HTML')


async def _ai_send_products(update, context, lang, products):
    """AI topgan real mahsulotlarni bosiladigan tugmalar sifatida yuboradi.
    Har bir tugma mavjud `prod_{id}` oqimini ochadi (ko'rish + buyurtma)."""
    rows = []
    for p in products:
        if not p.get('id'):
            continue
        label = p.get('name') or '—'
        if p.get('price'):
            label = f"{label} — {fmt_price(p['price'])}"
        if len(label) > 60:
            label = label[:57] + '…'
        rows.append([InlineKeyboardButton(label, callback_data=f"prod_{p['id']}")])
    if not rows:
        return
    rows.append([InlineKeyboardButton(t(lang, 'ai_exit'), callback_data="ai_exit")])
    await update.message.reply_text(
        t(lang, 'ai_found_products'),
        reply_markup=InlineKeyboardMarkup(rows)
    )


async def _ai_send_draft(update, context, lang, draft):
    """AI tuzgan e'lon qoralamasini ko'rsatadi va joylash tugmasini beradi."""
    context.user_data['ai_draft'] = draft
    name = draft.get('name') or '—'
    desc = draft.get('description') or ''
    cat = draft.get('category_name') or '—'
    price = draft.get('price')
    price_s = (fmt_price(price) if price else t(lang, 'ai_price_missing'))

    text = t(lang, 'ai_draft_card', name=name, price=price_s, category=cat, desc=desc)

    if price:
        kb = [
            [InlineKeyboardButton(t(lang, 'ai_add_photo_publish'), callback_data="ai_addphoto")],
            [InlineKeyboardButton(t(lang, 'ai_publish_nophoto'), callback_data="ai_publish")],
            [InlineKeyboardButton(t(lang, 'ai_discard'), callback_data="ai_exit")],
        ]
    else:
        # Narx yo'q — joylab bo'lmaydi; foydalanuvchi narx yozsin
        kb = [[InlineKeyboardButton(t(lang, 'ai_discard'), callback_data="ai_exit")]]

    await update.message.reply_text(
        text, reply_markup=InlineKeyboardMarkup(kb), parse_mode='HTML'
    )


def _ai_photos_kb(lang):
    """AI rasm yig'ish bosqichidagi inline tugmalar."""
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(t(lang, 'ai_photo_done_btn'), callback_data="ai_photos_done")],
        [InlineKeyboardButton(t(lang, 'ai_discard'), callback_data="ai_exit")],
    ])


async def _ai_publish_from_draft(update, context, draft, user, photos=None):
    """AI qoralamasini bazaga yozadi va rasmlarni saqlaydi (kanalga JOYLAMAYDI —
    bu reklama preview tasdiqidan keyin bo'ladi). Yaratilgan product_id ni qaytaradi."""
    photos = [p for p in (photos or []) if p][:db.MAX_PRODUCT_IMAGES]
    product_id = db.create_product(
        seller_id=user['id'],
        name=draft['name'],
        price=draft['price'],
        category_id=draft.get('category_id'),
        description=draft.get('description'),
        image_url=(photos[0] if photos else None),
    )
    if product_id and photos:
        db.set_product_images(product_id, photos)
    # AI rasm/qoralama holatlarini tozalaymiz
    for k in ('ai_draft', 'ai_awaiting_photos', 'ai_product_photos'):
        context.user_data.pop(k, None)
    return product_id


async def ai_publish_draft(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """AI tuzgan e'lonni RASMSIZ bazaga saqlaydi va reklama ko'rinishini ko'rsatadi."""
    query = update.callback_query
    lang = get_lang(update, context)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    draft = context.user_data.get('ai_draft')

    if not user or not draft or not draft.get('price') or not draft.get('name'):
        await query.answer(t(lang, 'ai_draft_expired'), show_alert=True)
        return
    await query.answer()

    product_id = await _ai_publish_from_draft(update, context, draft, user, photos=None)
    try:
        await query.edit_message_text(t(lang, 'ai_saved_preview', id=product_id))
    except Exception:
        pass
    if product_id:
        await show_ad_preview(update, context, product_id)


async def ai_addphoto_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """«📷 Rasm qo'shib joylash» — rasm yig'ish bosqichini boshlaydi."""
    query = update.callback_query
    lang = get_lang(update, context)
    draft = context.user_data.get('ai_draft')

    if not draft or not draft.get('price') or not draft.get('name'):
        await query.answer(t(lang, 'ai_draft_expired'), show_alert=True)
        return
    await query.answer()

    context.user_data['ai_awaiting_photos'] = True
    context.user_data['ai_product_photos'] = []
    try:
        await query.edit_message_text(t(lang, 'ai_send_photos'), reply_markup=_ai_photos_kb(lang))
    except Exception:
        await query.message.reply_text(t(lang, 'ai_send_photos'), reply_markup=_ai_photos_kb(lang))


async def ai_photo_collect(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """AI rejimida mahsulot rasmlarini yig'adi (global PHOTO/Document handler).
    Faqat 'ai_awaiting_photos' yoqilgan bo'lsa ishlaydi; aks holda e'tiborsiz."""
    if not context.user_data.get('ai_awaiting_photos'):
        # AI rasm bosqichida emasmiz. APP-ONLY: xaridor/sotuvchi botda ish qilmaydi —
        # rasm yuborsa launcher ko'rsatamiz (admin'gacha Faza 3 tegmaymiz).
        u = db.get_user_by_telegram_id(update.effective_user.id)
        if not u or u.get('role') != 'admin':
            await _go_to_app(update, context)
        return

    lang = get_lang(update, context)
    file_id = None
    if update.message.photo:
        photo = update.message.photo[-1]
        if photo.file_size and photo.file_size > 5 * 1024 * 1024:
            await update.message.reply_text(t(lang, 'photo_too_big'))
            return
        file_id = photo.file_id
    elif update.message.document:
        doc = update.message.document
        if not (doc.mime_type or '').startswith('image/'):
            await update.message.reply_text(t(lang, 'only_images'))
            return
        if doc.file_size and doc.file_size > 5 * 1024 * 1024:
            await update.message.reply_text(t(lang, 'photo_too_big_doc'))
            return
        file_id = doc.file_id
    else:
        return

    photos = context.user_data.setdefault('ai_product_photos', [])
    photos.append(file_id)
    n = len(photos)
    if n >= db.MAX_PRODUCT_IMAGES:
        await update.message.reply_text(
            t(lang, 'ai_photos_max_reached', max=db.MAX_PRODUCT_IMAGES),
            reply_markup=_ai_photos_kb(lang)
        )
    else:
        await update.message.reply_text(
            t(lang, 'ai_photos_added', n=n, max=db.MAX_PRODUCT_IMAGES),
            reply_markup=_ai_photos_kb(lang)
        )


async def ai_photos_done(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """«✅ Joylash» — yig'ilgan rasmlar bilan AI e'lonini joylaydi."""
    query = update.callback_query
    lang = get_lang(update, context)
    user = db.get_user_by_telegram_id(update.effective_user.id)
    draft = context.user_data.get('ai_draft')

    if not user or not draft or not draft.get('price') or not draft.get('name'):
        await query.answer(t(lang, 'ai_draft_expired'), show_alert=True)
        return

    photos = [p for p in context.user_data.get('ai_product_photos', []) if p]
    if not photos:
        # Hali rasm yo'q — yuborishni so'raymiz
        await query.answer(t(lang, 'ai_send_photo_hint'), show_alert=True)
        return
    await query.answer()

    product_id = await _ai_publish_from_draft(update, context, draft, user, photos=photos)

    msg = t(lang, 'ai_saved_preview', id=product_id)
    if photos:
        msg += t(lang, 'frag_photos_saved', n=len(photos))
    try:
        await query.edit_message_text(msg)
    except Exception:
        pass
    if product_id:
        await show_ad_preview(update, context, product_id)


async def ai_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await ai_assistant_start(update, context)


# ============================================================
# GENERIC BUTTON HANDLER
# ============================================================

async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query

    # Rate limiting: bir foydalanuvchi 0.5s dan tez callback yuborsa — e'tiborsiz
    uid = update.effective_user.id if update.effective_user else None
    if uid:
        now = _time.monotonic()
        last = context.user_data.get('_rl_btn', 0)
        if now - last < 0.5:
            try:
                await query.answer("⏳")
            except Exception:
                pass
            return
        context.user_data['_rl_btn'] = now

    data = query.data

    # === XAVFSIZLIK: admin amallari faqat admin uchun ===
    # button_handler barcha callback'larni yo'naltirgani uchun, admin amallarini
    # SHU YERDA tekshiramiz. Bu qo'shimcha himoya — oddiy foydalanuvchilar uchun
    # hech narsani o'zgartirmaydi, faqat admin bo'lmaganlarni to'xtatadi.
    ADMIN_CB_PREFIXES = (
        "admin_", "approve_seller_", "reject_seller_", "seller_req_", "excel_",
        "adisp_",
    )
    if data and data.startswith(ADMIN_CB_PREFIXES):
        _admin_user = db.get_user_by_telegram_id(uid) if uid else None
        _is_admin = bool(_admin_user) and (
            _admin_user.get("role") == "admin" or uid == ADMIN_ID
        )
        if not _is_admin:
            try:
                await query.answer(T(update, context, 'admin_only_action'), show_alert=True)
            except Exception:
                pass
            return

    await query.answer()

    # === APP-ONLY (Faza 1 — Xaridor): xaridor ishi botdan olib tashlangan ===
    # Eski xabarlardagi xaridor tugmalari bosilsa — "amal endi ilovada" ekrani.
    # (Buyurtma/savat/baho/xabar/bekor oqimlari ConversationHandler entry orqali
    #  allaqachon launcherga yo'naltirilgan; bu yerda non-conversation tugmalar.)
    BUYER_CB_EXACT = {
        "buyer_panel", "buyer_search_menu", "buyer_categories", "buyer_search",
        "buyer_shop_search", "buyer_orders", "buyer_profile", "buyer_reviews",
        "buyer_messages", "buyer_debts", "my_referral", "ai_assistant", "ai_exit",
        "cart_view", "cart_clear", "cart_clear_yes", "contact_admin",
        "join_with_code", "skip_search_location", "reapply_seller",
    }
    BUYER_CB_PREFIXES = (
        "cart_add_", "cart_inc_", "cart_dec_", "cart_reset_add_",
        "cvinc_", "cvdec_", "cvrm_",
        "order_detail_", "buyer_cancel_", "buyer_confirm_pickup_",
        "gbuyer_cancel_", "gbuyer_pickup_", "gorder_detail_",
        "debtbuyer_", "debtpayfull_", "debtpaypart_",
        "msgs_", "msg_", "call_", "recommend_", "share_link_",
        "shop_products_", "shop_ai_", "shop_list_", "cat_",
        # mahsulot kartochkasi + qidiruv (eski xabarlardagi tugmalar):
        "prod_", "pcomm_", "pg_", "sreg_", "sdist_", "srt_",
    )
    if data and (
        data in BUYER_CB_EXACT
        or data.startswith(BUYER_CB_PREFIXES)
        # shop_<id> detali (sotuvchi shop_paymode/shop_pending bilan to'qnashmasin)
        or (data.startswith("shop_") and data[5:6].isdigit())
    ):
        await _go_to_app(update, context)
        return

    # === APP-ONLY (Faza 2 — Sotuvchi): sotuvchi ishi botdan olib tashlangan ===
    # Panel/mahsulot/xodim/kanal/reklama boshqaruvi — hammasi ilovada.
    # SAQLANADI (bu gate USHLAMAYDI): buyurtma xabarnoma-javoblari va ularning
    # davomi — confirm_order_/cancel_order_/deliver_order_, g* variantlari,
    # crfwd_/gcrfwd_, ownappr_/ownrej_, setl_*/setlamt_, cclagree_/ccldeny_,
    # seller_order_/seller_gorder_ (detal — xabarnomadan ochilishi mumkin).
    # App buyurtma oqimi shularga tayanadi.
    SELLER_CB_EXACT = {
        "seller_panel", "seller_products", "seller_orders", "seller_profile",
        "seller_stats", "seller_export_excel", "seller_messages", "seller_reviews",
        "seller_scheduled", "seller_autoreposts", "seller_debts",
        "sellergrp_products", "sellergrp_sales", "sellergrp_customers", "sellergrp_settings",
        "staff_panel", "staff_list", "staff_add", "staff_add_nodept",
        "staff_invites", "staff_stats", "shop_paymode", "shop_pending",
        "sp_search", "seller_add_product", "edit_card_info", "edit_seller_region",
        "ai_publish", "ai_addphoto", "ai_photos_done",
        "adprev_publish", "adprev_regen", "adprev_long", "adprev_short",
        "adprev_edit", "adprev_skip", "adprev_schedule", "adprev_autorep",
        "schd_abort", "arep_abort",
        "switch_to_seller", "switch_to_seller_confirm", "do_switch_seller",
        "switch_to_buyer", "switch_to_buyer_confirm", "do_switch_buyer",
    }
    SELLER_CB_PREFIXES = (
        "sp_list_", "prod_menu_", "pstatus_", "edit_start_",
        "delete_prod_", "delete_confirm_",
        "toggle_stock_", "setstock_", "set_stock_",
        "arep_start_", "arep_off_", "arep_hour_", "arep_cancel_",
        "schd_date_", "schd_hour_", "schd_min_", "schd_cancel_",
        "staff_detail_", "staff_dept_", "staff_role_", "staff_reject_",
        "inv_cancel_", "staff_toggle_", "staff_pset_", "staff_perm_", "staff_rm_",
        "rvreply_", "airvpub_", "airvgen_",
        "sregset_", "sregdist_",
    )
    if data and (data in SELLER_CB_EXACT or data.startswith(SELLER_CB_PREFIXES)):
        await _go_to_app(update, context)
        return

    handlers = {
        "buyer_panel": buyer_panel,
        "buyer_search_menu": buyer_search_menu,
        "buyer_categories": buyer_categories,
        "buyer_search": buyer_search,
        "buyer_shop_search": buyer_shop_search,
        "buyer_orders": buyer_orders,
        "buyer_profile": buyer_profile,
        "seller_panel": seller_panel,
        "seller_products": seller_products,
        "seller_orders": seller_orders,
        "seller_profile": seller_profile,
        "admin_panel": admin_panel,
        "admingrp_people": admin_group_people,
        "admingrp_catalog": admin_group_catalog,
        "admingrp_manage": admin_group_manage,
        "admin_users": admin_users,
        "admin_shops": admin_shops,
        "admin_products": admin_products,
        "admin_orders": admin_orders,
        "admin_stats": admin_stats,
        "admin_channels": admin_channels,
        "admin_revenue": admin_revenue,
        "admin_broadcast": admin_broadcast_start,
        "admin_seller_requests": admin_seller_requests,
        "admin_disputes": admin_disputes,
        "admin_deleted_products": admin_deleted_products,
        "admin_analytics": admin_analytics,
        "edit_seller_region": edit_seller_region,
        "contact_admin": contact_admin_start,
        "admin_backup": admin_backup,
        "admin_settings": admin_settings,
        "admin_clean_cancelled": admin_clean_cancelled,
        "skip_search_location": skip_search_location,
        "switch_to_buyer": switch_to_buyer,
        "switch_to_seller": switch_to_seller,
        "switch_to_buyer_confirm": switch_to_buyer_confirm,
        "switch_to_seller_confirm": switch_to_seller_confirm,
        "do_switch_buyer": switch_to_buyer,
        "do_switch_seller": switch_to_seller,
        "my_referral": my_referral,
        "seller_stats": seller_stats,
        "seller_export_excel": seller_export_excel,
        "seller_messages": seller_messages,
        "seller_reviews": seller_reviews,
        "buyer_reviews": buyer_reviews,
        "buyer_messages": buyer_messages,
        "reapply_seller": reapply_seller,
        "admin_user_search": admin_user_search_start,
        "cart_view": cart_view,
        "cart_clear": cart_clear_ask,
        "cart_clear_yes": cart_clear_yes,
        "change_lang": change_language_menu,
        "ai_assistant": ai_assistant_start,
        "ai_exit": ai_assistant_exit,
        "ai_publish": ai_publish_draft,
        "ai_addphoto": ai_addphoto_start,
        "ai_photos_done": ai_photos_done,
        "adprev_publish": ad_preview_publish,
        "adprev_regen": ad_preview_regen,
        "adprev_long": ad_preview_set_length,
        "adprev_short": ad_preview_set_length,
        "adprev_edit": ad_preview_edit,
        "adprev_skip": ad_preview_skip,
        "adprev_schedule": ad_preview_schedule_start,
        "adprev_autorep": ad_preview_autorepost_start,
        "seller_scheduled": seller_scheduled_posts,
        "seller_autoreposts": seller_auto_reposts,
        "sellergrp_products": seller_group_products,
        "sellergrp_sales": seller_group_sales,
        "sellergrp_customers": seller_group_customers,
        "sellergrp_settings": seller_group_settings,
        "schd_abort": sched_abort_flow,
        "arep_abort": autorep_abort_flow,
        # MULTI-SOTUVCHI: ega paneli (exact-match — shop_ prefiks ziddiyatidan qochish uchun)
        "staff_panel": staff_panel,
        "staff_list": staff_list,
        "staff_add": staff_add,
        "staff_add_nodept": staff_add_nodept,
        "staff_invites": staff_invites,
        "staff_stats": staff_stats,
        "shop_paymode": shop_paymode_toggle,
        "shop_pending": shop_pending_products,
        "join_with_code": join_with_code_start,
    }

    if data in handlers:
        await handlers[data](update, context)
    elif data.startswith("setlang_"):
        await set_language(update, context)
    # === SAVAT (cart) va GURUH (savat buyurtmasi) callback'lari ===
    elif data.startswith("cart_reset_add_"):
        await cart_reset_add(update, context)
    elif data.startswith("cart_add_"):
        await cart_add(update, context)
    elif data.startswith("cart_inc_"):
        await cart_inc(update, context)
    elif data.startswith("cart_dec_"):
        await cart_dec(update, context)
    elif data.startswith("cvinc_"):
        await cart_view_inc(update, context)
    elif data.startswith("cvdec_"):
        await cart_view_dec(update, context)
    elif data.startswith("cvrm_"):
        await cart_view_remove(update, context)
    elif data.startswith("gcancel_"):
        await seller_reject_prompt(update, context)   # avval bekor sababini so'raymiz
    elif data.startswith(("gconfirm_", "gdeliver_")):
        await group_status_action(update, context)
    elif data.startswith(("rjok_", "rjback_")):
        await seller_reject_do(update, context)
    elif data.startswith("gcrfwd_"):
        await seller_forward_courier_group(update, context)
    elif data.startswith("seller_gorder_"):
        await seller_group_order_detail(update, context)
    elif data.startswith("gorder_detail_"):
        await buyer_group_order_detail(update, context)
    elif data.startswith("gbuyer_cancel_"):
        await buyer_cancel_group(update, context)
    elif data.startswith("gbuyer_pickup_"):
        await buyer_pickup_group(update, context)
    elif data.startswith("shop_products_"):
        await buyer_shop_products(update, context)
    elif data.startswith("shop_ai_"):
        await buyer_shop_ai_start(update, context)
    elif data.startswith("shop_list_"):
        await buyer_shop_list(update, context)
    elif data.startswith("shop_") and not data.startswith("shop_products_") \
            and not data.startswith("shop_list_") and not data.startswith("shop_ai_"):
        await buyer_shop_detail(update, context)
    elif data.startswith("cat_"):
        # cat_ID yoki cat_ID_pg_N
        await buyer_category_products(update, context)
    elif data.startswith("edit_start_"):
        await edit_product_hub(update, context)
    elif data.startswith("prod_menu_"):
        await seller_product_menu(update, context)
    elif data.startswith("sp_list_"):
        await seller_products_list(update, context)
    elif data.startswith("pstatus_"):
        await change_product_status(update, context)
    elif data == "sp_search":
        await seller_product_search_start(update, context)
    elif data.startswith(("ownappr_", "ownrej_")):
        await owner_review_product(update, context)
    elif data.startswith("staff_detail_"):
        await staff_detail(update, context)
    elif data.startswith("staff_dept_"):
        await staff_set_dept_prompt(update, context)
    elif data.startswith("staff_role_"):
        await staff_role_toggle(update, context)
    elif data.startswith("staff_reject_"):
        await staff_reject(update, context)
    elif data.startswith("inv_cancel_"):
        await invite_cancel(update, context)
    elif data.startswith("staff_toggle_"):
        await staff_toggle(update, context)
    elif data.startswith("staff_pset_"):
        await staff_perm_menu(update, context)
    elif data.startswith("staff_perm_"):
        await staff_perm_menu(update, context)
    elif data.startswith("staff_rm_"):
        await staff_remove(update, context)
    elif data == "noop":
        pass
    elif data.startswith("prod_"):
        await buyer_product_details(update, context)
    elif data.startswith("pcomm_"):
        await product_reviews_view(update, context)
    elif data.startswith("rvreply_"):
        await review_reply_start(update, context)
    elif data.startswith("airvpub_"):
        await ai_review_reply_publish(update, context)
    elif data.startswith("airvgen_"):
        await ai_review_reply_generate(update, context)
    elif data.startswith("schd_date_"):
        await sched_pick_date(update, context)
    elif data.startswith("schd_hour_"):
        await sched_pick_hour(update, context)
    elif data.startswith("schd_min_"):
        await sched_pick_minute(update, context)
    elif data.startswith("schd_cancel_"):
        await sched_cancel_post(update, context)
    elif data.startswith("arep_hour_"):
        await autorep_pick_hour(update, context)
    elif data.startswith("arep_cancel_"):
        await autorep_cancel(update, context)
    elif data.startswith("arep_start_"):
        await product_autorepost_start(update, context)
    elif data.startswith("arep_off_"):
        await product_autorepost_stop(update, context)
    elif data.startswith("toggle_stock_"):
        await toggle_product_stock(update, context)
    elif data.startswith("setstock_"):
        await set_stock_choice(update, context)
    elif data.startswith("set_stock_"):
        await set_stock_prompt(update, context)
    elif data.startswith("msgs_"):
        await view_messages(update, context)
    elif data.startswith("call_"):
        # tel: URL Telegram'da ishlamaydi — raqamni matn sifatida ko'rsatamiz
        product_id = int(data.split("_")[1])
        product = db.get_product_basic(product_id)
        if product:
            phone = product.get('phone_number') or '—'
            await update.callback_query.answer(f"📞 {phone}", show_alert=True)
    elif data.startswith("delete_prod_") or data.startswith("delete_confirm_"):
        await delete_product(update, context)
    elif data.startswith("order_detail_"):
        await buyer_order_detail(update, context)
    elif data.startswith("buyer_cancel_"):
        await buyer_cancel_order(update, context)
    elif data.startswith("buyer_confirm_pickup_"):
        await buyer_confirm_pickup(update, context)
    elif data.startswith("seller_order_"):
        await seller_order_detail(update, context)
    elif data.startswith("cancel_order_"):
        await seller_reject_prompt(update, context)   # avval bekor sababini so'raymiz
    elif data.startswith(("confirm_order_", "deliver_order_")):
        await update_order_status(update, context)
    elif data.startswith(("setl_paid_", "setl_debt_", "setl_inst_")):
        await settle_choice(update, context)
    elif data.startswith("setlamt_"):
        await settle_amount_choice(update, context)
    elif data == "seller_debts":
        await seller_debts(update, context)
    elif data == "buyer_debts":
        await buyer_debts(update, context)
    elif data.startswith("debtbuyer_"):
        await debt_buyer_detail(update, context)
    elif data.startswith("debtpayfull_"):
        await debt_pay_full(update, context)
    elif data.startswith("debtpaypart_"):
        await debt_pay_part_start(update, context)
    elif data.startswith("crfwd_"):
        await seller_forward_courier(update, context)
    elif data.startswith(("cclagree_", "ccldeny_")):
        await cancel_respond(update, context)
    elif data.startswith("admin_dispmsgs_"):
        await admin_dispute_messages(update, context)
    elif data.startswith("admin_disp_"):
        await admin_dispute_detail(update, context)
    elif data.startswith("adisp_"):
        await admin_resolve_dispute(update, context)
    elif data.startswith("admin_audit_"):
        await admin_audit_detail(update, context)
    elif data.startswith("admin_user_"):
        await admin_user_details(update, context)
    elif data.startswith("admin_block_"):
        await admin_block_user(update, context)
    elif data.startswith("admin_verify_"):
        await admin_verify_user(update, context)
    elif data.startswith("admin_askfill_"):
        await admin_request_fill(update, context)
    elif data.startswith("admin_sendfill_"):
        await admin_send_fill(update, context)
    elif data.startswith("excel_"):
        await admin_export_excel(update, context)
    elif data == "admin_clean_cancelled":
        await admin_clean_cancelled(update, context)
    elif data.startswith("recommend_"):
        await show_recommendations(update, context)
    elif data.startswith("share_link_"):
        await share_product_link(update, context)
    elif data.startswith("sregset_"):
        await seller_region_set(update, context)
    elif data.startswith("msg_"):
        # Buyurtmasiz sotuvchiga savol — mahsulot ID bo'yicha
        await product_ask_seller(update, context)
    elif data.startswith("sregdist_"):
        await seller_district_set(update, context)
    elif data.startswith("sreg_"):
        await search_region_select(update, context)
    elif data.startswith("sdist_"):
        await search_district_select(update, context)
    elif data.startswith("srt_"):
        await search_sort_change(update, context)
    elif data.startswith("pg_"):
        await search_page_change(update, context)
    elif data.startswith("admin_users_pg_"):
        await admin_users(update, context)
    elif data.startswith("admin_channels_pg_"):
        await admin_channels(update, context)
    elif data.startswith("admin_shopmod_"):
        await admin_shop_toggle_mod(update, context)
    elif data.startswith("admin_stafftog_"):
        await admin_staff_toggle(update, context)
    elif data.startswith("admin_shop_"):
        await admin_shop_detail(update, context)
    elif data.startswith("admin_order_"):
        await admin_order_detail(update, context)
    elif data.startswith("admin_force_cancel_"):
        await admin_force_cancel_order(update, context)
    elif data.startswith("approve_seller_"):
        await approve_seller(update, context)
    elif data.startswith("reject_seller_"):
        await reject_seller(update, context)
    elif data.startswith("seller_req_"):
        await admin_seller_request_detail(update, context)
    elif data == "noop":
        # Sahifa indikatori tugmasi — hech narsa qilmaymiz, faqat answer()
        await update.callback_query.answer()


# ============================================================
# TEXT HANDLER
# ============================================================

async def text_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    user = db.get_user_by_telegram_id(update.effective_user.id)

    # Zahira sonini belgilash rejimida bo'lsa — uni avval ishlov beramiz
    if context.user_data.get('setting_stock_for'):
        handled = await set_stock_submit(update, context)
        if handled:
            return

    # Pastki panel tugmalari (uz/ru) — kanonik kalitga aylantiramiz
    action = bottom_action(text)

    if action:
        if not user:
            await update.message.reply_text(T(update, context, 'not_registered_short'))
            return

        # Pastki menyu bosilganda — eski search/broadcast/AI holatlarini tozalaymiz
        context.user_data.pop('search_state', None)
        context.user_data.pop('search_query', None)
        context.user_data.pop('broadcasting', None)
        context.user_data.pop('ai_chat', None)
        context.user_data.pop('ai_draft', None)
        context.user_data.pop('ai_awaiting_photos', None)
        context.user_data.pop('ai_product_photos', None)
        context.user_data.pop('ad_preview', None)
        context.user_data.pop('ad_editing_caption', None)
        context.user_data.pop('awaiting_review_reply', None)
        ai_assistant.reset_history(context.user_data)

        # role o'rniga active_mode — bitta foydalanuvchi ikkala rejimda ishlashi mumkin
        active_mode = get_active_mode(user, context)

        # Rol talab qiladigan harakatlar.
        # APP-ONLY (Faza 1): xaridor amallari botdan olib tashlangan → _go_to_app.
        role_actions = {
            'btn_search_menu': (_go_to_app, 'buyer'),
            'btn_search':      (_go_to_app, 'buyer'),
            'btn_categories':  (_go_to_app, 'buyer'),
            'btn_my_orders':   (_go_to_app, 'buyer'),
            'btn_add_product': (_go_to_app, 'seller'),
            'btn_my_products': (_go_to_app, 'seller'),
            'btn_orders':      (_go_to_app, 'seller'),
        }

        if action == 'btn_profile':
            # APP-ONLY (Faza 1+2): profil ilovada (xaridor ham, sotuvchi ham)
            await _go_to_app(update, context)
        elif action == 'btn_home':
            if user['role'] == 'admin':
                await admin_panel(update, context)
            else:
                await _go_to_app(update, context)
        elif action == 'btn_contact_admin':
            # APP-ONLY: admin bilan bog'lanish ilovada
            await _go_to_app(update, context)
        else:
            fn, required_role = role_actions[action]
            if active_mode == required_role:
                await fn(update, context)
        return

    # Sotuvchi sharhga javob yozmoqda — kiritilgan matnni javob sifatida qabul qilamiz
    if context.user_data.get('awaiting_review_reply'):
        await review_reply_submit(update, context)
        return

    # Reklama matnini sotuvchi tahrirlamoqda — kiritilgan matnni qabul qilamiz
    if context.user_data.get('ad_editing_caption'):
        await ad_preview_caption_input(update, context)
        return

    # To'lov holati: sotuvchi qo'lda to'langan summani yozdi (qarz/bo'lib)
    if context.user_data.get('awaiting_settle_amount'):
        await settle_custom_amount_input(update, context)
        return

    # Qarz: qisman to'lov summasi kiritildi
    if context.user_data.get('awaiting_debt_payment'):
        await debt_payment_input(update, context)
        return

    # MULTI-SOTUVCHI: taklif uchun bo'lim nomi kiritildi
    if context.user_data.get('staff_invite_dept'):
        shop_id = context.user_data.pop('staff_invite_dept')
        shop = db.get_shop_by_id(shop_id)
        if shop and user:
            dept = (text.strip()[:60] or None)
            await _send_invite_link(update, context, shop, dept, user['id'])
        return

    # MULTI-SOTUVCHI: mavjud xodimga bo'lim biriktirish
    if context.user_data.get('staff_set_dept_for'):
        staff_id = context.user_data.pop('staff_set_dept_for')
        lang = get_lang(update, context)
        db.update_staff(staff_id, department=(text.strip()[:60] or None))
        await update.message.reply_text(
            t(lang, 'staff_dept_saved'),
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton(
                t(lang, 'back'), callback_data=f"staff_detail_{staff_id}")]]))
        return

    # MULTI-SOTUVCHI: taklif kodi bilan do'konga qo'shilish
    if context.user_data.get('joining_with_code'):
        context.user_data.pop('joining_with_code', None)
        await _handle_staff_deeplink(update, context, text.strip().upper(), user)
        return

    # AI yordamchi rejimi — erkin matn DeepSeek'ga uzatiladi
    if context.user_data.get('ai_chat'):
        if not user:
            context.user_data.pop('ai_chat', None)
            await update.message.reply_text(T(update, context, 'not_registered_short'))
            return
        # AI e'loni uchun rasm kutyapmiz — matn emas, rasm kerak
        if context.user_data.get('ai_awaiting_photos'):
            await update.message.reply_text(
                T(update, context, 'ai_send_photo_hint'),
                reply_markup=_ai_photos_kb(get_lang(update, context))
            )
            return
        await ai_handle_message(update, context)
        return

    # Admin bilan bog'lanish (pastki menyu orqali)
    if context.user_data.get('admin_searching_user'):
        if user and (user['role'] == 'admin' or update.effective_user.id == ADMIN_ID):
            await admin_user_search_result(update, context)
            return
        else:
            context.user_data.pop('admin_searching_user', None)

    if context.user_data.get('admin_searching_product'):
        if user and (user['role'] == 'admin' or update.effective_user.id == ADMIN_ID):
            await admin_product_search_result(update, context)
            return
        else:
            context.user_data.pop('admin_searching_product', None)

    if context.user_data.get('contacting_admin'):
        context.user_data.pop('contacting_admin', None)
        try:
            buyer_tg = f"@{user['telegram_username']}" if user and user.get('telegram_username') else "—"
            buyer_phone = user.get('phone_number') or '—' if user else '—'
            buyer_name = html.escape(user.get('name') or 'Anonim') if user else 'Anonim'
            role = user.get('role', 'buyer') if user else 'buyer'
            role_label = {'buyer': '🛒 Xaridor', 'seller': '🏪 Sotuvchi'}.get(role, role)
            await context.bot.send_message(
                chat_id=ADMIN_ID,
                text=(
                    f"📨 <b>Foydalanuvchidan xabar</b>\n\n"
                    f"{role_label}: {buyer_name}\n"
                    f"📞 {buyer_phone}\n"
                    f"📱 {buyer_tg}\n"
                    f"🆔 <code>{update.effective_user.id}</code>\n\n"
                    f"💬 {html.escape(text)}\n\n"
                    f"<i>Javob: /reply {update.effective_user.id} [matn]</i>"
                ),
                parse_mode='HTML'
            )
            await update.message.reply_text(T(update, context, 'contact_admin_sent'))
        except Exception as e:
            logging.error(f"Admin contact error: {e}")
            await update.message.reply_text(T(update, context, 'contact_admin_failed'))
        return

    # Admin broadcast
    if context.user_data.get('broadcasting'):
        context.user_data['broadcasting'] = False
        alang = get_lang(update, context)
        users = db.get_all_users()
        sent, failed = 0, 0
        sample_error = None
        failed_users = []   # kim olmaganini ko'rsatish uchun

        for u in users:
            try:
                await context.bot.send_message(
                    chat_id=u['telegram_id'],
                    text=t(u, 'admin_msg_prefix', text=text)
                )
                sent += 1
            except Exception as e:
                failed += 1
                failed_users.append(u)
                if sample_error is None:
                    sample_error = str(e)
                logging.error(f"Broadcast failed for {u['telegram_id']}: {e}")

        response = t(alang, 'broadcast_sent', n=sent)
        if failed:
            response += t(alang, 'broadcast_failed_n', n=failed)
            # Aniq kimlar olmagani — ism, username, ID
            response += t(alang, 'broadcast_failed_list_title')
            for u in failed_users[:30]:
                uname = u.get('telegram_username')
                tg = f"@{uname}" if uname else "—"
                response += t(alang, 'broadcast_failed_item',
                              name=html.escape(u.get('name') or t(alang, 'anonymous')),
                              tg=tg, id=u.get('telegram_id'))
            if sample_error:
                response += t(alang, 'broadcast_reason', err=sample_error[:200])
            response += t(alang, 'broadcast_reasons_common')
        await update.message.reply_text(response, parse_mode='HTML')
        return

    # Do'kon qidirish — foydalanuvchi do'kon nomini kiritdi
    if context.user_data.get('shop_search_state') == 'awaiting_query':
        context.user_data.pop('shop_search_state', None)
        query_text = text.strip()

        lang = get_lang(update, context)
        shops = db.search_shops(query=query_text)

        if not shops:
            await update.message.reply_text(
                t(lang, 'shop_not_found_q', q=query_text),
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton(t(lang, 'btn_retry_search'), callback_data="buyer_shop_search")],
                    [InlineKeyboardButton(t(lang, 'btn_home'), callback_data="buyer_panel")],
                ])
            )
            return

        context.user_data['_shop_results'] = shops
        await _render_shop_list(update.message, context, shops, page=0)
        return

    # Qidiruv — 1-bosqich: mahsulot nomi kiritildi, lokatsiyani so'raymiz
    if context.user_data.get('search_state') == 'awaiting_query':
        context.user_data['search_query'] = text.strip()
        context.user_data['search_state'] = 'awaiting_location'

        lang = get_lang(update, context)
        location_kb = ReplyKeyboardMarkup(
            [[KeyboardButton(t(lang, 'btn_send_location'), request_location=True)]],
            resize_keyboard=True,
            one_time_keyboard=True,
        )
        skip_kb = InlineKeyboardMarkup([[
            InlineKeyboardButton(t(lang, 'search_skip_location'), callback_data="skip_search_location")
        ]])
        await update.message.reply_text(
            t(lang, 'search_location_full'),
            reply_markup=location_kb,
        )
        await update.message.reply_text(
            t(lang, 'or_below'),
            reply_markup=skip_kb,
        )
        return

    # Qidiruv — 2-bosqich: foydalanuvchi lokatsiya o'rniga matn yubordi (manzil yozdi)
    # Bunga ham tayyor bo'lamiz — matnni e'tiborsiz qoldirib, faqat saqlangan so'rov bo'yicha qidiramiz
    if context.user_data.get('search_state') == 'awaiting_location':
        q_text = context.user_data.pop('search_query', None)
        context.user_data.pop('search_state', None)
        if q_text:
            await update.message.reply_text(
                T(update, context, 'text_instead_of_location'),
                reply_markup=ReplyKeyboardRemove(),
            )
            await _show_search_results(update, context, q_text)
        return

    # Noma'lum xabar — APP-ONLY: xaridor/sotuvchi botda matn yozmaydi.
    # Har qanday matnga launcher (chiroyli matn + 'Ilovaga kirish' tugma) bilan javob.
    # Admin (Faza 3 gacha) odatdagi javobni oladi.
    if not user or user.get('role') != 'admin':
        await _go_to_app(update, context)
        return
    await update.message.reply_text(T(update, context, 'unknown_command'))


async def location_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Lokatsiya xabari uchun — qidiruv 2-bosqichi va deeplink masofa so'rovi."""
    loc = update.message.location

    # Deeplink (kanal orqali ochilgan) mahsulot uchun masofa so'ralgan bo'lsa
    pending_pid = context.user_data.get('_dist_pending_product')
    if pending_pid and context.user_data.get('search_state') != 'awaiting_location':
        context.user_data.pop('_dist_pending_product', None)
        lang = get_lang(update, context)
        if loc:
            remember_buyer_geo(context, loc.latitude, loc.longitude)
            product = db.get_product_by_id(pending_pid)
            if product and product.get('shop_lat') and product.get('shop_lon'):
                km = haversine_km(loc.latitude, loc.longitude,
                                  product['shop_lat'], product['shop_lon'])
                if km is not None:
                    await update.message.reply_text(
                        t(lang, 'deeplink_distance_result', km=f"{km:.1f}"),
                        reply_markup=ReplyKeyboardRemove(),
                    )
                    return
        await update.message.reply_text(
            t(lang, 'location_saved_ok'), reply_markup=ReplyKeyboardRemove()
        )
        return

    if context.user_data.get('search_state') == 'awaiting_location':
        loc = update.message.location
        q_text = context.user_data.pop('search_query', None)
        context.user_data.pop('search_state', None)
        remember_buyer_geo(context, loc.latitude, loc.longitude)

        if not q_text:
            await update.message.reply_text(T(update, context, 'search_cancelled'), reply_markup=ReplyKeyboardRemove())
            return

        await update.message.reply_text(
            T(update, context, 'location_received_searching', q=q_text),
            reply_markup=ReplyKeyboardRemove()
        )
        await _show_search_results(update, context, q_text,
                                   buyer_lat=loc.latitude, buyer_lon=loc.longitude)
        return

    # APP-ONLY: hech qaysi oqimga tegishli bo'lmagan lokatsiya — launcher (non-admin)
    u = db.get_user_by_telegram_id(update.effective_user.id)
    if not u or u.get('role') != 'admin':
        await _go_to_app(update, context)
        return


# Eski/yaroqsiz tugma bosilganda Telegram qaytaradigan xabarlar — bular xato emas,
# kutilgan holat (xabar o'chirilgan, eskirgan yoki sanoq o'zgarmagan).
_STALE_CALLBACK_HINTS = (
    "message is not modified",
    "message to edit not found",
    "message can't be edited",
    "query is too old",
    "message to delete not found",
    "message identifier is not specified",
)


async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE):
    """Kutilmagan istisno bo'lsa — log'ga yozadi va foydalanuvchini xabardor qiladi.
    Eski xabardagi tugma bosilishi (stale callback) esa xato emas — jim yoki yumshoq
    ogohlantirish bilan o'tkazib yuboriladi."""
    err = context.error
    err_text = str(err or "").lower()

    # 1) Eski/yaroqsiz tugma (stale callback) — foydalanuvchiga yumshoq toast
    if isinstance(err, BadRequest) and any(h in err_text for h in _STALE_CALLBACK_HINTS):
        if "not modified" not in err_text:   # 'not modified' odatiy — jim o'tkazamiz
            logging.info(f"Stale callback o'tkazib yuborildi: {err}")
        try:
            if isinstance(update, Update) and update.callback_query:
                await update.callback_query.answer(
                    T(update, context, 'callback_stale'), show_alert=True)
        except Exception:
            pass
        return

    logging.error("Exception while handling update:", exc_info=err)

    # Sentry yoqilgan bo'lsa — xatoni o'sha yerga ham yuboramiz (tez ogohlantirish uchun)
    if _SENTRY_ENABLED:
        try:
            import sentry_sdk
            sentry_sdk.capture_exception(err)
        except Exception:
            pass

    # 2) Boshqa kutilmagan xatolar — foydalanuvchini xabardor qilamiz
    try:
        if isinstance(update, Update):
            if update.callback_query:
                try:
                    await update.callback_query.answer(
                        T(update, context, 'error_unexpected'), show_alert=True)
                except Exception:
                    pass
            elif update.effective_message:
                await update.effective_message.reply_text(
                    T(update, context, 'error_unexpected'))
    except Exception:
        pass


async def reminder_order_job(context: ContextTypes.DEFAULT_TYPE):
    """5 daqiqadan keyin sotuvchiga eslatma yuboradi — buyurtma hali tasdiqlanmagan bo'lsa."""
    try:
        data = context.job.data
        order_id = data['order_id']
        seller_tg = data.get('seller_tg')

        if not seller_tg:
            return

        order = db.get_order_by_id(order_id)
        if not order or order['status'] != 'pending':
            return  # Allaqachon tasdiqlangan yoki bekor qilingan

        product_name = html.escape(data.get('product_name') or '')
        buyer_name = html.escape(data.get('buyer_name') or '')
        total = data.get('total_price', 0)
        seller = db.get_user_by_id(order['seller_id'])
        slang = get_user_lang(seller) if seller else DEFAULT_LANG

        rtext = t(slang, 'job_reminder_seller',
                  pname=product_name, buyer=buyer_name, total=fmt_price(total))
        rkb = InlineKeyboardMarkup([
            [InlineKeyboardButton(t(slang, 'btn_confirm'), callback_data=f"confirm_order_{order_id}")],
            [InlineKeyboardButton(t(slang, 'btn_reject'), callback_data=f"cancel_order_{order_id}")],
        ])
        await context.bot.send_message(chat_id=seller_tg, text=rtext, parse_mode='HTML', reply_markup=rkb)
        # MULTI-SOTUVCHI: eslatma mahsulotni joylagan xodimga ham boradi
        product = db.get_product_basic(order.get('product_id')) if order.get('product_id') else None
        if product:
            await _fanout_order_to_staff(context, product, rtext, rkb, owner_tg=seller_tg)
        logging.info(f"Reminder: buyurtma {order_id} uchun sotuvchiga eslatma yuborildi.")
    except Exception as e:
        logging.error(f"reminder_order_job xatosi: {e}")


async def auto_cancel_order_job(context: ContextTypes.DEFAULT_TYPE):
    """10 daqiqadan keyin pending buyurtmani avtomatik bekor qiladi."""
    try:
        data = context.job.data
        order_id = data['order_id']
        buyer_tg = data['buyer_tg']
        seller_tg = data.get('seller_tg')

        order = db.get_order_by_id(order_id)
        if not order or order['status'] != 'pending':
            return  # Allaqachon tasdiqlangan yoki bekor qilingan

        # Buyurtmani bekor qilamiz
        db.update_order_status(order_id, 'cancelled')
        oid = fmt_order_id(order_id)

        # Xaridorga xabar (xaridor tilida)
        try:
            buyer = db.get_user_by_id(order['buyer_id'])
            await context.bot.send_message(
                chat_id=buyer_tg,
                text=t(buyer or 'uz', 'job_autocancel_buyer', oid=oid),
                parse_mode='HTML'
            )
        except Exception:
            pass

        # Sotuvchiga xabar (sotuvchi tilida)
        if seller_tg:
            try:
                seller = db.get_user_by_id(order['seller_id'])
                await context.bot.send_message(
                    chat_id=seller_tg,
                    text=t(seller or 'uz', 'job_autocancel_seller', oid=oid),
                    parse_mode='HTML'
                )
            except Exception:
                pass

        logging.info(f"Auto-cancel: buyurtma {order_id} 10 daqiqadan keyin bekor qilindi.")
    except Exception as e:
        logging.error(f"auto_cancel_order_job xatosi: {e}")


async def auto_backup_job(context: ContextTypes.DEFAULT_TYPE):
    """Har kuni avtomatik DB backup — adminga fayl sifatida yuboradi."""
    try:
        import datetime as dt

        # Kuniga FAQAT BIR MARTA — job ikki marta ishga tushsa yoki bot ikki instansda
        # ishlayotgan bo'lsa ham, backup adminga ikki marta ketmasligi uchun atomik qulf.
        today = dt.datetime.now(TZ_TASHKENT).strftime("%Y-%m-%d")
        if not db.claim_daily_once('last_auto_backup', today):
            logging.info("Avtomatik backup bugun allaqachon yuborilgan — takror o'tkazib yuborildi.")
            return

        ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = f"marketplace_backup_{ts}.db"

        ok = db.backup(backup_path)
        if not ok:
            logging.error("Avtomatik backup muvaffaqiyatsiz.")
            return

        # Fayl hajmini tekshiramiz
        file_size = os.path.getsize(backup_path)
        size_mb = file_size / (1024 * 1024)

        try:
            await context.bot.send_document(
                chat_id=ADMIN_ID,
                document=open(backup_path, 'rb'),
                filename=backup_path,
                caption=(
                    f"💾 <b>Avtomatik kunlik backup</b>\n\n"
                    f"📅 Sana: {dt.datetime.now().strftime('%d.%m.%Y %H:%M')}\n"
                    f"📦 Hajmi: {size_mb:.2f} MB"
                ),
                parse_mode='HTML'
            )
            logging.info(f"Avtomatik backup muvaffaqiyatli yuborildi: {backup_path} ({size_mb:.2f} MB)")
        except Exception as e:
            logging.error(f"Backup faylini adminga yuborib bo'lmadi: {e}")
        finally:
            try:
                os.remove(backup_path)
            except Exception:
                pass

    except Exception as e:
        logging.error(f"auto_backup_job xatosi: {e}")


async def cleanup_stale_orders_job(context: ContextTypes.DEFAULT_TYPE):
    """Kuniga bir marta — 3 kun ichida tasdiqlanmagan buyurtmalarni bekor qiladi
    va ikkala tomonga xabar yuboradi."""
    try:
        stale = db.auto_cancel_stale_orders(days=3)
        logging.info(f"Auto-cancel: {len(stale)} ta eski pending buyurtma bekor qilindi.")

        for s in stale:
            try:
                order = db.get_order_by_id(s['id'])
                if not order:
                    continue
                oid = fmt_order_id(s['id'])
                if order.get('buyer_tg'):
                    try:
                        buyer = db.get_user_by_id(order['buyer_id'])
                        await context.bot.send_message(
                            chat_id=order['buyer_tg'],
                            text=t(buyer or 'uz', 'stale_cancel_notify', oid=oid))
                    except Exception:
                        pass
                if order.get('seller_tg'):
                    try:
                        seller = db.get_user_by_id(order['seller_id'])
                        await context.bot.send_message(
                            chat_id=order['seller_tg'],
                            text=t(seller or 'uz', 'stale_cancel_notify', oid=oid))
                    except Exception:
                        pass
            except Exception as e:
                logging.error(f"Stale order notify failed: {e}")
    except Exception as e:
        logging.error(f"cleanup_stale_orders_job failed: {e}")


async def testpost(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Kanal ulanishini tekshirish (faqat admin).  /testpost"""
    if update.effective_user.id != ADMIN_ID:
        return
    if not CHANNEL_ID:
        await update.message.reply_text(
            "❌ CHANNEL_ID belgilanmagan.\n.env fayliga qo'shing:  CHANNEL_ID=@TezBozorUz24"
        )
        return
    try:
        await context.bot.send_message(
            chat_id=CHANNEL_ID,
            text="✅ TezBozor kanal ulanishi ishlayapti!"
        )
        await update.message.reply_text(f"✅ Yuborildi → {CHANNEL_ID}")
    except Exception as e:
        logging.error(f"testpost failed: {e}")
        await update.message.reply_text(
            f"❌ Xato: {e}\n\n"
            "Tekshiring: bot kanalda admin va 'Post Messages' ruxsati yoqilganmi?"
        )


# ============================================================
# RESTART: ochiq buyurtma taymerlarini qayta rejalashtirish
# ============================================================

def _reschedule_pending_order_timers(job_queue):
    """Bot restart/deploy'idan keyin 'pending' buyurtmalar uchun jonli teskari sanoq +
    avto-bekor jobini qayta tiklaydi. Joblar xotirada saqlangani uchun restartda yo'qoladi —
    bu ularni bazadan tiklaydi. Avto-bekor muddati (auto_cancel_at) DB'da saqlangani uchun
    teskari sanoq real (o'zgarmas) qoladi; muddati o'tib ketganlar darrov bekor qilinadi."""
    try:
        pend = db.get_pending_orders_for_reschedule()
        groups_done = set()
        resched = 0
        for o in pend:
            gid = o.get('order_group_id')
            if gid:
                if gid in groups_done:
                    continue
                groups_done.add(gid)
                _schedule_order_countdown(job_queue, group_id=gid, first=5)
            else:
                _schedule_order_countdown(job_queue, order_id=o['id'], first=5)
            resched += 1
        logging.info(f"Restart: {resched} ta ochiq buyurtma sanoq/taymeri qayta tiklandi.")
    except Exception as e:
        logging.error(f"Pending buyurtma taymerlarini qayta rejalashtirish xatosi: {e}")


def _reschedule_scheduled_posts(job_queue):
    """Bot restart/deploy'idan keyin kutilayotgan ('pending') rejalashtirilgan postlar
    joblarini qayta tiklaydi. Joblar xotirada saqlangani uchun restartda yo'qoladi —
    bu ularni bazadan tiklaydi. Vaqti o'tib ketganlar (downtime davrida) tez orada
    (5 soniya) joylanadi, shunda hech bir reja yo'qolmaydi."""
    try:
        from datetime import datetime, timezone
        pend = db.get_pending_scheduled_posts()
        now = datetime.now(timezone.utc)
        resched = 0
        for sp in pend:
            sa = sp.get('scheduled_at')
            if not sa:
                continue
            try:
                target = datetime.strptime(str(sa)[:19], "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
            except Exception:
                continue
            when = (target - now).total_seconds()
            if when < 0:
                when = 5  # downtime davrida o'tib ketgan — tez orada joylanadi
            job_queue.run_once(
                scheduled_post_job, when=when,
                data={'sched_id': sp['id']}, name=f"sched_post_{sp['id']}")
            resched += 1
        logging.info(f"Restart: {resched} ta rejalashtirilgan post qayta tiklandi.")
    except Exception as e:
        logging.error(f"Rejalashtirilgan postlarni qayta tiklash xatosi: {e}")


# ============================================================
# MAIN — HANDLER REGISTRATION (BUG FIX ASOSIY QISMI)
# ============================================================

def _validate_env():
    """Ishga tushishda muhit o'zgaruvchilarining holatini qisqacha log qiladi.
    Majburiylar yetishmasa ogohlantiradi (BOT_TOKEN allaqachon yuqorida tekshirilgan)."""
    logging.info("🔧 Muhit sozlamalari:")
    logging.info("   BOT_TOKEN: %s", "✅ bor" if TOKEN else "❌ yo'q")
    logging.info("   ADMIN_ID:  %s", f"✅ {ADMIN_ID}" if ADMIN_ID else "⚠️ o'rnatilmagan")
    logging.info("   CHANNEL_ID: %s", f"✅ {CHANNEL_ID}" if CHANNEL_ID else "⚠️ o'rnatilmagan")
    logging.info("   SENTRY_DSN: %s", "✅ yoqilgan" if os.getenv("SENTRY_DSN") else "— o'chiq")
    if not ADMIN_ID:
        logging.warning("⚠️ ADMIN_ID o'rnatilmagan — admin paneli hech kimga ochilmaydi. "
                        ".env faylida ADMIN_ID=... ni belgilang (.env.example'ga qarang).")


async def webapp_order_dispatch_job(context: ContextTypes.DEFAULT_TYPE):
    """Mini App (webapp) yaratgan buyurtmalarni topib, sotuvchiga bildirishnoma +
    jonli taymerni ishga tushiradi. Webapp alohida jarayon — uning yaratgan
    buyurtmasiga bot job-queue'si avtomatik ulanmaydi; shu job har ~12s skanlaydi.
    Bildirishnoma muvaffaqiyatli ketsa ham, ketmasa ham belgi tozalanadi (qayta
    spam qilmaslik uchun — ketмаса log'da ko'rinadi)."""
    try:
        ids = db.get_orders_awaiting_notify()
    except Exception as e:
        logging.error(f"webapp_order_dispatch_job: ro'yxat olinmadi: {e}")
        return
    handled_groups = set()
    for oid in ids:
        try:
            order = db.get_order_by_id(oid)
            gid = order.get('order_group_id') if order else None
            if gid:
                # Savat (guruh) buyurtmasi — butun guruhga BITTA xabar (bir marta)
                if gid not in handled_groups:
                    handled_groups.add(gid)
                    await _dispatch_group_notification(context, gid)
            else:
                await _dispatch_order_notification(context, oid)
        except Exception as e:
            logging.error(f"webapp_order_dispatch_job: order {oid} xabar xato: {e}")
        finally:
            try:
                db.clear_order_notify_pending(oid)
            except Exception:
                pass


async def _notify_courier_assigned(context, order_id):
    """#3 — buyurtmaga biriktirilgan KURYERga Telegram PUSH: tafsilotlar (mahsulot,
    xaridor + telefon, manzil, summa). Kuryer Mini App kuryer panelidan yetkazadi."""
    order = db.get_order_by_id(order_id)
    if not order:
        return
    cid = order.get('courier_id')
    if not cid:
        return
    courier = db.get_user_by_id(cid)
    if not courier or not courier.get('telegram_id'):
        return
    prod = db.get_product_by_id(order.get('product_id')) or {}
    buyer = db.get_user_by_id(order.get('buyer_id')) or {}
    lang = courier.get('language') or 'uz'
    name = html.escape(prod.get('name') or '—')
    bname = html.escape(buyer.get('name') or '—')
    bphone = html.escape(buyer.get('phone_number') or '—')
    addr = html.escape(order.get('delivery_address') or '—')
    if lang == 'ru':
        text = (f"🚴 <b>Вам назначен новый заказ!</b>\n\n"
                f"📦 {name} × {order.get('quantity')}\n"
                f"👤 {bname} · 📞 {bphone}\n"
                f"📍 {addr}\n"
                f"💵 {order.get('total_price')}\n\n"
                f"Откройте Mini App → 🚴 Панель курьера, чтобы доставить.")
    else:
        text = (f"🚴 <b>Sizga yangi buyurtma biriktirildi!</b>\n\n"
                f"📦 {name} × {order.get('quantity')}\n"
                f"👤 {bname} · 📞 {bphone}\n"
                f"📍 {addr}\n"
                f"💵 {order.get('total_price')}\n\n"
                f"Yetkazib berish uchun Mini App → 🚴 Kuryer panelini oching.")
    await context.bot.send_message(chat_id=courier['telegram_id'], text=text, parse_mode='HTML')


async def webapp_courier_notify_job(context: ContextTypes.DEFAULT_TYPE):
    """#3 — App'da buyurtmaga biriktirilgan kuryerlarga PUSH yuboradi (har ~12s).
    Webapp alohida jarayon — biriktirish app'da bo'ladi, bot shu job orqali xabar beradi.
    Yuborilsa ham, ketmasa ham belgi tozalanadi (qayta spam qilmaslik uchun)."""
    try:
        ids = db.get_orders_awaiting_courier_notify()
    except Exception as e:
        logging.error(f"webapp_courier_notify_job: ro'yxat olinmadi: {e}")
        return
    for oid in ids:
        try:
            await _notify_courier_assigned(context, oid)
        except Exception as e:
            logging.error(f"webapp_courier_notify_job: order {oid} kuryer PUSH xato: {e}")
        finally:
            try:
                db.clear_courier_notify(oid)
            except Exception:
                pass


async def webapp_scheduled_scan_job(context: ContextTypes.DEFAULT_TYPE):
    """Mini App yaratgan rejalashtirilgan postlarga publish jobini ulaydi (idempotent).
    Webapp alohida jarayon — uning yaratgan rejasiga bot job-queue'si avtomatik ulanmaydi;
    shu job har ~30s skanlaydi va jobi yo'qlariga run_once qo'yadi."""
    from datetime import datetime, timezone
    try:
        pend = db.get_pending_scheduled_posts()
    except Exception as e:
        logging.error(f"webapp_scheduled_scan_job: ro'yxat olinmadi: {e}")
        return
    jq = context.application.job_queue
    if not jq:
        return
    now = datetime.now(timezone.utc)
    for sp in pend:
        name = f"sched_post_{sp['id']}"
        if jq.get_jobs_by_name(name):
            continue  # allaqachon rejalashtirilgan (bot yoki avvalgi skan)
        try:
            target = datetime.strptime(str(sp.get('scheduled_at'))[:19],
                                       "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
        except Exception:
            continue
        when = max(1, (target - now).total_seconds())
        jq.run_once(scheduled_post_job, when=when, data={'sched_id': sp['id']}, name=name)
        logging.info(f"Mini App rejalashtirilgan post ulandi: sched {sp['id']} ({when:.0f}s)")


async def webapp_autorepost_scan_job(context: ContextTypes.DEFAULT_TYPE):
    """Mini App yaratgan avto qayta-reklamalarga kunlik jobni ulaydi (idempotent).
    Webapp alohida jarayon — bot job-queue'siga avtomatik ulanmaydi; har ~60s skanlaydi."""
    try:
        reposts = db.get_active_auto_reposts()
    except Exception as e:
        logging.error(f"webapp_autorepost_scan_job: {e}")
        return
    jq = context.application.job_queue
    if not jq:
        return
    for rp in reposts:
        name = f"autorep_{rp['id']}"
        if jq.get_jobs_by_name(name):
            continue
        try:
            hour = int(rp.get('hour') or 0)
        except Exception:
            continue
        _schedule_autorepost_job(jq, rp['id'], hour)
        logging.info(f"Mini App auto-repost ulandi: {rp['id']} (soat {hour})")


async def _post_init(application):
    """Bot ishga tushganda — chat Menu tugmasini to'g'ridan-to'g'ri Mini App'ga ulaymiz.
    Shunda foydalanuvchi buyer panelga kirmasdan, matn maydoni yonidagi doimiy tugma
    bilan bitta tegishda ilovani ochadi (app-first kirish)."""
    global BOT_USERNAME
    try:
        me = await application.bot.get_me()
        BOT_USERNAME = me.username
        logging.info(f"✅ BOT_USERNAME = @{BOT_USERNAME} (startapp deep-link'lar uchun).")
    except Exception as e:
        logging.warning(f"BOT_USERNAME olinmadi: {e}")
    try:
        if MINIAPP_URL:
            await application.bot.set_chat_menu_button(
                menu_button=MenuButtonWebApp(
                    text="🛍 Ilovani ochish",
                    web_app=WebAppInfo(url=MINIAPP_URL)))
            logging.info("✅ Menu tugmasi Mini App'ga ulandi (app-first kirish).")
        else:
            await application.bot.set_chat_menu_button(menu_button=MenuButtonCommands())
    except Exception as e:
        logging.warning(f"Menu tugmasi o'rnatilmadi: {e}")


def main():
    _validate_env()
    # Persistence — bot qayta ishga tushganda foydalanuvchi sessiyalari saqlanadi
    # (yarim qolgan ro'yxatdan o'tish, qidiruv state, va h.k.)
    persistence = PicklePersistence(filepath="tezbozor_state.pickle")
    app = (Application.builder().token(TOKEN).persistence(persistence)
           .post_init(_post_init).build())

    # ============================================================
    # GURUH/KANAL HIMOYASI — bot FAQAT shaxsiy (private) chatda javob beradi
    # ------------------------------------------------------------
    # Muammo: bot sotuvchining kanali/guruhiga ADMIN sifatida qo'shilgach, u
    # yerdagi HAR QANDAY xabarni qabul qiladi (admin bo'lgani uchun Telegram
    # privacy rejimi ishlamaydi). Natijada eng oxirdagi global text_handler har
    # bir guruh xabariga "Buyruqni tanlang" (unknown_command) menyusi bilan javob
    # berib, guruhni spam qilardi — sotuvchilar shundan norozi edi.
    #
    # Yechim: guruh / superguruh / kanaldan kelgan xabar, buyruq va inline tugma
    # bosishlarini ENG BOSHIDA (eng past guruh = eng yuqori ustuvorlik) to'xtatamiz.
    # Bot bu chatlarga UMUMAN javob bermaydi. Bot kanal/guruhda faqat o'zi mahsulot
    # POST qiladi (chiquvchi xabar — bu filtrga tushmaydi) va "🛒 Sotib olish"
    # tugmasi (deeplink URL) orqali xaridorni shaxsiy chatga olib kiradi.
    #
    # MUSTASNO: my_chat_member / chat_member (bot guruhga qo'shilishi/chiqarilishi)
    # — guruhni ulash mexanizmi shunga bog'liq, shuning uchun ularni TO'XTATMAYMIZ;
    # ular pastdagi ChatMemberHandler orqali ishlanaveradi.
    async def _guard_private_only(update: Update, context: ContextTypes.DEFAULT_TYPE):
        # A'zolik yangilanishlari — o'tkazib yuboramiz (ChatMemberHandler ushlaydi)
        if update.my_chat_member is not None or update.chat_member is not None:
            return
        chat = update.effective_chat
        # Shaxsiy chatdan boshqa har qanday joydan kelgan update'ni to'xtatamiz
        if chat is not None and chat.type != Chat.PRIVATE:
            raise ApplicationHandlerStop
        # Faollik kuzatuvi (throttled): shaxsiy chatdagi har qanday harakat (xabar/
        # buyruq/tugma) foydalanuvchini "faol" deb belgilaydi → faol vs bir martalik.
        if update.effective_user is not None:
            try:
                db.touch_user_activity(telegram_id=update.effective_user.id)
            except Exception:
                pass

    # group=-100 — qolgan barcha handler'lardan oldin ishlaydi
    app.add_handler(TypeHandler(Update, _guard_private_only), group=-100)

    # BUG FIX #5: global_fallbacks — jarayon ichida /start, /cancel yoki pastki menyu bosilsa to'xtatadi
    async def cancel_conversation(update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Har qanday conversation'ni to'xtatib, foydalanuvchini bosh sahifaga qaytaradi."""
        # Conversation state'larini tozalaymiz
        for key in ['adding_product', 'searching', 'search_state', 'search_query',
                    'broadcasting', 'setting_stock_for', 'attr_templates', 'attr_index',
                    'product_attrs', 'order_product', 'order_qty', 'shop_search_state',
                    '_shop_results']:
            context.user_data.pop(key, None)

        user = db.get_user_by_telegram_id(update.effective_user.id)
        if update.message:
            await update.message.reply_text(T(update, context, 'cancelled'))
        if user:
            if user['role'] == 'admin':
                await admin_panel(update, context)
            elif get_active_mode(user, context) == 'seller':
                await seller_panel(update, context)
            else:
                await buyer_panel(update, context)
        return ConversationHandler.END

    # Pastki menyu tugmalari (uz+ru) — conversation ichida bosilsa bekor qiladi
    BOTTOM_MENU_TEXTS = all_bottom_menu_texts()

    cancel_filter = filters.Regex(
        "^(" + "|".join(_re.escape(_btn) for _btn in BOTTOM_MENU_TEXTS) + ")$"
    )

    global_fallbacks = [
        CommandHandler("start", start),
        CommandHandler("cancel", cancel_conversation),
        CommandHandler("admin", admin_command),
        MessageHandler(cancel_filter, cancel_conversation),
        CallbackQueryHandler(button_handler),
    ]

    # --- Registration ConversationHandler ---
    registration_conv = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            SELECT_LANG:       [CallbackQueryHandler(registration_language, pattern="^reglang_")],
            PHONE:             [MessageHandler(filters.CONTACT | filters.TEXT & ~filters.COMMAND, registration_phone)],
            NAME:              [MessageHandler(filters.TEXT & ~filters.COMMAND, registration_name)],
            ROLE:              [CallbackQueryHandler(registration_role, pattern="^reg_")],
            SELLER_CATEGORY:   [CallbackQueryHandler(registration_seller_category, pattern="^regcat_")],
            SHOP_NAME:         [MessageHandler(filters.TEXT & ~filters.COMMAND, registration_shop_name)],
            SHOP_LANDMARK:     [MessageHandler(filters.TEXT & ~filters.COMMAND, registration_shop_landmark)],
            SHOP_ADDRESS:      [MessageHandler(filters.LOCATION | filters.TEXT & ~filters.COMMAND, registration_shop_address)],
            SHOP_ADDRESS_TEXT: [MessageHandler(filters.TEXT & ~filters.COMMAND, registration_shop_address_text)],
            WORKING_DAYS:      [MessageHandler(filters.TEXT & ~filters.COMMAND, registration_working_days)],
            WORKING_HOURS:     [MessageHandler(filters.TEXT & ~filters.COMMAND, registration_working_hours)],
            TELEGRAM_USERNAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, registration_telegram_username)],
        },
        fallbacks=global_fallbacks,
        allow_reentry=True,
        conversation_timeout=300,  # 5 daqiqadan keyin conversation avtomatik tugaydi
    )

    # --- Product add ConversationHandler ---
    _add_product_btn_re = "^(" + "|".join(_re.escape(s) for s in _lang_labels('btn_add_product')) + ")$"
    product_conv = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(_go_to_app, pattern="^seller_add_product$"),
            MessageHandler(filters.Regex(_add_product_btn_re), _go_to_app),
        ],
        states={
            PRODUCT_MODE:     [
                CallbackQueryHandler(seller_add_product_mode, pattern="^pmode_(classic|ai_guided|ai_smart)$"),
            ],
            PRODUCT_NAME:     [
                CallbackQueryHandler(add_product_nav, pattern="^addnav_"),
                MessageHandler(filters.TEXT & ~filters.COMMAND & ~cancel_filter, seller_add_product_name),
            ],
            PRODUCT_PRICE:    [
                CallbackQueryHandler(add_product_nav, pattern="^addnav_"),
                MessageHandler(filters.TEXT & ~filters.COMMAND & ~cancel_filter, seller_add_product_price),
            ],
            PRODUCT_STOCK:    [
                CallbackQueryHandler(seller_add_product_stock_choice, pattern="^apstock_(unlim|num)$"),
                CallbackQueryHandler(add_product_nav, pattern="^addnav_"),
                MessageHandler(filters.TEXT & ~filters.COMMAND & ~cancel_filter, seller_add_product_stock_text),
            ],
            PRODUCT_CATEGORY: [
                CallbackQueryHandler(seller_add_product_category, pattern="^prodcat_"),
                CallbackQueryHandler(add_product_nav, pattern="^addnav_"),
            ],
            PRODUCT_DESC:     [
                CallbackQueryHandler(add_product_nav, pattern="^addnav_"),
                MessageHandler(filters.TEXT & ~filters.COMMAND & ~cancel_filter, seller_add_product_desc),
            ],
            PRODUCT_PHOTO:    [
                MessageHandler(filters.PHOTO | filters.Document.IMAGE | filters.Sticker.ALL | (filters.TEXT & ~filters.COMMAND), seller_add_product_photo),
                CallbackQueryHandler(add_photo_more, pattern="^addphoto_more$"),
                CallbackQueryHandler(add_photo_done, pattern="^addphoto_done$"),
                CallbackQueryHandler(add_product_nav, pattern="^addnav_"),
            ],
            PRODUCT_ATTRS:    [
                CallbackQueryHandler(seller_add_product_attr_nav, pattern="^attrnav_(back|skip)$"),
                MessageHandler(filters.TEXT & ~filters.COMMAND & ~cancel_filter, seller_add_product_attr_text),
                CallbackQueryHandler(seller_add_product_attr_callback, pattern="^attr_"),
            ],
        },
        fallbacks=global_fallbacks,
    )

    # --- Product edit ConversationHandler (yangi: maydon tanlash usuli) ---
    # 'edit_start_' (tahrir oynasini ochish) button_handler orqali ishlaydi.
    # Bu conversation faqat aniq bir maydon tanlangach boshlanadi.
    edit_product_conv = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(_go_to_app, pattern=r"^ef_name_\d+$"),
            CallbackQueryHandler(_go_to_app, pattern=r"^ef_price_\d+$"),
            CallbackQueryHandler(_go_to_app, pattern=r"^ef_cat_\d+$"),
            CallbackQueryHandler(_go_to_app, pattern=r"^ef_desc_\d+$"),
            CallbackQueryHandler(_go_to_app, pattern=r"^ef_photos_\d+$"),
            CallbackQueryHandler(_go_to_app, pattern=r"^ea_\d+_"),
        ],
        states={
            EDIT_FIELD_NAME:     [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_field_name_save)],
            EDIT_FIELD_PRICE:    [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_field_price_save)],
            EDIT_FIELD_CATEGORY: [CallbackQueryHandler(edit_field_cat_save, pattern=r"^ecat_")],
            EDIT_FIELD_DESC:     [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_field_desc_save)],
            EDIT_FIELD_PHOTOS:   [
                MessageHandler(filters.PHOTO | filters.Document.IMAGE | filters.Sticker.ALL | (filters.TEXT & ~filters.COMMAND), edit_field_photos_collect),
                CallbackQueryHandler(edit_photos_more, pattern="^eph_more$"),
                CallbackQueryHandler(edit_photos_done, pattern="^eph_done$"),
            ],
            EDIT_FIELD_ATTR:     [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_attr_save)],
        },
        fallbacks=global_fallbacks,
    )

    # --- Buyer profile edit ---
    buyer_profile_edit_conv = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(_go_to_app, pattern="^edit_buyer_name$"),
            CallbackQueryHandler(_go_to_app, pattern="^edit_buyer_phone$"),
        ],
        states={
            EDIT_PROFILE_NAME:  [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_buyer_name_submit)],
            EDIT_PROFILE_PHONE: [MessageHandler(filters.CONTACT | filters.TEXT & ~filters.COMMAND, edit_buyer_phone_submit)],
        },
        fallbacks=global_fallbacks,
    )

    # --- Seller profile edit ---
    # BUG FIX #6: pattern regex to'g'irlandi — "edit_telegram" ham ushlanadi
    seller_profile_edit_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(
            _go_to_app,
            pattern="^(edit_shop_name|edit_shop_address|edit_shop_landmark|edit_working_days|edit_working_hours|edit_telegram)$"
        )],
        states={
            EDIT_SHOP_NAME:         [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_seller_field_submit)],
            EDIT_SHOP_ADDRESS:      [MessageHandler(filters.LOCATION | filters.TEXT & ~filters.COMMAND, edit_seller_field_submit)],
            EDIT_SHOP_ADDRESS_TEXT: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_shop_address_text_submit)],
            EDIT_SHOP_LANDMARK:     [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_seller_field_submit)],
            EDIT_WORKING_DAYS:      [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_seller_field_submit)],
            EDIT_WORKING_HOURS:     [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_seller_field_submit)],
            EDIT_TELEGRAM_USERNAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_seller_field_submit)],
        },
        fallbacks=global_fallbacks,
    )

    # --- Messaging ---
    message_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(_go_to_app, pattern="^order_msg_")],
        states={
            MESSAGE_TEXT: [MessageHandler(filters.TEXT & ~filters.COMMAND, message_send)],
        },
        fallbacks=global_fallbacks,
    )

    # --- Rating (3 qadam: mahsulot reytingi -> mahsulot izohi -> do'kon reytingi) ---
    rating_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(_go_to_app, pattern="^order_rate_")],
        states={
            PRODUCT_RATING:  [CallbackQueryHandler(rating_product_select, pattern="^prate_")],
            PRODUCT_COMMENT: [MessageHandler(filters.TEXT & ~filters.COMMAND, rating_product_comment)],
            SELLER_RATING:   [CallbackQueryHandler(rating_submit, pattern="^srate_")],
        },
        fallbacks=global_fallbacks,
    )

    # /admin va /recommend — conversation boshlamaydigan alohida buyruqlar
    app.add_handler(CommandHandler("admin", admin_command))
    # APP-ONLY: /recommend va /ai xaridor/sotuvchi amallari — endi ilovada
    app.add_handler(CommandHandler("recommend", _go_to_app))
    app.add_handler(CommandHandler("ai", _go_to_app))

    # --- Become seller (mavjud xaridorni sotuvchi qilish) ---
    # Ro'yxatdan o'tish handlerlarini qayta ishlatamiz — ular faqat context.user_data ga yozadi,
    # DB ga yozuv esa oxirgi qadamda (become_seller_finish) bo'ladi.
    become_seller_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(_go_to_app, pattern="^become_seller$")],
        states={
            SHOP_NAME:         [MessageHandler(filters.TEXT & ~filters.COMMAND, registration_shop_name)],
            SHOP_LANDMARK:     [MessageHandler(filters.TEXT & ~filters.COMMAND, registration_shop_landmark)],
            SHOP_ADDRESS:      [MessageHandler(filters.LOCATION | filters.TEXT & ~filters.COMMAND, registration_shop_address)],
            SHOP_ADDRESS_TEXT: [MessageHandler(filters.TEXT & ~filters.COMMAND, registration_shop_address_text)],
            WORKING_DAYS:      [MessageHandler(filters.TEXT & ~filters.COMMAND, registration_working_days)],
            WORKING_HOURS:     [MessageHandler(filters.TEXT & ~filters.COMMAND, registration_working_hours)],
            TELEGRAM_USERNAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, become_seller_finish)],
        },
        fallbacks=global_fallbacks,
    )

    # --- Buyurtma berish (Order Flow) ---
    order_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(_go_to_app, pattern=r"^order_\d+$")],
        states={
            ORDER_QUANTITY:      [MessageHandler(filters.TEXT & ~filters.COMMAND, order_quantity)],
            ORDER_DELIVERY_TYPE: [CallbackQueryHandler(order_delivery_type, pattern=r"^ord_dlv_|^ord_cancel$")],
            ORDER_ADDRESS:       [MessageHandler(filters.LOCATION | (filters.TEXT & ~filters.COMMAND), order_address)],
            ORDER_PAYMENT:       [CallbackQueryHandler(order_payment, pattern=r"^ord_pay_|^ord_cancel$")],
            ORDER_CONFIRM:       [CallbackQueryHandler(order_confirm, pattern=r"^ord_confirm_yes$|^ord_cancel$")],
        },
        fallbacks=global_fallbacks,
    )

    # ConversationHandler'lar — /start ichida boshlanadi
    app.add_handler(registration_conv)
    app.add_handler(become_seller_conv)
    app.add_handler(order_conv)
    app.add_handler(product_conv)
    app.add_handler(edit_product_conv)
    app.add_handler(buyer_profile_edit_conv)
    app.add_handler(seller_profile_edit_conv)
    app.add_handler(message_conv)
    app.add_handler(rating_conv)

    # --- Shartnomani bekor qilish (sabab tanlash oqimi) ---
    cancel_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(_go_to_app, pattern=r"^ccl_req_\d+$")],
        states={
            CANCEL_PICK_REASON: [CallbackQueryHandler(cancel_reason_pick, pattern=r"^ccl_rsn_|^ccl_air_|^ccl_abort$")],
            CANCEL_REASON_TEXT: [MessageHandler(filters.TEXT & ~filters.COMMAND, cancel_reason_text)],
        },
        fallbacks=global_fallbacks,
    )
    app.add_handler(cancel_conv)

    # --- Admin nizo bo'yicha tomonga (bot orqali) xabar yozishi ---
    admin_dm_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(admin_dispute_msg_start, pattern=r"^admindm_(buyer|seller)_\d+$")],
        states={
            ADMIN_DISPUTE_MSG: [MessageHandler(filters.TEXT & ~filters.COMMAND, admin_dispute_msg_send)],
        },
        fallbacks=global_fallbacks,
        conversation_timeout=120,
    )
    app.add_handler(admin_dm_conv)

    # --- Xaridor/sotuvchi admin xabariga javob berishi ---
    dispute_reply_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(dispute_reply_start, pattern=r"^dmreply_\d+$")],
        states={
            DMREPLY_MSG: [MessageHandler(filters.TEXT & ~filters.COMMAND, dispute_reply_send)],
        },
        fallbacks=global_fallbacks,
        conversation_timeout=300,
    )
    app.add_handler(dispute_reply_conv)

    # --- Karta ma'lumotlari tahrirlash ---
    card_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(_go_to_app, pattern="^edit_card_info$")],
        states={
            EDIT_CARD_TYPE:   [CallbackQueryHandler(edit_card_type,   pattern="^card_type_")],
            EDIT_CARD_NUMBER: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_card_number)],
            EDIT_CARD_OWNER:  [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_card_owner)],
        },
        fallbacks=global_fallbacks,
    )
    app.add_handler(card_conv)

    # --- Admin bilan bog'lanish ---
    contact_admin_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(_go_to_app, pattern="^contact_admin$")],
        states={
            CONTACT_ADMIN_MSG: [MessageHandler(filters.TEXT & ~filters.COMMAND, contact_admin_send)],
        },
        fallbacks=global_fallbacks,
    )
    app.add_handler(contact_admin_conv)

    # --- Sotuvchi kanalini ulash (forward orqali) ---
    link_channel_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(_go_to_app, pattern="^seller_link_channel$")],
        states={
            LINK_CHANNEL_WAIT: [
                MessageHandler(filters.FORWARDED, link_channel_receive),
                MessageHandler(filters.TEXT & ~filters.COMMAND & ~cancel_filter, link_channel_wait_hint),
            ],
        },
        fallbacks=global_fallbacks,
    )
    app.add_handler(link_channel_conv)

    # Sotuvchi kanallari menyusi + kanalni o'chirish — APP-ONLY (Faza 2): ilovada
    app.add_handler(CallbackQueryHandler(_go_to_app, pattern="^seller_channels_menu$"))
    app.add_handler(CallbackQueryHandler(_go_to_app, pattern="^seller_channels_recheck$"))
    app.add_handler(CallbackQueryHandler(_go_to_app, pattern="^chremove_"))

    # --- Sotuvchi guruhini ulash ---
    # Guruh ID si forward orqali ko'rinmaydi, shuning uchun "Guruh qo'shish" faqat
    # ko'rsatma beradi; haqiqiy bog'lanish bot guruhga qo'shilganda (my_chat_member) sodir bo'ladi.
    app.add_handler(CallbackQueryHandler(_go_to_app, pattern="^seller_link_group$"))
    app.add_handler(ChatMemberHandler(on_my_chat_member, ChatMemberHandler.MY_CHAT_MEMBER))

    # Admin mahsulot moderatsiyasi (detal, qidiruv, sotuvdan olib tashlash, rasm)
    app.add_handler(CallbackQueryHandler(admin_products, pattern=r"^admin_products(_pg_\d+)?$"))
    app.add_handler(CallbackQueryHandler(admin_product_search, pattern=r"^admin_product_search$"))
    app.add_handler(CallbackQueryHandler(admin_product_detail, pattern=r"^admin_prod_\d+$"))
    app.add_handler(CallbackQueryHandler(admin_product_remove, pattern=r"^admin_prodrm_\d+$"))
    app.add_handler(CallbackQueryHandler(admin_product_restore, pattern=r"^admin_prodrs_\d+$"))
    app.add_handler(CallbackQueryHandler(admin_product_image, pattern=r"^admin_prodimg_\d+$"))

    # Admin javob komandasi: /reply 123456789 matn
    app.add_handler(CommandHandler("reply", admin_reply_command))

    # Kanal ulanishini tekshirish (faqat admin): /testpost
    app.add_handler(CommandHandler("testpost", testpost))

    # --- Sotuvchi mahsulot qidirish ---
    sp_search_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(_go_to_app, pattern="^sp_search$")],
        states={
            SELLER_PRODUCT_SEARCH: [MessageHandler(filters.TEXT & ~filters.COMMAND, seller_product_search_result)],
        },
        fallbacks=global_fallbacks,
    )
    app.add_handler(sp_search_conv)

    # --- Savatni rasmiylashtirish (savat buyurtmasi) ---
    cart_checkout_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(_go_to_app, pattern="^cart_checkout$")],
        states={
            CART_DELIVERY_TYPE: [CallbackQueryHandler(cart_delivery_type, pattern=r"^cart_dlv_|^cart_cancel$")],
            CART_ADDRESS:       [MessageHandler(filters.LOCATION | (filters.TEXT & ~filters.COMMAND), cart_address)],
            CART_PAYMENT:       [CallbackQueryHandler(cart_payment, pattern=r"^cart_pay_|^cart_cancel$")],
            CART_CONFIRM:       [CallbackQueryHandler(cart_confirm, pattern=r"^cart_confirm_yes$|^cart_cancel$")],
        },
        fallbacks=global_fallbacks,
    )
    app.add_handler(cart_checkout_conv)

    # Umumiy handler'lar ENG OXIRIDA
    app.add_handler(CallbackQueryHandler(button_handler))
    app.add_handler(MessageHandler(filters.LOCATION, location_handler))
    # AI rejimida mahsulot rasmlarini yig'ish (faqat 'ai_awaiting_photos' yoqilganda ishlaydi)
    app.add_handler(MessageHandler(filters.PHOTO | filters.Document.IMAGE, ai_photo_collect))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, text_handler))

    # Global xato ushlovchi — kutilmagan exception bo'lsa, foydalanuvchini xabardor qiladi
    app.add_error_handler(error_handler)

    # Kunlik ishlaydigan job — eski 'pending' buyurtmalarni avtomatik bekor qilish.
    # Bot ishga tushgandan 60 soniyadan keyin birinchi marta, keyin har 24 soatda.
    if app.job_queue:
        app.job_queue.run_repeating(cleanup_stale_orders_job, interval=86400, first=60)
        logging.info("Stale orders cleanup job rejalashtirildi (har 24 soatda)")

        # Mini App (webapp) yaratgan buyurtmalarga sotuvchi bildirishnomasi (har 12s)
        app.job_queue.run_repeating(webapp_order_dispatch_job, interval=12, first=15)
        logging.info("Mini App buyurtma dispatch job rejalashtirildi (har 12s)")

        # #3 — App'da kuryerga biriktirilgan buyurtmalar uchun kuryer PUSH (har 12s)
        app.job_queue.run_repeating(webapp_courier_notify_job, interval=12, first=18)
        logging.info("Mini App kuryer biriktirish PUSH job rejalashtirildi (har 12s)")

        # Mini App yaratgan rejalashtirilgan postlarga publish jobini ulash (har 30s)
        app.job_queue.run_repeating(webapp_scheduled_scan_job, interval=30, first=20)
        logging.info("Mini App rejalashtirilgan post scan job rejalashtirildi (har 30s)")

        # Mini App yaratgan avto qayta-reklamalarga kunlik jobni ulash (har 60s)
        app.job_queue.run_repeating(webapp_autorepost_scan_job, interval=60, first=25)
        logging.info("Mini App auto-repost scan job rejalashtirildi (har 60s)")

        # Avtomatik backup — har kuni ertalab 06:00 (UTC) = 11:00 Toshkent
        from datetime import time as dt_time
        app.job_queue.run_daily(
            auto_backup_job,
            time=dt_time(hour=6, minute=0),
            name="daily_backup"
        )
        logging.info("Avtomatik backup job rejalashtirildi (har kuni 11:00 Toshkent)")

        # RESTART: ochiq buyurtmalar uchun eslatma/avto-bekor taymerlarini tiklaymiz
        _reschedule_pending_order_timers(app.job_queue)
        # RESTART: rejalashtirilgan postlarni qayta tiklaymiz
        _reschedule_scheduled_posts(app.job_queue)
        # RESTART: avto qayta-reklamalarni qayta tiklaymiz
        _reschedule_auto_reposts(app.job_queue)

    print("🚀 TezBozor Bot ishlamoqda...")
    # MUHIM: allowed_updates'ni ANIQ ko'rsatamiz — aks holda Telegram a'zolik
    # yangilanishlarini (my_chat_member) yubormaydi va bot guruhga qo'shilganini sezmaydi.
    # Update.ALL_TYPES barcha turdagi yangilanishlarni (jumladan my_chat_member) yoqadi.
    #
    # REJIM (dual-mode): WEBHOOK_URL env o'rnatilgan bo'lsa — webhook, aks holda polling.
    # DEFAULT = polling (WEBHOOK_URL yo'q) → hozirgi xatti-harakat AYNAN o'zgarmaydi.
    # Webhook cutover faqat VPS'da, nginx route + set_webhook bilan — WEBHOOK_MIGRATION.md.
    webhook_url = (os.getenv("WEBHOOK_URL") or "").strip()
    if webhook_url:
        port = int(os.getenv("WEBHOOK_PORT", "8443"))
        # url_path = WEBHOOK_PATH yoki URL'ning TO'LIQ yo'li (faqat oxirgi bo'lak EMAS —
        # nginx ko'p bo'lakli '/tg/<secret>' ni o'zgarmasdan uzatadi, PTB aynan shu
        # path'ni kutadi; aks holda 404). Boshidagi '/' olib tashlanadi.
        from urllib.parse import urlparse as _urlparse
        url_path = (os.getenv("WEBHOOK_PATH") or "").strip() or _urlparse(webhook_url).path.lstrip("/")
        secret = (os.getenv("WEBHOOK_SECRET") or "").strip() or None
        logging.info("Webhook rejimi: %s (127.0.0.1:%s, path=%s)", webhook_url, port, url_path)
        app.run_webhook(
            listen="127.0.0.1",          # nginx oldida turadi (TLS terminatsiya nginx'da)
            port=port,
            url_path=url_path,
            webhook_url=webhook_url,
            secret_token=secret,         # Telegram so'rovini header bilan tasdiqlaydi
            allowed_updates=Update.ALL_TYPES,
            drop_pending_updates=False,
        )
    else:
        app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()